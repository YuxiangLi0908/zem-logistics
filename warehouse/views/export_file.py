import io
import json
import os
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Any
import pandas as pd
import pytz
from asgiref.sync import sync_to_async
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib.postgres.aggregates import StringAgg
from django.contrib.staticfiles import finders
from django.core.exceptions import MultipleObjectsReturned, ObjectDoesNotExist
from django.db.models import (
    Case,
    CharField,
    Count,
    DateTimeField,
    F,
    FloatField,
    IntegerField,
    Max,
    Q,
    Sum,
    Value,
    When, Prefetch,
)
from django.db.models.functions import Cast, Concat, Round
from django.forms import model_to_dict
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render
from django.template.loader import get_template
from django.utils.decorators import method_decorator
from django.views import View
from xhtml2pdf import pisa

from warehouse.models.order import Order
from warehouse.models.vessel import Vessel
from warehouse.models.packing_list import PackingList
from warehouse.models.pallet import Pallet
from warehouse.models.po_check_eta import PoCheckEtaSeven
from warehouse.utils.constants import (
    ACCT_ACH_ROUTING_NUMBER,
    ACCT_BANK_NAME,
    ACCT_BENEFICIARY_ACCOUNT,
    ACCT_BENEFICIARY_ADDRESS,
    ACCT_BENEFICIARY_NAME,
    ACCT_SWIFT_CODE,
)


@method_decorator(login_required(login_url="login"), name="dispatch")
class ExportFile(View):
    template_main = {"DO": "export_file/do.html", "PL": "export_file/packing_list.html"}
    file_name = {"DO": "D/O", "PL": "拆柜单"}

    def get(self, request: HttpRequest) -> HttpResponse:
        name = request.GET.get("name")
        template_path = self.template_main[name]
        template = get_template(template_path)
        context = {"sample_data": "Hello, this is some sample data!"}
        html = template.render(context)

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="{self.file_name[name]}.pdf"'
        )
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            raise ValueError(
                "Error during PDF generation: %s" % pisa_status.err,
                content_type="text/plain",
            )
        return response


def export_bol(context: dict[str, Any]) -> HttpResponse:
    template_path = "export_file/bol_template.html"
    template = get_template(template_path)
    html = template.render(context)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="BOL_{context["batch_number"]}.pdf"'
    )
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        raise ValueError(
            "Error during PDF generation: %s" % pisa_status.err,
            content_type="text/plain",
        )
    return response


async def export_palletization_list_v2(request: HttpRequest) -> HttpResponse:
    """
    (新)拆柜单导出
    """
    status = request.POST.get("status")
    warehouse = request.POST.get("warehouse").split("-")[0].upper() if request.POST.get("warehouse") else ""
    container_number = request.POST.get("container_number")
    warehouse_unpacking_time = request.POST.get("first_time_download")

    try:
        warehouse_unpacking_time = datetime.strptime(warehouse_unpacking_time, "%Y-%m-%d %H:%M:%S").strftime(
            "%d/%m/%Y")
    except (ValueError, TypeError):
        warehouse_unpacking_time = "未获取到时间"

    TARGET_WAREHOUSES = {"GEU2", "GEU3", "GYR2", "GYR3", "IUSP", "ONT8", "LAX9", "LGB8", "SBD1"}  # 指定9个仓点
    UTC_TZ = pytz.UTC
    BASE_ETA = UTC_TZ.localize(datetime(2026, 1, 19))  # 带UTC时区的基准时间

    if status == "non_palletized":
        vessel_prefetch_queryset = Vessel.objects.all()
        packing_list = await sync_to_async(list)(
            PackingList.objects.select_related("container_number", "pallet", "shipment_batch_number")
            .prefetch_related(
                Prefetch(
                    "container_number__orders__vessel_id",
                    queryset=vessel_prefetch_queryset
                )
            )
            .filter(container_number__container_number=container_number)
            .annotate(
                custom_delivery_method=Case(
                    When(
                        Q(delivery_method="暂扣留仓(HOLD)")
                        | Q(delivery_method="暂扣留仓"),
                        then=Concat(
                            "delivery_method", Value("-"), "fba_id", Value("-"), "id"
                        ),
                    ),
                    When(
                        Q(delivery_method="客户自提") & ~Q(destination="客户自提"),
                        then=Concat(
                            "delivery_method",
                            Value("-"),
                            "destination",
                        ),
                    ),
                    When(
                        Q(delivery_method="客户自提") | Q(destination="客户自提"),
                        then=Concat(
                            "delivery_method",
                            Value("-"),
                            "destination",
                            Value("-"),
                            "shipping_mark",
                        ),
                    ),
                    default=F("delivery_method"),
                    output_field=CharField(),
                ),
                str_id=Cast("id", CharField()),
                str_fba_id=Cast("fba_id", CharField()),
                str_ref_id=Cast("ref_id", CharField()),
                str_shipping_mark=Cast("shipping_mark", CharField()),
                vessel_eta=F("container_number__orders__vessel_id__vessel_eta"),
            )
            .values(
                "container_number__container_number",
                "destination",
                "address",
                "zipcode",
                "contact_name",
                "custom_delivery_method",
                "note",
                "shipment_batch_number__shipment_batch_number",
                "PO_ID",
                "delivery_type",
                "shipment_batch_number__load_type",
                "vessel_eta",  # 带时区的datetime值
            )
            .annotate(
                fba_ids=StringAgg("str_fba_id", delimiter=",", distinct=True),
                ref_ids=StringAgg("str_ref_id", delimiter=",", distinct=True),
                shipping_marks=StringAgg(
                    "str_shipping_mark", delimiter=",", distinct=True
                ),
                ids=StringAgg("str_id", delimiter=",", distinct=True),
                pcs=Sum("pcs", output_field=IntegerField()),
                cbm=Sum("cbm", output_field=FloatField()),
                n_pallet=Count("pallet__pallet_id", distinct=True),
                weight_lbs=Sum("total_weight_lbs", output_field=FloatField()),
                plt_ids=StringAgg(
                    "str_id", delimiter=",", distinct=True, ordering="str_id"
                ),
            )
            .order_by("-cbm")
        )
    else:
        packing_list = []

    data = [i for i in packing_list]
    df = pd.DataFrame.from_records(data)
    df["拆柜备注"] = ""  # 初始化拆柜备注为空

    if not df.empty:
        df["vessel_eta_dt"] = pd.to_datetime(df["vessel_eta"], errors="coerce", utc=True)

        # 2. 定义判断条件（时区一致，可直接比较）
        mask = (
                (df["shipment_batch_number__load_type"] != "卡板")  # 卡板类型
                & (df["vessel_eta_dt"] >= BASE_ETA)  # 带时区的时间比较，无类型错误
                & (df["destination"].isin(TARGET_WAREHOUSES))  # 仓点在指定9个列表内
        )

        df.loc[mask, "拆柜备注"] = "100 height"

        # ===================== 核心修改：合并note到拆柜备注 =====================
        # 1. 处理note空值：将NaN/None转为空字符串
        df["note"] = df["note"].fillna("").astype(str)

        # 2. 定义拼接逻辑：
        # - 拆柜备注是100 height + note非空 → "100 height, note内容"
        # - 拆柜备注是空 + note非空 → "note内容"
        # - 其他情况保持原拆柜备注
        def merge_note_to_remark(remark, note):
            if note.strip() == "":  # note为空，直接返回原备注
                return remark
            if remark == "100 height":  # 有100 height，加逗号拼接
                return f"{remark}, {note.strip()}"
            else:  # 无100 height，直接用note
                return note.strip()

        # 应用拼接逻辑到拆柜备注列
        df["拆柜备注"] = df.apply(
            lambda row: merge_note_to_remark(row["拆柜备注"], row["note"]),
            axis=1
        )

        # 3. 清空原note列
        df["note"] = ""

    if not df.empty:
        df = df.rename(
            {
                "container_number__container_number": "container_number",
                "custom_delivery_method": "delivery_method",
                "shipping_marks": "shipping_mark",
                "n_pallet": "pl"
            },
            axis=1,
        )
        df["delivery_method"] = df["delivery_method"].apply(
            lambda x: x.split("-")[0] if x and isinstance(x, str) else x
        )
        df["pcs_original"] = df["pcs"].astype(str)  # 保存pcs原始值
        df["pcs"] = df["pcs_original"].copy()

        # 清空pcs和shipping_mark的逻辑（调整后）
        mask_base = (df["delivery_method"] == "卡车派送") & (df["delivery_type"] == "public")

        # 1. 从拆柜备注提取原note内容
        def extract_original_note(remark):
            if not isinstance(remark, str) or remark.strip() == "":
                return ""
            note_part = remark.replace("100 height, ", "").replace("100 height", "")
            return note_part.strip()

        df["original_note_from_remark"] = df["拆柜备注"].apply(extract_original_note)

        # 2. 判断特操
        mask_tecao = df["original_note_from_remark"].apply(
            lambda x: "特操" in x if x else False
        )
        mask_clear_pcs = mask_base & (~mask_tecao)
        df.loc[mask_clear_pcs, "pcs"] = ""

        # 3. 判断note是否为空
        mask_note_empty = (df["original_note_from_remark"] == "")
        mask_clear_mark = mask_base & mask_note_empty
        df.loc[mask_clear_mark, "shipping_mark"] = ""

        # 4. 清理临时列
        df = df.drop("original_note_from_remark", axis=1)

        df["pl"] = ""  # 清空打板数字段

        df = df[["destination", "delivery_method", "shipping_mark", "pcs", "pl", "note", "拆柜备注"]]
    else:
        df = pd.DataFrame(columns=["destination", "delivery_method", "shipping_mark", "pcs", "pl", "note", "拆柜备注"])

    df = df[["destination", "delivery_method", "shipping_mark", "拆柜备注", "pcs", "pl", "note"]]

    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "拆柜单"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header1_font = Font(size=15, bold=True)
    header2_font = Font(size=11, bold=True)
    data_font = Font(size=12, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    UNIFIED_ROW_HEIGHT = 30

    # 第一行合并单元格（柜号）
    ws.merge_cells('A1:C1')
    for col_idx in range(1, 4):
        cell = ws.cell(row=1, column=col_idx)
        cell.border = thin_border
    ws['A1'] = container_number or "未指定柜号"
    ws['A1'].font = header1_font
    ws['A1'].alignment = left_alignment
    ws.row_dimensions[1].height = UNIFIED_ROW_HEIGHT

    # 第一行合并单元格（拆柜时间）
    ws.merge_cells('D1:E1')
    for col_idx in range(4, 6):
        cell = ws.cell(row=1, column=col_idx)
        cell.border = thin_border
    ws['D1'] = warehouse_unpacking_time
    ws['D1'].font = header1_font
    ws['D1'].alignment = center_alignment

    # 第一行合并单元格（dock）
    ws.merge_cells('F1:G1')
    for col_idx in range(6, 8):
        cell = ws.cell(row=1, column=col_idx)
        cell.border = thin_border
    ws['F1'] = 'dock'
    ws['F1'].font = header1_font
    ws['F1'].alignment = center_alignment

    # 表头行（第二行）
    column_names = ["destination", "delivery_method", "shipping_mark", "拆柜备注", "pcs", "pl", "note"]
    for col_idx, name in enumerate(column_names, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = name
        cell.font = header2_font
        cell.border = thin_border
        cell.alignment = center_alignment
    ws.row_dimensions[2].height = UNIFIED_ROW_HEIGHT

    # 数据行
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 3):
        ws.row_dimensions[row_idx].height = UNIFIED_ROW_HEIGHT
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value if pd.notna(value) else ""
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 6:  # pl列左对齐
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment

    # 调整列宽
    total_columns = len(column_names)
    for col_idx in range(1, total_columns + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws[f"{column_letter}{row_idx}"]
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            cell_value = str(cell.value) if cell.value is not None else ""
            if len(cell_value) > max_length:
                max_length = len(cell_value)

        # 列宽适配
        if col_idx == 6:
            adjusted_width = max(10, min(max_length + 5, 30))
        else:
            adjusted_width = max(8, min(max_length + 2, 20))
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存并返回响应
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"{container_number if container_number else '拆柜单'}.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"

    return response

async def export_palletization_list(request: HttpRequest) -> HttpResponse:
    status = request.POST.get("status")
    container_number = request.POST.get("container_number")
    if status == "non_palletized":
        packing_list = await sync_to_async(list)(
            PackingList.objects.select_related("container_number", "pallet")
            .filter(container_number__container_number=container_number)
            .annotate(
                custom_delivery_method=Case(
                    When(
                        Q(delivery_method="暂扣留仓(HOLD)")
                        | Q(delivery_method="暂扣留仓"),
                        then=Concat(
                            "delivery_method", Value("-"), "fba_id", Value("-"), "id"
                        ),
                    ),
                    When(
                        Q(delivery_method="客户自提") | Q(destination="客户自提"),
                        then=Concat(
                            "delivery_method",
                            Value("-"),
                            "destination",
                            Value("-"),
                            "shipping_mark",
                        ),
                    ),
                    default=F("delivery_method"),
                    output_field=CharField(),
                ),
                str_id=Cast("id", CharField()),
                str_fba_id=Cast("fba_id", CharField()),
                str_ref_id=Cast("ref_id", CharField()),
                str_shipping_mark=Cast("shipping_mark", CharField()),
            )
            .values(
                "container_number__container_number",
                "destination",
                "address",
                "zipcode",
                "contact_name",
                "custom_delivery_method",
                "note",
                "shipment_batch_number__shipment_batch_number",
                "PO_ID",
                # "delivery_window_start",
                # "delivery_window_end",
            )
            .annotate(
                fba_ids=StringAgg("str_fba_id", delimiter=",", distinct=True),
                ref_ids=StringAgg("str_ref_id", delimiter=",", distinct=True),
                shipping_marks=StringAgg(
                    "str_shipping_mark", delimiter=",", distinct=True
                ),
                ids=StringAgg("str_id", delimiter=",", distinct=True),
                pcs=Sum("pcs", output_field=IntegerField()),
                cbm=Round(Sum("cbm"), 2, output_field=FloatField()),
                n_pallet=Count("pallet__pallet_id", distinct=True),
                weight_lbs=Sum("total_weight_lbs", output_field=FloatField()),
                plt_ids=StringAgg(
                    "str_id", delimiter=",", distinct=True, ordering="str_id"
                ),
            )
            .order_by("-cbm")
        )
    elif status == "palletized":
        packing_list = await sync_to_async(list)(
            Pallet.objects.select_related("container_number")
            .filter(container_number__container_number=container_number)
            .values(
                "container_number__container_number",
                "delivery_method",
                "destination",
                "fba_id",
                "ref_id",
                "shipping_mark",
                "note",
                "PO_ID",
                # "delivery_window_start",
                # "delivery_window_end",
            )
            .annotate(
                pcs=Sum("pcs", output_field=IntegerField()),
                cbm=Round(Sum("cbm"), 2, output_field=FloatField()),
                n_pallet=Count("pallet_id", distinct=True),
            )
            .order_by("-cbm")
        )
        packing_list_complement = await sync_to_async(list)(
            PackingList.objects.select_related("container_number", "pallet")
            .filter(container_number__container_number=container_number)
            .annotate(
                custom_delivery_method=Case(
                    When(
                        Q(delivery_method="暂扣留仓(HOLD)")
                        | Q(delivery_method="暂扣留仓"),
                        then=Concat(
                            "delivery_method", Value("-"), "fba_id", Value("-"), "id"
                        ),
                    ),
                    When(
                        Q(delivery_method="客户自提") | Q(destination="客户自提"),
                        then=Concat(
                            "delivery_method",
                            Value("-"),
                            "destination",
                            Value("-"),
                            "shipping_mark",
                        ),
                    ),
                    default=F("delivery_method"),
                    output_field=CharField(),
                ),
                str_id=Cast("id", CharField()),
                str_fba_id=Cast("fba_id", CharField()),
                str_ref_id=Cast("ref_id", CharField()),
                str_shipping_mark=Cast("shipping_mark", CharField()),
            )
            .values(
                "container_number__container_number",
                "destination",
                "address",
                "zipcode",
                "contact_name",
                "custom_delivery_method",
                "note",
                "shipment_batch_number__shipment_batch_number",
                "PO_ID",
                # "delivery_window_start",
                # "delivery_window_end",
            )
            .annotate(
                fba_ids=StringAgg("str_fba_id", delimiter=",", distinct=True),
                ref_ids=StringAgg("str_ref_id", delimiter=",", distinct=True),
                shipping_marks=StringAgg(
                    "str_shipping_mark", delimiter=",", distinct=True
                ),
                ids=StringAgg("str_id", delimiter=",", distinct=True),
                pcs=Sum("pcs", output_field=IntegerField()),
                cbm=Round(Sum("cbm"), 2, output_field=FloatField()),
                n_pallet=Count("pallet__pallet_id", distinct=True),
                weight_lbs=Sum("total_weight_lbs", output_field=FloatField()),
                plt_ids=StringAgg(
                    "str_id", delimiter=",", distinct=True, ordering="str_id"
                ),
            )
            .order_by("-cbm")
        )
        existing_po = {
            (plt["container_number__container_number"], plt["destination"])
            for plt in packing_list
        }
        for pl in packing_list_complement:
            po = (pl["container_number__container_number"], pl["destination"])
            if po not in existing_po:
                packing_list.append(
                    {
                        "container_number__container_number": pl[
                            "container_number__container_number"
                        ],
                        "delivery_method": pl["custom_delivery_method"],
                        "destination": pl["destination"],
                        "fba_id": pl["fba_ids"],
                        "ref_id": pl["ref_ids"],
                        "shipping_mark": pl["shipping_marks"],
                        "note": pl["note"],
                        "PO_ID": pl["PO_ID"],
                        "pcs": 0,
                        "cbm": 0,
                        "n_pallet": 0,
                        # "delivery_window_start": pl["delivery_window_start"],
                        # "delivery_window_end": pl["delivery_window_end"],
                    }
                )
    else:
        raise ValueError(f"Unknown container status: {status}\n{request.POST}")

    data = [i for i in packing_list]
    df = pd.DataFrame.from_records(data)
    df = df.rename(
        {
            "container_number__container_number": "container_number",
            "custom_delivery_method": "delivery_method",
            "fba_ids": "fba_id",
            "ref_ids": "ref_id",
            "shipping_marks": "shipping_mark",
        },
        axis=1,
    )
    df["delivery_method"] = df["delivery_method"].apply(lambda x: x.split("-")[0])
    df = df[
        [
            "container_number",
            "destination",
            "delivery_method",
            "fba_id",
            "ref_id",
            "shipping_mark",
            "pcs",
            "cbm",
            "n_pallet",
            "PO_ID",
            "note",
            # "delivery_window_start",
            # "delivery_window_end",
        ]
    ]
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={container_number}.xlsx"
    df.to_excel(excel_writer=response, index=False, columns=df.columns)
    return response

def export_po_check(request: HttpRequest) -> HttpResponse:
    pl_ids = request.POST.getlist("pl_ids")
    pls = [pl.split(",") for pl in pl_ids]
    selections = request.POST.getlist("is_selected")
    ids = [o for s, co in zip(selections, pls) for o in co if s == "on"]

    checkid_list = request.POST.getlist("ids")
    check_ids = [i.split(",") for i in checkid_list]
    check_id = [o for s, co in zip(selections, check_ids) for o in co if s == "on"]
    id_check_id_map = {
        int(id_val): check_val for id_val, check_val in zip(ids, check_id)
    }
    if ids:
        # 查找柜号下的pl
        packing_list = (
            PackingList.objects.select_related("container_number", "pallet")
            .filter(id__in=ids)
            .values(
                "id",
                "shipping_mark",
                "fba_id",
                "ref_id",
                "address",
                "zipcode",
                "container_number__container_number",
            )
            .annotate(
                total_pcs=Sum(
                    Case(
                        When(pallet__isnull=True, then=F("pcs")),
                        default=F("pallet__pcs"),
                        output_field=IntegerField(),
                    )
                ),
                total_n_pallet_act=Count("pallet__pallet_id", distinct=True),
                total_n_pallet_est=Sum("cbm", output_field=FloatField()) / 2,
                label=Max(
                    Case(
                        When(pallet__isnull=True, then=Value("EST")),
                        default=Value("ACT"),
                        output_field=CharField(),
                    )
                ),
            )
            .distinct()
            .order_by("destination", "container_number__container_number")
        )
    data = [i for i in packing_list]
    for item in data:
        item_id = item["id"]
        if item_id in id_check_id_map:
            item["check_id"] = id_check_id_map[item_id]
    keep = [
        "shipping_mark",
        "container_number__container_number",
        "fba_id",
        "ref_id",
        "total_pcs",
        "Pallet Count",
        "label",
        "check_id",
        "is_valid",
    ]
    df = pd.DataFrame.from_records(data)
    df["is_valid"] = None

    def get_est_pallet(n):
        if n < 1:
            return 1
        elif n % 1 >= 0.45:
            return int(n // 1 + 1)
        else:
            return int(n // 1)

    df["total_n_pallet_est"] = df["total_n_pallet_est"].apply(get_est_pallet)
    df["est"] = df["label"] == "EST"
    df["act"] = df["label"] == "ACT"
    df["Pallet Count"] = (
        df["total_n_pallet_act"] * df["act"] + df["total_n_pallet_est"] * df["est"]
    )
    df = df[keep].rename(
        {
            "fba_id": "PRO",
            "container_number__container_number": "BOL",
            "ref_id": "PO List (use , as separator) *",
            "total_pcs": "Carton Count",
        },
        axis=1,
    )
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f"attachment; filename=PO.csv"
    df.to_csv(path_or_buf=response, index=False)
    return response


def export_report(request: HttpRequest, export_format: str = "PO") -> HttpResponse:
    pl_ids = request.POST.getlist("pl_ids")
    pls = [pl.split(",") for pl in pl_ids]
    selections = request.POST.getlist("is_selected")
    ids = [o for s, co in zip(selections, pls) for o in co if s == "on"]
    return {"stauts": "ok"}


def export_po(request: HttpRequest, export_format: str = "PO") -> HttpResponse:
    ids = request.POST.get("pl_ids")
    ids = ids.replace("[", "").replace("]", "").split(", ")
    ids = [int(i) for i in ids]
    packing_list = (
        PackingList.objects.select_related("container_number", "pallet")
        .filter(id__in=ids)
        .values(
            "fba_id",
            "ref_id",
            "address",
            "zipcode",
            "destination",
            "delivery_method",
            "container_number__container_number",
            "shipping_mark",
        )
        .annotate(
            total_pcs=Sum(
                Case(
                    When(pallet__isnull=True, then=F("pcs")),
                    default=F("pallet__pcs"),
                    output_field=IntegerField(),
                )
            ),
            total_cbm=Sum(
                Case(
                    When(pallet__isnull=True, then=F("cbm")),
                    default=F("pallet__cbm"),
                    output_field=FloatField(),
                )
            ),
            total_weight_lbs=Sum(
                Case(
                    When(pallet__isnull=True, then=F("total_weight_lbs")),
                    default=F("pallet__weight_lbs"),
                    output_field=FloatField(),
                )
            ),
            total_n_pallet_act=Count("pallet__pallet_id", distinct=True),
            total_n_pallet_est=Sum("cbm", output_field=FloatField()) / 2,
            label=Max(
                Case(
                    When(pallet__isnull=True, then=Value("EST")),
                    default=Value("ACT"),
                    output_field=CharField(),
                )
            ),
        )
        .distinct()
        .order_by("destination", "container_number__container_number")
    )
    for p in packing_list:
        try:
            pl = PoCheckEtaSeven.objects.get(
                container_number__container_number=p[
                    "container_number__container_number"
                ],
                shipping_mark=p["shipping_mark"],
                fba_id=p["fba_id"],
                ref_id=p["ref_id"],
            )

            if not pl.last_eta_checktime and not pl.last_retrieval_checktime:
                p["check"] = "未校验"
            elif pl.last_retrieval_checktime and not pl.last_retrieval_status:
                if pl.handling_method:
                    p["check"] = "失效," + str(pl.handling_method)
                else:
                    p["check"] = "失效未处理"
            elif (
                not pl.last_retrieval_checktime
                and pl.last_eta_checktime
                and not pl.last_eta_status
            ):
                if pl.handling_method:
                    p["check"] = "失效," + str(pl.handling_method)
                else:
                    p["check"] = "失效未处理"
            else:
                p["check"] = "有效"
        except PoCheckEtaSeven.DoesNotExist:
            p["check"] = "未找到记录"
        except MultipleObjectsReturned:
            p["check"] = "唛头FBA_REF重复"
    data = [i for i in packing_list]
    if export_format == "PO":
        keep = [
            "fba_id",
            "container_number__container_number",
            "ref_id",
            "Pallet Count",
            "total_pcs",
            "label",
            "check",
        ]
    elif export_format == "FULL_TABLE":
        keep = [
            "container_number__container_number",
            "destination",
            "delivery_method",
            "fba_id",
            "ref_id",
            "total_cbm",
            "total_pcs",
            "total_weight_lbs",
            "Pallet Count",
            "label",
            "check",
        ]
    else:
        raise ValueError(f"unknown export_format option: {export_format}")
    df = pd.DataFrame.from_records(data)

    def get_est_pallet(n):
        if n < 1:
            return 1
        elif n % 1 >= 0.45:
            return int(n // 1 + 1)
        else:
            return int(n // 1)

    df["total_n_pallet_est"] = df["total_n_pallet_est"].apply(get_est_pallet)
    df["est"] = df["label"] == "EST"
    df["act"] = df["label"] == "ACT"
    df["Pallet Count"] = (
        df["total_n_pallet_act"] * df["act"] + df["total_n_pallet_est"] * df["est"]
    )
    df = df[keep].rename(
        {
            "fba_id": "PRO",
            "container_number__container_number": "BOL",
            "ref_id": "PO List (use , as separator) *",
            "total_pcs": "Carton Count",
        },
        axis=1,
    )
    if export_format == "FULL_TABLE":
        df = df.rename(
            {
                "total_cbm": "CBM",
                "total_weight_lbs": "WEIGHT(LBS)",
            },
            axis=1,
        )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=PO.xlsx"
    df.to_excel(excel_writer=response, index=False, columns=df.columns)
    return response


def export_do(request: HttpRequest) -> HttpResponse:
    selected_orders = json.loads(request.POST.get("selectedOrders", "[]"))
    if selected_orders:
        # 创建一个BytesIO对象来保存ZIP文件，我理解是前后端导出pdf只能响应一次，所以实现不了一次导出多个pdf，就将多个pdf合并为一个压缩包
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for container_number in selected_orders:
                pdf_response = export_do_branch(container_number)
                zip_file.writestr(f"DO_{container_number}.pdf", pdf_response.content)

        # 设置HTTP响应，格式为压缩包
        response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = 'attachment; filename="orders.zip"'
        zip_buffer.close()
        return response
    else:
        selected_orders = request.POST.get("container_number")
        pdf_response = export_do_branch(selected_orders)
        return pdf_response


def export_do_branch(container_number) -> Any:
    order = Order.objects.select_related(
        "container_number", "retrieval_id", "warehouse"
    ).get(container_number__container_number=container_number)
    container = order.container_number
    packing_list = PackingList.objects.filter(
        container_number__container_number=container_number
    )
    pcs, weight = 0, 0
    for pl in packing_list:
        pcs += pl.pcs if pl.pcs else 0
        weight += pl.total_weight_lbs if pl.total_weight_lbs else 0
    retrieval = order.retrieval_id
    vessel = order.vessel_id
    warehouse = order.warehouse
    detailedAddress = ""
    file_path = "warehouse/utils/fba_fulfillment_center.yaml"
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            fba_data = yaml.safe_load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"YAML文件未找到，请检查路径：{file_path}")
    except yaml.YAMLError as e:
        raise ValueError(f"YAML文件格式错误：{str(e)}")

    if retrieval and hasattr(retrieval, "retrieval_destination_area"):
        dest_area = retrieval.retrieval_destination_area.strip()  # 去除空格，避免匹配失败

        if dest_area in fba_data:
            fba_info = fba_data[dest_area]
            detailedAddress = f"{dest_area} {fba_info['location']}, {fba_info['city']}, {fba_info['state']} {fba_info['zipcode']}"
        else:
            detailedAddress = dest_area
    context = {
        "order": order,
        "retrieval": retrieval,
        "vessel": vessel,
        "detailedAddress": detailedAddress,
        "container": container,
        "warehouse": warehouse,
        "pcs": pcs,
        "weight": weight,
    }
    template_path = "export_file/do_template.html"
    template = get_template(template_path)
    html = template.render(context)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="DO_{container_number}.pdf"'
    )
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        raise ValueError(
            "Error during PDF generation: %s" % pisa_status.err,
            content_type="text/plain",
        )
    return response


def export_invoice(
    request: HttpRequest,
) -> tuple[HttpResponse, str, BytesIO, dict[Any, Any]]:
    customer = request.POST.get("customer")
    chinese_char = False if customer.isascii() else True
    invoice_statement_id = request.POST.get("invoice_statement_id")
    invoice_terms = request.POST.get("invoice_terms")
    invoice_date = request.POST.get("invoice_date")
    due_date = request.POST.get("due_date")
    container_number = request.POST.getlist("container_number")
    invoice_number = request.POST.getlist("invoice_number")
    rate = [float(r) for r in request.POST.getlist("rate")]
    amount = [float(r) for r in request.POST.getlist("amount")]
    total_amount = sum(amount)
    cnt = list(range(1, len(container_number) + 1))
    invoice_details = zip(cnt, container_number, invoice_number, rate, amount)

    context = {
        "customer": customer,
        "chinese_char": chinese_char,
        "invoice_details": invoice_details,
        "total_amount": total_amount,
        "invoice_statement_id": invoice_statement_id,
        "invoice_terms": invoice_terms,
        "invoice_date": invoice_date,
        "due_date": due_date,
        "container_number": container_number,
        "ACCT_ACH_ROUTING_NUMBER": ACCT_ACH_ROUTING_NUMBER,
        "ACCT_BANK_NAME": ACCT_BANK_NAME,
        "ACCT_BENEFICIARY_ACCOUNT": ACCT_BENEFICIARY_ACCOUNT,
        "ACCT_BENEFICIARY_ADDRESS": ACCT_BENEFICIARY_ADDRESS,
        "ACCT_BENEFICIARY_NAME": ACCT_BENEFICIARY_NAME,
        "ACCT_SWIFT_CODE": ACCT_SWIFT_CODE,
    }

    template_path = "export_file/invoice_template.html"
    template = get_template(template_path)
    html = template.render(context)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="invoice_{invoice_statement_id}_from_ZEM_ELITELINK LOGISTICS_INC.pdf"'
    )
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisa_status.err:
        raise ValueError(
            "Error during PDF generation: %s" % pisa_status.err,
            content_type="text/plain",
        )

    pdf_file = io.BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=pdf_file, link_callback=link_callback)
    if pisa_status.err:
        return HttpResponse(
            "Error during PDF generation: %s" % pisa_status.err,
            content_type="text/plain",
        )
    pdf_file.seek(0)
    return (
        response,
        f"invoice_{invoice_statement_id}_from_ZEM_ELITELINK LOGISTICS_INC.pdf",
        pdf_file,
        context,
    )


def link_callback(uri: Any, rel: Any) -> Any:
    if uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
        if not os.path.isfile(path):
            # Try to find the file using staticfiles finders
            result = finders.find(uri.replace(settings.STATIC_URL, ""))
            if not result:
                raise Exception(f"Static file not found: {uri}")
            if isinstance(result, (list, tuple)):
                result = result[0]
            path = result
        return path
    return uri
