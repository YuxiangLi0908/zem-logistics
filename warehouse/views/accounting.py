import ast
import io
import json
import math
import os
import re
import zipfile
from collections import defaultdict
from datetime import date, datetime, timedelta, time as datetime_time
from io import BytesIO
from itertools import chain, groupby
from operator import attrgetter

from asgiref.sync import sync_to_async, async_to_sync

from django.db import transaction
from django.db.models.fields.json import KeyTextTransform
from django.core.paginator import Paginator

from typing import Any

import openpyxl
import openpyxl.workbook
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import pandas as pd
import pytz
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.postgres.aggregates import ArrayAgg, StringAgg
from django.core.paginator import Paginator
from django.db import models, transaction
from django.db.models import (
    BooleanField,
    Case,
    CharField,
    Count,
    Exists,
    F,
    FloatField,
    IntegerField,
    OuterRef,
    Prefetch,
    Subquery,
    Sum,
    Value,
    When,
)
from django.db.models.fields.json import KeyTextTransform
from django.db.models.functions import Cast, Coalesce
from django.db.models.query import QuerySet
from django.http import HttpRequest, HttpResponse, HttpResponseForbidden
from django.shortcuts import render
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View
from office365.runtime.auth.user_credential import UserCredential
from office365.runtime.client_request_exception import ClientRequestException
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.sharing.links.kind import SharingLinkKind
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from simple_history.utils import bulk_create_with_history, bulk_update_with_history
from sqlalchemy.util import await_only

from warehouse.forms.order_form import OrderForm
from warehouse.models.container import Container
from warehouse.models.customer import Customer
from warehouse.models.fee_detail import FeeDetail
from warehouse.models.fleet_shipment_pallet import FleetShipmentPallet
from warehouse.models.invoice import (
    Invoice,
    InvoiceItem,
    InvoiceStatement,
    InvoiceStatus,
)
from warehouse.models.invoice_details import (
    InvoiceDelivery,
    InvoicePreport,
    InvoiceWarehouse,
)
from warehouse.models.order import Order
from warehouse.models.packing_list import PackingList
from warehouse.models.pallet import Pallet
from warehouse.models.quotation_master import QuotationMaster
from warehouse.models.transaction import Transaction
from warehouse.utils.constants import (
    ACCT_ACH_ROUTING_NUMBER,
    ACCT_BANK_NAME,
    ACCT_BENEFICIARY_ACCOUNT,
    ACCT_BENEFICIARY_ADDRESS,
    ACCT_BENEFICIARY_NAME,
    ACCT_SWIFT_CODE,
    APP_ENV,
    CARRIER_FLEET,
    CONTAINER_PICKUP_CARRIER,
    SP_DOC_LIB,
    SP_CLIENT_ID,
    SP_DOC_LIB,
    SP_PRIVATE_KEY,
    SP_SCOPE,
    SP_TENANT,
    SP_THUMBPRINT,
    SP_URL,
    SYSTEM_FOLDER,
)
from warehouse.views.export_file import export_invoice


@method_decorator(login_required(login_url="login"), name="dispatch")
class Accounting(View):
    template_pallet_data = "accounting/pallet_data.html"
    template_pl_data = "accounting/pl_data.html"
    template_invoice_management = "accounting/invoice_management.html"
    template_invoice_statement = "accounting/invoice_statement.html"
    template_invoice_container = "accounting/invoice_container.html"
    template_invoice_container_edit = "accounting/invoice_container_edit.html"
    template_invoice_preport = "accounting/invoice_preport.html"
    template_invoice_preport_edit = "accounting/invoice_preport_edit.html"
    template_invoice_warehouse = "accounting/invoice_warehouse.html"
    template_invoice_warehouse_edit = "accounting/invoice_warehouse_edit.html"
    template_invoice_delivery = "accounting/invoice_delivery.html"
    template_invoice_delievery_edit = "accounting/invoice_delivery_edit.html"
    template_invoice_delievery_public_edit = (
        "accounting/invoice_delivery_public_edit.html"
    )
    template_invoice_delievery_other_edit = (
        "accounting/invoice_delivery_other_edit.html"
    )
    template_invoice_confirm = "accounting/invoice_confirm.html"
    template_invoice_payable_confirm = "accounting/invoice_payable_confirm.html"
    template_invoice_confirm_edit = "accounting/invoice_confirm_edit.html"
    template_invoice_direct = "accounting/invoice_direct.html"
    template_invoice_direct_edit = "accounting/invoice_direct_edit.html"
    template_invoice_combina = "accounting/invoice_combina.html"
    template_invoice_combina_edit = "accounting/invoice_combina_edit.html"
    template_invoice_search = "accounting/invoice_search.html"
    template_invoice_payable = "accounting/invoice_payable.html"
    template_invoice_payable_edit = "accounting/invoice_payable_edit.html"
    template_invoice_payable_direct_edit = "accounting/invoice_payable_direct_edit.html"
    allowed_group = "accounting"
    warehouse_options = {
        "": "",
        "NJ-07001": "NJ-07001",
        "NJ-08817": "NJ-08817",
        "SAV-31326": "SAV-31326",
        "LA-91761": "LA-91761",
        "MO-62025": "MO-62025",
        "TX-77503": "TX-77503",
        "直送": "直送",
    }

    def get(self, request: HttpRequest) -> HttpResponse:
        # if not self._validate_user_group(request.user):
        #     return HttpResponseForbidden("You are not authenticated to access this page!")
        step = request.GET.get("step", None)
        if step == "pallet_data":
            template, context = self.handle_pallet_data_get()
            return render(request, template, context)
        elif step == "pl_data":
            template, context = self.handle_pl_data_get()
            return render(request, template, context)
        elif step == "invoice_direct":
            if self._validate_user_invoice_direct(request.user):
                template, context = self.handle_invoice_direct_get(request)
                return render(request, template, context)
            else:
                return HttpResponseForbidden(
                    "You are not authenticated to access this page!"
                )
        elif step == "invoice_combina":
            if self._validate_user_invoice_combina(request.user):
                template, context = self.handle_invoice_combina_get(request)
                return render(request, template, context)
            else:
                return HttpResponseForbidden(
                    "You are not authenticated to access this page!"
                )
        elif step == "invoice_preport":  # 提拆柜账单录入
            if self._validate_user_invoice_preport(request.user):
                template, context = self.handle_invoice_preport_get(request)
                return render(request, template, context)
            else:
                return HttpResponseForbidden(
                    "You are not authenticated to access this page!"
                )
        elif step == "invoice_warehouse":  # 库内账单录入
            if self._validate_user_invoice_warehouse(request.user):
                template, context = self.handle_invoice_warehouse_get(request)
                return render(request, template, context)
            else:
                return HttpResponseForbidden(
                    "You are not authenticated to access this page!"
                )
        elif step == "invoice_delivery":
            if self._validate_user_invoice_delivery(request.user):
                template, context = self.handle_invoice_delivery_get(request)
                return render(request, template, context)
            else:
                return HttpResponseForbidden(
                    "You are not authenticated to access this page!"
                )
        elif step == "invoice_payable":
            template, context = self.handle_invoice_payable_get(request)
            return render(request, template, context)

        elif step == "invoice_confirm":
            if self._validate_user_invoice_confirm(request.user):
                template, context = self.handle_invoice_confirm_get(request)
                return render(request, template, context)
            else:
                return HttpResponseForbidden(
                    "You are not authenticated to access this page!"
                )
        elif step == "invoice":
            template, context = self.handle_invoice_get()
            return render(request, template, context)
        elif step == "container_invoice":
            container_number = request.GET.get("container_number")
            template, context = self.handle_container_invoice_get(container_number)
            return render(request, template, context)
        elif step == "container_direct":
            template, context = self.handle_container_invoice_direct_get(request)
            return render(request, template, context)
        elif step == "container_combina":
            template, context = self.handle_container_invoice_combina_get(request)
            return render(request, template, context)
        elif step == "container_preport":
            template, context = self.handle_container_invoice_preport_get(request)
            return render(request, template, context)
        elif step == "container_warehouse":
            template, context = self.handle_container_invoice_warehouse_get(request)
            return render(request, template, context)
        elif step == "container_delivery":
            template, context = self.handle_container_invoice_delivery_get(request)
            return render(request, template, context)
        elif step == "container_confirm":
            template, context = self.handle_container_invoice_confirm_get(request)
            return render(request, template, context)
        elif step == "container_payable":
            template, context = self.handle_container_invoice_payable_get(
                request, False, False
            )
            return render(request, template, context)
        elif step == "container_payable_confirm":
            template, context = self.handle_container_invoice_payable_get(
                request, False, True
            )
            return render(request, template, context)
        elif step == "container_invoice_edit":
            container_number = request.GET.get("container_number")
            template, context = self.handle_container_invoice_edit_get(container_number)
            return render(request, template, context)
        elif step == "container_invoice_delete":
            template, context = self.handle_container_invoice_delete_get(request)
            return render(request, template, context)
        elif step == "invoice_search":
            template, context = self.handle_invoice_search_get(request)
            return render(request, template, context)
        else:
            raise ValueError(f"unknow request {step}")

    def post(self, request: HttpRequest) -> HttpResponse:
        # if not self._validate_user_group(request.user):
        #     return HttpResponseForbidden("You are not authenticated to access this page!")
        step = request.POST.get("step", None)
        if step == "pallet_data_search":
            start_date = request.POST.get("start_date")
            end_date = request.POST.get("end_date")
            template, context = self.handle_pallet_data_get(start_date, end_date)
            return render(request, template, context)
        elif step == "pl_data_search":
            start_date = request.POST.get("start_date")
            end_date = request.POST.get("end_date")
            container_number = request.POST.get("container_number")
            template, context = self.handle_pl_data_get(
                start_date, end_date, container_number
            )
            return render(request, template, context)
        elif step == "pallet_data_export":
            return self.handle_pallet_data_export_post(request)
        elif step == "pl_data_export":
            return self.handle_pl_data_export_post(request)
        elif step == "invoice_order_search":
            template, context = self.handle_invoice_order_search_post(request, "old")
            return render(request, template, context)
        elif step == "invoice_order_direct":
            template, context = self.handle_invoice_order_search_post(request, "direct")
            return render(request, template, context)
        elif step == "invoice_order_combina":
            template, context = self.handle_invoice_order_search_post(
                request, "combina"
            )
            return render(request, template, context)
        elif step == "invoice_order_preport":
            template, context = self.handle_invoice_order_search_post(
                request, "preport"
            )
            return render(request, template, context)
        elif step == "invoice_order_warehouse":
            template, context = self.handle_invoice_order_search_post(
                request, "warehouse"
            )
            return render(request, template, context)
        elif step == "invoice_order_delivery":
            template, context = self.handle_invoice_order_search_post(
                request, "delivery"
            )
            return render(request, template, context)
        elif step == "invoice_order_payable":
            template, context = self.handle_invoice_order_search_post(
                request, "payable"
            )
            return render(request, template, context)
        elif step == "invoice_order_confirm":
            template, context = self.handle_invoice_order_search_post(
                request, "confirm"
            )
            return render(request, template, context)
        elif step == "invoice_search":
            template, context = self.handle_invoice_order_search_post(request, "search")
            return render(request, template, context)
        elif step == "invoice_order_select":
            return self.handle_invoice_order_select_post(request)
        elif step == "export_invoice":
            return self.handle_export_invoice_post(request)
        elif step == "create_container_invoice":
            return self.handle_create_container_invoice_post(request)
        elif step == "container_invoice_edit":
            return self.handle_container_invoice_edit_post(request)
        elif step == "direct_save":
            template, context = self.handle_invoice_direct_save_post(request)
            return render(request, template, context)
        elif step == "preport_save":
            template, context = self.handle_invoice_preport_save_post(request)
            return render(request, template, context)
        elif step == "warehouse_save":
            template, context = self.handle_invoice_warehouse_save_post(request)
            return render(request, template, context)
        elif step == "payable_save":
            template, context = self.handle_invoice_payable_save_post(request)
            return render(request, template, context)
        elif step == "add_delivery_type":
            template, context = self.handle_invoice_delivery_type_save(request)
            return render(request, template, context)
        elif step == "update_delivery_invoice":
            template, context = self.handle_invoice_delivery_save(request)
            return render(request, template, context)
        elif step == "swicth_delivery_invoice":
            template, context = self.handle_swicth_delivery_save(request)
            return render(request, template, context)
        elif step == "confirm_save":
            template, context = self.handle_invoice_confirm_save(request)
            return render(request, template, context)
        elif step == "dismiss":
            template, context = self.handle_invoice_dismiss_save(request)
            return render(request, template, context)
        elif step == "redirect":
            template, context = self.handle_invoice_redirect_post(request)
            return render(request, template, context)
        elif step == "invoice_order_batch_export":
            return self.handle_invoice_order_batch_export(request)
        elif step == "invoice_order_delivered":
            return self.handle_invoice_order_batch_delivered(request)
        elif step == "invoice_order_reject":
            return self.handle_invoice_order_batch_reject(request)
        elif step == "migrate_payable_receivable_amount":
            template, context = self.migrate_payable_to_receivable()
            return render(request, template, context)
        elif step == "migrate_status":
            return self.migrate_status()
        elif step == "confirm_combina_save":
            template, context = self.handle_invoice_confirm_combina_save(request)
            return render(request, template, context)
        elif step == "adjustBalance":
            template, context = self.handle_adjust_balance_save(request)
            return render(request, template, context)
        elif step == "invoice_payable_carrier_export":
            return self.handle_carrier_invoice_export(request)
        elif step == "payable_confirm_phase":
            template, context = self.handle_invoice_payable_confirm_phase(request)
            return render(request, template, context)
        elif step == "export_carrier_payable":
            return self.handle_export_carrier_payable_phase(request)
        elif step == "check_unreferenced_delivery":
            template, context = self.handle_check_unreferenced_delivery(request)
            return render(request, template, context)
        else:
            raise ValueError(f"unknow request {step}")

    def handle_export_carrier_payable_phase(self, request) -> HttpRequest:
        confirm_phase = request.POST.get("confirm_phase")
        start_date_export = request.POST.get("start_date_export")
        end_date_export = request.POST.get("end_date_export")
        select_carrier = request.POST.get("select_carrier")
        if select_carrier == "ARM":
            select_carrier = "大方广"
        if not start_date_export or not select_carrier:
            raise ValueError("请选择月份和供应商")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{select_carrier}账单"
        # 查找该月份，该供应商的所有费用
        # 先查账单时间满足月份的，就是实际提柜时间+1个月
        if confirm_phase == "delivery":
            delivery_data = self._get_deliverys(
                start_date_export, end_date_export, select_carrier
            )
            # 设置表头
            headers = [
                "Carrier",
                "Pickup Number",
                "ISA",
                "柜号",
                "目的地",
                "板数",
                "价格",
                "单板成本",
                "车次总板数",
                "车次总成本",
            ]
            ws.append(headers)

            # 填充数据并处理合并单元格
            for fleet in delivery_data:
                carrier = fleet["carrier"]
                fleet_rows = fleet["fleets"]
                fleet_start_row = ws.max_row + 1

                for pickup_number, pickup_data in fleet_rows.items():
                    isa_rows = pickup_data["appointments"]
                    pickup_start_row = ws.max_row + 1

                    for isa, isa_data in isa_rows.items():
                        orders = isa_data["orders"]
                        isa_start_row = ws.max_row + 1

                        for order in orders:
                            row_data = [
                                carrier,
                                pickup_number,
                                isa if isa else "",
                                order["container_num"],
                                order["pallet_destination"],
                                order["cn_total_pallet"],
                                order["cn_total_expense"],
                                order["cn_per_expense"],
                                pickup_data["ISA_total_pallets"],
                                f"${pickup_data['ISA_total_expense']}",
                            ]
                            ws.append(row_data)

                        # 合并ISA列
                        if len(orders) > 1:
                            ws.merge_cells(
                                start_row=isa_start_row,
                                end_row=isa_start_row + len(orders) - 1,
                                start_column=3,
                                end_column=3,
                            )

                    pickup_end_row = ws.max_row
                    if pickup_end_row > pickup_start_row:
                        ws.merge_cells(
                            start_row=pickup_start_row,
                            end_row=pickup_end_row,
                            start_column=2,
                            end_column=2,
                        )

                        # 合并车次总板数和车次总成本列（按pickup级别合并）
                        ws.merge_cells(
                            start_row=pickup_start_row,
                            end_row=pickup_end_row,
                            start_column=9,
                            end_column=9,
                        )
                        ws.merge_cells(
                            start_row=pickup_start_row,
                            end_row=pickup_end_row,
                            start_column=10,
                            end_column=10,
                        )
                fleet_end_row = ws.max_row
                if fleet_end_row > fleet_start_row:
                    ws.merge_cells(
                        start_row=fleet_start_row,
                        end_row=fleet_end_row,
                        start_column=1,
                        end_column=1,
                    )
        else:
            if confirm_phase == "warehouse":
                # 筛选库内账单确认的，时间符合要求的
                order_list = Order.objects.filter(
                    retrieval_id__actual_retrieval_timestamp__gte=start_date_export,
                    retrieval_id__actual_retrieval_timestamp__lte=end_date_export,
                    payable_status__isnull=False,
                    payable_status__payable_status__has_key="warehouse",
                    payable_status__payable_status__warehouse="confirmed",
                )

                orders = []
                for order in order_list:
                    invoice = Invoice.objects.select_related(
                        "customer", "container_number"
                    ).get(container_number__container_number=order.container_number)
                    invoice_warehouse = InvoiceWarehouse.objects.get(
                        invoice_number__invoice_number=invoice.invoice_number,
                        invoice_type="payable",
                    )
                    if invoice_warehouse.carrier == select_carrier:
                        orders.append(order)
            elif confirm_phase == "pickup":
                # 否则就看payable_surcharge的preport_carrier里面能看到
                order_list = Order.objects.filter(
                    retrieval_id__actual_retrieval_timestamp__gte=start_date_export,
                    retrieval_id__actual_retrieval_timestamp__lte=end_date_export,
                    payable_status__isnull=False,
                    payable_status__payable_status__has_key="pickup",
                    payable_status__payable_status__pickup="confirmed",
                )
                orders = []
                for order in order_list:
                    if order.retrieval_id.retrieval_carrier == select_carrier:
                        orders.append(order)
            if len(orders) == 0:
                raise ValueError("未查询到符合条件的订单")

            all_fee_types = set()
            for order in orders:
                container_number = order.container_number.container_number
                invoice = Invoice.objects.select_related(
                    "customer", "container_number"
                ).get(container_number__container_number=container_number)
                invoice_warehouse = InvoiceWarehouse.objects.get(
                    invoice_number__invoice_number=invoice.invoice_number,
                    invoice_type="payable",
                )
                invoice_preport = InvoicePreport.objects.get(
                    invoice_number__invoice_number=invoice.invoice_number,
                    invoice_type="payable",
                )
                preport_other_fee = invoice_preport.other_fees
                warehouse_other_fee = invoice_warehouse.other_fees
                if confirm_phase == "warehouse":
                    all_fee_types.add("拆柜费")
                    if warehouse_other_fee:
                        all_fee_types.update(warehouse_other_fee.keys())
                elif confirm_phase == "pickup":
                    all_fee_types.update(["总费用"])
                    all_fee_types.add("基本费用")
                    all_fee_types.add("超重费")
                    all_fee_types.add("车架费")

                    # 其他自定义费用
                    if (
                        "other_fee" in preport_other_fee
                        and preport_other_fee["other_fee"]
                    ):
                        all_fee_types.update(preport_other_fee["other_fee"].keys())

            sorted_fee_types = sorted(all_fee_types)

            headers = ["柜号"] + sorted_fee_types
            ws.append(headers)

            # 填充数据
            for order in orders:
                container_number = order.container_number.container_number
                invoice = Invoice.objects.select_related(
                    "customer", "container_number"
                ).get(container_number__container_number=container_number)
                payable_surcharge = invoice.payable_surcharge
                total_amount = float(invoice.payable_total_amount or 0)
                palletization = float(invoice.payable_palletization or 0)

                # 初始化行数据，所有费用初始为0或空
                row_data = {fee: "" for fee in sorted_fee_types}
                row_data["柜号"] = container_number

                # 填充特定供应商的费用
                if select_carrier in ["BBR", "KNO"]:
                    row_data["拆柜费"] = invoice.payable_palletization or 0
                else:
                    row_data["总费用"] = total_amount - palletization
                    row_data["基本费用"] = invoice.payable_basic or 0
                    row_data["超重费"] = invoice.payable_overweight or 0
                    row_data["车架费"] = invoice.payable_chassis or 0

                    # 其他自定义费用
                    if (
                        "other_fee" in payable_surcharge
                        and payable_surcharge["other_fee"]
                    ):
                        for fee_name, fee_value in payable_surcharge[
                            "other_fee"
                        ].items():
                            row_data[fee_name] = fee_value
                if select_carrier in ["BBR", "KNO"]:
                    ordered_row = [row_data["柜号"], row_data["拆柜费"]]
                else:
                    fixed_columns = ["柜号", "总费用", "基本费用", "超重费", "车架费"]
                    dynamic_columns = [
                        col for col in row_data.keys() if col not in fixed_columns
                    ]
                    final_columns = fixed_columns + sorted(dynamic_columns)

                    ordered_row = [row_data.get(col, "") for col in final_columns]
                ws.append(ordered_row)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{select_carrier}_账单.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"

        wb.save(response)
        return response

    def _get_deliverys(self, start_date_export, end_date_export, select_carrier) -> Any:
        criteria = models.Q()
        criteria &= models.Q(
            models.Q(vessel_id__vessel_etd__gte=start_date_export),
            models.Q(vessel_id__vessel_etd__lte=end_date_export),
        )
        container_numbers = Order.objects.filter(criteria).values_list(
            "container_number", flat=True
        )
        delivery_po_ids = (
            Pallet.objects.filter(container_number__in=container_numbers)
            .values_list("PO_ID", flat=True)
            .distinct()
        )
        delivery_pending_orders = (
            FleetShipmentPallet.objects.select_related(
                "fleet_number",  # 确保预取关联的Fleet对象
                "shipment_batch_number",
                "container_number",
            )
            .filter(
                expense__isnull=False,
                PO_ID__in=delivery_po_ids,
            )
            .annotate(
                appointment_id=F("shipment_batch_number__appointment_id"),
                container_num=F("container_number__container_number"),
                pallet_destination=Subquery(
                    Pallet.objects.filter(
                        PO_ID=OuterRef("PO_ID"),
                        shipment_batch_number=OuterRef("shipment_batch_number"),
                    ).values("destination")[:1]
                ),
            )
            .order_by("fleet_number", "pickup_number", "container_num")
        )

        deliverys = {}
        for order in delivery_pending_orders:
            fleet_id = order.fleet_number_id
            if fleet_id not in deliverys:
                # 获取carrier信息
                carrier = order.fleet_number.carrier if order.fleet_number else None
                if carrier != select_carrier:
                    continue
                deliverys[fleet_id] = {
                    "fleets": {},
                    "total_pallets": 0,
                    "total_expense": 0,
                    "total_rows": 0,
                    "carrier": carrier,
                }

            if order.pickup_number not in deliverys[fleet_id]["fleets"]:
                deliverys[fleet_id]["fleets"][order.pickup_number] = {
                    "appointments": {},
                    "ISA_total_pallets": 0,
                    "ISA_total_expense": 0,
                    "total_rows": 0,
                }

            if (
                order.appointment_id
                not in deliverys[fleet_id]["fleets"][order.pickup_number][
                    "appointments"
                ]
            ):
                deliverys[fleet_id]["fleets"][order.pickup_number]["appointments"][
                    order.appointment_id
                ] = {"orders": [], "total_pallets": 0, "total_expense": 0, "rowspan": 0}

            # 计算单板成本并保留2位小数
            per_expense = (
                round(order.expense / int(order.total_pallet), 2)
                if order.total_pallet
                else 0
            )

            # 为每个order添加统计字段
            order_data = {
                "object": order,
                "cn_total_pallet": int(order.total_pallet) if order.total_pallet else 0,
                "cn_total_expense": order.expense or 0,
                "cn_per_expense": per_expense,
                "container_num": order.container_num,
                "pallet_destination": order.pallet_destination,
                "carrier": deliverys[fleet_id]["carrier"],
            }

            deliverys[fleet_id]["fleets"][order.pickup_number]["appointments"][
                order.appointment_id
            ]["orders"].append(order_data)

            deliverys[fleet_id]["fleets"][order.pickup_number]["appointments"][
                order.appointment_id
            ]["total_pallets"] += int(order_data["cn_total_pallet"])
            deliverys[fleet_id]["fleets"][order.pickup_number]["appointments"][
                order.appointment_id
            ]["total_expense"] += order_data["cn_total_expense"]
            deliverys[fleet_id]["fleets"][order.pickup_number]["appointments"][
                order.appointment_id
            ]["rowspan"] += 1

            deliverys[fleet_id]["fleets"][order.pickup_number][
                "ISA_total_pallets"
            ] += int(order_data["cn_total_pallet"])
            deliverys[fleet_id]["fleets"][order.pickup_number][
                "ISA_total_expense"
            ] += order_data["cn_total_expense"]
            deliverys[fleet_id]["fleets"][order.pickup_number]["total_rows"] += 1

            deliverys[fleet_id]["total_pallets"] += int(order_data["cn_total_pallet"])
            deliverys[fleet_id]["total_expense"] += order_data["cn_total_expense"]
            deliverys[fleet_id]["total_rows"] += 1

        # 转换数据结构
        delivery_pending_orders = [
            {
                "fleets": fleet_data["fleets"],
                "total_pallets": int(fleet_data["total_pallets"]),
                "total_expense": fleet_data["total_expense"],
                "carrier": fleet_data["carrier"],
            }
            for fleet_data in deliverys.values()
        ]
        return delivery_pending_orders

    def migrate_payable_to_receivable(self) -> tuple[Any, Any]:
        # with transaction.atomic():
        batch_size = 100
        invoices = Invoice.objects.filter(payable_basic__isnull=False)
        paginator = Paginator(invoices, batch_size)
        for page_num in range(1, paginator.num_pages + 1):
            with transaction.atomic():
                batch_invoices = paginator.page(page_num).object_list

                for invoice in batch_invoices:
                    # for invoice in invoices:
                    preport, created = InvoicePreport.objects.get_or_create(
                        invoice_number=invoice,
                        invoice_type="payable",
                        defaults={
                            "pickup": invoice.payable_basic,
                            "over_weight": invoice.payable_overweight,
                            "chassis": invoice.payable_chassis,
                            "amount": float(invoice.payable_total_amount)
                            - (
                                float(invoice.payable_palletization)
                                if invoice.payable_palletization is not None
                                else 0
                            ),
                            "other_fees": {
                                "other_fee": (
                                    invoice.payable_surcharge.get("other_fee")
                                    if invoice.payable_surcharge
                                    and "other_fee" in invoice.payable_surcharge
                                    else None
                                ),
                                "chassis_comment": (
                                    invoice.payable_surcharge.get("chassis_comment")
                                    if invoice.payable_surcharge
                                    and "chassis_comment" in invoice.payable_surcharge
                                    and invoice.payable_surcharge["chassis_comment"]
                                    is not None
                                    else None
                                ),
                            },
                        },
                    )
                    if not created:
                        preport.pickup = invoice.payable_basic
                        preport.over_weight = invoice.payable_overweight
                        preport.chassis = invoice.payable_chassis
                        preport.amount = float(invoice.payable_total_amount) - (
                            float(invoice.payable_palletization)
                            if invoice.payable_palletization is not None
                            else 0
                        )
                        other_fees = preport.other_fees if preport.other_fees else {}
                        if (
                            invoice.payable_surcharge
                            and "other_fee" in invoice.payable_surcharge
                        ):
                            other_fees["other_fee"] = invoice.payable_surcharge[
                                "other_fee"
                            ]
                        if (
                            invoice.payable_surcharge
                            and "chassis_comment" in invoice.payable_surcharge
                            and invoice.payable_surcharge["chassis_comment"] is not None
                        ):
                            other_fees["chassis_comment"] = invoice.payable_surcharge[
                                "chassis_comment"
                            ]
                        preport.other_fees = other_fees
                        preport.save()

                    wh, created_wh = InvoiceWarehouse.objects.get_or_create(
                        invoice_number=invoice,
                        invoice_type="payable",
                        defaults={
                            "palletization_fee": invoice.payable_palletization,
                            #'arrive_fee': invoice.arrive_fee,
                            "amount": invoice.payable_palletization,
                            "carrier": (
                                invoice.payable_surcharge.get("palletization_carrier")
                                if invoice.payable_surcharge
                                and "palletization_carrier" in invoice.payable_surcharge
                                else None
                            ),
                        },
                    )
                    if not created_wh:
                        wh.palletization_fee = invoice.payable_palletization
                        wh.amount = invoice.payable_palletization
                        if (
                            invoice.payable_surcharge
                            and "palletization_carrier" in invoice.payable_surcharge
                        ):
                            wh.carrier = invoice.payable_surcharge.get(
                                "palletization_carrier"
                            )

                        wh.save()
        context = {}
        return self.template_invoice_preport, context

    def handle_check_unreferenced_delivery(self, request) -> tuple[Any, Any]:
        unreferenced_invoices = InvoiceDelivery.objects.filter(
            pallet_delivery__isnull=True
        )
        # 先查看数量确认
        deleted_count, _ = unreferenced_invoices.delete()
        context = {}
        return self.template_invoice_delivery, context

    def get_special_stages(self, main_stage) -> tuple[str, str]:
        if main_stage == "delivery":
            return "warehouse_completed", "warehouse_completed"
        elif main_stage in ("tobeconfirmed", "confirmed"):
            return "delivery_completed", "delivery_completed"
        else:
            return "pending", "pending"  # 默认返回原状态

    def migrate_status(self) -> HttpRequest:
        start_date = date(2025, 9, 1)
        end_date = date(2025, 9, 30)

        # 先筛选符合 ETA 条件的 orders
        valid_orders = Order.objects.filter(
            vessel_id__vessel_eta__range=(start_date, end_date)
        ).select_related("vessel_id")

        # 再通过 Prefetch 预加载 container 的关联 order
        packinglists = (
            PackingList.objects.filter(delivery_type="other")
            .exclude(fba_id__icontains="FBA")
            .select_related("container_number")  # 单个 container 外键可 select_related
            .prefetch_related(
                Prefetch("container_number__order", queryset=valid_orders, to_attr="matched_orders")
            )
            .order_by("PO_ID")
        )

        if not packinglists.exists():
            return HttpResponse("未找到符合条件的 PackingList", status=404)

        # === 组装导出数据 ===
        data = []
        for pl in packinglists:
            container = pl.container_number
            vessel = None

            # container.matched_orders 是我们 prefetch 时命名的
            if hasattr(container, "matched_orders") and container.matched_orders:
                # 取第一个匹配的 order
                order = container.matched_orders[0]
                vessel = order.vessel_id

            data.append(
                {
                    "Shipping Mark": pl.shipping_mark,
                    "FBA ID": pl.fba_id,
                    "Destination": pl.destination,
                    "Contact Method": pl.contact_method,
                    "PCS": pl.pcs,
                    "CBM": pl.cbm,
                    "Weight (lbs)": pl.total_weight_lbs,
                    "Container Number": container.container_number if container else "",
                    "Vessel ETA": vessel.vessel_eta.strftime("%Y-%m-%d") if vessel and vessel.vessel_eta else "",
                }
            )

        # === DataFrame + 汇总 ===
        df = pd.DataFrame(data)
        total_cbm = df["CBM"].sum()

        summary_row = {col: "" for col in df.columns}
        summary_row["CBM"] = total_cbm
        df = pd.concat([df, pd.DataFrame([summary_row])], ignore_index=True)

        # === 导出 Excel ===
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="packinglist_vessel_eta_sep.xlsx"'
        df.to_excel(response, index=False)

        return response

    def replace_keywords(data):
        KEYWORD_MAPPING = {
            "等待费": "港口拥堵费",
            "查验": "查验费",
            "车架费": "托架费",
            "车架分离费": "托架提取费",
            "货柜储存费": "货柜放置费",
        }
        if isinstance(data, str):
            for old, new in KEYWORD_MAPPING.items():
                if old in data:
                    return new
            return data
        elif isinstance(data, dict):
            return {
                (KEYWORD_MAPPING.get(key, key) if isinstance(key, str) else key): value
                for key, value in data.items()
            }
        else:
            return data

    def handle_pallet_data_get(
        self, start_date: str = None, end_date: str = None
    ) -> tuple[Any, Any]:
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-30)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date
        pallet_data = (
            Order.objects.select_related(
                "container_number",
                "customer_name",
                "warehouse",
                "offload_id",
                "retrieval_id",
            )
            .filter(
                models.Q(offload_id__offload_required=True)
                & models.Q(offload_id__offload_at__isnull=False)
                & models.Q(retrieval_id__actual_retrieval_timestamp__isnull=False)
                & models.Q(offload_id__offload_at__gte=start_date)
                & models.Q(offload_id__offload_at__lte=end_date)
            )
            .order_by("offload_id__offload_at")
        )
        context = {
            "pallet_data": pallet_data,
            "start_date": start_date,
            "end_date": end_date,
        }
        return self.template_pallet_data, context

    def handle_pl_data_get(
        self,
        start_date: str = None,
        end_date: str = None,
        container_number: str = None,
    ) -> tuple[Any, Any]:
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-30)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date
        criteria = models.Q(
            container_number__order__eta__gte=start_date,
            container_number__order__eta__lte=end_date,
        )
        criteria |= models.Q(
            container_number__order__vessel_id__vessel_eta__gte=start_date,
            container_number__order__vessel_id__vessel_eta__lte=end_date,
        )
        if container_number == "None":
            container_number = None
        if container_number:
            criteria &= models.Q(container_number__container_number=container_number)
        pl_data = (
            PackingList.objects.select_related("container_number")
            .filter(criteria)
            .values(
                "container_number__container_number",
                "destination",
                "delivery_method",
                "cbm",
                "pcs",
                "total_weight_kg",
                "total_weight_lbs",
            )
            .order_by("container_number__container_number", "destination")
        )
        context = {
            "start_date": start_date,
            "end_date": end_date,
            "container_number": container_number,
            "pl_data": pl_data,
        }
        return self.template_pl_data, context

    def handle_invoice_get(
        self,
        start_date: str = None,
        end_date: str = None,
        customer: str = None,
    ) -> tuple[Any, Any]:
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-30)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date
        criteria = models.Q(
            models.Q(
                offload_id__offload_required=True, offload_id__offload_at__isnull=False
            )
            | models.Q(offload_id__offload_required=False)
        )
        criteria &= models.Q(created_at__gte=start_date)
        criteria &= models.Q(created_at__lte=end_date)
        if customer:
            criteria &= models.Q(customer_name__zem_name=customer)
        order = (
            Order.objects.select_related(
                "invoice_id",
                "customer_name",
                "container_number",
                "invoice_id__statement_id",
            )
            .filter(criteria)
            .order_by("created_at")
        )
        order_no_invoice = [o for o in order if o.invoice_id is None]
        order_invoice = [o for o in order if o.invoice_id]
        context = {
            "order_form": OrderForm(),
            "start_date": start_date,
            "end_date": end_date,
            "order": order,
            "customer": customer,
            "order_no_invoice": order_no_invoice,
            "order_invoice": order_invoice,
        }
        return self.template_invoice_management, context

    def handle_invoice_combina_get(
        self,
        request: HttpRequest,
        start_date: str = None,
        end_date: str = None,
        customer: str = None,
        warehouse: str = None,
    ) -> tuple[Any, Any]:
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date
        criteria = models.Q(
            cancel_notification=False,
            vessel_id__vessel_etd__gte=start_date,
            vessel_id__vessel_etd__lte=end_date,
            offload_id__offload_at__isnull=False,
        )
        if customer:
            criteria &= models.Q(customer_name__zem_name=customer)

        # 查找转运组合，没有生成账单的柜子
        order = Order.objects.select_related(
            "customer_name", "container_number", "retrieval_id"
        ).filter(
            criteria,
            models.Q(  # 考虑账单编辑点的是暂存的情况
                **{
                    "receivable_status__invoice_type": "receivable",
                    "receivable_status__stage__in": ["tobeconfirmed"],
                }
            ),
            order_type="转运组合",
            # container_number__account_order_type="转运组合",
        )
        order = self.process_orders_display_status(order, "receivable")
        # 已录入账单
        previous_order = (
            Order.objects.select_related(
                "invoice_id",
                "customer_name",
                "container_number",
                "invoice_id__statement_id",
                "receivable_status",
            )
            .values(
                "invoice_status",
                "container_number__container_number",
                "customer_name__zem_name",
                "created_at",
                "receivable_status",
            )
            .filter(
                criteria,
                order_type="转运组合",
                container_number__account_order_type="转运组合",
                **{
                    "receivable_status__isnull": False,
                    "receivable_status__invoice_type": "receivable",
                    "receivable_status__stage": "confirmed",
                },
            )
        )
        previous_order = self.process_orders_display_status(
            previous_order, "receivable"
        )
        context = {
            "order": order,
            "order_form": OrderForm(),
            "previous_order": previous_order,
            "start_date": start_date,
            "end_date": end_date,
            "customer": customer,
        }
        return self.template_invoice_combina, context

    def handle_invoice_direct_get(
        self,
        request: HttpRequest,
        start_date: str = None,
        end_date: str = None,
        customer: str = None,
        warehouse: str = None,
    ) -> tuple[Any, Any]:
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date
        criteria = models.Q(
            cancel_notification=False,
            vessel_id__vessel_etd__gte=start_date,
            vessel_id__vessel_etd__lte=end_date,
        )
        if customer:
            criteria &= models.Q(customer_name__zem_name=customer)

        status_field = "receivable_status"

        # 查找直送，没有生成账单的柜子
        order = Order.objects.select_related(
            "customer_name", "container_number", "retrieval_id"
        ).filter(
            criteria,
            models.Q(**{f"{status_field}__isnull": True})
            | models.Q(  # 考虑账单编辑点的是暂存的情况
                **{
                    "receivable_status__invoice_type": "receivable",
                    "receivable_status__stage__in": ["unstarted"],
                }
            ),
            order_type="直送",
        )
        order = self.process_orders_display_status(order, "receivable")
        # 已录入账单
        previous_order = (
            Order.objects.select_related(
                "invoice_id",
                "customer_name",
                "container_number",
                "invoice_id__statement_id",
                "receivable_status",
            )
            .values(
                "invoice_status",
                "container_number__container_number",
                "customer_name__zem_name",
                "created_at",
                "receivable_status",
            )
            .filter(
                criteria,
                order_type="直送",
                **{
                    "receivable_status__isnull": False,
                    "receivable_status__invoice_type": "receivable",
                },
            )
            .exclude(**{"receivable_status__stage__in": ["preport", "unstarted"]})
        )
        previous_order = self.process_orders_display_status(
            previous_order, "receivable"
        )
        context = {
            "order": order,
            "order_form": OrderForm(),
            "previous_order": previous_order,
            "start_date": start_date,
            "end_date": end_date,
            "customer": customer,
        }
        return self.template_invoice_direct, context

    def handle_invoice_search_get(
        self,
        request: HttpRequest,
        start_date: str = None,
        end_date: str = None,
        customer: str = None,
        warehouse: str = None,
    ) -> tuple[Any, Any]:
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date
        criteria = models.Q(
            vessel_id__vessel_etd__gte=start_date, vessel_id__vessel_etd__lte=end_date
        )
        if warehouse:
            criteria &= models.Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= models.Q(customer_name__zem_name=customer)
        invoice_type = request.POST.get("invoice_type") or "receivable"
        orders = Order.objects.select_related(
            "customer_name", "container_number", f"{invoice_type}_status"
        ).filter(criteria)
        orders = self.process_orders_display_status(orders, invoice_type)
        context = {
            "order": orders,
            "order_form": OrderForm(),
            "start_date": start_date,
            "end_date": end_date,
            "customer": customer,
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
            "invoice_type_filter": invoice_type,
        }
        return self.template_invoice_search, context

    def process_orders_display_status(self, orders, invoice_type):
        STAGE_MAPPING = {
            "unstarted": "未录入",
            "preport": "提拆柜录入阶段",
            "warehouse": "仓库录入阶段",
            "delivery": "派送录入阶段",
            "tobeconfirmed": "待财务确认",
            "confirmed": "财务已确认",
        }

        SUB_STAGE_MAPPING = {
            "pending": "仓库待处理",
            "warehouse_completed": "仓库已完成",
            "delivery_completed": "派送已完成",
            "warehouse_rejected": "仓库已驳回",
            "delivery_rejected": "派送已驳回",
        }

        is_dict_type = orders and isinstance(orders[0], dict)

        if is_dict_type:
            status_ids = [
                o[f"{invoice_type}_status"]
                for o in orders
                if o.get(f"{invoice_type}_status")
            ]
            status_objects = InvoiceStatus.objects.filter(id__in=status_ids).in_bulk()

        processed_orders = []
        for order in orders:
            if is_dict_type:
                status_id = order.get(f"{invoice_type}_status")
                status_obj = status_objects.get(status_id) if status_id else None

                if status_obj:
                    raw_stage = status_obj.stage
                    raw_public_stage = status_obj.stage_public
                    raw_other_stage = status_obj.stage_other
                    raw_is_rejected = status_obj.is_rejected
                    raw_reject_reason = status_obj.reject_reason
                else:
                    raw_stage = None
                    raw_public_stage = None
                    raw_other_stage = None
                    raw_is_rejected = False
                    raw_reject_reason = ""

                order_data = order.copy()
            else:
                status_obj = getattr(order, f"{invoice_type}_status", None)

                if status_obj:
                    raw_stage = status_obj.stage
                    raw_public_stage = status_obj.stage_public
                    raw_other_stage = status_obj.stage_other
                    raw_is_rejected = status_obj.is_rejected
                    raw_reject_reason = status_obj.reject_reason
                else:
                    raw_stage = None
                    raw_public_stage = None
                    raw_other_stage = None
                    raw_is_rejected = False
                    raw_reject_reason = ""

                order_data = order

            if raw_stage in ["warehouse", "delivery"]:
                stage1 = SUB_STAGE_MAPPING.get(raw_public_stage, str(raw_public_stage))
                stage2 = SUB_STAGE_MAPPING.get(raw_other_stage, str(raw_other_stage))
                display_stage = f"公仓: {stage1}\n私仓: {stage2}"
            else:
                display_stage = (
                    STAGE_MAPPING.get(raw_stage, str(raw_stage))
                    if raw_stage
                    else "未录入任何费用"
                )

            if is_dict_type:
                order_data.update(
                    {
                        "display_stage": display_stage,
                        "display_is_rejected": "已驳回" if raw_is_rejected else "正常",
                        "display_reject_reason": raw_reject_reason or " ",
                    }
                )
            else:
                order.display_stage = display_stage
                order.display_is_rejected = "已驳回" if raw_is_rejected else "正常"
                order.display_reject_reason = raw_reject_reason or " "

            processed_orders.append(order_data)

        return processed_orders

    # 港前账单，待开账单、已开账单、驳回账单
    def handle_invoice_preport_get(
        self,
        request: HttpRequest,
        start_date: str = None,
        end_date: str = None,
        customer: str = None,
        warehouse: str = None,
    ) -> tuple[Any, Any]:
        # 拆送——港前提拆柜费
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date

        criteria = (
            models.Q(cancel_notification=False)
            & (models.Q(order_type="转运") | models.Q(order_type="转运组合"))
            & models.Q(
                vessel_id__vessel_etd__gte=start_date,
                vessel_id__vessel_etd__lte=end_date,
            )
            & models.Q(offload_id__offload_at__isnull=False)
        )
        if warehouse:
            criteria &= models.Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= models.Q(customer_name__zem_name=customer)
        # 查找待录入账单（未操作过的）
        order = Order.objects.select_related(
            "invoice_id",
            "customer_name",
            "container_number",
            "invoice_id__statement_id",
            "offload_id",
        ).filter(
            criteria,
            models.Q(**{"receivable_status__isnull": True})
            | models.Q(  # 考虑账单编辑点的是暂存的情况
                **{
                    "receivable_status__invoice_type": "receivable",
                    "receivable_status__stage": "unstarted",
                }
            ),
        )
        # 查找驳回账单
        order_reject = Order.objects.filter(
            criteria,
            **{
                "receivable_status__invoice_type": "receivable",
                "receivable_status__is_rejected": True,
                "receivable_status__stage": "preport",
            },
        )
        order_reject = self.process_orders_display_status(order_reject, "receivable")
        # 查找待审核账单，给港前组长看
        order_pending = Order.objects.select_related(
            "invoice_id",
            "customer_name",
            "container_number",
            "invoice_id__statement_id",
        ).filter(
            criteria,
            **{
                "receivable_status__invoice_type": "receivable",
                "receivable_status__is_rejected": False,
                "receivable_status__stage": "preport",
            },
        )
        # 查找已录入账单
        previous_order = (
            Order.objects.select_related(
                "invoice_id",
                "customer_name",
                "container_number",
                "invoice_id__statement_id",
                "receivable_status",
            )
            .values(
                "invoice_status",
                "container_number__container_number",
                "customer_name__zem_name",
                "created_at",
                "receivable_status",
            )
            .filter(
                criteria,
                **{
                    "receivable_status__isnull": False,
                    "receivable_status__invoice_type": "receivable",
                },
            )
            .exclude(**{"receivable_status__stage__in": ["preport", "unstarted"]})
        )

        previous_order = self.process_orders_display_status(
            previous_order, "receivable"
        )
        groups = [group.name for group in request.user.groups.all()]
        if request.user.is_staff:
            groups.append("staff")
        context = {
            "order": order,
            "order_form": OrderForm(),
            "order_reject": order_reject,
            "order_pending": order_pending,
            "previous_order": previous_order,
            "start_date": start_date,
            "end_date": end_date,
            "customer": customer,
            "groups": groups,
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
        }
        return self.template_invoice_preport, context

    def handle_invoice_warehouse_get(
        self,
        request: HttpRequest,
        start_date: str = None,
        end_date: str = None,
        customer: str = None,
        warehouse: str = None,
    ) -> tuple[Any, Any]:
        # 库内操作费
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date
        criteria = models.Q(
            (models.Q(order_type="转运") | models.Q(order_type="转运组合")),
            models.Q(vessel_id__vessel_etd__gte=start_date),
            models.Q(vessel_id__vessel_etd__lte=end_date),
        )
        if warehouse and warehouse != "None":
            criteria &= models.Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= models.Q(customer_name__zem_name=customer)
        groups = [group.name for group in request.user.groups.all()]
        # 表示是否可以看见公仓和私仓
        display_mix = False
        if "NJ_mix_account" in groups:
            if warehouse == "NJ-07001":
                display_mix = True
        if "mix_account" in groups:
            display_mix = True

        delivery_type_filter = None
        if display_mix:  # 这个权限的，NJ公仓私仓都能看见
            delivery_type_filter = models.Q()
        else:
            if "warehouse_public" in groups and "warehouse_other" not in groups:
                delivery_type_filter = models.Q(
                    container_number__delivery_type__in=["public", "mixed"]
                )
            elif "warehouse_other" in groups and "warehouse_public" not in groups:
                delivery_type_filter = models.Q(
                    container_number__delivery_type__in=["other", "mixed"]
                )
            else:
                raise ValueError("没有权限")
        # 基础查询
        base_query = Order.objects.select_related(
            "invoice_id",
            "receivable_status",
            "customer_name",
            "container_number",
            "invoice_id__statement_id",
        ).filter(criteria)

        if delivery_type_filter:
            base_query = base_query.filter(delivery_type_filter)

        # 查找未操作过的
        if display_mix:
            order = base_query.filter(
                **{"receivable_status__stage": "warehouse"},
            ).order_by("receivable_status__reject_reason")
        else:
            if "warehouse_public" in groups and "warehouse_other" not in groups:
                # 如果是公仓人员
                order = (
                    base_query.filter(
                        **{"receivable_status__stage": "warehouse"},
                    )
                    .filter(
                        models.Q(**{"receivable_status__stage_public": "pending"})
                        | models.Q(
                            **{"receivable_status__stage_public": "warehouse_rejected"}
                        )
                    )
                    .order_by("receivable_status__reject_reason")
                )
            elif "warehouse_other" in groups and "warehouse_public" not in groups:
                # 如果是私仓人员
                order = (
                    base_query.filter(
                        **{"receivable_status__stage": "warehouse"},
                    )
                    .filter(
                        models.Q(**{"receivable_status__stage_other": "pending"})
                        | models.Q(
                            **{"receivable_status__stage_other": "warehouse_rejected"}
                        )
                    )
                    .order_by("receivable_status__reject_reason")
                )
        order = self.process_orders_display_status(order, "receivable")

        # 查找历史操作过的，状态是warehouse时，对应group的stage为completed，或者状态是库内之后的
        base_condition = ~models.Q(
            **{"receivable_status__stage__in": ["unstarted", "preport"]}
        )
        other_stages = models.Q(
            **{
                "receivable_status__stage__in": [
                    "delivery",
                    "tobeconfirmed",
                    "confirmed",
                ]
            }
        )

        if display_mix:
            warehouse_condition = models.Q(**{"receivable_status__stage": "delivery"})
        else:
            if "warehouse_public" in groups and "warehouse_other" not in groups:
                warehouse_condition = models.Q(
                    **{"receivable_status__stage_public": "warehouse_completed"}
                )
            elif "warehouse_other" in groups and "warehouse_public" not in groups:
                warehouse_condition = models.Q(
                    **{"receivable_status__stage_other": "warehouse_completed"}
                )
        previous_order = base_query.filter(
            base_condition,
            warehouse_condition | other_stages,  # 满足仓库条件或其他阶段
            **{"receivable_status__invoice_type": "receivable"},
        ).select_related(
            "customer_name", "container_number", "receivable_status", "payable_status"
        )
        previous_order = self.process_orders_display_status(
            previous_order, "receivable"
        )
        groups = [group.name for group in request.user.groups.all()]
        context = {
            "order": order,
            "order_form": OrderForm(),
            "start_date": start_date,
            "end_date": end_date,
            "customer": customer,
            "previous_order": previous_order,
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
            "groups": groups,
            "display_mix": display_mix,
        }
        return self.template_invoice_warehouse, context

    def handle_invoice_confirm_get(
        self,
        request: HttpRequest,
        start_date_confirm: str = None,
        end_date_confirm: str = None,
        customer: str = None,
        warehouse: str = None,
    ) -> tuple[Any, Any]:
        current_date = datetime.now().date()
        start_date_confirm = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date_confirm
            else start_date_confirm
        )
        end_date_confirm = (
            current_date.strftime("%Y-%m-%d")
            if not end_date_confirm
            else end_date_confirm
        )
        # 客服录入完毕的账单
        invoice_type = (
            request.POST.get("invoice_type")
            or request.GET.get("invoice_type")
            or "receivable"
        )
        criteria = models.Q()
        if warehouse:
            if warehouse == "直送":
                criteria &= models.Q(order_type="直送")
            else:
                # 这是非直送的筛选，标记一下加上直送的筛选
                criteria &= models.Q(
                    retrieval_id__retrieval_destination_precise=warehouse
                )
        if customer:
            criteria &= models.Q(customer_name__zem_name=customer)
        if invoice_type == "receivable":
            criteria = models.Q(
                models.Q(vessel_id__vessel_etd__gte=start_date_confirm),
                models.Q(vessel_id__vessel_etd__lte=end_date_confirm),
            )

            order = Order.objects.select_related(
                "customer_name", "container_number", "retrieval_id"
            ).filter(
                criteria,
                **{f"{invoice_type}_status__stage": "tobeconfirmed"},
            )

            previous_order = (
                Order.objects.select_related(
                    "customer_name", "container_number", "retrieval_id"
                )
                .values(
                    "container_number__container_number",
                    "customer_name__zem_name",
                    "created_at",
                    "invoice_id__invoice_date",
                    "order_type",
                    "retrieval_id__retrieval_destination_area",
                    "invoice_id__receivable_total_amount",
                    "invoice_id__receivable_preport_amount",
                    "invoice_id__receivable_warehouse_amount",
                    "invoice_id__receivable_delivery_amount",
                    "invoice_id__receivable_direct_amount",
                    "invoice_id__invoice_number",
                    "invoice_id__invoice_link",
                    "invoice_id__statement_id__invoice_statement_id",
                    "invoice_id__statement_id__statement_link",
                    "invoice_id__is_invoice_delivered",
                    "invoice_id__remain_offset",
                )
                .filter(criteria, **{"receivable_status__stage": "confirmed"})
            )
            previous_order = previous_order.annotate(
                total_amount=Case(
                    When(
                        order_type="转运",
                        then=F(f"invoice_id__receivable_preport_amount")
                        + F(f"invoice_id__receivable_warehouse_amount")
                        + F(f"invoice_id__receivable_delivery_amount"),
                    ),
                    When(
                        order_type="转运组合",
                        then=F(f"invoice_id__receivable_preport_amount")
                        + F(f"invoice_id__receivable_warehouse_amount")
                        + F(f"invoice_id__receivable_delivery_amount"),
                    ),
                    When(
                        order_type="直送",
                        then=F(f"invoice_id__receivable_direct_amount"),
                    ),
                    default=Value(0),
                    output_field=IntegerField(),
                )
            )

            previous_order = self.process_orders_display_status(
                previous_order, invoice_type
            )
            # 前端查询客户余额
            existing_customers = Customer.objects.all().order_by("zem_name")
            context = {
                "order": order,
                "previous_order": previous_order,
                "order_form": OrderForm(),
                "start_date_confirm": start_date_confirm,
                "end_date_confirm": end_date_confirm,
                "customer": customer,
                "warehouse_options": self.warehouse_options,
                "warehouse_filter": warehouse,
                "invoice_type_filter": invoice_type,
                "existing_customers": existing_customers,
            }
            return self.template_invoice_confirm, context
        else:
            return self._payable_confirm_get(request)

    def handle_invoice_payable_confirm_phase(self, request):
        # 判断是提拆/库内/派送哪个阶段的
        confirm_phase = request.POST.get("confirm_phase")
        cotainers = request.POST.getlist("containers")
        save_type = request.POST.get("save_type")
        reject_reason = request.POST.get("reject_reason")

        orders = Order.objects.filter(container_number__container_number__in=cotainers)
        container_numbers = [order.container_number for order in orders]
        invoice_statuses = InvoiceStatus.objects.filter(
            container_number__in=container_numbers, invoice_type="payable"
        ).select_related("container_number")

        status_map = {
            status.container_number.container_number: status
            for status in invoice_statuses
        }
        for order in orders:
            invoicestatus = status_map.get(order.container_number.container_number)
            if not invoicestatus:
                continue
            if save_type == "confirmed":
                invoicestatus.payable_status[confirm_phase] = "confirmed"
            elif save_type == "reject":
                invoicestatus.payable_status[confirm_phase] = "rejected"
                invoicestatus.is_rejected = True
                invoicestatus.reject_reason = reject_reason
            invoicestatus.save()
        return self._payable_confirm_get(request)

    def _payable_confirm_get(self, request):
        warehouse = request.POST.get("warehouse")
        customer = request.POST.get("customer")
        current_date = datetime.now().date()
        start_date_confirm = request.POST.get("start_date_confirm")
        end_date_confirm = request.POST.get("end_date_confirm")
        start_date_confirm = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date_confirm
            else start_date_confirm
        )
        end_date_confirm = (
            current_date.strftime("%Y-%m-%d")
            if not end_date_confirm
            else end_date_confirm
        )
        criteria = models.Q()
        if warehouse:
            if warehouse == "直送":
                criteria &= models.Q(order_type="直送")
            else:
                # 这是非直送的筛选，标记一下加上直送的筛选
                criteria &= models.Q(
                    retrieval_id__retrieval_destination_precise=warehouse
                )
        if customer:
            criteria &= models.Q(customer_name__zem_name=customer)
        criteria &= models.Q(
            models.Q(vessel_id__vessel_etd__gte=start_date_confirm),
            models.Q(vessel_id__vessel_etd__lte=end_date_confirm),
        )
        # 查找提拆的待开和已开
        payable_preport_subquery = InvoicePreport.objects.filter(
            invoice_number=OuterRef("invoice_id"), invoice_type="payable"
        )
        payable_other_fee_subquery = InvoicePreport.objects.filter(
            invoice_number=OuterRef("invoice_id"), invoice_type="payable"
        ).values("other_fees__other_fee")[:1]
        pick_pending_orders = (
            Order.objects.select_related(
                "customer_name", "container_number", "retrieval_id", "invoice_id"
            )
            .annotate(
                payable_pickup=Subquery(payable_preport_subquery.values("pickup")[:1]),
                payable_chassis=Subquery(
                    payable_preport_subquery.values("chassis")[:1]
                ),
                payable_over_weight=Subquery(
                    payable_preport_subquery.values("over_weight")[:1]
                ),
                payable_demurrage=Subquery(
                    payable_preport_subquery.values("demurrage")[:1]
                ),
                payable_per_diem=Subquery(
                    payable_preport_subquery.values("per_diem")[:1]
                ),
                payable_other_fee_data=Subquery(payable_other_fee_subquery),
            )
            .filter(
                criteria,
                models.Q(payable_status__stage="tobeconfirmed")
                | models.Q(payable_status__stage="confirmed"),
                payable_status__isnull=False,
                payable_status__payable_status__has_key="pickup",
                payable_status__payable_status__pickup="pending",
            )
            .values(
                "container_number__container_number",
                "customer_name__zem_name",
                "retrieval_id__retrieval_destination_area",
                "invoice_id__invoice_number",
                "payable_pickup",  # 提柜费
                "payable_over_weight",  # 超重费
                "payable_chassis",  # 车架费
                "payable_demurrage",  # 滞港费
                "payable_per_diem",  # 滞箱费
                "payable_other_fee_data",  # 其他费用
                "invoice_id__payable_preport_amount",  # 总费用
                "invoice_id",
                "retrieval_id__retrieval_carrier",
            )
        )

        pick_confirmed_orders = (
            Order.objects.select_related(
                "customer_name", "container_number", "retrieval_id", "invoice_id"
            )
            .annotate(
                payable_pickup=Subquery(payable_preport_subquery.values("pickup")[:1]),
                payable_chassis=Subquery(
                    payable_preport_subquery.values("chassis")[:1]
                ),
                payable_over_weight=Subquery(
                    payable_preport_subquery.values("over_weight")[:1]
                ),
                payable_demurrage=Subquery(
                    payable_preport_subquery.values("demurrage")[:1]
                ),
                payable_per_diem=Subquery(
                    payable_preport_subquery.values("per_diem")[:1]
                ),
                payable_other_fee_data=Subquery(payable_other_fee_subquery),
            )
            .filter(
                criteria,
                models.Q(payable_status__stage="tobeconfirmed")
                | models.Q(payable_status__stage="confirmed"),
                payable_status__isnull=False,
                payable_status__payable_status__has_key="pickup",
                payable_status__payable_status__pickup="confirmed",
            )
            .values(
                "container_number__container_number",
                "customer_name__zem_name",
                "retrieval_id__retrieval_destination_area",
                "invoice_id__invoice_number",
                "payable_pickup",  # 提柜费
                "payable_over_weight",  # 超重费
                "payable_chassis",  # 车架费
                "payable_demurrage",  # 滞港费
                "payable_per_diem",  # 滞箱费
                "payable_other_fee_data",  # 其他费用
                "invoice_id__payable_preport_amount",  # 总费用
                "retrieval_id__retrieval_carrier",
            )
        )

        # 查找提拆的待开和已开
        payable_warehouse_subquery = InvoiceWarehouse.objects.filter(
            invoice_number=OuterRef("invoice_id"), invoice_type="payable"
        )
        payable_other_fee_wh_subquery = InvoiceWarehouse.objects.filter(
            invoice_number=OuterRef("invoice_id"), invoice_type="payable"
        ).values("other_fees__other")[:1]
        warehouse_pending_orders = (
            Order.objects.select_related(
                "customer_name", "container_number", "retrieval_id", "invoice_id"
            )
            .annotate(
                payable_palletization_fee=Subquery(
                    payable_warehouse_subquery.values("palletization_fee")[:1]
                ),
                payable_arrive_fee=Subquery(
                    payable_warehouse_subquery.values("arrive_fee")[:1]
                ),
                payable_carrier=Subquery(
                    payable_warehouse_subquery.values("carrier")[:1]
                ),
                payable_other_fee_data=Subquery(payable_other_fee_wh_subquery),
            )
            .filter(
                criteria,
                models.Q(payable_status__stage="tobeconfirmed")
                | models.Q(payable_status__stage="confirmed"),
                payable_status__isnull=False,
                payable_status__payable_status__has_key="warehouse",
                payable_status__payable_status__warehouse="pending",
            )
            .values(
                "container_number__container_number",
                "customer_name__zem_name",
                "retrieval_id__retrieval_destination_area",
                "invoice_id__invoice_number",
                "payable_palletization_fee",  # 拆柜费
                "payable_arrive_fee",  # 入库费
                "payable_carrier",  # 供应商
                "payable_other_fee_data",  # 其他费用
                "invoice_id__payable_warehouse_amount",  # 总费用
                "invoice_id",
            )
        )
        warehouse_confirmed_orders = (
            Order.objects.select_related(
                "customer_name", "container_number", "retrieval_id", "invoice_id"
            )
            .annotate(
                payable_palletization_fee=Subquery(
                    payable_warehouse_subquery.values("palletization_fee")[:1]
                ),
                payable_arrive_fee=Subquery(
                    payable_warehouse_subquery.values("arrive_fee")[:1]
                ),
                payable_carrier=Subquery(
                    payable_warehouse_subquery.values("carrier")[:1]
                ),
                payable_other_fee_data=Subquery(payable_other_fee_wh_subquery),
            )
            .filter(
                criteria,
                models.Q(payable_status__stage="tobeconfirmed")
                | models.Q(payable_status__stage="confirmed"),
                payable_status__isnull=False,
                payable_status__payable_status__has_key="warehouse",
                payable_status__payable_status__warehouse="confirmed",
            )
            .values(
                "container_number__container_number",
                "customer_name__zem_name",
                "retrieval_id__retrieval_destination_area",
                "invoice_id__invoice_number",
                "payable_palletization_fee",  # 拆柜费
                "payable_arrive_fee",  # 入库费
                "payable_carrier",  # 供应商
                "payable_other_fee_data",  # 其他费用
                "invoice_id__payable_warehouse_amount",  # 总费用
                "invoice_id",
            )
        )

        # criteria和fleetshipmentpallet表直接关联较为复杂，所以先筛选pallet表，然后根据PO_ID去筛选
        container_numbers = Order.objects.filter(criteria).values_list(
            "container_number", flat=True
        )
        delivery_po_ids = (
            Pallet.objects.filter(container_number__in=container_numbers)
            .values_list("PO_ID", flat=True)
            .distinct()
        )
        delivery_pending_orders = (
            FleetShipmentPallet.objects.select_related(
                "fleet_number",  # 确保预取关联的Fleet对象
                "shipment_batch_number",
                "container_number",
            )
            .filter(
                expense__isnull=False,
                PO_ID__in=delivery_po_ids,
            )
            .annotate(
                appointment_id=F("shipment_batch_number__appointment_id"),
                container_num=F("container_number__container_number"),
                pallet_destination=Subquery(
                    Pallet.objects.filter(
                        PO_ID=OuterRef("PO_ID"),
                        shipment_batch_number=OuterRef("shipment_batch_number"),
                    ).values("destination")[:1]
                ),
            )
            .order_by("fleet_number", "pickup_number", "container_num")
        )

        deliverys = {}
        for order in delivery_pending_orders:
            fleet_id = order.fleet_number_id
            if fleet_id not in deliverys:
                # 获取carrier信息
                carrier = order.fleet_number.carrier if order.fleet_number else None

                deliverys[fleet_id] = {
                    "fleets": {},
                    "total_pallets": 0,
                    "total_expense": 0,
                    "total_rows": 0,
                    "carrier": carrier,
                }

            if order.pickup_number not in deliverys[fleet_id]["fleets"]:
                deliverys[fleet_id]["fleets"][order.pickup_number] = {
                    "appointments": {},
                    "ISA_total_pallets": 0,
                    "ISA_total_expense": 0,
                    "total_rows": 0,
                }

            if (
                order.appointment_id
                not in deliverys[fleet_id]["fleets"][order.pickup_number][
                    "appointments"
                ]
            ):
                deliverys[fleet_id]["fleets"][order.pickup_number]["appointments"][
                    order.appointment_id
                ] = {"orders": [], "total_pallets": 0, "total_expense": 0, "rowspan": 0}

            # 计算单板成本并保留2位小数
            per_expense = (
                round(order.expense / int(order.total_pallet), 2)
                if order.total_pallet
                else 0
            )

            # 为每个order添加统计字段
            order_data = {
                "object": order,
                "cn_total_pallet": int(order.total_pallet) if order.total_pallet else 0,
                "cn_total_expense": order.expense or 0,
                "cn_per_expense": per_expense,
                "container_num": order.container_num,
                "pallet_destination": order.pallet_destination,
                "carrier": deliverys[fleet_id]["carrier"],
            }

            deliverys[fleet_id]["fleets"][order.pickup_number]["appointments"][
                order.appointment_id
            ]["orders"].append(order_data)

            deliverys[fleet_id]["fleets"][order.pickup_number]["appointments"][
                order.appointment_id
            ]["total_pallets"] += int(order_data["cn_total_pallet"])
            deliverys[fleet_id]["fleets"][order.pickup_number]["appointments"][
                order.appointment_id
            ]["total_expense"] += order_data["cn_total_expense"]
            deliverys[fleet_id]["fleets"][order.pickup_number]["appointments"][
                order.appointment_id
            ]["rowspan"] += 1

            deliverys[fleet_id]["fleets"][order.pickup_number][
                "ISA_total_pallets"
            ] += int(order_data["cn_total_pallet"])
            deliverys[fleet_id]["fleets"][order.pickup_number][
                "ISA_total_expense"
            ] += order_data["cn_total_expense"]
            deliverys[fleet_id]["fleets"][order.pickup_number]["total_rows"] += 1

            deliverys[fleet_id]["total_pallets"] += int(order_data["cn_total_pallet"])
            deliverys[fleet_id]["total_expense"] += order_data["cn_total_expense"]
            deliverys[fleet_id]["total_rows"] += 1

        # 转换数据结构
        delivery_pending_orders = [
            {
                "fleets": fleet_data["fleets"],
                "total_pallets": int(fleet_data["total_pallets"]),
                "total_expense": fleet_data["total_expense"],
                "carrier": fleet_data["carrier"],
            }
            for fleet_data in deliverys.values()
        ]

        start_date_export = (current_date + timedelta(days=-15)).strftime("%Y-%m-%d")
        end_date_export = current_date.strftime("%Y-%m-%d")
        pickup_carriers = {
            "Kars": "Kars",
            "东海岸": "东海岸",
            "ARM": "ARM",
            "GM": "GM",
            "BEST": "BEST",
        }
        warehouse_carriers = {
            "BBR": "BBR",
            "KNO": "KNO",
        }
        existing_customers = Customer.objects.all().order_by("zem_name")
        context = {
            "pickup_carriers": pickup_carriers,
            "warehouse_carriers": warehouse_carriers,
            "pick_pending_orders": pick_pending_orders,
            "pick_confirmed_orders": pick_confirmed_orders,
            "warehouse_pending_orders": warehouse_pending_orders,
            "warehouse_confirmed_orders": warehouse_confirmed_orders,
            "delivery_pending_orders": delivery_pending_orders,
            "delivery_confirmed_orders": None,
            "warehouse_options": self.warehouse_options,
            "existing_customers": existing_customers,
            "order_form": OrderForm(),
            "start_date_confirm": start_date_confirm,
            "end_date_confirm": end_date_confirm,
            "start_date_export": start_date_export,
            "end_date_export": end_date_export,
            "invoice_type_filter": "payable",
            "CARRIER_FLEET": CARRIER_FLEET,
            "warehouse_filter": request.POST.get("warehouse_filter"),
        }
        return self.template_invoice_payable_confirm, context

    def handle_invoice_delivery_get(
        self,
        request: HttpRequest,
        start_date: str = None,
        end_date: str = None,
        customer: str = None,
        warehouse: str = None,
    ) -> tuple[Any, Any]:
        # 库内操作费
        current_date = datetime.now().date()
        start_date = request.POST.get("start_date") or (
            current_date + timedelta(days=-90)
        ).strftime("%Y-%m-%d")
        end_date = request.POST.get("end_date") or current_date.strftime("%Y-%m-%d")

        criteria = models.Q(
            (models.Q(order_type="转运") | models.Q(order_type="转运组合")),
            models.Q(vessel_id__vessel_etd__gte=start_date),
            models.Q(vessel_id__vessel_etd__lte=end_date),
            models.Q(offload_id__offload_at__isnull=False),
        )
        if warehouse is None:
            warehouse = request.POST.get("warehouse")
        if warehouse and warehouse != "None":
            criteria &= models.Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= models.Q(customer_name__zem_name=customer)

        groups = [group.name for group in request.user.groups.all()]
        delivery_type_filter = None
        display_mix = False
        if "mix_account" in groups:
            display_mix = True

        if display_mix:
            delivery_type_filter = models.Q()
            delivery_type = "mix"
        elif ("warehouse_other" in groups and "warehouse_public" not in groups) or (
            "NJ_mix_account" in groups
        ):
            delivery_type_filter = models.Q(
                container_number__delivery_type__in=["other", "mixed"]
            )
            delivery_type = "other"
        elif "warehouse_public" in groups and "warehouse_other" not in groups:
            delivery_type_filter = models.Q(
                container_number__delivery_type__in=["public", "mixed"]
            )
            delivery_type = "public"

        # 基础查询
        base_query = Order.objects.select_related(
            "invoice_id",
            "receivable_status",
            "customer_name",
            "container_number",
            "invoice_id__statement_id",
        ).filter(criteria)

        if delivery_type_filter:
            base_query = base_query.filter(delivery_type_filter)
        # 判断是否有暂扣的板子
        hold_subquery = Pallet.objects.filter(
            container_number=OuterRef("container_number"),  # 关联到Order的container
            delivery_method__contains="暂扣留仓",
        )
        if delivery_type != "mix":
            hold_subquery = hold_subquery.filter(delivery_type=delivery_type)

        # 查找未操作过的
        if "NJ_mix_account" in groups or (
            "warehouse_other" in groups and "warehouse_public" not in groups
        ):  # 这个权限的，要看NJ的私仓
            order = (
                base_query.filter(
                    models.Q(
                        **{"receivable_status__stage_other": "warehouse_completed"}
                    )
                    | models.Q(
                        **{"receivable_status__stage_other": "delivery_rejected"}
                    )
                )
                .annotate(
                    is_priority=Case(
                        When(
                            **{"receivable_status__stage_other": "delivery_rejected"},
                            then=Value(0),
                        ),
                        When(
                            **{"receivable_status__stage_other": "warehouse_completed"},
                            then=Value(1),
                        ),
                        output_field=IntegerField(),
                    ),
                    is_hold=Exists(hold_subquery),  # 表示柜子是否有暂扣的
                )
                .order_by("is_priority")
            )
        else:
            if display_mix:  # 这个权限的，都能看
                order = (
                    base_query.filter(
                        models.Q(**{"receivable_status__stage": "delivery"})
                    )
                    .annotate(
                        is_priority=Case(
                            When(
                                models.Q(
                                    **{
                                        "receivable_status__stage_other": "delivery_rejected"
                                    }
                                )
                                | models.Q(
                                    **{
                                        "receivable_status__stage_public": "delivery_rejected"
                                    }
                                ),
                                then=Value(0),
                            ),
                            default=Value(1),
                            output_field=IntegerField(),
                        ),
                        is_hold=Exists(hold_subquery),
                    )
                    .order_by("is_priority")
                )
            elif "warehouse_public" in groups and "warehouse_other" not in groups:
                # 只看公仓人员
                order = (
                    base_query.filter(
                        models.Q(
                            **{"receivable_status__stage_public": "warehouse_completed"}
                        )
                        | models.Q(
                            **{"receivable_status__stage_public": "delivery_rejected"}
                        )
                    )
                    .annotate(
                        is_priority=Case(
                            When(
                                **{
                                    "receivable_status__stage_public": "delivery_rejected"
                                },
                                then=Value(0),
                            ),
                            When(
                                **{
                                    "receivable_status__stage_public": "warehouse_completed"
                                },
                                then=Value(1),
                            ),
                            output_field=IntegerField(),
                        ),
                        is_hold=Exists(hold_subquery),
                    )
                    .order_by("is_priority")
                )
        order = self.process_orders_display_status(order, "receivable")

        # 查找历史操作过的
        base_condition = ~models.Q(
            **{"receivable_status__stage__in": ["unstarted", "preport"]}
        )
        other_stages = models.Q(
            **{"receivable_status__stage__in": ["tobeconfirmed", "confirmed"]}
        )
        delivery_completed_condition = models.Q()
        if "NJ_mix_account" in groups:
            delivery_completed_condition = models.Q(
                **{"receivable_status__stage_other": "delivery_completed"}
            )
        else:
            if display_mix:
                delivery_completed_condition = models.Q()
            elif "warehouse_public" in groups and "warehouse_other" not in groups:
                delivery_completed_condition = models.Q(
                    **{"receivable_status__stage_public": "delivery_completed"}
                )
            elif "warehouse_other" in groups and "warehouse_public" not in groups:
                delivery_completed_condition = models.Q(
                    **{"receivable_status__stage_other": "delivery_completed"}
                )
        previous_order = base_query.filter(
            base_condition,
            other_stages | delivery_completed_condition,  # 满足任意一个条件即可
            **{"receivable_status__invoice_type": "receivable"},
        ).select_related(
            "customer_name", "container_number", "receivable_status", "payable_status"
        )
        previous_order = self.process_orders_display_status(
            previous_order, "receivable"
        )
        context = {
            "order": order,
            "previous_order": previous_order,
            "order_form": OrderForm(),
            "start_date": start_date,
            "end_date": end_date,
            "customer": customer,
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
            "warehouse": warehouse,
            "modify_shipped_shipment": request.user.groups.filter(name="shipment_leader").exists(),
        }
        return self.template_invoice_delivery, context

    def handle_invoice_delivery_calc_expense(self, container_number: str) -> None:
        fleet_expenses = (
            FleetShipmentPallet.objects.values("PO_ID", "pickup_number")
            .annotate(total_expense=Sum("expense"), total_pallet=Sum("total_pallet"))
            .filter(container_number__container_number=container_number)
        )
        payable_delivery_amount = 0
        for expense_group in fleet_expenses:
            po_id = expense_group["PO_ID"]
            total_expense = expense_group["total_expense"]
            payable_delivery_amount += total_expense
            if not total_expense:
                raise ValueError(f"{expense_group['pickup_number']}车次成本未记录")
            total_pallet = expense_group["total_pallet"]
            # 找到该po_id的所有板子，根据外键找到对应的invoice_delivery
            pallets = Pallet.objects.filter(PO_ID=po_id).select_related(
                "invoice_delivery"
            )
            delivery_groups = {}
            # 将板数按照invoice_delivery去分组，记录每组的板数
            for pallet in pallets:
                if pallet.invoice_delivery:
                    delivery_id = pallet.invoice_delivery.id
                    if delivery_id not in delivery_groups:
                        delivery_groups[delivery_id] = {
                            "delivery": pallet.invoice_delivery,
                            "pallet_count": 0,
                            "pallets": [],
                        }
                    delivery_groups[delivery_id]["pallet_count"] += 1
                    delivery_groups[delivery_id]["pallets"].append(pallet)
            # 遍历每组的invoice_delivery，
            for group in delivery_groups.values():
                delivery = group["delivery"]
                pallet_count = group["pallet_count"]
                # 根据板数的比列，去计算成本
                allocated_expense = (pallet_count / total_pallet) * total_expense
                InvoiceDelivery.objects.filter(id=delivery.id).update(
                    expense=allocated_expense, total_pallet=pallet_count
                )
        # 给invoice的delivery_amount赋值
        invoice = Invoice.objects.get(
            container_number__container_number=container_number
        )
        invoice.payable_delivery_amount = payable_delivery_amount
        invoice.save()

    def handle_invoice_payable_save_post(self, request: HttpRequest) -> tuple[Any, Any]:
        data = request.POST.copy()
        save_type = data.get("save_type")
        container_number = data.get("container_number")
        invoice_status = InvoiceStatus.objects.get(
            container_number__container_number=container_number, invoice_type="payable"
        )
        if save_type == "return":  # 返回不需要保存数据
            return self.handle_invoice_payable_get(
                request,
                data.get("start_date"),
                data.get("end_date"),
                None,
                data.get("warehouse_filter"),
            )
        elif save_type == "reject":  # 财务驳回
            invoice_status.stage = "unstarted"
            invoice_status.is_rejected = True
            invoice_status.reject_reason = data.get("reject_reason")
            invoice_status.save()
            return self.handle_invoice_confirm_get(request)
        elif save_type == "reject_check":  # 初级审核驳回，驳回不改数据
            invoice_status.stage = "unstarted"
            invoice_status.is_rejected = True
            invoice_status.reject_reason = data.get("reject_reason")
            invoice_status.save()
            return self.handle_invoice_payable_get(
                request,
                data.get("start_date"),
                data.get("end_date"),
                None,
                data.get("warehouse_filter"),
            )
        elif save_type == "account_confirm":
            # 这里加一个方法，给invoice_delivery表计算每条记录的成本
            self.handle_invoice_delivery_calc_expense(container_number)
            invoice_status.stage = "confirmed"
            invoice_status.save()
            return self.handle_invoice_confirm_get(request)
        elif save_type == "check_confirm_without_data":  # 列表中确认通过，不需要改数据
            invoice_status.stage = "tobeconfirmed"
            invoice_status.save()
            return self.handle_invoice_payable_get(
                request,
                data.get("start_date"),
                data.get("end_date"),
                None,
                data.get("warehouse_filter"),
            )

        total_amount = data.get("total_amount")
        if not total_amount and total_amount == 0:
            raise ValueError("审核时，数据存储错误")
        # 直送的托架费提示信息
        chassis_comment = data.get("chassis_comment")
        # 将额外费用构建键值对
        pickup_fee_names = data.getlist("pickup_fee_name")
        pickup_fee_amounts = data.getlist("pickup_fee_amount")
        pt_amount = 0
        pickup_fees = {}
        for name, amount in zip(pickup_fee_names, pickup_fee_amounts):
            if name and amount:
                pt_amount += float(amount)
                pickup_fees[name] = float(amount)

        pallet_fee_names = data.getlist("pallet_fee_name")
        pallet_fee_amounts = data.getlist("pallet_fee_amount")

        wh_amount = 0
        pallet_fees = {}
        for name, amount in zip(pallet_fee_names, pallet_fee_amounts):
            if name and amount:
                wh_amount += float(amount)
                pallet_fees[name] = float(amount)
        # 存到invoice表里
        invoice = Invoice.objects.get(
            container_number__container_number=container_number
        )
        try:
            invoice_preports = InvoicePreport.objects.get(
                invoice_number__invoice_number=invoice.invoice_number,
                invoice_type="payable",
            )
        except InvoicePreport.DoesNotExist:
            invoice_preports = InvoicePreport(
                **{"invoice_number": invoice, "invoice_type": "payable"}
            )

        try:
            invoice_warehouse = InvoiceWarehouse.objects.get(
                invoice_number__invoice_number=invoice.invoice_number,
                invoice_type="payable",
            )
        except InvoiceWarehouse.DoesNotExist:
            invoice_warehouse = InvoiceWarehouse(
                **{"invoice_number": invoice, "invoice_type": "payable"}
            )
        if data.get("palletization_fee"):
            invoice_warehouse.palletization_fee = data.get("palletization_fee")
            wh_amount += float(data.get("palletization_fee"))
        if data.get("arrive_fee"):
            invoice_warehouse.arrive_fee = data.get("arrive_fee")
            wh_amount += float(data.get("arrive_fee"))
        # 拆柜供应商
        invoice_warehouse.carrier = data.get("palletization_carrier")
        invoice_warehouse.amount = wh_amount
        invoice.payable_warehouse_amount = wh_amount

        if data.get("basic_fee"):  # 提柜费
            invoice_preports.pickup = data.get("basic_fee")
            pt_amount += float(data.get("basic_fee"))
        if data.get("overweight_fee"):  # 超重费
            invoice_preports.over_weight = data.get("overweight_fee")
            pt_amount += float(data.get("overweight_fee"))
        if data.get("chassis_fee"):  # 车架费
            invoice_preports.chassis = data.get("chassis_fee")
            pt_amount += float(data.get("chassis_fee"))
        if data.get("demurrage_fee"):  # 滞港费
            invoice_preports.demurrage = data.get("demurrage_fee")
            pt_amount += float(data.get("demurrage_fee"))
        if data.get("per_diem_fee"):  # 滞箱费
            invoice_preports.per_diem = data.get("per_diem_fee")
            pt_amount += float(data.get("per_diem_fee"))
        if invoice.payable_palletization == "None":
            raise ValueError("未选择打板费")
        invoice_preports.amount = pt_amount
        invoice.payable_preport_amount = pt_amount
        invoice_preports.other_fees = {
            "other_fee": pickup_fees,
            "chassis_comment": chassis_comment,
        }
        invoice_preports.save()
        invoice_warehouse.other_fees = {
            "other_fee": pallet_fees,
        }
        invoice_warehouse.save()

        invoice.payable_total_amount = total_amount

        invoice.save()
        if save_type == "check_confirm":  # 初级审核，可以修改价格确认完就转到财务审核
            invoice_status.stage = "tobeconfirmed"
            invoice_status.save()
            return self.handle_invoice_payable_get(
                request,
                data.get("start_date"),
                data.get("end_date"),
                None,
                data.get("warehouse_filter"),
            )
        # 如果确认，就改变状态
        if save_type == "complete":
            invoice_status.stage = "preport"  # 这是转给rose去审核
            invoice_status.is_rejected = False
            invoice_status.reject_reason = ""

        # 修改payable_status字段
        for key, value in invoice_status.payable_status.items():
            if value != "confirmed":
                invoice_status.payable_status[key] = "pending"
        invoice_status.save()
        return self.handle_invoice_payable_get(
            request,
            data.get("start_date"),
            data.get("end_date"),
            None,
            data.get("warehouse_filter"),
        )

    def handle_invoice_warehouse_save_post(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        data = request.POST.copy()
        save_type = request.POST.get("save_type")
        container_number = data.get("container_number")
        invoice = Invoice.objects.select_related("container_number").get(
            container_number__container_number=container_number
        )

        invoice_warehouse = InvoiceWarehouse.objects.filter(
            invoice_number__invoice_number=invoice.invoice_number,
            invoice_type="receivable",
        )
        if not invoice_warehouse.exists():
            invoice_content = InvoiceWarehouse(
                **{
                    "invoice_number": invoice,
                }
            )
            invoice_content.save()
        delivery_type = request.POST.get("delivery_type")
        invoice_warehouse = InvoiceWarehouse.objects.get(
            invoice_number__invoice_number=invoice.invoice_number,
            invoice_type="receivable",
            delivery_type=delivery_type,
        )
        names = data.getlist("others_feename")[:-1]
        amounts = data.getlist("others_feeamount")[:-1]
        other_fees = dict(zip(names, map(float, amounts)))
        invoice_warehouse.other_fees = {k: v for k, v in other_fees.items() if k}
        exclude_fields = {
            "csrfmiddlewaretoken",
            "step",
            "warehouse",
            "container_number",
            "invoice_number",
            "others_feename",
            "others_feeamount",
        }

        # 附加项费用和附加项说明
        fields = [
            "sorting",
            "intercept",
            "po_activation",
            "self_pickup",
            "re_pallet",
            "counting",
            "warehouse_rent",
            "specified_labeling",
            "inner_outer_box",
            "pallet_label",
            "open_close_box",
            "destroy",
            "take_photo",
            "take_video",
            "repeated_operation_fee",
            "per_diem",
            "second_pickup",
        ]

        # 初始化单价和数量字典
        qty_data = {}
        rate_data = {}
        s_fields = fields
        for field in s_fields:
            # 保存数量
            quantity_key = f"{field}_quantity"
            if quantity_key in data:
                qty_data[field] = (
                    float(data.get(quantity_key, 0))
                    if data.get(quantity_key) is not None
                    else 0
                )

            # 保存单价
            price_key = f"{field}_price"
            if price_key in data:
                rate_data[field] = (
                    float(data.get(price_key, 1))
                    if data.get(price_key) is not None
                    else 1
                )

            # 保存原有字段
            if field in data and field not in exclude_fields and data[field]:
                setattr(invoice_warehouse, field, data[field])
        setattr(invoice_warehouse, "amount", data["amount"])
        # 保存单价和数量
        invoice_warehouse.qty = qty_data
        invoice_warehouse.rate = rate_data
        surcharges = {}
        surcharge_notes = {}
        for field in fields:

            surcharge_key = f"{field}_surcharge"
            note_key = f"{field}_surcharge_note"
            surcharge = data.get(surcharge_key, 0) or 0
            note = data.get(note_key, "")
            surcharges[field] = float(surcharge)
            surcharge_notes[field] = note
        invoice_warehouse.surcharges = surcharges
        invoice_warehouse.surcharge_notes = surcharge_notes
        invoice_warehouse.save()

        # 因为现在分公仓私仓两条记录，所以汇总的时候，要从数据库查一遍
        if save_type == "complete" or save_type == "account_comlete":
            invoice = Invoice.objects.select_related("container_number").get(
                container_number__container_number=container_number,
            )
            warehouse_amount = (
                InvoiceWarehouse.objects.filter(
                    invoice_number=invoice, invoice_type="receivable"
                ).aggregate(total_amount=Sum("amount"))["total_amount"]
                or 0
            )
            invoice.receivable_warehouse_amount = warehouse_amount
            invoice.save()

        order = Order.objects.select_related("retrieval_id", "container_number").get(
            container_number__container_number=container_number
        )

        if save_type == "complete":
            # 开始准备改变状态，先找到状态表
            invoice_status = InvoiceStatus.objects.get(
                container_number=order.container_number, invoice_type="receivable"
            )
            container_delivery_type = invoice_status.container_number.delivery_type
            # 如果这是被驳回的，就直接改主状态为待确认，其他不用动
            if invoice_status.is_rejected == True:
                invoice_status.stage = "tobeconfirmed"
                invoice_status.stage_other = "delivery_completed"
                invoice_status.stage_public = "delivery_completed"
                invoice_status.is_rejected = False
                invoice_status.reject_reason = ""
            else:
                if container_delivery_type in ["public", "other"]:
                    # 如果这个柜子只有一类仓，就直接改变状态
                    invoice_status.stage = "delivery"
                    invoice_status.is_rejected = False
                    invoice_status.reject_reason = ""
                    invoice_status.stage_public = "warehouse_completed"
                    invoice_status.stage_other = "warehouse_completed"
                elif container_delivery_type == "mixed":
                    if delivery_type == "public":
                        # 公仓组录完了，改变stage_public
                        if invoice_status.stage_public not in [
                            "delivery_completed",
                            "delivery_rejected",
                        ]:
                            invoice_status.stage_public = "warehouse_completed"
                            # 如果私仓也做完了，就改变主状态到派送阶段
                            if invoice_status.stage_other not in [
                                "pending",
                                "warehouse_rejected",
                            ]:
                                invoice_status.stage = "delivery"
                    elif delivery_type == "other":
                        # 私仓租录完了，改变stage_other
                        if invoice_status.stage_other not in [
                            "delivery_completed",
                            "delivery_rejected",
                        ]:
                            invoice_status.stage_other = "warehouse_completed"
                            # 如果公仓也做完了，就改变主状态
                            if invoice_status.stage_public not in [
                                "pending",
                                "warehouse_rejected",
                            ]:
                                invoice_status.stage = "delivery"
                    else:
                        raise ValueError("没有派送类别")
                    # 既有公仓权限，又有私仓权限的不知道咋处理，而且编辑页面也不好搞
                    invoice_status.is_rejected = False
                    invoice_status.reject_reason = ""
            invoice_status.save()
        elif save_type == "account_comlete":
            modified_get = request.GET.copy()
            modified_get["start_date_confirm"] = request.POST.get("start_date_confirm")
            modified_get["end_date_confirm"] = request.POST.get("end_date_confirm")
            new_request = request
            new_request.GET = modified_get
            return self.handle_container_invoice_confirm_get(new_request)

        order_form = OrderForm(request.POST)
        if order_form.is_valid():
            customer = order_form.cleaned_data.get("customer_name")
        else:
            customer = None
        return self.handle_invoice_warehouse_get(
            request,
            request.POST.get("start_date"),
            request.POST.get("end_date"),
            customer,
            request.POST.get("warehouse"),
        )

    def handle_invoice_direct_save_post(self, request: HttpRequest) -> tuple[Any, Any]:
        data = request.POST.copy()
        save_type = request.POST.get("save_type")
        container_number = data.get("container_number")
        direct_amount = request.POST.get("amount")
        invoice = Invoice.objects.select_related("container_number").get(
            container_number__container_number=container_number,
        )
        order = Order.objects.select_related("container_number").get(
            container_number__container_number=container_number
        )
        try:
            invoice_preports = InvoicePreport.objects.get(
                invoice_number__invoice_number=invoice.invoice_number,
                invoice_type="receivable",
            )
        except InvoicePreport.DoesNotExist:
            # 获取直送柜子的提拆柜费用
            # 如果之前没有录过费用，就根据报价表生成提+派送费用
            invoice_preports = InvoicePreport(
                **{"invoice_number": invoice, "invoice_type": "receivable"}
            )
            invoice_preports.save()

        names = data.getlist("others_feename")[:-1]
        amounts = data.getlist("others_feeamount")[:-1]
        other_fees = dict(zip(names, map(float, amounts)))
        invoice_preports.other_fees = {k: v for k, v in other_fees.items() if k}
        exclude_fields = {
            "csrfmiddlewaretoken",
            "step",
            "warehouse",
            "container_number",
            "invoice_number",
            "others_feename",
            "others_feeamount",
        }
        # 附加项费用和附加项说明
        fields = [
            "exam_fee",
            "second_pickup",
            "demurrage",
            "per_diem",
            "congestion_fee",
            "chassis",
            "prepull",
            "yard_storage",
            "handling_fee",
            "chassis_split",
            "over_weight",
        ]
        # 初始化单价和数量字典
        qty_data = {}
        rate_data = {}
        s_fields = ["pickup"] + ["amount"] + fields
        for field in s_fields:
            # 保存单价
            quantity_key = f"{field}_quantity"
            if quantity_key in data:
                qty_data[field] = float(data.get(quantity_key, 0)) or 0
            # 保存数量
            price_key = f"{field}_price"
            if price_key in data:
                rate_data[field] = float(data.get(price_key, 1)) or 1
            # 保存原有字段
            if field in data and field not in exclude_fields and data[field]:
                setattr(invoice_preports, field, data[field])
        # 保存单价和数量
        invoice_preports.qty = qty_data
        invoice_preports.rate = rate_data

        surcharges = {}
        surcharge_notes = {}
        for field in fields:
            surcharge_key = f"{field}_surcharge"
            note_key = f"{field}_surcharge_note"
            surcharge = request.POST.get(surcharge_key, 0) or 0
            note = request.POST.get(note_key, "")
            surcharges[field] = float(surcharge)
            surcharge_notes[field] = note
        invoice_preports.surcharges = surcharges
        invoice_preports.surcharge_notes = surcharge_notes
        invoice_preports.save()

        if save_type == "complete":  # 如果是普通账户确认，订单转为待财务确认状态
            # 更新invoice表和状态表
            invoice.receivable_direct_amount = direct_amount
            invoice_status, created = InvoiceStatus.objects.get_or_create(
                container_number=order.container_number, invoice_type="receivable"
            )
            invoice.save()
            invoice_status.stage = "tobeconfirmed"
            invoice_status.is_rejected = "False"
            invoice_status.reject_reason = ""
            invoice_status.save()
        elif save_type == "account_complete":
            # 如果是财务确认，订单转为已确认状态
            invoice = Invoice.objects.select_related("container_number").get(
                container_number__container_number=container_number
            )
            invoice.receivable_direct_amount = direct_amount
            invoice_status = InvoiceStatus.objects.get(
                container_number=order.container_number, invoice_type="receivable"
            )
            invoice.save()
            invoice_status.stage = "confirmed"
            invoice_status.is_rejected = "False"
            invoice_status.reject_reason = ""
            invoice_status.save()
        elif save_type == "reject":
            # 如果是财务拒绝，退回到未编辑状态，并记录驳回原因和驳回状态
            invoice_status = InvoiceStatus.objects.get(
                container_number=order.container_number, invoice_type="receivable"
            )
            invoice_status.stage = "unstarted"
            invoice_status.is_rejected = "True"
            invoice_status.reject_reason = data.get("invoice_reject_reason", "")
            invoice_status.save()

        if save_type in [
            "account_complete",
            "reject",
        ]:
            # 如果是从账单确认那里点进来的操作，就跳转回账单确认界面
            return self.handle_invoice_confirm_save(request)
        else:
            return self.handle_invoice_direct_get(
                request, request.POST.get("start_date"), request.POST.get("end_date")
            )

    def handle_invoice_preport_save_post(self, request: HttpRequest) -> tuple[Any, Any]:
        data = request.POST.copy()
        save_type = request.POST.get("save_type")
        container_number = data.get("container_number")
        invoice = Invoice.objects.select_related("container_number").get(
            container_number__container_number=container_number
        )
        preport_amount = data["amount"]
        # 提拆柜表费用记录
        invoice_preports = InvoicePreport.objects.get(
            invoice_number__invoice_number=invoice.invoice_number,
            invoice_type="receivable",
        )
        names = [name for name in data.getlist("others_feename") if name.strip()]
        amounts = [
            amount for amount in data.getlist("others_feeamount") if amount.strip()
        ]
        other_fees = dict(zip(names, map(float, amounts)))
        invoice_preports.other_fees = {k: v for k, v in other_fees.items() if k}
        exclude_fields = {
            "csrfmiddlewaretoken",
            "step",
            "warehouse",
            "container_number",
            "invoice_number",
            "pending",
            "others_feename",
            "others_feeamount",
        }
        # 附加项费用和附加项说明,qty多了一个pickup
        fields = [
            "chassis",
            "chassis_split",
            "prepull",
            "yard_storage",
            "handling_fee",
            "pier_pass",
            "congestion_fee",
            "hanging_crane",
            "dry_run",
            "exam_fee",
            "hazmat",
            "over_weight",
            "urgent_fee",
            "other_serive",
            "demurrage",
            "per_diem",
            "second_pickup",
        ]

        # 初始化单价和数量字典
        qty_data = {}
        rate_data = {}
        s_fields = ["pickup"] + ["amount"] + fields
        for field in s_fields:
            # 保存单价
            quantity_key = f"{field}_quantity"
            if quantity_key in data:
                qty_data[field] = float(data.get(quantity_key, 0)) or 0

            # 保存数量
            price_key = f"{field}_price"

            if price_key in data:
                rate_data[field] = float(data.get(price_key, 1)) or 1

            # 保存原有字段
            if field in data and field not in exclude_fields and data[field]:
                setattr(invoice_preports, field, data[field])
        # 保存单价和数量
        invoice_preports.qty = qty_data
        invoice_preports.rate = rate_data

        surcharges = {}
        surcharge_notes = {}
        for field in fields:

            surcharge_key = f"{field}_surcharge"
            note_key = f"{field}_surcharge_note"

            surcharge = request.POST.get(surcharge_key, 0) or 0
            note = request.POST.get(note_key, "")
            surcharges[field] = float(surcharge)
            surcharge_notes[field] = note
        invoice_preports.surcharges = surcharges
        invoice_preports.surcharge_notes = surcharge_notes
        invoice_preports.save()
        invoice_preports = InvoicePreport.objects.get(
            invoice_number__invoice_number=invoice.invoice_number,
            invoice_type="receivable",
        )
        # 只要更新了港前拆柜数据，就要计算一次总数，更新invoice的preport
        invoice = Invoice.objects.select_related("container_number").get(
            container_number__container_number=container_number
        )
        invoice.receivable_preport_amount = preport_amount
        invoice.save()
        # 账单状态记录
        order = Order.objects.select_related("retrieval_id", "container_number").get(
            container_number__container_number=container_number
        )

        invoice_status, created = InvoiceStatus.objects.get_or_create(
            container_number=order.container_number, invoice_type="receivable"
        )
        if data.get("pending") == "True":
            # 审核通过，进入库内账单录入
            invoice_status.stage = "warehouse"
            invoice_status.is_rejected = "False"
            invoice_status.reject_reason = ""
        elif data.get("pending") == "False":
            # 审核失败，驳回账单
            invoice_status.is_rejected = "True"
            invoice_status.reject_reason = data.get("invoice_reject_reason", "")
        else:
            # 提拆柜录入完毕,如果是complete表示客服录入完成，订单状态进入下一步，否则不改状态
            if save_type == "complete":
                # 如果这是被财务驳回的，就直接改主状态为待确认，其他不用动
                if invoice_status.is_rejected == True and (
                    invoice_status.stage_public == "delivery_completed"
                    or invoice_status.stage_other == "delivery_completed"
                ):
                    invoice_status.stage = "tobeconfirmed"
                    invoice_status.reject_reason = ""
                else:
                    invoice_status.stage = "preport"
                    invoice_status.reject_reason = ""
                invoice_status.is_rejected = "False"
            elif (
                save_type == "account_complete"
            ):  # 如果是财务从确认界面跳转过来的，就要return回账单确认界面
                modified_get = request.GET.copy()
                modified_get["start_date_confirm"] = request.POST.get(
                    "start_date_confirm"
                )
                modified_get["end_date_confirm"] = request.POST.get("end_date_confirm")
                new_request = request
                new_request.GET = modified_get
                return self.handle_container_invoice_confirm_get(new_request)

        order.save()
        invoice_status.save()
        groups = [group.name for group in request.user.groups.all()]
        if request.user.is_staff:
            groups.append("staff")

        return self.handle_invoice_preport_get(
            request, request.POST.get("start_date"), request.POST.get("end_date")
        )

    def handle_invoice_delivery_type_save(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        invoice = Invoice.objects.get(
            container_number__container_number=container_number
        )
        selections = request.POST.getlist("is_type_added")
        plt_ids = request.POST.getlist("added_plt_ids")
        plt_ids = [id for s, id in zip(selections, plt_ids) if s == "on"]
        alter_type = request.POST.get("alter_type")

        # 如果是改公仓/私仓，是一个方法，否则是选派送类别
        if alter_type == "transferDes":
            # 更改公仓/私仓类别
            plt_delivery_type = request.POST.getlist("plt_delivery_type")
            plt_delivery_type = [
                des for s, des in zip(selections, plt_delivery_type) if s == "on"
            ]
            for i in range(len((plt_ids))):
                ids = plt_ids[i].split(",")
                ids = [int(id) for id in ids]
                pallet = Pallet.objects.filter(id__in=ids)
                updated_pallets = []
                for plt in pallet:
                    plt.delivery_type = plt_delivery_type[i]
                    updated_pallets.append(plt)
                bulk_update_with_history(
                    updated_pallets, Pallet, fields=["delivery_type"]
                )

        else:
            total_cbm = request.POST.getlist("cbm")
            total_cbm = [cbm for s, cbm in zip(selections, total_cbm) if s == "on"]
            total_weight_lbs = request.POST.getlist("weight_lbs")
            total_weight_lbs = [
                weight for s, weight in zip(selections, total_weight_lbs) if s == "on"
            ]
            destination = request.POST.getlist("destination")
            destination = [des for s, des in zip(selections, destination) if s == "on"]
            # 更改amazon/walmart等类别
            delivery_type = request.POST.getlist("delivery_type")
            delivery_type = [
                des for s, des in zip(selections, delivery_type) if s == "on"
            ]
            zipcode = request.POST.getlist("zipcode")
            zipcode = [code for s, code in zip(selections, zipcode) if s == "on"]
            total_pallet = request.POST.getlist("total_pallet")
            total_pallet = [n for s, n in zip(selections, total_pallet) if s == "on"]

            # 将前端的每一条记录存为invoice_delivery的一条
            for i in range(len((plt_ids))):
                ids = plt_ids[i].split(",")
                ids = [int(id) for id in ids]
                pallets = Pallet.objects.filter(id__in=ids)
                current_date = datetime.now().date()
                invoice_delivery = f"{current_date.strftime('%Y-%m-%d').replace('-', '')}-{alter_type}-{destination[i]}-{len(pallets)}"
                invoice_content = InvoiceDelivery(
                    **{
                        "invoice_delivery": invoice_delivery,
                        "invoice_number": invoice,
                        "delivery_type": delivery_type[i],
                        "type": alter_type,  # 沃尔玛/亚马逊等
                        "zipcode": zipcode[i],
                        "destination": destination[i],
                        "total_pallet": total_pallet[i],
                        "total_cbm": total_cbm[i],
                        "total_weight_lbs": total_weight_lbs[i],
                    }
                )
                invoice_content.save()
                # 找单价
                order = Order.objects.select_related(
                    "retrieval_id", "container_number", "vessel_id"
                ).get(container_number__container_number=container_number)
                container_type = order.container_number.container_type
                warehouse = order.retrieval_id.retrieval_destination_area
                vessel_etd = order.vessel_id.vessel_etd
                cutoff_date = date(2025, 4, 1)
                cutoff_datetime = datetime.combine(cutoff_date, datetime_time.min).replace(
                    tzinfo=pytz.UTC
                )
                is_new_rule = vessel_etd >= cutoff_datetime
                customer = order.customer_name
                customer_name = customer.zem_name
                fee_details = self._get_fee_details(
                    warehouse, vessel_etd, customer_name
                )
                self._calculate_and_set_delivery_cost(
                    invoice_content, container_type, fee_details, warehouse, is_new_rule
                )
                if invoice_content.cost is not None:
                    if invoice_content.type != "combine":
                        invoice_content.total_cost = float(
                            invoice_content.cost
                        ) * float(invoice_content.total_pallet)
                    elif invoice_content.type == "combine":  # 组合柜总价=单价
                        invoice_content.total_cost = float(invoice_content.cost)
                    else:
                        invoice_content.total_cost = 0
                invoice_content.save()

                updated_pallets = []
                for plt in pallets:
                    invoices_to_check = set()

                    old_invoice_delivery = plt.invoice_delivery
                    if old_invoice_delivery:
                        invoices_to_check.add(plt.invoice_delivery)
                        # old_invoice_delivery.delete()
                    # try:
                    #     invoice_delivery = plt.invoice_delivery
                    #     if invoice_delivery and hasattr(invoice_delivery, "delete"):
                    #         invoice_delivery.delete()
                    # except InvoiceDelivery.DoesNotExist:
                    #     pass
                    # pallet指向InvoiceDelivery表
                    plt.invoice_delivery = invoice_content
                    updated_pallets.append(plt)
                for inv in invoices_to_check:
                    inv.delete()
                bulk_update_with_history(
                    updated_pallets, Pallet, fields=["invoice_delivery"]
                )

        return self.handle_container_invoice_delivery_get(request)

    def handle_invoice_confirm_save(self, request: HttpRequest) -> tuple[Any, Any]:
        save_type = request.POST.get("save_type", None)
        if save_type == "reject":
            return self.handle_invoice_confirm_get(
                request,
                request.POST.get("start_date_confirm"),
                request.POST.get("end_date_confirm"),
            )
        invoice_type = "receivable"
        container_number = request.POST.get("container_number")
        order = Order.objects.select_related("retrieval_id", "container_number").get(
            container_number__container_number=container_number
        )
        invoice = Invoice.objects.get(
            container_number__container_number=container_number
        )
        # 更新状态
        invoice_status = InvoiceStatus.objects.get(
            container_number=order.container_number, invoice_type="receivable"
        )
        invoice_status.stage = "confirmed"
        invoice_status.save()

        context = self._parse_invoice_excel_data(order, invoice, invoice_type)
        workbook, invoice_data = self._generate_invoice_excel(context)
        invoice.invoice_date = invoice_data["invoice_date"]
        invoice.invoice_link = invoice_data["invoice_link"]
        if order.order_type == "直送":
            invoice.receivable_total_amount = float(
                invoice.receivable_direct_amount or 0
            )
            invoice.remain_offset = float(invoice.receivable_direct_amount or 0)
        else:
            invoice.receivable_total_amount = (
                float(invoice.receivable_preport_amount or 0)
                + float(invoice.receivable_warehouse_amount or 0)
                + float(invoice.receivable_delivery_amount or 0)
            )
            invoice.remain_offset = (
                float(invoice.receivable_preport_amount or 0)
                + float(invoice.receivable_warehouse_amount or 0)
                + float(invoice.receivable_delivery_amount or 0)
            )
        invoice.save()
        receivable_status = order.receivable_status
        receivable_status.stage = "confirmed"
        receivable_status.save()
        order.save()
        return self.handle_invoice_confirm_get(
            request,
            request.POST.get("start_date_confirm"),
            request.POST.get("end_date_confirm"),
        )

    def sum_exe_number(self, warehouse, other_fees_dict, row_data, valid_headers):
        if warehouse.other_fees and isinstance(warehouse.other_fees, dict):
            for key, value in warehouse.other_fees.items():
                if key == "other_fee" and isinstance(value, dict):
                    other_fees_dict.update(value)
                    if other_fees_dict:
                        for k, v in other_fees_dict.items():
                            row_data[k] = v
                            valid_headers.append(k)
                elif isinstance(value, (int, float)):
                    other_fees_dict[key] = value

    def handle_carrier_invoice_export(self, request: HttpRequest) -> HttpResponse:
        select_month = request.POST.get("select_month")
        select_carrier = request.POST.get("select_carrier")
        if not select_month or not select_carrier:
            raise ValueError("请选择月份和供应商")

        year, month = map(int, select_month.split("-"))
        month = month - 1
        # 查找该月份，该供应商的所有费用
        # 先查账单时间满足月份的，就是实际提柜时间+1个月
        order_list = Order.objects.filter(
            retrieval_id__actual_retrieval_timestamp__year=year,
            retrieval_id__actual_retrieval_timestamp__month=month,
            invoice_id__isnull=False,
        ).exclude(payable_status__stage="unstarted", payable_status__isnull=False)
        order_list = order_list.order_by(
            "warehouse", "retrieval_id__actual_retrieval_timestamp"
        )
        orders = []

        if select_carrier in ["BBR", "KNO"]:
            # 如果是BBR和KNO，就只读InvoiceWarehouse表
            for order in order_list:
                try:
                    invoice = Invoice.objects.get(
                        container_number__container_number=order.container_number
                    )
                    invoice_warehouse = InvoiceWarehouse.objects.get(
                        invoice_number__invoice_number=invoice.invoice_number,
                        invoice_type="payable",
                    )
                    if invoice_warehouse.carrier == select_carrier:
                        orders.append((order, invoice, invoice_warehouse, None))
                except InvoiceWarehouse.DoesNotExist:
                    continue
        else:
            if select_carrier == "ARM":
                s_carrier = "大方广"
            elif select_carrier == "Kars":
                s_carrier = "kars"
            else:
                s_carrier = select_carrier

            for order in order_list:
                if order.retrieval_id.retrieval_carrier != s_carrier:
                    continue
                try:
                    invoice = Invoice.objects.get(
                        container_number__container_number=order.container_number
                    )
                    invoice_preport = InvoicePreport.objects.get(
                        invoice_number__invoice_number=invoice.invoice_number,
                        invoice_type="payable",
                    )
                    try:
                        invoice_warehouse = InvoiceWarehouse.objects.get(
                            invoice_number__invoice_number=invoice.invoice_number,
                            invoice_type="payable",
                        )
                    except InvoiceWarehouse.DoesNotExist:
                        invoice_warehouse = None
                    orders.append((order, invoice, invoice_warehouse, invoice_preport))
                except InvoicePreport.DoesNotExist:
                    continue
        if not orders:
            raise ValueError("未查询到符合条件的订单")
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{select_carrier}_{select_month}账单"

        pallet_name = str(
            InvoiceWarehouse._meta.get_field("palletization_fee").verbose_name
        )
        arrive_name = str(InvoiceWarehouse._meta.get_field("arrive_fee").verbose_name)
        fixed_fee_types = []
        if select_carrier in ["BBR", "KNO"]:
            fixed_fee_types = [pallet_name, arrive_name, "总费用"]
        else:
            fixed_fee_types = [
                "基本费用",
                "超重费",
                "车架费",
                pallet_name,
                arrive_name,
                "总费用",
            ]

        valid_headers = ["柜号", "提柜时间", "仓库", "柜型", "总费用"]

        rows = []
        for order, invoice, warehouse, preport in orders:
            row_data = {}
            row_data["柜号"] = order.container_number.container_number
            row_data["提柜时间"] = (
                order.retrieval_id.actual_retrieval_timestamp.strftime("%Y-%m-%d")
            )
            row_data["仓库"] = order.warehouse.name
            row_data["柜型"] = order.container_number.container_type
            other_fees_dict = {}

            if select_carrier in ["BBR", "KNO"]:
                if warehouse:
                    row_data[pallet_name] = warehouse.palletization_fee or 0
                    row_data[arrive_name] = warehouse.arrive_fee or 0

                    # 计算其他费用
                    self.sum_exe_number(
                        warehouse, other_fees_dict, row_data, valid_headers
                    )
                    row_data["总费用"] = getattr(warehouse, "amount", 0) or 0
            else:
                total_amount = 0

                if preport:
                    total_amount += float(preport.amount or 0)
                    row_data["基本费用"] = preport.pickup or 0
                    row_data["超重费"] = preport.over_weight or 0
                    row_data["车架费"] = preport.chassis or 0

                    # 计算preport的其他费用
                    if preport.other_fees and isinstance(preport.other_fees, dict):
                        for key, value in preport.other_fees.items():
                            if key == "other_fee" and isinstance(value, dict):
                                other_fees_dict.update(value)
                                if other_fees_dict:
                                    for k, v in other_fees_dict.items():
                                        row_data[k] = v
                                        valid_headers.append(k)
                            elif isinstance(value, (int, float)):
                                other_fees_dict[key] = value

                if warehouse:
                    total_amount += float(getattr(warehouse, "amount", 0) or 0)
                    row_data[pallet_name] = warehouse.palletization_fee or 0
                    row_data[arrive_name] = warehouse.arrive_fee or 0
                    # 计算warehouse的其他费用
                    other_fees_dict = {}
                    self.sum_exe_number(
                        warehouse, other_fees_dict, row_data, valid_headers
                    )
                row_data["总费用"] = total_amount
            for fee_type in fixed_fee_types:
                if fee_type not in row_data:
                    row_data[fee_type] = 0
            rows.append(row_data)
        valid_headers = list(dict.fromkeys(valid_headers))
        for fee_type in fixed_fee_types:
            if fee_type not in ["柜号", "总费用"]:
                # 检查该列是否有非零值
                has_value = any(row.get(fee_type, 0) not in (0, "0") for row in rows)
                if has_value:
                    valid_headers.append(fee_type)
        valid_headers = list(dict.fromkeys(valid_headers))
        # 写入表头
        ws.append(valid_headers)

        # 写入数据行
        for row in rows:
            row_data = [row.get(col, 0) for col in valid_headers]
            ws.append(row_data)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{select_carrier}_账单_{select_month}.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"

        wb.save(response)
        return response

    def handle_invoice_confirm_combina_save(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        invoice_number = request.POST.get("invoice_number")
        invoice = Invoice.objects.get(invoice_number=invoice_number)
        total_fee = float(request.POST.get("totalAmount", 0))
        base_fee = float(request.POST.get("base_fee", 0))
        overweight_fee = float(request.POST.get("overweight_fee", 0))
        overpallet_fee = float(request.POST.get("overpallet_fee", 0))
        overregion_pickup_fee = float(request.POST.get("overregion_pickup_fee", 0))
        overregion_delivery_fee = float(request.POST.get("overregion_delivery_fee", 0))

        plts_by_destination = (
            Pallet.objects.filter(container_number__container_number=container_number)
            .values("destination")
            .annotate(total_cbm=Sum("cbm"), total_weight=Sum("weight_lbs"))
        )
        invoice_item_data = []

        combina_data_des_key = request.POST.getlist("combina_data_des_key")
        combina_data_des_cbm = request.POST.getlist("combina_data_des_cbm")
        combina_data_des_price = request.POST.getlist("combina_data_des_price")
        combina_data_des_location = request.POST.getlist("combina_data_des_location")
        combina_data_des_rate = request.POST.getlist("combina_data_des_rate")
        base_location = []
        for i in range(len(combina_data_des_key)):
            location = combina_data_des_location[i].split(",")
            for num in range(len(location)):
                cbm_d = 0.000
                weight_d = 0.000
                base_location.append(location[num])
                if num == 0:
                    qty = float(combina_data_des_rate[i])
                    rate = float(combina_data_des_price[i])
                    amount = qty * rate
                else:
                    qty = rate = amount = 0.00
                for item in plts_by_destination:

                    cleaned_item_dest = item["destination"].strip()
                    cleaned_location = location[num].strip()
                    if cleaned_location in cleaned_item_dest:
                        cbm_d = item["total_cbm"]
                        weight_d = item["total_weight"]
                        break
                invoice_item_data.append(
                    {
                        "invoice_number": invoice,
                        "description": "派送费",
                        "warehouse_code": location[num],
                        "cbm": cbm_d,
                        "weight": weight_d,
                        "qty": qty,
                        "rate": rate,
                        "amount": amount,
                        "note": combina_data_des_key[i],
                    }
                )

        if overweight_fee > 0:
            overweight_extra_weight = request.POST.get("overweight_extra_weight")
            invoice_item_data.append(
                {
                    "invoice_number": invoice,
                    "description": "超重费",
                    "warehouse_code": None,
                    "cbm": None,
                    "weight": overweight_extra_weight,
                    "qty": 1.0,
                    "rate": overweight_fee,
                    "amount": overweight_fee,
                    "note": None,
                }
            )

        if overpallet_fee > 0:
            current_pallets = request.POST.get("current_pallets")
            limit_pallets = request.POST.get("limit_pallets")
            over_count = float(current_pallets) - float(limit_pallets)
            invoice_item_data.append(
                {
                    "invoice_number": invoice,
                    "description": "超板费",
                    "warehouse_code": None,
                    "cbm": None,
                    "weight": None,
                    "qty": over_count,
                    "rate": overpallet_fee / over_count if over_count > 0 else 0,
                    "amount": overpallet_fee,
                    "note": None,
                }
            )

        if overregion_pickup_fee > 0:
            overregion_pickup_non_combina_cbm_ratio = request.POST.get(
                "overregion_pickup_non_combina_cbm_ratio"
            )
            overregion_pickup_non_combina_base_fee = request.POST.get(
                "overregion_pickup_non_combina_base_fee"
            )
            invoice_item_data.append(
                {
                    "invoice_number": invoice,
                    "description": "提拆费",
                    "warehouse_code": None,
                    "cbm": None,
                    "weight": None,
                    "qty": float(overregion_pickup_non_combina_cbm_ratio) / 100,
                    "rate": overregion_pickup_non_combina_base_fee,
                    "amount": overregion_pickup_fee,
                    "note": None,
                }
            )
        if overregion_delivery_fee > 0:
            overregion_delivery_destination = request.POST.getlist(
                "overregion_delivery_destination"
            )
            overregion_delivery_pallets = request.POST.getlist(
                "overregion_delivery_pallets"
            )
            overregion_delivery_cbm = request.POST.getlist("overregion_delivery_cbm")
            overregion_delivery_price = request.POST.getlist(
                "overregion_delivery_price"
            )
            overregion_delivery_subtotal = request.POST.getlist(
                "overregion_delivery_subtotal"
            )
            for i in range(len(overregion_delivery_destination)):
                if overregion_delivery_destination[i] not in base_location:
                    invoice_item_data.append(
                        {
                            "invoice_number": invoice,
                            "description": "超区派送费",
                            "warehouse_code": overregion_delivery_destination[i],
                            "cbm": overregion_delivery_cbm[i],
                            "weight": None,
                            "qty": overregion_delivery_pallets[i],
                            "rate": overregion_delivery_price[i],
                            "amount": overregion_delivery_subtotal[i],
                            "note": None,
                        }
                    )

        # 客服手动录入的额外费用
        i = 0
        while f"port_fees[{i}][price]" in request.POST:
            rate = float(request.POST.get(f"port_fees[{i}][price]"))
            qty = float(request.POST.get(f"port_fees[{i}][quantity]"))
            amount = float(request.POST.get(f"port_fees[{i}][value]", 0)) + (
                float(request.POST.get(f"port_fees[{i}][surcharge]"))
                if request.POST.get(f"port_fees[{i}][surcharge]")
                else 0
            )
            if qty > 0 and rate > 0 and amount > 0:
                invoice_item_data.append(
                    {
                        "invoice_number": invoice,
                        "description": request.POST.get(f"port_fees[{i}][name]"),
                        "warehouse_code": None,
                        "cbm": None,
                        "weight": None,
                        "rate": rate,
                        "qty": qty,
                        "amount": amount,
                        "note": request.POST.get(f"port_fees[{i}][surcharge_note]"),
                    }
                )
            i += 1

        j = 0
        while f"warehouse_fees[{j}][price]" in request.POST:
            rate = float(request.POST.get(f"warehouse_fees[{j}][price]"))
            qty = float(request.POST.get(f"warehouse_fees[{j}][quantity]"))
            amount = float(request.POST.get(f"warehouse_fees[{j}][value]", 0)) + (
                float(request.POST.get(f"warehouse_fees[{j}][surcharge]"))
                if request.POST.get(f"warehouse_fees[{j}][surcharge]")
                else 0
            )
            if qty > 0 and rate > 0 and amount > 0:
                invoice_item_data.append(
                    {
                        "invoice_number": invoice,
                        "description": request.POST.get(f"warehouse_fees[{j}][name]"),
                        "warehouse_code": None,
                        "cbm": None,
                        "weight": None,
                        "rate": rate,
                        "qty": qty,
                        "amount": amount,
                        "note": request.POST.get(
                            f"warehouse_fees[{j}][surcharge_note]"
                        ),
                    }
                )
            j += 1

        k = 0
        while f"deliverys[{k}][destination]" in request.POST:
            qty = float(request.POST.get(f"deliverys[{k}][total_pallet]"))
            rate = float(request.POST.get(f"deliverys[{k}][cost]"))
            amount = float(request.POST.get(f"deliverys[{k}][total_cost]"))
            if qty > 0 and amount > 0:
                invoice_item_data.append(
                    {
                        "invoice_number": invoice,
                        "description": "超出仓点",
                        "warehouse_code": request.POST.get(
                            f"deliverys[{k}][destination]"
                        ),
                        "cbm": request.POST.get(f"deliverys[{k}][total_cbm]"),
                        "weight": request.POST.get(f"deliverys[{k}][total_weight_lbs]"),
                        "qty": qty,
                        "rate": rate,
                        "amount": amount,
                        "note": request.POST.get(f"deliverys[{k}][note]"),
                    }
                )
            k += 1

        invoice_item = InvoiceItem.objects.filter(
            invoice_number__invoice_number=invoice.invoice_number
        )
        invoice_item.delete()

        invoice_item_instances = [
            InvoiceItem(**inv_itm_data) for inv_itm_data in invoice_item_data
        ]
        bulk_create_with_history(invoice_item_instances, InvoiceItem)

        order = Order.objects.get(container_number__container_number=container_number)

        context = self._parse_invoice_excel_data(order, invoice, "receivable")
        workbook, invoice_data = self._generate_invoice_excel(context)
        invoice.invoice_date = invoice_data["invoice_date"]
        invoice.invoice_link = invoice_data["invoice_link"]
        invoice.receivable_total_amount = total_fee
        invoice.remain_offset = total_fee
        invoice.save()

        if not order.invoice_id:
            order.invoice_id = invoice
            order.save()
        save_type = request.POST.get("save_type")
        if save_type == "complete":  # 暂存的不改变状态
            receivable_status = order.receivable_status
            receivable_status.stage = "confirmed"
            receivable_status.save()
        order.save()
        is_from_account_confirmation = request.POST.get("is_from_account_confirmation")
        if is_from_account_confirmation:
            # 返回账单确认界面
            return self.handle_invoice_confirm_get(
                request,
                request.POST.get("start_date_confirm"),
                request.POST.get("end_date_confirm"),
            )
        else:
            # 返回组合柜列表
            return self.handle_invoice_combina_get(
                request,
                request.POST.get("start_date_confirm"),
                request.POST.get("end_date_confirm"),
                request.POST.get("customer"),
                request.POST.get("warehouse"),
            )

    def handle_invoice_dismiss_save(self, request: HttpRequest) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        status = request.POST.get("status")
        start_date_confirm = request.POST.get("start_date_confirm")
        end_date_confirm = request.POST.get("end_date_confirm")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        reject_reason = request.POST.get("reject_reason")
        order = Order.objects.select_related("container_number").get(
            container_number__container_number=container_number
        )
        # 更新状态
        invoice_type = request.POST.get("invoice_type")
        if invoice_type == "receivable":
            invoice_status = InvoiceStatus.objects.get(
                container_number=order.container_number, invoice_type="receivable"
            )
        elif invoice_type == "payable":
            invoice_status = InvoiceStatus.objects.get(
                container_number=order.container_number, invoice_type="payable"
            )
        invoice_status.stage = status
        if status == "warehouse":
            delivery_type = request.POST.get("delivery_type")
            if delivery_type == "public":
                invoice_status.stage_public = "warehouse_rejected"
            elif delivery_type == "other":
                invoice_status.stage_other = "warehouse_rejected"
        elif status == "delivery":
            # 检查时驳回公仓还是私仓
            reject_type = request.POST.get("reject_type")
            delivery_type = request.POST.get("delivery_type")
            if reject_type == "public" or delivery_type == "public":
                invoice_status.stage_public = "delivery_rejected"
            else:
                invoice_status.stage_other = "delivery_rejected"
        invoice_status.is_rejected = "True"
        invoice_status.reject_reason = reject_reason
        invoice_status.save()
        if (
            start_date_confirm
            and end_date_confirm
            and start_date_confirm != "None"
            and end_date_confirm != "None"
        ):
            return self.handle_invoice_confirm_get(
                request, start_date_confirm, end_date_confirm
            )
        elif start_date and end_date:
            return self.handle_invoice_combina_get(request, start_date, end_date)
        else:
            raise ValueError("缺少起止日期")

    def handle_invoice_redirect_post(self, request: HttpRequest) -> tuple[Any, Any]:
        status = request.POST.get("status")
        if status == "preport":
            return self.handle_container_invoice_preport_get(request)
        elif status == "warehouse":
            delivery_type = request.POST.get("delivery_type")
            return self.handle_container_invoice_warehouse_get(request, delivery_type)
        elif status == "delivery":
            return self.handle_container_invoice_delivery_get(request)
        elif status == "direct":
            return self.handle_container_invoice_direct_get(request)

    def handle_swicth_delivery_save(self, request: HttpRequest) -> tuple[Any, Any]:
        # 这个方法是，切换公仓派送的类别，如果当前是组合柜，就切换到亚马逊/沃尔玛，反之相同
        container_number = request.POST.get("container_number")
        public_pallets = Pallet.objects.filter(
            container_number__container_number=container_number,
            delivery_type="public",
            invoice_delivery__isnull=False,
        )
        has_combine = public_pallets.filter(invoice_delivery__type="combine").exists()
        # has_other = public_pallets.exclude(invoice_delivery__type="combine").exists()

        # 决定切换方向：如果有组合柜且没有其他类型，切换到非组合柜；否则切换到组合柜
        iscombina = not has_combine

        order = Order.objects.select_related(
            "retrieval_id", "container_number", "vessel_id"
        ).get(container_number__container_number=container_number)
        warehouse = order.retrieval_id.retrieval_destination_area
        vessel_etd = order.vessel_id.vessel_etd
        is_transfer = order.order_type == "转运"
        customer = order.customer_name
        customer_name = customer.zem_name
        fee_details = self._get_fee_details(warehouse, vessel_etd, customer_name)
        self._auto_classify_pallet(
            container_number, fee_details, warehouse, True, iscombina, is_transfer
        )
        return self.handle_container_invoice_delivery_get(request)

    def handle_invoice_delivery_save(self, request: HttpRequest) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        delivery_type = request.POST.get("delivery_type")
        total_cost = request.POST.getlist("total_cost")
        cost = request.POST.getlist("cost")
        total_pallet = request.POST.getlist("total_pallet")
        po_activation = request.POST.getlist("po_activation")
        type_value = request.POST.get("type")
        redirect_step = request.POST.get("redirect_step")
        invoice = Invoice.objects.select_related("customer", "container_number").get(
            container_number__container_number=container_number
        )
        # 如果派送方式都填完了，invoice记录派送价格和账单状态
        if type_value == "amount":
            delivery_amount = (
                InvoiceDelivery.objects.filter(
                    invoice_number=invoice,
                    invoice_type="receivable",
                ).aggregate(total_amount=Sum("total_cost"))["total_amount"]
                or 0
            )
            invoice.receivable_delivery_amount = delivery_amount
            # 派送账单确认的时候，就计算一遍总费用
            invoice.receivable_total_amount = (
                float(invoice.receivable_preport_amount or 0)
                + float(invoice.receivable_warehouse_amount or 0)
                + float(invoice.receivable_delivery_amount or 0)
            )
            invoice.save()

            order = Order.objects.select_related(
                "retrieval_id", "container_number"
            ).get(container_number__container_number=container_number)
            if redirect_step == "False":
                invoice_status = InvoiceStatus.objects.get(
                    container_number=order.container_number,
                    invoice_type="receivable",
                )
                container_delivery_type = invoice_status.container_number.delivery_type
                groups = [group.name for group in request.user.groups.all()]
                # 如果不是从财务确认界面跳转来的，才需要改变状态
                if (
                    "mix_account" in groups and "NJ_mix_account" not in groups
                ):  # 公仓//私仓权限都有的
                    invoice_status.stage_public = "delivery_completed"
                    invoice_status.stage_other = "delivery_completed"
                    invoice_status.stage = "tobeconfirmed"
                    invoice_status.is_rejected = False
                    invoice_status.reject_reason = ""
                else:
                    if container_delivery_type in ["public", "other"]:
                        # 如果这个柜子只有一类仓，就直接改变状态
                        invoice_status.stage = "tobeconfirmed"
                        invoice_status.is_rejected = False
                        invoice_status.reject_reason = ""
                        invoice_status.stage_public = "delivery_completed"
                        invoice_status.stage_other = "delivery_completed"
                    elif container_delivery_type == "mixed":
                        if delivery_type == "public":
                            # 公仓组录完了，改变stage_public
                            invoice_status.stage_public = "delivery_completed"
                            # 如果私仓也做完了，就改变主状态到派送阶段
                            if invoice_status.stage_other == "delivery_completed":
                                invoice_status.stage = "tobeconfirmed"
                        elif delivery_type == "other":
                            # 私仓租录完了，改变stage_other
                            invoice_status.stage_other = "delivery_completed"
                            # 如果公仓也做完了，就改变主状态
                            if invoice_status.stage_public == "delivery_completed":
                                invoice_status.stage = "tobeconfirmed"
                        invoice_status.is_rejected = False
                        invoice_status.reject_reason = ""

                invoice_status.save()
        else:
            # 记录其中一种派送方式到invoice_delivery表
            plt_ids = request.POST.getlist("plt_ids")
            new_plt_ids = [ast.literal_eval(sub_plt_id) for sub_plt_id in plt_ids]
            expense = request.POST.getlist("expense")
            # 有的类型有备注，有的没有
            note = request.POST.getlist("note") if "note" in request.POST else None
            # 将前端的每一条记录存为invoice_delivery的一条
            for i in range(len((new_plt_ids))):
                ids = [int(id) for id in new_plt_ids[i]]
                pallet = Pallet.objects.filter(id__in=ids)
                # 因为每一条记录中所有的板子都是对应一条invoice_delivery，建表的时候就是这样存的，所以取其中一个的外键就可以
                pallet_obj = pallet[0]
                invoice_content = pallet_obj.invoice_delivery
                # 除价格外，其他在新建记录的时候就存了
                if total_cost[i] is None:
                    raise ValueError("总价为空")
                invoice_content.total_cost = total_cost[i]
                invoice_content.cost = cost[i]
                invoice_content.total_pallet = total_pallet[i]
                if expense[i]:
                    invoice_content.expense = expense[i]
                if po_activation[i]:
                    invoice_content.po_activation = po_activation[i]
                if note:
                    if "None" in note[i]:
                        invoice_content.note = None
                    else:
                        invoice_content.note = note[i]
                invoice_content.save()
        # 如果是财务确认界面跳转的，需要重定向到财务确认界面，并且执行派送界面的账单确认操作
        if redirect_step == "True":
            # 派送界面，一种派送方式点确认后，自动计算总费用
            return self.handle_container_invoice_confirm_get(request)
        else:
            return self.handle_invoice_delivery_get(request)

    def handle_container_invoice_warehouse_get(
        self,
        request: HttpRequest,
        delivery_type: str = None,
    ) -> tuple[Any, Any]:
        container_number = request.GET.get("container_number")
        order = Order.objects.select_related("retrieval_id", "container_number").get(
            container_number__container_number=container_number
        )
        warehouse = order.retrieval_id.retrieval_destination_area
        vessel_etd = order.vessel_id.vessel_etd

        customer = order.customer_name
        quotation = (
            QuotationMaster.objects.filter(
                effective_date__lte=vessel_etd,
                is_user_exclusive=True,
                exclusive_user=customer.zem_name,
            )
            .order_by("-effective_date")
            .first()
        )
        if not quotation:
            quotation = (
                QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=False,  # 非用户专属的通用报价单
                )
                .order_by("-effective_date")
                .first()
            )
        if not quotation:
            raise ValueError("找不到报价表")
        WAREHOUSE_FEE = FeeDetail.objects.get(
            quotation_id=quotation.id, fee_type="warehouse"
        )
        # 提拆、打托缠膜费用

        # FS_constrain = {
        #     key: float(re.search(r'\d+(\.\d+)?', value).group())
        #     for key, value in WAREHOUSE_FEE.details.items()
        #     if '/' in str(value) and re.search(r'\d+(\.\d+)?', value)
        # }
        FS_constrain = {}
        for (
            key,
            value,
        ) in (
            WAREHOUSE_FEE.details.items()
        ):  # 把details里面的键值对改成值是纯数字的，用于在费用表单提交前，验证数据合规性
            if not isinstance(value, dict):
                match = re.findall(r"\$(\d+(\.\d+)?)", str(value))
                if match and len(match) == 1:
                    if "（" in str(key):
                        key = key.split("（")[0]
                    FS_constrain[key] = float(match[0][0])
        fs_json = json.dumps(FS_constrain, ensure_ascii=False)
        # 其他费用
        FS = {
            "sorting": f"{WAREHOUSE_FEE.details.get('分拣费', 'N/A')}",  # 分拣费
            "intercept": f"{WAREHOUSE_FEE.details.get('拦截费', 'N/A')}",  # 拦截费
            "po_activation": f"{WAREHOUSE_FEE.details.get('亚马逊PO激活', 'N/A')}",  # 拦截费
            "self_pickup": f"{WAREHOUSE_FEE.details.get('客户自提', 'N/A')}",  # 客户自提
            "re_pallet": f"{WAREHOUSE_FEE.details.get('重新打板', 'N/A')}",  # 重新打板
            "counting": f"{WAREHOUSE_FEE.details.get('货品清点费', 'N/A')}",  # 货品清点费
            "warehouse_rent": WAREHOUSE_FEE.details.get("仓租", "N/A"),  # 仓租
            "specified_labeling": f"{WAREHOUSE_FEE.details.get('指定贴标', 'N/A')}",  # 指定贴标
            "inner_outer_box": f"{WAREHOUSE_FEE.details.get('内外箱', 'N/A')}",  # 内外箱
            "pallet_label": f"{WAREHOUSE_FEE.details.get('托盘标签', 'N/A')}",  # 内外箱
            "open_close_box": f"{WAREHOUSE_FEE.details.get('开封箱', 'N/A')}",  # 开封箱
            "destroy": f"{WAREHOUSE_FEE.details.get('销毁', 'N/A')}",  # 销毁
            "take_photo": f"{WAREHOUSE_FEE.details.get('拍照', 'N/A')}",  # 拍照
            "take_video": f"{WAREHOUSE_FEE.details.get('拍视频', 'N/A')}",  # 拍视频
            "repeated_operation_fee": f"{WAREHOUSE_FEE.details.get('重复操作费', 'N/A')}",  # 重复操作费
        }

        invoice = Invoice.objects.select_related("customer", "container_number").get(
            container_number__container_number=container_number
        )
        groups = [group.name for group in request.user.groups.all()]
        if not delivery_type:
            delivery_type = request.GET.get("delivery_type")
        if delivery_type is None:
            # 确定delivery_type
            if "warehouse_public" in groups and "warehouse_other" not in groups:
                delivery_type = "public"
            elif "warehouse_other" in groups and "warehouse_public" not in groups:
                delivery_type = "other"
        # 不需要赋值单价的字段
        excluded_fields = {
            "id",
            "invoice_number",
            "invoice_type",
            "delivery_type",
            "amount",
            "qty",
            "rate",
            "other_fees",
            "surcharges",
            "surcharge_notes",
            "history",
        }

        try:
            invoice_warehouse = InvoiceWarehouse.objects.get(
                invoice_number__invoice_number=invoice.invoice_number,
                invoice_type="receivable",
                delivery_type=delivery_type,
            )
        except InvoiceWarehouse.DoesNotExist:
            invoice_warehouse, created = InvoiceWarehouse.objects.get_or_create(
                invoice_number=invoice,
                invoice_type="receivable",
                delivery_type=delivery_type,
            )
            qty_data, rate_data = self._extract_unit_price(
                model=InvoiceWarehouse,
                unit_prices=FS_constrain,
                pickup_fee=None,
                excluded_fields=excluded_fields,
            )
            invoice_warehouse.qty = qty_data
            invoice_warehouse.rate = rate_data
            invoice_warehouse.save()

            context = {
                "warehouse": request.GET.get("warehouse"),
                "invoice": invoice,
                "invoice_type": "receivable",
                "container_number": container_number,
                "FS": FS,
                "fs_json": fs_json,
                "delivery_type": delivery_type,
                "start_date": request.GET.get("start_date"),
                "end_date": request.GET.get("end_date"),
                "invoice_type": "receivable",
                "invoice_warehouse": invoice_warehouse,
                "surcharges": invoice_warehouse.surcharges,
                "surcharge_notes": invoice_warehouse.surcharge_notes,
            }
            return self.template_invoice_warehouse_edit, context
        # 如果单价和数量都为空的话，就初始化
        if not invoice_warehouse.qty and not invoice_warehouse.rate:
            qty_data, rate_data = self._extract_unit_price(
                model=InvoiceWarehouse,
                unit_prices=FS_constrain,
                pickup_fee=None,
                excluded_fields=excluded_fields,
            )
            invoice_warehouse.qty = qty_data
            invoice_warehouse.rate = rate_data
            invoice_warehouse.save()
        step = request.POST.get("step")
        redirect_step = step == "redirect"
        context = {
            "warehouse": request.GET.get("warehouse"),
            "invoice_warehouse": invoice_warehouse,
            "invoice": invoice,
            "container_number": container_number,
            "surcharges": invoice_warehouse.surcharges,
            "surcharge_notes": invoice_warehouse.surcharge_notes,
            "start_date": request.GET.get("start_date"),
            "end_date": request.GET.get("end_date"),
            "redirect_step": redirect_step,
            "FS": FS,
            "fs_json": fs_json,
            "status": order.receivable_status.stage,
            "start_date_confirm": request.POST.get("start_date_confirm") or None,
            "end_date_confirm": request.POST.get("end_date_confirm") or None,
            "invoice_type": "receivable",
            "delivery_type": delivery_type,
            "invoice_warehouse": invoice_warehouse,
        }
        return self.template_invoice_warehouse_edit, context

    def get_orders(self, criteria):
        """审核应付看到的"""
        payable_preport_subq = InvoicePreport.objects.filter(
            invoice_number=OuterRef("invoice_id"),
            invoice_type="payable"
        ).only("pickup", "over_weight", "chassis", "demurrage", "per_diem", "other_fees")

        warehouse_subq = InvoiceWarehouse.objects.filter(
            invoice_number=OuterRef("invoice_id"),
            invoice_type="payable"
        ).annotate(
            total_fee=Coalesce(F('palletization_fee'), 0.0) + Coalesce(F('arrive_fee'), 0.0)
        ).only("total_fee", "other_fees")

        order_pending = (
            Order.objects.select_related(
                "customer_name",
                "container_number",
                "invoice_id",
                "retrieval_id",
                "vessel_id",
                "payable_status"
            )
            .filter(
                criteria,
                payable_status__stage="preport",
                payable_status__invoice_type="payable"
            )
            .annotate(
                reject_priority=Case(
                    When(payable_status__is_rejected=True, then=Value(1)),
                    default=Value(2),
                    output_field=IntegerField(),
                ),
                basic_fee=Coalesce(
                    Cast(Subquery(payable_preport_subq.values("pickup")[:1]), FloatField()),
                    0.0
                ),
                overweight_fee=Coalesce(
                    Cast(Subquery(payable_preport_subq.values("over_weight")[:1]), FloatField()),
                    0.0
                ),
                chassis_fee=Coalesce(
                    Cast(Subquery(payable_preport_subq.values("chassis")[:1]), FloatField()),
                    0.0
                ),
                demurrage_fee=Coalesce(
                    Cast(Subquery(payable_preport_subq.values("demurrage")[:1]), FloatField()),
                    0.0
                ),
                per_diem_fee=Coalesce(
                    Cast(Subquery(payable_preport_subq.values("per_diem")[:1]), FloatField()),
                    0.0
                ),
                palletization_fee=Coalesce(
                    Cast(Subquery(warehouse_subq.values("total_fee")[:1]), FloatField()),
                    0.0
                ),
                total_amount=Coalesce(
                    Cast(F("invoice_id__payable_total_amount"), FloatField()),
                    0.0
                ),
                preport_other_fees=Subquery(payable_preport_subq.values("other_fees")[:1]),
                warehouse_other_fees=Subquery(warehouse_subq.values("other_fees")[:1])
            )
            .order_by("reject_priority")
        )

        orders = list(order_pending.iterator(chunk_size=200))

        def parse_fee_data(data):
            """提取other_fee的工具函数，减少重复代码"""
            if not data:
                return {}
            try:
                parsed = json.loads(data) if isinstance(data, str) else data
                return parsed.get("other_fee", {}) if isinstance(parsed, dict) else {}
            except (json.JSONDecodeError, TypeError):
                return {}

        for order in orders:
            preport_fees = order.preport_other_fees
            warehouse_fees = order.warehouse_other_fees
            order.other_fee = {
                **parse_fee_data(preport_fees), **parse_fee_data(warehouse_fees)
            }

        # 已审核账单查询
        pre_order_pending = Order.objects.select_related(
            "customer_name",
            "container_number",
            "invoice_id",
            "invoice_id__statement_id",
            "payable_status",
            "vessel_id",
        ).filter(
            criteria,
            payable_status__stage__in=["tobeconfirmed", "confirmed"],
            payable_status__isnull=False,
        ).only(
            "id", "customer_name", "container_number", "invoice_id",
            "payable_status", "vessel_id"
        )

        return orders, pre_order_pending

    def handle_invoice_payable_get(
        self,
        request: HttpRequest,
        start_date: str = None,
        end_date: str = None,
        customer: str = None,
        warehouse: str = None,
        preport_carrier: str = None,
    ) -> tuple[Any, Any]:
        # 库内操作费
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date

        criteria = (
            models.Q(cancel_notification=False)
            & models.Q(
                retrieval_id__actual_retrieval_timestamp__gte=start_date,
                retrieval_id__actual_retrieval_timestamp__lte=end_date,
            )
            & models.Q(retrieval_id__empty_returned=True)
        )
        if warehouse and warehouse != "None":
            if warehouse == "直送":
                criteria &= models.Q(order_type="直送")
            else:
                # 这是非直送的筛选，标记一下加上直送的筛选
                criteria &= models.Q(
                    retrieval_id__retrieval_destination_precise=warehouse
                )
        if customer:
            criteria &= models.Q(customer_name__zem_name=customer)
        if preport_carrier:
            criteria &= models.Q(retrieval_id__retrieval_carrier=preport_carrier)
        # 待录入的订单
        order = (
            Order.objects.select_related(
                "invoice_id",
                "customer_name",
                "container_number",
                "invoice_id__statement_id",
                "vessel_id",
                "retrieval_id",
            )
            .filter(criteria)
            .filter(
                models.Q(payable_status__is_rejected=True)
                | models.Q(payable_status__isnull=True)
                | models.Q(payable_status__stage="unstarted")
            )
            .annotate(
                reject_priority=Case(
                    When(payable_status__is_rejected=True, then=Value(1)),
                    default=Value(2),
                    output_field=IntegerField(),
                )
            )
            .order_by("reject_priority")
        )
        # 先判断权限，如果是初级审核应付账单权限，状态就是preport
        order_pending = None
        pre_order_pending = None
        previous_order = None

        is_payable_check = self._validate_user_invoice_payable_check(request.user)
        if is_payable_check:  # 审核应付看到的
            #将应付的费用直接加到审核的列表上
            order_pending, pre_order_pending =self.get_orders(criteria)

        if not is_payable_check or request.user.is_staff:
            # 查找客服已录入账单
            previous_order = Order.objects.select_related(
                "customer_name",
                "container_number",
                "invoice_id",
                "invoice_id__statement_id",
                "payable_status",
                "vessel_id",
            ).filter(
                criteria,
                ~models.Q(payable_status__is_rejected=True),
                payable_status__stage__in=["preport", "tobeconfirmed", "confirmed"],
                payable_status__isnull=False,
            )
            previous_order = self.process_orders_display_status(
                previous_order, "payable"
            )
        groups = [group.name for group in request.user.groups.all()]
        if request.user.is_staff:
            groups.append("staff")

        # 导出excel的筛选框
        current_year = datetime.now().year
        current_month = datetime.now().month + 1
        months = []
        for i in range(12):
            month = (current_month - i - 1) % 12 + 1
            year = current_year if (current_month - i - 1) >= 0 else current_year - 1
            months.append(
                {
                    "value": f"{year}-{month:02d}",
                    "label": f"{year}年{month}月",
                    "selected": (month == current_month and year == current_year),
                }
            )
        carriers = {
            "": "",
            "Kars": "Kars",
            "东海岸": "东海岸",
            "ARM": "ARM",
            "BBR": "BBR",
            "KNO": "KNO",
            "GM": "GM",
            "BEST": "BEST",
        }
        context = {
            "order": order,
            "order_form": OrderForm(),
            "previous_order": previous_order,
            "order_pending": order_pending,
            "pre_order_pending": pre_order_pending,
            "start_date": start_date,
            "end_date": end_date,
            "customer": customer,
            "groups": groups,
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
            "is_payable_check": is_payable_check,
            "months": reversed(months) if months else [],
            "carriers": carriers,
            "preport_carrier": CONTAINER_PICKUP_CARRIER,
            "preport_carrier_filter": preport_carrier,
        }
        return self.template_invoice_payable, context

    def handle_container_invoice_confirm_get(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.GET.get("container_number")
        start_date_confirm = request.GET.get("start_date_confirm")
        end_date_confirm = request.GET.get("end_date_confirm")
        invoice_type = request.GET.get("invoice_type")
        invoice = Invoice.objects.get(
            container_number__container_number=container_number
        )
        order = Order.objects.select_related("container_number").get(
            container_number__container_number=container_number
        )
        if invoice_type == "receivable":
            # 这里要区分一下，如果是组合柜的柜子，跳转就直接跳转到组合柜计算界面
            if (
                order.order_type == "转运组合"
                and order.container_number.account_order_type == "转运组合"
            ):
                # 这里表示是组合柜的方式计算，因为如果是报的组合柜但是不符合组合柜要求，那么account_order_type就是转运了
                setattr(request, "is_from_account_confirmation", True)
                return self.handle_container_invoice_combina_get(request)
            else:
                invoice_preports = InvoicePreport.objects.get(
                    invoice_number__invoice_number=invoice.invoice_number,
                    invoice_type=invoice_type,
                )
                if order.order_type in ["转运", "转运组合"]:
                    invoice_warehouses = InvoiceWarehouse.objects.filter(
                        invoice_number__invoice_number=invoice.invoice_number,
                        invoice_type=invoice_type,
                    )
                    invoice_delivery = InvoiceDelivery.objects.filter(
                        invoice_number__invoice_number=invoice.invoice_number,
                        invoice_type=invoice_type,
                    )
                    amazon = []
                    local = []
                    combine = []
                    walmart = []
                    selfdelivery = []
                    upsdelivery = []
                    selfpickup = []
                    for delivery in invoice_delivery:
                        if delivery.type == "amazon":
                            amazon.append(delivery)
                        elif delivery.type == "local":
                            local.append(delivery)
                        elif delivery.type == "combine":
                            combine.append(delivery)
                        elif delivery.type == "walmart":
                            walmart.append(delivery)
                        elif delivery.type == "selfdelivery":
                            selfdelivery.append(delivery)
                        elif delivery.type == "upsdelivery":
                            upsdelivery.append(delivery)
                        elif delivery.type == "selfpickup":
                            selfpickup.append(delivery)
                    container = Container.objects.get(container_number=container_number)
                    if (
                        order.order_type == "转运组合"
                        and container.account_order_type == "转运"
                    ):
                        non_combina_reason = container.non_combina_reason
                    else:
                        non_combina_reason = None
                    context = {
                        "invoice": invoice,
                        "order_type": order.order_type,
                        "invoice_preports": invoice_preports,
                        "invoice_warehouses": invoice_warehouses,
                        "amazon": amazon,
                        "local": local,
                        "combine": combine,
                        "walmart": walmart,
                        "selfdelivery": selfdelivery,
                        "upsdelivery": upsdelivery,
                        "selfpickup": selfpickup,
                        "container_number": container_number,
                        "start_date_confirm": start_date_confirm,
                        "end_date_confirm": end_date_confirm,
                        "invoice_type": invoice_type,
                        "non_combina_reason": non_combina_reason,
                        "delivery_amount": getattr(
                            invoice, "receivable_delivery_amount", 0
                        ),
                        "total_amount": getattr(invoice, "receivable_total_amount", 0),
                    }
                    return self.template_invoice_confirm_edit, context
                elif order.order_type == "直送":
                    modified_get = request.GET.copy()
                    modified_get["start_date_confirm"] = request.GET.get(
                        "start_date_confirm"
                    )
                    modified_get["end_date_confirm"] = request.GET.get(
                        "end_date_confirm"
                    )
                    modified_get["confirm_step"] = True
                    new_request = request
                    new_request.GET = modified_get
                    return self.handle_container_invoice_direct_get(new_request)
        else:
            return self.handle_container_invoice_payable_get(request, True, False)

    def handle_container_invoice_delivery_get(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.GET.get("container_number")
        invoice_type = request.GET.get("invoice_type")
        invoice = Invoice.objects.select_related("customer", "container_number").get(
            container_number__container_number=container_number
        )
        order = Order.objects.select_related(
            "retrieval_id", "container_number", "vessel_id"
        ).get(container_number__container_number=container_number)
        warehouse = order.retrieval_id.retrieval_destination_area
        vessel_etd = order.vessel_id.vessel_etd
        is_transfer = False
        if order.order_type == "转运":
            is_transfer = True
        customer = order.customer_name
        customer_name = customer.zem_name
        fee_details = self._get_fee_details(warehouse, vessel_etd, customer_name)
        # 这里新加入一个功能，是自动将派送类别归类，要先看下
        invoice_status = InvoiceStatus.objects.get(
            container_number=order.container_number, invoice_type="receivable"
        )
        if invoice_status.stage not in ["confirmed", "tobeconfirmed"]:
            groups = [group.name for group in request.user.groups.all()]
            if "warehouse_other" not in groups:
                # 私仓组，不用关公仓归类的情况
                has_unclassified_public_pallets = (
                    Pallet.objects.filter(
                        container_number__container_number=container_number,
                        delivery_type="public",
                        invoice_delivery__isnull=True,
                    )
                    .exclude(delivery_method="暂扣留仓(HOLD)")
                    .exists()
                )
                # 如果有未绑定派送类型的时候，才去归类
                if has_unclassified_public_pallets:
                    self._auto_classify_pallet(
                        container_number,
                        fee_details,
                        warehouse,
                        False,
                        False,
                        is_transfer,
                    )
        # 把pallet汇总
        base_query = Pallet.objects.prefetch_related(
            "container_number",
            "container_number__order",
            "container_number__order__warehouse",
            "container_number__order__customer_name",
            "invoice_delivery",
            Prefetch("invoice_delivery__pallet_delivery", to_attr="delivered_pallets"),
        ).filter(container_number__container_number=container_number)
        common_annotations = {
            "str_id": Cast("id", CharField()),
            "is_hold": Case(
                When(delivery_method__contains="暂扣留仓", then=Value(True)),
                default=Value(False),
                output_field=BooleanField(),
            ),
            "has_delivery": Case(
                When(
                    models.Q(invoice_delivery__isnull=False)
                    | models.Q(
                        invoice_delivery__isnull=True,
                        delivery_method__contains="暂扣留仓",
                    ),
                    then=Value(True),
                ),
                default=Value(False),
                output_field=BooleanField(),
            ),
            "is_self_pickup": Case(
                When(delivery_method__contains="客户自提", then=Value(True)),
                default=Value(False),
                output_field=BooleanField(),
            ),
        }
        common_values = [
            "container_number__container_number",
            "destination",
            "zipcode",
            "address",
            "delivery_method",
            "invoice_delivery__type",
            "delivery_type",
            "is_hold",
            "has_delivery",
        ]
        common_aggregates = {
            "ids": StringAgg("str_id", delimiter=",", distinct=True, ordering="str_id"),
            "total_cbm": Sum("cbm", output_field=FloatField()),
            "total_weight_lbs": Sum("weight_lbs", output_field=FloatField()),
            "total_pallet": Count("pallet_id", distinct=True),
        }
        common_ordering = [
            Case(
                When(is_hold=True, then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            ),
            F("invoice_delivery__type").asc(nulls_first=True),
        ]
        if "NJ" in warehouse:
            pallets = (
                base_query.annotate(**common_annotations)
                .values(*common_values)
                .annotate(**common_aggregates)
                .order_by(*common_ordering)
            )
        else:  # sav和la的自提要按唛头分组
            self_pickup_pallets = (
                base_query.annotate(**common_annotations)
                .filter(is_self_pickup=True)
                .annotate(invoice_delivery_type=F("invoice_delivery__type"))
                .values(
                    *common_values, "shipping_mark", "invoice_delivery_type"
                )  # 自提组需要包含shipping_mark
                .annotate(**common_aggregates)
                .order_by(*common_ordering)
            )
            non_self_pickup_pallets = (
                base_query.annotate(**common_annotations)
                .filter(is_self_pickup=False)
                .annotate(invoice_delivery_type=F("invoice_delivery__type"))
                .values(
                    *common_values, "invoice_delivery_type"
                )  # 非自提组不需要shipping_mark
                .annotate(**common_aggregates)
                .order_by(*common_ordering)
            )
            pallets = list(self_pickup_pallets) + list(non_self_pickup_pallets)

        has_delivery = True
        for plt in pallets:
            if plt["has_delivery"] == False:
                has_delivery = False
                break
        # 需要重新规范板数，就是total_n_pallet
        cutoff_date = date(2025, 4, 1)
        cutoff_datetime = datetime.combine(cutoff_date,datetime_time.min).replace(
            tzinfo=pytz.UTC
        )
        is_new_rule = vessel_etd >= cutoff_datetime

        # 将已经归类的板子分组统计
        delivery_groups = self._process_delivery_records(
            invoice.invoice_number,
            pallets,
            order.container_number.container_type,
            fee_details,
            warehouse,
            is_new_rule,
        )

        groups = [group.name for group in request.user.groups.all()]

        step = request.POST.get("step")
        redirect_step = (step == "redirect") or (
            request.POST.get("redirect_step") == "True"
        )

        container = Container.objects.get(container_number=container_number)
        if order.order_type == "转运组合" and container.account_order_type == "转运":
            non_combina_reason = container.non_combina_reason
        else:
            non_combina_reason = None

        context = {
            "warehouse": request.GET.get("warehouse"),
            "invoice": invoice,
            "container_number": container_number,
            "pallet": pallets,
            "amazon": delivery_groups["amazon"],
            "local": delivery_groups["local"],
            "combine": delivery_groups["combine"],
            "walmart": delivery_groups["walmart"],
            "selfdelivery": delivery_groups["selfdelivery"],
            "upsdelivery": delivery_groups["upsdelivery"],
            "selfpickup": delivery_groups["selfpickup"],
            "redirect_step": redirect_step,
            "start_date": request.GET.get("start_date") or None,
            "end_date": request.GET.get("end_date") or None,
            "start_date_confirm": request.POST.get("start_date_confirm") or None,
            "end_date_confirm": request.POST.get("end_date_confirm") or None,
            "invoice_type": invoice_type,
            "non_combina_reason": non_combina_reason,
            "delivery_types": [
                ("", ""),
                ("公仓", "public"),
                ("其他", "other"),
            ],
        }
        if "mix_account" in groups:  # 如果公仓私仓都能看，就进总页面
            context["delivery_type"] = "mixed"
            context["has_delivery"] = has_delivery
            return self.template_invoice_delievery_edit, context
        else:
            if "NJ_mix_account" in groups or (
                "warehouse_other" in groups and "warehouse_public" not in groups
            ):  # 只看私仓
                pallet = self._filter_pallets(pallets, "other")
                has_delivery = True
                for plt in pallet:
                    if plt["has_delivery"] == False:
                        has_delivery = False
                        break
                context["delivery_type"] = "other"
                context["pallet"] = pallet
                context["has_delivery"] = has_delivery
                return self.template_invoice_delievery_other_edit, context
            elif "warehouse_public" in groups and "warehouse_other" not in groups:
                pallet = self._filter_pallets(pallets, "public")
                has_delivery = True
                for plt in pallet:
                    if plt["has_delivery"] == False:
                        has_delivery = False
                        break
                context["pallet"] = pallet
                context["delivery_type"] = "public"
                context["has_delivery"] = has_delivery
                return self.template_invoice_delievery_public_edit, context
            else:
                raise ValueError("没有权限")

    def _auto_classify_pallet(
        self,
        container_number: str,
        fee_details: dict,
        warehouse: str,
        is_specific_type: bool,
        iscombina: bool,
        is_transfer: bool,
    ) -> None:
        # 如果是切换模式，先删除所有相关的invoice_delivery记录
        if is_specific_type:
            delete_query = Pallet.objects.filter(
                container_number__container_number=container_number,
                delivery_type="public",
                invoice_delivery__isnull=False,
            ).exclude(delivery_method="暂扣留仓(HOLD)")

            invoice_delivery_ids = list(
                delete_query.values_list("invoice_delivery_id", flat=True).distinct()
            )
            # 先解除关联
            delete_query.update(invoice_delivery=None)
            # 删除invoice_delivery记录
            if invoice_delivery_ids:
                with transaction.atomic():
                    InvoiceDelivery.objects.filter(id__in=invoice_delivery_ids).delete()
        # 查找公仓、没有派送类别、没有暂扣的板子
        base_query = (
            Pallet.objects.prefetch_related(
                "container_number",
                "container_number__order",
                "container_number__order__warehouse",
                "container_number__order__customer_name",
                "invoice_delivery",
                Prefetch(
                    "invoice_delivery__pallet_delivery", to_attr="delivered_pallets"
                ),
            )
            .filter(
                container_number__container_number=container_number,
                delivery_type="public",
            )
            .exclude(delivery_method="暂扣留仓(HOLD)")  # 排除暂扣留仓的
        )
        if not is_specific_type:
            # 如果没指定类型，只看没有处理的板子，指定了的话，要看全部的，因为是公仓全部转类型
            base_query = base_query.filter(invoice_delivery__isnull=True)
        common_values = [
            "container_number__container_number",
            "destination",
            "zipcode",
            "address",
            "delivery_method",
            "invoice_delivery__type",
            "delivery_type",
            "PO_ID",
        ]
        common_aggregates = {
            "total_cbm": Sum("cbm", output_field=FloatField()),
            "total_weight_lbs": Sum("weight_lbs", output_field=FloatField()),
            "total_pallet": Count("pallet_id", distinct=True),
            "pallet_ids": ArrayAgg("pallet_id"),
        }
        pallets = base_query.values(*common_values).annotate(**common_aggregates)
        if not pallets.exists():
            return

        container = Container.objects.get(container_number=container_number)
        container_type_temp = 0 if container.container_type == "40HQ/GP" else 1
        invoice = Invoice.objects.select_related("customer", "container_number").get(
            container_number__container_number=container_number
        )
        current_date = datetime.now().date()

        # 如果没有指定类型，自动判断是否为组合柜
        if not is_specific_type and not is_transfer:
            iscombina = self._is_combina(container_number)

        for pallet in pallets:
            destination = pallet["destination"]
            delivery_type = None
            cost = 0

            if iscombina:
                # 按组合柜算，就去组合柜表找
                rules = fee_details.get(f"{warehouse}_COMBINA").details
                for region, region_data in rules.items():
                    for item in region_data:
                        if destination in item["location"]:
                            delivery_type = "combine"
                            cost = item["prices"][container_type_temp]
                            break
                    if delivery_type:
                        break
            else:
                # 去亚马逊/沃尔玛表找
                rules = fee_details.get(f"{warehouse}_PUBLIC").details
                # details = rules.get(f"{warehouse}_AMAZON", rules) if "LA" not in warehouse else rules
                details = (
                    {"LA_AMAZON": rules}
                    if "LA" in warehouse and "LA_AMAZON" not in rules
                    else rules
                )
                for category, zones in details.items():
                    for zone, locations in zones.items():
                        if destination in locations:
                            if "AMAZON" in category:
                                delivery_type = "amazon"
                                cost = zone
                            elif "WALMART" in category:
                                delivery_type = "walmart"
                                cost = zone
                    if delivery_type:
                        break
            if not delivery_type:
                continue
            invoice_delivery_id = f"{current_date.strftime('%Y%m%d')}-{delivery_type}-{destination}-{pallet['total_pallet']}"
            existing_deliveries = InvoiceDelivery.objects.filter(
                invoice_delivery=invoice_delivery_id
            )

            # 删除所有已存在的相同invoice_delivery记录
            if existing_deliveries.exists():
                # 先解除关联这些invoice_delivery的所有托盘
                Pallet.objects.filter(invoice_delivery__in=existing_deliveries).update(
                    invoice_delivery=None
                )
                # 删除invoice_delivery记录
                existing_deliveries.delete()
            invoice_content = InvoiceDelivery(
                invoice_delivery=invoice_delivery_id,
                invoice_number=invoice,
                delivery_type="public",
                type=delivery_type,
                zipcode=pallet["zipcode"],
                cost=cost,
                total_cost=float(cost) * float(pallet["total_pallet"]),
                destination=destination,
                total_pallet=pallet["total_pallet"],
                total_cbm=pallet["total_cbm"],
                total_weight_lbs=pallet["total_weight_lbs"],
            )
            invoice_content.save()

            # 更新托盘的invoice_delivery关联
            if pallet["pallet_ids"]:
                Pallet.objects.filter(pallet_id__in=pallet["pallet_ids"]).update(
                    invoice_delivery=invoice_content
                )

    def _is_combina(self, container_number: str) -> bool:
        container = Container.objects.get(container_number=container_number)
        order = Order.objects.select_related(
            "retrieval_id", "container_number", "vessel_id"
        ).get(container_number__container_number=container_number)
        customer = order.customer_name
        customer_name = customer.zem_name
        # 从报价表找+客服录的数据
        warehouse = order.retrieval_id.retrieval_destination_area
        vessel_etd = order.vessel_id.vessel_etd

        container_type = container.container_type
        #  基础数据统计
        plts = Pallet.objects.filter(
            container_number__container_number=container_number
        ).aggregate(
            unique_destinations=Count("destination", distinct=True),
            total_weight=Sum("weight_lbs"),
            total_cbm=Sum("cbm"),
            total_pallets=Count("id"),
        )
        plts["total_cbm"] = round(plts["total_cbm"], 2)
        plts["total_weight"] = round(plts["total_weight"], 2)
        # 获取匹配的报价表
        matching_quotation = (
            QuotationMaster.objects.filter(
                effective_date__lte=vessel_etd,
                is_user_exclusive=True,
                exclusive_user=customer_name,
            )
            .order_by("-effective_date")
            .first()
        )
        if not matching_quotation:
            matching_quotation = (
                QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=False,  # 非用户专属的通用报价单
                )
                .order_by("-effective_date")
                .first()
            )
        # 获取组合柜规则
        stipulate = FeeDetail.objects.get(
            quotation_id=matching_quotation.id, fee_type="COMBINA_STIPULATE"
        ).details
        combina_fee = FeeDetail.objects.get(
            quotation_id=matching_quotation.id, fee_type=f"{warehouse}_COMBINA"
        ).details
        if isinstance(combina_fee, str):
            combina_fee = json.loads(combina_fee)
        # 看是否超出组合柜限定仓点,NJ/SAV是14个
        default_combina = stipulate["global_rules"]["max_mixed"]["default"]
        exceptions = stipulate["global_rules"]["max_mixed"].get("exceptions", {})
        combina_threshold = exceptions.get(warehouse, default_combina) if exceptions else default_combina

        default_threshold = stipulate["global_rules"]["bulk_threshold"]["default"]
        exceptions = stipulate["global_rules"]["bulk_threshold"].get("exceptions", {})
        uncombina_threshold = exceptions.get(warehouse, default_threshold) if exceptions else default_threshold
        if plts["unique_destinations"] > uncombina_threshold:
            container.account_order_type = "转运"
            container.non_combina_reason = (
                f"总仓点超过{uncombina_threshold}个"
            )
            container.save()
            return False  # 不是组合柜

        # 按区域统计
        destinations = (
            Pallet.objects.filter(container_number__container_number=container_number)
            .values_list("destination", flat=True)
            .distinct()
        )
        plts_by_destination = (
            Pallet.objects.filter(container_number__container_number=container_number)
            .values("destination")
            .annotate(total_cbm=Sum("cbm"))
        )
        total_cbm_sum = sum(item["total_cbm"] for item in plts_by_destination)
        # 区分组合柜区域和非组合柜区域
        container_type_temp = 0 if container_type == "40HQ/GP" else 1
        matched_regions = self.find_matching_regions(
            plts_by_destination, combina_fee, container_type_temp, total_cbm_sum, combina_threshold
        )
        # 判断是否混区，False表示满足混区条件
        is_mix = self.is_mixed_region(
            matched_regions["matching_regions"], warehouse, vessel_etd
        )
        if is_mix:
            container.account_order_type = "转运"
            container.non_combina_reason = "混区不符合标准"
            container.save()
            return False
        # 非组合柜区域
        non_combina_region_count = len(matched_regions["non_combina_dests"].keys())
        # 组合柜区域
        combina_region_count = len(matched_regions["combina_dests"])

        if combina_region_count + non_combina_region_count != len(destinations):
            raise ValueError("计算组合柜和非组合柜区域有误")
        if non_combina_region_count > (
            uncombina_threshold
            - combina_threshold
        ):
            # 当非组合柜的区域数量超出时，不能按转运组合
            container.account_order_type = "转运"
            container.non_combina_reason = "非组合柜区的数量不符合标准"
            container.save()
            return False
        container.non_combina_reason = None
        container.account_order_type = "转运组合"
        container.save()
        return True

    def _filter_pallets(self, pallets: Any, delivery_type: str) -> Any:
        if isinstance(pallets, QuerySet):
            return pallets.filter(delivery_type=delivery_type)
        elif isinstance(pallets, list):
            return [
                pallet
                for pallet in pallets
                if (
                    # 支持对象属性、字典键、或 getattr 安全访问
                    pallet.get("delivery_type") == delivery_type
                    if isinstance(pallet, dict)
                    else getattr(pallet, "delivery_type", None) == delivery_type
                )
            ]
        else:
            raise ValueError("pallets must be QuerySet or list")

    def _get_fee_details(self, warehouse: str, vessel_etd, customer_name: str) -> dict:
        if not vessel_etd:
            raise ValueError('柜子缺少ETD')
        try:
            quotation = (
                QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=True,
                    exclusive_user=customer_name,
                )
                .order_by("-effective_date")
                .first()
            )
            if not quotation:
                quotation = (
                    QuotationMaster.objects.filter(
                        effective_date__lte=vessel_etd,
                        is_user_exclusive=False,  # 非用户专属的通用报价单
                    )
                    .order_by("-effective_date")
                    .first()
                )
            if not quotation:
                raise ValueError("找不到报价表")
            id = quotation.id
            fee_types = {
                "NJ": ["NJ_LOCAL", "NJ_PUBLIC", "NJ_COMBINA"],
                "SAV": ["SAV_PUBLIC", "SAV_COMBINA"],
                "LA": ["LA_PUBLIC", "LA_COMBINA"],
            }.get(warehouse, [])

            return {
                fee.fee_type: fee
                for fee in FeeDetail.objects.filter(
                    quotation_id=id, fee_type__in=fee_types
                )
            }
        except QuotationMaster.DoesNotExist:
            raise ValueError("没有找到有效的报价单")

    def _process_delivery_records(
        self,
        invoice_number: str,
        pallets: Any,
        container_type: str,
        fee_details: dict,
        warehouse: str,
        is_new_rule: bool,
    ) -> dict:
        invoice_deliveries = InvoiceDelivery.objects.prefetch_related(
            Prefetch("pallet_delivery", queryset=Pallet.objects.all())
        ).filter(invoice_number__invoice_number=invoice_number)

        delivery_groups = {
            "amazon": [],
            "local": [],
            "combine": [],
            "walmart": [],
            "selfdelivery": [],
            "upsdelivery": [],
            "selfpickup": [],
            "invoice_delivery": invoice_deliveries,
        }
        # 没有单价的找单价，再根据type汇总派送方式
        for delivery in invoice_deliveries:
            pallets_in_delivery = delivery.pallet_delivery.all()
            pallet_ids = [str(p.id) for p in pallets_in_delivery]
            setattr(delivery, "plt_ids", pallet_ids)
            if delivery.cost is None:
                self._calculate_and_set_delivery_cost(
                    delivery, container_type, fee_details, warehouse, is_new_rule
                )
            if delivery.type in delivery_groups:
                delivery_groups[delivery.type].append(delivery)

        return delivery_groups

    # 根据报价表找单价
    def _calculate_and_set_delivery_cost(
        self,
        delivery: InvoiceDelivery,
        container_type: str,
        fee_details: dict,
        warehouse: str,
        is_new_rule: bool,
    ) -> None:
        destination = (
            delivery.destination.split("-")[1]
            if "-" in delivery.destination
            else delivery.destination
        )
        cost = 0
        is_niche_warehouse = True  # 本地派送算冷门仓点
        # 根据报价表找单价，ups和自发没有报价，所以不用找
        if delivery.type == "amazon":
            if "LA" in warehouse:
                amazon_data = fee_details.get(f"{warehouse}_PUBLIC").details
            else:
                amazon_data = fee_details.get(f"{warehouse}_PUBLIC").details.get(
                    f"{warehouse}_AMAZON"
                )
            for k, v in amazon_data.items():
                if destination in v:
                    cost = k
                    break
            niche_warehouse = fee_details.get(f"{warehouse}_PUBLIC").niche_warehouse
            if destination in niche_warehouse:
                is_niche_warehouse = True
            else:
                is_niche_warehouse = False
        elif delivery.type == "local" and warehouse == "NJ":
            local_data = fee_details.get("NJ_LOCAL").details
            for k, v in local_data.items():
                if str(delivery.zipcode) in map(str, v["zipcodes"]):
                    n_pallet = int(delivery.total_pallet)  # 板数
                    costs = v["prices"]
                    if n_pallet <= 5:
                        cost = int(costs[0])
                    elif n_pallet >= 5:
                        cost = int(costs[1])
                    break
        elif delivery.type == "combine":
            combine_data = fee_details.get(f"{warehouse}_COMBINA").details
            cost = None
            for region, price_groups in combine_data.items():
                for price_group in price_groups:
                    if destination in price_group["location"]:
                        if "45HQ/GP" in container_type:
                            cost = int(price_group["prices"][1])
                        elif "40HQ/GP" in container_type:
                            cost = int(price_group["prices"][0])
                        break
                if cost is not None:
                    break
            niche_warehouse = fee_details.get(f"{warehouse}_PUBLIC").niche_warehouse
            if destination in niche_warehouse:
                is_niche_warehouse = True
            else:
                is_niche_warehouse = False
        elif delivery.type == "walmart":
            if "LA" in warehouse:
                walmart_data = fee_details.get(f"{warehouse}_PUBLIC").details
            else:
                walmart_data = fee_details.get(f"{warehouse}_PUBLIC").details.get(
                    f"{warehouse}_WALMART"
                )
            for price, locations in walmart_data.items():
                if str(destination).upper() in [loc.upper() for loc in locations]:
                    cost = price
                    break
            niche_warehouse = fee_details.get(f"{warehouse}_PUBLIC").niche_warehouse
            if destination in niche_warehouse:
                is_niche_warehouse = True
            else:
                is_niche_warehouse = False
        if cost is not None:
            delivery.cost = cost
        else:
            delivery.cost = 0
        # 计算板数
        if delivery.type == "local":
            is_new_rule = False
        total_pallet = self._calculate_total_pallet(
            delivery.total_cbm, is_new_rule, is_niche_warehouse
        )
        delivery.total_pallet = total_pallet
        delivery.save()

    def _calculate_total_pallet(
        self, cbm: float, is_new_rule: bool, is_niche_warehouse: bool
    ) -> float:
        raw_p = float(cbm) / 1.8
        integer_part = int(raw_p)
        decimal_part = raw_p - integer_part
        # 本地派送的按照4.1之前的规则
        if decimal_part > 0:
            if is_new_rule:  # etd4.1之后的
                if is_niche_warehouse:
                    additional = 1 if decimal_part > 0.5 else 0.5
                else:
                    additional = 1 if decimal_part > 0.5 else 0
            else:
                if decimal_part > 0:
                    if is_niche_warehouse:
                        additional = 1
                    else:
                        additional = 1 if decimal_part > 0.9 else 0.5
            total_pallet = integer_part + additional
        elif decimal_part == 0:
            total_pallet = integer_part
        else:
            ValueError("板数计算错误")
        return total_pallet

    def find_matching_regions(
        self,
        plts_by_destination: dict,
        combina_fee: dict,
        container_type,
        total_cbm_sum: FloatField,
        combina_threshold: int,
    ) -> dict:
        matching_regions = defaultdict(float)  # 各区的cbm总和
        des_match_quote = {}  # 各仓点的匹配详情
        destination_matches = set()  # 组合柜的仓点
        non_combina_dests = {}  # 非组合柜的仓点
        price_display = defaultdict(
            lambda: {"price": 0.0, "location": set()}
        )  # 各区的价格和仓点
        dest_cbm_list = []  # 临时存储初筛组合柜内的cbm和匹配信息

        for plts in plts_by_destination:
            destination = plts["destination"]
            # 如果是沃尔玛的，只保留后面的名字，因为报价表里就是这么保留的
            dest = destination.replace("沃尔玛", "").split("-")[-1].strip()
            cbm = plts["total_cbm"]
            dest_matches = []
            matched = False
            # 遍历所有区域和location
            for region, fee_data_list in combina_fee.items():
                for fee_data in fee_data_list:
                    # 如果匹配到组合柜仓点，就登记到组合柜集合中
                    if dest in fee_data["location"]:
                        temp_cbm = matching_regions[region] + cbm
                        matching_regions[region] = temp_cbm
                        dest_matches.append(
                            {
                                "region": region,
                                "location": dest,
                                "prices": fee_data["prices"],
                                "cbm": cbm,
                            }
                        )
                        price_display[region]["price"] = fee_data["prices"][
                            container_type
                        ]
                        price_display[region]["location"].add(dest)
                        matched = True
            if not matched:
                # 非组合柜仓点
                non_combina_dests[dest] = {"cbm": cbm}
            # 记录匹配结果
            if dest_matches:
                des_match_quote[dest] = dest_matches
                # 将组合柜内的记录下来，后续方便按照cbm排序
                dest_cbm_list.append(
                    {"dest": dest, "cbm": cbm, "matches": dest_matches}
                )
                destination_matches.add(dest)

        if len(destination_matches) > combina_threshold:
            # 按cbm降序排序，将cbm大的归到非组合
            sorted_dests = sorted(dest_cbm_list, key=lambda x: x["cbm"], reverse=True)
            # 重新将排序后的前12个加入里面
            destination_matches = set()
            matching_regions = defaultdict(float)
            price_display = defaultdict(lambda: {"price": 0.0, "location": set()})
            for item in sorted_dests[:combina_threshold]:
                dest = item["dest"]
                destination_matches.add(dest)

                # 重新计算各区域的CBM总和
                for match in item["matches"]:
                    region = match["region"]
                    matching_regions[region] += item["cbm"]
                    price_display[region]["price"] = match["prices"][container_type]
                    price_display[region]["location"].add(dest)

            # 其余仓点转为非组合柜
            for item in sorted_dests[combina_threshold:]:
                non_combina_dests[item["dest"]] = {"cbm": item["cbm"]}
                # 将cbm大的从组合柜集合中删除
                des_match_quote.pop(item["dest"], None)

        # 下面开始计算组合柜和非组合柜各仓点占总体积的比例
        total_ratio = 0.0
        ratio_info = []

        # 处理组合柜仓点的cbm_ratio
        for dest, matches in des_match_quote.items():
            cbm = matches[0]["cbm"]  # 同一个dest的cbm在所有matches中相同
            ratio = round(cbm / total_cbm_sum, 4)
            total_ratio += ratio
            ratio_info.append((dest, ratio, cbm, True))  # 最后一个参数表示是否是组合柜
            for match in matches:
                match["cbm_ratio"] = ratio

        # 处理非组合柜仓点的cbm_ratio
        for dest, data in non_combina_dests.items():
            cbm = data["cbm"]
            ratio = round(cbm / total_cbm_sum, 4)
            total_ratio += ratio
            ratio_info.append((dest, ratio, cbm, False))
            data["cbm_ratio"] = ratio

        # 处理四舍五入导致的误差
        if abs(total_ratio - 1.0) > 0.0001:  # 考虑浮点数精度
            # 找到CBM最大的仓点
            ratio_info.sort(key=lambda x: x[2], reverse=True)
            largest_dest, largest_ratio, largest_cbm, is_combi = ratio_info[0]

            # 调整最大的仓点的ratio
            diff = 1.0 - total_ratio
            if is_combi:
                for match in des_match_quote[largest_dest]:
                    match["cbm_ratio"] = round(match["cbm_ratio"] + diff, 4)
            else:
                non_combina_dests[largest_dest]["cbm_ratio"] = round(
                    non_combina_dests[largest_dest]["cbm_ratio"] + diff, 4
                )
        return {
            "des_match_quote": des_match_quote,
            "matching_regions": matching_regions,
            "combina_dests": destination_matches,
            "non_combina_dests": non_combina_dests,
            "price_display": price_display,
        }

    def is_mixed_region(self, matched_regions, warehouse, vessel_etd) -> bool:
        regions = list(matched_regions.keys())
        # LA仓库的特殊规则：CDEF区不能混
        if warehouse == "LA":
            if vessel_etd.month > 7 or (
                vessel_etd.month == 7 and vessel_etd.day >= 15
            ):  # 715之后没有混区限制
                return False
            if len(regions) <= 1:  # 只有一个区，就没有混区的情况
                return False
            if set(regions) == {"A区", "B区"}:  # 如果只有A区和B区，也满足混区规则
                return False
            return True
        # 其他仓库无限制
        return False

    def handle_container_invoice_combina_get(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.GET.get("container_number")
        container = Container.objects.get(container_number=container_number)
        order = Order.objects.select_related(
            "retrieval_id", "container_number", "vessel_id"
        ).get(container_number__container_number=container_number)

        invoice = Invoice.objects.select_related("customer", "container_number").get(
            container_number__container_number=container_number
        )

        invoice_status = InvoiceStatus.objects.get(
            container_number=order.container_number, invoice_type="receivable"
        )
        is_from_account_confirmation = getattr(
            request, "is_from_billing_confirmation", False
        )
        context = {
            "invoice_number": invoice.invoice_number,
            "container_number": container_number,
            "start_date_confirm": request.GET.get("start_date_confirm"),
            "end_date_confirm": request.GET.get("end_date_confirm"),
            "start_date": request.GET.get("start_date"),
            "end_date": request.GET.get("end_date"),
            "is_from_account_confirmation": is_from_account_confirmation,
        }
        # 查看是不是财务未确认状态，未确认就从报价表找+客服录的数据，确认了就从invoice_item表找
        if invoice_status.stage == "confirmed":
            invoice_item = InvoiceItem.objects.filter(
                invoice_number__invoice_number=invoice.invoice_number
            )
            context["invoice"] = invoice
            context["invoice_item"] = invoice_item
            return self.template_invoice_container_edit, context
        # 从报价表找+客服录的数据
        warehouse = order.retrieval_id.retrieval_destination_area
        vessel_etd = order.vessel_id.vessel_etd

        container_type = container.container_type
        # 1. 基础数据统计
        plts = Pallet.objects.filter(
            container_number__container_number=container_number
        ).aggregate(
            unique_destinations=Count("destination", distinct=True),
            total_weight=Sum("weight_lbs"),
            total_cbm=Sum("cbm"),
            total_pallets=Count("id"),
        )
        plts["total_cbm"] = round(plts["total_cbm"], 2)
        plts["total_weight"] = round(plts["total_weight"], 2)
        # 3. 获取匹配的报价表
        customer = order.customer_name
        customer_name = customer.zem_name
        matching_quotation = (
            QuotationMaster.objects.filter(
                effective_date__lte=vessel_etd,
                is_user_exclusive=True,
                exclusive_user=customer_name,
            )
            .order_by("-effective_date")
            .first()
        )
        if not matching_quotation:
            matching_quotation = (
                QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=False,  # 非用户专属的通用报价单
                )
                .order_by("-effective_date")
                .first()
            )
        if not matching_quotation:
            context["reason"] = "找不到匹配报价表"
            return self.template_invoice_combina_edit, context
        # 4. 获取费用规则
        PICKUP_FEE = FeeDetail.objects.get(
            quotation_id=matching_quotation.id, fee_type="preport"
        )
        combina_fee = FeeDetail.objects.get(
            quotation_id=matching_quotation.id, fee_type=f"{warehouse}_COMBINA"
        ).details
        customer = order.customer_name
        customer_name = customer.zem_name
        fee_details = self._get_fee_details(warehouse, vessel_etd, customer_name)
        stipulate = FeeDetail.objects.get(
            quotation_id=matching_quotation.id, fee_type="COMBINA_STIPULATE"
        ).details
        if isinstance(combina_fee, str):
            combina_fee = json.loads(combina_fee)
        # 2. 检查基本条件
        if plts["unique_destinations"] == 0:
            context["reason"] = "未录入拆柜数据"
            return self.template_invoice_combina_edit, context

        default_combina = stipulate["global_rules"]["max_mixed"]["default"]
        exceptions = stipulate["global_rules"]["max_mixed"].get("exceptions", {})
        combina_threshold = exceptions.get(warehouse, default_combina) if exceptions else default_combina

        default_threshold = stipulate["global_rules"]["bulk_threshold"]["default"]
        exceptions = stipulate["global_rules"]["bulk_threshold"].get("exceptions", {})
        uncombina_threshold = exceptions.get(warehouse, default_threshold) if exceptions else default_threshold

        if (
            plts["unique_destinations"]
            > uncombina_threshold
        ):
            container.account_order_type = "转运"
            container.save()
            context["reason"] = (
                f"超过{uncombina_threshold}个仓点"
            )
            return self.template_invoice_combina_edit, context

        # 按区域统计
        destinations = (
            Pallet.objects.filter(container_number__container_number=container_number)
            .values_list("destination", flat=True)
            .distinct()
        )
        plts_by_destination = (
            Pallet.objects.filter(container_number__container_number=container_number)
            .values("destination")
            .annotate(total_cbm=Sum("cbm"))
        )
        # 这里之前是
        total_cbm_sum = sum(item["total_cbm"] for item in plts_by_destination)
        # 区分组合柜区域和非组合柜区域
        container_type_temp = 0 if container_type == "40HQ/GP" else 1
        matched_regions = self.find_matching_regions(
            plts_by_destination, combina_fee, container_type_temp, total_cbm_sum, combina_threshold
        )
        # 判断是否混区，除了LA的CDEF不能混，别的都能混
        is_mix = self.is_mixed_region(
            matched_regions["matching_regions"], warehouse, vessel_etd
        )

        # 非组合柜区域
        non_combina_region_count = len(matched_regions["non_combina_dests"].keys())
        # 组合柜区域
        combina_region_count = len(matched_regions["combina_dests"])
        # 组合柜对应的区
        des_match_quote = matched_regions["des_match_quote"]
        # 组合柜区域，各区及总cbm
        matching_regions = matched_regions["matching_regions"]
        if combina_region_count + non_combina_region_count != len(destinations):
            raise ValueError("计算组合柜和非组合柜区域有误")

        non_combina_cbm_ratio = round(
            sum(
                data["cbm_ratio"]
                for data in matched_regions["non_combina_dests"].values()
            ),
            4,
        )
        non_combina_cbm = round(
            sum(data["cbm"] for data in matched_regions["non_combina_dests"].values()),
            4,
        )

        if combina_region_count > combina_threshold or non_combina_region_count > ( uncombina_threshold - combina_threshold ):
            container.account_order_type = "转运"
            container.save()
            if combina_region_count > combina_threshold:
                # reason = '不满足组合柜区域要求'
                reason = f"规定{combina_threshold}组合柜区,但实际有{combina_region_count}个:matched_regions['combina_dests']，所以按照转运方式统计价格"
            elif non_combina_region_count > (
                uncombina_threshold
                - combina_threshold
            ):
                stipulate_non_combina = (
                    uncombina_threshold
                    - combina_threshold
                )
                reason = f"规定{stipulate_non_combina}个非组合柜区，总共{uncombina_threshold}个区，组合柜{combina_threshold}个区，但是有{non_combina_region_count}个：{list(matched_regions['non_combina_dests'].keys())}，所以按照转运方式统计价格"
                # reason = '不满足组合柜区域要求'
            actual_fees = self._combina_get_extra_fees(invoice)
            context["reason"] = reason
            context["extra_fees"] = actual_fees
            return self.template_invoice_combina_edit, context
        # 7.2 计算基础费用
        base_fee = 0
        extra_fees = {
            "overweight": 0,
            "overpallets": 0,
            "overregion_pickup": 0,
            "overregion_delivery": 0,
        }
        base_fee = 0.0
        price_display = matched_regions["price_display"]
        price_display_new = None
        if not is_mix:
            # 单一区域情况
            if len(matched_regions["price_display"]) == 1:
                region_dict = matched_regions["price_display"]
                region = list(region_dict.keys())[0]
                base_fee = combina_fee[region][0]["prices"][
                    0 if container_type == "40HQ/GP" else 1
                ]
                total_ratio = sum(
                    match["cbm_ratio"]
                    for dest in des_match_quote.keys()
                    for match in des_match_quote[dest]
                    if match["region"] == region
                )
                base_fee *= total_ratio
                price_display_new = [
                    {
                        "key": region,
                        "cbm": round(matching_regions[region], 2),
                        "rate": round(
                            sum(
                                match["cbm_ratio"]
                                for dest in data["location"]
                                for match in des_match_quote.get(dest, [])
                            ),
                            4,
                        ),
                        "price": data["price"],
                        "location": ", ".join(data["location"]),
                    }
                    for region, data in price_display.items()
                ]
            else:  # 允许混区的情况
                price_display_new = None

                base_fee = round(base_fee, 2)

                price_display_new = [
                    {
                        "key": region,
                        "cbm": round(matching_regions[region], 2),
                        "rate": round(
                            sum(
                                match[
                                    "cbm_ratio"
                                ]  # 遍历所有匹配的 location 并累加 cbm_ratio
                                for dest in data[
                                    "location"
                                ]  # 遍历当前 region 的所有 location
                                for match in des_match_quote.get(
                                    dest, []
                                )  # 获取该 location 的所有匹配项
                            ),
                            4,
                        ),
                        "price": data["price"],
                        "location": ", ".join(data["location"]),
                    }
                    for region, data in price_display.items()
                ]
                base_fee = round(
                    sum(item["price"] * item["rate"] for item in price_display_new), 2
                )
        if not price_display_new:
            container.account_order_type = "转运"
            container.save()
            reason = "混区不符合规定"
            context["reason"] = reason
            return self.template_invoice_combina_edit, context
        # 7.3 检查超限情况
        # 超重检查
        if plts["total_weight"] > stipulate["global_rules"]["weight_limit"]["default"]:
            extra_fees["overweight"] = "需人工录入"  # 实际业务中应有默认费率

        # 超板检查——确定上限的板数
        if container_type == "40HQ/GP":
            std_plt = stipulate["global_rules"]["std_40ft_plts"]
            if warehouse == "LA" and "LA_std_40ft_plts" in stipulate["global_rules"]:
                std_plt = stipulate["global_rules"]["LA_std_40ft_plts"]
        else:
            std_plt = stipulate["global_rules"]["std_45ft_plts"]
            if warehouse == "LA" and "LA_std_45ft_plts" in stipulate["global_rules"]:
                std_plt = stipulate["global_rules"]["LA_std_45ft_plts"]
        exceptions_dict = std_plt["exceptions"]

        max_pallets = 0
        for exception_des, exception_plt in exceptions_dict.items():
            exception_warehouse = exception_des.split("_")[0]
            exception_regions = list(exception_des.split("_")[1])
            # 检查是否匹配当前 warehouse 和 region
            if exception_warehouse == warehouse and any(
                region in exception_regions
                for region, value in matched_regions["matching_regions"].items()
            ):
                max_pallets = exception_plt
                break
        if max_pallets == 0:
            max_pallets = std_plt["default"]
        # 处理超的板数
        # 先计算实际板数
        total_pallets = math.ceil(plts["total_cbm"] / 1.8)  # 取上限
        if total_pallets > max_pallets:
            over_count = total_pallets - max_pallets
        else:
            over_count = 0
        # 找每个仓点的单价，倒序排序，方便计算超板的（没有超板的也要查，可能前端会改板数）
        plts_by_destination = (
            Pallet.objects.filter(container_number__container_number=container_number)
            .values("destination")
            .annotate(
                total_cbm=Sum("cbm"),
                price=Value(None, output_field=models.FloatField()),
                is_fixed_price=Value(False, output_field=BooleanField()),
                total_pallet=Count("id", output_field=FloatField()),
            )
        )  # 形如{'destination': 'A', 'total_cbm': 10.5，'price':31.5,'is_fixed_price':True},
        plts_by_destination = self._calculate_delivery_fee_cost(
            fee_details, warehouse, plts_by_destination, destinations, over_count
        )
        max_price = 0
        max_single_price = 0
        for plt_d in plts_by_destination:
            if plt_d["is_fixed_price"]:  # 一口价的不用乘板数
                max_price = max(float(plt_d["price"]), max_price)
                max_single_price = max(max_price, max_single_price)
            else:
                max_price = max(float(plt_d["price"]) * over_count, max_price)
                max_single_price = max(float(plt_d["price"]), max_single_price)
        extra_fees["overpallets"] = max_price

        # 计算非组合柜费用的提拆费和派送费
        if non_combina_region_count:
            # 提拆费，要计算下非组合柜区域占当前柜子的cbm比例*对应的提拆费
            container_type = order.container_number.container_type
            match = re.match(r"\d+", container_type)
            if match:
                pick_subkey = match.group()
                # 这个提拆费是从组合柜规则的warehouse_pricing的nonmix_40ft 45ft取
                c_type = f"nonmix_{pick_subkey}ft"
                try:
                    pickup_fee = stipulate["warehouse_pricing"][warehouse][c_type]
                except KeyError as e:
                    error_msg = f"缺少{pick_subkey}柜型的报价配置"
                    raise ValueError(error_msg)

            extra_fees["overregion_pickup"] = non_combina_cbm_ratio * pickup_fee
            # 派送费
            for item in matched_regions["non_combina_dests"]:
                # 计算改区域的板数
                plts_by_destination_overregion = (
                    Pallet.objects.filter(
                        container_number__container_number=container_number,
                        destination__in=matched_regions["non_combina_dests"].keys(),
                    )
                    .values("destination")
                    .annotate(
                        total_cbm=Sum("cbm"),
                        total_pallet=Count("id", output_field=FloatField()),
                        total_weight=Sum("weight_lbs"),
                        price=Value(None, output_field=models.FloatField()),
                        is_fixed_price=Value(False, output_field=BooleanField()),
                    )
                )

                plts_by_destination_overregion = self._calculate_delivery_fee_cost(
                    fee_details,
                    warehouse,
                    plts_by_destination_overregion,
                    destinations,
                    None,
                )

                sum_price = 0
                for plt_d in plts_by_destination_overregion:
                    if plt_d["is_fixed_price"]:  # 一口价的不用乘板数
                        sum_price += float(plt_d["price"])
                    else:
                        sum_price += float(plt_d["price"]) * plt_d["total_pallet"]
            extra_fees["overregion_delivery"] = sum_price
        else:
            pickup_fee = 0
        # 超仓点的加收费用
        addition_fee = 0
        if "tiered_pricing" in stipulate:
            region_count = combina_region_count + non_combina_region_count
            if warehouse in stipulate["tiered_pricing"]:
                for rule in stipulate["tiered_pricing"][warehouse]:
                    min_points = rule.get("min_points")
                    max_points = rule.get("max_points")
                    if int(min_points) <= region_count <= int(max_points):
                        addition_fee = {
                            "min_points": int(min_points),
                            "max_points": int(max_points),
                            "add_fee": rule.get("fee"),
                        }
        else:
            addition_fee = None

        display_data = {
            # 基础信息
            "plts_by_destination": plts_by_destination,
            "container_info": {
                "number": container_number,
                "type": container_type,
                "warehouse": warehouse,
            },
            # 组合柜信息
            "combina_data": {"base_fee": base_fee, "regions": [], "destinations": []},
            # 超限费用
            "extra_fees": {
                "overweight": {
                    "is_over": plts["total_weight"]
                    > stipulate["global_rules"]["weight_limit"]["default"],
                    "current_weight": plts["total_weight"],
                    "limit_weight": stipulate["global_rules"]["weight_limit"][
                        "default"
                    ],
                    "extra_weight": round(
                        plts["total_weight"]
                        - stipulate["global_rules"]["weight_limit"]["default"],
                        2,
                    ),
                    "fee": extra_fees["overweight"],
                    "input_field": True,  # 显示输入框
                },
                "overpallets": {
                    "is_over": total_pallets > max_pallets,
                    "current_pallets": total_pallets,
                    "limit_pallets": max_pallets,
                    "over_count": over_count if total_pallets > max_pallets else 0,
                    "pallet_details": [],
                    "max_price_used": None,
                    "fee": extra_fees["overpallets"],
                },
                "overregion": {
                    "pickup": {
                        "non_combina_cbm": non_combina_cbm,
                        "total_cbm": plts["total_cbm"],
                        "ratio": non_combina_cbm_ratio * 100,
                        "base_fee": pickup_fee,
                        "fee": extra_fees["overregion_pickup"],
                    },
                    "delivery": {
                        "fee": extra_fees["overregion_delivery"],
                        "details": [],
                    },
                },
                "addition_fee": addition_fee,
            },
        }
        display_data["combina_data"]["destinations"] = price_display_new

        # 填充超板费详细信息，不超板也要展示详情，因为前端可以修改超的板数
        # if total_pallets > max_pallets:
        for plt in plts_by_destination:
            display_data["extra_fees"]["overpallets"]["pallet_details"].append(
                {
                    "destination": plt["destination"],
                    "price": plt["price"],
                    "is_fixed_price": plt["is_fixed_price"],
                    "is_max_used": float(plt["price"])
                    == max_single_price,  # 标记是否被采用
                }
            )
        display_data["extra_fees"]["overpallets"]["max_price_used"] = max_price

        # 填充超区派送费详细信息
        if non_combina_region_count:
            is_overregion = True
            for plt in plts_by_destination_overregion:
                display_data["extra_fees"]["overregion"]["delivery"]["details"].append(
                    {
                        "destination": plt["destination"],
                        "pallets": plt["total_pallet"],
                        "price": plt["price"],
                        "cbm": round(plt["total_cbm"], 2),
                        "subtotal": float(plt["price"]) * plt["total_pallet"],
                    }
                )
        else:
            is_overregion = False
        total_amount = (
            base_fee
            + extra_fees["overpallets"]
            + extra_fees["overregion_pickup"]
            + extra_fees["overregion_delivery"]
        )
        # 港前-仓库-派送录入的费用显示到界面上
        actual_fees = self._combina_get_extra_fees(invoice)

        # 8. 返回结果
        context.update(
            {
                "display_data": display_data,
                "total_amount": total_amount,
                "invoice_number": invoice.invoice_number,
                "container_number": container_number,
                "is_overregion": is_overregion,
                "extra_fees": actual_fees,
                "destination_matches": matched_regions["combina_dests"],
                "non_combina_dests": list(matched_regions["non_combina_dests"].keys()),
            }
        )
        return self.template_invoice_combina_edit, context

    def _combina_get_extra_fees(self, invoice) -> Any:
        extra_fees = {"preports": [], "warehouse": [], "deliverys": []}
        preports_fee = InvoicePreport.objects.get(
            invoice_number__invoice_number=invoice.invoice_number,
            invoice_type="receivable",
        )

        preport_fields = [
            "pickup",
            "chassis",
            "chassis_split",
            "prepull",
            "yard_storage",
            "handling_fee",
            "pier_pass",
            "congestion_fee",
            "hanging_crane",
            "dry_run",
            "exam_fee",
            "hazmat",
            "over_weight",
            "urgent_fee",
            "other_serive",
            "demurrage",
            "per_diem",
            "second_pickup",
        ]

        for field in preport_fields:
            value = getattr(preports_fee, field)
            if value is not None and value != 0:
                extra_fees["preports"].append(
                    {
                        "name": InvoicePreport._meta.get_field(field).verbose_name,
                        "value": value,
                        "rate": preports_fee.rate.get(field, ""),
                        "qty": preports_fee.qty.get(field, ""),
                        "surcharge": preports_fee.surcharges.get(field, ""),
                        "surcharge_note": preports_fee.surcharge_notes.get(field, ""),
                    }
                )
        if preports_fee.other_fees:
            for name, value in preports_fee.other_fees.items():
                if value and value != 0:
                    extra_fees["preports"].append(
                        {
                            "name": name,
                            "value": value,
                            "rate": value,
                            "qty": 1,
                            "surcharge": None,
                            "surcharge_note": None,
                        }
                    )

        warehouse_fees = InvoiceWarehouse.objects.filter(
            invoice_number__invoice_number=invoice.invoice_number,
            invoice_type="receivable",
        )
        warehouse_fields = [
            "sorting",
            "intercept",
            "po_activation",
            "self_pickup",
            "split_delivery",
            "re_pallet",
            "handling",
            "counting",
            "warehouse_rent",
            "specified_labeling",
            "inner_outer_box",
            "inner_outer_box_label",
            "pallet_label",
            "open_close_box",
            "destroy",
            "take_photo",
            "take_video",
            "repeated_operation_fee",
        ]

        for field in warehouse_fields:
            total_value = 0
            total_surcharge = 0
            surcharge_notes = []

            field_rate = None
            field_qty = None
            for warehouse_fee in warehouse_fees:
                value = getattr(warehouse_fee, field)
                if value is not None and value != 0:
                    total_value += value
                    surcharge = warehouse_fee.surcharges.get(field, 0)
                    if surcharge:
                        total_surcharge += surcharge
                    note = warehouse_fee.surcharge_notes.get(field, "")
                    if note:
                        surcharge_notes.append(note)
                    if field_rate is None:
                        field_rate = warehouse_fee.rate.get(field, "")
                        field_qty = warehouse_fee.qty.get(field, "")
            if total_value != 0:
                extra_fees["warehouse"].append(
                    {
                        "name": InvoiceWarehouse._meta.get_field(field).verbose_name,
                        "value": total_value,
                        "rate": field_rate,
                        "qty": field_qty,
                        "surcharge": total_surcharge,
                        "surcharge_note": "; ".join(filter(None, surcharge_notes)),
                    }
                )

        deliverys = InvoiceDelivery.objects.filter(
            invoice_number=invoice,
            invoice_type="receivable",
        ).exclude(type="combine")
        DELIVERY_TYPE_MAPPING = {
            "selfdelivery": "自发",
            "upsdelivery": "UPS",
            "amazon": "亚马逊",
            "walmart": "沃尔玛",
            "local": "本地派送",
        }
        for delivery in deliverys:
            extra_fees["deliverys"].append(
                {
                    "destination": delivery.destination,
                    "total_pallet": delivery.total_pallet,
                    "total_cbm": delivery.total_cbm,
                    "cost": delivery.cost,
                    "total_weight_lbs": delivery.total_weight_lbs,
                    "total_cost": delivery.total_cost,
                    "note": delivery.note,
                    "type": DELIVERY_TYPE_MAPPING.get(delivery.type, delivery.type),
                }
            )
        return extra_fees

    def _calculate_delivery_fee_cost(
        self,
        fee_details: dict,
        warehouse: str,
        plts_by_destination: list,
        is_new_rule: bool,
        over_count: float,
    ) -> str:
        is_niche_warehouse = True
        if "LA" in warehouse:
            amazon_data = fee_details.get(f"{warehouse}_PUBLIC").details
            walmart_data = None
        else:
            amazon_data = fee_details.get(f"{warehouse}_PUBLIC").details.get(
                f"{warehouse}_AMAZON"
            )
            walmart_data = fee_details.get(f"{warehouse}_PUBLIC").details.get(
                f"{warehouse}_WALMART"
            )
        if warehouse == "NJ":
            local_data = fee_details.get("NJ_LOCAL").details
        for pl in plts_by_destination:
            # 从亚马逊、沃尔玛、本地报价表中挨个找
            # 先找亚马逊
            for price, locations in amazon_data.items():
                if pl["destination"] in locations:
                    pl["price"] = price
                    break
            if pl["price"]:  # 说明这个仓点在这个报价表里，确实是不是冷门仓点
                niche_warehouse = fee_details.get(f"{warehouse}_PUBLIC").niche_warehouse
                if pl["destination"] in niche_warehouse:
                    is_niche_warehouse = True
                else:
                    is_niche_warehouse = False

            # 再找沃尔玛
            if not pl["price"] and walmart_data:
                for price, locations in walmart_data.items():
                    if str(pl["destination"]).upper() in [
                        loc.upper() for loc in locations
                    ]:
                        pl["price"] = price
                        break
                if pl["price"]:
                    niche_warehouse = fee_details.get(
                        f"{warehouse}_PUBLIC"
                    ).niche_warehouse
                    if pl["destination"] in niche_warehouse:
                        is_niche_warehouse = True
                    else:
                        is_niche_warehouse = False
            # 再找本地派送
            if not pl["price"] and warehouse == "NJ":
                destination = re.sub(r"[^0-9]", "", str(pl["destination"]))
                for price, locations in local_data.items():
                    if str(destination) in map(str, locations["zipcodes"]):
                        pl["is_fixed_price"] = (
                            True  # 表示一口价，等会就不会再乘以板数了
                        )
                        costs = locations["prices"]
                        if not over_count:
                            total_pallet = self._calculate_total_pallet(
                                pl["total_cbm"], is_new_rule, is_niche_warehouse
                            )
                        else:
                            total_pallet = over_count
                        if total_pallet <= 5:
                            pl["price"] = int(costs[0])
                        elif total_pallet >= 5:
                            pl["price"] = int(costs[1])
                        break
            if not pl["price"]:
                pl["price"] = 0
            if not over_count:
                total_pallet = self._calculate_total_pallet(
                    pl["total_cbm"], is_new_rule, is_niche_warehouse
                )
            else:
                total_pallet = over_count
            pl["total_pallet"] = total_pallet
        return plts_by_destination

    def handle_container_invoice_direct_get(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.GET.get("container_number")
        order = Order.objects.select_related(
            "retrieval_id", "container_number", "receivable_status", "payable_status"
        ).get(container_number__container_number=container_number)
        warehouse = order.retrieval_id.retrieval_destination_area
        vessel_etd = order.vessel_id.vessel_etd
        customer = order.customer_name
        customer_name = customer.zem_name
        quotation = (
            QuotationMaster.objects.filter(
                effective_date__lte=vessel_etd,
                is_user_exclusive=True,
                exclusive_user=customer_name,
            )
            .order_by("-effective_date")
            .first()
        )
        if not quotation:
            quotation = (
                QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=False,  # 非用户专属的通用报价单
                )
                .order_by("-effective_date")
                .first()
            )
        if not quotation:
            raise ValueError("找不到报价表")
        PICKUP_FEE = FeeDetail.objects.get(quotation_id=quotation.id, fee_type="direct")
        # 提拆、打托缠膜费用
        pickup_fee = 0
        pickup = PICKUP_FEE.details["pickup"]
        for fee, location in pickup.items():
            if warehouse in location:
                pickup_fee = fee
        # 其他费用
        destination = order.retrieval_id.retrieval_destination_area
        new_destination = destination.replace(" ", "") if destination else ""
        second_delivery = PICKUP_FEE.details.get("二次派送")
        second_pickup = None
        for fee, location in second_delivery.items():
            if new_destination in location:
                second_pickup = fee
        FS = {
            "exam_fee": f"{PICKUP_FEE.details.get('查验柜运费', 'N/A')}",  # 查验费
            "second_delivery": second_pickup,  # 二次派送
            "demurrage": f"{PICKUP_FEE.details.get('滞港费', 'N/A')}",  # 滞港费
            "per_diem": f"{PICKUP_FEE.details.get('滞箱费', 'N/A')}",  # 滞箱费
            "congestion_fee": f"{PICKUP_FEE.details.get('港口拥堵费', 'N/A')}",  # 港口拥堵费
            "chassis": f"{PICKUP_FEE.details.get('车架费', 'N/A')}",  # 车架费
            "prepull": f"{PICKUP_FEE.details.get('预提费', 'N/A')}",  # 预提费
            "yard_storage": f"{PICKUP_FEE.details.get('货柜储存费', 'N/A')}",  # 货柜储存费
            "chassis_split": f"{PICKUP_FEE.details.get('车架分离费', 'N/A')}",  # 车架分离费
            "over_weight": f"{PICKUP_FEE.details.get('超重费', 'N/A')}",  # 超重费
        }
        FS_constrain = {}
        for key, value in PICKUP_FEE.details.items():
            if not isinstance(value, dict):
                match = re.findall(r"\$(\d+(\.\d+)?)", str(value))
                if match and len(match) == 1:
                    FS_constrain[key] = float(match[0][0])
        fs_json = json.dumps(FS_constrain, ensure_ascii=False)
        try:
            invoice = Invoice.objects.select_related(
                "customer", "container_number"
            ).get(container_number__container_number=container_number)
        except Invoice.DoesNotExist:
            current_date = datetime.now().date()
            order_id = str(order.id)
            customer_id = order.customer_name.id
            invoice = Invoice(
                **{
                    "invoice_number": f"{current_date.strftime('%Y-%m-%d').replace('-', '')}C{customer_id}{order_id}",
                    "customer": order.customer_name,
                    "container_number": order.container_number,
                }
            )
            invoice.save()
            order.invoice_id = invoice

        # 建立invoicestatus表
        try:
            invoice_status = InvoiceStatus.objects.get(
                container_number=order.container_number, invoice_type="receivable"
            )
        except InvoiceStatus.DoesNotExist:
            invoice_status = InvoiceStatus(
                container_number=order.container_number,
                invoice_type="receivable",
            )
            invoice_status.save()
            order.receivable_status = invoice_status
        order.save()

        invoice_preports, created = InvoicePreport.objects.get_or_create(
            invoice_number=invoice,
            invoice_type="receivable",
            defaults={
                "pickup": pickup_fee,
            },
        )
        # 如果单价和数量都为空的话，就初始化
        renamed_FS_constrain = (
            {  # 因为报价表中，直送和提拆名字不一致，但是表名一致，名称就无法匹配
                (
                    "港口拥堵费"
                    if key == "等待费"
                    else (
                        "查验费"
                        if "查验" in key
                        else (
                            "托架费"
                            if "车架费" in key
                            else (
                                "托架提取费"
                                if key == "车架分离费"
                                else "货柜放置费" if key == "货柜储存费" else key
                            )
                        )
                    )
                ): value
                for key, value in FS_constrain.items()
            }
        )
        if not invoice_preports.qty and not invoice_preports.rate:
            # 提取单价信息
            excluded_fields = {
                "id",
                "invoice_number",
                "invoice_type",
                "amount",
                "qty",
                "rate",
                "other_fees",
                "surcharges",
                "surcharge_notes",
                "history",
            }
            qty_data, rate_data = self._extract_unit_price(
                model=InvoicePreport,
                unit_prices=renamed_FS_constrain,
                pickup_fee=pickup_fee,
                excluded_fields=excluded_fields,
            )
            invoice_preports.qty = qty_data
            invoice_preports.rate = rate_data
            invoice_preports.save()
        else:
            qty_data = invoice_preports.qty
            rate_data = invoice_preports.rate

        context = {
            "warehouse": warehouse,
            "invoice_preports": invoice_preports,
            "container_number": container_number,
            "start_date": request.GET.get("start_date"),
            "end_date": request.GET.get("end_date"),
            "surcharges": invoice_preports.surcharges,
            "surcharges_notes": invoice_preports.surcharge_notes,
            "FS": FS,
            "fs_json": fs_json,
            "status": (order.receivable_status.stage),
            "start_date_confirm": request.GET.get("start_date_confirm") or None,
            "end_date_confirm": request.GET.get("end_date_confirm") or None,
            "confirm_step": request.GET.get("confirm_step") or None,
            "invoice_type": "receivable",
            "qty_data": qty_data,
            "rate_data": rate_data,
        }
        return self.template_invoice_direct_edit, context

    def _extract_unit_price(self, model, unit_prices, pickup_fee, excluded_fields):
        # 构建qty JSON
        qty_data = {}
        rate_data = {}
        # 遍历模型的所有FloatField字段
        for field in model._meta.get_fields():
            if not (
                isinstance(field, models.FloatField)
                and field.name not in excluded_fields
            ):
                continue
            price = unit_prices.get(field.verbose_name, 1.0)
            rate_data[field.name] = float(price) if price not in [None, "N/A"] else 1.0
            qty_data[field.name] = 0
        if pickup_fee:
            rate_data["pickup"] = pickup_fee
            qty_data["pickup"] = 1
        return qty_data, rate_data

    def _create_invoice_defaults(self, order: Order) -> dict:
        current_date = datetime.now().date()
        order_id = str(order.id)
        customer_id = order.customer_name.id

        return {
            "invoice_number": f"{current_date.strftime('%Y%m%d')}C{customer_id}{order_id}",
            "customer": order.customer_name,
            "container_number": order.container_number,
            "receivable_total_amount": 0.0,
            "receivable_preport_amount": 0.0,
            "receivable_warehouse_amount": 0.0,
            "receivable_delivery_amount": 0.0,
            "receivable_direct_amount": 0.0,
            "payable_total_amount": 0.0,
            "payable_basic": 0.0,
            "payable_chassis": 0.0,
            "payable_overweight": 0.0,
            "payable_palletization": 0.0,
            "remain_offset": 0.0,
        }

    def handle_container_invoice_payable_get(
        self, request: HttpRequest, account_confirm: str, is_confirm
    ) -> tuple[Any, Any]:
        container_number = request.GET.get("container_number")
        order = Order.objects.select_related(
            "retrieval_id", "container_number", "warehouse"
        ).get(container_number__container_number=container_number)

        invoice_status, created = InvoiceStatus.objects.get_or_create(
            container_number=order.container_number,
            invoice_type="payable",
            defaults={"stage": "unstarted"},
        )

        if created:
            order.payable_status = invoice_status
            order.save()

        invoice, created = Invoice.objects.get_or_create(
            container_number=order.container_number,
            defaults=self._create_invoice_defaults(order),
        )

        if created:
            order.invoice_id = invoice
            order.save()

        # 确定是否可编辑
        payable_check = self._validate_user_invoice_payable_check(request.user)
        # 未存储过，或者存储了被驳回，或者是组长权限其未被财务确认
        is_editable = (
            invoice_status.stage == "unstarted"
            or invoice_status.is_rejected
            or (payable_check and invoice_status.stage == "tobeconfirmed")
        )
        # 构建基础上下文
        context = {
            "start_date": request.GET.get("start_date"),
            "end_date": request.GET.get("end_date"),
            "warehouse_filter": request.GET.get("warehouse_filter"),
            "is_editable": is_editable,
            "invoice_stage": invoice_status.stage,
            "is_confirm": is_confirm,
            "user_is_leader": payable_check,
            "is_rejected": invoice_status.is_rejected,
            "reject_reason": (
                invoice_status.reject_reason if invoice_status.is_rejected else None
            ),
            "account_confirm": account_confirm,
            "payable_check": payable_check,
            "invoice_number": invoice.invoice_number,
            "container_number": container_number,
            "container_type": order.container_number.container_type,
            "warehouse": order.retrieval_id.retrieval_destination_area,
            "warehouse_precise": order.retrieval_id.retrieval_destination_precise,
            "payable_total_amount": invoice.payable_total_amount,
        }

        preport_carrier = order.retrieval_id.retrieval_carrier
        if preport_carrier == "大方广":
            preport_carrier = "ARM"
        elif preport_carrier == "kars":
            preport_carrier = "Kars"
        context["preport_carrier"] = preport_carrier

        if order.order_type == "直送":
            return self._handle_direct_order(
                context, order, invoice, invoice_status, is_editable
            )

        return self._handle_regular_order(
            context, order, invoice, invoice_status, is_editable
        )

    def _handle_direct_order(
        self, context, order, invoice, invoice_status, is_editable
    ):
        # 处理直送账单
        vessel_etd = order.vessel_id.vessel_etd
        carrier = order.retrieval_id.retrieval_carrier

        # 如果已保存且不是驳回状态，从数据库读取
        if invoice_status.stage != "unstarted" and not invoice_status.is_rejected:
            try:
                invoice_preport = InvoicePreport.objects.get(
                    invoice_number=invoice, invoice_type="payable"
                )
                basic_fee = invoice_preport.pickup
                chassis_fee = (
                    invoice_preport.chassis
                )  # 等待费，报价表里只有等待费的提示信息，所以只要录过的才有等待费
                chassis_comment = invoice_preport.other_fees.get(
                    "chassis_comment", ""
                )  # 不管录没录，都显示提示信息
                context.update(
                    {
                        "chassis_fee": chassis_fee,
                    }
                )
                is_save_invoice = True
            except InvoicePreport.DoesNotExist:
                basic_fee = 0
                chassis_comment = ""
                is_save_invoice = False
        else:
            is_save_invoice = True
            # 从报价表获取
            DETAILS = self._get_feetail(vessel_etd, "PAYABLE_DIRECT")
            destination = order.retrieval_id.retrieval_destination_area

            result = None
            for warehouse, destinations in DETAILS.items():
                if destination in destinations:
                    destination_carriers = destinations[destination]
                    if carrier in destination_carriers:
                        result = destination_carriers[carrier]
                        break
            # 应付报价表，只有两个信息，一个是价格，一个是提示信息
            if result:
                basic_fee = result["price"]
                chassis_comment = result.get("chassis", "")
            else:
                basic_fee = 0
                chassis_comment = "无相关信息"

        context.update(
            {
                "basic_fee": basic_fee,
                "chassis_comment": chassis_comment,
                "is_save_invoice": is_save_invoice,
                "is_editable":is_editable,
            }
        )

        return self.template_invoice_payable_direct_edit, context

    def _handle_regular_order(
        self, context, order, invoice, invoice_status, is_editable
    ):
        # 处理转运&组合账单
        vessel_etd = order.vessel_id.vessel_etd
        if not vessel_etd:
            raise ValueError('缺失ETD时间！')
        act_pick_time = order.retrieval_id.actual_retrieval_timestamp
        warehouse = context["warehouse"]
        warehouse_precise = context["warehouse_precise"]
        preport_carrier = context["preport_carrier"]
        is_first_find = False
        # 初始化费用变量
        fees = {
            "basic_fee": None,
            "overweight_fee": None,
            "chassis_fee": None,
            "arrive_fee": None,
            "palletization_fee": None,
            "palletization_carrier": None,
            "demurrage_fee": None,
            "per_diem_fee": None,
        }
        # 如果有值，从数据库读数据
        try:
            invoice_preport = InvoicePreport.objects.get(
                invoice_number=invoice, invoice_type="payable"
            )
            invoice_warehouse = InvoiceWarehouse.objects.get(
                invoice_number=invoice, invoice_type="payable"
            )

            fees.update(
                {
                    "basic_fee": invoice_preport.pickup,
                    "overweight_fee": (
                        invoice_preport.over_weight
                        if invoice_preport.over_weight
                        and float(invoice_preport.over_weight) > 0
                        else None
                    ),
                    "chassis_fee": (
                        invoice_preport.chassis
                        if invoice_preport.chassis
                        and float(invoice_preport.chassis) > 0
                        else None
                    ),
                    "demurrage_fee": (
                        invoice_preport.demurrage
                        if invoice_preport.demurrage
                        and float(invoice_preport.demurrage) > 0
                        else None
                    ),
                    "per_diem_fee": (
                        invoice_preport.per_diem
                        if invoice_preport.per_diem
                        and float(invoice_preport.per_diem) > 0
                        else None
                    ),
                    "palletization_fee": (
                        invoice_warehouse.palletization_fee
                        if invoice_warehouse.palletization_fee
                        and float(invoice_warehouse.palletization_fee) > 0
                        else None
                    ),
                    "arrive_fee": (
                        invoice_warehouse.arrive_fee
                        if invoice_warehouse.arrive_fee
                        and float(invoice_warehouse.arrive_fee) > 0
                        else None
                    ),
                    "palletization_carrier": invoice_warehouse.carrier,
                }
            )
        except (InvoicePreport.DoesNotExist, InvoiceWarehouse.DoesNotExist):
            is_first_find = True
        # 正常情况下，第一次录，preport等表是空的，但是如果有异常，将状态改为未录入，那么也是初次查找
        if invoice_status.stage == "unstarted" and not invoice_status.is_rejected:
            is_first_find = True
        DETAILS = self._get_feetail(vessel_etd, "PAYABLE")
        precise_warehouse = warehouse_precise.replace("-", " ")
        pallet_details = None
        # 从报价表获取费用标准
        try:
            if warehouse == "NJ":
                if "08817" in warehouse_precise:
                    try:
                        pickup_details = DETAILS[warehouse]["NJ 08817"][preport_carrier]
                        search_carrier = DETAILS[warehouse]["NJ 08817"]
                        # 获取拆柜供应商选项
                        pallet_details = {
                            carrier: value
                            for carrier, details in search_carrier.items()
                            for key in ["palletization", "arrive_warehouse"]
                            if (value := details.get(key)) is not None and value != "/"
                        }
                    except KeyError:
                        pickup_details = None
                        search_carrier = None
                        pallet_details = None
                        fees["basic_fee"] = 0
                    
                else:
                    try:
                        pickup_details = DETAILS[warehouse]["NJ 07001"][preport_carrier]
                        search_carrier = DETAILS[warehouse]["NJ 07001"]
                        # 获取拆柜供应商选项
                        pallet_details = {
                            carrier: value
                            for carrier, details in search_carrier.items()
                            for key in ["palletization", "arrive_warehouse"]
                            if (value := details.get(key)) is not None and value != "/"
                        }
                    except KeyError:
                        pickup_details = None
                        search_carrier = None
                        pallet_details = None
                        fees["basic_fee"] = 0         
            else:
                try:
                    pickup_details = DETAILS[warehouse][precise_warehouse][preport_carrier]
                except KeyError:
                    fees["basic_fee"] = 0
                    pickup_details = None
                pallet_details = None  # 因为只有NJ的有两个拆柜供应商可以切换
            if pickup_details and is_first_find:  # 从未存储过时，才从报价表计算费用
                # 计算基础费用
                if "40" in context["container_type"]:
                    fees["basic_fee"] = pickup_details["basic_40"]
                else:
                    fees["basic_fee"] = pickup_details["basic_45"]

                # 计算超重费
                actual_weight = order.container_number.weight_lbs
                if (
                    pickup_details.get("overweight") not in (None, "/")
                    and float(actual_weight) > 42000
                ):
                    fees["overweight_fee"] = pickup_details.get("overweight")

                # 计算车架费
                cutoff_date = timezone.datetime(2025, 9, 1, tzinfo=timezone.utc)
                if act_pick_time and act_pick_time < cutoff_date:
                    fees = self._calculate_chassis_fee(fees, pickup_details, order)
                else:                   
                    fees = self._calculate_chassis_fee_91(fees, pickup_details, order)

                # 没有拆柜费时，获取入库费
                if "palletization" not in pickup_details or pickup_details.get(
                    "palletization"
                ) in (None, "/"):
                    if pickup_details.get("arrive_warehouse") not in (None, "/"):
                        fees["arrive_fee"] = pickup_details.get("arrive_warehouse")

        except (KeyError, TypeError) as e:
            raise ValueError("报价表查找失败", e)
        
        context["pallet_details"] = pallet_details

        # 获取其他费用
        pickup_other_fee, pallet_other_fee = self._get_other_fees(invoice)
        # 更新时间数据
        time_data = self._calculate_time_data(order)
        context.update(
            {
                "pickup_other_fee": pickup_other_fee,
                "pallet_other_fee": pallet_other_fee,
            }
        )
        context.update(time_data)
        context.update(fees)
        context["is_save_invoice"] = not is_first_find
        context["actual_weight"] = order.container_number.weight_lbs

        return self.template_invoice_payable_edit, context

    def _calculate_chassis_fee_91(self, fees, pickup_details, order):
        #就用还空和实际提柜时间比较，5天内免费
        if pickup_details.get("chassis") not in (None, "/") and pickup_details.get(
            "chassis_free_day"
        ) not in (None, "/"):
            empty_returned_at = order.retrieval_id.empty_returned_at
            act_pick_time = order.retrieval_id.actual_retrieval_timestamp

            returned_date = empty_returned_at.date()
            arrive_date = act_pick_time.date()
            delta_days = (returned_date - arrive_date).days
            delta = delta_days + 1
            if delta > 5:
                fees["chassis_fee"] = (delta - 5) * pickup_details.get("chassis")
        return fees


    def _calculate_chassis_fee(self, fees, pickup_details, order):
        # 计算车架费
        if pickup_details.get("chassis") not in (None, "/") and pickup_details.get(
            "chassis_free_day"
        ) not in (None, "/"):

            arrive_at = order.retrieval_id.arrive_at
            lfd = order.retrieval_id.temp_t49_lfd
            empty_returned_at = order.retrieval_id.empty_returned_at

            if lfd and arrive_at and empty_returned_at:
                arrive_date = arrive_at.date()
                returned_date = empty_returned_at.date()

                if arrive_date < lfd.date():
                    delta = returned_date - arrive_date
                else:
                    delta = returned_date - lfd.date()

                actual_day = delta.days + 1
                free_day = actual_day - int(pickup_details.get("chassis_free_day"))

                if free_day < 0:
                    fees["chassis_fee"] = 0
                else:
                    fees["chassis_fee"] = free_day * pickup_details.get("chassis")

        return fees

    def _get_other_fees(self, invoice):
        # 获取其他费用
        try:
            invoice_preport = InvoicePreport.objects.get(
                invoice_number=invoice, invoice_type="payable"
            )
            pickup_other_fee = invoice_preport.other_fees.get("other_fee", {})
        except InvoicePreport.DoesNotExist:
            pickup_other_fee = {}

        try:
            invoice_warehouse = InvoiceWarehouse.objects.get(
                invoice_number=invoice, invoice_type="payable"
            )
            pallet_other_fee = invoice_warehouse.other_fees.get("other_fee", {})
        except InvoiceWarehouse.DoesNotExist:
            pallet_other_fee = {}

        return pickup_other_fee, pallet_other_fee

    def _calculate_time_data(self, order):
        # 计算时间数据
        arrive_at = order.retrieval_id.arrive_at
        lfd = order.retrieval_id.temp_t49_lfd
        empty_returned_at = order.retrieval_id.empty_returned_at

        actual_day = None
        if lfd and arrive_at and empty_returned_at:
            arrive_date = arrive_at.date()
            returned_date = empty_returned_at.date()

            if arrive_date < lfd.date():
                delta = returned_date - arrive_date
            else:
                delta = returned_date - lfd.date()

            actual_day = delta.days + 1

        return {
            "actual_day": actual_day,
            "lfd": lfd,
            "arrive_date": arrive_at.date() if arrive_at else None,
            "returned_date": empty_returned_at.date() if empty_returned_at else None,
            "empty_returned_at": empty_returned_at,
            "actual_retrieval_timestamp": order.retrieval_id.actual_retrieval_timestamp,
        }

    def _get_feetail(self, vessel_etd, TABLENAME: str) -> Any:
        quotation = (
            QuotationMaster.objects.filter(
                effective_date__lte=vessel_etd, quote_type="payable"
            )
            .order_by("-effective_date")
            .first()
        )
        if not quotation:
            raise ValueError("找不到报价表")
        PAYABLE_FEE = FeeDetail.objects.get(
            quotation_id=quotation.id, fee_type=TABLENAME
        )
        # 规则详情
        DETAILS = PAYABLE_FEE.details
        return DETAILS

    def handle_container_invoice_preport_get(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.GET.get("container_number")
        order = Order.objects.select_related(
            "retrieval_id", "container_number", "warehouse"
        ).get(container_number__container_number=container_number)
        # 查看仓库和柜型，计算提拆费
        warehouse = order.retrieval_id.retrieval_destination_area
        container_type = order.container_number.container_type

        # if warehouse != "LA" and order.order_type == "转运组合":
        #     # 转运组合的，判断下是否符合组合柜规则,因为LA的建单会人工确认是不是组合柜，所以LA的不用管
        #     iscombina = self._is_combina(container_number)
        #     # order_type用于在提拆录入费用界面显示，防止不符合组合柜规定但建单是组合柜的，被提拆同事当组合柜录
        #     order_type = order.container_number.account_order_type
        # else:
        #     order_type = order.order_type
        #     iscombina = False
        #统一查一遍是否符合组合柜规则，
        iscombina = self._is_combina(container_number)
        if order.order_type == "转运组合":
            order_type = order.container_number.account_order_type
        else:
            order_type = order.order_type

        vessel_etd = order.vessel_id.vessel_etd
        customer = order.customer_name
        customer_name = customer.zem_name
        quotation = (
            QuotationMaster.objects.filter(
                effective_date__lte=vessel_etd,
                is_user_exclusive=True,
                exclusive_user=customer_name,
            )
            .order_by("-effective_date")
            .first()
        )
        if not quotation:
            quotation = (
                QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=False,  # 非用户专属的通用报价单
                )
                .order_by("-effective_date")
                .first()
            )
        if not quotation:
            raise ValueError("找不到报价表")
        PICKUP_FEE = FeeDetail.objects.get(
            quotation_id=quotation.id, fee_type="preport"
        )
        # 提拆、打托缠膜费用
        match = re.match(r"\d+", container_type)
        if match:
            pick_subkey = match.group()
            try:
                pickup_fee = PICKUP_FEE.details[warehouse][pick_subkey]
            except KeyError as e:
                pickup_fee = 0

        FS_constrain = {  # 把details里面的键值对改成值是纯数字的，用于在费用表单提交前，验证数据合规性
            key: float(re.search(r"\d+(\.\d+)?", value).group())
            for key, value in PICKUP_FEE.details.items()
            if not isinstance(value, dict)
            and "/" in str(value)
            and re.search(r"\d+(\.\d+)?", value)
        }
        fs_json = json.dumps(FS_constrain, ensure_ascii=False)
        # 其他费用
        FS = {
            "chassis": f"{PICKUP_FEE.details.get('托架费', 'N/A')}",  # 托架费
            "chassis_split": f"{PICKUP_FEE.details.get('托架提取费', 'N/A')}",  # 托架提取费
            "prepull": f"{PICKUP_FEE.details.get('预提费', 'N/A')}",  # 预提费
            "yard_storage": f"{PICKUP_FEE.details.get('货柜放置费', 'N/A')}",  # 货柜放置费
            "handling": f"{PICKUP_FEE.details.get('操作处理费', 'N/A')}",  # 操作处理费
            "pier_pass": PICKUP_FEE.details.get("码头", "N/A"),  # 码头费
            "congestion": f"{PICKUP_FEE.details.get('港口拥堵费', 'N/A')}",  # 港口拥堵费
            "hanging_crane": f"{PICKUP_FEE.details.get('火车站吊柜费', 'N/A')}",  # 火车站吊柜费
            "dry_run": f"{PICKUP_FEE.details.get('空跑费', 'N/A')}",  # 空跑费
            "exam_fee": f"{PICKUP_FEE.details.get('查验费', 'N/A')}",  # 查验费
            "hazmat": f"{PICKUP_FEE.details.get('危险品', 'N/A')}",  # 危险品
            "over_weight": f"{PICKUP_FEE.details.get('超重费', 'N/A')}",  # 超重费
            "urgent_fee": f"{PICKUP_FEE.details.get('加急费', 'N/A')}",  # 加急费
            "other_serive": f"{PICKUP_FEE.details.get('其他服务', 'N/A')}",  # 其他服务
        }
        # 提拆柜费用读取对应表
        try:
            invoice = Invoice.objects.select_related(
                "customer", "container_number"
            ).get(container_number__container_number=container_number)
        except Invoice.DoesNotExist:
            # 没有账单就创建
            order = Order.objects.select_related(
                "customer_name", "container_number"
            ).get(container_number__container_number=container_number)
            current_date = datetime.now().date()
            order_id = str(order.id)
            customer_id = order.customer_name.id
            invoice = Invoice(
                **{
                    "invoice_number": f"{current_date.strftime('%Y-%m-%d').replace('-', '')}C{customer_id}{order_id}",
                    "customer": order.customer_name,
                    "container_number": order.container_number,
                    "receivable_total_amount": 0.0,
                    "receivable_preport_amount": 0.0,
                    "receivable_warehouse_amount": 0.0,
                    "receivable_delivery_amount": 0.0,
                    "receivable_direct_amount": 0.0,
                    "payable_total_amount": 0.0,
                    "payable_basic": 0.0,
                    "payable_chassis": 0.0,
                    "payable_overweight": 0.0,
                    "payable_palletization": 0.0,
                    "remain_offset": 0.0,
                }
            )
            invoice.save()
            order.invoice_id = invoice

        # 建立invoicestatus表
        try:
            invoice_status = InvoiceStatus.objects.get(
                container_number=order.container_number, invoice_type="receivable"
            )
        except InvoiceStatus.DoesNotExist:
            invoice_status = InvoiceStatus(
                container_number=order.container_number,
                invoice_type="receivable",
            )
            invoice_status.save()
            order.receivable_status = invoice_status
        order.save()
        # 建立invoicepreport表
        invoice_preports, created = InvoicePreport.objects.get_or_create(
            invoice_number=invoice,
            invoice_type="receivable",
            defaults={
                "pickup": pickup_fee,
            },
        )
        # 如果单价和数量都为空的话，就初始化
        if not invoice_preports.qty and not invoice_preports.rate:
            # 提取单价信息
            excluded_fields = {
                "id",
                "invoice_number",
                "invoice_type",
                "amount",
                "qty",
                "rate",
                "other_fees",
                "surcharges",
                "surcharge_notes",
                "history",
            }
            qty_data, rate_data = self._extract_unit_price(
                model=InvoicePreport,
                unit_prices=FS_constrain,
                pickup_fee=pickup_fee,
                excluded_fields=excluded_fields,
            )
            invoice_preports.qty = qty_data
            invoice_preports.rate = rate_data
            invoice_preports.save()
        else:
            qty_data = invoice_preports.qty
            rate_data = invoice_preports.rate
        groups = [group.name for group in request.user.groups.all()]
        if request.user.is_staff:
            groups.append("staff")
        step = request.POST.get("step")
        redirect_step = step == "redirect"
        if not iscombina:
            container = Container.objects.get(container_number=container_number)
            non_combina_reason = container.non_combina_reason
        else:
            non_combina_reason = None
        context = {
            "warehouse": warehouse,
            "order_type": order_type,
            "container_type": container_type,
            "reject_reason": order.invoice_reject_reason,
            "invoice_preports": invoice_preports,
            "surcharges": invoice_preports.surcharges,
            "surcharges_notes": invoice_preports.surcharge_notes,
            "container_number": container_number,
            "groups": groups,
            "start_date": request.GET.get("start_date"),
            "end_date": request.GET.get("end_date"),
            "FS": FS,
            "fs_json": fs_json,
            "status": order.receivable_status.stage,
            "redirect_step": redirect_step,
            "start_date_confirm": request.POST.get("start_date_confirm") or None,
            "end_date_confirm": request.POST.get("end_date_confirm") or None,
            "invoice_type": "receivable",
            "qty_data": qty_data,
            "rate_data": rate_data,
            "non_combina_reason": non_combina_reason,
        }
        return self.template_invoice_preport_edit, context

    def handle_container_invoice_get(self, container_number: str) -> tuple[Any, Any]:
        order = Order.objects.select_related("offload_id").get(
            container_number__container_number=container_number
        )
        if order.offload_id.offload_at == None:
            packing_list = (
                PackingList.objects.select_related("container_number", "pallet")
                .filter(container_number__container_number=container_number)
                .values("container_number__container_number", "destination")
                .annotate(
                    total_cbm=Sum("cbm", output_field=FloatField()),
                    total_weight=Sum("total_weight_lbs", output_field=FloatField()),
                    total_n_pallet=Count("id", distinct=True),
                )
                .order_by("destination", "-total_cbm")
            )
        else:
            packing_list = (
                Pallet.objects.select_related("container_number")
                .filter(container_number__container_number=container_number)
                .values("container_number__container_number", "destination")
                .annotate(
                    total_cbm=Sum("cbm", output_field=FloatField()),
                    total_weight=Sum("weight_lbs", output_field=FloatField()),
                    total_n_pallet=Count("pallet_id", distinct=True),
                )
                .order_by("destination", "-total_cbm")
            )
        for pl in packing_list:
            c_p = math.ceil(pl["total_cbm"] / 1.8)
            w_p = math.ceil(pl["total_weight"] / 1000)
            pl["total_n_pallet"] = max(c_p, w_p)
            # if pl["total_cbm"] > 1:
            #     pl["total_n_pallet"] = round(pl["total_cbm"] / 2)
            # elif pl["total_cbm"] >= 0.6 and pl["total_cbm"] <= 1:
            #     pl["total_n_pallet"] = 0.5
            # else:
            #     pl["total_n_pallet"] = 0.25
        context = {
            "order": order,
            "packing_list": packing_list,
        }
        return self.template_invoice_container, context

    def handle_container_invoice_edit_get(
        self, container_number: str
    ) -> tuple[Any, Any]:
        invoice = Invoice.objects.select_related("customer", "container_number").get(
            container_number__container_number=container_number
        )
        invoice_item = InvoiceItem.objects.filter(
            invoice_number__invoice_number=invoice.invoice_number
        )
        context = {
            "invoice": invoice,
            "invoice_item": invoice_item,
        }
        return self.template_invoice_container_edit, context

    def handle_container_invoice_delete_get(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        invoice_number = request.GET.get("invoice_number")
        invoice = Invoice.objects.select_related("container_number").get(
            invoice_number=invoice_number
        )
        invoice_item = InvoiceItem.objects.filter(
            invoice_number__invoice_number=invoice_number
        )
        container_number = invoice.container_number.container_number
        # delete file from sharepoint
        try:
            self._delete_file_from_sharepoint(
                "invoice", f"INVOICE-{container_number}.xlsx"
            )
        except FileNotFoundError as e:
            pass
        except RuntimeError as e:
            raise RuntimeError(e)
        # delete invoice item
        invoice_item.delete()
        # delete invoice
        invoice.delete()
        return self.handle_invoice_get()

    def handle_pallet_data_export_post(self, request: HttpRequest) -> HttpResponse:
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        pallet_data = (
            Order.objects.select_related(
                "container_number",
                "customer_name",
                "warehouse",
                "offload_id",
                "retrieval_id",
            )
            .filter(
                models.Q(offload_id__offload_required=True)
                & models.Q(offload_id__offload_at__isnull=False)
                & models.Q(retrieval_id__actual_retrieval_timestamp__isnull=False)
                & models.Q(offload_id__offload_at__gte=start_date)
                & models.Q(offload_id__offload_at__lte=end_date)
            )
            .order_by("offload_id__offload_at")
        )
        data = [
            {
                "货柜号": d.container_number.container_number,
                "客户": d.customer_name.zem_name,
                "入仓仓库": d.warehouse.name,
                "柜型": d.container_number.container_type,
                "拆柜完成时间": d.offload_id.offload_at.strftime("%Y-%m-%d %H:%M:%S"),
                "打板数": d.offload_id.total_pallet,
            }
            for d in pallet_data
        ]
        df = pd.DataFrame.from_records(data)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f"attachment; filename=pallet_data_{start_date}_{end_date}.xlsx"
        )
        df.to_excel(excel_writer=response, index=False, columns=df.columns)
        return response

    def handle_pl_data_export_post(self, request: HttpRequest) -> HttpResponse:
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        container_number = request.POST.get("container_number")
        _, context = self.handle_pl_data_get(start_date, end_date, container_number)
        data = [
            {
                "货柜号": d["container_number__container_number"],
                "目的地": d["destination"],
                "派送方式": d["delivery_method"],
                "CBM": d["cbm"],
                "箱数": d["pcs"],
                "总重KG": d["total_weight_kg"],
                "总重lbs": d["total_weight_lbs"],
            }
            for d in context["pl_data"]
        ]
        df = pd.DataFrame.from_records(data)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f"attachment; filename=packing_list_data_{start_date}_{end_date}.xlsx"
        )
        df.to_excel(excel_writer=response, index=False, columns=df.columns)
        return response

    def handle_invoice_order_search_post(
        self, request: HttpRequest, status
    ) -> tuple[Any, Any]:
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        order_form = OrderForm(request.POST)
        warehouse = request.POST.get("warehouse_filter")
        if order_form.is_valid():
            customer = order_form.cleaned_data.get("customer_name")
        else:
            customer = None
        preport_carrier = request.POST.get("preport_carrier", None)
        if status == "direct":
            return self.handle_invoice_direct_get(
                request, start_date, end_date, customer, warehouse
            )
        elif status == "combina":
            return self.handle_invoice_combina_get(
                request, start_date, end_date, customer, warehouse
            )
        elif status == "preport":
            return self.handle_invoice_preport_get(
                request, start_date, end_date, customer, warehouse
            )
        elif status == "warehouse":
            return self.handle_invoice_warehouse_get(
                request, start_date, end_date, customer, warehouse
            )
        elif status == "delivery":
            return self.handle_invoice_delivery_get(
                request, start_date, end_date, customer, warehouse
            )
        elif status == "payable":
            return self.handle_invoice_payable_get(
                request, start_date, end_date, customer, warehouse, preport_carrier
            )
        elif status == "confirm":
            start_date_confirm = request.POST.get("start_date_confirm")
            end_date_confirm = request.POST.get("end_date_confirm")
            return self.handle_invoice_confirm_get(
                request, start_date_confirm, end_date_confirm, customer, warehouse
            )
        elif status == "search":
            return self.handle_invoice_search_get(
                request, start_date, end_date, customer, warehouse
            )
        else:
            return self.handle_invoice_get(start_date, end_date, customer)

    def handle_invoice_order_select_post(self, request: HttpRequest) -> HttpResponse:
        selected_orders = json.loads(request.POST.get("selectedOrders", "[]"))
        selected_orders = list(set(selected_orders))
        invoice_type = request.POST.get("invoice_type")
        if selected_orders:
            order = Order.objects.select_related(
                "customer_name", "container_number", "invoice_id"
            ).filter(container_number__container_number__in=selected_orders)
            order_id = [o.id for o in order]
            customer = order[0].customer_name
            current_date = datetime.now().date().strftime("%Y-%m-%d")
            invoice_statement_id = (
                f"{current_date.replace('-', '')}S{customer.id}{max(order_id)}"
            )
            context = {
                "order": order,
                "customer": customer,
                "invoice_statement_id": invoice_statement_id,
                "current_date": current_date,
                "invoice_type": invoice_type,
            }
            return render(request, self.template_invoice_statement, context)
        else:
            template, context = self.handle_invoice_order_search_post(request)
            return render(request, template, context)

    def handle_create_container_invoice_post(
        self, request: HttpRequest
    ) -> HttpResponse:
        container_number = request.POST.get("container_number")
        if self._check_invoice_exist(container_number):
            raise RuntimeError(f"货柜-{container_number}已生成invoice!")
        description = request.POST.getlist("description")
        warehouse_code = request.POST.getlist("warehouse_code")
        cbm = request.POST.getlist("cbm")
        weight = request.POST.getlist("weight")
        qty = request.POST.getlist("qty")
        rate = request.POST.getlist("rate")
        amount = request.POST.getlist("amount")
        note = request.POST.getlist("note")
        order = Order.objects.select_related("customer_name", "container_number").get(
            container_number__container_number=container_number
        )
        context = {
            "order": order,
            "container_number": container_number,
            "data": zip(
                description, warehouse_code, cbm, weight, qty, rate, amount, note
            ),
        }
        workbook, invoice_data = self._generate_invoice_excel(context)
        invoice = Invoice(
            **{
                "invoice_number": invoice_data["invoice_number"],
                "invoice_date": invoice_data["invoice_date"],
                "invoice_link": invoice_data["invoice_link"],
                "customer": context["order"].customer_name,
                "container_number": context["order"].container_number,
                "receivable_total_amount": invoice_data["total_amount"],
                "remain_offset": invoice_data["total_amount"],
            }
        )
        invoice.save()
        order.invoice_id = invoice
        order.save()
        invoice_item_data = []
        for d, wc, c, w, q, r, a, n in zip(
            description, warehouse_code, cbm, weight, qty, rate, amount, note
        ):
            invoice_item_data.append(
                {
                    "invoice_number": invoice,
                    "description": d,
                    "warehouse_code": wc,
                    "cbm": c if c else None,
                    "weight": w if w else None,
                    "qty": q if q else None,
                    "rate": r if r else None,
                    "amount": a if a else None,
                    "note": n if n else "",
                }
            )
        invoice_item_instances = [
            InvoiceItem(**inv_itm_data) for inv_itm_data in invoice_item_data
        ]
        bulk_create_with_history(invoice_item_instances, InvoiceItem)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f"attachment; filename=INVOICE-{container_number}.xlsx"
        )
        workbook.save(response)
        return response

    def handle_export_invoice_post(self, request: HttpRequest) -> HttpResponse:
        resp, file_name, pdf_file, context = export_invoice(request)
        pdf_file.seek(0)
        invoice = Invoice.objects.select_related("statement_id").filter(
            models.Q(container_number__container_number__in=context["container_number"])
        )
        invoice_statement = InvoiceStatement.objects.filter(
            models.Q(
                invoice__container_number__container_number__in=context[
                    "container_number"
                ]
            )
        )
        for invc_stmt in invoice_statement.distinct():
            try:
                self._delete_file_from_sharepoint(
                    "invoice_statement",
                    f"invoice_{invc_stmt.invoice_statement_id}_from_ZEM_ELITELINK LOGISTICS_INC.pdf",
                )
            except:
                pass
        invoice_statement.delete()
        link = self._upload_excel_to_sharepoint(
            pdf_file, "invoice_statement", file_name
        )
        invoice_statement = InvoiceStatement(
            **{
                "invoice_statement_id": context["invoice_statement_id"],
                "statement_amount": context["total_amount"],
                "statement_date": context["invoice_date"],
                "due_date": context["due_date"],
                "invoice_terms": context["invoice_terms"],
                "customer": Customer.objects.get(accounting_name=context["customer"]),
                "statement_link": link,
            }
        )
        invoice_statement.save()
        for invc in invoice:
            invc.statement_id = invoice_statement
        bulk_update_with_history(invoice, Invoice, fields=["statement_id"])
        return resp

    def handle_container_invoice_edit_post(self, request: HttpRequest) -> HttpResponse:
        invoice_number = request.POST.get("invoice_number")
        invoice = Invoice.objects.select_related("container_number").get(
            invoice_number=invoice_number
        )
        container_number = invoice.container_number.container_number
        description = request.POST.getlist("description")
        warehouse_code = request.POST.getlist("warehouse_code")
        cbm = request.POST.getlist("cbm")
        weight = request.POST.getlist("weight")
        qty = request.POST.getlist("qty")
        rate = request.POST.getlist("rate")
        amount = request.POST.getlist("amount")
        note = request.POST.getlist("note")
        order = Order.objects.select_related("customer_name").get(
            invoice_id__invoice_number=invoice_number
        )
        context = {
            "order": order,
            "container_number": container_number,
            "data": zip(
                description, warehouse_code, cbm, weight, qty, rate, amount, note
            ),
        }

        # delete old file from sharepoint
        try:
            self._delete_file_from_sharepoint(
                "invoice", f"INVOICE-{container_number}.xlsx"
            )
        except FileNotFoundError as e:
            pass
        except RuntimeError as e:
            raise RuntimeError(f"Error: {e}")
        # create new file and upload to sharepoint
        workbook, invoice_data = self._generate_invoice_excel(context)
        # update invoice information
        # invoice.invoice_number = invoice_data["invoice_number"]
        # invoice.invoice_date = invoice_data["invoice_date"]
        invoice.invoice_link = invoice_data["invoice_link"]
        invoice.receivable_total_amount = invoice_data["total_amount"]
        invoice.remain_offset = invoice_data["total_amount"]
        invoice.save()
        # update invoice item information
        InvoiceItem.objects.filter(
            invoice_number__invoice_number=invoice_number
        ).delete()
        invoice_item_data = []
        for d, wc, c, q, r, a, n in zip(
            description, warehouse_code, cbm, qty, rate, amount, note
        ):
            invoice_item_data.append(
                {
                    "invoice_number": invoice,
                    "description": d,
                    "warehouse_code": wc,
                    "cbm": c if c else None,
                    "qty": q if q else None,
                    "rate": r if r else None,
                    "amount": a if a else None,
                    "note": n if n else None,
                }
            )
        invoice_item_instances = [
            InvoiceItem(**inv_itm_data) for inv_itm_data in invoice_item_data
        ]
        bulk_create_with_history(invoice_item_instances, InvoiceItem)
        # export new file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f"attachment; filename=INVOICE-{container_number}.xlsx"
        )
        workbook.save(response)
        return response

    def handle_invoice_order_batch_reject(self, request: HttpRequest) -> HttpResponse:
        selected_orders = json.loads(request.POST.get("selectedOrders", "[]"))
        selected_orders = list(set(selected_orders))
        invoice_status = InvoiceStatus.objects.filter(
            container_number__container_number__in=selected_orders
        )
        for item in invoice_status:
            item.stage = "tobeconfirmed"
            item.save()

        # 重开账单，需要撤销通知客户
        invoices = Invoice.objects.prefetch_related(
            "order", "order__container_number", "container_number"
        ).filter(container_number__container_number__in=selected_orders)
        for invoice in invoices:
            invoice.is_invoice_delivered = False
            invoice.save()
        return self.handle_invoice_confirm_get(
            request,
            request.POST.get("start_date_confirm"),
            request.POST.get("end_date_confirm"),
        )

    def handle_invoice_order_batch_delivered(
        self, request: HttpRequest
    ) -> HttpResponse:
        selected_orders = json.loads(request.POST.get("selectedOrders", "[]"))
        selected_orders = list(set(selected_orders))
        invoices = Invoice.objects.prefetch_related(
            "order", "order__container_number", "container_number"
        ).filter(container_number__container_number__in=selected_orders)
        for invoice in invoices:
            invoice.is_invoice_delivered = True
            invoice.save()
        return self.handle_invoice_confirm_get(
            request,
            request.POST.get("start_date_confirm"),
            request.POST.get("end_date_confirm"),
        )

    def handle_invoice_order_batch_export(self, request: HttpRequest) -> HttpResponse:
        selected_orders = json.loads(request.POST.get("selectedOrders", "[]"))
        selected_orders = list(set(selected_orders))
        orders = Order.objects.select_related(
            "retrieval_id", "container_number"
        ).filter(container_number__container_number__in=selected_orders)
        invoices = Invoice.objects.prefetch_related(
            "order", "order__container_number", "container_number"
        ).filter(container_number__container_number__in=selected_orders)
        data = [
            (
                order,
                invoices.get(
                    container_number__container_number=order.container_number.container_number
                ),
            )
            for order in orders
        ]
        invoice_type = request.POST.get("invoice_type")
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for order, invoice in data:
                context = self._parse_invoice_excel_data(order, invoice, invoice_type)
                workbook, _ = self._generate_invoice_excel(
                    context, save_to_sharepoint=False
                )
                excel_file = io.BytesIO()
                workbook.save(excel_file)
                excel_file.seek(0)  # Go to the beginning of the in-memory file
                zip_file.writestr(
                    f"INVOICE-{order.container_number.container_number}.xlsx",
                    excel_file.read(),
                )
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.read(), content_type="application/zip")
        response["Content-Disposition"] = 'attachment; filename="invoices.zip"'
        return response

    def handle_adjust_balance_save(self, request: HttpRequest) -> tuple[Any, Any]:
        customer_id = request.POST.get("customerId")
        customer = Customer.objects.get(id=customer_id)
        amount = float(request.POST.get("usdamount"))
        note = request.POST.get("note")
        user = request.user if request.user.is_authenticated else None

        selected_orders = json.loads(request.POST.get("selectedOrders", "[]"))
        selected_orders = list(set(selected_orders))
        # 查账单，按待核销金额从小到大排序
        invoices = Invoice.objects.filter(
            container_number__container_number__in=selected_orders
        ).order_by("remain_offset")
        sum_offset = 0.0
        for invoice in invoices:
            if amount <= 0:
                break
            offset_amount = min(amount, invoice.remain_offset)
            sum_offset += offset_amount
            invoice.remain_offset -= offset_amount
            invoice.save()
            amount -= offset_amount

        transaction = Transaction.objects.create(
            customer=customer,
            amount=sum_offset,
            transaction_type="write_off",
            note=note,
            created_by=user,
            created_at=timezone.now(),
        )

        customer.balance = customer.balance - sum_offset
        customer.save()
        return self.handle_invoice_confirm_get(
            request,
            request.POST.get("start_date_confirm"),
            request.POST.get("end_date_confirm"),
        )

    def _generate_invoice_excel(
        self,
        context: dict[Any, Any],
        save_to_sharepoint: bool = True,
    ) -> tuple[openpyxl.workbook.Workbook, dict[Any, Any]]:
        current_date = datetime.now().date()
        order_id = str(context["order"].id)
        customer_id = context["order"].customer_name.id
        invoice_number = f"{current_date.strftime('%Y-%m-%d').replace('-', '')}C{customer_id}{order_id}"
        workbook = openpyxl.Workbook()  # 创建一个工作簿对象
        worksheet = workbook.active  # 获取工作簿的活动工作表
        worksheet.title = "Sheet1"  # 给表命名
        cells_to_merge = [  # 要合并的单元格
            "A1:E1",
            "A3:A4",
            "B3:D3",
            "B4:D4",
            "E3:E4",
            "F3:I4",
            "A5:A6",
            "B5:D5",
            "B6:D6",
            "E5:E6",
            "F5:I6",
            "A9:B9",
            "A10:B10",
            "F1:I1",
            "C1:E1",
            "A2:I2",
            "A7:I7",
            "A8:I8",
            "C9:I9",
            "C10:I10",
            "A11:I11",
        ]
        self._merge_ws_cells(worksheet, cells_to_merge)  # 进行合并

        worksheet.column_dimensions["A"].width = 18
        worksheet.column_dimensions["B"].width = 18
        worksheet.column_dimensions["C"].width = 15
        worksheet.column_dimensions["D"].width = 7
        worksheet.column_dimensions["E"].width = 8
        worksheet.column_dimensions["F"].width = 7
        worksheet.column_dimensions["G"].width = 11
        worksheet.column_dimensions["H"].width = 11
        worksheet.column_dimensions["I"].width = 11
        worksheet.row_dimensions[1].height = 40

        worksheet["A1"] = "Zem Elitelink Logistics Inc"
        worksheet["A3"] = "NJ"
        worksheet["B3"] = "27 Engelhard Ave. Avenel NJ 07001"
        worksheet["B4"] = "Contact: Marstin Ma 929-810-9968"
        worksheet["A5"] = "SAV"
        worksheet["B5"] = "1001 Trade Center Pkwy, Rincon, GA 31326, USA"
        worksheet["B6"] = "Contact: Ken 929-329-4323"
        worksheet["A7"] = "E-mail: OFFICE@ZEMLOGISTICS.COM"
        worksheet["A9"] = "BILL TO"
        worksheet["A10"] = context["order"].customer_name.zem_name
        worksheet["F1"] = "Invoice"
        worksheet["E3"] = "Date"
        worksheet["F3"] = current_date.strftime("%Y-%m-%d")
        worksheet["E5"] = "Invoice #"
        worksheet["F5"] = invoice_number

        worksheet["A1"].font = Font(size=20)
        worksheet["F1"].font = Font(size=28)
        worksheet["A3"].alignment = Alignment(vertical="center")
        worksheet["A5"].alignment = Alignment(vertical="center")
        worksheet["E3"].alignment = Alignment(vertical="center")
        worksheet["E5"].alignment = Alignment(vertical="center")
        worksheet["F3"].alignment = Alignment(vertical="center")
        worksheet["F5"].alignment = Alignment(vertical="center")

        worksheet.append(
            [
                "CONTAINER #",
                "DESCRIPTION",
                "WAREHOUSE CODE",
                "CBM",
                "WEIGHT",
                "QTY",
                "RATE",
                "AMOUNT",
                "NOTE",
            ]
        )  # 添加表头
        invoice_item_starting_row = 12
        invoice_item_row_count = 0
        row_count = 13
        total_amount = 0.0
        total_cbm = 0.0
        total_weight = 0.0
        for d, wc, cbm, weight, qty, r, amt, n in context["data"]:
            if r == {}:
                continue
            worksheet.append(
                [context["container_number"], d, wc, cbm, weight, qty, r, amt, n]
            )  # 添加数据
            total_amount += float(amt)  # 计算总金额
            total_cbm += float(cbm) if cbm else 0
            total_weight += float(weight) if weight else 0
            row_count += 1
            invoice_item_row_count += 1
        worksheet.append(
            [
                "Total",
                None,
                None,
                total_cbm,
                total_weight,
                None,
                None,
                total_amount,
                None,
            ]
        )  # 工作表末尾添加总金额
        invoice_item_row_count += 1
        for row in worksheet.iter_rows(  # 单元格设置样式
            min_row=invoice_item_starting_row,
            max_row=invoice_item_starting_row + invoice_item_row_count,
            min_col=1,
            max_col=9,
        ):
            for cell in row:
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
        self._merge_ws_cells(worksheet, [f"A{row_count}:C{row_count}"])
        self._merge_ws_cells(worksheet, [f"F{row_count}:G{row_count}"])
        worksheet[f"A{row_count}"].alignment = Alignment(horizontal="center")
        worksheet[f"G{row_count}"].number_format = numbers.FORMAT_NUMBER_00
        worksheet[f"G{row_count}"].alignment = Alignment(horizontal="left")
        row_count += 1
        self._merge_ws_cells(worksheet, [f"A{row_count}:I{row_count}"])
        row_count += 1

        bank_info = [
            f"Beneficiary Name: {ACCT_BENEFICIARY_NAME}",
            f"Bank Name: {ACCT_BANK_NAME}",
            f"SWIFT Code: {ACCT_SWIFT_CODE}",
            f"ACH/Wire Transfer Routing Number: {ACCT_ACH_ROUTING_NUMBER}",
            f"Beneficiary Account #: {ACCT_BENEFICIARY_ACCOUNT}",
            f"Beneficiary Address: {ACCT_BENEFICIARY_ADDRESS}",
            f"Email:FINANCE@ZEMLOGISTICS.COM",
            f"phone: 929-810-9968",
        ]
        for c in bank_info:
            worksheet.append([c])
            self._merge_ws_cells(worksheet, [f"A{row_count}:I{row_count}"])
            row_count += 1
        self._merge_ws_cells(worksheet, [f"A{row_count}:I{row_count}"])

        excel_file = io.BytesIO()  # 创建一个BytesIO对象
        workbook.save(excel_file)  # 将workbook保存到BytesIO中
        excel_file.seek(0)  # 将文件指针移动到文件开头
        if save_to_sharepoint:
            invoice_link = self._upload_excel_to_sharepoint(
                excel_file, "invoice", f"INVOICE-{context['container_number']}.xlsx"
            )
        else:
            invoice_link = ""

        worksheet["A9"].font = Font(color="00FFFFFF")
        worksheet["A9"].fill = PatternFill(
            start_color="00000000", end_color="00000000", fill_type="solid"
        )
        invoice_data = {
            "invoice_number": invoice_number,
            "invoice_date": current_date.strftime("%Y-%m-%d"),
            "invoice_link": invoice_link,
            "total_amount": total_amount,
        }
        return workbook, invoice_data

    def _merge_ws_cells(
        self, ws: openpyxl.worksheet.worksheet, cells: list[str]
    ) -> None:
        for c in cells:
            ws.merge_cells(c)

    def _parse_invoice_excel_data(
        self, order: Order, invoice: Invoice, invoice_type: str
    ) -> dict[str, Any]:
        description = []
        warehouse_code = []
        cbm = []
        weight = []
        qty = []
        rate = []
        amount = []
        note = []
        if order.order_type == "直送":
            invoice_preport = InvoicePreport.objects.get(
                invoice_number__invoice_number=invoice.invoice_number,
                invoice_type=invoice_type,
            )
            for field in invoice_preport._meta.fields:
                if isinstance(field, models.FloatField) and field.name != "amount":
                    value = getattr(invoice_preport, field.name)
                    if value not in [None, 0]:
                        if field.verbose_name == "操作处理费":
                            description.append("等待费")
                        elif field.verbose_name == "提拆/打托缠膜":
                            description.append("提派")
                        elif field.verbose_name == "货柜放置费":
                            description.append("货柜储存费")
                        else:
                            description.append(field.verbose_name)
                        surcharge = invoice_preport.surcharges.get(field.name, 0)
                        if surcharge > 0:
                            value += surcharge
                            note.append(
                                f"{invoice_preport.surcharge_notes.get(field.name)}: ${surcharge}"
                            )
                        else:
                            surcharge_note = invoice_preport.surcharge_notes.get(
                                field.name
                            )
                            if surcharge_note and surcharge_note.strip():
                                note.append(surcharge_note)
                            else:
                                note.append("")
                        warehouse_code.append("")
                        cbm.append("")
                        weight.append("")
                        qty.append(invoice_preport.qty[field.name])
                        rate.append(invoice_preport.rate[field.name])
                        amount.append(value)

            for k, v in invoice_preport.other_fees.items():
                description.append(k)
                amount.append(v)
                warehouse_code.append("")
                cbm.append("")
                weight.append("")
                qty.append(1)
                rate.append(v)
                note.append("")
        else:
            if (
                order.order_type != "转运"
                and order.container_number.account_order_type == "转运组合"
            ):
                # 组合柜就从invoiceItem表找就行了，转运的才去三个表找
                invoice_item = InvoiceItem.objects.filter(
                    invoice_number__invoice_number=invoice.invoice_number
                )
                if invoice_item is None:
                    raise ValueError("缺少账单明细表")
                for item in invoice_item:
                    description.append(item.description)
                    warehouse_code.append(item.warehouse_code)
                    cbm.append(item.cbm)
                    weight.append(item.weight)
                    qty.append(item.qty)
                    rate.append(item.rate)
                    amount.append(item.amount)
                    note.append(item.note)
            else:
                invoice_preport = InvoicePreport.objects.get(
                    invoice_number__invoice_number=invoice.invoice_number,
                    invoice_type=invoice_type,
                )
                invoice_warehouse = InvoiceWarehouse.objects.filter(
                    invoice_number__invoice_number=invoice.invoice_number
                )
                invoice_delivery = InvoiceDelivery.objects.filter(
                    invoice_number__invoice_number=invoice.invoice_number
                )

                for field in invoice_preport._meta.fields:
                    if (
                        isinstance(field, models.FloatField)
                        and field.name != "amount"
                        and field.name != "other_fees"
                    ):
                        value = getattr(invoice_preport, field.name)
                        if value not in [None, 0, {}]:
                            description.append(field.verbose_name)
                            warehouse_code.append("")
                            cbm.append("")
                            weight.append("")
                            qty.append(invoice_preport.qty[field.name])
                            rate.append(invoice_preport.rate[field.name])
                            amount.append(value)
                            #note.append("")
                        #if field.verbose_name == "港口拥堵费":
                            note.append(invoice_preport.surcharge_notes.get(field.name))
                for k, v in invoice_preport.other_fees.items():
                    if v not in [None, 0, {}]:
                        description.append(k)
                        amount.append(v)
                        warehouse_code.append("")
                        cbm.append("")
                        weight.append("")
                        qty.append(1)
                        rate.append(v)
                        note.append("")
                for warehouse in invoice_warehouse:
                    for field in warehouse._meta.fields:
                        if (
                            isinstance(field, models.FloatField)
                            and field.name != "amount"
                            and field.name
                            not in ["palletization_fee", "arrive_fee", "carrier"]
                        ):
                            value = getattr(warehouse, field.name)
                            if value not in [None, 0]:
                                description.append(field.verbose_name)
                                warehouse_code.append("")
                                cbm.append("")
                                weight.append("")
                                qty.append(warehouse.qty[field.name])
                                rate.append(warehouse.rate[field.name])
                                amount.append(value)
                                #note.append("")
                                note.append(warehouse.surcharge_notes.get(field.name))
                for warehouse in invoice_warehouse:
                    for k, v in warehouse.other_fees.items():
                        if v not in [None, 0]:
                            description.append(k)
                            warehouse_code.append("")
                            cbm.append("")
                            weight.append("")
                            qty.append(1)
                            rate.append(v)
                            amount.append(v)
                            #note.append("")
                            note.append(warehouse.surcharge_notes.get(field.name))
                for delivery in invoice_delivery:
                    if delivery.total_cost is None:
                        raise ValueError("派送费为空")
                    description.append("派送费")
                    warehouse_code.append(delivery.destination.upper())
                    cbm.append(delivery.total_cbm)
                    weight.append(delivery.total_weight_lbs)
                    qty.append(delivery.total_pallet)
                    amount.append(delivery.total_cost)
                    note.append(delivery.note)
                    
                    try:
                        rate.append(int(delivery.total_cost / delivery.total_pallet))
                    except:
                        rate.append("")
        context = {
            "order": order,
            "container_number": order.container_number.container_number,
            "data": zip(
                description, warehouse_code, cbm, weight, qty, rate, amount, note
            ),
        }
        return context

    def _get_sharepoint_auth(self) -> ClientContext:
        ctx = ClientContext(SP_URL).with_client_certificate(
            SP_TENANT,
            SP_CLIENT_ID,
            SP_THUMBPRINT,
            private_key=SP_PRIVATE_KEY,
            scopes=[SP_SCOPE],
        )
        return ctx

    def _upload_excel_to_sharepoint(
        self, file: BytesIO, schema: str, file_name: str
    ) -> str:
        conn = self._get_sharepoint_auth()
        file_path = os.path.join(SP_DOC_LIB, f"{SYSTEM_FOLDER}/{schema}/{APP_ENV}")
        sp_folder = conn.web.get_folder_by_server_relative_url(file_path)
        resp = sp_folder.upload_file(f"{file_name}", file).execute_query()
        link = (
            resp.share_link(SharingLinkKind.OrganizationView)
            .execute_query()
            .value.to_json()["sharingLinkInfo"]["Url"]
        )
        return link

    def _delete_file_from_sharepoint(
        self,
        schema: str,
        file_name: str,
    ) -> None:
        conn = self._get_sharepoint_auth()
        file_path = os.path.join(
            SP_DOC_LIB, f"{SYSTEM_FOLDER}/{schema}/{APP_ENV}/{file_name}"
        )
        try:
            conn.web.get_file_by_server_relative_url(
                file_path
            ).delete_object().execute_query()
        except ClientRequestException as e:
            if e.response.status_code == 404:
                raise FileNotFoundError(e)
            else:
                raise RuntimeError(e)

    # 按照权限分组，有三个分组：客服组（添加账单详情）、组长组（确认客服组操作）、财务组（确认账单）
    def _validate_user_group(self, user: User) -> bool:
        if user.groups.filter(name=self.allowed_group).exists():
            return True
        else:
            return False

    def _validate_user_invoice_direct(self, user: User) -> bool:
        if user.is_staff or user.groups.filter(name="invoice_direct").exists():
            return True
        else:
            return False

    def _validate_user_invoice_combina(self, user: User) -> bool:
        if user.is_staff or user.groups.filter(name="invoice_combina").exists():
            return True
        else:
            return False

    def _validate_user_invoice_preport(self, user: User) -> bool:
        if user.is_staff or user.groups.filter(name="invoice_preport").exists():
            return True
        elif user.groups.filter(name="invoice_preport_leader").exists():
            return True
        else:
            return False

    def _validate_user_invoice_warehouse(self, user: User) -> bool:
        if user.is_staff or user.groups.filter(name="invoice_warehouse").exists():
            return True
        else:
            return False

    def _validate_user_invoice_delivery(self, user: User) -> bool:
        if user.is_staff or user.groups.filter(name="invoice_delivery").exists():
            return True
        else:
            return False

    def _validate_user_invoice_confirm(self, user: User) -> bool:
        if user.is_staff or user.groups.filter(name="invoice_confirm").exists():
            return True
        else:
            return False

    def _check_invoice_exist(self, container_number: str) -> bool:
        return Invoice.objects.filter(
            container_number__container_number=container_number
        ).exists()

    def _validate_user_invoice_payable_check(self, user: User) -> bool:
        # if user.groups.filter(name="invoice_paybale_check").exists():
        if user.is_staff or user.groups.filter(name="invoice_paybale_check").exists():
            return True
        else:
            return False
