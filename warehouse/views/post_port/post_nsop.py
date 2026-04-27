from datetime import datetime, timedelta, time
from typing import Any, Dict, List, Tuple
from django.db.models import Prefetch, F, Subquery, OuterRef, Exists
from collections import OrderedDict, defaultdict
import pandas as pd
import json
import uuid
import pytz
import os
import random
import xml.etree.ElementTree as ET
import platform
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from django.db import transaction
import re
import base64
import io
import zipfile
import barcode
from PIL import Image
from django.template.loader import get_template
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from xhtml2pdf import pisa
from barcode.writer import ImageWriter
from django.db.models.functions import Ceil, Length
from django.utils.safestring import mark_safe
from django.utils.html import escape
from asgiref.sync import sync_to_async
from django.contrib.postgres.aggregates import StringAgg
from django.db.models.functions import Round, Cast, Coalesce
from django.core.exceptions import ObjectDoesNotExist
from simple_history.utils import bulk_update_with_history
from django.db import models
from django.db.models.expressions import ExpressionWrapper
import math  
from django.db.models import (
    Case,
    CharField,
    BooleanField,
    Count,
    F,
    Func,
    FloatField,
    IntegerField,
    Max,
    Q,
    Sum,
    Value,
    When,
)
from warehouse.utils.config import app_config
import asyncio
import aiohttp
from django.http import JsonResponse, HttpResponseForbidden
from io import BytesIO
from django.core.exceptions import MultipleObjectsReturned
from django.db.models.functions import Cast, Concat
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect, render
from django.views import View
from django.utils import timezone
from datetime import timedelta, datetime, date
from dateutil.parser import parse
from django.core.serializers.json import DjangoJSONEncoder
from django.contrib.auth.models import User
from warehouse.forms.warehouse_form import ZemWarehouseForm
from warehouse.models.packing_list import PackingList
from warehouse.models.pallet import Pallet
from asgiref.sync import sync_to_async
from warehouse.models.quotation_master import QuotationMaster
from warehouse.models.fleet import Fleet
from warehouse.models.order import Order
from warehouse.models.po_check_eta import PoCheckEtaSeven
from warehouse.models.shipment import Shipment
from warehouse.models.fee_detail import FeeDetail
from warehouse.models.container import Container
from warehouse.models.invoicev2 import (
    Invoicev2,
    InvoiceItemv2,
    InvoiceStatusv2,
)
from django.contrib import messages
from warehouse.models.transfer_location import TransferLocation
from warehouse.views.post_port.shipment.fleet_management import FleetManagement
from warehouse.views.post_port.shipment.shipping_management import ShippingManagement
from warehouse.views.receivable_accounting import ReceivableAccounting
from warehouse.views.po import PO
from warehouse.views.export_file import link_callback
from warehouse.utils.constants import (
    LOAD_TYPE_OPTIONS,
    amazon_fba_locations,
    NJ_DES,SAV_DES,LA_DES,
    DELIVERY_METHOD_OPTIONS
)
FOUR_MAJOR_WAREHOUSES = ["ONT8", "LAX9", "LGB8", "SBD1"]


class PostNsop(View):
    template_main_dash = "post_port/new_sop/01_appointment/01_appointment_management.html"
    template_td_shipment = "post_port/new_sop/02_shipment/02_td_shipment.html"
    template_td_unshipment = "post_port/new_sop/02_1_shipment/unscheduled_section.html"
    template_fleet_schedule = "post_port/new_sop/03_fleet_schedule/03_fleet_schedule.html"
    template_unscheduled_pos_all = "post_port/new_sop/01_unscheduled_pos_all/01_unscheduled_main.html"
    template_ltl_pos_all = "post_port/new_sop/05_ltl_pos_all/05_ltl_main.html"
    template_ltl_history_pos = "post_port/new_sop/06_ltl_history_pos/06_ltl_main.html"
    template_history_shipment = "post_port/new_sop/04_history_shipment/04_history_shipment_main.html"
    template_batch_shipment = "post_port/new_sop/leader_check/batch_shipment.html"
    template_fleet_check = "post_port/new_sop/leader_check/fleet_check.html"
    template_master_shipment_check = "post_port/new_sop/leader_check/master_shipment_check.html"
    template_client_exception = "post_port/new_sop/leader_check/client_exception.html"
    template_pod_reupload = "post_port/new_sop/leader_check/pod_reupload.html"
    template_fleet_po_check = "post_port/new_sop/leader_check/fleet_po_check.html"
    template_bol = "export_file/bol_base_template.html"
    template_bol_pickup = "export_file/bol_template.html"
    template_la_bol_pickup = "export_file/LA_bol_template.html"
    template_ltl_label = "export_file/ltl_label.html"
    template_ltl_bol = "export_file/ltl_bol.html"
    template_ltl_bol_multi = "export_file/ltl_bol_multi.html"
    area_options = {"NJ": "NJ", "SAV": "SAV", "LA": "LA", "MO": "MO", "TX": "TX", "LA": "LA"}
    warehouse_options = {"":"", "NJ-07001": "NJ-07001", "SAV-31326": "SAV-31326", "LA-91761": "LA-91761", "LA-91748": "LA-91748", "LA-91766": "LA-91766", "LA-91730": "LA-91730"}
    load_type_options = {
        "еН°жЭњ": "еН°жЭњ",
        "еЬ∞жЭњ": "еЬ∞жЭњ",
    }
    account_options = {
        "": "",
        "Carrier Central1": "Carrier Central1",
        "Carrier Central2": "Carrier Central2",
        "ZEM-AMF": "ZEM-AMF",
        "ARM-AMF": "ARM-AMF",
        "walmart": "walmart",
    }
    arm_account_options = {
        "": "",
        "Carrier Central ADWG": "Carrier Central ADWG",
        "Carrier Central1": "Carrier Central1",
        "Carrier Central2": "Carrier Central2",
        "ZEM-AMF": "ZEM-AMF",
        "ARM-AMF": "ARM-AMF",
        "walmart": "walmart",
    }
    abnormal_fleet_options = {
        "": "",
        "еПЄжЬЇжЬ™жМЙжЧґжПРиіІ": "еПЄжЬЇжЬ™жМЙжЧґжПРиіІ",
        "йАБдїУиҐЂжЛТжФґ": "йАБдїУиҐЂжЛТжФґ",
        "жЬ™йАБиЊЊ": "жЬ™йАБиЊЊ",
        "еЕґеЃГ": "еЕґеЃГ",
    }
    carrier_options = {
        "": "",
        "Arm-AMF": "Arm-AMF",
        "Zem-AMF": "Zem-AMF",
        "ASH": "ASH",
        "Arm": "Arm",
        "ZEM": "ZEM",
        "LiFeng": "LiFeng",
        "LLC": "LLC",
        "FWT": "FWT",
        "е§Іж£ЃжЮЧ": "е§Іж£ЃжЮЧ",
        "Sunzong": "Sunzong",
        "pengfeng": "pengfeng",
        "fortune": "fortune",
        "Allways": "Allways",
    }
    shipment_type_options = {
        "FTL": "FTL",
        "е§ЦйЕН": "е§ЦйЕН",
        # "LTL": "LTL",     
        # "ењЂйАТ": "ењЂйАТ",
        # "еЃҐжИЈиЗ™жПР": "еЃҐжИЈиЗ™жПР",
    }
    RE_PUBLIC_WH = re.compile(
        r"^[A-Z]{4}$|"               # 4дљНзЇѓе≠ЧжѓН
        r"^[A-Z]{3}\d{1,2}$|"         # 3е≠ЧжѓН + 1жИЦ2дљНжХ∞е≠Ч (жЬАеЄЄиІБ)
        r"^[A-Z]{3}\d[A-Z]$|"         # 3е≠ЧжѓН + 1жХ∞е≠Ч + 1е≠ЧжѓН
        r"^[A-Z]{2}\d{2}$"            # 2е≠ЧжѓН + 2жХ∞е≠Ч
    )
    PUBLIC_KEYWORDS = {"WALMART", "ж≤Ге∞ФзОЫ", "AMAZON", "дЇЪй©ђйАК"}
    
    async def get(self, request: HttpRequest) -> HttpResponse:
        if not await self._user_authenticate(request):
            return redirect("login")
        step = request.GET.get("step")
        if step == "appointment_management":
            template, context = await self.handle_appointment_management_get(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "schedule_shipment":
            template, context = await self.handle_td_shipment_get(request)
            return render(request, template, context)
        elif step == "schedule_unshipment":
            template, context = await self.handle_td_unshipment_get(request)
            return render(request, template, context)
        elif step == "fleet_management":
            template, context = await self.handle_fleet_management_get(request)
            return render(request, template, context)
        elif step == "fleet_leader_check":
            if not await self._validate_user_check_po_group(request.user):
                return HttpResponseForbidden(
                    "You are not authenticated to access this page!"
                )
            template, context = await self.handle_fleet_leader_check_get(request)
            return render(request, template, context)       
        elif step == "pod_reupload_check":
            if not await self._validate_user_check_po_group(request.user):
                return HttpResponseForbidden(
                    "You are not authenticated to access this page!"
                )
            return render(request, self.template_pod_reupload, {})       
        elif step == "fleet_po_check":
            if not await self._validate_user_check_po_group(request.user):
                return HttpResponseForbidden(
                    "You are not authenticated to access this page!"
                )
            return render(request, self.template_fleet_po_check, {}) 
        elif step == "master_shipment_check":
            return render(request, self.template_master_shipment_check, {})
        elif step == "client_exception":
            return render(request, self.template_client_exception, {})
        elif step == "unscheduled_pos_all":
            template, context = await self.handle_unscheduled_pos_all_get(request)
            return render(request, template, context)
        elif step == "LTL_pallets":
            context = {
                "warehouse_options": self.warehouse_options
            }
            return render(request, self.template_ltl_pos_all, context)
        elif step == "LTL_history_po":
            context = {
                "warehouse_options": self.warehouse_options
            }
            return render(request, self.template_ltl_history_pos, context)
        elif step == "history_shipment":
            context = {"warehouse_options": self.warehouse_options}
            return render(request, self.template_history_shipment, context)
        elif step == "batch_shipment":
            context = {"warehouse_options": self.warehouse_options}
            return render(request, self.template_batch_shipment, context)
        elif step == "download_batch_shipment_template":
            return await self.handle_download_batch_shipment_template(request)
        
        else:
            raise ValueError('иЊУеЕ•йФЩиѓѓ')

    async def post(self, request: HttpRequest) -> HttpResponse:
        if not await self._user_authenticate(request):
            return redirect("login")
        
        # ж£АжЯ•жШѓеР¶жШѓJSONиѓЈж±В
        content_type = request.content_type or ''
        if content_type.startswith('application/json'):
            try:
                import json
                data = json.loads(request.body.decode('utf-8'))
                step = data.get('step')
            except (json.JSONDecodeError, UnicodeDecodeError):
                step = None
        else:
            step = request.POST.get("step")
        
        print('step', step)
        
        # е§ДзРЖжЯЬеПЈдїУзВєжЯ•иѓҐ
        if step == "search_by_container_and_destination":
            return await self.handle_search_by_container_and_destination(request)
        

        if step == "appointment_management_warehouse":
            template, context = await self.handle_appointment_management_post(request)
            return render(request, template, context)
        elif step == "unscheduled_pos_warehouse":
            template, context = await self.handle_unscheduled_pos_post(request)
            return render(request, template, context)
        elif step == "ltl_warehouse":
            template, context = await self.handle_ltl_unscheduled_pos_post(request)
            return render(request, template, context)
        elif step == "ltl_history_warehouse":
            template, context = await self.handle_ltl_history_pos_post(request)
            return render(request, template, context)
        elif step == "td_shipment_warehouse":
            template, context = await self.handle_td_shipment_post(request)
            return render(request, template, context)
        elif step == "td_unshipment_warehouse":
            template, context = await self.handle_td_unshipment_post(request)
            return render(request, template, context)
        elif step == "fleet_schedule_warehouse":
            template, context = await self.handle_fleet_schedule_post(request)
            return render(request, template, context)
        elif step == "history_shipment_warehouse":
            template, context = await self.handle_history_shipment_post(request)
            return render(request, template, context)
        elif step == "export_pos":
            return await self.handle_export_pos(request)
        elif step == "shipment_export_po":
            return await self.handle_shipment_export_pos(request)    
        elif step =="fleet_export_po":
            return await self.handle_fleet_export_pos(request) 
        elif step == "appointment_time_modify":
            template, context = await self.handle_appointment_time(request)
            return render(request, template, context)
        elif step == "update_fleet":
            fm = FleetManagement()
            context = await fm.handle_update_fleet_post(request,'post_nsop')
            page = request.POST.get("page")
            if page == "arm_appointment":
                template, context = await self.handle_unscheduled_pos_post(request)
            else:
                template, context = await self.handle_td_shipment_post(request)
            context.update({"success_messages": "жЫіжЦ∞еЗЇеЇУиљ¶жђ°жИРеКЯ!"}) 
            return render(request, template, context)
        elif step == "fleet_confirmation":
            template, context = await self.handle_fleet_confirmation_post(request)
            return render(request, template, context) 
        elif step == "cancel_maersk_shipment":
            return await self.handle_cancel_maersk_shipment(request)
        elif step == "ltl_cancel_shipment":
            template, context = await self.handle_ltl_cancel_shipment_post(request)
            return render(request, template, context)
        elif step == "cancel_fleet":
            fm = FleetManagement()
            context = await fm.handle_cancel_fleet_post(request,'post_nsop')
            
            page = request.POST.get("page")
            if page == "arm_appointment":
                template, context = await self.handle_unscheduled_pos_post(request,context)
            elif page == "ltl_readyShip":
                template, context = await self.handle_ltl_unscheduled_pos_post(request)
            else:
                template, context = await self.handle_td_shipment_post(request)
            context.update({"success_messages": 'еПЦжґИжЙєжђ°жИРеКЯ!'})  
            return render(request, template, context)
        elif step == "confirm_delivery":
            fm = FleetManagement()
            context = await fm.handle_confirm_delivery_post(request,'post_nsop')
            page = request.POST.get("page")
            if page == "arm_appointment":
                template, context = await self.handle_unscheduled_pos_post(request)
            elif page == "ltl_delivery_section":
                template, context = await self.handle_ltl_unscheduled_pos_post(request)
            else:
                template, context = await self.handle_fleet_schedule_post(request)
            context.update({"success_messages": 'з°ЃиЃ§йАБиЊЊжИРеКЯ!'})  
            return render(request, template, context)
        elif step == "abnormal_fleet":
            fm = FleetManagement()
            page = request.POST.get("page")
            context = await fm.handle_abnormal_fleet_post(request,'post_nsop')
            
            if page == "arm_appointment":
                template, context = await self.handle_unscheduled_pos_post(request)
            elif page == "ltl_pos_all":
                # жККзЇ¶зЫіжО•еПЦжґИ
                sm = ShippingManagement()
                info = await sm.handle_cancel_abnormal_appointment_post(request,'post_nsop')     
                template, context = await self.handle_ltl_unscheduled_pos_post(request)
            else:
                template, context = await self.handle_fleet_schedule_post(request)
            context.update({"success_messages": 'еЉВеЄЄе§ДзРЖжИРеКЯ!'})  
            return render(request, template, context)
        elif step == "set_shipping_no_link":
            shipment_batch_number = request.POST.get("shipment_batch_number")
            if shipment_batch_number:
                shipment = await sync_to_async(Shipment.objects.get)(
                    shipment_batch_number=shipment_batch_number
                )
                shipment.shipping_order_link = "No Link"
                await sync_to_async(shipment.save)()
            template, context = await self.handle_ltl_unscheduled_pos_post(request)
            context.update({"success_messages": 'еЈ≤иЃЊзљЃдЄЇдЄНеЫЮдЉ†!'})           
            return render(request, template, context)
        elif step == "shipping_order_upload" or step == "batch_shipping_order_upload":
            template, context = await self.handle_shipping_order_upload(request)
            return render(request, template, context)
        elif step == "batch_pod_upload":
            fm = FleetManagement()
            context = await fm.handle_pod_upload_post(request,'post_nsop')
            template, context = await self.handle_fleet_schedule_post(request)
            context.update({"success_messages": 'жЙєйЗПPODдЄКдЉ†жИРеКЯ!'})
            return render(request, template, context)
        elif step == "pod_upload":
            fm = FleetManagement()
            context = await fm.handle_pod_upload_post(request,'post_nsop')
            page = request.POST.get("page")
            if page == "arm_appointment":
                template, context = await self.handle_unscheduled_pos_post(request)
            elif page == "ltl_pod_section":
                template, context = await self.handle_ltl_unscheduled_pos_post(request)
            elif page == "master_shipment_check":
                template, context = await self.handle_master_shipment_check_post(request)
            else:
                template, context = await self.handle_fleet_schedule_post(request)

            context.update({"success_messages": 'PODдЄКдЉ†жИРеКЯ!'})           
            return render(request, template, context)
        elif step == "get_maersk_quote":
            return await self.handle_get_maersk_quote(request)
        elif step == "get_maersk_tracking":
            return await self.handle_get_maersk_tracking(request)
        elif step == "check_business_residential":
            return await self.handle_check_business_residential(request)
        elif step == 'maersk_schedule_post':
            return await self.handle_maersk_schedule_post(request)
        elif step == "bind_group_shipment":
            template, context = await self.handle_appointment_post(request)
            return render(request, template, context) 
        elif step =="ltl_bind_group_shipment":
            template, context = await self.handle_ltl_bind_group_shipment(request)
            return render(request, template, context) 
        elif step == "unassign_shipment":
            template, context = await self.handle_cancel_appointment_post(request)
            return render(request, template, context) 
        elif step == "one_fleet_departure":
            template, context = await self.handle_one_fleet_departure_post(request)
            return render(request, template, context)
        elif step == "add_pallet":
            template, context = await self.handle_add_pallet_post(request)
            return render(request, template, context)      
        elif step == "fleet_add_pallet":
            template, context = await self.handle_fleet_add_pallet_post(request)
            return render(request, template, context)
        elif step == "sp_notified_customer":
            template, context = await self.handle_sp_notified_customer_post(request)
            return render(request, template, context)       
        elif step == "fl_notified_customer":
            template, context = await self.handle_fl_notified_customer_post(request)
            return render(request, template, context)
        elif step == "search_addable_po":
            template, context = await self.handle_search_addable_po_post(request)
            return render(request, template, context)
        elif step == "shipment_add_pallet":
            template, context = await self.handle_shipment_add_pallet_post(request)
            return render(request, template, context)
        elif step == "modify_intelligent_po":
            template, context = await self.handle_modify_intelligent_po_post(request)
            return render(request, template, context)
        elif step == "upload_check_po":
            po_cl = PO()
            request.POST = request.POST.copy()
            request.POST['time_code'] = 'eta' 
            info = await po_cl.handle_upload_check_po_post(request,'post_nsop')
            context = {'success_messages':'ж†°й™МзїУжЮЬдЄКдЉ†жИРеКЯпЉБ'}
            page = request.POST.get("page")
            if page == "arm_appointment":
                template, context = await self.handle_unscheduled_pos_post(request,context)
            else:
                template, context = await self.handle_appointment_management_post(request,context)
            return render(request, template, context)    
        elif step == "create_empty_appointment":
            sm = ShippingManagement()
            info = await sm.handle_create_empty_appointment_post(request,'post_nsop') 
            context = {'success_messages':'е§ЗзЇ¶зЩїиЃ∞жИРеКЯпЉБ'}
            page = request.POST.get("page")
            if page == "arm_appointment":
                template, context = await self.handle_unscheduled_pos_post(request,context)
            else:
                template, context = await self.handle_appointment_management_post(request,context)
            return render(request, template, context)   
        elif step == "download_empty_appointment_template":
            sm = ShippingManagement()
            return await sm.handle_download_empty_appointment_template_post()  
        elif step == "upload_and_create_empty_appointment":
            sm = ShippingManagement()
            info = await sm.handle_upload_and_create_empty_appointment_post(request,'post_nsop') 
            context = {'success_messages':'е§ЗзЇ¶жЙєйЗПзЩїиЃ∞жИРеКЯпЉБ'}
            page = request.POST.get("page")
            if page == "arm_appointment":
                template, context = await self.handle_unscheduled_pos_post(request,context)
            else:
                template, context = await self.handle_appointment_management_post(request,context)
            return render(request, template, context)   
        elif step == "edit_appointment":
            template, context = await self.handle_edit_appointment_post(request)
            return render(request, template, context) 
        elif step == "save_external_shipment":
            template, context = await self.handle_save_external_shipment_post(request)
            return render(request, template, context) 
        elif step == "edit_note_sp":
            template, context = await self.handle_edit_note_sp_post(request)
            return render(request, template, context) 
        elif step == "export_virtual_fleet_pos":
            return await self.handle_export_virtual_fleet_pos_post(request)
        elif step == "multi_group_booking":
            template, context = await self.handle_multi_group_booking(request)
            return render(request, template, context) 
        elif step == "update_fleet_info":
            template, context = await self.handle_update_fleet_info(request)
            return render(request, template, context)      
        elif step == "fix_shipment_exceptions":
            solution = request.POST.get("solution")
            if solution != "keep_old":
                template, context = await self.handle_edit_appointment_post(request,"fleet_departure")
                return render(request, template, context) 
            sm = ShippingManagement()
            info = await sm.handle_fix_shipment_exceptions_post(request,'post_nsop')  
            shipment_batch_number = request.POST.get("shipment_batch_number")
            
            if solution == "keep_old":
                context = {"success_messages": f"{shipment_batch_number}еЈ≤жФєдЄЇж≠£еЄЄзКґжАБпЉБ"}
            else:
                context = {"success_messages": f"{shipment_batch_number}зЇ¶еЈ≤дњЃжФєпЉМеПѓдї•ж≠£еЄЄдљњзФ®пЉБ"}
            template, context = await self.handle_fleet_schedule_post(request,context)
            return render(request, template, context) 
        elif step == "cancel_abnormal_appointment":
            shipment_batch_number = request.POST.get("batch_number")
            sm = ShippingManagement()
            info = await sm.handle_cancel_abnormal_appointment_post(request,'post_nsop')     
            context = {"success_messages": f"{shipment_batch_number}зЇ¶еЈ≤еПЦжґИдЄНеПѓзФ®пЉМжЙАжЬЙpoеЈ≤иІ£зїСпЉБ"}
            template, context = await self.handle_fleet_schedule_post(request,context)
            return render(request, template, context) 
        elif step == "export_bol":
            return await self.handle_bol_post(request)
        elif step =="export_bol_fleet":
            return await self.handle_bol_fleet_post(request)
        elif step == "verify_ltl_cargo":
            template, context = await self.handle_verify_ltl_cargo(request)
            return render(request, template, context)  
        elif step == "export_ltl_unscheduled":
            return await self.export_ltl_unscheduled(request)
        elif step == "save_fleet_cost":
            template, context = await self.handle_save_fleet_cost(request)
            return render(request, template, context)  
        elif step =="save_selfpick_cargo":
            template, context = await self.handle_save_selfpick_cargo(request)
            return render(request, template, context) 
        elif step == "save_selfdel_cargo":
            template, context = await self.handle_save_selfdel_cargo(request)
            return render(request, template, context) 
        elif step == "export_ltl_label":
            return await self.export_ltl_label(request)
        elif step == "export_ltl_bol":
            return await self.export_ltl_bol(request)
        elif step == "upload_self_pickup_file":
            return await self.handle_bol_upload_post(request)
        elif step == "save_shipping_tracking":
            template, context = await self.handle_save_shipping_tracking(request)
            return render(request, template, context) 
        elif step == "update_pod_status":
            template, context = await self.handle_update_pod_status(request)
            return render(request, template, context) 
        elif step == "save_releaseCommand":
            template, context = await self.handle_save_releaseCommand(request)
            return render(request, template, context) 
        elif step == "shipment_note_save":
            template, context = await self.handle_save_shipment_note(request)
            return render(request, template, context)
        elif step == "query_quotation":
            result = await self.handle_query_quotation(request)
            if isinstance(result, HttpResponse):
                return result
            template, context = result
            return render(request, template, context) 
        elif step == "export_maersk_label":
            return await self.handle_export_maersk_label(request)
        elif step == "export_maersk_bol":
            return await self.handle_export_maersk_bol(request)
        elif step == "po_invalid_save":
            template, context = await self.handle_po_invalid_save(request)
            return render(request, template, context) 
        elif step == "batch_update_delivery_method":
            template, context = await self.handle_batch_update_delivery_method(request)
            return render(request, template, context)
        elif step == "batch_one_pick_multi_drop":
            template, context = await self.handle_batch_one_pick_multi_drop(request)
            return render(request, template, context)
        elif step == "batch_shipment_upload":
            template, context = await self.handle_batch_shipment_upload(request)
            return render(request, template, context)
        elif step == "batch_shipment_confirm":
            template, context = await self.handle_batch_shipment_confirm(request)
            return render(request, template, context)
        elif step == "fleet_leader_check_filter":
            template, context = await self.handle_fleet_leader_check_get(request)
            return render(request, template, context)
        elif step == "fleet_leader_check":
            template, context = await self.handle_fleet_leader_check_post(request)
            return render(request, template, context)
        elif step == "fleet_po_search":
            template, context = await self.handle_fleet_po_search_post(request)
            return render(request, template, context)
        elif step == "master_shipment_search":
            template, context = await self.handle_master_shipment_check_post(request)
            return render(request, template, context)
        elif step == "client_exception_search":
            template, context = await self.handle_client_exception_search_post(request)
            return render(request, template, context)
        elif step == "load_exceptions":
            return await self.handle_load_exceptions_post(request)
        elif step == "add_exception":
            return await self.handle_add_exception_post(request)
        elif step == "delete_exception":
            return await self.handle_delete_exception_post(request)
        elif step == "create_fictional_master":
            template, context = await self.handle_create_fictional_master_post(request)
            return render(request, template, context)
        elif step == "bind_existing_shipment":
            template, context = await self.handle_bind_existing_shipment_post(request)
            return render(request, template, context)
        elif step == "save_virtual_shipment_time":
            template, context = await self.handle_save_virtual_shipment_time_post(request)
            return render(request, template, context)
        elif step == "unbind_master_shipment":
            template, context = await self.handle_unbind_master_shipment_post(request)
            return render(request, template, context)
        elif step == "pod_reupload_search":
            template, context = await self.handle_pod_reupload_post(request)
            return render(request, template, context)
        elif step == "pod_reupload":
            template, context = await self.handle_pod_reupload_upload_post(request)
            return render(request, template, context)
        elif step == "fleet_po_delete":
            template, context = await self.handle_fleet_po_delete_post(request)
            return render(request, template, context)
        elif step == "add_po_query_plt":
            template, context = await self.handle_add_po_query_plt(request)
            return render(request, template, context)
        elif step == "add_pallets_to_shipment":
            template, context = await self.handle_add_pallets_to_shipment(request)
            return render(request, template, context)
        else:
            raise ValueError('иЊУеЕ•йФЩиѓѓ',step)
    
    async def _validate_user_check_po_group(self, user: User) -> bool:
        is_staff = await sync_to_async(lambda: user.is_staff)()
        if is_staff:
            return True
        return await sync_to_async(
            lambda: user.groups.filter(name="shipment_po_check").exists()
        )()
    
    async def handle_batch_update_delivery_method(self, request: HttpRequest):
        '''жЙєйЗПдњЭе≠ШжіЊйАБжЦєеЉП'''
        context = {}
        batch_data_json = request.POST.get('batch_methods')
        try:
            batch_data = json.loads(batch_data_json)

            for entry in batch_data:
                raw_id = entry.get('cargo_id', '')
                new_method = entry.get('delivery_method', '')

                if not raw_id:
                    continue

                if raw_id.startswith('plt_'):
                    # --- е§ДзРЖжЙУжЭњжХ∞жНЃ (Pallet) ---
                    # зІїйЩ§еЙНзЉАеєґеИЖеЙ≤жИР ID еИЧи°®
                    pallet_ids = raw_id.replace('plt_', '').split(',')
                    # жЙєйЗПжЫіжЦ∞еѓєеЇФзЪДжЙШзЫШ
                    await Pallet.objects.filter(id__in=pallet_ids).aupdate(delivery_method=new_method)
                else:
                    # --- е§ДзРЖжЬ™жЙУжЭњжХ∞жНЃ (PackingList) ---
                    # йАЪеЄЄжШѓжЩЃйАЪиіІзЙ©зЪД ID еИЧи°®
                    cargo_ids = raw_id.split(',')
                    await PackingList.objects.filter(id__in=cargo_ids).aupdate(delivery_method=new_method)
            context.update({'success_messages':f"жИРеКЯжЙєйЗПдњЃжФє {len(batch_data)} й°єжіЊйАБжЦєеЉП"})
        
        except Exception as e:
            context.update({'error_messages':f"дњЃжФє {len(batch_data)} й°єжіЊйАБжЦєеЉПе§±иі•пЉМеОЯеЫ†жШѓ{e}"})

        # йЗНеЃЪеРСеЫЮеОЯй°µйЭҐеєґдњЭжМБељУеЙНзЪДдїУеЇУеТМж†Зз≠Њй°µ
        return await self.handle_ltl_unscheduled_pos_post(request, context)

    async def handle_batch_one_pick_multi_drop(self, request: HttpRequest):
        '''LTL дЄАжПРе§ЪеНЄж†ЗиЃ∞'''
        context = {}
        batch_json = request.POST.get('batch_correlation_data', '[]')
        try:
            correlation_list = json.loads(batch_json)
        except (json.JSONDecodeError, TypeError):
            correlation_list = []

        if correlation_list:
            # 1. зФЯжИР Seed: еПЦжЙАжЬЙйАЙдЄ≠й°єеФЫе§ізЪДеЙН5дљНеТМеРО5дљНжЛЉжО•
            # дЊЛе¶В: ZSDMI5120800011 -> ZSDMI0011
            seed_parts = []
            for item in correlation_list:
                mark = str(item.get('shipping_mark', '')).strip()
                if len(mark) > 10:
                    seed_parts.append(f"{mark[:5]}{mark[-5:]}")
                else:
                    seed_parts.append(mark)
            
            seed_str = "|".join(seed_parts)
            
            # 2. йАЪињЗ Seed зФЯжИРеФѓдЄАзЪД UUID (uuid5 еЯЇдЇОеСљеРНз©ЇйЧіеТМе≠Чзђ¶дЄ≤зФЯжИР)
            # ињЩж†Је¶ВжЮЬзФ®жИЈеЬ®дЄНеРМжЧґйЧізВєйАЙдЇЖеРМдЄАзїДеФЫе§іпЉМзФЯжИРзЪДеЕ≥иБФIDзРЖиЃЇдЄКжШѓдЄАиЗізЪД
            correlation_id = str(uuid.uuid5(uuid.NAMESPACE_DNS, seed_str))

            # 3. еЉВж≠•еЊ™зОѓжЫіжЦ∞
            for entry in correlation_list:
                raw_id = entry.get('cargo_id', '')
                if not raw_id:
                    continue

                if raw_id.startswith('plt_'):
                    # --- е§ДзРЖжЙУжЭњжХ∞жНЃ (Pallet) ---
                    pallet_ids = raw_id.replace('plt_', '').split(',')

                    existing_records = await sync_to_async(list)(
                        Pallet.objects.filter(id__in=pallet_ids).values('id', 'ltl_correlation_id')
                    )
                    if not existing_records:
                        continue
                        
                    # ж£АжЯ•жШѓеР¶йЬАи¶БеПЦжґИеЕ≥иБФпЉИе¶ВжЮЬжЙАжЬЙйАЙдЄ≠зЪДиЃ∞ељХйГљжЬЙltl_correlation_idеАЉпЉЙ
                    has_correlation = any(record['ltl_correlation_id'] for record in existing_records)
                    
                    # еП™еПЦжґИеЈ≤жЬЙеЕ≥иБФзЪДиЃ∞ељХ
                    if has_correlation:
                        # жЙЊеЗЇжЬЙеАЉзЪДиЃ∞ељХID
                        await Pallet.objects.filter(id__in=pallet_ids).aupdate(ltl_correlation_id='')
                        context.update({'success_messages': f"жИРеКЯеПЦжґИдЄАжПРе§ЪеНЄеЕ≥иБФ"})
                    else:
                        # дљњзФ®еЉВж≠• aupdate жЙєйЗПжЫіжЦ∞
                        await Pallet.objects.filter(id__in=pallet_ids).aupdate(ltl_correlation_id=correlation_id)
                        context.update({'success_messages':f"жИРеКЯеЕ≥иБФдЄАжПРе§ЪеНЄ"})
                else:
                    # --- е§ДзРЖжЬ™жЙУжЭњжХ∞жНЃ (PackingList) ---
                    cargo_ids = raw_id.split(',')

                    existing_records = await sync_to_async(list)(
                        PackingList.objects.filter(id__in=cargo_ids).values('id', 'ltl_correlation_id')
                    )
                    if not existing_records:
                        continue
                    has_correlation = any(record['ltl_correlation_id'] for record in existing_records)
                    
                    # е¶ВжЮЬеОЯжЬђжЬЙеАЉпЉМеИЩиЃЊзљЃдЄЇз©ЇпЉИеПЦжґИжУНдљЬпЉЙ
                    if has_correlation:
                        await PackingList.objects.filter(id__in=cargo_ids).aupdate(ltl_correlation_id='')
                        context.update({'success_messages': f"жИРеКЯеПЦжґИдЄАжПРе§ЪеНЄеЕ≥иБФ"})
                    else:
                        # дљњзФ®еЉВж≠• aupdate жЙєйЗПжЫіжЦ∞
                        await PackingList.objects.filter(id__in=cargo_ids).aupdate(ltl_correlation_id=correlation_id)
                        context.update({'success_messages':f"жИРеКЯеЕ≥иБФдЄАжПРе§ЪеНЄ"})
        return await self.handle_ltl_unscheduled_pos_post(request, context)
    
    async def handle_batch_shipment_upload(self, request: HttpRequest):
        '''е§ДзРЖжЙєйЗПйҐДзЇ¶еЗЇеЇУжЦЗдїґдЄКдЉ†'''
        context = {}
        
        if 1:
            # ж£АжЯ•жШѓеР¶жЬЙжЦЗдїґдЄКдЉ†
            if 'excel_file' not in request.FILES:
                context['error'] = 'иѓЈйАЙжЛ©и¶БдЄКдЉ†зЪДExcelжЦЗдїґ'
                return self.template_batch_shipment, context
            
            # иОЈеПЦдЄКдЉ†зЪДжЦЗдїґ
            excel_file = request.FILES['excel_file']
            
            # дљњзФ®pandasиѓїеПЦExcelжЦЗдїґ
            df = pd.read_excel(excel_file)
            
            # ж£АжЯ•и°®е§іжШѓеР¶ж≠£з°Ѓ
            required_columns = ['жЯЬеПЈ', 'дїУзВє', 'CBM', 'еН°жЭњ', 'йҐДзЇ¶жЧґйЧі', 'ISA', 'pickup time', 'PickUp number', 'Shipment ID', 'йҐДзЇ¶иі¶еПЈ', 'йҐДзЇ¶з±їеЮЛ', 'и£Еиљ¶з±їеЮЛ']
            for col in required_columns:
                if col not in df.columns:
                    context['error'] = f'жЦЗдїґзЉЇе∞СењЕи¶БзЪДеИЧ: {col}'
                    return self.template_batch_shipment, context
            
            # е§ДзРЖжХ∞жНЃпЉМжМЙиљ¶зїДеТМз©Їи°МеИЖзїД
            groups = []
            current_group = {
                'containers': [],
                'appointment_time': None,
                'isa': None,
                'pickup_time': None,
                'pickup_number': None,
                'shipment_id': None,
                'appointment_account': None,
                'appointment_type': None,
                'loading_type': None
            }
            current_car_group = 1
            is_same_car = False
            has_processed_first_group = False
            
            for index, row in df.iterrows():
                # ж£АжЯ•жШѓеР¶дЄЇз©Їи°М
                if row.isnull().all():
                    # з©Їи°МпЉМзїУжЭЯељУеЙНзїДеєґеЉАеІЛжЦ∞зїД
                    if current_group['containers'] or any([
                        current_group['appointment_time'],
                        current_group['isa'],
                        current_group['appointment_account'],
                        current_group['appointment_type'],
                        current_group['loading_type']
                    ]):
                        current_group['car_group'] = current_car_group
                        groups.append(current_group)
                        current_group = {
                            'containers': [],
                            'appointment_time': None,
                            'isa': None,
                            'pickup_time': None,
                            'pickup_number': None,
                            'shipment_id': None,
                            'appointment_account': None,
                            'appointment_type': None,
                            'loading_type': None
                        }
                        # з©Їи°Ми°®з§Їиљ¶зїДзїУжЭЯ
                        current_car_group += 1
                        is_same_car = False
                        has_processed_first_group = True
                else:
                    # ж£АжЯ•жШѓеР¶жШѓдЄАжПРxеНЄи°М
                    is_multi_drop = False
                    for col in row:
                        if isinstance(col, str) and any(phrase in col for phrase in ['дЄАжПРдЄ§еНЄ', 'дЄАжПРдЄЙеНЄ', 'дЄАжПРе§ЪеНЄ']):
                            is_multi_drop = True
                            break
                    
                    if is_multi_drop:
                        # дЄАжПРxеНЄи°®з§ЇељУеЙНзЇ¶зїУжЭЯпЉМдљЖиљ¶зїДдЄНзїУжЭЯ
                        if current_group['containers'] or any([
                            current_group['appointment_time'],
                            current_group['isa'],
                            current_group['appointment_account'],
                            current_group['appointment_type'],
                            current_group['loading_type']
                        ]):
                            current_group['car_group'] = current_car_group
                            groups.append(current_group)
                            current_group = {
                                'containers': [],
                                'appointment_time': None,
                                'isa': None,
                                'pickup_time': None,
                                'pickup_number': None,
                                'shipment_id': None,
                                'appointment_account': None,
                                'appointment_type': None,
                                'loading_type': None
                            }
                            is_same_car = True
                            has_processed_first_group = True
                    else:
                        # йЭЮз©Їи°МпЉМе§ДзРЖжХ∞жНЃ
                        container_number = row.get('жЯЬеПЈ')
                        warehouse = row.get('дїУзВє')
                        cbm = row.get('CBM')
                        pallet = row.get('еН°жЭњ')
                        
                        # ж£АжЯ•жШѓеР¶жШѓйҐДзЇ¶дњ°жБѓи°МпЉИеМЕеРЂйҐДзЇ¶жЧґйЧіз≠ЙпЉЙ
                        has_appointment_info = pd.notna(row.get('йҐДзЇ¶жЧґйЧі')) or pd.notna(row.get('ISA'))
                        
                        if has_appointment_info and not is_same_car:
                            # е¶ВжЮЬељУеЙНжЬЙеЃєеЩ®жХ∞жНЃпЉМдњЭе≠ШеИ∞дЄКдЄАдЄ™еИЖзїД
                            if current_group['containers'] or any([
                                current_group['appointment_time'],
                                current_group['isa'],
                                current_group['appointment_account'],
                                current_group['appointment_type'],
                                current_group['loading_type']
                            ]):
                                current_group['car_group'] = current_car_group
                                groups.append(current_group)
                                current_group = {
                                    'containers': [],
                                    'appointment_time': None,
                                    'isa': None,
                                    'pickup_time': None,
                                    'pickup_number': None,
                                    'shipment_id': None,
                                    'appointment_account': None,
                                    'appointment_type': None,
                                    'loading_type': None
                                }
                                # еЉАеІЛжЦ∞зЪДиљ¶зїД
                                if has_processed_first_group:
                                    current_car_group += 1
                        
                        # ж£АжЯ•ењЕе°Ђй°є
                        if container_number and warehouse and cbm and pallet:
                            # ињЩжШѓжЯЬеПЈдњ°жБѓи°М
                            # жПРеПЦеН°жЭњзЪДжХіжХ∞еАЉзФ®дЇОжѓФиЊГ
                            pallet_int = None
                            try:
                                if isinstance(pallet, (int, float)):
                                    pallet_int = int(pallet)
                                else:
                                    # е∞ЭиѓХдїОе≠Чзђ¶дЄ≤дЄ≠жПРеПЦжХ∞е≠Ч
                                    import re
                                    match = re.search(r'\d+', str(pallet))
                                    if match:
                                        pallet_int = int(match.group())
                            except (ValueError, TypeError):
                                pass
                            
                            # ж£АжЯ•жШѓеР¶еЈ≤е≠ШеЬ®зЫЄеРМзЪДжЯЬеПЈеТМдїУзВєзїДеРИ
                            is_duplicate = False
                            for existing_container in current_group['containers']:
                                if existing_container['container_number'] == container_number and existing_container['warehouse'] == warehouse:
                                    is_duplicate = True
                                    break
                            if not is_duplicate:
                                current_group['containers'].append({
                                    'container_number': container_number,
                                    'warehouse': warehouse,
                                    'cbm': cbm,
                                    'pallet': pallet,
                                    'pallet_int': pallet_int
                                })
                        
                        # жПРеПЦйҐДзЇ¶дњ°жБѓпЉИеП™жПРеПЦйЭЮз©ЇеАЉпЉЙ
                        if pd.notna(row.get('йҐДзЇ¶жЧґйЧі')):
                            appointment_time = row.get('йҐДзЇ¶жЧґйЧі')
                            # иљђжНҐжЧґйЧіж†ЉеЉП
                            try:
                                if isinstance(appointment_time, str):
                                    # е§ДзРЖе≠Чзђ¶дЄ≤ж†ЉеЉПзЪДжЧґйЧі
                                    # е∞ЭиѓХдЄНеРМзЪДжЧґйЧіж†ЉеЉП
                                    for fmt in ['%B %d, %Y, %I:%M %p', '%B %d, %Y, midnight', '%B %d, %Y, noon']:
                                        try:
                                            dt = datetime.strptime(appointment_time, fmt)
                                            # иљђжНҐдЄЇеРОзЂѓжЬЯжЬЫзЪДж†ЉеЉП
                                            current_group['appointment_time'] = dt.strftime('%Y-%m-%d %H:%M')
                                            break
                                        except ValueError:
                                            continue
                                    else:
                                        # е¶ВжЮЬжЙАжЬЙж†ЉеЉПйГље§±иі•пЉМдњЭжМБеОЯеАЉ
                                        current_group['appointment_time'] = appointment_time
                                else:
                                    # е§ДзРЖdatetimeеѓєи±°
                                    current_group['appointment_time'] = appointment_time.strftime('%Y-%m-%d %H:%M')
                            except Exception:
                                # е¶ВжЮЬиљђжНҐе§±иі•пЉМдњЭжМБеОЯеАЉ
                                current_group['appointment_time'] = appointment_time
                        if pd.notna(row.get('ISA')):
                            current_group['isa'] = row.get('ISA')
                        if pd.notna(row.get('pickup time')):
                            pickup_time = row.get('pickup time')
                            # иљђжНҐжЧґйЧіж†ЉеЉП
                            try:
                                if isinstance(pickup_time, str):
                                    # е§ДзРЖе≠Чзђ¶дЄ≤ж†ЉеЉПзЪДжЧґйЧі
                                    # е∞ЭиѓХдЄНеРМзЪДжЧґйЧіж†ЉеЉП
                                    for fmt in ['%B %d, %Y, %I:%M %p', '%B %d, %Y, midnight', '%B %d, %Y, noon']:
                                        try:
                                            dt = datetime.strptime(pickup_time, fmt)
                                            # иљђжНҐдЄЇеРОзЂѓжЬЯжЬЫзЪДж†ЉеЉП
                                            current_group['pickup_time'] = dt.strftime('%Y-%m-%d %H:%M')
                                            break
                                        except ValueError:
                                            continue
                                    else:
                                        # е¶ВжЮЬжЙАжЬЙж†ЉеЉПйГље§±иі•пЉМдњЭжМБеОЯеАЉ
                                        current_group['pickup_time'] = pickup_time
                                else:
                                    # е§ДзРЖdatetimeеѓєи±°
                                    current_group['pickup_time'] = pickup_time.strftime('%Y-%m-%d %H:%M')
                            except Exception:
                                # е¶ВжЮЬиљђжНҐе§±иі•пЉМдњЭжМБеОЯеАЉ
                                current_group['pickup_time'] = pickup_time
                        if pd.notna(row.get('PickUp number')):
                            current_group['pickup_number'] = row.get('PickUp number')
                        if pd.notna(row.get('Shipment ID')):
                            try:
                                # е∞ЖShipment IDиљђжНҐдЄЇжХіжХ∞
                                current_group['shipment_id'] = int(row.get('Shipment ID'))
                            except (ValueError, TypeError):
                                # е¶ВжЮЬиљђжНҐе§±иі•пЉМдњЭжМБеОЯеАЉ
                                current_group['shipment_id'] = row.get('Shipment ID')
                        if pd.notna(row.get('йҐДзЇ¶иі¶еПЈ')):
                            current_group['appointment_account'] = row.get('йҐДзЇ¶иі¶еПЈ')
                        if pd.notna(row.get('йҐДзЇ¶з±їеЮЛ')):
                            current_group['appointment_type'] = row.get('йҐДзЇ¶з±їеЮЛ')
                        if pd.notna(row.get('и£Еиљ¶з±їеЮЛ')):
                            current_group['loading_type'] = row.get('и£Еиљ¶з±їеЮЛ')
                        if pd.notna(row.get('е§Зж≥®')):
                            current_group['note'] = row.get('е§Зж≥®')
                        if pd.notna(row.get('еПСиіІдїУеЇУ')):
                            current_group['origin'] = row.get('еПСиіІдїУеЇУ')
                        
                        # йЗНзљЃis_same_carж†ЗењЧ
                        if has_appointment_info:
                            is_same_car = False
                            has_processed_first_group = True
            
            # жЈїеК†жЬАеРОдЄАзїД
            if current_group['containers'] or any([
                current_group['appointment_time'],
                current_group['isa'],
                current_group['appointment_account'],
                current_group['appointment_type'],
                current_group['loading_type']
            ]):
                current_group['car_group'] = current_car_group
                groups.append(current_group)
            
            # дЄЇз©ЇзЪДpickup_numberзФЯжИРеАЉ
            for group in groups:
                # жЈїеК†зЫЃзЪДеЬ∞дњ°жБѓпЉИеПЦзђђдЄАдЄ™еЃєеЩ®зЪДдїУзВєпЉЙ
                if group.get('containers'):
                    group['destination'] = group['containers'][0].get('warehouse', '')
                else:
                    group['destination'] = ''
                
                if not group.get('pickup_number'):
                    # еЯЇз°АеЙНзЉА
                    prefix = 'ZEM-RC-'
                    
                    # иОЈеПЦељУе§©жЬИжЧ•пЉИMMDDж†ЉеЉПпЉЙ
                    today = datetime.now()
                    month = str(today.month).zfill(2)
                    day = str(today.day).zfill(2)
                    month_day = month + day
                    
                    # йҐДзЇ¶иі¶еПЈе§ДзРЖ
                    shipment_account = group.get('appointment_account', '')
                    account_part = ''
                    if 'Central' in shipment_account or 'walmart' in shipment_account:
                        account_part = 'ASH'
                    else:
                        # зФ® - еИЖзїДпЉМеПЦзђђдЄАдЄ™зїД
                        parts = shipment_account.split('-')
                        account_part = parts[0] if parts else shipment_account
                    
                    # зЫЃзЪДеЬ∞е§ДзРЖпЉИеПЦзђђдЄАдЄ™еЃєеЩ®зЪДдїУзВєпЉЙ
                    destination_part = ''
                    if group.get('containers'):
                        destination = group['containers'][0].get('warehouse', '')
                        if '-' in destination:
                            # е¶ВжЮЬеМЕеРЂ -пЉМеПЦ - еРОйЭҐзЪДеЖЕеЃє
                            parts = destination.split('-')
                            destination_part = '-'.join(parts[1:]).replace(' ', '')
                        else:
                            # е¶ВжЮЬдЄНеМЕеРЂ -пЉМдљњзФ®жХідЄ™зЫЃзЪДеЬ∞
                            destination_part = destination.replace(' ', '')
                    
                    # зФЯжИР4дљНйЪПжЬЇжХ∞е≠Ч
                    random_num = str(random.randint(1000, 9999))
                    
                    # зїДеРИжИРеЃМжХізЪД pickupNumber
                    pickup_number = f"{prefix}{month_day}{account_part}-{destination_part}-{random_num}"
                    group['pickup_number'] = pickup_number
            
            # й™МиѓБжѓПзїДжХ∞жНЃ
            valid_groups = []
            for group in groups:
                
                # ж£АжЯ•йҐДзЇ¶дњ°жБѓжШѓеР¶еЃМжХіпЉИpickup_timeеТМpickup_numberеПѓдї•дЄЇз©ЇпЉЙ
                required_appointment_fields = [
                    'appointment_time',
                    'isa',
                    'appointment_account',
                    'appointment_type',
                    'loading_type'
                ]
                
                missing_fields = []
                for field in required_appointment_fields:
                    if not group.get(field):
                        missing_fields.append(field)
                
                # жХ∞жНЃй™МиѓБ
                validation_errors = []
                field_errors = {
                    'isa': False,
                    'appointment_account': False,
                    'appointment_type': False,
                    'loading_type': False,
                    'appointment_time': False,
                    'pickup_time': False,
                    'origin': False
                }
                
                # й™МиѓБISAжШѓеР¶дЄЇжХіжХ∞
                isa = group.get('isa')
                if isa:
                    try:
                        # иљђжНҐдЄЇжХіжХ∞
                        isa_int = int(isa)
                        group['isa'] = isa_int
                        
                        # ж£АжЯ•ISAжШѓеР¶еЈ≤е≠ШеЬ®
                        
                        try:
                            from asgiref.sync import sync_to_async
                            existed_appointment = await sync_to_async(Shipment.objects.get)(
                                appointment_id=isa_int
                            )
                            # ж£АжЯ•жШѓеР¶еЈ≤зЩїиЃ∞
                            if existed_appointment.in_use:
                                validation_errors.append(f'ISA {isa_int} еЈ≤зїПиҐЂдљњзФ®дЇЖ!')
                                field_errors['isa'] = True
                            # ж£АжЯ•жШѓеР¶еЈ≤еПЦжґИ
                            elif existed_appointment.is_canceled:
                                validation_errors.append(f'ISA {isa_int} еЈ≤зїПе≠ШеЬ®еєґдЄФиҐЂеПЦжґИдЇЖ!')
                                field_errors['isa'] = True
                            # ж£АжЯ•жШѓеР¶ињЗжЬЯ
                            elif existed_appointment.shipment_appointment.replace(tzinfo=pytz.UTC) < timezone.now():
                                validation_errors.append(f'ISA {isa_int} йҐДзЇ¶жЧґйЧіжШѓ{existed_appointment.shipment_appointment}е∞ПдЇОељУеЙНжЧґйЧіпЉМеЈ≤ињЗжЬЯ!')
                                field_errors['isa'] = True
                            # ж£АжЯ•зЫЃзЪДеЬ∞жШѓеР¶дЄАиЗі
                            elif group.get('destination'):
                                existing_dest = existed_appointment.destination.replace("Walmart", "").replace("WALMART", "").replace("-", "").upper()
                                current_dest = group.get('destination').replace("Walmart", "").replace("WALMART", "").replace("-", "").upper()
                                if existing_dest != current_dest:
                                    validation_errors.append(f"ISA {isa_int} зЩїиЃ∞зЪДзЫЃзЪДеЬ∞жШѓ {existed_appointment.destination} пЉМж≠§жђ°зЩїиЃ∞зЪДзЫЃзЪДеЬ∞жШѓ {group.get('destination')}!")
                                    field_errors['isa'] = True
                        except Shipment.DoesNotExist:
                            # ISAдЄНе≠ШеЬ®пЉМзїІзї≠е§ДзРЖ
                            pass
                            
                    except (ValueError, TypeError):
                        validation_errors.append('ISAењЕй°їжШѓжХіжХ∞')
                        field_errors['isa'] = True
                
                # й™МиѓБйҐДзЇ¶иі¶еПЈ
                appointment_account = group.get('appointment_account')
                valid_accounts = ['Carrier Central1', 'Carrier Central2', 'ZEM-AMF', 'ARM-AMF', 'walmart']
                if appointment_account and appointment_account not in valid_accounts:
                    validation_errors.append('йҐДзЇ¶иі¶еПЈењЕй°їжШѓCarrier Central1гАБCarrier Central2гАБZEM-AMFгАБARM-AMFгАБwalmartдЄ≠зЪДдЄАдЄ™')
                    field_errors['appointment_account'] = True
                
                # й™МиѓБеПСиіІдїУеЇУ
                origin = group.get('origin')
                valid_origins = ['NJ-07001', 'SAV-31326', 'LA-91761']
                if not origin:
                    validation_errors.append('еПСиіІдїУеЇУдЄНиГљдЄЇз©Ї')
                    field_errors['origin'] = True
                elif origin not in valid_origins:
                    validation_errors.append('еПСиіІдїУеЇУењЕй°їжШѓNJ-07001гАБSAV-31326гАБLA-91761дЄ≠зЪДдЄАдЄ™')
                    field_errors['origin'] = True
                
                # й™МиѓБйҐДзЇ¶з±їеЮЛ
                appointment_type = group.get('appointment_type')
                valid_types = ['FTL', 'LTL', 'е§ЦйЕН', 'ењЂйАТ', 'еЃҐжИЈиЗ™жПР']
                if appointment_type and appointment_type not in valid_types:
                    validation_errors.append('йҐДзЇ¶з±їеЮЛењЕй°їжШѓFTLгАБLTLгАБе§ЦйЕНгАБењЂйАТгАБеЃҐжИЈиЗ™жПРдЄ≠зЪДдЄАдЄ™')
                    field_errors['appointment_type'] = True
                
                # й™МиѓБи£Еиљ¶з±їеЮЛ
                loading_type = group.get('loading_type')
                valid_loading_types = ['еН°жЭњ', 'еЬ∞жЭњ']
                if loading_type and loading_type not in valid_loading_types:
                    validation_errors.append('и£Еиљ¶з±їеЮЛењЕй°їжШѓеН°жЭњжИЦеЬ∞жЭњ')
                    field_errors['loading_type'] = True
                
                # й™МиѓБйҐДзЇ¶жЧґйЧіж†ЉеЉП
                appointment_time = group.get('appointment_time')
                if appointment_time:
                    # е§ДзРЖTimestampз±їеЮЛ
                    if hasattr(appointment_time, 'strftime'):
                        appointment_time = appointment_time.strftime('%Y-%m-%d %H:%M')
                    # иљђжНҐдЄЇе≠Чзђ¶дЄ≤
                    appointment_time_str = str(appointment_time)
                    # жФѓжМБе§ЪзІНжЧ•жЬЯж†ЉеЉП
                    formats = ['%Y-%m-%d %H:%M', '%Y-%m-%d', '%Y/%m/%d %H:%M', '%Y/%m/%d']
                    valid = False
                    for fmt in formats:
                        try:
                            datetime.strptime(appointment_time_str, fmt)
                            valid = True
                            break
                        except ValueError:
                            pass
                    if not valid:
                        validation_errors.append('йҐДзЇ¶жЧґйЧіж†ЉеЉПдЄНж≠£з°ЃпЉМиѓЈдљњзФ®YYYY-MM-DD HH:MMгАБYYYY-MM-DDгАБYYYY/MM/DD HH:MMжИЦYYYY/MM/DDж†ЉеЉП')
                        field_errors['appointment_time'] = True
                
                # й™МиѓБpickup timeж†ЉеЉП
                pickup_time = group.get('pickup_time')
                if pickup_time:
                    # е§ДзРЖTimestampз±їеЮЛ
                    if hasattr(pickup_time, 'strftime'):
                        pickup_time = pickup_time.strftime('%Y-%m-%d %H:%M')
                    # иљђжНҐдЄЇе≠Чзђ¶дЄ≤
                    pickup_time_str = str(pickup_time)
                    # жФѓжМБе§ЪзІНжЧ•жЬЯж†ЉеЉП
                    formats = ['%Y-%m-%d %H:%M', '%Y-%m-%d', '%Y/%m/%d %H:%M', '%Y/%m/%d']
                    valid = False
                    for fmt in formats:
                        try:
                            datetime.strptime(pickup_time_str, fmt)
                            valid = True
                            break
                        except ValueError:
                            pass
                    if not valid:
                        validation_errors.append('pickup timeж†ЉеЉПдЄНж≠£з°ЃпЉМиѓЈдљњзФ®YYYY-MM-DD HH:MMгАБYYYY-MM-DDгАБYYYY/MM/DD HH:MMжИЦYYYY/MM/DDж†ЉеЉП')
                        field_errors['pickup_time'] = True
                
                # й™МиѓБжЯЬе≠Рдњ°жБѓ
                from asgiref.sync import sync_to_async
                
                async def validate_container(container):
                    container_number = container.get('container_number')
                    warehouse = container.get('warehouse')
                    pallet_count = container.get('pallet')
                    pallet_count_int = container.get('pallet_int')
                    
                    # еИЭеІЛеМЦеЃєеЩ®й™МиѓБдњ°жБѓ
                    container['validation'] = {
                        'status': 'ж≠£з°Ѓ',
                        'message': '',
                        'ids': []
                    }
                    
                    # еРМж≠•еЗљжХ∞зФ®дЇОжХ∞жНЃеЇУжЯ•иѓҐ
                    def check_pallet_records():
                        return Pallet.objects.filter(
                            container_number__container_number=container_number,
                            destination=warehouse
                        ).exclude(delivery_method__contains="жЪВжЙ£зХЩдїУ")
                    
                    def check_packinglist_records():
                        return PackingList.objects.filter(
                            container_number__container_number=container_number,
                            destination=warehouse
                        ).exclude(delivery_method__contains="жЪВжЙ£зХЩдїУ")
                    
                    def get_shipment(shipment_id):
                        try:
                            return Shipment.objects.get(id=shipment_id)
                        except Shipment.DoesNotExist:
                            return None
                    
                    # еЕИеОїPalletи°®жЯ•жЙЊ
                    pallet_records = await sync_to_async(check_pallet_records)()
                    
                    # еИЭеІЛеМЦзїЯиЃ°дњ°жБѓ
                    container['stats'] = {
                        'total_weight': 0,
                        'total_pcs': 0,
                        'total_cbm': 0,
                        'total_pallet': 0,
                        'pallet_only': True
                    }
                    
                    # жЈїеК†жШѓеР¶жЙУжЭњж†ЗиЃ∞
                    container['is_pallet'] = False
                    
                    if await sync_to_async(pallet_records.exists)():
                        container['is_pallet'] = True
                        # ж£АжЯ•жХ∞йЗПжШѓеР¶е∞ПдЇОеН°жЭњеАЉ
                        found_count = await sync_to_async(lambda: pallet_records.count())()
                        if pallet_count_int is not None and found_count < pallet_count_int:
                            container['validation']['status'] = 'йФЩиѓѓ'
                            container['validation']['message'] = f'еЃЮйЩЕжЭњжХ∞дЄЇ{found_count}жЭњ'
                        else:
                            # ж£АжЯ•shipment_batch_number_idжШѓеР¶жЬЙеАЉ
                            pallet_list = await sync_to_async(list)(pallet_records)
                            # зїЯиЃ°palletи°®зЪДдњ°жБѓ
                            total_weight = 0
                            total_pcs = 0
                            total_cbm = 0
                            total_pallet = len(pallet_list)
                            
                            for record in pallet_list:
                                total_weight += record.weight_lbs or 0
                                total_pcs += record.pcs or 0
                                total_cbm += record.cbm or 0
                                if record.shipment_batch_number_id:
                                    # жЯ•жЙЊеѓєеЇФзЪДshipment
                                    shipment = await sync_to_async(get_shipment)(record.shipment_batch_number_id)
                                    if shipment:
                                        container['validation']['status'] = 'йФЩиѓѓ'
                                        container['validation']['message'] = f'жЭње≠РеЈ≤жЬЙзЇ¶пЉМзЇ¶жШѓ{shipment.shipment_batch_number}'
                                        break
                            # е¶ВжЮЬж≤°жЬЙйФЩиѓѓпЉМжФґйЫЖid
                            if container['validation']['status'] == 'ж≠£з°Ѓ':
                                container['validation']['ids'] = [f'plt_id{record.id}' for record in pallet_list]
                                # жЫіжЦ∞зїЯиЃ°дњ°жБѓ
                                container['stats'] = {
                                    'total_weight': total_weight,
                                    'total_pcs': total_pcs,
                                    'total_cbm': total_cbm,
                                    'total_pallet': total_pallet,
                                    'pallet_only': True
                                }
                    else:
                        # еОїPackinglistи°®жЯ•жЙЊ
                        packinglist_records = await sync_to_async(check_packinglist_records)()
                        if await sync_to_async(packinglist_records.exists)():
                            # ж£АжЯ•жХ∞йЗПжШѓеР¶е∞ПдЇОеН°жЭњеАЉ
                            found_count = await sync_to_async(lambda: packinglist_records.count())()
                            # ж£АжЯ•shipment_batch_number_idжШѓеР¶жЬЙеАЉ
                            packinglist_list = await sync_to_async(list)(packinglist_records)
                            # зїЯиЃ°packinglistи°®зЪДдњ°жБѓ
                            total_weight = 0
                            total_pcs = 0
                            total_cbm = 0
                            total_pallet = 0
                            
                            for record in packinglist_list:
                                total_weight += record.total_weight_lbs or 0
                                total_pcs += record.pcs or 0
                                total_cbm += record.cbm or 0
                                if record.shipment_batch_number_id:
                                    # жЯ•жЙЊеѓєеЇФзЪДshipment
                                    shipment = await sync_to_async(get_shipment)(record.shipment_batch_number_id)
                                    if shipment:
                                        container['validation']['status'] = 'йФЩиѓѓ'
                                        container['validation']['message'] = f'жЭње≠РеЈ≤жЬЙзЇ¶пЉМзЇ¶жШѓ{shipment.shipment_batch_number}'
                                        break
                            # иЃ°зЃЧжЭњжХ∞пЉИжАїCBM/1.8пЉЙ
                            if total_cbm > 0:
                                total_pallet = round(total_cbm / 1.8, 1)
                            
                            # е¶ВжЮЬж≤°жЬЙйФЩиѓѓпЉМжФґйЫЖid
                            if container['validation']['status'] == 'ж≠£з°Ѓ':
                                container['validation']['ids'] = [f'pl_id{record.id}' for record in packinglist_list]
                                # жЫіжЦ∞зїЯиЃ°дњ°жБѓ
                                container['stats'] = {
                                    'total_weight': total_weight,
                                    'total_pcs': total_pcs,
                                    'total_cbm': total_cbm,
                                    'total_pallet': total_pallet,
                                    'pallet_only': False
                                }
                        else:
                            # дЄ§дЄ™и°®йГљж≤°жЙЊеИ∞
                            container['validation']['status'] = 'йФЩиѓѓ'
                            container['validation']['message'] = 'жЬ™жЙЊеИ∞еѓєеЇФзЪДжЭњжХ∞иЃ∞ељХ'
                
                # й™МиѓБжѓПдЄ™еЃєеЩ®
                for container in group.get('containers', []):
                    await validate_container(container)
                
                # зїЯиЃ°жХідЄ™еИЖзїДзЪДжАїйЗНйЗПгАБжАїдїґжХ∞гАБжАїCBMеТМжАїжЭњжХ∞
                total_weight = 0
                total_pcs = 0
                total_cbm = 0
                total_pallet = 0
                pallet_only = True
                
                for container in group.get('containers', []):
                    stats = container.get('stats', {})
                    total_weight += stats.get('total_weight', 0)
                    total_pcs += stats.get('total_pcs', 0)
                    total_cbm += stats.get('total_cbm', 0)
                    total_pallet += stats.get('total_pallet', 0)
                    if not stats.get('pallet_only', True):
                        pallet_only = False
                
                # жЈїеК†зїЯиЃ°дњ°жБѓеИ∞еИЖзїД
                group['total_weight'] = round(total_weight, 3)
                group['total_pcs'] = total_pcs
                group['total_cbm'] = round(total_cbm, 3)
                group['total_pallet'] = total_pallet
                group['pallet_only'] = pallet_only
                
                # жЈїеК†й™МиѓБйФЩиѓѓдњ°жБѓеИ∞еИЖзїД
                group['validation_errors'] = validation_errors
                group['field_errors'] = field_errors
                
                if not missing_fields and group.get('containers'):
                    valid_groups.append(group)
            
            # жМЙиљ¶зїДеПЈеИЖзїДпЉМж£АжЯ•иљ¶зїДеЖЕжШѓеР¶жЬЙйФЩиѓѓ
            car_group_errors = {}
            for group in valid_groups:
                car_group = group.get('car_group')
                # ж£АжЯ•еИЖзїДжШѓеР¶жЬЙйФЩиѓѓ
                has_error = False
                if group.get('validation_errors'):
                    has_error = True
                else:
                    # ж£АжЯ•жЯЬе≠РжШѓеР¶жЬЙйФЩиѓѓ
                    for container in group.get('containers', []):
                        if container.get('validation', {}).get('status') == 'йФЩиѓѓ':
                            has_error = True
                            break
                # жЫіжЦ∞иљ¶зїДйФЩиѓѓзКґжАБ
                if car_group not in car_group_errors:
                    car_group_errors[car_group] = False
                if has_error:
                    car_group_errors[car_group] = True
            
            # дЄЇжѓПдЄ™еИЖзїДжЈїеК†иљ¶зїДйФЩиѓѓж†ЗењЧ
            for group in valid_groups:
                car_group = group.get('car_group')
                group['car_group_has_error'] = car_group_errors.get(car_group, False)
            
            # дњЭе≠Ше§ДзРЖзїУжЮЬ
            context['groups'] = valid_groups
            context['success'] = f'жЦЗдїґиІ£жЮРеЃМжИРпЉМеЕ±иІ£жЮРеЗЇ {len(valid_groups)} зїДжЬЙжХИжХ∞жНЃ'
            
        # except Exception as e:
        #     context['error'] = f'жЦЗдїґе§ДзРЖе§±иі•: {str(e)}'
        
        return self.template_batch_shipment, context
    
    async def handle_batch_shipment_confirm(self, request: HttpRequest):
        '''е§ДзРЖжЙєйЗПйҐДзЇ¶еЗЇеЇУз°ЃиЃ§еТМеНХдЄ™еИЖзїДйҐДзЇ¶еЗЇеЇУ'''
        context = {}
        
        # ж£АжЯ•жШѓеНХдЄ™йҐДзЇ¶ињШжШѓжЙєйЗПйҐДзЇ¶
        is_single = 'group_appointment_time' in request.POST
        groups = []
        
        if is_single:
            # еНХдЄ™еИЖзїДйҐДзЇ¶
            # жЮДеїЇеИЖзїДжХ∞жНЃ
            shipment_id_str = request.POST.get('group_shipment_id')
            shipment_id = None
            if shipment_id_str:
                try:
                    shipment_id = int(shipment_id_str)
                except (ValueError, TypeError):
                    shipment_id = shipment_id_str
            
            group = {
                'appointment_time': request.POST.get('group_appointment_time'),
                'isa': request.POST.get('group_isa'),
                'pickup_time': request.POST.get('group_pickup_time'),
                'pickup_number': request.POST.get('group_pickup_number'),
                'shipment_id': shipment_id,
                'appointment_account': request.POST.get('group_appointment_account'),
                'appointment_type': request.POST.get('group_appointment_type'),
                'loading_type': request.POST.get('group_loading_type'),
                'car_group': request.POST.get('group_car_group'),
                'total_weight': float(request.POST.get('group_total_weight', 0)),
                'total_pcs': int(request.POST.get('group_total_pcs', 0)),
                'total_cbm': float(request.POST.get('group_total_cbm', 0)),
                'total_pallet': float(request.POST.get('group_total_pallet', 0)),
                'pallet_only': request.POST.get('group_pallet_only', 'true').lower() == 'true',
                'note': request.POST.get('group_note', ''),
                'origin': request.POST.get('group_origin', ''),
                'containers': []
            }
            
            # иІ£жЮРжЯЬе≠Рдњ°жБѓпЉИеП™йЬАи¶БIDпЉЙ
            container_count = int(request.POST.get('group_container_count', 0))
            for j in range(1, container_count + 1):
                # дїОеЙНзЂѓдЉ†йАТзЪДеЃєеЩ®дњ°жБѓдЄ≠жПРеПЦID
                ids_str = request.POST.get(f'group_container_{j}_ids', '')
                ids = ids_str.split(',') if ids_str else []
                
                container = {
                    'ids': ids
                }
                group['containers'].append(container)
            
            groups.append(group)
        else:
            # жЙєйЗПйҐДзЇ¶
            # иОЈеПЦеИЖзїДжХ∞йЗП
            group_count = int(request.POST.get('group_count', 0))
            
            if group_count == 0:
                context['error'] = 'ж≤°жЬЙи¶Бе§ДзРЖзЪДйҐДзЇ¶жХ∞жНЃ'
                return self.template_batch_shipment, context
            
            # иІ£жЮРеИЖзїДжХ∞жНЃ
            for i in range(1, group_count + 1):
                # жЮДеїЇеИЖзїДжХ∞жНЃ
                # е§ДзРЖShipment IDпЉМиљђжНҐдЄЇжХіжХ∞
                shipment_id_str = request.POST.get(f'group_{i}_shipment_id')
                shipment_id = None
                if shipment_id_str:
                    try:
                        shipment_id = int(shipment_id_str)
                    except (ValueError, TypeError):
                        shipment_id = shipment_id_str
                
                group = {
                    'appointment_time': request.POST.get(f'group_{i}_appointment_time'),
                    'appointment_id': request.POST.get(f'group_{i}_isa'),
                    'pickup_time': request.POST.get(f'group_{i}_pickup_time'),
                    'pickup_number': request.POST.get(f'group_{i}_pickup_number'),
                    'shipment_id': shipment_id,
                    'appointment_account': request.POST.get(f'group_{i}_appointment_account'),
                    'appointment_type': request.POST.get(f'group_{i}_appointment_type'),
                    'load_type': request.POST.get(f'group_{i}_loading_type'),
                    'car_group': request.POST.get(f'group_{i}_car_group'),
                    'destination': request.POST.get(f'group_{i}_destination'),
                    'total_weight': float(request.POST.get(f'group_{i}_total_weight', 0)),
                    'total_pcs': int(request.POST.get(f'group_{i}_total_pcs', 0)),
                    'total_cbm': float(request.POST.get(f'group_{i}_total_cbm', 0)),
                    'total_pallet': float(request.POST.get(f'group_{i}_total_pallet', 0)),
                    'pallet_only': request.POST.get(f'group_{i}_pallet_only', 'true').lower() == 'true',
                    'note': request.POST.get(f'group_{i}_note', ''),
                    'origin': request.POST.get(f'group_{i}_origin', ''),
                    'containers': []
                }
                
                # иІ£жЮРжЯЬе≠Рдњ°жБѓпЉИеП™йЬАи¶БIDпЉЙ
                container_count = int(request.POST.get(f'group_{i}_container_count', 0))
                for j in range(1, container_count + 1):
                    # дїОеЙНзЂѓдЉ†йАТзЪДеЃєеЩ®дњ°жБѓдЄ≠жПРеПЦID
                    ids_str = request.POST.get(f'group_{i}_container_{j}_ids', '')
                    ids = ids_str.split(',') if ids_str else []
                    
                    container = {
                        'ids': ids
                    }
                    group['containers'].append(container)
                
                groups.append(group)
        
        # й™МиѓБйҐДзЇ¶дњ°жБѓеТМжЯЬе≠Рдњ°жБѓ
        valid_groups = []
        for group in groups:
            # жФґйЫЖйФЩиѓѓдњ°жБѓ
            errors = []
            
            # ж£АжЯ•ењЕе°Ђе≠ЧжЃµ
            required_fields = ['appointment_time', 'appointment_id', 'appointment_account', 'appointment_type', 'load_type']
            for field in required_fields:
                if not group.get(field):
                    errors.append(f'зЉЇе∞СењЕе°Ђе≠ЧжЃµ: {field}')
            
            # й™МиѓБISAжШѓеР¶дЄЇжХіжХ∞
            try:
                if group.get('isa'):
                    int(group.get('isa'))
            except (ValueError, TypeError):
                errors.append('ISAењЕй°їжШѓжХіжХ∞')
            
            # й™МиѓБйҐДзЇ¶иі¶еПЈ
            valid_accounts = ['Carrier Central1', 'Carrier Central2', 'ZEM-AMF', 'ARM-AMF', 'walmart']
            if group.get('appointment_account') not in valid_accounts:
                errors.append(f"йҐДзЇ¶иі¶еПЈењЕй°їжШѓдї•дЄЛдєЛдЄА: {', '.join(valid_accounts)}")
            
            # й™МиѓБйҐДзЇ¶з±їеЮЛ
            valid_types = ['FTL', 'LTL', 'е§ЦйЕН', 'ењЂйАТ', 'еЃҐжИЈиЗ™жПР']
            if group.get('appointment_type') not in valid_types:
                errors.append(f"йҐДзЇ¶з±їеЮЛењЕй°їжШѓдї•дЄЛдєЛдЄА: {', '.join(valid_types)}")
            
            # й™МиѓБи£Еиљ¶з±їеЮЛ
            valid_loading_types = ['еН°жЭњ', 'еЬ∞жЭњ']
            if group.get('load_type') not in valid_loading_types:
                errors.append(f"и£Еиљ¶з±їеЮЛењЕй°їжШѓдї•дЄЛдєЛдЄА: {', '.join(valid_loading_types)}")
            
            # й™МиѓБеПСиіІдїУеЇУ
            valid_origins = ['NJ-07001', 'SAV-31326', 'LA-91761']
            origin = group.get('origin')
            if not origin:
                errors.append('еПСиіІдїУеЇУдЄНиГљдЄЇз©Ї')
            elif origin not in valid_origins:
                errors.append(f"еПСиіІдїУеЇУењЕй°їжШѓдї•дЄЛдєЛдЄА: {', '.join(valid_origins)}")
            
            # й™МиѓБжЯЬе≠Рдњ°жБѓ
            if not group.get('containers') or len(group.get('containers', [])) == 0:
                errors.append('зЉЇе∞СжЯЬе≠Рдњ°жБѓ')
            
            if not errors:
                valid_groups.append(group)
            else:
                context['error'] = 'йҐДзЇ¶дњ°жБѓжЬЙиѓѓ: ' + '; '.join(errors)
                return self.template_batch_shipment, context
    
        sm = ShippingManagement()
        
        # жМЙиљ¶зїДеПЈеИЖзїД
        car_groups = {}
        for group in valid_groups:
            car_group = group.get('car_group', '1')
            if car_group not in car_groups:
                car_groups[car_group] = []
            car_groups[car_group].append(group)
        
        # е§ДзРЖжѓПдЄ™иљ¶зїД
        results = []
        group_index = 1
        
        for car_group, car_group_groups in car_groups.items():
            success_appointment_ids = []
            
            for group in car_group_groups:
                # жПРеПЦзЫЃзЪДеЬ∞дњ°жБѓпЉИзЫіжО•дљњзФ®йҐДзЇ¶дњ°жБѓйЗМзЪДзЫЃзЪДеЬ∞еАЉпЉЙ
                destination = group.get('destination', '')
                # зФЯжИРжЙєжђ°еПЈ
                shipment_batch_number = await self.generate_unique_batch_number(destination)
                
                # зФЯжИРpickup number
                pickup_number_raw = group.get('pickup_number')
                pickup_number = await self._get_unique_pickup_number(pickup_number_raw)
                
                # жХізРЖжЯЬе≠РID
                selected = []  # plеЉАе§ізЪДID
                selected_plt = []  # pltеЉАе§ізЪДID
                
                for container in group.get('containers', []):
                    for id_str in container.get('ids', []):
                        if id_str.startswith('pl_id'):
                            # еОїжОЙpl_idеЙНзЉАпЉМиљђжНҐдЄЇint
                            try:
                                pl_id = int(''.join(filter(str.isdigit, id_str)))
                                selected.append(pl_id)
                            except (ValueError, IndexError):
                                pass
                        elif id_str.startswith('plt_id'):
                            # еОїжОЙplt_idеЙНзЉАпЉМиљђжНҐдЄЇint
                            try:
                                plt_id = int(''.join(filter(str.isdigit, id_str)))
                                selected_plt.append(plt_id)
                            except (ValueError, IndexError):
                                pass

                # еЗЖе§Зshipment_data
                shipment_data = {
                    'shipment_batch_number': shipment_batch_number,
                    'destination': destination,
                    'total_weight': group.get('total_weight', 0),  
                    'total_cbm': group.get('total_cbm', 0), 
                    'total_pallet': group.get('total_pallet', 0), 
                    'total_pcs': group.get('total_pcs', 0), 
                    'shipment_type': group.get('appointment_type'),
                    'shipment_account': group.get('appointment_account'),
                    'appointment_id': group.get('appointment_id'),
                    'shipment_cargo_id': group.get('shipment_id'), 
                    'shipment_appointment': group.get('appointment_time'),
                    'load_type': group.get('load_type'),
                    'origin': group.get('origin', ''), 
                    'note': group.get('note', ''),
                    'address': '',  # йЬАи¶БдїОйЕНзљЃжИЦжХ∞жНЃеЇУдЄ≠иОЈеПЦ
                    'pickup_number': pickup_number,
                    'pickup_time': group.get('pickup_time'),
                }

                # еИЫеїЇдЄАдЄ™жЦ∞зЪДPOSTе≠ЧеЕЄпЉМеП™еМЕеРЂељУеЙНзїДзЪДдњ°жБѓ
                import json
                post_data = {
                    'shipment_data': json.dumps(shipment_data),
                    'batch_number': shipment_batch_number,
                    'address': shipment_data['address'],
                    'pl_ids': selected,
                    'plt_ids': selected_plt,
                    'type': 'td',
                    'origin': group.get('origin', ''),
                    'load_type': group.get('load_type'),
                    'note': group.get('note'),
                    'destination': group.get('destination'),
                    'shipment_type': group.get('appointment_type'),
                    'appointment_id': group.get('appointment_id'),
                    'shipment_cargo_id': group.get('shipment_id'),
                    'pickup_number': group.get('pickup_number'),
                    'pickup_time': group.get('pickup_time'),
                    'shipment_appointment': group.get('appointment_time'),
                    'csrfmiddlewaretoken': request.POST.get('csrfmiddlewaretoken')
                }
                # жЫіжЦ∞request.POST
                request.POST = post_data
                
                # ж†єжНЃshipment_typeйАЙжЛ©дЄНеРМзЪДе§ДзРЖжЦєж≥Х
                shipment_type = group.get('appointment_type')

                if shipment_type in ['FTL', 'ењЂйАТ', 'е§ЦйЕН']:
                    # и∞ГзФ®ShippingManagementзЪДhandle_appointment_postжЦєж≥Х
                    info = await sm.handle_appointment_post(request, 'post_nsop')
                    # info =  {'success': 'йҐДзЇ¶еЗЇеЇУжИРеКЯ'}
                    # ж£АжЯ•зїУжЮЬ
                    if info:
                        status = 'жИРеКЯ'
                        message = 'йҐДзЇ¶еЗЇеЇУжИРеКЯ'
                        # жФґйЫЖжИРеКЯзЪДйҐДзЇ¶ID
                        success_appointment_ids.append(group.get('appointment_id'))
                    else:
                        status = 'е§±иі•'
                        message = info.get('message', 'йҐДзЇ¶еЗЇеЇУе§±иі•')
                elif shipment_type in ['LTL', 'еЃҐжИЈиЗ™жПР']:
                    # еЗЖе§ЗLTLеТМеЃҐжИЈиЗ™жПРзЪДжХ∞жНЃ
                    # еЯЇдЇОзО∞жЬЙзЪДpost_dataеИЫеїЇжЦ∞зЪДpost_data
                    ltl_post_data = post_data.copy()
                    
                    # ињЩйЗМйЬАи¶БдїОgroupдЄ≠иОЈеПЦзЫЄеЕ≥дњ°жБѓпЉМжЪВжЧґдљњзФ®йїШиЃ§еАЉ
                    # еРОзї≠йЬАи¶БдњЃжФєеЙНзЂѓпЉМеЬ®йҐДзЇ¶дњ°жБѓдЄ≠еМЕеРЂжЫіе§Ъиѓ¶зїЖдњ°жБѓ
                    ltl_post_data['destination'] = destination
                    ltl_post_data['address'] = ''  # йЬАи¶БдїОеЙНзЂѓиОЈеПЦ
                    ltl_post_data['carrier'] = 'Maersk'
                    ltl_post_data['shipment_appointment'] = group.get('appointment_time')
                    ltl_post_data['arm_bol'] = ''  # йЬАи¶БдїОеЙНзЂѓиОЈеПЦ
                    ltl_post_data['arm_pro'] = ''  # йЬАи¶БзФЯжИР
                    ltl_post_data['shipment_type'] = shipment_type
                    ltl_post_data['auto_fleet'] = 'true'  # йїШиЃ§еАЉ
                    ltl_post_data['fleet_cost'] = '0'  # йїШиЃ§еАЉ
                    ltl_post_data['maersk_batch_number'] = shipment_batch_number
                    ltl_post_data['note'] = ''  # йЬАи¶БдїОеЙНзЂѓиОЈеПЦ
                    ltl_post_data['warehouse'] = destination  # еБЗиЃЊwarehouseеТМdestinationзЫЄеРМ
                    
                    # жЫіжЦ∞request.POST
                    request.POST = ltl_post_data
                    # и∞ГзФ®зІБдїУзїСеЃЪйАїиЊС
                    #_, context = await self.handle_ltl_bind_group_shipment(request)
                    context =  {'success': 'йҐДзЇ¶еЗЇеЇУжИРеКЯ'}
                    # ж£АжЯ•зїУжЮЬ
                    if context.get('success'):
                        status = 'жИРеКЯ'
                        message = 'йҐДзЇ¶еЗЇеЇУжИРеКЯ'
                        # жФґйЫЖжИРеКЯзЪДйҐДзЇ¶IDпЉИињЩйЗМйЬАи¶Бж†єжНЃеЃЮйЩЕињФеЫЮеАЉи∞ГжХіпЉЙ
                        # жЪВжЧґеБЗиЃЊињФеЫЮзЪДcontextдЄ≠еМЕеРЂappointment_id
                        if context.get('appointment_id'):
                            success_appointment_ids.append(context.get('appointment_id'))
                    else:
                        status = 'е§±иі•'
                        message = context.get('error', 'йҐДзЇ¶еЗЇеЇУе§±иі•')
                # except Exception as e:
                #     status = 'е§±иі•'
                #     message = f'йҐДзЇ¶еЗЇеЇУе§±иі•: {str(e)}'
                
                # иЃ∞ељХзїУжЮЬ
                results.append({
                    'group_index': group_index,
                    'car_group': car_group,
                    'appointment_time': group.get('appointment_time'),
                    'isa': group.get('isa'),
                    'pickup_time': group.get('pickup_time'),
                    'pickup_number': pickup_number,
                    'appointment_account': group.get('appointment_account'),
                    'appointment_type': group.get('appointment_type'),
                    'loading_type': group.get('loading_type'),
                    'container_count': len(group.get('containers', [])),
                    'shipment_batch_number': shipment_batch_number,
                    'status': status,
                    'message': message
                })
                
                group_index += 1
            
            # зФЯжИРиљ¶жђ°пЉИе¶ВжЮЬиљ¶зїДеЖЕжЬЙе§ЪдЄ™зЇ¶дЄФеЕ®йГ®жИРеКЯпЉЙ
            if len(car_group_groups) > 1 and success_appointment_ids:
                fleet_number = await self._add_appointments_to_fleet(success_appointment_ids)
                # еПѓдї•еЬ®ињЩйЗМжЈїеК†иљ¶жђ°зФЯжИРжИРеКЯзЪДжґИжБѓ
                #fleet_number = 'FLLLLLO88'
                # е∞Жfleet_numberжЈїеК†еИ∞иѓ•иљ¶зїДзЪДжЙАжЬЙзїУжЮЬдЄ≠
                for result in results:
                    if result['car_group'] == car_group:
                        result['fleet_number'] = fleet_number
        
        context['results'] = results
        if is_single:
            context['success'] = 'еНХдЄ™йҐДзЇ¶еЃМжИР'
        else:
            context['success'] = f'жЙєйЗПйҐДзЇ¶еЃМжИРпЉМеЕ±е§ДзРЖ {len(valid_groups)} зїДжХ∞жНЃ'
        
        return self.template_batch_shipment, context
    
    async def handle_download_batch_shipment_template(self, request: HttpRequest):
        '''дЄЛиљљжЙєйЗПйҐДзЇ¶еЗЇеЇУж®°жЭњжЦЗдїґ'''
        import os
        from django.http import FileResponse
        
        # ж®°жЭњжЦЗдїґиЈѓеЊД
        template_path = os.path.join('warehouse', 'templates', 'export_file', 'batch_shipment_template.xlsx')
        
        # ж£АжЯ•жЦЗдїґжШѓеР¶е≠ШеЬ®
        if not os.path.exists(template_path):
            from django.http import HttpResponse
            return HttpResponse('ж®°жЭњжЦЗдїґдЄНе≠ШеЬ®', status=404)
        
        # жЙУеЉАжЦЗдїґеєґињФеЫЮ
        try:
            f = open(template_path, 'rb')
            response = FileResponse(f)
            response['Content-Disposition'] = 'attachment; filename="batch_shipment_template.xlsx"'
            response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response
        except Exception as e:
            from django.http import HttpResponse
            return HttpResponse(f'дЄЛиљље§±иі•: {str(e)}', status=500)
    
    async def handle_maersk_schedule_post(self, request: HttpRequest) -> JsonResponse:
        """е§ДзРЖMaerskйҐДзЇ¶дЄЛеНХ"""
        try:     
            # 1. жПРеПЦеПВжХ∞
            quote_id = request.POST.get('quote_id')
            service_code = request.POST.get('service_code')
            consignee_json = request.POST.get('consignee_json')
            line_items_json = request.POST.get('line_items_json')
            schedule_json = request.POST.get('schedule_json')
            warehouse = request.session.get('warehouse') or request.POST.get('warehouse')
            if not all([quote_id, service_code, consignee_json, line_items_json, schedule_json]):
                return JsonResponse({'success': False, 'message': 'ењЕи¶БеПВжХ∞зЉЇе§±'}, status=400)

            consignee = json.loads(consignee_json)
            line_items_raw = json.loads(line_items_json)
            schedule = json.loads(schedule_json)

            # 2. жЮДйА† Shipper дњ°жБѓ (зђ¶еРИ ShipmentAddress еЃЪдєЙ)
            shipper_config = app_config.WAREHOUSE_ADDRESS.get(warehouse, {})
            if not shipper_config:
                 # йїШиЃ§дљњзФ® LA-91761 е¶ВжЮЬжЙЊдЄНеИ∞
                 shipper_config = app_config.WAREHOUSE_ADDRESS.get('LA-91761', {})
            
            shipper_address = {
                "name": shipper_config.get('name'),
                "address1": shipper_config.get('address1'),
                "city": shipper_config.get('city'),
                "regionCode": shipper_config.get('regionCode'),
                "postalCode": shipper_config.get('postalCode'),
                "countryCode": shipper_config.get('countryCode', 'US'),
                "phone": "909-320-8774" # йїШиЃ§жИЦдїОйЕНзљЃиОЈеПЦ
            }

            # 3. жЮДйА† LineItems (ж†ЉеЉПеМЦ)
            line_items = []
            for item in line_items_raw:
                line_items.append({
                    "description": item.get('description') or 'Pallet',
                    "pieces": int(item.get('pieces') or item.get('Pieces') or 1),
                    "length": int(item.get('length') or item.get('Length') or 0),
                    "width": int(item.get('width') or item.get('Width') or 0),
                    "height": int(item.get('height') or item.get('Height') or 0),
                    "weight": int(float(item.get('weight') or item.get('Weight') or 0)),
                    # дљњзФ®еЙНзЂѓдЉ†еЕ•зЪДеМЕи£ЕжЦєеЉПпЉМйїШиЃ§ Pallet
                    "packaging": item.get('packaging') or item.get('Packaging') or "Pallet", 
                    "weightUnit": "lb",
                    "dimensionalUnit": "in"
                })

            # 4. и∞ГзФ® Maersk дЄЛеНХ API
            # жЮДйА† Consignee Address (зђ¶еРИ ShipmentAddress еЃЪдєЙ)
            consignee_address = {
                "name": consignee.get('company') or consignee.get('name'), # еЕђеПЄеРНдЄЇ name
                "contact": consignee.get('name'), # иБФз≥їдЇЇдЄЇ contact
                "address1": consignee.get('address1'),
                "city": consignee.get('city'),
                "regionCode": consignee.get('state'),
                "postalCode": consignee.get('zipcode'),
                "countryCode": "US",
                "phone": consignee.get('phone'),
                "email": consignee.get('email')
            }
            
            # е§ДзРЖ Consignee References (зЫЃзЪДеЬ∞е§ДзРЖ)
            warehouse_type = request.POST.get('warehouse_type', 'public')
            fleet_cost = schedule.get('cost', '')
            destination = schedule.get('destination', '')
            destination_name = destination
            
            # зФЯжИР batch_number зФ®дЇОдљЬдЄЇ references
            batch_number_ref = ""
            if warehouse_type == 'private' and len(destination) > 8:
                destination_name = destination[:8]
            else:
                destination_name = destination
                
            batch_number_ref = await self.generate_unique_batch_number(destination_name)
            if not batch_number_ref:
                return JsonResponse({'success': False, 'message': 'жЙєжђ°еПЈзФЯжИРе§±иі•'}, status=400)

            # --- е§ДзРЖзЙєжЃКжЬНеК° ---
            # 1. Liftgate (дїО schedule жПРеПЦ is_liftgate)
            consignee_accessorials = []
            if schedule.get('is_liftgate'):
                consignee_accessorials.append("Liftgate")
            
            # 2. Insurance (дїО schedule жПРеПЦ is_insurance еТМ insurance_amount)
            monetary = None
            if schedule.get('is_insurance'):
                try:
                    declared_value = float(schedule.get('insurance_amount', 0))
                    monetary = {
                        "isDeclaredValueInsurance": True,
                        "declaredValue": declared_value
                    }
                except (ValueError, TypeError):
                    pass # е¶ВжЮЬйЗСйҐЭжЧ†жХИпЉМењљзХ•дњЭйЩ©йАЙй°є

            payload = {
                "quoteId": int(quote_id),
                "serviceCode": service_code,
                "shipper": {
                    "address": shipper_address,
                    "references": [batch_number_ref]
                },
                "consignee": {
                    "address": consignee_address,
                    "accessorials": consignee_accessorials  # еҐЮеК† consignee accessorials
                },
                "lineItems": line_items,
                "shipDate": schedule.get('pickup_time'), # ењЕе°Ђ
                "shipReadyTime": "09:00:00", # йїШиЃ§еАЉ
                "shipCloseTime": "17:00:00", # йїШиЃ§еАЉ
                "accessorials": [] 
            }

            # 3. жЈїеК† Insurance еИ∞ payload
            if monetary:
                payload["monetary"] = monetary

            # 4. E0 жЬНеК°йЬАи¶БйҐДзЇ¶жЧґйЧі (дЉ†еЕ• SpecialInstructions)
            if service_code == 'E0':
                # йҐДзЇ¶жЧґйЧідї•е≠Чзђ¶дЄ≤ељҐеЉПдЉ†еЕ•
                payload["SpecialInstructions"] = schedule.get('pickup_time')

            api_url = "https://zem-maersk-gateway.kindmoss-a5050a64.eastus.azurecontainerapps.io/shipment"
            api_key = os.environ.get("MAERSK_API_KEY")
            headers = {
                "Content-Type": "application/json",
                "x-api-key": api_key
            }

            pro_number = None
            async with aiohttp.ClientSession() as session:
                async with session.post(api_url, json=payload, headers=headers) as response:
                    if response.status == 200:
                        res_data = await response.json()
                        pro_number = res_data.get('housebill')
                    else:
                        text = await response.text()
                        return JsonResponse({'success': False, 'message': f'Maersk APIдЄЛеНХе§±иі•: {response.status} - {text}'}, status=response.status)

            if not pro_number:
                return JsonResponse({'success': False, 'message': 'дЄЛеНХжИРеКЯдљЖжЬ™иОЈеПЦеИ∞PROеПЈ'}, status=500)

            # 5. ж†єжНЃдїУеЇУз±їеЮЛеИЖеПСе§ДзРЖ
            warehouse_type = request.POST.get('warehouse_type', 'public')
            
            if warehouse_type == 'private':  
                # --- зІБдїУйҐДзЇ¶еЗЇеЇУйАїиЊС (и∞ГзФ® handle_ltl_bind_group_shipment) ---
                post_data = request.POST.copy()
                post_data['destination'] = schedule.get('destination')
                post_data['address'] = f"{consignee.get('address1')}, {consignee.get('city')}, {consignee.get('state')} {consignee.get('zipcode')}"
                post_data['carrier'] = 'Maersk'
                post_data['shipment_appointment'] = schedule.get('pickup_time')
                post_data['arm_bol'] = schedule.get('bol', '')
                post_data['arm_pro'] = pro_number
                post_data['shipment_type'] = schedule.get('shipment_type', 'LTL')
                # зІБдїУ auto_fleet жО•жФґ 'true'/'false' е≠Чзђ¶дЄ≤
                post_data['auto_fleet'] = 'true' if schedule.get('auto_schedule') == 'жШѓ' else 'false'
                post_data['fleet_cost'] = fleet_cost
                # дЉ†йАТеЈ≤зФЯжИРзЪД batch_number 
                post_data['maersk_batch_number'] = batch_number_ref
                # дЉ†йАТе§Зж≥®
                post_data['note'] = schedule.get('note', '')
                post_data['warehouse'] = warehouse

                request.POST = post_data
                # и∞ГзФ®зІБдїУзїСеЃЪйАїиЊС
                template_name, context = await self.handle_ltl_bind_group_shipment(request)
                
                if 'error_messages' in context:
                    return JsonResponse({'success': False, 'message': f"зІБдїУйҐДзЇ¶е§±иі•: {context['error_messages']}"}, status=400)
                
                return JsonResponse({
                    'success': True, 
                    'pro_number': pro_number, 
                    'batch_number': batch_number_ref, # зІБдїУйАїиЊСжЪВдЄНињФеЫЮ batch_number зїЩеЙНзЂѓеЉєз™ЧжШЊз§Ї
                    'cost': fleet_cost,
                    'message': 'дЄЛеНХеПКзІБдїУйҐДзЇ¶жИРеКЯ'
                })
                
            else:
                # --- еЕђдїУйАїиЊС (и∞ГзФ® handle_appointment_post) ---
                # жЮДйА† request.POST
                # зФЯжИРеФѓдЄАзЪД appointment_id
                appointment_id = f"M-{pro_number}"
                
                # еЗЖе§ЗеПВжХ∞
                post_data = request.POST.copy()
                post_data['shipment_type'] = schedule.get('shipment_type', 'LTL')
                post_data['appointment_id'] = appointment_id
                post_data['destination'] = schedule.get('destination')
                post_data['address'] = f"{consignee.get('address1')}, {consignee.get('city')}, {consignee.get('state')} {consignee.get('zipcode')}"
                post_data['pickup_time'] = schedule.get('pickup_time') # жПРиіІжЧґйЧі
                post_data['shipment_appointment'] = schedule.get('pickup_time') # дєЯжШѓйҐДзЇ¶жЧґйЧі
                post_data['note'] = schedule.get('note', '')
                post_data['maersk_batch_number'] = batch_number_ref
                # cargo_ids еТМ plt_ids еЈ≤зїПе≠ШеЬ®дЇО post_data дЄ≠
                
                # жЫњжНҐ request.POST
                request.POST = post_data
                
                # и∞ГзФ® handle_appointment_post
                template_name, context = await self.handle_appointment_post(request)
                
                if 'error_messages' in context:
                    return JsonResponse({'success': False, 'message': f"йҐДзЇ¶дњЭе≠Ше§±иі•: {context['error_messages']}"}, status=400)
                
                success_msg = context.get('success_messages', '')
                # жПРеПЦзФЯжИРзЪД batch_number (йАЪеЄЄеЬ® success_messages дЄ≠ "зїСеЃЪжИРеКЯпЉМжЙєжђ°еПЈжШѓXXX")

                # 6. жЫіжЦ∞ PRO еПЈеИ∞ Pallet
                # ж†єжНЃ batch_number жИЦ appointment_id жЯ•жЙЊ Shipment
                try:
                    shipment = await sync_to_async(Shipment.objects.get)(appointment_id=appointment_id)
                    
                    # жЫіжЦ∞ Pallet зЪД ltl_pro_num
                    # ж≥®жДПпЉЪPallet еЕ≥иБФзЪДжШѓ shipment_batch_number (Shipment еѓєи±°)
                    await sync_to_async(Pallet.objects.filter(shipment_batch_number=shipment).update)(
                        ltl_pro_num=pro_number
                    )
                    
                    # 7. иЗ™еК®жОТиљ¶ (е¶ВжЮЬйЬАи¶Б)
                    if schedule.get('auto_schedule') == 'жШѓ':
                        # и∞ГзФ® _add_appointments_to_fleet
                        fleet_number = await self._add_appointments_to_fleet([appointment_id])
                        
                        # жЫіжЦ∞ Fleet дњ°жБѓ (жѓФе¶В carrier, pro)
                        fleet = await sync_to_async(Fleet.objects.get)(fleet_number=fleet_number)
                        fleet.carrier = 'Maersk'
                        fleet.fleet_cost = fleet_cost
                        await sync_to_async(fleet.save)()
                        
                except Exception as e:
                    # иЃ∞ељХи≠¶еСКдљЖдЄНдЄ≠жЦ≠жµБз®ЛпЉМеЫ†дЄЇдЄЛеНХеЈ≤жИРеКЯ
                    print(f"жЫіжЦ∞PROжИЦжОТиљ¶е§±иі•: {e}")

                return JsonResponse({
                    'success': True, 
                    'pro_number': pro_number, 
                    'batch_number': batch_number_ref,
                    'cost': fleet_cost,
                    'message': 'дЄЛеНХеПКйҐДзЇ¶жИРеКЯ'
                })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': f'з≥їзїЯеЉВеЄЄ: {str(e)}'}, status=500)

    async def handle_cancel_maersk_shipment(self, request: HttpRequest) -> JsonResponse:
        """еПЦжґИMaerskдЄЛеНХ"""
        try:
            fleet_number = request.POST.get('fleet_number')
            if not fleet_number:
                return JsonResponse({'success': False, 'message': 'жЬ™жПРдЊЫиљ¶жђ°еПЈ'}, status=400)

            # 1. жЯ•жЙЊ Fleet еТМ Shipment
            try:
                fleet = await sync_to_async(Fleet.objects.get)(fleet_number=fleet_number)
            except Fleet.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'жЙЊдЄНеИ∞иѓ•иљ¶жђ°'}, status=404)

            # иОЈеПЦеЕ≥иБФзЪД shipments
            shipments = await sync_to_async(list)(
                Shipment.objects.filter(fleet_number=fleet)
            )
            
            # иОЈеПЦ PRO еПЈ
            pro_number = None
            if shipments:
                for s in shipments:
                    if s.ARM_PRO:
                        pro_number = s.ARM_PRO
                        break
            
            if not pro_number:
                return JsonResponse({'success': False, 'message': 'зЉЇе∞С PRO еПЈз†БпЉМжЧ†ж≥ХеПЦжґИ Maersk иЃҐеНХ'}, status=400)

            # 2. и∞ГзФ® Maersk API еПЦжґИ
            api_url = "https://zem-maersk-gateway.kindmoss-a5050a64.eastus.azurecontainerapps.io/shipment/void"
            api_key = os.environ.get("MAERSK_API_KEY")
            
            params = {
                "pro_number": pro_number,
                "control_station": "GOP"
            }
            headers = {
                "x-api-key": api_key
            }

            async with aiohttp.ClientSession() as session:
                async with session.post(api_url, params=params, headers=headers) as response:
                    if response.status == 200:
                        # 3. жЬђеЬ∞еПЦжґИ
                        await self._cancel_maersk_fleet(fleet_number)
                        return JsonResponse({
                            'success': True, 
                            'pro_number': pro_number,
                            'message': 'Maersk иЃҐеНХеЈ≤еПЦжґИ'
                        })
                    else:
                        text = await response.text()
                        return JsonResponse({'success': False, 'message': f'APIиѓЈж±Ве§±иі•: {response.status} - {text}'}, status=response.status)

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': f'з≥їзїЯеЉВеЄЄ: {str(e)}'}, status=500)

    async def _cancel_maersk_fleet(self, fleet_number):
        """жЬђеЬ∞еПЦжґИиљ¶жђ°йАїиЊС"""
        shipment = await sync_to_async(Shipment.objects.get)(fleet_number__fleet_number=fleet_number)
        await sync_to_async(shipment.delete)()

        fleet = await sync_to_async(Fleet.objects.get)(fleet_number=fleet_number)
        if fleet.departured_at is not None:
            raise RuntimeError(
                f"Shipment with batch number {fleet_number} has been shipped!"
            )
        await sync_to_async(fleet.delete)()

    async def handle_get_maersk_quote(self, request: HttpRequest) -> JsonResponse:
        """и∞ГзФ®Maersk APIиОЈеПЦжК•дїЈ"""
        try:
            origin_zip = request.POST.get('origin_zip')
            dest_zip = request.POST.get('dest_zip')
            ship_date = request.POST.get('ship_date')
            need_liftgate_raw = request.POST.get('need_liftgate', 'еР¶')
            need_liftgate = 'true' if str(need_liftgate_raw).strip() in ('жШѓ', 'true', 'True', '1') else 'false'

            if dest_zip and not dest_zip.strip().isdigit():
                # и°®з§ЇеЙНзЂѓжЯ•иѓҐзЪДжШѓдїУзВєпЉМйЬАи¶БеОїдЇЪй©ђйАКжЦЗдїґйЗМжЯ•жЙЊйВЃзЉЦ
                if dest_zip in amazon_fba_locations:
                    fba = amazon_fba_locations[dest_zip]
                    dest_zip = fba['zipcode']
                else:
                    return JsonResponse({'success': False, 'message': 'ж≤°жЬЙжЯ•еИ∞иѓ•зЫЃзЪДеЬ∞зЪДйВЃзЉЦ'}, status=400)
            
            # жЦ∞еҐЮеПВжХ∞пЉЪзїУжЮДеМЦзЪДиіІзЙ©жШОзїЖ
            line_items_json = request.POST.get('line_items_json')
            
            # жЧІеПВжХ∞пЉИдњЭзХЩеЕЉеЃєпЉЙ
            cargo_details = request.POST.get('cargo_details')
            total_weight = request.POST.get('total_weight')

            if not all([origin_zip, dest_zip, ship_date]):
                return JsonResponse({'success': False, 'message': 'еЯЇз°АеПВжХ∞дЄНеЃМжХі'}, status=400)

            line_items = []
            
            if line_items_json:
                try:
                    parsed_items = json.loads(line_items_json)
                    for item in parsed_items:
                        line_items.append({
                            "description": item.get('description') or 'Pallet',
                            "pieces": int(item.get('pieces', 1)),
                            "length": int(item.get('length', 0)),
                            "width": int(item.get('width', 0)),
                            "height": int(item.get('height', 0)),
                            "weight": int(float(item.get('weight', 0))),
                            "packaging": (item.get('packaging') or 'Pallet')
                        })
                except json.JSONDecodeError:
                    return JsonResponse({'success': False, 'message': 'иіІзЙ©жШОзїЖжХ∞жНЃж†ЉеЉПйФЩиѓѓ'}, status=400)
            elif cargo_details:
                # иІ£жЮРиіІзЙ©иѓ¶жГЕ (жЧІйАїиЊС)
                # ж†ЉеЉП: 32*35*60*3жЭњ жИЦ 32*35*60
                
                # е¶ВжЮЬжЬЙе§ЪдЄ™е∞ЇеѓЄи°МпЉМжМЙжНҐи°Мзђ¶еИЖеЙ≤
                detail_lines = [line.strip() for line in cargo_details.split('\n') if line.strip()]
                
                # иЃ°зЃЧжАїжЭњжХ∞зФ®дЇОеИЖжСКйЗНйЗП
                total_pallets = 0
                parsed_details = []
                
                for line in detail_lines:
                    # еМєйЕН L*W*H*Count жИЦ L*W*H
                    match = re.match(r'(\d+)\*(\d+)\*(\d+)(?:\*(\d+))?', line)
                    if match:
                        l, w, h, count = match.groups()
                        count = int(count) if count else 1
                        total_pallets += count
                        parsed_details.append({'l': int(l), 'w': int(w), 'h': int(h), 'count': count})
                
                if total_pallets == 0:
                    return JsonResponse({'success': False, 'message': 'жЧ†ж≥ХиІ£жЮРжЙШзЫШе∞ЇеѓЄдњ°жБѓ'}, status=400)

                # еИЖжСКйЗНйЗП
                try:
                    weight_val = float(total_weight) if total_weight else 0
                except ValueError:
                    weight_val = 0
                    
                weight_per_pallet = int(weight_val / total_pallets) if total_pallets > 0 else 0
                
                for item in parsed_details:
                    # еѓєдЇОжѓПдЄАзІНе∞ЇеѓЄпЉМзФЯжИРеѓєеЇФжХ∞йЗПзЪДLineItem
                    for _ in range(item['count']):
                        line_items.append({
                            "description": "Pallet",
                            "pieces": 1,
                            "length": item['l'],
                            "width": item['w'],
                            "height": item['h'],
                            "weight": weight_per_pallet,
                            "packaging": "Pallet"
                        })
            else:
                 return JsonResponse({'success': False, 'message': 'зЉЇе∞СиіІзЙ©жШОзїЖ'}, status=400)

            ship_date_formatted = ship_date
            try:
                if ship_date and '-' in ship_date:
                    parts = ship_date.split('-')
                    if len(parts) == 3:
                        ship_date_formatted = f"{parts[1].zfill(2)}/{parts[2].zfill(2)}/{parts[0]}"
            except Exception:
                ship_date_formatted = ship_date

            # жЮДеїЇдїЕеРЂиЃ°дїЈе≠ЧжЃµзЪДеИЧи°®пЉМйБњеЕНдЄЛжЄЄй™МиѓБе§±иі•
            rating_items = [
                {
                    "description": it["description"],
                    "pieces": it["pieces"],
                    "length": it["length"],
                    "width": it["width"],
                    "height": it["height"],
                    "weight": it["weight"],
                }
                for it in line_items
            ]
            payload = {
                "shipDate": ship_date_formatted,
                "origin_zip": origin_zip,
                "dest_zip": dest_zip,
                "lineItems": rating_items,
                "liftgate": need_liftgate,
                "declaredValue": None,
                "insuranceValue": None,
                "debrisRemoval": None
            }
            api_url = "https://zem-maersk-gateway.kindmoss-a5050a64.eastus.azurecontainerapps.io/rating"
            api_key = os.environ.get("MAERSK_API_KEY")
            
            if not api_key:
                 return JsonResponse({'success': False, 'message': 'жЬ™йЕНзљЃAPI Key'}, status=500)

            headers = {
                "Content-Type": "application/json",
                "x-api-key": api_key
            }

            async with aiohttp.ClientSession() as session:
                async with session.post(api_url, json=payload, headers=headers) as response:
                    if response.status == 200:
                        data = await response.json()
                        # е∞ЖиѓЈж±ВеПВжХ∞дЄ≠зЪД lineItems ж≥®еЕ•еИ∞еУНеЇФжХ∞жНЃдЄ≠пЉМдї•дЊњеЙНзЂѓдљњзФ®
                        data['lineItems'] = line_items
                        data['need_liftgate'] = need_liftgate
                        return JsonResponse({'success': True, 'data': data})
                    else:
                        text = await response.text()
                        return JsonResponse({'success': False, 'message': f'APIи∞ГзФ®е§±иі•: {response.status} - {text}'}, status=response.status)

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    async def handle_get_maersk_tracking(self, request: HttpRequest) -> JsonResponse:
        try:
            pro_number = (request.POST.get('pro_number') or '').strip()
            fleet_number = (request.POST.get('fleet_number') or '').strip()
            if not pro_number and fleet_number:
                pro_number = await sync_to_async(
                    lambda: (
                        Shipment.objects.filter(
                            fleet_number__fleet_number=fleet_number,
                            carrier='Maersk',
                            ARM_PRO__isnull=False,
                        )
                        .exclude(ARM_PRO='')
                        .values_list('ARM_PRO', flat=True)
                        .distinct()
                        .first()
                    )
                )()

            if not pro_number:
                return JsonResponse({'success': False, 'message': 'зЉЇе∞С pro_number'}, status=400)

            api_url = "https://zem-maersk-gateway.kindmoss-a5050a64.eastus.azurecontainerapps.io/tracking"
            api_key = '2Tdtqrj4dqnooXIJi4ReCVrMGW3ehJnC'

            headers = {
                "Content-Type": "application/json",
                "x-api-key": api_key
            }
            payload = {"pro_number": pro_number}

            async with aiohttp.ClientSession() as session:
                async with session.post(api_url, json=payload, headers=headers) as response:
                    if response.status == 200:
                        data = await response.json()
                        return JsonResponse({'success': True, 'data': data, 'pro_number': pro_number})
                    text = await response.text()
                    return JsonResponse(
                        {'success': False, 'message': f'APIи∞ГзФ®е§±иі•: {response.status} - {text}'},
                        status=response.status
                    )
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    async def handle_check_business_residential(self, request: HttpRequest) -> HttpResponse:
        '''ж£АжЯ•еХЖзІБеЬ∞еЭА'''
        try:
            cargo_id = request.POST.get("cargo_id", "").strip()
            address = request.POST.get("address", "").strip()
            city = request.POST.get("city", "").strip()
            state = request.POST.get("state", "").strip()
            zip_code = request.POST.get("zip", "").strip()
            country = request.POST.get("country", "US").strip()

            if not zip_code:
                return JsonResponse({'success': False, 'message': 'йВЃзЉЦжШѓењЕе°Ђй°є'}, status=400)

            api_url = "https://zem-maersk-gateway.kindmoss-a5050a64.eastus.azurecontainerapps.io/get_rdi"
            api_key = '2Tdtqrj4dqnooXIJi4ReCVrMGW3ehJnC'

            headers = {
                "Content-Type": "application/json",
                "x-api-key": api_key
            }
            
            payload = {
                "street": address,
                "city": city,
                "state": state,
                "zipcode": zip_code,
                "country": country
            }

            rdi_value = None
            async with aiohttp.ClientSession() as session:
                async with session.post(api_url, json=payload, headers=headers) as response:
                    if response.status == 200:
                        data = await response.json()
                        rdi_value = data.get('rdi', '')
                    else:
                        text = await response.text()
                        # е§ДзРЖRDIжЙЊдЄНеИ∞зЪДжГЕеЖµпЉМињЩдЄНжШѓйФЩиѓѓпЉМеП™жШѓиѓ•еЬ∞еЭАжЯ•дЄНеИ∞
                        if response.status == 404 and 'RDI value not found' in text:
                            return JsonResponse(
                                {'success': False, 'message': 'ељУеЙНеЬ∞еЭАжЯ•дЄНеИ∞жШѓеХЖдЄЪеЬ∞еЭАињШжШѓзІБдЇЇеЬ∞еЭА'},
                                status=200
                            )
                        # еЕґдїЦйФЩиѓѓжЙНињФеЫЮзЬЯж≠£зЪДйФЩиѓѓдњ°жБѓ
                        return JsonResponse(
                            {'success': False, 'message': f'APIи∞ГзФ®е§±иі•: {response.status} - {text}'},
                            status=response.status
                        )

            # дњЭе≠ШеИ∞жХ∞жНЃеЇУ
            if cargo_id:
                if cargo_id.startswith('plt_'):
                    # PalletиЃ∞ељХ
                    plt_ids = cargo_id[4:].split(',')
                    for plt_id in plt_ids:
                        try:
                            plt_id_int = int(plt_id.strip())
                            await sync_to_async(Pallet.objects.filter(id=plt_id_int).update)(
                                ltl_address=address,
                                ltl_city=city,
                                ltl_state=state,
                                ltl_zipcode=zip_code,
                                ltl_address_type=rdi_value
                            )
                        except Exception:
                            pass
                else:
                    # PackingListиЃ∞ељХ
                    pl_ids = cargo_id.split(',')
                    for pl_id in pl_ids:
                        try:
                            pl_id_int = int(pl_id.strip())
                            await sync_to_async(PackingList.objects.filter(id=pl_id_int).update)(
                                ltl_address=address,
                                ltl_city=city,
                                ltl_state=state,
                                ltl_zipcode=zip_code,
                                ltl_address_type=rdi_value
                            )
                        except Exception:
                            pass

            return JsonResponse({'success': True, 'data': {'rdi': rdi_value}})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    async def handle_search_by_container_and_destination(self, request: HttpRequest) -> JsonResponse:
        """ж†єжНЃжЯЬеПЈеТМдїУзВєжЯ•иѓҐPackingListеТМPallet"""
        try:
            import json
            data = json.loads(request.body.decode('utf-8'))
            search_data = data.get('data', [])

            all_packinglists = []
            all_pallets = []

            for item in search_data:
                destination = item.get('destination', '').strip()
                container_nos_str = item.get('container_nos', '').strip()

                # иІ£жЮРжЯЬеПЈпЉЪжФѓжМБжНҐи°МжИЦйАЧеПЈеИЖйЪФ
                container_nos = []
                if container_nos_str:
                    # еЕИжМЙжНҐи°МеИЖеЙ≤пЉМеЖНжМЙйАЧеПЈеИЖеЙ≤
                    for part in container_nos_str.split('\n'):
                        container_nos.extend([cn.strip() for cn in part.split(',') if cn.strip()])

                # жЮДеїЇйАЪзФ®жЯ•иѓҐжЭ°дїґ
                base_criteria = models.Q()
                if not destination:
                    continue
                    
                if not container_nos:
                    continue
                base_criteria &= models.Q(
                    destination=destination,
                    container_number__container_number__in=container_nos,
                    container_number__orders__cancel_notification=False,
                    shipment_batch_number__isnull=True
                ) & ~models.Q(container_number__orders__order_type='зЫійАБ')

                # PackingListйҐЭе§ЦжЭ°дїґпЉЪcontainer_number__orders__offload_id__offload_at__isnull=True
                pl_criteria = base_criteria & models.Q(container_number__orders__offload_id__offload_at__isnull=True)

                # PalletйҐЭе§ЦжЭ°дїґпЉЪcontainer_number__orders__offload_id__offload_at__isnull=False
                plt_criteria = base_criteria & models.Q(container_number__orders__offload_id__offload_at__isnull=False)
                
                # жЯ•иѓҐPackingList
                packinglists = await sync_to_async(list)(
                    PackingList.objects.prefetch_related(
                        "container_number",
                        "container_number__orders",
                        "container_number__orders__warehouse",
                        "container_number__orders__offload_id",
                        "container_number__orders__customer_name",
                    )
                    .filter(pl_criteria)
                    .annotate(
                        str_id=Cast("id", CharField()),
                        str_fba_id=Cast("fba_id", CharField()),
                        str_ref_id=Cast("ref_id", CharField()),
                        str_container_number=Cast("container_number__container_number", CharField()),
                        data_source=Value("PACKINGLIST", output_field=CharField()),
                    )
                    .values(
                        "destination",
                        "container_number__container_number",
                        "delivery_method",
                        "note",
                        "PO_ID",
                        "data_source",
                    )
                    .annotate(
                        fba_ids=StringAgg(
                            "str_fba_id",
                            delimiter=",",
                            distinct=True,
                            ordering="str_fba_id",
                        ),
                        ref_ids=StringAgg(
                            "str_ref_id",
                            delimiter=",",
                            distinct=True,
                            ordering="str_ref_id",
                        ),
                        ids=StringAgg(
                            "str_id", delimiter=",", distinct=True, ordering="str_id"
                        ),
                        total_pcs=Sum("pcs", output_field=FloatField()),
                        total_cbm = Round(Sum("cbm", output_field=FloatField()), 3),
                        total_weight_lbs=Round(Sum("total_weight_lbs", output_field=FloatField()),3),
                        total_n_pallet_est= Ceil(Sum("cbm", output_field=FloatField()) / 1.8),  # дљњзФ®еОЯеІЛcbmе≠ЧжЃµиЃ°зЃЧ
                        total_n_pallet_act= Ceil(Sum("cbm", output_field=FloatField()) / 1.8),  # дљњзФ®еОЯеІЛcbmе≠ЧжЃµиЃ°зЃЧ
                        label=Value("EST"),
                    )
                )
                all_packinglists.extend(packinglists)

                # жЯ•иѓҐPallet
                pallets = await sync_to_async(list)(
                    Pallet.objects.prefetch_related(
                        "container_number",
                        "container_number__orders",
                        "container_number__orders__warehouse",
                        "container_number__orders__offload_id",
                        "container_number__orders__customer_name",
                    )
                    .filter(plt_criteria)
                    .annotate(
                        str_id=Cast("id", CharField()),
                        str_fba_id=Cast("fba_id", CharField()),
                        str_ref_id=Cast("ref_id", CharField()),
                        str_container_number=Cast("container_number__container_number", CharField()),
                        data_source=Value("PALLET", output_field=CharField()),
                    )
                    .values(
                        "destination",
                        "container_number__container_number",
                        "delivery_method",
                        "note",
                        "data_source",
                        "PO_ID",
                    ).annotate(
                        fba_ids=StringAgg(
                            "str_fba_id",
                            delimiter=",",
                            distinct=True,
                            ordering="str_fba_id",
                        ),
                        ref_ids=StringAgg(
                            "str_ref_id",
                            delimiter=",",
                            distinct=True,
                            ordering="str_ref_id",
                        ),
                        plt_ids=StringAgg(
                            "str_id", delimiter=",", distinct=True, ordering="str_id"
                        ),
                        total_pcs=Sum("pcs", output_field=IntegerField()),
                        total_cbm = Round(Sum("cbm", output_field=FloatField()), 3),
                        total_weight_lbs=Round(Sum("weight_lbs", output_field=FloatField()),3),
                        total_n_pallet_act=Count("pallet_id", distinct=True),
                    )
                )
                all_pallets.extend(pallets)

            # еРИеєґжЙАжЬЙжХ∞жНЃеєґиІДиМГж†ЉеЉП
            all_data = all_packinglists + all_pallets
            
            # иІДиМГжХ∞жНЃж†ЉеЉПпЉМз°ЃдњЭдЄО openBindModal еЗљжХ∞жЬЯжЬЫзЪДж†ЉеЉПдЄАиЗі
            formatted_data = []
            for item in all_data:
                # ж†єжНЃ data_source з°ЃеЃЪ cargo_id еТМ plt_ids
                if item.get('data_source') == 'PACKINGLIST':
                    cargo_id = item.get('ids', '')  # дљњзФ®иБЪеРИеРОзЪДidsе≠ЧжЃµ
                    plt_ids = ''
                else:
                    cargo_id = ''
                    plt_ids = item.get('plt_ids', '')  # дљњзФ®иБЪеРИеРОзЪДplt_idsе≠ЧжЃµ
                
                # иОЈеПЦжЯЬеПЈдњ°жБѓ
                container_no = item.get('str_container_number', '')
                if not container_no and item.get('container_number__container_number'):
                    container_no = str(item.get('container_number__container_number'))
                
                # иОЈеПЦеЕ•дїУжЧґйЧіпЉИйЬАи¶БдїОеЕ≥иБФзЪДoffload_idиОЈеПЦпЉЙ
                offload_time = '-'  # йїШиЃ§еАЉ
                
                formatted_item = {
                    'ids': cargo_id,  # cargo_id
                    'plt_ids': plt_ids,
                    'total_weight': item.get('total_weight_lbs', 0),
                    'total_cbm': item.get('total_cbm', 0),
                    'total_pallet': item.get('total_n_pallet_act', item.get('total_n_pallet_est', 0)),
                    'ref_ids': item.get('ref_ids', '-'),
                    'fba_ids': item.get('fba_ids', '-'),
                    'container_no': container_no,
                    'is_dropped': False,  # йїШиЃ§еАЉ
                    'offload_time': offload_time,
                    'customer_name': item.get('customer_name', ''),  # еПѓиГљйЬАи¶БдїОеЕ≥иБФжЯ•иѓҐиОЈеПЦ
                    'delivery_method': item.get('delivery_method', ''),
                    'destination': item.get('destination', ''),  # жЈїеК†зЫЃзЪДеЬ∞е≠ЧжЃµ
                    'cns': container_no,  # жЯЬеПЈеИЧи°®
                    'total_n_pallet_act': item.get('total_n_pallet_act', 0),
                    'total_n_pallet_est': item.get('total_n_pallet_est', 0),
                    'total_weight_lbs': item.get('total_weight_lbs', 0),
                    'data_source': item.get('data_source', '')
                }
                formatted_data.append(formatted_item)
            
            return JsonResponse({'success': True, 'data': formatted_data})
        except Exception as e:
            import traceback
            error_msg = f'жЯ•иѓҐе§±иі•: {str(e)}'
            return JsonResponse({'success': False, 'message': error_msg}, status=500)

    async def handle_bol_upload_post(self, request: HttpRequest) -> HttpResponse:
        '''еЃҐжИЈиЗ™жПРзЪДBOLжЦЗдїґдЄЛиљљ'''
        fleet_number = request.POST.get("fleet_number")
        customerInfo = request.POST.get("arm_pickup_data")
        notes = ""
        pickup_number = ""
        shipment_batch_number = ""

        # е¶ВжЮЬеЬ®зХМйЭҐиЊУеЕ•дЇЖпЉМе∞±зФ®зХМйЭҐжЈїеК†еРОзЪДеАЉ
        if customerInfo and customerInfo != "[]":
            customer_info = json.loads(customerInfo)
            arm_pickup = [
                [
                    "container",
                    "destination",
                    "mark",
                    "pallet",
                    "pcs",
                    "carrier",
                    "pickup",
                ]
            ]
            for row in customer_info:
                # жККжПРиіІжЧґйЧідњЃжФєж†ЉеЉП
                pickup_time = row.get('appointment_datetime', '').strip()
                s_time = pickup_time.split("T")[0]
                dt = datetime.strptime(s_time, "%Y-%m-%d")
                new_string = dt.strftime("%m-%d")

                destination_raw = row.get('zipcode', '').strip()
                destination = re.sub(r"[\u4e00-\u9fff]", " ", destination_raw)
                if not shipment_batch_number:
                    shipment_batch_number = (
                        row.get("shipment_batch_number__shipment_batch_number")
                        or row.get("shipment_batch_number")
                        or ""
                    ).strip()
                # жЛЉжО•е§Зж≥®
                note = row.get('shipment_batch_number__note') or row.get('note') or ''
                notes += note
                arm_pickup.append(
                    [
                        row.get('container_number', '').strip(),
                        destination,
                        row.get('shipping_mark', '').strip(),
                        row.get('total_pallet', '').strip(),
                        row.get('total_pcs', '').strip(),
                        row.get('carrier', '').strip(),
                        s_time,
                    ]
                )

        else:  # ж≤°жЬЙе∞±дїОжХ∞жНЃеЇУжЯ•
            arm_pickup = await sync_to_async(list)(
                Pallet.objects.select_related(
                    "container_number__container_number",
                    "shipment_batch_number__fleet_number",
                )
                .filter(shipment_batch_number__fleet_number__fleet_number=fleet_number)
                .values(
                    "container_number__container_number",
                    "destination",
                    "shipping_mark",
                    "shipment_batch_number__shipment_batch_number",
                    "shipment_batch_number__fleet_number__fleet_type",
                    "shipment_batch_number__fleet_number__carrier",
                    "shipment_batch_number__fleet_number__appointment_datetime",
                    "shipment_batch_number__fleet_number__pickup_number",  # жПРеПЦpickup_number
                    "shipment_batch_number__note",
                )
                .annotate(
                    total_pcs=Count("pcs", distinct=True),
                    total_pallet=Count("pallet_id", distinct=True),
                )
            )
            if arm_pickup:
                new_list = []
                for p in arm_pickup:
                    # дњЭе≠Шpickup_numberпЉИдїОжХ∞жНЃеЇУжПРеПЦпЉЙ
                    pickup_number = p["shipment_batch_number__fleet_number__pickup_number"] or ""
                    if not shipment_batch_number:
                        shipment_batch_number = (
                            p.get("shipment_batch_number__shipment_batch_number") or ""
                        )
                    p_time = p["shipment_batch_number__fleet_number__appointment_datetime"]

                    # жПРеПЦеєігАБжЬИгАБжЧ•
                    year = p_time.year
                    month = p_time.month
                    day = p_time.day
                    p_time = f"{year}-{month}-{day}"
                    destination = re.sub(r"[\u4e00-\u9fff]", " ", p["destination"])
                    new_list.append(
                        [
                            p["container_number__container_number"],
                            destination,
                            p["shipping_mark"],
                            p["total_pallet"],
                            p["total_pcs"],
                            p["shipment_batch_number__fleet_number__carrier"],
                            p_time,
                        ]
                    )
                    notes += p["shipment_batch_number__note"] or ""  # жЛЉжО•е§Зж≥®
                arm_pickup = [
                                 [
                                     "container",
                                     "destination",
                                     "mark",
                                     "pallet",
                                     "pcs",
                                     "carrier",
                                     "pickup",
                                 ]
                             ] + new_list
            else:
                raise ValueError("жЯЬе≠РжЬ™жЛЖжЯЬпЉМиѓЈж†ЄеЃЮ")
            s_time = arm_pickup[1][-1]
            dt = datetime.strptime(s_time, "%Y-%m-%d")
            new_string = dt.strftime("%m-%d")

        # BOLйЬАи¶БеЬ®еРОйЭҐеК†дЄАдЄ™жЛ£иіІеНХ
        df = pd.DataFrame(arm_pickup[1:], columns=arm_pickup[0])

        # жЈїеК†жНҐи°МеЗљжХ∞
        def wrap_text(text, max_length=11):
            """е∞ЖжЦЗжЬђжМЙжЬАе§ІйХњеЇ¶жНҐи°М"""
            if not isinstance(text, str):
                text = str(text)

            if len(text) <= max_length:
                return text

            # жМЙжЬАе§ІйХњеЇ¶еИЖеЙ≤жЦЗжЬђ
            wrapped_lines = []
            for i in range(0, len(text), max_length):
                wrapped_lines.append(text[i:i + max_length])
            return '\n'.join(wrapped_lines)

        # еѓєDataFrameеЇФзФ®жНҐи°Ме§ДзРЖ
        df_wrapped = df.applymap(wrap_text)

        files = request.FILES.getlist("files")
        if files:
            system_name = platform.system()
            zh_font_path = None

            # вЬЕ жМЙз≥їзїЯз±їеЮЛиЃЊзљЃйїШиЃ§иЈѓеЊД
            if system_name == "Windows":
                zh_font_path = "C:/Windows/Fonts/msyh.ttc"  # еЊЃиљѓйЫЕйїС
            else:  # Linux
                # Linux йАЪеЄЄзФ® Noto жИЦжАЭжЇРйїСдљУе≠ЧдљУ
                possible_fonts = [
                    "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
                    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
                    "/usr/share/fonts/truetype/arphic/uming.ttc",  # е§ЗзФ®
                    "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",  # жЦЗж≥Йй©њеЊЃз±≥йїС
                ]
                for path in possible_fonts:
                    if os.path.exists(path):
                        zh_font_path = path
                        break

            # вЬЕ ж£АжЯ•е≠ЧдљУжЦЗдїґжШѓеР¶е≠ШеЬ®пЉМеР¶еИЩйААеЫЮйїШиЃ§иЛ±жЦЗе≠ЧдљУ
            if zh_font_path and os.path.exists(zh_font_path):
                zh_font = fm.FontProperties(fname=zh_font_path)
                plt.rcParams["font.family"] = zh_font.get_name()
            else:
                plt.rcParams["font.family"] = "DejaVu Sans"

            plt.rcParams["axes.unicode_minus"] = False  # йШ≤ж≠ҐиіЯеПЈдє±з†Б

            for file in files:
                # иЃЊзљЃйАЪзФ®е≠ЧдљУйБњеЕНи≠¶еСК
                # plt.rcParams['font.family'] = ['sans-serif']
                # plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial', 'Helvetica']

                # дњЭжМБеОЯжЭ•зЪДA4е∞ЇеѓЄ
                fig, ax = plt.subplots(figsize=(10.4, 8.5))
                #ax.axis("tight")
                ax.axis("off")
                # з®НеЊЃеЗПе∞Пй°ґйГ®иЊєиЈЭпЉМдЄЇж†ЗйҐШзХЩеЗЇдЄАзВєз©ЇйЧі
                fig.subplots_adjust(top=1.45)  # дїО1.5еЊЃи∞ГеИ∞1.45

                # еЬ®и°®ж†ЉдЄКжЦєжЈїеК†ж†ЗйҐШ
                ax.text(
                    0.5,
                    0.93,
                    "Pickup List",
                    fontsize=14,
                    fontweight="bold",
                    ha="center",
                    va="top",
                    transform=ax.transAxes,
                )

                # еЬ®ж†ЗйҐШдЄЛжЦєжЈїеК†Pickup Number
                ax.text(
                    0.5,
                    0.89,
                    f"Pickup Number: {pickup_number}",
                    fontsize=11,
                    ha="center",
                    va="top",
                    transform=ax.transAxes,
                )
                def get_line_count(text):
                    return str(text).count("\n") + 1

                row_line_counts = [
                    max(get_line_count(cell) for cell in row)
                    for row in df_wrapped.values
                ]
                max_line_count = max(row_line_counts) if row_line_counts else 1
                EXTRA_PADDING = 0.003 * max_line_count
                BASE_ROW_HEIGHT = 0.028
                HEADER_HEIGHT = 0.05
                
                # жХ∞жНЃи°МжАїйЂШеЇ¶
                data_height = sum(
                    BASE_ROW_HEIGHT * line_count + EXTRA_PADDING
                    for line_count in row_line_counts
                )

                # и°®е§ійЂШеЇ¶
                total_table_height = HEADER_HEIGHT + data_height
                # еИЫеїЇи°®ж†Љ - дњЭжМБеОЯжЭ•зЪДдљНзљЃеТМиЃЊзљЃ
                TABLE_TOP_Y = 0.85  # и°®ж†Љй°ґйГ®еЫЇеЃЪеЬ®ж†ЗйҐШдЄЛжЦє
                table_y = TABLE_TOP_Y - total_table_height

                the_table = ax.table(
                    cellText=df_wrapped.values,
                    colLabels=df_wrapped.columns,
                    cellLoc="center",
                    bbox=[0.1, table_y, 0.8, total_table_height],
                )

                

                # иЃЊзљЃи°®ж†Љж†ЈеЉП - дњЭжМБеОЯжЭ•зЪДиЃЊзљЃпЉМеП™еҐЮеК†и°МйЂШ
                for (row, col), cell in the_table.get_celld().items():
                    cell.set_fontsize(10)
                    cell.set_text_props(wrap=True)

                    if row == 0:
                        cell.set_height(HEADER_HEIGHT)
                    else:
                        line_count = row_line_counts[row - 1]
                        cell.set_height(
                            BASE_ROW_HEIGHT * line_count + EXTRA_PADDING
                        )

                    # еИЧеЃљдњЭжМБдљ†еОЯжЭ•зЪДйАїиЊС
                    if col in (0, 1, 2):
                        cell.set_width(0.15)
                    elif col in (3, 4):
                        cell.set_width(0.06)
                    else:
                        cell.set_width(0.12)
                
                # ========= 8пЄПвГ£ иЃ°зЃЧи°®ж†ЉеЇХйГ®дљНзљЃ =========
                renderer = fig.canvas.get_renderer()
                table_bbox = the_table.get_window_extent(renderer=renderer)
                table_bbox = table_bbox.transformed(ax.transAxes.inverted())
                table_bottom = table_bbox.y0
                
                # ========= 9пЄПвГ£ Notes =========
                notes_y = table_y - 0.04
                ax.text(
                    0.05,
                    notes_y,
                    f"Notes: {notes}",
                    fontsize=10,
                    ha="left",
                    va="top",
                    transform=ax.transAxes,
                )

                pickup_y = notes_y - 0.03
                ax.text(
                    0.05,
                    pickup_y,
                    f"pickup_number: {pickup_number}",
                    fontsize=10,
                    ha="left",
                    va="top",
                    transform=ax.transAxes,
                )

                # ========= рЯФЯ дњЭе≠Ши°®ж†Љ PDF =========
                buf_table = io.BytesIO()
                fig.savefig(buf_table, format="pdf", bbox_inches="tight")
                plt.close(fig)
                buf_table.seek(0)

                # ========= 1пЄПвГ£1пЄПвГ£ еРИеєґеОЯ PDF =========
                merger = PdfMerger()
                merger.append(PdfReader(io.BytesIO(file.read())))
                merger.append(PdfReader(buf_table))

                output_buf = io.BytesIO()
                merger.write(output_buf)
                output_buf.seek(0)

                file_name = file.name

        response = HttpResponse(output_buf.getvalue(), content_type="application/octet-stream")
        def sanitize_filename_component(value: str) -> str:
            value_str = str(value or "").strip()
            value_str = re.sub(r"\s+", "", value_str)
            value_str = re.sub(r'[\\/:*?"<>|]+', "_", value_str)
            value_str = re.sub(r"\++", "+", value_str)
            value_str = re.sub(r"[^A-Za-z0-9._+\-]+", "_", value_str)
            return value_str

        container_no = ""
        destination_name = ""
        shipping_mark = ""
        if isinstance(arm_pickup, list) and len(arm_pickup) > 1 and isinstance(arm_pickup[1], list):
            container_no = arm_pickup[1][0] if len(arm_pickup[1]) > 0 else ""
            destination_name = arm_pickup[1][1] if len(arm_pickup[1]) > 1 else ""
            shipping_mark = arm_pickup[1][2] if len(arm_pickup[1]) > 2 else ""

        filename_parts = [
            sanitize_filename_component(new_string),
            sanitize_filename_component(container_no),
            sanitize_filename_component(destination_name),
            sanitize_filename_component(shipping_mark),
            "BOL",
            sanitize_filename_component(shipment_batch_number),
            "LTL",
        ]
        filename_parts = [p for p in filename_parts if p]
        output_filename = "+".join(filename_parts) + ".pdf"
        response["Content-Disposition"] = (
            f'attachment; filename="{output_filename}"'
        )
        response["X-Content-Type-Options"] = "nosniff"
        return response
    
    async def export_ltl_bol(self, request: HttpRequest) -> HttpResponse:
        '''еѓЉеЗЇLTL BOL'''
        fleet_number = request.POST.get("fleet_number")
        arm_pickup_data = request.POST.get("arm_pickup_data")
        warehouse = request.POST.get("warehouse")
        contact_flag = False  # и°®з§ЇеЬ∞еЭАж†Пз©ЇеЗЇжЭ•пЉМеЃҐжЬНжЙЛеК®PдЄКеОї
        contact = {}
        arm_pickup_groups = []

        def safe_int(value, default: int = 0) -> int:
            if value is None:
                return default
            if isinstance(value, int):
                return value
            value_str = str(value).strip()
            if value_str == "" or value_str.lower() == "none":
                return default
            try:
                return int(float(value_str))
            except Exception:
                return default

        def parse_contact_from_address(address: str) -> dict:
            address = re.sub("[\u4e00-\u9fff]", " ", address)
            address = re.sub(r"\uFF0C", ",", address)
            parts = [p.strip() for p in address.split(";")]
            return {
                "company": parts[0] if len(parts) > 0 else "",
                "Road": parts[1] if len(parts) > 1 else "",
                "city": parts[2] if len(parts) > 2 else "",
                "name": parts[3] if len(parts) > 3 else "",
                "phone": parts[4] if len(parts) > 4 else "",
            }

        def format_two_per_line(values: List[str]) -> str:
            cleaned = []
            seen = set()
            for v in values:
                v_str = str(v).strip()
                if not v_str or v_str.lower() == "none":
                    continue
                if v_str in seen:
                    continue
                seen.add(v_str)
                cleaned.append(escape(v_str))

            lines = []
            for i in range(0, len(cleaned), 2):
                lines.append(", ".join(cleaned[i : i + 2]))
            return mark_safe("<br>".join(lines)) if lines else ""

        def extract_pickup_attachments(rows: List[dict]) -> Tuple[List[dict], List[bytes]]:
            attachments = []
            pdf_attachments: List[bytes] = []
            seen = set()
            for r in rows:
                image_value = r.get("pickup_image", "")
                if isinstance(image_value, str):
                    image_value_str = image_value.strip()
                    if (
                        image_value_str
                        and image_value_str.startswith("data:image/")
                        and ";base64," in image_value_str
                        and len(image_value_str) <= 3_000_000
                        and image_value_str not in seen
                    ):
                        seen.add(image_value_str)
                        attachments.append({"kind": "image", "src": image_value_str})

                file_value = r.get("pickup_file_content", "")
                if isinstance(file_value, str):
                    file_value_str = file_value.strip()
                    if file_value_str:
                        if file_value_str.startswith("data:application/pdf;base64,"):
                            base64_part = file_value_str.split(",", 1)[1] if "," in file_value_str else ""
                            if base64_part and len(base64_part) <= 20_000_000:
                                try:
                                    pdf_bytes = base64.b64decode(base64_part, validate=False)
                                    if not pdf_bytes.startswith(b"%PDF"):
                                        continue
                                    dedupe_key = f"PDF::{hash(pdf_bytes)}"
                                    if dedupe_key not in seen:
                                        seen.add(dedupe_key)
                                        pdf_attachments.append(pdf_bytes)
                                    continue
                                except Exception:
                                    continue
                            continue
                        if len(file_value_str) > 30_000:
                            file_value_str = file_value_str[:30_000] + "\n...[truncated]"
                        safe_html = escape(file_value_str).replace("\r\n", "\n").replace(
                            "\n", "<br>"
                        )
                        dedupe_key = f"FILE::{safe_html}"
                        if dedupe_key in seen:
                            continue
                        seen.add(dedupe_key)
                        attachments.append(
                            {
                                "kind": "file",
                                "html": mark_safe(safe_html),
                            }
                        )
            return attachments, pdf_attachments

        if arm_pickup_data and arm_pickup_data != "[]":
            loaded = json.loads(arm_pickup_data)
            raw_groups = []
            if isinstance(loaded, dict) and isinstance(loaded.get("data"), list):
                raw_groups = [loaded.get("data", [])]
            elif isinstance(loaded, list):
                if loaded and all(isinstance(x, list) for x in loaded):
                    raw_groups = loaded
                elif loaded and all(isinstance(x, dict) for x in loaded) and any(
                    isinstance(x.get("data"), list) for x in loaded
                ):
                    raw_groups = [
                        x.get("data", []) if isinstance(x, dict) else [] for x in loaded
                    ]
                else:
                    raw_groups = [loaded]

            arm_pickup = []
            for group in raw_groups:
                if not isinstance(group, list):
                    continue
                group_contact_flag = False
                group_contact = {}
                group_rows = []
                for row in group:
                    if not isinstance(row, dict):
                        continue
                    address = row.get("address", "")
                    row_contact_flag = False
                    row_contact = {}
                    if address:
                        row_contact_flag = True
                        row_contact = parse_contact_from_address(address)
                        group_contact_flag = True
                        group_contact = row_contact
                        contact_flag = True
                        contact = group_contact
                    row_dict = {
                        "container_number__container_number": str(
                            row.get("container_number__container_number", "")
                        ).strip(),
                        "destination": str(row.get("destination", "")).strip(),
                        "shipping_mark": str(row.get("shipping_mark", "")).strip(),
                        "shipment_batch_number__ARM_PRO": str(
                            row.get("shipment_batch_number__ARM_PRO", "")
                        ).strip(),
                        "total_pallet": safe_int(row.get("total_pallet", 0)),
                        "total_pcs": safe_int(row.get("total_pcs", 0)),
                        "shipment_batch_number__fleet_number__carrier": str(
                            row.get("shipment_batch_number__fleet_number__carrier", "")
                        ).strip(),
                        "shipment_batch_number__note": str(
                            row.get("shipment_batch_number__note", "")
                        ).strip(),
                        "slot": str(row.get("slot", "")).strip(),
                        "pickup_image": str(row.get("pickup_image", "")).strip(),
                        "pickup_image_orientation": str(
                            row.get("pickup_image_orientation", "horizontal")
                        ).strip(),
                        "pickup_file_content": str(
                            row.get("pickup_file_content", "")
                        ),
                    }
                    for possible_key in (
                        "arm_pickup_group",
                        "pickup_group",
                        "group_id",
                        "group",
                        "pickup_number",
                        "bol_group",
                    ):
                        possible_value = row.get(possible_key)
                        if possible_value is None:
                            continue
                        possible_value_str = str(possible_value).strip()
                        if possible_value_str:
                            row_dict[possible_key] = possible_value_str
                    if row_contact_flag:
                        row_dict["__contact_flag"] = True
                        row_dict["__contact"] = row_contact
                    group_rows.append(row_dict)
                    arm_pickup.append(row_dict)
                if group_rows:
                    if len(raw_groups) == 1:
                        grouped = {}
                        grouped_order = []
                        for r in group_rows:
                            group_id = ""
                            # дЉШеЕИдљњзФ® shipment_batch_number ињЫи°МеИЖзїД
                            shipment_batch_number = r.get("shipment_batch_number__shipment_batch_number", "")
                            if shipment_batch_number and shipment_batch_number != "None":
                                key = f"BATCH::{shipment_batch_number}"
                            else:
                                # е¶ВжЮЬж≤°жЬЙ shipment_batch_numberпЉМдљњзФ®еОЯжЬЙзЪДеИЖзїДйАїиЊС
                                for possible_key in (
                                    "arm_pickup_group",
                                    "pickup_group",
                                    "group_id",
                                    "group",
                                    "pickup_number",
                                    "bol_group",
                                ):
                                    if r.get(possible_key):
                                        group_id = str(r.get(possible_key)).strip()
                                        break
                                arm_pro_value = r.get("shipment_batch_number__ARM_PRO", "")
                                if group_id:
                                    key = f"GROUP::{group_id}"
                                elif arm_pro_value and arm_pro_value != "None":
                                    key = f"ARM_PRO::{arm_pro_value}"
                                else:
                                    key = f"CN_DEST::{r.get('container_number__container_number','')}|{r.get('destination','')}"
                            
                            if key not in grouped:
                                grouped[key] = []
                                grouped_order.append(key)
                            grouped[key].append(r)

                        if len(grouped_order) > 1:
                            for key in grouped_order:
                                rows = grouped[key]
                                first = rows[0] if rows else {}
                                arm_pickup_groups.append(
                                    {
                                        "rows": rows,
                                        "contact_flag": bool(
                                            first.get("__contact_flag", False)
                                        ),
                                        "contact": first.get("__contact", {}),
                                    }
                                )
                        else:
                            arm_pickup_groups.append(
                                {
                                    "rows": group_rows,
                                    "contact_flag": group_contact_flag,
                                    "contact": group_contact,
                                }
                            )
                    else:
                        arm_pickup_groups.append(
                            {
                                "rows": group_rows,
                                "contact_flag": group_contact_flag,
                                "contact": group_contact,
                            }
                        )
        else:  # ж≤°жЬЙе∞±дїОжХ∞жНЃеЇУжЯ•
            arm_pickup = await sync_to_async(list)(
                Pallet.objects.select_related(
                    "container_number__container_number",
                    "shipment_batch_number__fleet_number",
                )
                .filter(shipment_batch_number__fleet_number__fleet_number=fleet_number)
                .values(
                    "container_number__container_number",
                    "shipment_batch_number__shipment_appointment",
                    "shipment_batch_number__ARM_PRO",
                    "shipment_batch_number__fleet_number__carrier",
                    "shipment_batch_number__fleet_number__appointment_datetime",
                    "shipment_batch_number__fleet_number__fleet_type",
                    "destination",
                    "shipping_mark",
                    "shipment_batch_number__note",
                    "slot",
                    "shipment_batch_number__shipment_batch_number",  # жЈїеК† shipment_batch_number
                )
                .annotate(
                    total_pcs=Sum("pcs"),
                    total_pallet=Count("pallet_id", distinct=True),
                    total_weight=Sum("weight_lbs"),
                    total_cbm=Sum("cbm"),
                )
            )
            if arm_pickup:
                # жМЙзЕІ shipment_batch_number ињЫи°МеИЖзїД
                grouped_by_batch = {}
                for item in arm_pickup:
                    batch_number = item.get("shipment_batch_number__shipment_batch_number")
                    if batch_number not in grouped_by_batch:
                        grouped_by_batch[batch_number] = []
                    grouped_by_batch[batch_number].append(item)
                
                # жЮДеїЇ arm_pickup_groups
                arm_pickup_groups = []
                for batch_number, rows in grouped_by_batch.items():
                    arm_pickup_groups.append({
                        "rows": rows, 
                        "contact_flag": contact_flag, 
                        "contact": contact
                    })

        fleet = await sync_to_async(Fleet.objects.get)(fleet_number=fleet_number)
        pickup_time_str = fleet.appointment_datetime
        pickup_time = pickup_time_str.strftime("%Y-%m-%d")
        pallet = 0
        pcs = 0
        shipping_mark = ""
        notes = set()
        arm_pro = ""
        carrier = ""
        container_number = ""
        destination = ""
        for arm in arm_pickup:
            arm_pro = arm.get("shipment_batch_number__ARM_PRO", "")
            carrier = arm.get("shipment_batch_number__fleet_number__carrier", "")
            pallet += safe_int(arm.get("total_pallet", 0))
            pcs += safe_int(arm.get("total_pcs", 0))
            container_number = arm.get("container_number__container_number", "")
            destination = arm.get("destination", "")
            shipping_mark += arm.get("shipping_mark", "")
            notes.add(arm.get("shipment_batch_number__note", ""))
            marks = arm.get("shipping_mark", "")
            new_marks = ""
            if marks:
                array = marks.split(",")
                if len(array) > 1:
                    parts = []
                    for i in range(0, len(array)):
                        part = ",".join(array[i : i + 1])
                        parts.append(part)
                    new_marks = "\n".join(parts)
                else:
                    new_marks = marks
            arm["shipping_mark"] = new_marks
        group_container_number = format_two_per_line(
            [a.get("container_number__container_number", "") for a in arm_pickup]
        )
        all_marks = []
        for a in arm_pickup:
            marks_value = a.get("shipping_mark", "")
            for part in re.split(r"[\n,]+", str(marks_value)):
                part_str = part.strip()
                if part_str:
                    all_marks.append(part_str)
        group_shipping_mark = format_two_per_line(all_marks)
        pickup_attachments, pickup_pdfs = extract_pickup_attachments(arm_pickup)
        notes_str = "<br>".join(filter(None, notes))
        barcode_type = "code128"
        barcode_class = barcode.get_barcode_class(barcode_type)

        def generate_barcode_base64(content: str) -> str:
            my_barcode = barcode_class(content, writer=ImageWriter())
            buffer = io.BytesIO()
            my_barcode.write(buffer, options={"dpi": 600})
            buffer.seek(0)
            image = Image.open(buffer)
            width, height = image.size
            new_height = int(height * 0.7)
            cropped_image = image.crop((0, 0, width, new_height))
            new_buffer = io.BytesIO()
            cropped_image.save(new_buffer, format="PNG")
            return base64.b64encode(new_buffer.getvalue()).decode("utf-8")

        def wrap_text_by_length(text, length=20):
            """
            жѓПйЪФжМЗеЃЪйХњеЇ¶еЉЇеИґжНҐи°М
            """
            if not text or not str(text).strip():
                return mark_safe("  ")
            
            text = str(text).strip()
            # жѓП length дЄ™е≠Чзђ¶еИЗеИЖдЄАжђ°
            lines = [text[i:i+length] for i in range(0, len(text), length)]
            
            # е¶ВжЮЬжШѓзФЯжИР PDF/жЙУеН∞пЉМеїЇиЃЃзФ® <br>
            return mark_safe("<br>".join([escape(line) for line in lines]))
        is_multi_arm_pickup = len(arm_pickup) > 1
  
        if is_multi_arm_pickup:
            # дЄАжПРе§ЪеНЄзФ®еНХзЛђж®°жЭњ

            # еҐЮеК†и£ЕиіІй°ЇеЇП
            count = len(arm_pickup)
            for i, arm in enumerate(arm_pickup):
                if i == count - 1:
                    arm['multi_drop'] = 'outside'
                else:
                    arm['multi_drop'] = f'inside {i + 1}'
            # еҐЮеК†е§Зж≥®еНХзЛђи°®ж†Љ
            notes_table = []
            for item in arm_pickup:
                raw_note = item.get("shipment_batch_number__note") or ""
                # и∞ГзФ®жКШи°МеЗљжХ∞пЉМеБЗиЃЊ note еИЧеЬ® A4 зЇЄеЃљеЇ¶дЄЛе§ІзЇ¶ 20-25 дЄ™дЄ≠жЦЗе≠Чзђ¶жНҐи°М
                formatted_note = wrap_text_by_length(raw_note, length=20) 
                
                notes_table.append({
                    "container_no": item.get("container_number__container_number") or "",
                    "destination": item.get("destination") or "",
                    "note": formatted_note  # еЈ≤зїПжШѓеЄ¶жЬЙ <br> зЪД mark_safe е≠Чзђ¶дЄ≤
                })
            bol_pages = []
            all_pickup_pdfs: List[bytes] = []
            for group in arm_pickup_groups:
                group_rows = group.get("rows", [])
                if not group_rows:
                    continue
                group_arm_pro = ""
                group_carrier = ""
                group_container_number = ""
                group_destination = ""
                group_pallet = 0
                group_pcs = 0
                group_container_numbers = []
                group_marks = []
                for row in group_rows:
                    if not group_arm_pro:
                        group_arm_pro = row.get("shipment_batch_number__ARM_PRO", "")
                    if not group_carrier:
                        group_carrier = row.get(
                            "shipment_batch_number__fleet_number__carrier", ""
                        )
                    if not group_container_number:
                        group_container_number = row.get(
                            "container_number__container_number", ""
                        )
                    if not group_destination:
                        group_destination = row.get("destination", "")
                    group_pallet += safe_int(row.get("total_pallet", 0))
                    group_pcs += safe_int(row.get("total_pcs", 0))
                    group_container_numbers.append(
                        row.get("container_number__container_number", "")
                    )
                    row_marks_value = row.get("shipping_mark", "")
                    for part in re.split(r"[\n,]+", str(row_marks_value)):
                        part_str = part.strip()
                        if part_str:
                            group_marks.append(part_str)

                if not group_arm_pro or group_arm_pro == "None":
                    if 'иЗ™жПР' in group_destination:
                        group_destination = 'client pickup'
                    barcode_content = f"{group_container_number}|{group_destination}"
                else:
                    barcode_content = f"{group_arm_pro}"
                if not group_carrier or group_carrier == "None" or group_carrier == "":
                    group_carrier = "nocarrier"
                group_attachments, group_pdfs = extract_pickup_attachments(group_rows)
                all_pickup_pdfs.extend(group_pdfs)
                bol_pages.append(
                    {
                        "warehouse": warehouse,
                        "arm_pro": group_arm_pro,
                        "carrier": group_carrier,
                        "pallet": group_pallet,
                        "pcs": group_pcs,
                        "container_number": format_two_per_line(group_container_numbers),
                        "shipping_mark": format_two_per_line(group_marks),
                        "pickup_attachments": group_attachments,
                        "pickup_has_pdf": bool(group_pdfs),
                        "barcode": generate_barcode_base64(barcode_content),
                        "contact": group.get("contact", {}),
                        "contact_flag": group.get("contact_flag", False),
                        "pickup_time": pickup_time,
                    }
                )

            context = {
                "warehouse": warehouse,
                "bol_pages": bol_pages,
                "arm_pickup": arm_pickup,
                "notes_table": notes_table,
            }
            template = get_template(self.template_ltl_bol_multi)
            html = template.render(context)
            pickup_pdfs = all_pickup_pdfs
        else:
            if arm_pro == "" or arm_pro == "None" or arm_pro is None:
                if 'иЗ™жПР' in destination:
                    destination = 'client pickup'
                barcode_content = f"{container_number}|{destination}"
            else:
                barcode_content = f"{arm_pro}"
            barcode_base64 = generate_barcode_base64(barcode_content)

            if not carrier or carrier == "None" or carrier == "":
                carrier = "nocarrier"

            context = {
                "warehouse": warehouse,
                "arm_pro": arm_pro,
                "carrier": carrier,
                "pallet": pallet,
                "pcs": pcs,
                "container_number": group_container_number,
                "shipping_mark": group_shipping_mark,
                "barcode": barcode_base64,
                "pickup_attachments": pickup_attachments,
                "pickup_has_pdf": bool(pickup_pdfs),
                "arm_pickup": arm_pickup,
                "contact": contact,
                "contact_flag": contact_flag,
                "pickup_time": pickup_time,
                "notes": notes_str,
            }
            template = get_template(self.template_ltl_bol)
            html = template.render(context)

        pdf_buffer = io.BytesIO()
        def sanitize_filename(value: str) -> str:
            """жЄЕзРЖжЦЗдїґеРНдЄ≠зЪДжНҐи°МеТМйЭЮж≥Хе≠Чзђ¶"""
            if not value:
                return ""
            # еОїжОЙжНҐи°Мзђ¶
            value = value.replace("\n", "_").replace("\r", "_")
            # жЫњжНҐ Windows дЄНеЕБиЃЄзЪДе≠Чзђ¶
            value = re.sub(r'[\\/:*?"<>|]+', "_", value)
            # йШ≤ж≠ҐжЦЗдїґеРНе§™йХњ
            return value[:100]
        safe_shipping_mark = sanitize_filename(shipping_mark)
        safe_destination = sanitize_filename(destination)
        if is_multi_arm_pickup:
            segments = []
            for group in arm_pickup_groups:
                rows = group.get("rows", [])
                if not rows:
                    continue
                containers = []
                container_seen = set()
                marks = []
                mark_seen = set()
                for r in rows:
                    cn = str(r.get("container_number__container_number", "")).strip()
                    if cn and cn.lower() != "none" and cn not in container_seen:
                        container_seen.add(cn)
                        containers.append(cn)
                    sm = r.get("shipping_mark", "")
                    for part in re.split(r"[\n,]+", str(sm)):
                        part_str = part.strip()
                        if (
                            part_str
                            and part_str.lower() != "none"
                            and part_str not in mark_seen
                        ):
                            mark_seen.add(part_str)
                            marks.append(part_str)
                container_part = sanitize_filename(containers[0]) if containers else "NA"
                mark_part = sanitize_filename(marks[0]) if marks else "NA"
                segments.append(f"{container_part}-{mark_part}")
            filename_base = "+".join(filter(None, segments))
            filename_base = sanitize_filename(filename_base)
            if not filename_base:
                safe_fleet_number = sanitize_filename(fleet_number)
                filename_base = f"{safe_fleet_number}+MULTI"
            content_disposition = f'attachment; filename="{filename_base}+BOL.pdf"'
        else:
            content_disposition = (
                f'attachment; filename="{container_number}+{safe_destination}+{safe_shipping_mark}+BOL.pdf"'
            )

        pisa_status = pisa.CreatePDF(html, dest=pdf_buffer, link_callback=link_callback)
        if pisa_status.err:
            raise ValueError(
                "Error during PDF generation: %s" % pisa_status.err,
                content_type="text/plain",
            )
        pdf_buffer.seek(0)
        base_pdf_bytes = pdf_buffer.getvalue()
        merged_bytes = base_pdf_bytes

        if pickup_pdfs:
            try:
                base_reader = PdfReader(io.BytesIO(base_pdf_bytes))
                insert_at = None
                anchor = "__ZEM_PICKUP_PDF_ANCHOR__"
                fallback_matches: List[int] = []
                for idx, page in enumerate(base_reader.pages):
                    page_text = page.extract_text() or ""
                    normalized = re.sub(r"\s+", "", page_text).upper()
                    if anchor in (page_text or ""):
                        insert_at = idx
                        break
                    if "CONTAINERS:" in normalized and "SHIPPING_MARK:" in normalized:
                        fallback_matches.append(idx)
                if insert_at is None and fallback_matches:
                    insert_at = fallback_matches[-1]

                writer = PdfWriter()
                for idx, page in enumerate(base_reader.pages):
                    if insert_at is not None and idx == insert_at:
                        for pdf_bytes in pickup_pdfs:
                            try:
                                att_reader = PdfReader(io.BytesIO(pdf_bytes))
                                for att_page in att_reader.pages:
                                    writer.add_page(att_page)
                            except Exception:
                                continue
                    writer.add_page(page)

                if insert_at is None:
                    for pdf_bytes in pickup_pdfs:
                        try:
                            att_reader = PdfReader(io.BytesIO(pdf_bytes))
                            for att_page in att_reader.pages:
                                writer.add_page(att_page)
                        except Exception:
                            continue

                out = io.BytesIO()
                writer.write(out)
                merged_bytes = out.getvalue()
            except Exception:
                merged_bytes = base_pdf_bytes

        response = HttpResponse(merged_bytes, content_type="application/pdf")
        response["Content-Disposition"] = content_disposition
        response["X-Content-Type-Options"] = "nosniff"
        return response
    
    async def handle_export_maersk_label(self, request: HttpRequest) -> HttpResponse:
        '''Maersk LabelиОЈеПЦ'''
        fleet_number = request.POST.get("fleet_number")
        e_label_type = request.POST.get("eLabelType", "Label4x6")
        
        # Get Shipment info
        shipment_data = await sync_to_async(list)(
            Shipment.objects.filter(fleet_number__fleet_number=fleet_number).values('ARM_PRO', 'destination','origin')
        )
        
        if not shipment_data:
            return JsonResponse({'success': False, 'message': 'жЬ™жЙЊеИ∞иѓ•иљ¶жђ°зЪДShipmentиЃ∞ељХ'}, status=400)
            
        shipment = shipment_data[0]
        arm_pro = shipment.get('ARM_PRO')
        origin = shipment.get('origin')
        
        if not arm_pro:
             return JsonResponse({'success': False, 'message': 'зЉЇе∞СPROеПЈз†Б (ARM_PRO)'}, status=400)
             
        # Extract zip from destination (digits only)
        szip = "".join(filter(str.isdigit, origin or ""))
        if not szip:
             return JsonResponse({'success': False, 'message': 'жЧ†ж≥ХдїОдїУеЇУжПРеПЦйВЃзЉЦ (szip)'}, status=400)
             
        # Call Maersk Gateway API
        api_url = "https://zem-maersk-gateway.kindmoss-a5050a64.eastus.azurecontainerapps.io/label"
        api_key = os.environ.get("MAERSK_API_KEY")

        params = {
            "shawb": arm_pro,
            "eLabelType": e_label_type,
            "szip": szip
        }
        headers = {
            "Content-Type": "application/json",
            "x-api-key": api_key
        }
        
        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(api_url, json=params, headers=headers) as response:
                    if response.status == 200:

                        # 1пЄПвГ£ иѓїеПЦжО•еП£ињФеЫЮеЖЕеЃє
                        text = await response.text()

                        # 2пЄПвГ£ е¶ВжЮЬжШѓиҐЂеМЕи£єзЪДе≠Чзђ¶дЄ≤пЉМеЕИеПНеЇПеИЧеМЦ
                        if text.startswith('"'):
                            text = json.loads(text)

                        try:
                            # 3пЄПвГ£ иІ£жЮР XML
                            root = ET.fromstring(text)

                            count = 0
                            for elem in root.iter():
                                count += 1
                                if count >= 20:
                                    break

                            # 4пЄПвГ£ ењљзХ• namespace еЉЇеИґжЯ•жЙЊ DataStream_Byte
                            data_node = None
                            for elem in root.iter():
                                if elem.tag.endswith("DataStream_Byte"):
                                    data_node = elem
                                    break

                            if data_node is not None and data_node.text:

                                base64_str = data_node.text.strip()

                                # 5пЄПвГ£ Base64 иІ£з†Б
                                decoded_bytes = base64.b64decode(base64_str)

                                if not decoded_bytes.startswith(b'%PDF'):
                                    print("вЪ†пЄП и≠¶еСКпЉЪжЦЗдїґе§ідЄНжШѓ %PDFпЉМеПѓиГљдЄНжШѓPDFжЦЗдїґ")

                                # 6пЄПвГ£ дњЭе≠Ши∞ГиѓХжЦЗдїґ
                                with open("debug_label.pdf", "wb") as f:
                                    f.write(decoded_bytes)

                                # 7пЄПвГ£ ињФеЫЮзїЩжµПиІИеЩ®
                                response_obj = HttpResponse(decoded_bytes, content_type='application/pdf')
                                response_obj['Content-Disposition'] = f'attachment; filename="Maersk_Label_{arm_pro}.pdf"'
                                return response_obj

                            else:
                                return HttpResponse(
                                    "<script>alert('жЬ™жЙЊеИ∞PDFжЦЗдїґеЖЕеЃє');history.back();</script>"
                                )

                        except Exception as e:
                            return HttpResponse(
                                f"<script>alert('XMLиІ£жЮРе§±иі•: {str(e)}');history.back();</script>"
                            )

                    else:
                        text = await response.text()
                        try:
                            error_json = json.loads(text)
                            error_msg = error_json.get('detail', text)
                        except:
                            error_msg = text

                        return HttpResponse(
                            f"<script>alert('APIи∞ГзФ®е§±иі•: {response.status} - {error_msg}');history.back();</script>"
                        )
        except Exception as e:
            # return JsonResponse({'success': False, 'message': f'з≥їзїЯйФЩиѓѓ: {str(e)}'}, status=500)
            return HttpResponse(
                f"<script>alert('з≥їзїЯйФЩиѓѓ: {str(e)}');history.back();</script>", 
            )

    async def handle_export_maersk_bol(self, request: HttpRequest) -> HttpResponse:
        '''Maersk BOLиОЈеПЦ'''
        fleet_number = request.POST.get("fleet_number")
        
        # Get Shipment info
        shipment_data = await sync_to_async(list)(
            Shipment.objects.filter(fleet_number__fleet_number=fleet_number).values('ARM_PRO', 'destination','origin')
        )
        
        if not shipment_data:
            return HttpResponse(
                "<script>alert('жЬ™жЙЊеИ∞иѓ•иљ¶жђ°зЪДShipmentиЃ∞ељХ');history.back();</script>", 
                content_type="text/html"
            )
            
        shipment = shipment_data[0]
        arm_pro = shipment.get('ARM_PRO')
        origin = shipment.get('origin')
        
        if not arm_pro:
             return HttpResponse(
                "<script>alert('зЉЇе∞СPROеПЈз†Б (ARM_PRO)');history.back();</script>", 
                content_type="text/html"
            )
             
        # Extract zip from destination (digits only)
        szip = "".join(filter(str.isdigit, origin or ""))
        if not szip:
             return HttpResponse(
                "<script>alert('жЧ†ж≥ХдїОдїУеЇУжПРеПЦйВЃзЉЦ (szip)');history.back();</script>", 
            )
             
        # Call Maersk Gateway API
        api_url = "https://zem-maersk-gateway.kindmoss-a5050a64.eastus.azurecontainerapps.io/bol"
        api_key = os.environ.get("MAERSK_API_KEY")
        
        params = {
            "shawb": arm_pro,
            "szip": szip
        }
        
        headers = {
            "Content-Type": "application/json",
            "x-api-key": api_key
        }
        
        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(api_url, json=params, headers=headers) as response:
                    if response.status == 200:

                        # 1пЄПвГ£ иѓїеПЦжО•еП£ињФеЫЮеЖЕеЃє
                        text = await response.text()

                        # 2пЄПвГ£ е¶ВжЮЬжШѓиҐЂеМЕи£єзЪДе≠Чзђ¶дЄ≤пЉМеЕИеПНеЇПеИЧеМЦ
                        if text.startswith('"'):
                            text = json.loads(text)

                        try:
                            # 3пЄПвГ£ иІ£жЮР XML
                            root = ET.fromstring(text)

                            count = 0
                            for elem in root.iter():
                                print(elem.tag)
                                count += 1
                                if count >= 20:
                                    break

                            # 4пЄПвГ£ ењљзХ• namespace еЉЇеИґжЯ•жЙЊ DataStream_Byte
                            data_node = None
                            for elem in root.iter():
                                if elem.tag.endswith("DataStream_Byte"):
                                    data_node = elem
                                    break

                            if data_node is not None and data_node.text:

                                base64_str = data_node.text.strip()

                                # 5пЄПвГ£ Base64 иІ£з†Б
                                decoded_bytes = base64.b64decode(base64_str)

                                if not decoded_bytes.startswith(b'%PDF'):
                                    print("вЪ†пЄП и≠¶еСКпЉЪжЦЗдїґе§ідЄНжШѓ %PDFпЉМеПѓиГљдЄНжШѓPDFжЦЗдїґ")

                                # 6пЄПвГ£ ињљеК†йЭЮMaerskжЛ£иіІеНХй°µпЉИдљњзФ® ltl_bol ж®°жЭњзЪД Pickup List йГ®еИЖпЉЙ
                                arm_pickup = await sync_to_async(list)(
                                    Pallet.objects.select_related(
                                        "container_number__container_number",
                                        "shipment_batch_number__fleet_number",
                                    )
                                    .filter(shipment_batch_number__fleet_number__fleet_number=fleet_number)
                                    .values(
                                        "container_number__container_number",
                                        "destination",
                                        "shipping_mark",
                                        "shipment_batch_number__ARM_PRO",
                                        "shipment_batch_number__fleet_number__carrier",
                                        "slot",
                                        "shipment_batch_number__note",
                                    )
                                    .annotate(
                                        total_pcs=Sum("pcs"),
                                        total_pallet=Count("pallet_id", distinct=True),
                                    )
                                )
                                notes = set()
                                for row in arm_pickup:
                                    v = row.get("shipment_batch_number__note") or ""
                                    if v:
                                        notes.add(v)
                                notes_str = "<br>".join(filter(None, notes))
                                pickup_time = datetime.now().strftime("%Y-%m-%d")
                                context = {
                                    "warehouse": origin or "",
                                    "arm_pickup": arm_pickup,
                                    "notes": notes_str,
                                    "pickup_attachments": [],
                                    "pickup_has_pdf": False,
                                    "container_number": "",
                                    "shipping_mark": "",
                                }
                                template = get_template("export_file/ltl_bol.html")
                                html = template.render(context)
                                pickup_pdf_buf = io.BytesIO()
                                pisa.CreatePDF(html, dest=pickup_pdf_buf, link_callback=link_callback)
                                pickup_pdf_buf.seek(0)
                                pickup_pdf_bytes = pickup_pdf_buf.getvalue()

                                base_reader = PdfReader(io.BytesIO(pickup_pdf_bytes))
                                pickup_pages = []
                                start_idx = None
                                for idx, page in enumerate(base_reader.pages):
                                    txt = page.extract_text() or ""
                                    if "Pickup List" in txt:
                                        start_idx = idx
                                        break
                                if start_idx is None:
                                    pickup_pages = base_reader.pages
                                else:
                                    pickup_pages = base_reader.pages[start_idx:]

                                merger = PdfMerger()
                                merger.append(PdfReader(io.BytesIO(decoded_bytes)))
                                temp_bufs = []
                                for pg in pickup_pages:
                                    out = PdfWriter()
                                    out.add_page(pg)
                                    tmp = io.BytesIO()
                                    out.write(tmp)
                                    tmp.seek(0)
                                    temp_bufs.append(tmp)
                                for b in temp_bufs:
                                    merger.append(PdfReader(b))
                                out_all = io.BytesIO()
                                merger.write(out_all)
                                out_all.seek(0)

                                response_obj = HttpResponse(out_all.getvalue(), content_type='application/pdf')
                                response_obj['Content-Disposition'] = f'attachment; filename="Maersk_BOL_{arm_pro}.pdf"'
                                return response_obj

                            else:
                                return HttpResponse(
                                    "<script>alert('жЬ™жЙЊеИ∞PDFжЦЗдїґеЖЕеЃє');history.back();</script>"
                                )

                        except Exception as e:
                            return HttpResponse(
                                f"<script>alert('XMLиІ£жЮРе§±иі•: {str(e)}');history.back();</script>"
                            )

                    else:
                        text = await response.text()
                        try:
                            error_json = json.loads(text)
                            error_msg = error_json.get('detail', text)
                        except:
                            error_msg = text

                        return HttpResponse(
                            f"<script>alert('APIи∞ГзФ®е§±иі•: {response.status} - {error_msg}');history.back();</script>"
                        )
        except Exception as e:
            return HttpResponse(
                f"<script>alert('з≥їзїЯйФЩиѓѓ: {str(e)}');history.back();</script>", 
            )

    async def export_ltl_label(self, request: HttpRequest) -> HttpResponse:
        '''жЦ∞еКЯиГљLTLзЪДLABELжЦЗдїґдЄЛиљљ'''
        fleet_number = request.POST.get("fleet_number")
        arm_pickup_data = request.POST.get("arm_pickup_data")
        contact_flag = False  # и°®з§ЇеЬ∞еЭАж†Пз©ЇеЗЇжЭ•пЉМеЃҐжЬНжЙЛеК®PдЄКеОї
        contact = ""
        if arm_pickup_data and arm_pickup_data != "[]":
            arm_pickup_data = json.loads(arm_pickup_data)
            for row in arm_pickup_data:
                address = row.get("address", "")
                if address:
                    contact_flag = True
                    address = re.sub("[\u4e00-\u9fff]", " ", address)
                    address = re.sub(r"\uFF0C", ",", address)
                    parts = [p.strip() for p in address.split(";")]
                    contact = {
                        "company": parts[0] if len(parts) > 0 else "",
                        "Road":    parts[1] if len(parts) > 1 else "",
                        "city":    parts[2] if len(parts) > 2 else "",
                        "name":    parts[3] if len(parts) > 3 else "",
                        "phone":   parts[4] if len(parts) > 4 else "",
                    }
        
        arm_pickup = await sync_to_async(list)(
            Pallet.objects.select_related(
                "container_number__container_number",
                "shipment_batch_number__fleet_number",
            )
            .filter(shipment_batch_number__fleet_number__fleet_number=fleet_number)
            .values(
                "container_number__container_number",
                "shipment_batch_number__shipment_appointment",
                "shipment_batch_number__ARM_PRO",
                "shipment_batch_number__fleet_number__carrier",
                "shipment_batch_number__fleet_number__appointment_datetime",
                "destination",
                "shipping_mark",
            )
            .annotate(
                total_pcs=Sum("pcs"),
                total_pallet=Count("pallet_id", distinct=True),
                total_weight=Sum("weight_lbs"),
                total_cbm=Sum("cbm"),
            )
        )
        if not arm_pickup:
            raise ValueError('иѓ•иљ¶жђ°дЄЛжЬ™жЯ•еИ∞жЭње≠РиЃ∞ељХпЉБ')
        pallets = 0
        for arm in arm_pickup:
            if not arm["shipment_batch_number__shipment_appointment"]:
                raise ValueError(f'{fleet_number}зЪДзЇ¶ж≤°жЬЙжЧґйЧі!')
            arm_pro = arm["shipment_batch_number__ARM_PRO"]
            carrier = arm["shipment_batch_number__fleet_number__carrier"]
            pickup_time = arm["shipment_batch_number__shipment_appointment"]
            container_number = arm["container_number__container_number"]
            destination = arm["destination"]
            shipping_mark = arm["shipping_mark"]
            pallets += arm["total_pallet"]
        pickup_time_str = str(pickup_time)
        date_str = datetime.strptime(pickup_time_str[:19], "%Y-%m-%d %H:%M:%S")
        pickup_time = date_str.strftime("%Y-%m-%d")

        # зФЯжИРжЭ°ељҐз†Б
        barcode_type = "code128"
        barcode_class = barcode.get_barcode_class(barcode_type)
        if arm_pro == "" or arm_pro == "None" or arm_pro == None:
            barcode_content = f"{container_number}|{shipping_mark}"
        else:
            barcode_content = f"{arm_pro}"
        my_barcode = barcode_class(
            barcode_content, writer=ImageWriter()
        )  # е∞ЖжЭ°ељҐз†БиљђжНҐдЄЇеЫЊеГПељҐеЉП
        buffer = io.BytesIO()  # еИЫеїЇзЉУеЖ≤еМЇ
        my_barcode.write(buffer, options={"dpi": 600})  # зЉУеЖ≤еМЇе≠ШеВ®еЫЊеГП
        buffer.seek(0)
        image = Image.open(buffer)
        width, height = image.size
        new_height = int(height * 0.7)
        cropped_image = image.crop((0, 0, width, new_height))
        new_buffer = io.BytesIO()
        cropped_image.save(new_buffer, format="PNG")

        barcode_base64 = base64.b64encode(new_buffer.getvalue()).decode("utf-8")
        data = [
            {
                "warehouse": request.POST.get("warehouse"),
                "arm_pro": arm_pro,
                "barcode": barcode_base64,
                "carrier": carrier,
                "contact": contact,
                "contact_flag": contact_flag,
                "fraction": f"{i + 1}/{pallets}",
            }
            for i in range(pallets)
        ]
        context = {"data": data}
        template = get_template(self.template_ltl_label)
        html = template.render(context)
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="{pickup_time}+{container_number}+{destination}+{shipping_mark}+LABEL.pdf"'
        )
        pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
        if pisa_status.err:
            raise ValueError(
                "Error during PDF generation: %s" % pisa_status.err,
                content_type="text/plain",
            )
        return response

    
    async def handle_bol_fleet_post(self, request: HttpRequest) -> HttpResponse:
        #еЗЖе§ЗеПВжХ∞
        pickup_list = request.POST.get("pickup_list_data")
        april_pickup_list = None

        if pickup_list:
            outer_list = json.loads(pickup_list)
            april_pickup_list = []
            for group in outer_list:
                if not isinstance(group, dict):
                    continue

                april_pickup_list.append({
                    "container_number__container_number": group.get("container_number", "").strip(),
                    "destination": group.get("destination", "").strip(),
                    "cbm": group.get("cbm", "").strip(),
                    "pallet": group.get("pallet_count", "").strip(),
                    "loading_note": group.get("loading_note", "").strip(),
                })
        mutable_post = request.POST.copy()
        fleet_number = request.POST.get("fleet_number")
        mutable_post["customerInfo"] = None
        mutable_post["pickupList"] = None

        request.POST = mutable_post
        fm = FleetManagement()
        fleet = await sync_to_async(Fleet.objects.get)(fleet_number=fleet_number)
            
        #if fleet.fleet_type == 'FTL': 
        shipment = await sync_to_async(list)(Shipment.objects.filter(fleet_number=fleet))
        if len(shipment) > 1:
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                for s in shipment:
                    s_number = s.shipment_batch_number
                    mutable_post["shipment_batch_number"] = s_number
                    pdf_response = await fm.handle_export_bol_post(request, april_pickup_list)
                    zip_file.writestr(f"BOL_{s_number}.pdf", pdf_response.content)
            response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
            response["Content-Disposition"] = 'attachment; filename="orders.zip"'
            zip_buffer.close()
            return response
        else:
            mutable_post["shipment_batch_number"] = shipment[0].shipment_batch_number
        return await fm.handle_export_bol_post(request, april_pickup_list)
        # else:
        #     raise ValueError('еЗЇеЇУз±їеЮЛеЉВеЄЄпЉБ')
        
    async def handle_bol_post(self, request: HttpRequest) -> Any:
        fm = FleetManagement()
        mutable_post = request.POST.copy()
        mutable_post["customerInfo"] = None
        mutable_post["pickupList"] = None
        warehouse = mutable_post["warehouse"]
        context = {}
        for key, code in self.warehouse_options.items():
            if key in warehouse:
                mutable_post["warehouse"] = code

        appointment_id = request.POST.get("appointment_id")
        mutable_post["appointment_id"] = appointment_id
        shipment = await sync_to_async(
            lambda: Shipment.objects.select_related("fleet_number").get(
                appointment_id=appointment_id
            )
        )()
        
        if shipment.shipment_type == 'еЃҐжИЈиЗ™жПР':
            raise ValueError("иѓ•йҐДзЇ¶жЙєжђ°йҐДзЇ¶з±їеЮЛдЄЇеЃҐжИЈиЗ™жПРпЉМдЄНжФѓжМБеЃҐжПРзЪДBOLдЄЛиљљпЉБ")
        if shipment.fleet_number:
            mutable_post["fleet_number"] = shipment.fleet_number
        else:
            raise ValueError("иѓ•йҐДзЇ¶жЙєжђ°е∞ЪжЬ™жОТиљ¶")
        request.POST = mutable_post

        shipment_batch_number = shipment.shipment_batch_number
        fleet_number = shipment.fleet_number.fleet_number

        mutable_post = request.POST.copy()
        mutable_post['shipment_batch_number'] = shipment_batch_number
        mutable_post['fleet_number'] = fleet_number
        request.POST = mutable_post
        return await fm.handle_export_bol_post(request)
    
    async def generate_unique_batch_number(self,destination):
        """зФЯжИРеФѓдЄАзЪДshipment_batch_number"""
        current_time = datetime.now()

        for i in range(10):
            batch_id = (
                destination
                + current_time.strftime("%m%d%H%M%S")
                + str(uuid.uuid4())[:2].upper()
            )
            batch_number = batch_id.replace(" ", "").replace("/", "-").upper()      
            exists = await sync_to_async(
                Shipment.objects.filter(shipment_batch_number=batch_number).exists
            )()

            if not exists:
                return batch_number
        raise ValueError('жЙєжђ°еПЈеІЛзїИйЗНе§Н')
    
    async def handle_modify_intelligent_po_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        all_suggestions_raw = request.POST.get('all_suggestions')
        if all_suggestions_raw:
            all_suggestions = json.loads(all_suggestions_raw)

        suggestion_data_raw = request.POST.get('suggestion_data')
        if suggestion_data_raw:
            suggestion_data = json.loads(suggestion_data_raw)
            suggestion_id = suggestion_data['suggestion_id']
            cargos = suggestion_data.get('cargos', [])
            intelligent_cargos = suggestion_data.get('intelligent_cargos', [])

        selected_ids_raw = request.POST.get("selected_ids")
        selected_ids = json.loads(selected_ids_raw) if selected_ids_raw else []
        
        selected_cargos = [c for c in intelligent_cargos if c['unique_id'] in selected_ids]
        cargos.extend(selected_cargos)

         # === жЫіжЦ∞ primary_group зЪДзїЯиЃ°жХ∞жНЃ ===
        total_pallets = sum(c.get('total_n_pallet_act', 0) or c.get('total_n_pallet_est', 0) for c in cargos)
        total_cbm = sum(c.get('total_cbm', 0) for c in cargos)

        primary_group = suggestion_data.get('primary_group', {})
        if primary_group:
            # жЫіжЦ∞дЄїзїДзЪДжЭњжХ∞еТМCBM
            primary_group['total_pallets'] = total_pallets
            primary_group['total_cbm'] = total_cbm
        new_intelligent_cargos = [c for c in intelligent_cargos if c['unique_id'] not in selected_ids]

        suggestion_data['cargos'] = cargos
        suggestion_data['intelligent_cargos'] = new_intelligent_cargos

        # жЫњжНҐжОЙ all_suggestions дЄ≠еѓєеЇФй°є
        for i, s in enumerate(all_suggestions):
            if s['suggestion_id'] == suggestion_id:
                all_suggestions[i] = suggestion_data
                break
        return await self.handle_td_unshipment_post(request,{},all_suggestions)

    async def handle_fleet_add_pallet_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        context = {}
        appointment_id = request.POST.get("add_pallet_ISA")
        plt_ids = request.POST.getlist("plt_ids")
        actual_shipped_pallet = request.POST.getlist("actual_shipped_pallet")
        shipped_pallet_ids = []
        
        for plt_id, p_shipped in zip(plt_ids, actual_shipped_pallet):
            # жЄЕзРЖжХ∞жНЃпЉЪзІїйЩ§з©Їе≠Чзђ¶дЄ≤еТМNone
            if not plt_id or not p_shipped:
                continue

            p_shipped_int = int(float(p_shipped))
            
            # еИЖеЙ≤plt_idsеєґжЄЕзРЖз©ЇеАЉ
            plt_id_list = [pid.strip() for pid in plt_id.split(',') if pid.strip()]
            
            if not plt_id_list:
                continue
                
            # еПЦеЙНp_shipped_intдЄ™еЕГзі†
            shipped_count = min(p_shipped_int, len(plt_id_list))  # йШ≤ж≠ҐзіҐеЉХиґКзХМ
            shipped_pallet_ids += plt_id_list[:shipped_count]
        pallets = await sync_to_async(list)(
            Pallet.objects.select_related("container_number").filter(
                id__in=shipped_pallet_ids
            )
        )
        total_weight, total_cbm, total_pcs = 0.0, 0.0, 0
        for plt in pallets:
            total_weight += plt.weight_lbs
            total_pcs += plt.pcs
            total_cbm += plt.cbm
        # жЯ•жЙЊиѓ•еЗЇеЇУжЙєжђ°,е∞ЖйЗНйЗПз≠Йдњ°жБѓеК†еИ∞еЗЇеЇУжЙєжђ°дЄК
        try:
            shipment = await Shipment.objects.select_related("fleet_number").aget(appointment_id=appointment_id)
        except ObjectDoesNotExist:
            context.update({"error_messages": f"{appointment_id}йҐДзЇ¶еПЈжЙЊдЄНеИ∞"})
            return await self.handle_fleet_schedule_post(request,context)
        fleet = shipment.fleet_number
        fleet.total_weight += total_weight
        fleet.total_pcs += total_pcs
        fleet.total_cbm += total_cbm
        fleet.total_pallet += len(shipped_pallet_ids)
        await sync_to_async(fleet.save)()
        # жЯ•жЙЊиѓ•еЗЇеЇУжЙєжђ°дЄЛзЪДзЇ¶пЉМжККеК†е°ЮзЪДжЯЬе≠РжЭњжХ∞еК†еИ∞еРМдЄАдЄ™зЫЃзЪДеЬ∞зЪДзЇ¶
        
        plt_to_update = []
        for pallet in pallets:
            pallet.shipment_batch_number = shipment
            plt_to_update.append(pallet)
        if plt_to_update:
            result = await sync_to_async(bulk_update_with_history)(
                plt_to_update,
                Pallet,
                fields=["shipment_batch_number"],
            )
        if 'error_messages' not in context:
            context.update({"success_messages": f"{appointment_id}еК†е°ЮжИРеКЯпЉБ"})
        page = request.POST.get("page")
        if page == "arm_appointment":
            return await self.handle_unscheduled_pos_post(request,context)
        else:
            return await self.handle_fleet_schedule_post(request,context)
    
    async def handle_fl_notified_customer_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        page = request.POST.get("page")
        fleet_number = request.POST.get("fleet_number")
        context = {}
        if not bool(fleet_number) or not fleet_number or 'None' in fleet_number:
            context.update({
                'error_messages':'fleet_numberдЄЇз©ЇпЉБ',
            })
            if page == "arm_appointment":
                return await self.handle_unscheduled_pos_post(request,context)
            else:
                return await self.handle_td_shipment_post(request, context)
        try:
            fleet = await sync_to_async(Fleet.objects.get)(
                fleet_number=fleet_number
            )
            shipments = await sync_to_async(list)(
                Shipment.objects.filter(fleet_number=fleet)
            )
            if not shipments:
                context.update({
                    'error_messages': f"жЙЊдЄНеИ∞иљ¶жђ°еПЈдЄЇ {fleet_number} зЫЄеЕ≥иБФзЪДйҐДзЇ¶жЙєжђ°иЃ∞ељХ",
                })
            else:
                for shipment in shipments:
                    shipment.is_notified_customer = True
                    await sync_to_async(shipment.save)()
                context = {'success_messages': f"{fleet_number}йАЪзЯ•еЃҐжИЈжИРеКЯпЉБ"}
        except Fleet.DoesNotExist:
            context.update({
                'error_messages': f"жЙЊдЄНеИ∞иљ¶жђ°еПЈдЄЇ {fleet_number} зЪДиЃ∞ељХ",
            })
        except Exception as e:
            context.update({
                'error_messages': f"йАЪзЯ•еЃҐжИЈе§±иі•: {str(e)}",
            })
        
        if page == "arm_appointment":
            return await self.handle_unscheduled_pos_post(request,context)
        else:
            return await self.handle_td_shipment_post(request, context)
        
    async def handle_sp_notified_customer_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        appointment_id = request.POST.get("appointment_id")
        context = {}
        if not bool(appointment_id) or not appointment_id or 'None' in appointment_id:
            context.update({
                'error_messages':'ISAдЄЇз©ЇпЉБ',
            })
            return await self.handle_td_shipment_post(request, context)
        try:
            shipment = await sync_to_async(Shipment.objects.get)(
                appointment_id=appointment_id
            )
            shipment.is_notified_customer = True
            await sync_to_async(shipment.save)()
            context = {'success_messages': f"{appointment_id}йАЪзЯ•еЃҐжИЈжИРеКЯпЉБ"}
        except Shipment.DoesNotExist:
            context.update({
                'error_messages': f"жЙЊдЄНеИ∞йҐДзЇ¶еПЈдЄЇ {appointment_id} зЪДиЃ∞ељХ",
            })
        except Exception as e:
            context.update({
                'error_messages': f"йАЪзЯ•еЃҐжИЈе§±иі•: {str(e)}",
            })
        page = request.POST.get("page")
        if page == "arm_appointment":
            return await self.handle_unscheduled_pos_post(request,context)
        else:
            return await self.handle_td_shipment_post(request, context)
    
    async def handle_search_addable_po_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("warehouse")
        destination = request.POST.get("destination")
        appointment_id = request.POST.get("appointment_id")
        context = {}
        if not bool(appointment_id) or not appointment_id or 'None' in appointment_id:
            context.update({
                'error_messages':'ISAдЄЇз©ЇпЉБ',
                "show_add_po_inventory_modal": False,
            })
        criteria_p = models.Q(
            (
                models.Q(container_number__orders__order_type="иљђињР")
                | models.Q(container_number__orders__order_type="иљђињРзїДеРИ")
            ),
            shipment_batch_number__isnull=True,
            container_number__orders__created_at__gte="2024-09-01",
        )
        pl_criteria = criteria_p & models.Q(
            container_number__orders__offload_id__offload_at__isnull=True,
            container_number__orders__retrieval_id__retrieval_destination_precise=warehouse,
            #destination=destination,
            delivery_type='public',
        )
        plt_criteria = criteria_p & models.Q(
            container_number__orders__offload_id__offload_at__isnull=False,
            location__startswith=warehouse,
            #destination=destination,
            delivery_type='public',
        )
        packing_list_not_scheduled = await self._get_packing_list(
            request.user,pl_criteria, plt_criteria
        )
        # ињЗжї§жОЙжѓПзїДдЄ≠ is_pass=False зЪДеЖЕеЃє
        filtered_packing_list = []
        for po in packing_list_not_scheduled:
            if getattr(po, 'is_pass', True):  # е¶ВжЮЬ is_pass дЄЇ True жИЦиАЕдЄНе≠ШеЬ®иѓ•е±ЮжАІпЉМеИЩдњЭзХЩ
                filtered_packing_list.append(po)

        packing_list_not_scheduled = filtered_packing_list
        
        context.update({
            "warehouse": warehouse,
            "destination": destination,
            "appointment_id": request.POST.get("appointment_id"),
            "packing_list_not_scheduled": packing_list_not_scheduled,
            "active_tab": request.POST.get("active_tab"),       
        })
        if 'show_add_po_inventory_modal' not in context:
            context.update({"show_add_po_inventory_modal": True})# вЖР жОІеИґжШѓеР¶зЫіжО•еЉєеЗЇвАЬжЈїеК†POвАЭеЉєз™Ч
        page = request.POST.get("page")
        if page == "arm_appointment":
            return await self.handle_unscheduled_pos_post(request,context)
        else:
            return await self.handle_td_shipment_post(request, context)
    
    async def handle_add_pallet_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("warehouse")
        destination = request.POST.get("destination")
        active_tab = request.POST.get("active_tab")
        step = request.POST.get("step")
        criteria_plt = models.Q(
            models.Q(
                shipment_batch_number__fleet_number__fleet_number__isnull=True
            ) | models.Q(
                shipment_batch_number__fleet_number__fleet_number__isnull=False,
                shipment_batch_number__fleet_number__departured_at__isnull=True
            ),
            location=warehouse,
            destination=destination,
            container_number__orders__offload_id__offload_at__isnull=False,
        )
        plt_unshipped = await self._get_packing_list(
            request.user,
            models.Q(
                container_number__orders__offload_id__offload_at__isnull=False,
            )& models.Q(pk=0),
            criteria_plt
        )
        context = {
            "warehouse": warehouse,
            "destination": destination,
            "plt_unshipped": plt_unshipped,
            "step": step,  # вЖР еЙНзЂѓйЭ†ињЩдЄ™еИ§жЦ≠и¶БдЄНи¶БеЉєз™Ч
            "active_tab": active_tab,          # вЖР зФ®жЭ•жОІеИґеЙНзЂѓжЙУеЉАеУ™дЄ™ж†Зз≠Њй°µ
            "show_add_po_modal": True,   # вЖР жОІеИґжШѓеР¶зЫіжО•еЉєеЗЇвАЬжЈїеК†POвАЭеЉєз™Ч
            "add_po_title": "еК†е°Ю",
            "add_pallet_ISA": request.POST.get('add_pallet_ISA')
        }
        page = request.POST.get("page")
        if page == "arm_appointment":
            return await self.handle_unscheduled_pos_post(request,context)
        else:
            return await self.handle_fleet_schedule_post(request, context)

    async def handle_one_fleet_departure_post(
        self, request: HttpRequest, context: str | None = None
    ) -> tuple[str, dict[str, Any]]:
        if not context:
            context = {}
        request.POST = request.POST.copy()
        if "plt_ids" in request.POST:
            try:
                raw_plt = request.POST.get("plt_ids")
                plt_list = json.loads(raw_plt)  # => [["5022,5023,5024,5025,5026"], ["3343,3344"]]
                flattened_plt = [",".join(inner) if isinstance(inner, list) else str(inner) for inner in plt_list]
                request.POST.setlist("plt_ids", flattened_plt)
            except Exception as e:
                raise ValueError("вЪ†пЄП plt_ids иІ£жЮРйФЩиѓѓ:", e)

        for key in ["scheduled_pallet", "actual_shipped_pallet"]:
            if key in request.POST:
                raw_value = request.POST.get(key)  # дЊЛе¶В '5,2'
                parts = [v.strip() for v in raw_value.split(",") if v.strip()]
                request.POST.setlist(key, parts)
        
        fleet_num_str = request.POST.get("fleet_number")
        if fleet_num_str:
            batch_numbers = await sync_to_async(
                lambda: list(
                    Shipment.objects
                    .filter(fleet_number__fleet_number=fleet_num_str)  # иЈ®и°®жЯ•иѓҐ Fleet ж®°еЮЛзЪД fleet_number е≠ЧжЃµ
                    .exclude(shipment_batch_number__isnull=True)       # жОТйЩ§з©ЇеАЉ
                    .values_list('shipment_batch_number', flat=True)
                    .distinct()
                )
            )()
        else:
            raise ValueError('зЉЇе∞Сfleet_numberеАЉ')

        request.POST = request.POST.copy()
        request.POST.setlist('batch_number', batch_numbers)
        fm = FleetManagement()
        context_new = await fm.handle_fleet_departure_post(request,'post_nsop')
        context.update(context_new)
        page = request.POST.get("page")

        # з°ЃиЃ§еЗЇеЇУ-еИЖжСКжИРжЬђ
        fleet = await sync_to_async(Fleet.objects.get)(fleet_number=fleet_num_str)
        if fleet.fleet_cost is not None and fleet.fleet_cost > 0:
            await fm.insert_fleet_shipment_pallet_fleet_cost(
                request, fleet_num_str, fleet.fleet_cost
            )
        if page == "arm_appointment":
            return await self.handle_unscheduled_pos_post(request,context)
        elif page == "ltl_readyShip":
            return await self.handle_ltl_unscheduled_pos_post(request)       
        else:
            return await self.handle_fleet_schedule_post(request,context)         
    
    async def handle_export_virtual_fleet_pos_post(
        self, request: HttpRequest
    ) ->  HttpResponse:
        cargo_ids_str = request.POST.get("cargo_ids", "").strip()
        pl_ids = [
            int(i.strip()) for i in cargo_ids_str.split(",") if i.strip().isdigit()
        ]
        
        plt_ids_str = request.POST.get("plt_ids", "").strip()
        plt_ids = [
            int(i.strip()) for i in plt_ids_str.split(",") if i.strip().isdigit()
        ]
        if not pl_ids and not plt_ids:
            raise ValueError("ж≤°жЬЙиОЈеПЦеИ∞дїїдљХ ID")
        
        all_data = []

        if plt_ids:
            pallet_data = await sync_to_async(
                lambda: list(
                    Pallet.objects.select_related("container_number")
                    .filter(id__in=plt_ids,is_dropped_pallet=False)
                    .values(
                        "id",
                        "shipping_mark",
                        "fba_id",
                        "ref_id",
                        "address",
                        "zipcode",
                        "container_number__container_number",
                        "destination",
                        "cbm",
                        "pcs",
                        "PO_ID"
                    )
                    .annotate(
                        total_pcs=Sum("pcs"),
                        total_n_pallet_act=Count("id"),
                        total_cbm=Sum("cbm", output_field=FloatField()),
                        label=Value('ACT')
                    )
                    .distinct()
                    .order_by("destination", "container_number__container_number")
                )
            )()
            
            additional_packinglist_ids, unmatched_pallet_data = await self.find_packinglist_ids_by_pallet_data(pallet_data)
            # еРИеєґpackinglist IDs
            all_packinglist_ids = list(set(pl_ids + additional_packinglist_ids))
            # е§ДзРЖжЬ™еМєйЕНзЪДpalletжХ∞жНЃпЉИжЈїеК†ж†ЗиЃ∞пЉЙ
            if unmatched_pallet_data:
                for item in unmatched_pallet_data:
                    item = dict(item)
                    ref_ids = str(item["ref_id"]).split(",") if item["ref_id"] else [""]
                    
                    for ref_id in ref_ids:
                        new_row = item.copy()
                        new_row["ref_id"] = ref_id.strip()
                        new_row["check_id"] = "жЬ™жЙЊеИ∞еѓєеЇФPOиЃ∞ељХпЉМиѓЈжЙЛеК®е§ДзРЖ"
                        new_row["is_unmatched"] = True
                        all_data.append(new_row)
        else:
            all_packinglist_ids = pl_ids
        
        
        packinglist_data = await sync_to_async(
            lambda: list(
                PackingList.objects.select_related("container_number")
                .filter(id__in=all_packinglist_ids)
                .values(
                    "id",
                    "shipping_mark",
                    "fba_id",
                    "ref_id",
                    "address",
                    "zipcode",
                    "container_number__container_number",
                    "destination",
                    "cbm"
                )
                .annotate(
                    total_pcs=Sum("pcs"),
                    total_n_pallet_est=Sum("cbm", output_field=FloatField()) / 2,
                    total_cbm=Sum("cbm", output_field=FloatField()),
                    label=Value("EST"),
                )
                .distinct()
                .order_by("destination", "container_number__container_number")
            )
        )()
        pl_ids_list = [pl["id"] for pl in packinglist_data]
        check_map = await sync_to_async(
            lambda: {
                p.packing_list_id: p.id
                for p in PoCheckEtaSeven.objects.filter(packing_list_id__in=pl_ids_list)
            }
        )()

        # е±ХеЉАжХ∞жНЃпЉЪе∞Жref_idжМЙйАЧеПЈеИЖеЙ≤жИРе§Ъи°М
        for item in packinglist_data:
            check_id = check_map.get(item["id"])
            if check_id:
                item["check_id"] = check_id  # еМєйЕНеИ∞е∞±жШЊз§ЇID
            else:
                item["check_id"] = "жЬ™жЙЊеИ∞ж†°й™МиЃ∞ељХ"  # еМєйЕНдЄНеИ∞е∞±жШЊз§ЇжПРз§Ї
            
        
        # еРИеєґжХ∞жНЃ
        all_data += packinglist_data
        df = pd.DataFrame.from_records(all_data)

        df["total_n_pallet_est"] = df["total_cbm"] / 2
        def get_est_pallet(n):
            if n < 1:
                return 1
            elif n % 1 >= 0.45:
                return int(n // 1 + 1)
            else:
                return int(n // 1)

        df["total_n_pallet_est"] = df["total_n_pallet_est"].apply(get_est_pallet)
        df["is_valid"] = None
        df["is_est"] = df["label"] == "EST"
        df["Pallet Count"] = df.apply(
            lambda row: row["total_n_pallet_est"] if row["is_est"] else max(1, row.get("total_n_pallet_act", 1)),
            axis=1
        ).astype(int)

        keep = [
            "shipping_mark",
            "container_number__container_number",
            "fba_id",
            "ref_id",
            "total_pcs",
            "Pallet Count",
            "label",
            "is_valid",
            "check_id",
            "total_cbm",
            "destination",
        ]
        df = df[keep].rename(
            {
                "fba_id": "PRO",
                "container_number__container_number": "BOL",
                "ref_id": "PO List (use , as separator) *",
                "total_pcs": "Carton Count",
                "total_cbm": "Total CBM",
                "destination": "Destination",
                "check_id": "Check Result",
            },
            axis=1,
        )
        
        # еѓЉеЗЇ CSV

        if len(df) == 0:
            raise ValueError('ж≤°жЬЙжХ∞жНЃ',len(df))
        # е¶ВжЮЬеП™жЬЙдЄАдЄ™ DestinationпЉМдњЭжМБеОЯжЭ•ињФеЫЮеНХ CSV
        grouped_by_dest = {}
        for _, row in df.iterrows():
            dest = row["Destination"]
            if dest not in grouped_by_dest:
                grouped_by_dest[dest] = []
            grouped_by_dest[dest].append(row.to_dict())
        
        # е¶ВжЮЬеП™жЬЙдЄАдЄ™ DestinationпЉМињФеЫЮеНХ CSV
        if len(grouped_by_dest) == 1:
            dest_name = list(grouped_by_dest.keys())[0]
            df_single = pd.DataFrame(grouped_by_dest[dest_name])
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = f"attachment; filename=PO_{dest_name}.csv"
            df_single.to_csv(path_or_buf=response, index=False)
            return response
        
        # е§ЪдЄ™ Destination жЙУеМЕ zip
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for dest, rows in grouped_by_dest.items():
                df_dest = pd.DataFrame.from_records(rows)
                
                csv_buffer = io.BytesIO()
                df_dest.to_csv(csv_buffer, index=False, encoding='utf-8-sig')
                csv_buffer.seek(0)
            
                safe_dest = "".join(c for c in dest if c.isalnum() or c in (' ', '-', '_')).rstrip()
                filename = f"{safe_dest}.csv" if safe_dest else f"destination_{hash(dest)}.csv"
                
                zf.writestr(filename, csv_buffer.getvalue())

        zip_buffer.seek(0)
        zip_bytes = zip_buffer.getvalue()

        response = HttpResponse(zip_bytes, content_type="application/zip")
        response["Content-Disposition"] = "attachment; filename=PO_virtual_fleet.zip"
        response["Content-Length"] = len(zip_bytes)
        return response
    
    async def handle_update_fleet_info(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        """жЫіжЦ∞ fleet еЯЇз°Адњ°жБѓ"""
        context = {}

        # иОЈеПЦиѓЈж±ВдЄ≠зЪДе≠ЧжЃµ
        fleet_number = request.POST.get("fleet_number", "").strip()
        warehouse = request.POST.get("warehouse", "").strip()
        carrier = request.POST.get("carrier", "").strip()
        third_party_address = request.POST.get("third_party_address", "").strip()
        pickup_number = request.POST.get("pickup_number", "").strip()
        license_plate = request.POST.get("license_plate", "").strip()
        motor_carrier_number = request.POST.get("motor_carrier_number", "").strip()
        dot_number = request.POST.get("dot_number", "").strip()
        appointment_datetime_str = request.POST.get("appointment_datetime", "").strip()
        note = request.POST.get("note", "").strip()

        # жЯ•жЙЊ Fleet
        fleet = await sync_to_async(lambda: Fleet.objects.filter(fleet_number=fleet_number).first())()
        if not fleet:
            context["error_messages"] = f"Fleet {fleet_number} дЄНе≠ШеЬ®"
            return await self.handle_td_shipment_post(request, context)

        # иІ£жЮРжЧґйЧіе≠Чзђ¶дЄ≤
        appointment_datetime = None
        if appointment_datetime_str:
            try:
                # ж†ЉеЉПдЊЛе¶В 2025-10-11T16:09
                appointment_datetime = datetime.strptime(appointment_datetime_str, "%Y-%m-%dT%H:%M")
            except Exception as e:
                context["error_messages"] = f"жЧґйЧіж†ЉеЉПйФЩиѓѓ: {appointment_datetime_str} ({e})"
                return await self.handle_td_shipment_post(request, context)

        fleet.origin = warehouse or fleet.origin
        fleet.carrier = carrier or fleet.carrier
        fleet.third_party_address = third_party_address or fleet.third_party_address
        fleet.pickup_number = pickup_number or fleet.pickup_number
        fleet.license_plate = license_plate or fleet.license_plate
        fleet.motor_carrier_number = motor_carrier_number or fleet.motor_carrier_number
        fleet.dot_number = dot_number or fleet.dot_number
        fleet.note = note or fleet.note
        fleet.is_virtual = False
        if appointment_datetime:
            fleet.appointment_datetime = appointment_datetime

        await sync_to_async(fleet.save)()

        context["message"] = f"Fleet {fleet_number} дњ°жБѓеЈ≤жИРеКЯжЫіжЦ∞гАВ"
        context["fleet_number"] = fleet_number
        return await self.handle_td_shipment_post(request, context)


    async def handle_multi_group_booking(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        """е§ДзРЖе§ЪзїДйҐДзЇ¶еЗЇеЇУ"""
        booking_data_str = request.POST.get('booking_data')
        warehouse = request.POST.get('warehouse')
        context = {}
        
        if not booking_data_str:
            context.update({"error_messages": "ж≤°жЬЙжФґеИ∞йҐДзЇ¶жХ∞жНЃ"})
            return await self.handle_td_shipment_post(request, context)
        
        try:
            booking_data = json.loads(booking_data_str)
        except json.JSONDecodeError:
            context.update({"error_messages": "йҐДзЇ¶жХ∞жНЃж†ЉеЉПйФЩиѓѓ"})
            return await self.handle_td_shipment_post(request, context)
        
        if not isinstance(booking_data, list) or len(booking_data) == 0:
            context.update({"error_messages": "йҐДзЇ¶жХ∞жНЃдЄЇз©ЇжИЦж†ЉеЉПдЄНж≠£з°Ѓ"})
            return await self.handle_td_shipment_post(request, context)
        
        # й™МиѓБжѓПзїДзЪДidпЉМе¶ВжЮЬжЬЙйЧЃйҐШпЉМзЫіжО•жК•йФЩзїЩеЙНзЂѓ
        appointment_ids = [group.get('appointment_id', '') for group in booking_data if group.get('appointment_id')]
        
        # жЙєйЗПй™МиѓБжЙАжЬЙйҐДзЇ¶еПЈ
        error_messages = await self._batch_validate_appointments(appointment_ids, booking_data)
        
        # е¶ВжЮЬжЬЙй™МиѓБйФЩиѓѓпЉМзЫіжО•ињФеЫЮ
        if error_messages:
            context.update({"error_messages": mark_safe("<br>".join(error_messages))})
            return await self.handle_td_shipment_post(request, context)
        
        # е≠ШеВ®е§ДзРЖзїУжЮЬ
        success_groups = []
        failed_groups = []
        success_appointment_ids = []
        # дЄЇжѓПдЄ™е§ІзїДеИЖеИЂе§ДзРЖйҐДзЇ¶
        for group_index, group_data in enumerate(booking_data, 1):
            # еЗЖе§Зи∞ГзФ® handle_appointment_post жЙАйЬАзЪДеПВжХ∞
            cargo_ids = group_data.get('cargo_ids_checked', '')
            plt_ids = group_data.get('plt_ids_checked', '')
            destination = group_data.get('destination', '')
            appointment_id = group_data.get('appointment_id', '')
            shipment_cargo_id = group_data.get('shipment_cargo_id', '')
            shipment_type = group_data.get('shipment_type', '')
            shipment_account = group_data.get('shipment_account', '')
            shipment_appointment = group_data.get('shipment_appointment', '')
            load_type = group_data.get('load_type', '')
            origin = group_data.get('origin', '')
            if not origin:
                origin = warehouse
            note = group_data.get('note', '')
            suggestion_id = group_data.get('suggestion_id')
            pickup_time = group_data.get('pickup_time')
            pickup_number = group_data.get('pickupNumber')
            result = await self._process_single_group_booking(request, suggestion_id, cargo_ids, plt_ids, destination, appointment_id,shipment_cargo_id,shipment_type,
                                                              shipment_account,shipment_appointment,load_type,origin,note,pickup_time,pickup_number)
            
            if result['success']:
                success_groups.append({
                    'suggestion_id': result.get('suggestion_id'),
                    'appointment_id': result.get('appointment_id'),
                    'batch_number': result.get('shipment_batch_number')
                })
                appointment_id = int(str(appointment_id).strip())
                success_appointment_ids.append(appointment_id)
            else:
                failed_groups.append({
                    'suggestion_id': result.get('suggestion_id'),
                    'appointment_id': result.get('appointment_id'),
                    'batch_number': result.get('shipment_batch_number'),
                    'error': result.get('error', 'жЬ™зЯ•йФЩиѓѓ')
                })
                       
        # жЮДеїЇињФеЫЮжґИжБѓ
        messages = []
        if success_groups:
            success_msg = mark_safe(f"жИРеКЯйҐДзЇ¶ {len(success_groups)} дЄ™е§ІзїД: <br>")
            success_msg += ", ".join([f"(жЙєжђ°еПЈпЉЪ{group['batch_number']},йҐДзЇ¶еПЈ:{group['appointment_id']})" for group in success_groups])
            messages.append(mark_safe(success_msg + "<br>"))
            
        if failed_groups:
            failed_msg = mark_safe(f"йҐДзЇ¶е§±иі• {len(failed_groups)} дЄ™е§ІзїД: <br>")
            failed_details = []
            for group in failed_groups:
                detail = f"(жЙєжђ°еПЈпЉЪ{group['batch_number']},йҐДзЇ¶еПЈ:{group['appointment_id']}) - {group['error']}"
                failed_details.append(detail)
            failed_msg += "; ".join(failed_details)
            if not success_msg:
                success_msg = None
            messages.append(mark_safe(success_msg + "<br>"))
        
        # е≠ШеВ®жИРеКЯеИЫеїЇзЪДshipment IDsпЉМжЦєдЊњеРОзї≠зЇ¶иљ¶дљњзФ®
        if success_appointment_ids:
            # жОТиљ¶
            fleet_number = await self._add_appointments_to_fleet(success_appointment_ids)
            success_msg = f"жИРеКЯжОТиљ¶пЉМиљ¶жђ°еПЈжШѓ {fleet_number}"
            messages.append(success_msg)
        if messages:
            context.update({"success_messages": mark_safe("<br>".join(messages))})
            
        template_name = request.POST.get('template_name')
        if template_name and template_name == "unshipment":
            return await self.handle_td_unshipment_post(request,context)
        return await self.handle_td_shipment_post(request, context)

    async def _batch_validate_appointments(self, appointment_ids: list, booking_data: list) -> list[str]:
        """жЙєйЗПй™МиѓБжЙАжЬЙйҐДзЇ¶еПЈ"""
        error_messages = []
        
        if not appointment_ids:
            return error_messages
        
        try:
            # жЙєйЗПжЯ•иѓҐжЙАжЬЙйҐДзЇ¶еПЈ
            existed_appointments = await sync_to_async(list)(
                Shipment.objects.filter(appointment_id__in=appointment_ids)
            )
            
            # еИЫеїЇжШ†е∞Де≠ЧеЕЄдЊњдЇОжЯ•жЙЊ
            appointment_dict = {appt.appointment_id: appt for appt in existed_appointments}
            
            # й™МиѓБжѓПдЄ™йҐДзЇ¶еПЈ
            for appointment_id in appointment_ids:
                existed_appointment = appointment_dict.get(appointment_id)
                if not existed_appointment:
                    continue  # йҐДзЇ¶еПЈдЄНе≠ШеЬ®пЉМй™МиѓБйАЪињЗ
                
                # жЙЊеИ∞еѓєеЇФзЪДзїДжХ∞жНЃ
                group_data = next((group for group in booking_data if group.get('appointment_id') == appointment_id), None)
                if not group_data:
                    continue
                    
                destination = group_data.get('destination', '')
                
                # й™МиѓБйҐДзЇ¶зКґжАБ
                if existed_appointment.in_use:
                    error_messages.append(f"ISA {appointment_id} еЈ≤зїПиҐЂдљњзФ®!")
                elif existed_appointment.is_canceled:
                    error_messages.append(f"ISA {appointment_id} еЈ≤зїПеПЦжґИ!")
                elif (existed_appointment.shipment_appointment.replace(tzinfo=pytz.UTC) < timezone.now()):
                    error_messages.append(f"ISA {appointment_id} йҐДзЇ¶жЧґйЧіе∞ПдЇОељУеЙНжЧґйЧіпЉМеЈ≤ињЗжЬЯ!")
                elif (existed_appointment.destination.replace("Walmart", "").replace("WALMART", "").replace("-", "").upper() != 
                    destination.replace("Walmart", "").replace("WALMART", "").replace("-", "").upper()):
                    error_messages.append(f"ISA {appointment_id} зЩїиЃ∞зЪДзЫЃзЪДеЬ∞жШѓ {existed_appointment.destination}пЉМж≠§жђ°зЩїиЃ∞зЪДзЫЃзЪДеЬ∞жШѓ {destination}!")
                    
        except Exception as e:
            error_messages.append(f"й™МиѓБйҐДзЇ¶еПЈжЧґеЗЇйФЩ: {str(e)}")
        
        return error_messages
    
    async def _add_appointments_to_fleet(self,appointment_ids):
        current_time = datetime.now()
        fleet_number = (
            "F"
            + current_time.strftime("%m%d%H%M%S")
            + str(uuid.uuid4())[:2].upper()
        )
        shipment_info = await sync_to_async(list)(
            Shipment.objects.filter(appointment_id__in=appointment_ids)
            .values('id', 'shipment_type', 'origin')
            .distinct()
        )
        shipment_ids = [item['id'] for item in shipment_info]
        shipment_types = list(set(item['shipment_type'] for item in shipment_info))
        origins = list(set(item['origin'] for item in shipment_info))

        total_weight, total_cbm, total_pcs, total_pallet = 0.0, 0.0, 0, 0
        #иЃ∞ељХжАїжХ∞
        if shipment_ids:
            # иОЈеПЦжЙАжЬЙPalletиЃ∞ељХ
            pallet_records = await sync_to_async(list)(
                Pallet.objects.filter(
                    shipment_batch_number_id__in=shipment_ids,
                    container_number__orders__offload_id__offload_at__isnull=False
                )
            )
            
            # иОЈеПЦжЙАжЬЙPackingListиЃ∞ељХ  
            packinglist_records = await sync_to_async(list)(
                PackingList.objects.filter(
                    shipment_batch_number_id__in=shipment_ids,
                    container_number__orders__offload_id__offload_at__isnull=True
                )
            )
            
            # ж±ЗжАїPalletжХ∞жНЃ
            for p in pallet_records:
                total_weight += p.weight_lbs or 0
                total_cbm += p.cbm or 0
                total_pcs += p.pcs or 0
                total_pallet += 1
            
            # ж±ЗжАїPackingListжХ∞жНЃ
            for pl in packinglist_records:
                total_weight += pl.total_weight_lbs or 0
                total_cbm += pl.cbm or 0
                total_pcs += pl.pcs or 0
                total_pallet += round(pl.cbm /1.8)
        
        fleet_data = {
            "fleet_number": fleet_number,
            "fleet_type": shipment_types[0] if shipment_types else None,
            "origin": origins[0] if origins else None,
            "total_weight": total_weight,
            "total_cbm": total_cbm,
            "total_pallet": total_pallet,
            "total_pcs": total_pcs,
            "is_virtual": True,
        }
        fleet = Fleet(**fleet_data)
        await sync_to_async(fleet.save)()

        if shipment_ids:
            await sync_to_async(
                Shipment.objects.filter(id__in=shipment_ids).update
            )(fleet_number=fleet)
        return fleet_number

    async def _process_single_group_booking(self, request: HttpRequest, suggestion_id, cargo_ids, plt_ids,destination, appointment_id,shipment_cargo_id,shipment_type,
                                                              shipment_account,shipment_appointment,load_type,origin,note,pickup_time,pickup_number) -> dict:
        """е§ДзРЖеНХдЄ™е§ІзїДзЪДйҐДзЇ¶еЗЇеЇУ"""       
        new_post = {}
        cargo_id_list = []
        if cargo_ids and cargo_ids.strip():
            cargo_id_list = [int(id.strip()) for id in cargo_ids.split(',') if id.strip()]
        
        plt_id_list = []
        if plt_ids and plt_ids.strip():
            plt_id_list = [int(id.strip()) for id in plt_ids.split(',') if id.strip()]
        # иЃЊзљЃиіІзЙ©IDеПВжХ∞пЉИдЄОhandle_appointment_postдњЭжМБдЄАиЗіпЉЙ
        
        
        total_weight, total_cbm, total_pcs, total_pallet = 0.0, 0.0, 0, 0
        pallet_records = await sync_to_async(list)(
            Pallet.objects.filter(
                id__in=plt_id_list,
            )
        )
       
        # иОЈеПЦжЙАжЬЙPackingListиЃ∞ељХ  
        packinglist_records = await sync_to_async(list)(
            PackingList.objects.filter(
                id__in=cargo_id_list,
            )
        )
        # ж±ЗжАїPalletжХ∞жНЃ
        for p in pallet_records:
            total_weight += p.weight_lbs or 0
            total_cbm += p.cbm or 0
            total_pcs += p.pcs or 0
            total_pallet += 1
        
        # ж±ЗжАїPackingListжХ∞жНЃ
        for pl in packinglist_records:
            total_weight += pl.total_weight_lbs or 0
            total_cbm += pl.cbm or 0
            total_pcs += pl.pcs or 0
            total_pallet += round(pl.cbm /1.8)
        # иЃЊзљЃйҐДзЇ¶дњ°жБѓеПВжХ∞
        shipment_batch_number = await self.generate_unique_batch_number(destination)
        address = await self.get_address(destination)
        shipment_data = {
            'shipment_batch_number': shipment_batch_number,
            'destination': destination,
            'total_weight': total_weight,
            'total_cbm': total_cbm,
            'total_pallet': total_pallet,
            'total_pcs': total_pcs,
            'total_pallet': total_pallet,
            'shipment_type': shipment_type,
            'shipment_account': shipment_account,
            'appointment_id': appointment_id,
            'shipment_cargo_id': shipment_cargo_id,
            'shipment_appointment': shipment_appointment,
            'load_type': load_type,
            'origin': origin,
            'note': note,
            'address': address,
            'pickup_time': pickup_time,
            'pickup_number': pickup_number,
        }
        new_post = {**new_post, **shipment_data}
        new_post['shipment_data'] = str(shipment_data)
        new_post['pl_ids'] = cargo_id_list
        new_post['plt_ids'] = plt_id_list
        new_post['type'] = 'td' 
        new_post['batch_number'] = shipment_batch_number  

        # еИЫеїЇжЦ∞зЪДHttpRequestеѓєи±°
        new_request = HttpRequest()
        new_request.method = 'POST'     
        new_request.POST = new_post     
        
        try:
            # зЫіжО•и∞ГзФ® sm.handle_appointment_post_tuple
            sm = ShippingManagement()
            info = await sm.handle_appointment_post(new_request, 'post_nsop')
            appointment_id = appointment_id
            return {
                'success': True,
                'appointment_id': appointment_id,
                'suggestion_id': suggestion_id,
                'shipment_batch_number': shipment_batch_number,
            }
            
        except Exception as e:
            return {
                'success': False,
                'appointment_id': appointment_id,
                'suggestion_id': suggestion_id,
                'shipment_batch_number': shipment_batch_number,
                'error': f"йҐДзЇ¶е§±иі•: {str(e)}"
            }
    
    async def handle_query_quotation(
        self, request: HttpRequest
    ) -> HttpResponse | tuple[str, dict[str, Any]]:
        '''жЯ•иѓҐе§ЪдЄ™дїУзВєзЪДжК•дїЈ'''
        context = {}
        is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
        cargo_ids = request.POST.get("cargo_ids", "")
        plt_ids = request.POST.get("plt_ids", "")
        cargo_id_list = [int(i) for i in cargo_ids.split(",") if i]
        plt_id_list = [int(i) for i in plt_ids.split(",") if i]
        if not cargo_ids and not plt_ids:
            message = "жЬ™жПРдЊЫIDпЉМжЧ†ж≥ХжЯ•иѓҐжК•дїЈ"
            if is_ajax:
                return JsonResponse({"success": False, "message": message}, status=400)
            context.update({"error_messages": message})
            return await self.handle_td_shipment_post(request, context)
        
        pls = await sync_to_async(
            lambda: list(
                PackingList.objects
                .filter(id__in=cargo_id_list)
                .select_related('container_number')
                .values('PO_ID','container_number__container_number','destination')
                .annotate(
                    total_cbm=Round(Sum('cbm'), 2, output_field=FloatField()),
                    total_pallets=Ceil(
                        Round(Sum('cbm'), 2, output_field=FloatField()) / 1.8
                    ),
                    source=Value('packinglist')
                )
                .order_by('PO_ID')
            )
        )()
        plts = await sync_to_async(
            lambda: list(
                Pallet.objects
                .filter(id__in=plt_id_list)
                .select_related('container_number')
                .values('PO_ID','container_number__container_number','destination','location')
                .annotate(
                    total_cbm=Round(Sum('cbm'), 2, output_field=FloatField()),
                    total_pallets=Count('id'), 
                    source=Value('pallet')
                )
                .order_by('PO_ID')
            )
        )()

        combined_list = pls + plts
        quotation_table_data = []
        quote_total = 0.0
        # жЯ•жЙЊжК•дїЈ
        if not combined_list:
            message = f"{combined_list}жШѓз©ЇзЪД"
            if is_ajax:
                return JsonResponse({"success": False, "message": message}, status=400)
            context = {"error_messages": message}
            return await self.handle_td_shipment_post(request, context)
        
        for po in combined_list:
            # йБНеОЖдЊЭжђ°жЯ•жЙЊжК•дїЈ
            container_number = po['container_number__container_number']
            
            destination_str = po['destination']
            destination = await self._process_destination_wlm(destination_str)
            order = await sync_to_async(
                lambda cn=container_number: Order.objects.select_related(
                    'retrieval_id',  # йҐДеК†иљљretrieval_id
                    'vessel_id',
                    'customer_name'   # йҐДеК†иљљcustomer_name
                ).filter(
                    container_number__container_number=cn
                ).first()
            )()
            # 2026/4/8 claireиѓіжК•дїЈеЇФиѓ•йГљи¶БжМЙзЕІйҐДжК•зЪДдїУеЇУжЯ•жЙЊпЉМдЄНжМЙзЕІеЃЮйЩЕжЙАеЬ®еЬ∞жЯ•жЙЊ
            warehouse = order.retrieval_id.retrieval_destination_area
            # if po['source'] == 'packinglist':
            #     warehouse = order.retrieval_id.retrieval_destination_area
            # else:
            #     warehouse = po['location'].split('-')[0]
            #     if not warehouse:
            #         message = f"{container_number}зЪДжЭње≠РзЉЇе∞СеЃЮйЩЕдїУеЇУдљНзљЃпЉБ"
            #         if is_ajax:
            #             return JsonResponse({"success": False, "message": message}, status=400)
            #         context = {"error_messages": message}
            #         return await self.handle_td_shipment_post(request, context)

            customer_name = order.customer_name.zem_name if order.customer_name else None
            #жЯ•жЙЊжК•дїЈи°®
            quotations = await self._get_fee_details(order, warehouse, customer_name)
            if isinstance(quotations, dict) and quotations.get("error_messages"):
                message = quotations["error_messages"]
                if is_ajax:
                    return JsonResponse({"success": False, "message": message}, status=400)
                context = {"error_messages": message}
                return await self.handle_td_shipment_post(request, context)
            fee_details = quotations['fees']
            
            is_combina = False
            is_combina_reason = None
            if order.order_type == "иљђињРзїДеРИ":
                container = await sync_to_async(
                    lambda: Container.objects.get(container_number=container_number)
                )()
                if container.manually_order_type == "иљђињРзїДеРИ":
                    is_combina = True
                elif container.manually_order_type == "иљђињР":
                    is_combina = False
                else:
                    combina_context, is_combina, is_combina_reason = await self._is_combina(
                        container, order, warehouse
                    )
                    if (
                        isinstance(combina_context, dict)
                        and combina_context.get("error_messages")
                    ):
                        message = combina_context["error_messages"]
                        if is_ajax:
                            return JsonResponse({"success": False, "message": message}, status=400)
                        context = {"error_messages": message}
                        return await self.handle_td_shipment_post(request, context)
            
            non_combina_table = True
            
            if is_combina:
                #зїДеРИжЯЬиЃ°зЃЧ
                combina_key = f"{warehouse}_COMBINA"
                if combina_key not in fee_details:
                    message = f"{warehouse}_COMBINA-{container_number}жЬ™жЙЊеИ∞зїДеРИжЯЬжК•дїЈи°®пЉБ"
                    if is_ajax:
                        return JsonResponse({"success": False, "message": message}, status=400)
                    context = {"error_messages": message}
                    return await self.handle_td_shipment_post(request, context)
                
                rules = fee_details.get(combina_key).details
                
                container_type_temp = 0 if "40" in container.container_type else 1
                total_container_cbm = await sync_to_async(
                    lambda: PackingList.objects
                        .filter(container_number=container)
                        .aggregate(
                            total_cbm_sum=Round(
                                Sum('cbm'), 
                                2, 
                                output_field=FloatField()
                            )
                        )['total_cbm_sum'] or 0.0
                )()
                cbm = float(po['total_cbm'])
                temp_table = await self._process_combina_quote(po, cbm, float(total_container_cbm), rules, container_type_temp, warehouse, quotations.get("filename"))
                if temp_table:
                    non_combina_table = False
                    quotation_table_data.append(temp_table)
                
                rules = fee_details.get(combina_key).details
            
            if non_combina_table:
                # дЄНзЃ°зђ¶дЄНзђ¶еРИиљђињРпЉМж≤°жМЙзїДеРИжЯЬиЃ°еИ∞иієпЉМе∞±жМЙиљђињРжЦєеЉПиЃ°зЃЧ
                
                public_key = f"{warehouse}_PUBLIC"
                if public_key not in fee_details:
                    context = {"error_messages": f'{warehouse}_PUBLIC-{container_number}жЬ™жЙЊеИ∞дЇЪй©ђйАКж≤Ге∞ФзОЫжК•дїЈи°®пЉБ'}
                    return await self.handle_td_shipment_post(request, context)
                
                rules = fee_details.get(f"{warehouse}_PUBLIC").details
                niche_warehouse = fee_details.get(f"{warehouse}_PUBLIC").niche_warehouse
                if destination in niche_warehouse:
                    is_niche_warehouse = True
                else:
                    is_niche_warehouse = False

                #LAеТМеЕґдїЦзЪДе≠ШеВ®ж†ЉеЉПжЬЙзВєеМЇеИЂ
                details = (
                    {"LA_AMAZON": rules}
                    if "LA" in warehouse and "LA_AMAZON" not in rules
                    else rules
                )
                delivery_category = None
                rate_found = False
                for category, zones in details.items():
                    for zone, locations in zones.items():
                        if destination in locations:
                            if "AMAZON" in category:
                                delivery_category = "amazon"
                                rate = zone
                                rate_found = True
                            elif "WALMART" in category:
                                delivery_category = "walmart"
                                rate = zone
                                rate_found = True
                    if rate_found:
                        break

                if rate_found:
                    # жЙЊеИ∞жК•дїЈ
                    rate = float(rate) if rate else 0.0
                    quotation_table_data.append({
                        'container_number': po['container_number__container_number'],
                        'order_type': order.order_type,
                        'is_combina': is_combina,
                        'is_combina_reason': is_combina_reason,
                        'destination': destination,                         
                        'cbm': po['total_cbm'],
                        'total_pallets': po['total_pallets'], 
                        'rate': rate, 
                        'amount': rate * po['total_pallets'],
                        'type': delivery_category,
                        'region': None,
                        'warehouse': warehouse, 
                        'is_niche_warehouse': is_niche_warehouse,  
                        'quotation_name': quotations['filename'],  
                    })
                    quote_total += rate * po['total_pallets']
                else:            
                    quotation_table_data.append({
                        'container_number': po['container_number__container_number'],
                        'order_type': order.order_type,
                        'is_combina': is_combina,
                        'is_combina_reason': is_combina_reason,
                        'destination': destination,                         
                        'cbm': po['total_cbm'],
                        'total_pallets': po['total_pallets'], 
                        'rate': None, 
                        'amount': None,
                        'type': 'жЬ™жЙЊеИ∞',
                        'region': None,
                        'warehouse': warehouse, 
                        'is_niche_warehouse': None,  
                        'quotation_name': quotations.get("filename"),  
                    })
            
        quotation_table_data = sorted(
            quotation_table_data,
            key=lambda x: x.get("type") or "",
        )
        context = {"quotation_table_data": quotation_table_data}
        if is_ajax:
            template = get_template("post_port/new_sop/02_shipment/modals/quote_table_ajax.html")
            html = template.render(context, request)
            return JsonResponse({"success": True, "html": html})
        return await self.handle_td_shipment_post(request, context)
    
    async def _process_destination_wlm(self,destination):
        """е§ДзРЖзЫЃзЪДеЬ∞е≠ЧжЃµ"""
        if destination and '-' in destination:
            parts = destination.split('-')
            if len(parts) > 1:
                return parts[1]
        return destination
    
    async def _process_combina_quote(self, po, cbm, total_cbm, rules, container_type_temp, warehouse, filename):
        """жМЙзїДеРИжЯЬжЦєеЉПжЯ•жЙЊдїУзВєжК•дїЈ """

        #жФєеЙНеТМжФєеРОзЪД
        destination_origin, destination = self._process_destination(po['destination'])
        
        # ж£АжЯ•жШѓеР¶е±ЮдЇОзїДеРИеМЇеЯЯ
        price = 0
        is_combina_region = False
        region = None
        for region, region_data in rules.items():
            for item in region_data:
                rule_locations = item.get("location", [])
                if isinstance(rule_locations, str):
                    rule_locations = [rule_locations] # зїЯдЄАиљђжИРеИЧи°®е§ДзРЖ

                if any(destination == loc.replace(" ", "").upper() for loc in rule_locations):
                    is_combina_region = True
                    price = item["prices"][container_type_temp]
                    region = region
                    break
            if is_combina_region:
                break
        if destination == "UPS":
            is_combina_region = False
        
        if is_combina_region:
            '''жМЙзїДеРИжЯЬиЃ°иіє'''
            return ({
                'container_number': po['container_number__container_number'],
                'order_type': 'иљђињРзїДеРИ',
                'is_combina': True,
                'destination': po['destination'],                         
                'cbm': cbm,
                'total_pallets': po['total_pallets'], 
                'rate': price, 
                'amount': round(price * cbm / total_cbm,2),
                'type': "зїДеРИжЯЬ",
                'region': region,
                'warehouse': warehouse, 
                'is_niche_warehouse': None,  
                'quotation_name': filename,  
            })
        else:
            return None
        
    def _process_destination(self, destination_origin):
        """е§ДзРЖзЫЃзЪДеЬ∞е≠Чзђ¶дЄ≤"""
        def clean_all_spaces(s):
            if not s:  # е§ДзРЖNone/з©Їе≠Чзђ¶дЄ≤
                return ""
            # еМєйЕНжЙАжЬЙз©Їж†Љз±їеЮЛпЉЪ
            # \xa0 йЭЮдЄ≠жЦ≠з©Їж†Љ | \u3000 дЄ≠жЦЗеЕ®иІТз©Їж†Љ | \s жЩЃйАЪз©Їж†Љ/еИґи°®зђ¶/жНҐи°Мз≠Й
            import re
            cleaned = re.sub(r'[\xa0\u3000\s]+', '', str(s))
            return cleaned
        
        destination_origin = str(destination_origin)

        # еМєйЕНж®°еЉПпЉЪжМЙ"жФє"жИЦ"йАБ"еИЖеЙ≤пЉМеИЖеЙ≤зђ¶жФЊеЬ®зђђдЄАзїДзЪДжЬЂе∞Њ
        if "жФє" in destination_origin or "йАБ" in destination_origin:
            # жЙЊеИ∞зђђдЄАдЄ™"жФє"жИЦ"йАБ"зЪДдљНзљЃ
            first_change_pos = min(
                (destination_origin.find(char) for char in ["жФє", "йАБ"] 
                if destination_origin.find(char) != -1),
                default=-1
            )
            
            if first_change_pos != -1:
                # зђђдЄАйГ®еИЖпЉЪеИ∞зђђдЄАдЄ™"жФє"жИЦ"йАБ"пЉИеМЕеРЂеИЖйЪФзђ¶пЉЙ
                first_part = destination_origin[:first_change_pos + 1]
                # зђђдЇМйГ®еИЖпЉЪеЙ©дЄЛзЪДйГ®еИЖ
                second_part = destination_origin[first_change_pos + 1:]
                
                # е§ДзРЖзђђдЄАйГ®еИЖпЉЪжМЙ"-"еИЖеЙ≤еПЦеРОйЭҐзЪДйГ®еИЖ
                if "-" in first_part:
                    if first_part.upper().startswith("UPS-"):
                        first_result = first_part
                    else:
                        first_result = first_part.split("-", 1)[1]
                else:
                    first_result = first_part
                
                # е§ДзРЖзђђдЇМйГ®еИЖпЉЪжМЙ"-"еИЖеЙ≤еПЦеРОйЭҐзЪДйГ®еИЖ
                if "-" in second_part:
                    if second_part.upper().startswith("UPS-"):
                        second_result = second_part
                    else:
                        second_result = second_part.split("-", 1)[1]
                else:
                    second_result = second_part
                
                second_result = second_result.replace(" ", "").upper()
                return clean_all_spaces(first_result), clean_all_spaces(second_result)
            else:
                raise ValueError(first_change_pos)
        
        # е¶ВжЮЬдЄНеМЕеРЂ"жФє"жИЦ"йАБ"жИЦиАЕж≤°жЬЙжЙЊеИ∞
        # еП™е§ДзРЖзђђдЇМйГ®еИЖпЉИеБЗиЃЊзђђдЄАйГ®еИЖдЄЇз©ЇпЉЙ
        if "-" in destination_origin:
            if destination_origin.upper().startswith("UPS-"):
                second_result = destination_origin
            else:
                second_result = destination_origin.split("-", 1)[1]
            
        else:
            second_result = destination_origin
        
        second_result = second_result.replace(" ", "").upper()
        return None, clean_all_spaces(second_result)
    
    async def _is_combina(self, container: Container, order: Order, warehouse) -> Any:
        context = {}
        
        customer_name = order.customer_name.zem_name
        vessel_etd = order.vessel_id.vessel_etd
        container_type = container.container_type
        has_pallet = True
        #  еЯЇз°АжХ∞жНЃзїЯиЃ°
        plts = await sync_to_async(
            lambda: Pallet.objects.filter(
                container_number=container
            ).aggregate(
                unique_destinations=Count("destination", distinct=True),
                total_weight=Sum("weight_lbs"),
                total_cbm=Sum("cbm"),
                total_pallets=Count("id"),
            )
        )()
        if plts['total_pallets'] == 0:
            has_pallet = False
            plts = await sync_to_async(
                lambda: PackingList.objects.filter(
                    container_number=container
                ).aggregate(
                    unique_destinations=Count("destination", distinct=True),
                    total_weight=Sum("total_weight_lbs"),
                    total_cbm=Sum("cbm"),
                    total_pallets=Coalesce(
                        Round(
                            Cast(Sum("cbm"), output_field=FloatField()) / 1.8,
                            output_field=IntegerField()
                        ),
                        0  # йїШиЃ§еАЉпЉМељУSum("cbm")дЄЇNoneжЧґиЃЊдЄЇ0
                    )
                )
            )()
        plts["total_cbm"] = round(float(plts.get("total_cbm") or 0.0), 2)
        plts["total_weight"] = round(float(plts.get("total_weight") or 0.0), 2)
        if plts["total_cbm"] == 0.0:
            raise ValueError(f"{container.container_number} total_cbmжШѓ0")
        # иОЈеПЦеМєйЕНзЪДжК•дїЈи°®
        matching_quotation = await sync_to_async(
            lambda: QuotationMaster.objects.filter(
                effective_date__lte=vessel_etd,
                is_user_exclusive=True,
                exclusive_user=customer_name,
                quote_type='receivable',
            )
            .order_by("-effective_date")
            .first()
        )()
        if not matching_quotation:
            matching_quotation = await sync_to_async(
                lambda: QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=False,
                    quote_type='receivable',
                )
                .order_by("-effective_date")
                .first()
            )()
        if not matching_quotation:
            context.update({"error_messages": f"жЙЊдЄНеИ∞{container.container_number}еПѓзФ®зЪДжК•дїЈи°®пЉБ"})
            return context, None, None
        # иОЈеПЦзїДеРИжЯЬиІДеИЩ
        try:
            stipulate_fee_detail = await sync_to_async(
                lambda: FeeDetail.objects.get(
                    quotation_id=matching_quotation.id, 
                    fee_type="COMBINA_STIPULATE"
                )
            )()
            stipulate = stipulate_fee_detail.details
        except FeeDetail.DoesNotExist:
            context.update({
                "error_messages": f"жК•дїЈи°®гАК{matching_quotation.filename}гАЛ-{matching_quotation.id}дЄ≠жЙЊдЄНеИ∞<жК•дїЈи°®иІДеИЩ>еИЖи°®пЉМиѓЈжИ™ж≠§еЫЊзїЩжКАжЬѓеСШпЉБ"
            })
            return context, None, None
        
        combina_fee_detail = await sync_to_async(
            lambda: FeeDetail.objects.get(
                quotation_id=matching_quotation.id, 
                fee_type=f"{warehouse}_COMBINA"
            )
        )()
        combina_fee = combina_fee_detail.details
        if isinstance(combina_fee, str):
            combina_fee = json.loads(combina_fee)

        # зЬЛжШѓеР¶иґЕеЗЇзїДеРИжЯЬйЩРеЃЪдїУзВє,NJ/SAVжШѓ14дЄ™
        warehouse_specific_key = f'{warehouse}_max_mixed'
        if warehouse_specific_key in stipulate.get("global_rules", {}):
            combina_threshold = stipulate["global_rules"][warehouse_specific_key]["default"]
        else:
            combina_threshold = stipulate["global_rules"]["max_mixed"]["default"]

        warehouse_specific_key1 = f'{warehouse}_bulk_threshold'
        if warehouse_specific_key1 in stipulate.get("global_rules", {}):
            uncombina_threshold = stipulate["global_rules"][warehouse_specific_key1]["default"]
        else:
            uncombina_threshold = stipulate["global_rules"]["bulk_threshold"]["default"]

        if plts["unique_destinations"] > uncombina_threshold:
            container.account_order_type = "иљђињР"
            container.non_combina_reason = (
                f"жАїдїУзВєиґЕињЗ{uncombina_threshold}дЄ™"
            )
            await sync_to_async(container.save)()
            return context, False, f"жАїдїУзВєиґЕињЗ{uncombina_threshold}дЄ™" # дЄНжШѓзїДеРИжЯЬ

        # жМЙеМЇеЯЯзїЯиЃ°
        if has_pallet:
            destinations = await sync_to_async(
                lambda: list(Pallet.objects.filter(container_number=container)
                            .values_list("destination", flat=True)
                            .distinct())
            )()
            plts_by_destination = await sync_to_async(
                lambda: list(Pallet.objects.filter(container_number=container)
                            .values("destination")
                            .annotate(total_cbm=Sum("cbm")))
            )()
        else:
            destinations = await sync_to_async(
                lambda: list(PackingList.objects.filter(container_number=container)
                            .values_list("destination", flat=True)
                            .distinct())
            )()
            plts_by_destination = await sync_to_async(
                lambda: list(PackingList.objects.filter(container_number=container)
                            .values("destination")
                            .annotate(total_cbm=Sum("cbm")))
            )()
        total_cbm_sum = sum(item["total_cbm"] for item in plts_by_destination)
        # еМЇеИЖзїДеРИжЯЬеМЇеЯЯеТМйЭЮзїДеРИжЯЬеМЇеЯЯ
        container_type_temp = 0 if "40" in container_type else 1
        matched_regions = self.find_matching_regions(
            plts_by_destination, combina_fee, container_type_temp, total_cbm_sum, combina_threshold
        )
        # еИ§жЦ≠жШѓеР¶жЈЈеМЇпЉМFalseи°®з§Їжї°иґ≥жЈЈеМЇжЭ°дїґ
        is_mix = self.is_mixed_region(
            matched_regions["matching_regions"], warehouse, vessel_etd
        )
        if is_mix:
            container.account_order_type = "иљђињР"
            container.non_combina_reason = "жЈЈеМЇдЄНзђ¶еРИж†ЗеЗЖ"
            await sync_to_async(container.save)()
            return context, False, "жЈЈеМЇдЄНзђ¶еРИж†ЗеЗЖ"
        
        filtered_non_destinations = [key for key in matched_regions["non_combina_dests"].keys() if "UPS" not in key]
        # йЭЮзїДеРИжЯЬеМЇеЯЯ
        non_combina_region_count = len(filtered_non_destinations)
        # зїДеРИжЯЬеМЇеЯЯ
        combina_region_count = len(matched_regions["combina_dests"])

        filtered_destinations = self._filter_ups_destinations(destinations)
        if combina_region_count + non_combina_region_count != len(filtered_destinations):
            raise ValueError(
                f"иЃ°зЃЧзїДеРИжЯЬеТМйЭЮзїДеРИжЯЬеМЇеЯЯжЬЙиѓѓ\n"
                f"зїДеРИжЯЬзЫЃзЪДеЬ∞пЉЪ{matched_regions['combina_dests']}пЉМжХ∞йЗПпЉЪ{combina_region_count}\n"
                f"йЭЮзїДеРИжЯЬзЫЃзЪДеЬ∞пЉЪ{filtered_non_destinations}пЉМжХ∞йЗПпЉЪ{non_combina_region_count}\n"
                f"зЫЃзЪДеЬ∞йЫЖеРИпЉЪ{filtered_destinations}\n"
                f"зЫЃзЪДеЬ∞жАїжХ∞пЉЪ{len(filtered_destinations)}"
            )
        sum_region_count = non_combina_region_count + combina_region_count
        if sum_region_count > uncombina_threshold:
            # ељУйЭЮзїДеРИжЯЬзЪДеМЇеЯЯжХ∞йЗПиґЕеЗЇжЧґпЉМдЄНиГљжМЙиљђињРзїДеРИ
            container.account_order_type = "иљђињР"
            container.non_combina_reason = f"жАїеМЇжХ∞йЗПдЄЇ{sum_region_count},и¶Бж±ВжШѓ{uncombina_threshold}"
            await sync_to_async(container.save)()
            return context, False,f"жАїеМЇжХ∞йЗПдЄЇ{sum_region_count},и¶Бж±ВжШѓ{uncombina_threshold}"
        container.non_combina_reason = None
        container.account_order_type = "иљђињРзїДеРИ"
        await sync_to_async(container.save)()
        return context, True, None
    
    def find_matching_regions(
        self,
        plts_by_destination: dict,
        combina_fee: dict,
        container_type,
        total_cbm_sum: FloatField,
        combina_threshold: int,
    ) -> dict:
        matching_regions = defaultdict(float)  # еРДеМЇзЪДcbmжАїеТМ
        des_match_quote = {}  # еРДдїУзВєзЪДеМєйЕНиѓ¶жГЕ
        destination_matches = set()  # зїДеРИжЯЬзЪДдїУзВє
        non_combina_dests = {}  # йЭЮзїДеРИжЯЬзЪДдїУзВє
        price_display = defaultdict(
            lambda: {"price": 0.0, "location": set()}
        )  # еРДеМЇзЪДдїЈж†ЉеТМдїУзВє
        dest_cbm_list = []  # дЄіжЧґе≠ШеВ®еИЭз≠ЫзїДеРИжЯЬеЖЕзЪДcbmеТМеМєйЕНдњ°жБѓ

        region_counter = {}
        region_price_map = {}
        for plts in plts_by_destination:
            destination = plts["destination"]
            if ('UPS' in destination) or ('FEDEX' in destination):
                continue
            # е¶ВжЮЬжШѓж≤Ге∞ФзОЫзЪДпЉМеП™дњЭзХЩеРОйЭҐзЪДеРНе≠ЧпЉМеЫ†дЄЇжК•дїЈи°®йЗМе∞±жШѓињЩдєИдњЭзХЩзЪД
            clean_dest = destination.replace("ж≤Ге∞ФзОЫ", "").strip()

            if clean_dest.upper().startswith("UPS-"):
                dest = clean_dest
            else:
                dest = clean_dest.split("-")[-1].strip()

            cbm = plts["total_cbm"]
            dest_matches = []
            matched = False
            # йБНеОЖжЙАжЬЙеМЇеЯЯеТМlocation
            for region, fee_data_list in combina_fee.items():           
                for fee_data in fee_data_list:
                    prices_obj = fee_data["prices"]
                    price = self._extract_price(prices_obj, container_type)
                    
                    # е¶ВжЮЬеМєйЕНеИ∞зїДеРИжЯЬдїУзВєпЉМе∞±зЩїиЃ∞еИ∞зїДеРИжЯЬйЫЖеРИдЄ≠
                    if dest in fee_data["location"]:
                        # еИЭеІЛеМЦ
                        if region not in region_price_map:
                            region_price_map[region] = [price]
                            region_counter[region] = 0
                            actual_region = region
                        else:
                            # е¶ВжЮЬиѓ• region дЄЛеЈ≤жЬЙзЫЄеРМдїЈж†Љ вЖТ дЄНеК†зЉЦеПЈ
                            found = None
                            for r_key, r_val in price_display.items():
                                if r_key.startswith(region) and r_val["price"] == price:
                                    found = r_key
                                    break
                            if found:
                                actual_region = found
                            else:                                
                                # жЦ∞дїЈж†Љ вЖТ йЬАи¶БзЉЦеПЈ
                                region_counter[region] += 1
                                actual_region = f"{region}{region_counter[region]}"
                                region_price_map[region].append(price)

                        temp_cbm = matching_regions.get(actual_region, 0) + cbm
                        matching_regions[actual_region] = temp_cbm
                        dest_matches.append(
                            {
                                "region": actual_region,
                                "location": dest,
                                "prices": fee_data["prices"],
                                "cbm": cbm,
                            }
                        )
                        if actual_region not in price_display:
                            price_display[actual_region] = {
                                "price": price,
                                "location": set([dest]),
                            }
                        else:
                            # дЄНи¶Би¶ЖзЫЦпЉМжЫіжЦ∞йЫЖеРИ
                            price_display[actual_region]["location"].add(dest)
                        matched = True
            
            if not matched:
                # йЭЮзїДеРИжЯЬдїУзВє
                non_combina_dests[dest] = {"cbm": cbm}
            # иЃ∞ељХеМєйЕНзїУжЮЬ
            if dest_matches:
                des_match_quote[dest] = dest_matches
                # е∞ЖзїДеРИжЯЬеЖЕзЪДиЃ∞ељХдЄЛжЭ•пЉМеРОзї≠жЦєдЊњжМЙзЕІcbmжОТеЇП
                dest_cbm_list.append(
                    {"dest": dest, "cbm": cbm, "matches": dest_matches}
                )
                destination_matches.add(dest)
        if len(destination_matches) > combina_threshold:
            # жМЙcbmйЩНеЇПжОТеЇПпЉМе∞Жcbmе§ІзЪДељТеИ∞йЭЮзїДеРИ
            sorted_dests = sorted(dest_cbm_list, key=lambda x: x["cbm"], reverse=True)
            # йЗНжЦ∞е∞ЖжОТеЇПеРОзЪДеЙН12дЄ™еК†еЕ•йЗМйЭҐ
            destination_matches = set()
            matching_regions = defaultdict(float)
            price_display = defaultdict(lambda: {"price": 0.0, "location": set()})
            for item in sorted_dests[:combina_threshold]:
                dest = item["dest"]
                destination_matches.add(dest)

                # йЗНжЦ∞иЃ°зЃЧеРДеМЇеЯЯзЪДCBMжАїеТМ
                for match in item["matches"]:
                    region = match["region"]
                    matching_regions[region] += item["cbm"]
                    price_display[region]["price"] = self._extract_price(match["prices"], container_type)
                    
                    price_display[region]["location"].add(dest)

            # еЕґдљЩдїУзВєиљђдЄЇйЭЮзїДеРИжЯЬ
            for item in sorted_dests[combina_threshold:]:
                non_combina_dests[item["dest"]] = {"cbm": item["cbm"]}
                # е∞Жcbmе§ІзЪДдїОзїДеРИжЯЬйЫЖеРИдЄ≠еИ†йЩ§
                des_match_quote.pop(item["dest"], None)

        # дЄЛйЭҐеЉАеІЛиЃ°зЃЧзїДеРИжЯЬеТМйЭЮзїДеРИжЯЬеРДдїУзВєеН†жАїдљУзІѓзЪДжѓФдЊЛ
        total_ratio = 0.0
        ratio_info = []

        # е§ДзРЖзїДеРИжЯЬдїУзВєзЪДcbm_ratio
        for dest, matches in des_match_quote.items():
            cbm = matches[0]["cbm"]  # еРМдЄАдЄ™destзЪДcbmеЬ®жЙАжЬЙmatchesдЄ≠зЫЄеРМ
            ratio = round(cbm / total_cbm_sum, 4)
            total_ratio += ratio
            ratio_info.append((dest, ratio, cbm, True))  # жЬАеРОдЄАдЄ™еПВжХ∞и°®з§ЇжШѓеР¶жШѓзїДеРИжЯЬ
            for match in matches:
                match["cbm_ratio"] = ratio

        # е§ДзРЖйЭЮзїДеРИжЯЬдїУзВєзЪДcbm_ratio
        for dest, data in non_combina_dests.items():
            cbm = data["cbm"]
            ratio = round(cbm / total_cbm_sum, 4)
            total_ratio += ratio
            ratio_info.append((dest, ratio, cbm, False))
            data["cbm_ratio"] = ratio

        # е§ДзРЖеЫЫиИНдЇФеЕ•еѓЉиЗізЪДиѓѓеЈЃ
        if abs(total_ratio - 1.0) > 0.0001:  # иАГиЩСжµЃзВєжХ∞з≤ЊеЇ¶
            # жЙЊеИ∞CBMжЬАе§ІзЪДдїУзВє
            ratio_info.sort(key=lambda x: x[2], reverse=True)
            largest_dest, largest_ratio, largest_cbm, is_combi = ratio_info[0]

            # и∞ГжХіжЬАе§ІзЪДдїУзВєзЪДratio
            diff = 1.0 - total_ratio
            if is_combi:
                for match in des_match_quote[largest_dest]:
                    match["cbm_ratio"] = round(match["cbm_ratio"] + diff, 4)
            else:
                non_combina_dests[largest_dest]["cbm_ratio"] = round(
                    non_combina_dests[largest_dest]["cbm_ratio"] + diff, 4
                )
        return {
            "des_match_quote": des_match_quote,
            "matching_regions": matching_regions,
            "combina_dests": destination_matches,
            "non_combina_dests": non_combina_dests,
            "price_display": price_display,
        }
    
    def _extract_price(self, prices_obj, container_type):
        """
        еЃЙеЕ®еЬ∞дїО prices_obj дЄ≠жПРеПЦжХ∞еАЉ priceпЉЪ
        - е¶ВжЮЬ prices_obj жШѓ dictпЉМжМЙйФЃеПЦпЉИcontainer_type еПѓдЄЇе≠Чзђ¶дЄ≤жИЦжХіеЮЛпЉЙгАВ
        - е¶ВжЮЬжШѓ list/tupleпЉМдЄФ container_type жШѓ intпЉМеИЩе∞ЭиѓХеПЦ prices_obj[container_type]гАВ
        иЛ•иґКзХМжИЦиѓ•й°єдЄНжШѓжХ∞еАЉпЉМеИЩеЫЮйААеИ∞еИЧи°®дЄ≠зђђдЄАдЄ™жХ∞еАЉй°єгАВ
        - е¶ВжЮЬжШѓеНХеАЉпЉИint/floatпЉЙпЉМзЫіжО•ињФеЫЮгАВ
        - еЕґеЃГжГЕеЖµињФеЫЮ NoneгАВ
        """
        # дЉШеЕИе§ДзРЖ dict
        if isinstance(prices_obj, dict):
            # еЕБиЃЄ container_type жШѓ str жИЦ intпЉИint иљђдЄЇзіҐеЉХзЪДжГЕеЖµдЄНеЄЄиІБпЉЙ
            val = prices_obj.get(container_type)
            if isinstance(val, (int, float)):
                return val
            # е¶ВжЮЬеПЦеИ∞зЪДдЄНжШѓжХ∞е≠ЧпЉМе∞ЭиѓХжЙЊ dict зЪДзђђдЄАдЄ™жХ∞е≠ЧеАЉдљЬдЄЇеЫЮйАА
            for v in prices_obj.values():
                if isinstance(v, (int, float)):
                    return v
            return None

        # list/tuple жМЙ index йАЙ
        if isinstance(prices_obj, (list, tuple)):
            # ељУ container_type жШѓжХіжХ∞зіҐеЉХжЧґпЉМдЉШеЕИдљњзФ®иѓ•зіҐеЉХ
            if isinstance(container_type, int):
                try:
                    candidate = prices_obj[container_type]
                    if isinstance(candidate, (int, float)):
                        return candidate
                except Exception:
                    pass
            # еЫЮйААпЉЪйАЙзђђдЄАдЄ™жХ∞е≠Чй°є
            first_num = next((x for x in prices_obj if isinstance(x, (int, float))), None)
            return first_num

        # зЫіжО•жШѓжХ∞е≠Ч
        if isinstance(prices_obj, (int, float)):
            return prices_obj

        # еЕґдїЦпЉИе≠Чзђ¶дЄ≤з≠ЙпЉЙпЉМдЄНиГљдљЬдЄЇ price
        return None
    
    def _filter_ups_destinations(self, destinations):
        """ињЗжї§жОЙеМЕеРЂUPSзЪДзЫЃзЪДеЬ∞пЉМжФѓжМБеИЧи°®еТМQuerySet"""
        if hasattr(destinations, '__iter__') and not isinstance(destinations, (str, dict)):
            destinations_list = list(destinations)
        else:
            destinations_list = destinations
        filtered_destinations = [
            dest.strip() for dest in destinations_list 
            if dest is not None 
            and 'UPS' not in str(dest).upper() 
            and 'FEDEX' not in str(dest).upper()
        ]
        return list(dict.fromkeys(filtered_destinations))
    
    def is_mixed_region(self, matched_regions, warehouse, vessel_etd) -> bool:
        regions = list(matched_regions.keys())
        # LAдїУеЇУзЪДзЙєжЃКиІДеИЩпЉЪCDEFеМЇдЄНиГљжЈЈ
        if warehouse == "LA":
            if vessel_etd.year > 2025: 
                return False
            if vessel_etd.month > 7 or (
                vessel_etd.month == 7 and vessel_etd.day >= 15
            ):  # 715дєЛеРОж≤°жЬЙжЈЈеМЇйЩРеИґ
                return False
            if len(regions) <= 1:  # еП™жЬЙдЄАдЄ™еМЇпЉМе∞±ж≤°жЬЙжЈЈеМЇзЪДжГЕеЖµ
                return False
            if set(regions) == {"AеМЇ", "BеМЇ"}:  # е¶ВжЮЬеП™жЬЙAеМЇеТМBеМЇпЉМдєЯжї°иґ≥жЈЈеМЇиІДеИЩ
                return False
            return True
        # еЕґдїЦдїУеЇУжЧ†йЩРеИґ
        return False
    
    async def _get_fee_details(self, order: Order, warehouse, customer_name) -> dict:
        context = {}
        quotation, quotation_error = await self._get_quotation_for_order(order, customer_name, 'receivable')
        if quotation_error:
            context.update({"error_messages": quotation_error})
            return context
        id = quotation.id

        fee_types = {
            "NJ": ["NJ_LOCAL", "NJ_PUBLIC", "NJ_COMBINA"],
            "SAV": ["SAV_PUBLIC", "SAV_COMBINA"],
            "LA": ["LA_PUBLIC", "LA_COMBINA"],
        }.get(warehouse, [])

        fees_list = await sync_to_async(
            lambda qid=id, ft=fee_types: list(
                FeeDetail.objects.filter(
                    quotation_id=qid, 
                    fee_type__in=ft
                )
            )
        )()
        fees_dict = {fee.fee_type: fee for fee in fees_list}
        return {
            "quotation": quotation,
            "fees": fees_dict,
            "filename": quotation.filename,
        }
    
    async def _get_quotation_for_order(self, order: Order, customer_name, quote_type: str = 'receivable') :
        """иОЈеПЦиЃҐеНХеѓєеЇФзЪДжК•дїЈи°®"""
        vessel_etd = order.vessel_id.vessel_etd
        
        # еЕИжЯ•жЙЊзФ®жИЈдЄУе±ЮжК•дїЈи°®
        quotation = await sync_to_async(
            lambda: QuotationMaster.objects.filter(
                effective_date__lte=vessel_etd,
                is_user_exclusive=True,
                exclusive_user=customer_name,
                quote_type=quote_type,
            )
            .order_by("-effective_date")
            .first()
        )()
        quotation = await sync_to_async(
            lambda: QuotationMaster.objects.filter(
                effective_date__lte=vessel_etd,
                is_user_exclusive=True,
                exclusive_user=customer_name,
                quote_type=quote_type,
            )
            .order_by("-effective_date")
            .first()
        )()
        
        if not quotation:
            # жЯ•жЙЊйАЪзФ®жК•дїЈи°®
            quotation = await sync_to_async(
                lambda: QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=False,
                    quote_type=quote_type,
                )
                .order_by("-effective_date")
                .first()
            )()
        
        if quotation:
            return quotation, None
        else:
            error_msg = f"жЙЊдЄНеИ∞зФЯжХИжЧ•жЬЯеЬ®{vessel_etd}дєЛеЙНзЪД{quote_type}жК•дїЈи°®"
            return None, error_msg
        
    async def handle_edit_note_sp_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        cargo_ids = request.POST.get("cargo_ids", "")
        plt_ids = request.POST.get("plt_ids", "")
        note_sp = request.POST.get("note_sp", "").strip()
        context = {}

        cargo_id_list = [int(i) for i in cargo_ids.split(",") if i]
        plt_id_list = [int(i) for i in plt_ids.split(",") if i]
        if not cargo_ids and not plt_ids:
            context.update({'error_messages': "жЬ™жПРдЊЫдїїдљХиЃ∞ељХIDпЉМжЧ†ж≥ХжЫіжЦ∞е§Зж≥®"})
            return await self.handle_td_shipment_post(request, context)
        # жЫіжЦ∞ PackingList
        if cargo_id_list:
            updated_count = await sync_to_async(
                lambda: PackingList.objects.filter(id__in=cargo_id_list).update(note_sp=note_sp)
            )()
            if updated_count == 0:
                context.update({'error_messages': "жЫіжЦ∞е§Зж≥®е§±иі•пЉБ"})
                return await self.handle_td_shipment_post(request,context)

        # жЫіжЦ∞ Pallet
        if plt_id_list:
            updated_count = await sync_to_async(
                lambda: Pallet.objects.filter(id__in=plt_id_list).update(note_sp=note_sp)
            )()
            if updated_count == 0:
                context.update({'error_messages': "жЫіжЦ∞е§Зж≥®е§±иі•пЉБ"})
                return await self.handle_td_shipment_post(request,context)
        context.update({'success_messages':"жЫіжЦ∞е§Зж≥®жИРеКЯпЉБ"}) 
        page = request.POST.get("page")
        if page == "arm_appointment":
            return await self.handle_unscheduled_pos_post(request,context)
        else:
            return await self.handle_td_shipment_post(request,context)


    async def handle_edit_appointment_post(
        self, request: HttpRequest, name: str | None = None
    ) -> tuple[str, dict[str, Any]]:
        context = {}
        if name == "fleet_departure":
            shipment_batch_number = request.POST.get('shipment_batch_number', '').strip()
            old_shipments = await sync_to_async(list)(
                Shipment.objects.filter(shipment_batch_number=shipment_batch_number)
            ) 
            appointment_id_new = request.POST.get('appointment_id', '').strip()
        else:
            shipment_id = request.POST.get('shipment_id', '').strip()
            appointment_id_old = request.POST.get('appointment_id', '').strip()
            shipment_batch_number = request.POST.get('shipment_batch_number', '').strip()
            if shipment_id:
                old_shipments = await sync_to_async(list)(
                    Shipment.objects.filter(id=shipment_id)
                )
                appointment_id_old = old_shipments[0].appointment_id
            elif appointment_id_old:
                old_shipments = await sync_to_async(list)(
                    Shipment.objects.filter(appointment_id=appointment_id_old)
                )
            elif shipment_batch_number:
                old_shipments = await sync_to_async(list)(
                    Shipment.objects.filter(shipment_batch_number=shipment_batch_number)
                )
                appointment_id_old = old_shipments[0].appointment_id
            else:
                old_shipments = []
            appointment_id_new = request.POST.get('appointment_id_input', '').strip()
        if not old_shipments:
            key = shipment_id or appointment_id_old or shipment_batch_number
            context.update({'error_messages':f"жЬ™жЙЊеИ∞ ID={key}!"})     
        if len(old_shipments) > 1:
            key = shipment_id or appointment_id_old or shipment_batch_number
            context.update({'error_messages':f"жЙЊеИ∞е§ЪжЭ°зЫЄеРМ ID={key}зЪДиЃ∞ељХпЉМиѓЈж£АжЯ•жХ∞жНЃ!"})   
        
        old_shipment = old_shipments[0]
        if name == "fleet_departure":
            appointment_id_old = old_shipments[0].appointment_id
    
        shipment_appointment = request.POST.get('shipment_appointment')
        pickup_time = request.POST.get('pickup_time')
        pickup_number = request.POST.get('pickup_number')
        destination = request.POST.get('destination').strip()
        load_type = request.POST.get('load_type')
        origin = request.POST.get('origin')
        shipment_cargo_id = request.POST.get('shipment_cargo_id')
        page = request.POST.get("page")
        if appointment_id_new == appointment_id_old:
            old_shipment.shipment_appointment = shipment_appointment if shipment_appointment else None
            old_shipment.destination = destination
            old_shipment.load_type = load_type
            old_shipment.origin = origin
            old_shipment.pickup_time = pickup_time if pickup_time else None
            old_shipment.pickup_number = pickup_number
            old_shipment.shipment_cargo_id = shipment_cargo_id           
            old_shipment.is_canceled = False
            old_shipment.status = ""
            if page != "01_appointment" and page != "arm_appointment":
                old_shipment.in_use = True
            await sync_to_async(old_shipment.save)()         
            context.update( {'success_messages':f"{appointment_id_old}йҐДзЇ¶дњ°жБѓдњЃжФєжИРеКЯпЉБ"})
            if name == "fleet_departure":
                return await self.handle_fleet_schedule_post(request,context)
            
            if page == "arm_appointment":
                return await self.handle_unscheduled_pos_post(request,context)
            elif page == "01_appointment":
                return await self.handle_appointment_management_post(request,context)
            return await self.handle_td_shipment_post(request,context)
        else:
            context = await self._check_ISA_is_repetition(appointment_id_new,destination)
            if context.get('success_messages'):
                old_shipment.appointment_id = appointment_id_new
                old_shipment.destination = destination
                old_shipment.load_type = load_type
                old_shipment.origin = origin
                old_shipment.pickup_time = pickup_time if pickup_time else None
                old_shipment.pickup_number = pickup_number
                old_shipment.shipment_cargo_id = shipment_cargo_id
                old_shipment.is_canceled = False
                old_shipment.status = ""
                old_shipment.shipment_appointment = shipment_appointment if shipment_appointment else None
                if page != "01_appointment" and page != "arm_appointment":
                    old_shipment.in_use = True
                await sync_to_async(old_shipment.save)()
            if name == "fleet_departure":
                return await self.handle_fleet_schedule_post(request,context)
            page = request.POST.get("page")
            if page == "arm_appointment":
                return await self.handle_unscheduled_pos_post(request,context)
            elif page == "01_appointment":
                return await self.handle_appointment_management_post(request,context)
            return await self.handle_td_shipment_post(request,context)
    
    async def handle_save_external_shipment_post(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        """дњЭе≠Ше§ЦйЕНзЪДзЇ¶зЪДдњЃжФє"""
        context = {}
        shipment_id = request.POST.get('shipment_id', '').strip()
        appointment_id_new = request.POST.get('appointment_id', '').strip()
        shipment_appointment = request.POST.get('shipment_appointment')
        pickup_time = request.POST.get('pickup_time')
        
        if not shipment_id:
            context.update({'error_messages': 'жЬ™жПРдЊЫ shipment_id!'})
            return await self.handle_td_shipment_post(request, context)
        
        old_shipments = await sync_to_async(list)(
            Shipment.objects.filter(id=shipment_id)
        )
        
        if not old_shipments:
            context.update({'error_messages': f"жЬ™жЙЊеИ∞ ID={shipment_id}!"})
            return await self.handle_td_shipment_post(request, context)
        
        old_shipment = old_shipments[0]
        appointment_id_old = old_shipment.appointment_id
        
        if appointment_id_new == appointment_id_old:
            # йҐДзЇ¶еПЈдЄНеПШпЉМеП™жЫіжЦ∞еЕґдїЦе≠ЧжЃµ
            old_shipment.shipment_appointment = shipment_appointment if shipment_appointment else None
            old_shipment.pickup_time = pickup_time if pickup_time else None
            await sync_to_async(old_shipment.save)()
            context.update({'success_messages': f"{appointment_id_old}йҐДзЇ¶дњ°жБѓдњЃжФєжИРеКЯпЉБ"})
        else:
            # йҐДзЇ¶еПЈеПШеМЦпЉМж£АжЯ•йЗНе§Н
            context = await self._check_ISA_is_repetition(appointment_id_new, old_shipment.destination)
            if context.get('success_messages'):
                old_shipment.appointment_id = appointment_id_new
                old_shipment.shipment_appointment = shipment_appointment if shipment_appointment else None
                old_shipment.pickup_time = pickup_time if pickup_time else None
                await sync_to_async(old_shipment.save)()
                if not context.get('success_messages'):
                    context.update({'success_messages': f"{appointment_id_old}йҐДзЇ¶дњ°жБѓдњЃжФєжИРеКЯпЉБ"})
        
        return await self.handle_td_shipment_post(request, context)
    
    async def _check_ISA_is_repetition(self,appointment_id,destination):
        context = {}
        try:
            existed_appointment = await sync_to_async(Shipment.objects.get)(
                appointment_id=appointment_id
            )
            if existed_appointment:
                if existed_appointment.in_use:             
                    context.update({'error_messages': f"{appointment_id}еЈ≤зїПеЬ®дљњзФ®дЄ≠пЉБ"})
                elif existed_appointment.is_canceled:
                    context.update({'error_messages': f"{appointment_id}еЈ≤зїПиҐЂеПЦжґИпЉБ"})
                elif (
                    existed_appointment.shipment_appointment.replace(tzinfo=pytz.UTC)
                    < timezone.now()
                ):
                    context.update({'error_messages': f"{appointment_id}еЬ®е§ЗзЇ¶дЄ≠зЩїиЃ∞зЪДжЧґйЧіжЧ©дЇОељУеЙНжЧґйЧіпЉМиѓЈеЕИдњЃжФєе§ЗзЇ¶пЉБ"})
                elif existed_appointment.destination != destination:
                    context.update({'error_messages': f"{appointment_id}еЬ®е§ЗзЇ¶дЄ≠зЩїиЃ∞зЪДзЫЃзЪДеЬ∞еТМжЬђжђ°дњЃжФєзЪДзЫЃзЪДеЬ∞дЄНеРМпЉБ"})
        except Shipment.DoesNotExist:
            # е¶ВжЮЬжЯ•дЄНеИ∞иЃ∞ељХпЉМзЫіжО•ињФеЫЮжИРеКЯжґИжБѓ
            context.update({'success_messages':f"еЈ≤жЫіжНҐдЄЇжЦ∞зЪД{appointment_id}пЉБ"})
        return context
    
    async def handle_cancel_appointment_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        appointment_id = request.POST.get('appointment_id')
        shipment = await sync_to_async(Shipment.objects.get)(
            appointment_id=appointment_id
        )
        shipment_batch_number = shipment.shipment_batch_number

        request.POST = request.POST.copy()
        request.POST['shipment_batch_number'] = shipment_batch_number
        request.POST['type'] = 'td'     
        sm = ShippingManagement()
        context = await sm.handle_cancel_post(request,'post_nsop')   
        page = request.POST.get("page")
        if page == "arm_appointment":
            return await self.handle_unscheduled_pos_post(request,context)    
        else:  
            return await self.handle_td_shipment_post(request,context)
    
    async def handle_shipment_add_pallet_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        context = {}
        appointment_id = request.POST.get("appointment_id")
        pl_ids_str = request.POST.getlist("cargo_ids")
        selected = []
        if pl_ids_str:
            for id_str in pl_ids_str:
                if id_str.strip():
                    selected.extend([int(x.strip()) for x in id_str.split(',') if x.strip().isdigit()])

        plt_ids_str = request.POST.getlist("plt_ids")
        selected_plt = []
        if plt_ids_str:
            for id_str in plt_ids_str: 
                if id_str.strip():
                    selected_plt.extend([int(x.strip()) for x in id_str.split(',') if x.strip().isdigit()])

        if not selected and not selected_plt:
            context.update({"error_messages": f"{appointment_id}ж≤°жЬЙжЙЊеИ∞и¶БжЈїеК†poзЪДidпЉБ"})
            return await self.handle_td_shipment_post(request,context)

        shipment = await sync_to_async(Shipment.objects.get)(
            appointment_id=appointment_id
        )

        shipment_batch_number = shipment.shipment_batch_number
        request.POST = request.POST.copy()
        request.POST['alter_type'] = 'add'
        request.POST['pl_ids'] = selected
        request.POST['plt_ids'] = selected_plt
        request.POST['shipment_batch_number'] = shipment_batch_number           
        sm = ShippingManagement()
        info = await sm.handle_alter_po_shipment_post(request,'post_nsop') 
        
        context.update({"success_messages": f"{appointment_id}жЈїеК†жИРеКЯпЉБ"})
        page = request.POST.get("page")
        if page == "arm_appointment":
            return await self.handle_unscheduled_pos_post(request,context)
        else:
            return await self.handle_td_shipment_post(request,context)
    
    async def _get_unique_pickup_number(self, pickup_number: str) -> str:
        """
        иОЈеПЦеФѓдЄАзЪДpickup_numberпЉМе¶ВжЮЬжЬЙйЗНе§НеИЩиЗ™еК®жЈїеК†еЇПеПЈ
        """
        if not pickup_number:
            # е¶ВжЮЬpickup_numberдЄЇз©ЇпЉМињФеЫЮдЄАдЄ™йїШиЃ§еАЉ
            import random
            prefix = 'ZEM-RC-' 
            today = datetime.now()
            month = str(today.month).zfill(2)
            day = str(today.day).zfill(2)
            month_day = month + day
            random_num = str(random.randint(1000, 9999))
            return f"{prefix}{month_day}-{random_num}"
        
        base_number = pickup_number
        counter = 1
        pickup_number = f"{base_number}-{counter}"
        
        while await Shipment.objects.filter(pickup_number=pickup_number).aexists():
            counter += 1
            pickup_number = f"{base_number}-{counter}"
                 
        return pickup_number

    async def _get_pl_plt_total_weight(self, ids: list[int], plt_ids: list[int]) -> Tuple[float, float, int, int]:
        total_weight = 0.0
        total_cbm = 0.0
        total_pcs = 0
        total_pallet = 0
        
        # иЃ°зЃЧPackingListзЪДзїЯиЃ°дњ°жБѓ
        if ids:
            # дљњзФ®aggregateиЃ°зЃЧжАїеТМ
            packinglist_stats = await sync_to_async(
                PackingList.objects.filter(id__in=ids).aggregate
            )(
                total_weight_sum=Sum('total_weight_lbs'),
                total_cbm_sum=Sum('cbm'),
                total_pcs_sum=Sum('pcs'),
            )
            
            total_weight += packinglist_stats.get('total_weight_sum') or 0.0
            total_cbm += packinglist_stats.get('total_cbm_sum') or 0.0
            total_pcs += packinglist_stats.get('total_pcs_sum') or 0
            # PackingListзЪДжЭњжХ∞йЬАи¶БзЙєжЃКиЃ°зЃЧпЉЪcbm/1.8еПЦдЄКйЩР
            if packinglist_stats.get('total_cbm_sum'):
                total_pallet += math.ceil(packinglist_stats['total_cbm_sum'] / 1.8)
        
        # иЃ°зЃЧPalletзЪДзїЯиЃ°дњ°жБѓ
        if plt_ids:
            # дљњзФ®aggregateиЃ°зЃЧжАїеТМ
            pallet_stats = await sync_to_async(
                Pallet.objects.filter(id__in=plt_ids).aggregate
            )(
                total_weight_sum=Sum('weight_lbs'),
                total_cbm_sum=Sum('cbm'),
                total_pcs_sum=Sum('pcs'),
                pallet_count=Count('id'),  # жѓПдЄ™palletзЃЧдЄАдЄ™жЭњ
            )
            
            total_weight += pallet_stats.get('total_weight_sum') or 0.0
            total_cbm += pallet_stats.get('total_cbm_sum') or 0.0
            total_pcs += pallet_stats.get('total_pcs_sum') or 0
            total_pallet += pallet_stats.get('pallet_count') or 0
        
        # еЫЫиИНдЇФеЕ•е§ДзРЖ
        total_weight = round(total_weight, 2)
        total_cbm = round(total_cbm, 3)
        total_pallet = math.ceil(total_pallet)  # з°ЃдњЭжЭњжХ∞жШѓжХіжХ∞
        return (float(total_weight), float(total_cbm), int(total_pcs),  int(total_pallet) )

    async def handle_ltl_bind_group_shipment(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:    
        context = {}
        # иОЈеПЦи°®еНХжХ∞жНЃ
        cargo_ids = request.POST.get('cargo_ids', '').strip()
        plt_ids = request.POST.get('plt_ids', '').strip()
        destination = request.POST.get('destination', '').strip()
        address = request.POST.get('address', '').strip()
        carrier = request.POST.get('carrier', '').strip()
        supplier = request.POST.get('supplier', '').strip()
        shipment_appointment = request.POST.get('shipment_appointment', '').strip()
        if not shipment_appointment:
            raise ValueError("жПРиіІжЧґйЧідЄНиГљдЄЇз©ЇпЉБ")
        arm_bol = request.POST.get('arm_bol', '').strip()
        arm_pro = request.POST.get('arm_pro', '').strip()
        is_print_label = request.POST.get('is_print_label', 'false').strip() == 'true'
        shipment_type = request.POST.get('shipment_type', '').strip()
        warehouse = request.POST.get('warehouse')
        auto_fleet = request.POST.get('auto_fleet', 'true').strip()
        auto_fleet_bool = auto_fleet.lower() == 'true'
        note = request.POST.get('note', '').strip()

        nj_shipped = False
        if shipment_type == "еЃҐжИЈиЗ™жПР" and "NJ" in warehouse: 
            nj_shipped = True
            auto_fleet_bool = True
        
        # иІ£жЮРIDеИЧи°®
        packinglist_ids = []
        pallet_ids = []
        
        if cargo_ids:
            packinglist_ids = [int(id.strip()) for id in cargo_ids.split(',') if id.strip()]
        
        if plt_ids:
            pallet_ids = [int(id.strip()) for id in plt_ids.split(',') if id.strip()]
        
        if not packinglist_ids and not pallet_ids:
            context = {'error_messages':'иѓЈйАЙжЛ©и¶БзїСеЃЪзЪДиіІзЙ©пЉБ'}
            return await self.handle_ltl_unscheduled_pos_post(request, context)
        
        #иЃ°зЃЧжАїйЗНйЗПз≠ЙеПВжХ∞
        total_weight, total_cbm, total_pcs, total_pallet = await self._get_pl_plt_total_weight(packinglist_ids,pallet_ids)
        # жЧґйЧіе≠ЧжЃµе§ДзРЖ
        try:
            pickup_time = timezone.make_aware(datetime.fromisoformat(shipment_appointment.replace('Z', '')))
        except (ValueError, TypeError):
            context = {'error_messages':'жПРиіІжЧґйЧіж†ЉеЉПйФЩиѓѓпЉБ'}
            return await self.handle_ltl_unscheduled_pos_post(request, context)
        
        if auto_fleet_bool:
            current_time = datetime.now()
            try:
                shipment_appointment_dt = datetime.fromisoformat(shipment_appointment.replace('Z', ''))
                month_day = shipment_appointment_dt.strftime("%m%d")
            except:
                month_day = current_time.strftime("%m%d")
            pickupNumber = "ZEM" + "-" + warehouse + "-" + "" + month_day + carrier + destination

            fleet_cost = (request.POST.get("fleet_cost", ""))
            if not fleet_cost:
                fleet_cost = 0.0
            else:
                fleet_cost = float(fleet_cost)
            fleet = Fleet(
                **{
                    "carrier": request.POST.get("carrier").strip(),
                    "Supplier": supplier,
                    "fleet_type": shipment_type,
                    "pickup_number": pickupNumber,
                    "appointment_datetime": shipment_appointment,  # иљ¶жђ°зЪДжПРиіІжЧґйЧі
                    "fleet_number": "FO"
                    + current_time.strftime("%m%d%H%M%S")
                    + str(uuid.uuid4())[:2].upper(),
                    "scheduled_at": current_time,
                    "total_weight": total_weight,
                    "total_cbm": total_cbm,
                    "total_pallet": total_pallet,
                    "total_pcs": total_pcs,
                    "origin": warehouse,
                    "fleet_cost": fleet_cost,
                }
            )
            # NJдїУзЪДеЃҐжИЈиЗ™жПРеТМUPSпЉМйГљдЄНйЬАи¶Бз°ЃиЃ§еЗЇеЇУеТМз°ЃиЃ§еИ∞иЊЊпЉМеЃҐжИЈиЗ™жПРйЬАи¶БPODдЄКдЉ†
            if nj_shipped: 
                fleet.departured_at = shipment_appointment
                fleet.arrived_at = shipment_appointment
            await sync_to_async(fleet.save)()

        maersk_batch_number = request.POST.get('maersk_batch_number')
        if maersk_batch_number:
             batch_number = maersk_batch_number
        else:
            if len(destination) > 8:
                destination_name = destination[:8]
            else:
                destination_name = destination
            batch_number = await self.generate_unique_batch_number(destination_name)

        shipment_appointment_tz = self._parse_datetime(shipment_appointment)
        tzinfo = self._parse_tzinfo(warehouse)
        shipmentappointment_utc = self._parse_ts(shipment_appointment, tzinfo)
        current_time = timezone.now()
        # еИЫеїЇShipmentиЃ∞ељХ
        shipment_data = {
            'shipment_batch_number': batch_number,
            'shipment_type': shipment_type,
            'destination': destination,
            'address': address,
            'carrier': carrier,
            'ARM_BOL': arm_bol,
            'ARM_PRO': arm_pro,
            'is_print_label': is_print_label,
            'pickup_time': pickup_time,
            'shipment_schduled_at': current_time,
            'shipment_appointment': shipment_appointment,
            'shipment_appointment_tz': shipment_appointment_tz,
            'shipment_appointment_utc': shipmentappointment_utc,
            'total_weight': total_weight,
            'total_cbm': total_cbm,
            'total_pallet': total_pallet,
            'total_pcs': total_pcs,
            'origin': warehouse,
            'note': note,
        }
        if auto_fleet_bool:
            shipment_data['fleet_number'] = fleet
        if nj_shipped: 
            # еЃҐжИЈиЗ™жПРзЪДйҐДзЇ¶еЃМи¶БзЫіжО•иЈ≥еИ∞PODдЄКдЉ†,жЧґйЧіжМЙйҐДиЃ°жПРиіІжЧґйЧі
            shipment_data.update({
                'is_shipped': True,
                'shipped_at': shipment_appointment,
                'shipped_at_utc': shipmentappointment_utc,
                'is_arrived': True,
                'arrived_at': shipment_appointment,
                'arrived_at_utc': shipmentappointment_utc
            })
        shipment = await sync_to_async(Shipment.objects.create)(**shipment_data)
        
        if packinglist_ids:
            # зЫіжО•дљњзФ®update()жЦєж≥ХжЙєйЗПжЫіжЦ∞
            await sync_to_async(
                PackingList.objects.filter(id__in=packinglist_ids).update
            )(
                shipment_batch_number=shipment,
                master_shipment_batch_number=shipment
            )

        if pallet_ids:
            # зЫіжО•дљњзФ®update()жЦєж≥ХжЙєйЗПжЫіжЦ∞
            await sync_to_async(
                Pallet.objects.filter(id__in=pallet_ids).update
            )(
                shipment_batch_number=shipment,
                master_shipment_batch_number=shipment
            )
        
        success_msg = f'йҐДзЇ¶еЗЇеЇУзїСеЃЪжИРеКЯ! <br>жЙєжђ°еПЈжШѓ:{batch_number}!'
        if auto_fleet_bool:
            success_msg += ' <br>еЈ≤иЗ™еК®жОТиљ¶пЉБ'
        else:
            success_msg += ' <br>йЬАжЙЛеК®жОТиљ¶пЉБ'
        context = {'success_messages': mark_safe(success_msg)}
        return await self.handle_ltl_unscheduled_pos_post(request, context)
    
    def _parse_datetime(self, datetime_string: str) -> tuple[str, str]:
        try:
            dt = datetime.fromisoformat(datetime_string)
            return ""   # дљ†еОЯеЗљжХ∞еП™ињФеЫЮ timezoneпЉМињЩйЗМдњЭжМБзїУжЮДдЄНз†іеЭП
        except ValueError:
            pass
        datetime_pattern = re.compile(
            r"""
            (?P<month>\d{1,2})          # Month (1 or 2 digits)
            [/\-]                       # Separator (/, -)
            (?P<day>\d{1,2})            # Day (1 or 2 digits)
            [/\-]                       # Separator (/, -)
            (?P<year>\d{4})             # Year (4 digits)
            [, ]*                       # Optional comma or space
            (?P<hour>\d{1,2})           # Hour (1 or 2 digits)
            [:]                         # Colon
            (?P<minute>\d{2})           # Minute (2 digits)
            [ ]*                        # Optional space
            (?P<period>AM|PM)           # AM or PM
            [ ,]*                       # Optional comma or space
            (?P<timezone>[A-Z+\-:\d]*)? # Optional timezone
            """,
            re.VERBOSE | re.IGNORECASE,
        )
        match = datetime_pattern.search(datetime_string)
        if not match:
            raise ValueError(f"Invalid datetime format: {datetime_string}")

        parts = match.groupdict()
        month, day, year = int(parts["month"]), int(parts["day"]), int(parts["year"])
        hour, minute = int(parts["hour"]), int(parts["minute"])
        period = parts["period"].upper()
        timezone = parts["timezone"].strip() if parts["timezone"] else ""
        return timezone
    
    async def handle_pl_ids_to_plt_ids(self, selected: list, selected_plt: list) -> tuple:
        """
        зКґжАБиљђжНҐж†ЄењГйАїиЊСпЉЪ
        1. ж£АжЯ•йАЙдЄ≠зЪД PL жШѓеР¶еЈ≤жЙУжЭњгАВ
        2. иЛ•еЈ≤жЙУжЭњпЉМдїО selected дЄ≠зІїйЩ§иѓ• PL IDгАВ
        3. е∞ЖеѓєеЇФзЪД Pallet ID еК†еЕ• selected_pltгАВ
        ињФеЫЮ: (е§ДзРЖеРОзЪДselected, е§ДзРЖеРОзЪДselected_plt)
        """
        if not selected:
            return selected, selected_plt

        # 1. жЙєйЗПиОЈеПЦйАЙдЄ≠ PackingList зЪД PO_ID жШ†е∞Д
        # дљњзФ® values_list жПРйЂШжЯ•иѓҐжХИзОЗ
        pl_data = await sync_to_async(list)(
            PackingList.objects.filter(id__in=selected)
            .values("id", "PO_ID", "container_number")
        )

        if not pl_data:
            return selected, selected_plt

        # жЮДеїЇжЯ•иѓҐжЭ°дїґж±†пЉЪжѓПдЄАдЄ™жЭ°дїґйГљжШѓ (PO_ID AND жЯЬеПЈ)
        condition_queries = Q()
        pl_info_map = {} # иЃ∞ељХ pl_id еѓєеЇФзЪДзЙєеЊБпЉМжЦєдЊњеРОзї≠еЙФйЩ§

        for item in pl_data:
            po = item["PO_ID"]
            container = item["container_number"]
            if po and container:
                # жЮДеїЇе§НеРИжЯ•иѓҐжЭ°дїґ
                condition_queries |= Q(PO_ID=po, container_number=container)
                pl_info_map[item["id"]] = (po, container)

        if not pl_info_map:
            return selected, selected_plt
        # 2. з≤Њз°ЃжЯ•жЙЊ Pallet и°®дЄ≠гАРеЃМеЕ®еМєйЕНгАСињЩдЇЫзїДеРИзЪДиЃ∞ељХ
        pallets = await sync_to_async(list)(
            Pallet.objects.filter(condition_queries)
            .values("id", "PO_ID", "container_number")
        )

        if not pallets:
            return selected, selected_plt

        # 3. зїЯиЃ°еУ™дЇЫ (PO_ID, жЯЬеПЈ) зїДеРИеЈ≤зїПжЙУжЭњдЇЖ
        finished_combinations = {
            (p["PO_ID"], p["container_number"]) for p in pallets
        }
        
        # жФґйЫЖеМєйЕНеИ∞зЪДжЙАжЬЙ Pallet ID
        new_found_plt_ids = [p["id"] for p in pallets]

        # 4. жЙІи°МзКґжАБиљђжНҐпЉЪ
        # е¶ВжЮЬжЯРжЭ° PL зЪД (PO, жЯЬеПЈ) зїДеРИеЬ® Pallet и°®йЗМжЙЊеИ∞дЇЖпЉМе∞±дїО selected дЄ≠зІїйЩ§
        updated_selected = [
            pl_id for pl_id in selected 
            if pl_id not in pl_info_map or pl_info_map[pl_id] not in finished_combinations
        ]

        # еРИеєґ Pallet ID еИЧи°®еєґеОїйЗН
        updated_selected_plt = list(set(selected_plt + new_found_plt_ids))
        return updated_selected, updated_selected_plt

    async def handle_appointment_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:  
        '''жЄѓеРОзЪДйҐДзЇ¶еЗЇеЇУеКЯиГљ''' 
        context = {}
        shipment_type = request.POST.get('shipment_type')
        page = request.POST.get("page")
        if shipment_type == "FTL":         
            pickup_number_raw = request.POST.get('pickupNumber')        
            if page == "arm_appointment":
                pickup_number = await self._get_unique_pickup_number(pickup_number_raw)
            else:
                pickup_number = pickup_number_raw
        else:
            pickup_number = None
        appointment_id = request.POST.get('appointment_id')
        shipment_batch_number_in = (request.POST.get('shipment_batch_number') or '').strip()
        ids = request.POST.get("cargo_ids")
        plt_ids = request.POST.get("plt_ids")
        selected = [int(i) for i in ids.split(",") if i]
        selected_plt = [int(i) for i in plt_ids.split(",") if i]

        # е§ДзРЖжЙУжЭњжХ∞жНЃдЄЇеРМж≠•еИ∞йҐДзЇ¶еЗЇеЇУзХМйЭҐзЪДйЧЃйҐШ
        selected, selected_plt = await self.handle_pl_ids_to_plt_ids(selected, selected_plt)
        if not selected and not selected_plt:
            context.update({"error_messages": f"ж≤°жЬЙйАЙжЛ©POпЉБ"}) 
            if page == "arm_appointment":          
                return await self.handle_unscheduled_pos_post(request,context)
            else:
                return await self.handle_td_shipment_post(request,context)
        
        operation_type = request.POST.get('operation_type')
        shipment_cargo_id = request.POST.get('shipment_cargo_id')
        
        if operation_type == "remove_po":            
            try:
                if appointment_id:
                    shipment = await sync_to_async(Shipment.objects.get)(appointment_id=appointment_id)
                elif shipment_batch_number_in:
                    shipment = await sync_to_async(Shipment.objects.get)(shipment_batch_number=shipment_batch_number_in)
                else:
                    raise ObjectDoesNotExist
            except ObjectDoesNotExist:
                context.update({"error_messages": "жЬ™жЙЊеИ∞еѓєеЇФйҐДзЇ¶иЃ∞ељХ"})
                if page == "arm_appointment":
                    return await self.handle_unscheduled_pos_post(request, context)
                return await self.handle_td_shipment_post(request, context)

            shipment_batch_number = shipment.shipment_batch_number
            
            request.POST = request.POST.copy()
            request.POST['alter_type'] = 'remove'
            request.POST['pl_ids'] = selected
            request.POST['plt_ids'] = selected_plt
            request.POST['shipment_batch_number'] = shipment_batch_number    
            sm = ShippingManagement()
            info = await sm.handle_alter_po_shipment_post(request,'post_nsop') 
            context.update({"success_messages": f"еИ†йЩ§йГ®еИЖPOпЉМжЙєжђ°еПЈжШѓ{shipment_batch_number}"})
            if page == "arm_appointment":
                return await self.handle_unscheduled_pos_post(request,context)
            else:
                return await self.handle_td_shipment_post(request,context)
        
        destination = request.POST.get('destination')                  
        
        if selected or selected_plt:
            packing_list_selected = await self._get_packing_list(
                request.user,
                models.Q(id__in=selected)
                & models.Q(shipment_batch_number__isnull=True),
                models.Q(id__in=selected_plt)
                & models.Q(shipment_batch_number__isnull=True),
            )
            total_weight, total_cbm, total_pcs, total_pallet = 0.0, 0.0, 0, 0
            for pl in packing_list_selected:
                total_weight += (
                    pl.get("total_weight_lbs") if pl.get("total_weight_lbs") else 0
                )
                total_cbm += pl.get("total_cbm") if pl.get("total_cbm") else 0
                total_pcs += pl.get("total_pcs") if pl.get("total_pcs") else 0
                if pl.get("label") == "ACT":
                    total_pallet += pl.get("total_n_pallet_act")
                else:
                    if pl.get("total_n_pallet_est < 1"):
                        total_pallet += 1
                    elif pl.get("total_n_pallet_est") % 1 >= 0.45:
                        total_pallet += int(pl.get("total_n_pallet_est") // 1 + 1)
                    else:
                        total_pallet += int(pl.get("total_n_pallet_est") // 1)
            address = request.POST.get('address')
            if not address:
                address = await self.get_address(destination)
                if not address:
                    raise ValueError('ж≤°жЙЊеИ∞еЬ∞еЭА')
            
            #еЕИеОїжЯ•иѓҐдЄАдЄЛshipmentи°®пЉМжЬЙж≤°жЬЙињЩдЄ™иЃ∞ељХпЉМе∞±жШѓзђђдЄАжђ°йҐДзЇ¶еЗЇеЇУпЉМе¶ВжЮЬжЬЙе∞±жШѓдњЃжФє
            try:
                if appointment_id:
                    shipment = await sync_to_async(Shipment.objects.get)(appointment_id=appointment_id)
                elif shipment_batch_number_in:
                    shipment = await sync_to_async(Shipment.objects.get)(shipment_batch_number=shipment_batch_number_in)
                else:
                    raise ObjectDoesNotExist
                if shipment.shipment_batch_number:  #еЈ≤зїПжЬЙжЙєжђ°еПЈдЇЖпЉМиѓіжШОињЩжШѓдњЃжФєPOзЪД
                    shipment_batch_number = shipment.shipment_batch_number
                else:
                    shipment_batch_number = await self.generate_unique_batch_number(destination)

                #дЄНзЃ°дєЛеЙНжАОдєИж†ЈпЉМзЫЃеЙНйГљжШѓи¶БйЗНжЦ∞жМЙplt_ids/pl_idsйЗНжЦ∞зїСеЃЪпЉМжЙАдї•и¶БжККдї•еЙНдЄїзЇ¶/зЇ¶зїСеЃЪињЩдЄ™зЪДиІ£зїС               
                if selected_plt: 
                    await sync_to_async(
                        Pallet.objects.filter(master_shipment_batch_number=shipment).update
                    )(master_shipment_batch_number=None)
                    await sync_to_async(
                        Pallet.objects.filter(shipment_batch_number=shipment).update
                    )(shipment_batch_number=None)
                if selected:  #дЄНзЃ°дєЛеЙНжАОдєИж†ЈпЉМзЫЃеЙНйГљжШѓи¶БйЗНжЦ∞жМЙplt_ids/pl_idsйЗНжЦ∞зїСеЃЪпЉМжЙАдї•и¶БжККдї•еЙНзЪДиІ£зїС
                    await sync_to_async(
                        PackingList.objects.filter(master_shipment_batch_number=shipment).update
                    )(master_shipment_batch_number=None)
                    await sync_to_async(
                        PackingList.objects.filter(shipment_batch_number=shipment).update
                    )(shipment_batch_number=None)
            except ObjectDoesNotExist:
                #жЙЊдЄНеИ∞пЉМйВ£е∞±жЦ∞еїЇдЄАжЭ°иЃ∞ељХ
                shipment_batch_number = await self.generate_unique_batch_number(destination)
                               
            except MultipleObjectsReturned:
                context.update({"error_messages": f"е≠ШеЬ®е§ЪжЭ°йЗНе§НзЪД{appointment_id}!"})  
                if page == "arm_appointment":
                    return await self.handle_unscheduled_pos_post(request,context)
                else:
                    return await self.handle_td_shipment_post(request,context)          

            shipment_type = request.POST.get('shipment_type')
            shipment_data = {
                'shipment_batch_number': shipment_batch_number,
                'destination': destination,
                'total_weight': total_weight,
                'total_cbm': total_cbm,
                'total_pallet': total_pallet,
                'total_pcs': total_pcs,
                'total_pallet': total_pallet,
                'shipment_type': shipment_type,
                'shipment_account': request.POST.get('shipment_account'),
                'appointment_id': appointment_id,
                'shipment_cargo_id': shipment_cargo_id,
                'load_type': request.POST.get('load_type'),
                'origin': request.POST.get('warehouse'),
                'note': request.POST.get('note'),
                'address': address,
                'pickup_number': pickup_number,
                'pickup_time': request.POST.get('pickup_time'),
            }
            request.POST = request.POST.copy()

            if shipment_type != 'е§ЦйЕН':
                shipment_data['shipment_appointment'] = request.POST.get('shipment_appointment')
                request.POST['shipment_appointment'] = request.POST.get('shipment_appointment')
            
            request.POST['shipment_data'] = str(shipment_data)
            request.POST['batch_number'] = shipment_batch_number   
            request.POST['address'] = address      
            request.POST['pl_ids'] = selected
            request.POST['plt_ids'] = selected_plt
            request.POST['type'] = 'td'
            request.POST['origin'] = request.POST.get('warehouse')  
            request.POST['load_type'] = request.POST.get('load_type')  
            request.POST['note'] = request.POST.get('note')  
            request.POST['destination'] = destination
            request.POST['shipment_type'] = request.POST.get('shipment_type')  
            request.POST['appointment_id'] = request.POST.get('appointment_id')  
            request.POST['shipment_cargo_id'] = request.POST.get('shipment_cargo_id')  
            request.POST['pickup_number'] = pickup_number  
            request.POST['pickup_time'] = request.POST.get('pickup_time') 
            
            sm = ShippingManagement()
            info = await sm.handle_appointment_post(request,'post_nsop') 
            context.update({"success_messages": f"зїСеЃЪжИРеКЯпЉМжЙєжђ°еПЈжШѓ{shipment_batch_number},pickupNumberжШѓ{pickup_number}"})
            if page == "arm_appointment":
                return await self.handle_unscheduled_pos_post(request,context)
            else:
                return await self.handle_td_shipment_post(request,context)
        else:
            context.update({"error_messages": f"ж≤°жЬЙйАЙжЛ©POпЉБ"}) 
        # template_name = request.POST.get('template_name')
        # if template_name and template_name == "unshipment":
        #     return await self.handle_td_unshipment_post(request,context)
        if page == "arm_appointment":          
            return await self.handle_unscheduled_pos_post(request,context)
        else:
            return await self.handle_td_shipment_post(request,context)
    
    async def handle_fleet_confirmation_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:     
        '''з°ЃиЃ§еЗЇеЇУзХМйЭҐ'''
        page = request.POST.get("page")
        selected_ids_str = request.POST.get("selected_ids")
        fleet_cost = request.POST.get("fleet_cost")

        if fleet_cost:
            try:
                fleet_cost_value = float(fleet_cost)
            except (ValueError, TypeError):
                fleet_cost_value = 0.0
        else:
            fleet_cost_value = 0.0  
        error_message = None
        if selected_ids_str:
            try:
                selected_ids_list = json.loads(selected_ids_str)
                selected_ids = [int(id) for id in selected_ids_list]
            except (json.JSONDecodeError, ValueError) as e:
                # е§ДзРЖиІ£жЮРйФЩиѓѓ
                error_message = f"selected_ids еПВжХ∞ж†ЉеЉПйФЩиѓѓ: {e}"
                # ж†єжНЃдљ†зЪДйФЩиѓѓе§ДзРЖжЦєеЉПйАЙжЛ©
                raise ValueError(error_message)
        if page == "arm_appointment":
            selected_ids = await sync_to_async(list)(
                Shipment.objects.filter(appointment_id__in=selected_ids).values_list('id', flat=True)
            )

        request.POST = request.POST.copy()
        fm = FleetManagement()
        if selected_ids:
            #еЕИзФЯжИРfleet_number
            current_time = datetime.now()
            fleet_number = (
                "F"
                + current_time.strftime("%m%d%H%M%S")
                + str(uuid.uuid4())[:2].upper()
            )
            shipment_selected = await sync_to_async(list)(
                Shipment.objects.filter(id__in=selected_ids)
            )
            fleet_type = None
            shipment_types = set()
            total_weight, total_cbm, total_pcs, total_pallet = 0.0, 0.0, 0, 0
            for s in shipment_selected:
                shipment_types.add(s.shipment_type)
                total_weight += s.total_weight
                total_cbm += s.total_cbm
                total_pcs += s.total_pcs
                total_pallet += s.total_pallet
            if len(shipment_types) > 1:
                error_message = f"йАЙдЄ≠зЪДйҐДзЇ¶жЙєжђ°еМЕеРЂдЄНеРМзЪД shipment_type: {list(shipment_types)}"
                raise ValueError(error_message)
            else:
                fleet_type = shipment_selected[0].shipment_type if shipment_selected else None
            fleet_data_dict = {
                'fleet_number': fleet_number,
                'fleet_type': fleet_type,
                'origin': request.POST.get('warehouse'),
                'total_weight': total_weight,
                'total_cbm': total_cbm,
                'total_pallet': total_pallet,
                'total_pcs': total_pcs,
                'fleet_cost': fleet_cost_value,
            }
            
            request.POST['fleet_number'] = fleet_number
            request.POST['fleet_type'] = fleet_type
        request.POST['fleet_data'] = str(fleet_data_dict)
        request.POST['fleet_cost'] = fleet_cost_value
        request.POST['selected_ids'] = selected_ids
        
        info = await fm.handle_fleet_confirmation_post(request,'post_nsop')
        context = {}
        if error_message:
            context.update({"error_messages": error_message}) 
        _, context = await self.handle_td_shipment_post(request, context)
        context.update({"success_messages": f'жОТиљ¶жИРеКЯ!жЙєжђ°еПЈжШѓпЉЪ{fleet_number}'})   
        
        if page == "arm_appointment":
            return await self.handle_unscheduled_pos_post(request,context)
        elif page == "ltl_unscheduledFleet":
            return await self.handle_ltl_unscheduled_pos_post(request,context)
        return await self.handle_td_shipment_post(request, context)

    async def handle_appointment_time(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        shipment_id = request.POST.get("shipment_id")
        shipment = await sync_to_async(Shipment.objects.get)(
            id=shipment_id
        )
        operation = request.POST.get("operation")
        if operation == "edit":
            appointmentTime = request.POST.get("appointmentTime")
            naive_datetime = parse(appointmentTime).replace(tzinfo=None)
            shipment.shipment_appointment = naive_datetime
            await sync_to_async(shipment.save)()
            context = {'success_messages':'е§ЗзЇ¶дњЃжФєжИРеКЯпЉБ'}
        elif operation == "delete":
            shipment.is_canceled = True
            await sync_to_async(shipment.delete)()
            context = {'success_messages':'е§ЗзЇ¶еИ†йЩ§жИРеКЯпЉБ'}
        page = request.POST.get("page")
        if page == "arm_appointment":
            return await self.handle_unscheduled_pos_post(request,context)
        else:
            return await self.handle_appointment_management_post(request,context)

    async def handle_fleet_export_pos(self, request: HttpRequest) -> HttpResponse:
        '''еЈ≤жОТиљ¶зЪДPOеѓЉеЗЇ'''
        fleet_number = request.POST.get("export_po_fleet_number")
        if not fleet_number:
            raise ValueError('ж≤°жЬЙиОЈеПЦеИ∞иљ¶жђ°еПЈ')
        
        try:
            shipments = [shipment async for shipment in Shipment.objects.filter(fleet_number__fleet_number=fleet_number)]
        except Shipment.DoesNotExist:
            raise ValueError(f'ж≤°жЬЙжЙЊеИ∞{fleet_number}зЪДзЇ¶')
        
        pl_ids = [
            pk async for pk in PackingList.objects.filter(
                shipment_batch_number__in=shipments
            ).values_list('id', flat=True)
        ]
        plt_ids = [
            pk async for pk in Pallet.objects.filter(
                shipment_batch_number__in=shipments
            ).values_list('id', flat=True)
        ]
        if not pl_ids and not plt_ids:
            raise ValueError(f'{fleet_number}иљ¶йЗМж≤°жЬЙдїїдљХpoж≤°жЬЙиОЈеПЦеИ∞id')
        return await self._execute_export_logic(pl_ids, plt_ids)

    async def handle_shipment_export_pos(self, request: HttpRequest) -> HttpResponse:
        '''еЈ≤жОТзЇ¶зЪДPOеѓЉеЗЇ'''
        appointment_id = request.POST.getlist("appointment_id")[0]
        if not appointment_id:
            raise ValueError('ж≤°жЬЙиОЈеПЦеИ∞зЇ¶зЪДid')
        
        try:
            shipment = await Shipment.objects.aget(appointment_id=appointment_id)
        except Shipment.DoesNotExist:
            raise ValueError(f'ж≤°жЬЙжЙЊеИ∞{appointment_id}зЪДзЇ¶')
        
        pl_ids = [
            pk async for pk in PackingList.objects.filter(
                shipment_batch_number=shipment
            ).values_list('id', flat=True)
        ]
        plt_ids = [
            pk async for pk in Pallet.objects.filter(
                shipment_batch_number=shipment
            ).values_list('id', flat=True)
        ]
        if not pl_ids and not plt_ids:
            raise ValueError(f'{appointment_id}зЇ¶йЗМж≤°жЬЙдїїдљХpoж≤°жЬЙиОЈеПЦеИ∞id')
        return await self._execute_export_logic(pl_ids, plt_ids)
        

    async def handle_export_pos(self, request: HttpRequest) -> HttpResponse:
        cargo_ids_str_list = request.POST.getlist("cargo_ids")
        pl_ids = [
            int(pl_id) 
            for sublist in cargo_ids_str_list 
            for pl_id in sublist.split(",") 
            if pl_id.strip()  # йЭЮз©ЇжЙНиљђжНҐ
        ]
        plt_ids_str_list = request.POST.getlist("plt_ids")
        plt_ids = [
            int(plt_id) 
            for sublist in plt_ids_str_list 
            for plt_id in sublist.split(",") 
            if plt_id.strip()  # йЭЮз©ЇжЙНиљђжНҐ
        ]

        if not pl_ids and not plt_ids:
            raise ValueError('ж≤°жЬЙиОЈеПЦеИ∞id')
        return await self._execute_export_logic(pl_ids, plt_ids)
    
    async def _execute_export_logic(self, pl_ids: list, plt_ids: list) -> HttpResponse:
        all_data = []
        #еЕИжККplt_idжЙЊеИ∞еѓєеЇФзЪДpl_id
        if plt_ids:
            pallet_data = await sync_to_async(
                lambda: list(
                    Pallet.objects.select_related("container_number")
                    .filter(id__in=plt_ids,is_dropped_pallet=False)
                    .values(
                        "id",
                        "shipping_mark",
                        "fba_id",
                        "ref_id",
                        "address",
                        "zipcode",
                        "container_number__container_number",
                        "destination",
                        "cbm",
                        "pcs",
                        "PO_ID"
                    )
                    .annotate(
                        total_pcs=Sum("pcs"),
                        total_n_pallet_act=Count("id"),
                        total_cbm=Sum("cbm", output_field=FloatField()),
                        label=Value('ACT')
                    )
                    .distinct()
                    .order_by("destination", "container_number__container_number")
                )
            )()

            additional_packinglist_ids, unmatched_pallet_data = await self.find_packinglist_ids_by_pallet_data(pallet_data)

            # еРИеєґpackinglist IDs
            all_packinglist_ids = list(set(pl_ids + additional_packinglist_ids))
            # е§ДзРЖжЬ™еМєйЕНзЪДpalletжХ∞жНЃпЉИжЈїеК†ж†ЗиЃ∞пЉЙ
            if unmatched_pallet_data:
                for item in unmatched_pallet_data:
                    item = dict(item)
                    ref_ids = str(item["ref_id"]).split(",") if item["ref_id"] else [""]
                    
                    for ref_id in ref_ids:
                        new_row = item.copy()
                        new_row["ref_id"] = ref_id.strip()
                        new_row["check_id"] = "жЬ™жЙЊеИ∞еѓєеЇФPOиЃ∞ељХпЉМиѓЈжЙЛеК®е§ДзРЖ"
                        new_row["is_unmatched"] = True
                        new_row["total_n_pallet_est"] = new_row.get("total_n_pallet_act", 0)
                        all_data.append(new_row)
        else:
            all_packinglist_ids = pl_ids

        # жЯ•жЙЊжЯЬеПЈдЄЛзЪДpl
        packing_list = await sync_to_async(
            lambda: list(
                PackingList.objects.select_related("container_number", "pallet")
                .filter(id__in=all_packinglist_ids)
                .values(
                    "id",
                    "shipping_mark",
                    "fba_id",
                    "ref_id",
                    "address",
                    "zipcode",
                    "container_number__container_number",
                    "destination",
                    "cbm"
                )
                .annotate(
                    total_pcs=Sum(
                        Case(
                            When(pallet__isnull=True, then=F("pcs")),
                            default=F("pallet__pcs"),
                            output_field=IntegerField(),
                        )
                    ),
                    total_n_pallet_est=Sum("cbm", output_field=FloatField()) / 2,
                    total_cbm=Sum("cbm", output_field=FloatField()),
                    label=Max(
                        Case(
                            When(pallet__isnull=True, then=Value("EST")),
                            default=Value("ACT"),
                            output_field=CharField(),
                        )
                    ),
                )
                .distinct()
                .order_by("destination", "container_number__container_number")
            )
        )()
        
        pl_ids_list = [pl["id"] for pl in packing_list]
        check_map = await sync_to_async(
            lambda: {
                p.packing_list_id: p.id
                for p in PoCheckEtaSeven.objects.filter(packing_list_id__in=pl_ids_list)
            }
        )()

        # е±ХеЉАжХ∞жНЃпЉЪе∞Жref_idжМЙйАЧеПЈеИЖеЙ≤жИРе§Ъи°М
        data = []
        for item in packing_list:
            item = dict(item)  # еЫ†дЄЇ values() ињФеЫЮзЪДжШѓ ValuesQuerySet
            item["check_id"] = check_map.get(item["id"])  # е¶ВжЮЬж≤°жЬЙеѓєеЇФиЃ∞ељХе∞±ињФеЫЮ None
            data.append(item)
        all_data += data
        if len(all_data) == 1:
            raise ValueError('ж≤°жЬЙжЙЊеИ∞дїїдљХжХ∞жНЃпЉБ')
        keep = [
            "shipping_mark",
            "container_number__container_number",
            "fba_id",
            "ref_id",
            "total_pcs",
            "Pallet Count",
            "label",
            "check_id",
            "is_valid",
            "total_cbm",
            "destination", 
        ]
        df = pd.DataFrame.from_records(all_data)

        
        df["is_valid"] = None

        def get_est_pallet(n):
            if n < 1:
                return 1
            elif n % 1 >= 0.45:
                return int(n // 1 + 1)
            else:
                return int(n // 1)
        if 'total_n_pallet_est' not in df.columns:
            df['total_n_pallet_est'] = 0
        else:
            df["total_n_pallet_est"] = df["total_n_pallet_est"].apply(get_est_pallet)

        if 'label' not in df.columns:
            df["label"] = "EST"
        df["est"] = df["label"] == "EST"
        df["Pallet Count"] = (
            df["total_n_pallet_est"] * df["est"]
        )
        df = df[keep].rename(
            {
                "fba_id": "PRO",
                "container_number__container_number": "BOL",
                "ref_id": "PO List (use , as separator) *",
                "total_pcs": "Carton Count",
                "total_cbm": "Total CBM",
                "destination": "Destination",
            },
            axis=1,
        )
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f"attachment; filename=PO.csv"
        df.to_csv(path_or_buf=response, index=False)
        return response

    async def find_packinglist_ids_by_pallet_data(self, pallet_data):
        """
        ж†єжНЃpalletжХ∞жНЃжЯ•жЙЊеѓєеЇФзЪДpackinglist IDs
        иІДеИЩпЉЪж†єжНЃPO_IDзЫЄеРМпЉМдЄФpackinglistзЪДfba_idеТМref_idеМЕеРЂеЬ®palletиЃ∞ељХдЄ≠
        """
        packinglist_ids = []
        unmatched_pallet_records = []
        
        for pallet_item in pallet_data:
            PO_ID = pallet_item.get('PO_ID')
            pallet_fba_id = pallet_item.get('fba_id', '').strip()
            pallet_ref_id = pallet_item.get('ref_id', '').strip()
            
            if not PO_ID:
                unmatched_pallet_records.append(pallet_item)
                continue
                
            # жЮДеїЇжЯ•иѓҐжЭ°дїґ
            base_query = models.Q(PO_ID=PO_ID)
            all_po_matching_packinglists = await sync_to_async(list)(
                PackingList.objects.filter(base_query).values('id', 'fba_id', 'ref_id')
            )
            matching_packinglists = [
                packinglist['id'] 
                for packinglist in all_po_matching_packinglists 
                if (not pallet_fba_id or (packinglist['fba_id'].strip() or '') in pallet_fba_id) and 
                (not pallet_ref_id or (packinglist['ref_id'].strip() or '') in pallet_ref_id)
            ]
            
            if matching_packinglists:
                packinglist_ids.extend(matching_packinglists)
            else:
                unmatched_pallet_records.append(pallet_item)
        
        # еОїйЗНеєґињФеЫЮ
        return list(set(packinglist_ids)), unmatched_pallet_records


    #ињЩдЄ™жШѓжМЙзЕІжЛњзЇ¶зЪДж®°жЭњеОїеѓЉеЗЇ   
    async def handle_export_pos_get_appointment(self, request: HttpRequest) -> HttpResponse:
        cargo_ids_str_list = request.POST.getlist("cargo_ids")
        pallet_ids = request.POST.getlist("plt_ids")

        packinglist_ids = []
        if cargo_ids_str_list and cargo_ids_str_list[0]:
            packinglist_ids = cargo_ids_str_list[0].split(',')
            packinglist_ids = [int(x) for x in packinglist_ids if x]
        else:
            packinglist_ids = []     
            
        if pallet_ids and pallet_ids[0]:
            pallet_ids = pallet_ids[0].split(',')
        else:
            pallet_ids = []
        if len(packinglist_ids) == 0 and pallet_ids == 0:
            raise ValueError('ж≤°жЬЙжЙЊеИ∞PO')
        all_data = []
        unmatched_pallet_data = []  # е≠ШеВ®жЙЊдЄНеИ∞еѓєеЇФpackinglistзЪДpalletиЃ∞ељХ
        if pallet_ids:
            pallet_data = await sync_to_async(list)(
                Pallet.objects.select_related("container_number")
                .filter(id__in=pallet_ids)
                .values(
                    "fba_id",
                    "ref_id",
                    "address",
                    "zipcode",
                    "destination",
                    "delivery_method",
                    "container_number__container_number",
                    "shipping_mark",
                )
                .annotate(
                    total_pcs=Sum("pcs", output_field=IntegerField()),
                    total_cbm=Sum("cbm", output_field=FloatField()),
                    total_weight_lbs=Sum("weight_lbs", output_field=FloatField()),
                    total_n_pallet_act=Count("pallet_id", distinct=True),
                    label=Value("ACT"),
                )
                .distinct()
                .order_by("destination", "container_number__container_number")
            )

            # ж†єжНЃpalletжХ∞жНЃжЯ•жЙЊеѓєеЇФзЪДpackinglist IDs
            additional_packinglist_ids, unmatched_pallet_data = await self.find_packinglist_ids_by_pallet_data(pallet_data)
            # еРИеєґpackinglist IDs
            all_packinglist_ids = list(set(packinglist_ids + additional_packinglist_ids))
            # е§ДзРЖжЬ™еМєйЕНзЪДpalletжХ∞жНЃпЉИжЈїеК†ж†ЗиЃ∞пЉЙ
            if unmatched_pallet_data:
                expanded_unmatched_data = []
                for item in unmatched_pallet_data:
                    item = dict(item)
                    ref_ids = str(item["ref_id"]).split(",") if item["ref_id"] else [""]
                    
                    for ref_id in ref_ids:
                        new_row = item.copy()
                        new_row["ref_id"] = ref_id.strip()
                        new_row["check"] = "жЬ™жЙЊеИ∞еѓєеЇФPOиЃ∞ељХпЉМиѓЈжЙЛеК®е§ДзРЖ"  # жЈїеК†зЙєжЃКж†ЗиЃ∞
                        new_row["is_unmatched"] = True  # ж†ЗиЃ∞дЄЇжЬ™еМєйЕНиЃ∞ељХ
                        expanded_unmatched_data.append(new_row)
                all_data += expanded_unmatched_data
        else:
            all_packinglist_ids = packinglist_ids

        if all_packinglist_ids:
            packing_list_data = await sync_to_async(list)(
                PackingList.objects.select_related("container_number", "pallet")
                .filter(id__in=all_packinglist_ids)
                .values(
                    "fba_id",
                    "ref_id",
                    "address",
                    "zipcode",
                    "destination",
                    "delivery_method",
                    "container_number__container_number",
                    "shipping_mark",
                )
                .annotate(
                    total_pcs=Sum("pcs"),
                    total_cbm=Sum("cbm"),
                    total_weight_lbs=Sum("total_weight_lbs"),
                    total_n_pallet_est=Sum("cbm", output_field=FloatField()) / 2,
                    label=Value("EST"),
                )
                .distinct()
                .order_by("destination", "container_number__container_number")
            )
            all_data += packing_list_data
        
        
        for p in all_data:
            if p.get("is_unmatched"):
                continue
            try:
                pl = await sync_to_async(PoCheckEtaSeven.objects.get)(
                    container_number__container_number=p["container_number__container_number"],
                    shipping_mark=p["shipping_mark"],
                    fba_id=p["fba_id"],
                    ref_id=p["ref_id"],
                )

                if not pl.last_eta_checktime and not pl.last_retrieval_checktime:
                    p["check"] = "жЬ™ж†°й™М"
                elif pl.last_retrieval_checktime and not pl.last_retrieval_status:
                    if pl.handling_method:
                        p["check"] = "е§±жХИ," + str(pl.handling_method)
                    else:
                        p["check"] = "е§±жХИжЬ™е§ДзРЖ"
                elif (
                    not pl.last_retrieval_checktime
                    and pl.last_eta_checktime
                    and not pl.last_eta_status
                ):
                    if pl.handling_method:
                        p["check"] = "е§±жХИ," + str(pl.handling_method)
                    else:
                        p["check"] = "е§±жХИжЬ™е§ДзРЖ"
                else:
                    p["check"] = "жЬЙжХИ"
            except PoCheckEtaSeven.DoesNotExist:
                p["check"] = "жЬ™жЙЊеИ∞иЃ∞ељХ"
            except MultipleObjectsReturned:
                p["check"] = "еФЫе§іFBA_REFйЗНе§Н"
        data = [i for i in all_data]
        export_format = request.POST.get("export_format", "PO")
        if export_format == "PO":
            keep = [
                "fba_id",
                "container_number__container_number",
                "destination",
                "total_cbm",
                "ref_id",
                "Pallet Count",
                "total_pcs",
                "label",
                "check",
            ]
        elif export_format == "FULL_TABLE":
            keep = [
                "container_number__container_number",
                "destination",
                "delivery_method",
                "fba_id",
                "ref_id",
                "total_cbm",
                "total_pcs",
                "total_weight_lbs",
                "Pallet Count",
                "label",
                "check",
            ]
        else:
            raise ValueError(f"unknown export_format option: {export_format}")
        df = pd.DataFrame.from_records(data)

        if "total_n_pallet_est" not in df.columns:
            df["total_n_pallet_est"] = 0
        if "total_n_pallet_act" not in df.columns:
            df["total_n_pallet_act"] = 0
        def get_est_pallet(n):
            if pd.isna(n) or n is None:
                return 0
            if n < 1:
                return 1
            elif n % 1 >= 0.45:
                return int(n // 1 + 1)
            else:
                return int(n // 1)

        if "total_n_pallet_est" in df.columns:
            df["total_n_pallet_est"] = df["total_n_pallet_est"].apply(get_est_pallet)

        df["est"] = df["label"] == "EST"
        df["act"] = df["label"] == "ACT"

        df["Pallet Count"] = (
            df.get("total_n_pallet_act", 0) * df["act"]
            + df.get("total_n_pallet_est", 0) * df["est"]
        )
        df = df[keep].rename(
            {
                "fba_id": "PRO",
                "container_number__container_number": "BOL",
                "ref_id": "PO List (use , as separator) *",
                "total_pcs": "Carton Count",
            },
            axis=1,
        )
        if export_format == "FULL_TABLE":
            df = df.rename(
                {
                    "total_cbm": "CBM",
                    "total_weight_lbs": "WEIGHT(LBS)",
                },
                axis=1,
            )
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename=PO.xlsx"
        df.to_excel(excel_writer=response, index=False, columns=df.columns)
        return response


    async def handle_unscheduled_pos_all_get(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        context = {
            "warehouse_options": self.warehouse_options
        }
        #if await self._validate_user_four_major_whs(request.user):
        #дЄНзЬЛжЭГйЩРдЇЖпЉМе∞±йїШиЃ§жЙУеЉАе∞±жШѓеЫЫе§ІдїУзЪД
        warehouse = 'LA-91761'
        new_request = HttpRequest()
        new_request.method = request.method
        new_request.POST = request.POST.copy()
        new_request.POST['warehouse'] = warehouse
        new_request.user = request.user
        return await self.handle_unscheduled_pos_post(new_request)
    
    async def handle_appointment_management_get(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        context = {"warehouse_options": self.warehouse_options}
        return self.template_main_dash, context

    async def handle_td_unshipment_get(    
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        context = {"warehouse_options": self.warehouse_options}
        return self.template_td_unshipment, context

    async def handle_td_shipment_get(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        context = {"warehouse_options": self.warehouse_options}
        return self.template_td_shipment, context
    

    async def handle_fleet_leader_check_get(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        """иОЈеПЦйЬАи¶Бж†ЄеѓєзЪДиљ¶йШЯдњ°жБѓ"""
        # иОЈеПЦз≠ЫйАЙеПВжХ∞
        start_date = request.POST.get('start_date') or request.GET.get('start_date')
        end_date = request.POST.get('end_date') or request.GET.get('end_date')
        destination = request.POST.get('destination') or request.GET.get('destination')
        
        # е¶ВжЮЬж≤°жЬЙиЃЊзљЃжЧґйЧіпЉМйїШиЃ§дљњзФ®жЬАињСдЄАдЄ™жЬИ
        if not start_date and not end_date:
            today = timezone.now().date()
            one_month_ago = today - timezone.timedelta(days=30)
            start_date = one_month_ago.strftime('%Y-%m-%d')
            end_date = today.strftime('%Y-%m-%d')
        
        # жЮДеїЇжЯ•иѓҐжЭ°дїґ
        query_conditions = models.Q(
            check_by_leader=False,
            fleet_type__in=['FTL','е§ЦйЕН','ењЂйАТ']
        )
        
        # жЈїеК†жЧґйЧіз≠ЫйАЙжЭ°дїґ
        if start_date:
            start_datetime = timezone.datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            query_conditions &= models.Q(appointment_datetime__gte=start_datetime)
            
        if end_date:
            end_datetime = timezone.datetime.strptime(end_date, '%Y-%m-%d').replace(
                hour=23, minute=59, second=59, microsecond=999999, tzinfo=timezone.utc
            )
            query_conditions &= models.Q(appointment_datetime__lte=end_datetime)
        
        # жЯ•жЙЊйЬАи¶Бж†ЄеѓєзЪДfleetиЃ∞ељХ
        fleets = await sync_to_async(list)(Fleet.objects.filter(query_conditions).order_by('appointment_datetime'))
        
        # е¶ВжЮЬиЃЊзљЃдЇЖзЫЃзЪДдїУз≠ЫйАЙпЉМињЫдЄАж≠•ињЗжї§shipment
        if destination:
            filtered_fleets = []
            for fleet in fleets:
                # ж£АжЯ•иѓ•fleetжШѓеР¶жЬЙзђ¶еРИжЭ°дїґзЪДshipment
                has_matching_shipment = await sync_to_async(Shipment.objects.filter(
                    fleet_number__fleet_number=fleet.fleet_number,
                    destination__icontains=destination
                ).exists)()
                
                if has_matching_shipment:
                    filtered_fleets.append(fleet)
            
            fleets = filtered_fleets
        
        # дЄЇжѓПдЄ™fleetиОЈеПЦзЫЄеЕ≥зЪДshipmentгАБpackinglistеТМpallet
        fleet_data = []
        for fleet in fleets:
            # иОЈеПЦзЫЄеЕ≥зЪДshipment
            shipment_query = Shipment.objects.filter(
                fleet_number__fleet_number=fleet.fleet_number
            )
            
            # е¶ВжЮЬиЃЊзљЃдЇЖзЫЃзЪДдїУпЉМжЈїеК†зЫЃзЪДдїУз≠ЫйАЙжЭ°дїґ
            if destination:
                shipment_query = shipment_query.filter(destination__icontains=destination)
            
            shipments = await sync_to_async(list)(shipment_query)

            shipment_data = []
            for shipment in shipments:
                # зЫіжО•дљњзФ®valuesеТМannotateињЫи°МеИЖзїДзїЯиЃ°
                pallets = await sync_to_async(list)(Pallet.objects.filter(
                    shipment_batch_number__shipment_batch_number=shipment.shipment_batch_number,
                    container_number__orders__offload_id__offload_at__isnull=False
                ).select_related('container_number')
                .values(
                    'PO_ID', 
                    'shipment_batch_number__shipment_batch_number',
                    'shipment_batch_number',
                    'master_shipment_batch_number',
                    'container_number__container_number',
                    'destination',
                    'is_dropped_pallet'
                ).annotate(
                    pallet_count=Count('id')
                ))
                
                shipment_data.append({
                    'shipment_batch_number': shipment.shipment_batch_number,
                    'appointment_id': shipment.appointment_id,
                    'pallets': pallets
                })
            
            
            fleet_data.append({
                'fleet_number': fleet.fleet_number,
                'appointment_datetime': fleet.appointment_datetime,
                'shipments': shipment_data
            })
        
        context = {
            'fleets': fleet_data,
            'warehouse_options': self.warehouse_options,
            'start_date': start_date,
            'end_date': end_date,
            'destination': destination
        }
        return self.template_fleet_check, context
    
    async def handle_fleet_leader_check_post(self, request):
        """е§ДзРЖиљ¶йШЯйШЯйХњж†ЄеѓєжПРдЇ§"""
        try:
            fleet_number = request.POST.get('fleet_number')
            if not fleet_number:
                messages.error(request, 'иљ¶жђ°еПЈдЄНиГљдЄЇз©Ї')
                template, context = await self.handle_fleet_leader_check_get(request)
                return template, context
            
            # жЯ•жЙЊеєґжЫіжЦ∞fleetиЃ∞ељХ
            fleet = await sync_to_async(Fleet.objects.get)(fleet_number=fleet_number)
            fleet.check_by_leader = True
            await sync_to_async(fleet.save)()
            
            messages.success(request, f'иљ¶жђ° {fleet_number} еЈ≤ж†ЄеѓєжЧ†иѓѓ')
            template, context = await self.handle_fleet_leader_check_get(request)
            return template, context
        except Fleet.DoesNotExist:
            messages.error(request, f'иљ¶жђ° {fleet_number} дЄНе≠ШеЬ®')
            template, context = await self.handle_fleet_leader_check_get(request)
            return template, context
        
    async def handle_fleet_management_get(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        context = {"warehouse_options": self.warehouse_options}
        return self.template_fleet_schedule, context
    
    async def _update_shipment_totals(self, sp_base_q):
        """
        жЫіжЦ∞shipmentзЪДжАїйЗНйЗПгАБжАїдљУзІѓеТМжАїжЭњжХ∞
        """
        shipment_list = await sync_to_async(list)(
            Shipment.objects.filter(sp_base_q).order_by("pickup_time", "shipment_appointment")
        )

        for shipment in shipment_list:
            total_weight = 0
            total_cbm = 0
            total_pallet = 0
            
            # жЯ•иѓҐеЕ≥иБФзЪДpackinglist
            packinglists = await sync_to_async(list)(
                PackingList.objects.filter(
                    shipment_batch_number=shipment,
                    container_number__orders__offload_id__offload_at__isnull=True
                )
            )
            for packinglist in packinglists:
                if packinglist.total_weight_lbs:
                    total_weight += packinglist.total_weight_lbs
                if packinglist.cbm:
                    total_cbm += packinglist.cbm
                    total_pallet += math.ceil(packinglist.cbm / 2)
            
            # жЯ•иѓҐеЕ≥иБФзЪДpallet
            pallets = await sync_to_async(list)(
                Pallet.objects.filter(
                    shipment_batch_number=shipment,
                    container_number__orders__offload_id__offload_at__isnull=False
                )
            )
            for pallet in pallets:
                if pallet.weight_lbs:
                    total_weight += pallet.weight_lbs
                if pallet.cbm:
                    total_cbm += pallet.cbm
                total_pallet += 1
            # жЫіжЦ∞shipmentзЪДжАїйЗНйЗПеТМжАїдљУзІѓ
            shipment.total_weight = float(round(total_weight, 3))
            shipment.total_cbm = float(round(total_cbm, 3))
            shipment.total_pallet = int(total_pallet)
            shipment.shipped_weight = float(round(total_weight, 3))
            shipment.shipped_cbm = float(round(total_cbm, 3))
            shipment.shipped_pallet = int(total_pallet)
            await sync_to_async(shipment.save, thread_sensitive=True)()
              
    
    async def _fl_unscheduled_data(
        self, request: HttpRequest, warehouse:str, four_major_whs: str | None = None, group: str | None = None
    ) -> tuple[str, dict[str, Any]]:
        '''жЦ∞жЄѓеРОеЕђдїУжЬ™жОТиљ¶еТМеЈ≤жОТиљ¶жЯ•жЙЊ'''
        target_date = datetime(2025, 10, 10)
        base_q = models.Q(
            origin=warehouse,
            fleet_number__isnull=True,
            in_use=True,
            is_canceled=False,
            is_notified_customer=True,
            is_virtual_sp=False,
        )
        if group and 'ltl' in group.lower():  # е¶ВжЮЬgroupеМЕеРЂltlпЉИдЄНеМЇеИЖе§Іе∞ПеЖЩпЉЙ
            sp_base_q = base_q & models.Q(shipment_type__in=['LTL', 'еЃҐжИЈиЗ™жПР'])
        else:
            sp_base_q = base_q & models.Q(shipment_type="FTL")

        if four_major_whs == "four_major_whs":          
            sp_base_q &= models.Q(destination__in=FOUR_MAJOR_WAREHOUSES)

        await self._update_shipment_totals(sp_base_q)

        shipment_list = await sync_to_async(list)(
            Shipment.objects.filter(sp_base_q).order_by("pickup_time", "shipment_appointment")
        )

        f_base_q = models.Q(
            origin=warehouse,
            departured_at__isnull=True,
            is_canceled=False,
        )
        if group and 'ltl' in group.lower():  # е¶ВжЮЬgroupеМЕеРЂltlпЉИдЄНеМЇеИЖе§Іе∞ПеЖЩпЉЙ
            ltl_fleet_numbers = await sync_to_async(list)(
                Shipment.objects.filter(
                    destination__regex=r'\d.*\d.*\d.*\d.*\d'
                ).values_list("fleet_number", flat=True).distinct()
            )

            fl_base_q = f_base_q & models.Q(
                models.Q(fleet_type="еЃҐжИЈиЗ™жПР") |
                models.Q(
                    fleet_type="LTL",
                    fleet_number__in=ltl_fleet_numbers
                )
            )
        else:
            shipment_q = Shipment.objects.annotate(
                letters_count=Length(
                    Func(
                        F("destination"),
                        Value("[^A-Za-z]"),
                        Value(""),
                        Value("g"),
                        function="regexp_replace"
                    )
                ),
                numbers_count=Length(
                    Func(
                        F("destination"),
                        Value("[^0-9]"),
                        Value(""),
                        Value("g"),
                        function="regexp_replace"
                    )
                )
            ).filter(
                fleet_number=OuterRef("pk"),
                letters_count__gte=3,
                numbers_count__lt=3
            )


            fl_base_q = f_base_q & models.Q(
                models.Q(fleet_type="FTL") |
                (models.Q(fleet_type="LTL") & models.Q(Exists(shipment_q)))
            )
        if four_major_whs == "four_major_whs":
            fl_base_q &= models.Q(shipment__destination__in=FOUR_MAJOR_WAREHOUSES)
        
        #еЕИжЯ•дЄАдЄЛжЬЙж≤°жЬЙж≤°з±їеЮЛзЪДиљ¶жђ°пЉМи°•дЄКз±їеЮЛ    
        await self._update_fleets_type(fl_base_q,target_date)

        fleet = await sync_to_async(list)(
            Fleet.objects.filter(fl_base_q).filter(
                Q(appointment_datetime__gt=target_date) | Q(appointment_datetime__isnull=True)
            )
            .prefetch_related("shipment")
            .annotate(
                shipment_batch_numbers=StringAgg(
                    "shipment__shipment_batch_number", delimiter=","
                ),
                appointment_ids=StringAgg("shipment__appointment_id", delimiter=","),
                all_appointment_times=StringAgg(
                    Func(
                        F("shipment__shipment_appointment"),
                        Value("YYYY-MM-DD HH24:MI"),  # иІДеЃЪж†ЉеЉПпЉЪеєі-жЬИ-жЧ• жЧґ:еИЖ
                        function="to_char"
                    ),
                    delimiter="\n",
                    distinct=True  # еїЇиЃЃеК†дЄКпЉМйШ≤ж≠ҐеРМдЄАиЊЖиљ¶е§ЪдЄ™POеѓЉиЗіжЧґйЧіи°МйЗНе§НжШЊз§Ї
                ),
            )
            .order_by("appointment_datetime")
        )
        # еЬ®иОЈеПЦfleetеИЧи°®еРОпЉМжЈїеК†еЕЈдљУжЯЬеПЈгАБдїУзВєз≠Йиѓ¶жГЕ
        for fleet_obj in fleet:
            detailed_shipments = []
            
            # иОЈеПЦиѓ•иљ¶йШЯзЪДжЙАжЬЙshipment
            shipments = await sync_to_async(list)(fleet_obj.shipment.all())
            if shipments:
                all_notified = all(shipment.is_notified_customer for shipment in shipments)
                fleet_obj.is_notified_customer = all_notified
                # еПЦзђђдЄАжЭ° shipment зЪД shipment_type дљЬдЄЇи£Еиљљз±їеЮЛ
                fleet_obj.load_type = shipments[0].load_type
            else:
                fleet_obj.is_notified_customer = False
                fleet_obj.load_type = ""

            for shipment in shipments:
                shipment_batch_number = shipment.shipment_batch_number
                
                packinglists = await sync_to_async(list)(
                    PackingList.objects.filter(
                        shipment_batch_number__shipment_batch_number=shipment_batch_number,
                        container_number__orders__offload_id__offload_at__isnull=True
                    ).select_related('container_number')
                    .values('container_number__container_number', 'destination')
                    .annotate(
                        total_cbm=Sum('cbm'),
                        pallet_count=ExpressionWrapper(
                            Sum('cbm') / 2, 
                            output_field=FloatField()
                        )
                    )
                )
                pallets = await sync_to_async(list)(
                    Pallet.objects.filter(
                        shipment_batch_number__shipment_batch_number=shipment_batch_number,
                        container_number__orders__offload_id__offload_at__isnull=False
                    ).select_related('container_number')
                    .values('container_number__container_number', 'destination','is_dropped_pallet')
                    .annotate(
                        total_cbm=Sum('cbm'),
                        pallet_count=Count('id')  # palletзЪДжЭњжХ∞е∞±жШѓжХ∞йЗП
                    )
                )
                
                # жЮДеїЇзїЯдЄАж†ЉеЉПзЪДжХ∞жНЃ
                for item in packinglists:
                    detailed_shipments.append({
                        "type": "EST",
                        "container_number": item["container_number__container_number"],
                        "destination": item["destination"],
                        "cbm": float(item["total_cbm"]) if item["total_cbm"] else 0,
                        "pallet_count": math.ceil(float(item["pallet_count"])) if item["pallet_count"] else 0,  # еРСдЄКеПЦжХі
                        "shipment_batch_number": shipment_batch_number if shipment_batch_number else "",
                        "appointment_id": shipment.appointment_id,
                        "scheduled_time": shipment.shipment_appointment.strftime("%Y-%m-%d %H:%M") if shipment.shipment_appointment else "",
                        "note": shipment.note or "",
                        "is_dropped_pallet": False,
                    })
                
                for item in pallets:
                    detailed_shipments.append({
                        "type": "ACT",
                        "container_number": item["container_number__container_number"],
                        "destination": item["destination"],
                        "cbm": float(item["total_cbm"]) if item["total_cbm"] else 0,
                        "pallet_count": item["pallet_count"] or 0,
                        "shipment_batch_number": shipment_batch_number if shipment_batch_number else "",
                        "appointment_id": shipment.appointment_id,
                        "scheduled_time": shipment.shipment_appointment.strftime("%Y-%m-%d %H:%M") if shipment.shipment_appointment else "",
                        "note": shipment.note or "",
                        "is_dropped_pallet": item["is_dropped_pallet"],
                    })         
            fleet_obj.detailed_shipments = json.dumps(detailed_shipments) 
        context = {
            "shipment_list": shipment_list,
            "fleet_list": fleet,
        }
        return context

    async def _update_fleets_type(self, fl_base_q, target_date):
        fleets_without_type = await sync_to_async(list)(
            Fleet.objects.filter(fl_base_q).filter(
                Q(appointment_datetime__gt=target_date) | Q(appointment_datetime__isnull=True)
            ).filter(
                Q(fleet_type__isnull=True) | Q(fleet_type='')  # зЫіжО•з≠ЫйАЙеЗЇж≤°жЬЙз±їеЮЛзЪДиљ¶йШЯ
            )
        )
        
        for fleet in fleets_without_type:
            # иОЈеПЦеЕ≥иБФзЪДshipment
            shipments_list = []
            async for shipment in Shipment.objects.filter(fleet_number=fleet):
                shipments_list.append(shipment)
            
            if not shipments_list:
                continue
            
            # иОЈеПЦжЙАжЬЙshipment_typeеєґеОїйЗН
            shipment_types = set()
            for shipment in shipments_list:
                if shipment.shipment_type:
                    shipment_types.add(shipment.shipment_type)
            
            if not shipment_types:
                continue
            
            fleet_type = list(shipment_types)[0]
            fleet.fleet_type = fleet_type
            await fleet.asave()
    
    async def _fl_delivery_get(
        self, warehouse:str, four_major_whs: str | None = None, group:str |None = None
    ) -> dict[str, Any]:
        
        base_criteria = models.Q(
            is_arrived=False,
            is_canceled=False,
            is_shipped=True,
            origin=warehouse,
            fleet_number__isnull=False,
        ) & ~Q(status="Exception")

        if group and 'ltl' in group.lower():  # е¶ВжЮЬgroupеМЕеРЂltlпЉИдЄНеМЇеИЖе§Іе∞ПеЖЩпЉЙ
            criteria = base_criteria & models.Q(
                shipment_type__in=['LTL', 'еЃҐжИЈиЗ™жПР'],
                shipping_order_link__isnull=False
            )
        else:
            criteria = base_criteria & models.Q(shipment_type="FTL")

        if four_major_whs == "four_major_whs":
            criteria &= models.Q(destination__in=FOUR_MAJOR_WAREHOUSES)

        shipments = await sync_to_async(list)(
            Shipment.objects.prefetch_related("fleet_number")
            .filter(criteria)
            .order_by("shipped_at")
        )
        for s in shipments:
            pl_pairs = await sync_to_async(list)(
                PackingList.objects.select_related("container_number")
                .filter(
                    container_number__orders__offload_id__offload_at__isnull=False,
                    shipment_batch_number=s
                )
                .values("container_number__container_number", "shipping_mark")
                .distinct()
            )
            plt_pairs = await sync_to_async(list)(
                Pallet.objects.select_related("container_number")
                .filter(
                    container_number__orders__offload_id__offload_at__isnull=False,
                    shipment_batch_number=s
                )
                .values("container_number__container_number", "shipping_mark")
                .distinct()
            )
            pairs_set = set()
            for d in pl_pairs + plt_pairs:
                cont = d.get("container_number__container_number")
                mark = d.get("shipping_mark")
                if cont and mark:
                    pairs_set.add((cont, mark))
            s.container_mark_pairs = [
                {"container_number": c, "shipping_mark": m} for c, m in sorted(pairs_set)
            ]
        shipment_fleet_dict = {}
        for s in shipments:
            if s.fleet_number is None:
                continue
            if s.shipment_appointment is None:
                shipment_appointment = ""
            else:
                shipment_appointment = s.shipment_appointment.replace(
                    microsecond=0
                ).isoformat()
            if s.fleet_number.fleet_number not in shipment_fleet_dict:
                shipment_fleet_dict[s.fleet_number.fleet_number] = [
                    {
                        "shipment_batch_number": s.shipment_batch_number,
                        "appointment_id": s.appointment_id,
                        "destination": s.destination,
                        "carrier": s.carrier,
                        "shipment_appointment": shipment_appointment,
                        "origin": s.origin,
                    }
                ]
            else:
                shipment_fleet_dict[s.fleet_number.fleet_number].append(
                    {
                        "shipment_batch_number": s.shipment_batch_number,
                        "appointment_id": s.appointment_id,
                        "destination": s.destination,
                        "carrier": s.carrier,
                        "shipment_appointment": shipment_appointment,
                        "origin": s.origin,
                    }
                )
        context = {
            "shipments": shipments, #еЊЕз°ЃиЃ§йАБиЊЊжЙєжђ°
            "abnormal_fleet_options": self.abnormal_fleet_options,
            #"shipment": json.dumps(shipment_fleet_dict),
        }
        return context
    
    async def _fl_pod_get(
        self, warehouse:str, four_major_whs: str | None = None, group: str | None = None
    ) -> dict[str, Any]: 

        criteria = models.Q(
            models.Q(models.Q(pod_link__isnull=True) | models.Q(pod_link="")),
            shipped_at__isnull=False,
            arrived_at__isnull=False,
            shipment_schduled_at__gte="2024-12-01",
            origin=warehouse,
        )
        if group and 'ltl' in group.lower():  # е¶ВжЮЬgroupеМЕеРЂltlпЉИдЄНеМЇеИЖе§Іе∞ПеЖЩпЉЙ
            criteria = criteria & models.Q(shipment_type__in=['LTL', 'еЃҐжИЈиЗ™жПР'])
        else:
            criteria = criteria & (
                models.Q(shipment_type='FTL')
                |
                models.Q(shipment_type='е§ЦйЕН', shipment_schduled_at__gte='2026-04-27')
            )

        if four_major_whs == "four_major_whs":
            criteria &= models.Q(destination__in=FOUR_MAJOR_WAREHOUSES)
        shipments = await sync_to_async(list)(
            Shipment.objects.select_related("fleet_number")
            .filter(criteria)
            .order_by("shipped_at")
        )
        for shipment in shipments:
            # иОЈеПЦдЄОиѓ•shipmentеЕ≥иБФзЪДжЙАжЬЙpallet
            pallets = await sync_to_async(list)(
                Pallet.objects.filter(shipment_batch_number=shipment)
                .select_related('container_number')
            )
            
            customer_names = set()
            
            for pallet in pallets:
                if pallet.container_number:
                    # иОЈеПЦдЄОиѓ•containerеЕ≥иБФзЪДжЙАжЬЙorder
                    orders = await sync_to_async(list)(
                        Order.objects.filter(container_number=pallet.container_number)
                        .select_related('customer_name')
                    )
                    
                    for order in orders:
                        if order.customer_name:
                            customer_names.add(order.customer_name.zem_name)
            
            # е∞ЖеЃҐжИЈеРНзФ®йАЧеПЈжЛЉжО•пЉМеєґжЈїеК†еИ∞shipmentеѓєи±°дЄК
            shipment.customer = ", ".join(customer_names) if customer_names else "жЧ†еЃҐжИЈдњ°жБѓ"
        context = {
            "fleet": shipments,
        }
        return context
    
    async def _shipment_exceptions_data(
        self, warehouse:str
    ) -> dict[str, Any]:
        shipment = await sync_to_async(list)(
            Shipment.objects.filter(
                status="Exception",
                is_canceled=False,
                origin=warehouse,
            ).order_by("shipment_appointment")
        )
        shipment_data = {
            s.shipment_batch_number: {
                "origin": s.origin,
                "load_type": s.load_type,
                "note": s.note if s.note and s.note != 'NaN' else '',
                "destination": s.destination,
                "address": s.address,
                "origin": s.origin,
            }
            for s in shipment
        }
        unused_appointment = await sync_to_async(list)(
            Shipment.objects.filter(in_use=False, is_canceled=False)
        )
        unused_appointment = {
            s.appointment_id: {
                "destination": s.destination.strip(),
                "shipment_appointment": s.shipment_appointment.replace(
                    microsecond=0
                ).isoformat(),
            }
            for s in unused_appointment
        }
        context = {
            "exception_shipments": shipment,           
            "unused_appointment": json.dumps(unused_appointment),
            "shipment_data": json.dumps(shipment_data),          
        }
        return context
    
    async def  handle_fleet_schedule_post(
        self, request: HttpRequest, context: dict| None = None
    ) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("warehouse")
        #еЉВеЄЄзЇ¶
        exception_sp = await self._shipment_exceptions_data(warehouse)
        #еЊЕеЗЇеЇУ
        ready_to_ship_data = await self._sp_ready_to_ship_data(warehouse,request.user)
        sum_fleet = []
        for i in ready_to_ship_data:
            sum_fleet.append(i['fleet_number'])
        # еЊЕйАБиЊЊ
        delivery_data = await self._fl_delivery_get(warehouse)
        #еЊЕдЉ†POD
        pod_data = await self._fl_pod_get(warehouse)
        
        summary = {
            'exception_count': len(exception_sp["exception_shipments"]),
            'ready_to_ship_count': len(ready_to_ship_data),
            'ready_count': len(delivery_data['shipments']),
            'pod_count': len(pod_data['fleet']),
        }
        if not context:
            context = {}

        context.update({
            'delivery_shipments': delivery_data['shipments'],
            'pod_shipments': pod_data['fleet'],
            'ready_to_ship_data': ready_to_ship_data,
            'sum_fleet': sum_fleet,
            'summary': summary,        
            'warehouse_options': self.warehouse_options,
            "account_options": self.account_options,
            "warehouse": warehouse,
            "carrier_options": self.carrier_options,
            "abnormal_fleet_options": self.abnormal_fleet_options,
            "exception_shipments": exception_sp["exception_shipments"],
            "shipment_data": exception_sp["shipment_data"],
            "unused_appointment": exception_sp["unused_appointment"],
            'current_time': timezone.now(),
            "load_type_options": LOAD_TYPE_OPTIONS,
            "account_options": self.account_options,
            "shipment_type_options": self.shipment_type_options
        })     
        return self.template_fleet_schedule, context

    async def handle_td_unshipment_post(
        self, request: HttpRequest, context: dict| None = None, matching_suggestions: dict | None = None,
    ) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("warehouse")
        if not warehouse:
            if context:
                context.update({
                    "error_messages": "жЬ™йАЙжЛ©дїУеЇУ!",
                    'warehouse_options': self.warehouse_options
                })
            else:
                context = {
                    "error_messages":"жЬ™йАЙжЛ©дїУеЇУ!",
                    'warehouse_options': self.warehouse_options,
                }
            return self.template_td_unshipment, context
        st_type = request.POST.get("st_type", "pallet")
        # зФЯжИРеМєйЕНеїЇиЃЃ
        max_cbm, max_pallet = await self.get_capacity_limits(st_type)

        # иОЈеПЦдЄЙз±їжХ∞жНЃпЉЪжЬ™жОТзЇ¶гАБеЈ≤жОТзЇ¶гАБеЊЕеЗЇеЇУ
        if not matching_suggestions:
            sp_result = await self.sp_unscheduled_data(
                warehouse, st_type, max_cbm, max_pallet, request.user
            )
            matching_suggestions = sp_result.get("matching_suggestions", [])
            destination_list = sp_result.get("destination_list", [])
        else:
            destination_list = []

        if not context:
            context = {}
        else:
            # йШ≤ж≠ҐдЉ†еЕ•зЪД context иҐЂжДПе§ЦдњЃжФє
            context = context.copy()

        context.update({
            'warehouse': warehouse,
            'st_type': st_type,
            'matching_suggestions': matching_suggestions,
            'destination_list': destination_list,
            'max_cbm': max_cbm,
            'max_pallet': max_pallet,
            'warehouse_options': self.warehouse_options,
            "account_options": self.account_options,
            "load_type_options": LOAD_TYPE_OPTIONS,
            "shipment_type_options": self.shipment_type_options,
            "carrier_options": self.carrier_options,
            'active_tab': request.POST.get('active_tab')
        }) 
        context["matching_suggestions_json"] = json.dumps(matching_suggestions, cls=DjangoJSONEncoder)
        context["warehouse_json"] = json.dumps(warehouse, cls=DjangoJSONEncoder)
        return self.template_td_unshipment, context
    
    async def _history_scheduled_data(self, warehouse: str, user, start_date, end_date) -> list:
        """иОЈеПЦеЈ≤жОТзЇ¶жХ∞жНЃ - жМЙshipment_batch_numberеИЖзїД"""
        # иОЈеПЦжЬЙshipment_batch_numberдљЖfleet_numberдЄЇз©ЇзЪДиіІзЙ©
        base_q = models.Q(
            shipment_appointment__gte=start_date,
            shipment_appointment__lte=end_date,
        )
        if "LA" in warehouse:
            # LAдїУеЇУпЉЪзЫЃзЪДеЬ∞и¶БеЬ®LA_DESеЖЕпЉМжИЦиАЕдЄНеЬ®NJ_DESеТМSAV_DESеЖЕ
            base_q &= (
                models.Q(destination__in=LA_DES) |
                ~models.Q(destination__in=NJ_DES + SAV_DES)
            )
        elif "NJ" in warehouse:
            # NJдїУеЇУпЉЪзЫЃзЪДеЬ∞и¶БеЬ®NJ_DESеЖЕпЉМжИЦиАЕдЄНеЬ®LA_DESеТМSAV_DESеЖЕ
            base_q &= (
                models.Q(destination__in=NJ_DES) |
                ~models.Q(destination__in=LA_DES + SAV_DES)
            )
        elif "SAV" in warehouse:
            # SAVдїУеЇУпЉЪзЫЃзЪДеЬ∞и¶БеЬ®SAV_DESеЖЕпЉМжИЦиАЕдЄНеЬ®LA_DESеТМNJ_DESеЖЕ
            base_q &= (
                models.Q(destination__in=SAV_DES) |
                ~models.Q(destination__in=LA_DES + NJ_DES)
            )
        else:
            # еЕґдїЦдїУеЇУпЉЪдљњзФ®йїШиЃ§йАїиЊСпЉМжИЦиАЕж†єжНЃеЕЈдљУйЬАж±Ви∞ГжХі
            pass
        shipment_list = await sync_to_async(list)(
            Shipment.objects.filter(base_q)
            .order_by("shipment_appointment")
        )
        result = []
    
        for shipment in shipment_list:
            # еИЭеІЛеМЦзїЯиЃ°еПШйЗП
            total_weight = 0.0
            total_cbm = 0.0
            total_pallet = 0  # дїО0еЉАеІЛиЃ°жХ∞
            
            # жЮДеїЇеЯЇз°Аshipmentдњ°жБѓ
            shipment_data = {
                'shipment_batch_number': shipment.shipment_batch_number,
                'appointment_id': shipment.appointment_id,
                'shipment_cargo_id': shipment.shipment_cargo_id,
                'destination': shipment.destination,
                'shipment_appointment': shipment.shipment_appointment,
                'pickup_time': shipment.pickup_time,
                'load_type': shipment.load_type,
                'shipment_account': shipment.shipment_account,
                'origin': shipment.origin,
                'pickup_time': shipment.pickup_time,
                'pickup_number': shipment.pickup_number,
                'in_use': shipment.in_use,
                'is_canceled': shipment.is_canceled,
                'is_notified_customer': shipment.is_notified_customer,
                'cargos': []  # зїСеЃЪзЪДPOеИЧи°®
            }
            
            # жЯ•иѓҐеЕ≥иБФзЪДpackinglist (жЬЙеЕ•дїУжЧґйЧізЪД)
            packinglists = await sync_to_async(list)(
                PackingList.objects.select_related("container_number").filter(
                    shipment_batch_number=shipment,
                    container_number__orders__offload_id__offload_at__isnull=False
                )
            )
            
            # жЯ•иѓҐеЕ≥иБФзЪДpallet (жЬЙеЕ•дїУжЧґйЧізЪД)
            pallets = await sync_to_async(list)(
                Pallet.objects.select_related("container_number").filter(
                    shipment_batch_number=shipment,
                    container_number__orders__offload_id__offload_at__isnull=False
                )
            )
            containers = []
            # е§ДзРЖpackinglistжХ∞жНЃеєґиЃ°зЃЧзїЯиЃ°
            for pl in packinglists:
                # иЃ°зЃЧйЗНйЗП
                pl_weight = pl.total_weight_lbs or 0.0
                total_weight += pl_weight
                
                # иЃ°зЃЧдљУзІѓ
                pl_cbm = pl.cbm or 0.0
                total_cbm += pl_cbm
                
                # иЃ°зЃЧжЭњжХ∞ - дЉШеЕИдљњзФ®еЃЮйЩЕжЭњжХ∞пЉМе¶ВжЮЬж≤°жЬЙеИЩж†єжНЃCBMдЉ∞зЃЧ
                pl_pallet = math.ceil(pl_cbm / 2) if pl_cbm > 0 else 0
                
                total_pallet += pl_pallet
                
                cargo_data = {
                    'ids': str(pl.id),
                    'plt_ids': '',
                    'ref_ids': pl.ref_id or '',
                    'fba_ids': pl.fba_id or '',
                    'cns': pl.container_number.container_number if pl.container_number else '',
                    'destination': pl.destination or '',
                    'total_pallet': total_pallet,
                    'total_weight_lbs': pl_weight,
                    'total_cbm': pl_cbm,
                    'label': 'EST',
                    'offload_time': None,
                    'note_sp': pl.note_sp or ''
                }
                shipment_data['cargos'].append(cargo_data)
                containers.append(f"{pl.container_number.container_number} - {pl.destination}")
            
            # е§ДзРЖpalletжХ∞жНЃеєґиЃ°зЃЧзїЯиЃ°
            for pallet in pallets:
                # иЃ°зЃЧйЗНйЗП
                pallet_weight = pallet.weight_lbs or 0.0
                total_weight += pallet_weight
                
                # иЃ°зЃЧдљУзІѓ
                pallet_cbm = pallet.cbm or 0.0
                total_cbm += pallet_cbm
                
                # жѓПдЄ™palletзЃЧ1жЭњ
                pallet_count = 1
                total_pallet += pallet_count
                
                cargo_data = {
                    'ids': '',
                    'plt_ids': str(pallet.id),
                    'ref_ids': pallet.ref_id or '',
                    'fba_ids': pallet.fba_id or '',
                    'cns': pallet.container_number.container_number if pallet.container_number else '',
                    'destination': pallet.destination or '',
                    'total_pallet': total_pallet,  # еЃЮйЩЕиЃ°зЃЧзЪДжЭњжХ∞
                    'total_weight_lbs': pallet_weight,
                    'total_cbm': pallet_cbm,
                    'label': 'ACT',
                    'offload_time': None,
                    'note_sp': pallet.note_sp or ''
                }
                shipment_data['cargos'].append(cargo_data)
                containers.append(f"{pallet.container_number.container_number} - {pallet.destination}")
            containers = list(set(containers))
            new_containers = '\n'.join(set(containers))
            # жЫіжЦ∞shipmentзЪДжАїзїЯиЃ°дњ°жБѓ
            shipment_data.update({
                'total_weight': float(round(total_weight, 3)),
                'total_cbm': float(round(total_cbm, 3)),
                'total_pallet': int(total_pallet),
                'shipped_weight': float(round(total_weight, 3)),
                'shipped_cbm': float(round(total_cbm, 3)),
                'shipped_pallet': int(total_pallet),
                'container_numbers': new_containers
            })
            
            # еЉВж≠•жЫіжЦ∞жХ∞жНЃеЇУдЄ≠зЪДshipmentиЃ∞ељХ
            shipment.total_weight = total_weight
            shipment.total_cbm = total_cbm
            shipment.total_pallet = total_pallet
            shipment.shipped_weight = total_weight
            shipment.shipped_cbm = total_cbm
            shipment.shipped_pallet = total_pallet
            await sync_to_async(shipment.save)()
            
            result.append(shipment_data)
    

        return result
    
    async def _history_scheduled_fleet_data(self, warehouse: str, user, start_date, end_date) -> list:
        fl_base_q = models.Q(
            origin=warehouse,
            fleet_type="FTL",
            appointment_datetime__gte=start_date,
            appointment_datetime__lte=end_date,
        )
        fleet = await sync_to_async(list)(
            Fleet.objects.filter(fl_base_q)
            .prefetch_related("shipment")
            .annotate(
                shipment_batch_numbers=StringAgg(
                    "shipment__shipment_batch_number", delimiter=","
                ),
                appointment_ids=StringAgg("shipment__appointment_id", delimiter=","),
            )
            .order_by("appointment_datetime")
        )
        # еЬ®иОЈеПЦfleetеИЧи°®еРОпЉМжЈїеК†еЕЈдљУжЯЬеПЈгАБдїУзВєз≠Йиѓ¶жГЕ
        for fleet_obj in fleet:
            detailed_shipments = []
            
            # иОЈеПЦиѓ•иљ¶йШЯзЪДжЙАжЬЙshipment
            shipments = await sync_to_async(list)(fleet_obj.shipment.all())
            if shipments:
                all_notified = all(shipment.is_notified_customer for shipment in shipments)
                fleet_obj.is_notified_customer = all_notified
                # еПЦзђђдЄАжЭ° shipment зЪД shipment_type дљЬдЄЇи£Еиљљз±їеЮЛ
                fleet_obj.shipment_type = shipments[0].shipment_type
            else:
                fleet_obj.is_notified_customer = False
                fleet_obj.shipment_type = ""

            for shipment in shipments:
                shipment_batch_number = shipment.shipment_batch_number
                
                packinglists = await sync_to_async(list)(
                    PackingList.objects.filter(
                        shipment_batch_number__shipment_batch_number=shipment_batch_number,
                        container_number__orders__offload_id__offload_at__isnull=True
                    ).select_related('container_number')
                    .values('container_number__container_number', 'destination')
                    .annotate(
                        total_cbm=Sum('cbm'),
                        pallet_count=ExpressionWrapper(
                            Sum('cbm') / 2, 
                            output_field=FloatField()
                        )
                    )
                )
                pallets = await sync_to_async(list)(
                    Pallet.objects.filter(
                        shipment_batch_number__shipment_batch_number=shipment_batch_number,
                        container_number__orders__offload_id__offload_at__isnull=False
                    ).select_related('container_number')
                    .values('container_number__container_number', 'destination','is_dropped_pallet')
                    .annotate(
                        total_cbm=Sum('cbm'),
                        pallet_count=Count('id')  # palletзЪДжЭњжХ∞е∞±жШѓжХ∞йЗП
                    )
                )
                
                # жЮДеїЇзїЯдЄАж†ЉеЉПзЪДжХ∞жНЃ
                for item in packinglists:
                    detailed_shipments.append({
                        "type": "EST",
                        "container_number": item["container_number__container_number"],
                        "destination": item["destination"],
                        "cbm": float(item["total_cbm"]) if item["total_cbm"] else 0,
                        "pallet_count": math.ceil(float(item["pallet_count"])) if item["pallet_count"] else 0,  # еРСдЄКеПЦжХі
                        "shipment_batch_number": shipment_batch_number if shipment_batch_number else "",
                        "appointment_id": shipment.appointment_id,
                        "scheduled_time": shipment.shipment_appointment.strftime("%Y-%m-%d %H:%M") if shipment.shipment_appointment else "",
                        "note": shipment.note or "",
                        "is_dropped_pallet": False,
                    })
                
                for item in pallets:
                    detailed_shipments.append({
                        "type": "ACT",
                        "container_number": item["container_number__container_number"],
                        "destination": item["destination"],
                        "cbm": float(item["total_cbm"]) if item["total_cbm"] else 0,
                        "pallet_count": item["pallet_count"] or 0,
                        "shipment_batch_number": shipment_batch_number if shipment_batch_number else "",
                        "appointment_id": shipment.appointment_id,
                        "scheduled_time": shipment.shipment_appointment.strftime("%Y-%m-%d %H:%M") if shipment.shipment_appointment else "",
                        "note": shipment.note or "",
                        "is_dropped_pallet": item["is_dropped_pallet"],
                    })         
            fleet_obj.detailed_shipments = json.dumps(detailed_shipments) 
        return fleet
    
    async def handle_history_shipment_post(
        self, request: HttpRequest, context: dict| None = None, 
    ) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("warehouse")
        if not warehouse:
            if context:
                context.update({
                    "error_messages": "жЬ™йАЙжЛ©дїУеЇУ!",
                    'warehouse_options': self.warehouse_options
                })
            else:
                context = {
                    "error_messages":"жЬ™йАЙжЛ©дїУеЇУ!",
                    'warehouse_options': self.warehouse_options,
                }
            return self.template_history_shipment, context
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-30)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date
        
        #еЈ≤жОТзЇ¶
        scheduled_data = await self._history_scheduled_data(warehouse, request.user, start_date, end_date)

        #еЈ≤жОТиљ¶
        schedule_fleet_data = await self._history_scheduled_fleet_data(request, warehouse,start_date, end_date)
        if not context:
            context = {}
        else:
            # йШ≤ж≠ҐдЉ†еЕ•зЪД context иҐЂжДПе§ЦдњЃжФє
            context = context.copy()
        summary = {
            'shipments':len(scheduled_data),
            'fleets': len(schedule_fleet_data),
        }
        context.update({
            'warehouse': warehouse,
            'scheduled_data': scheduled_data,
            'fleet_list': schedule_fleet_data,   #еЈ≤жОТиљ¶
            'warehouse_options': self.warehouse_options,
            "account_options": self.account_options,
            "load_type_options": self.load_type_options,
            "shipment_type_options": self.shipment_type_options,
            "carrier_options": self.carrier_options,
            'active_tab': request.POST.get('active_tab'),
            'summary': summary,
            'start_date': start_date,
            'end_date': end_date,
        }) 
        return self.template_history_shipment, context
    
    async def handle_td_shipment_post(
        self, request: HttpRequest, context: dict| None = None, matching_suggestions: dict | None = None,
    ) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("warehouse")
        if not warehouse:
            if context:
                context.update({
                    "error_messages": "жЬ™йАЙжЛ©дїУеЇУ!",
                    'warehouse_options': self.warehouse_options
                })
            else:
                context = {
                    "error_messages":"жЬ™йАЙжЛ©дїУеЇУ!",
                    'warehouse_options': self.warehouse_options,
                }
            return self.template_td_shipment, context
        st_type = request.POST.get("st_type", "pallet")
        # зФЯжИРеМєйЕНеїЇиЃЃ
        max_cbm, max_pallet = await self.get_capacity_limits(st_type)
        
        # жЬ™жОТзЇ¶
        if not matching_suggestions:
            sp_result = await self.sp_unscheduled_data(
                warehouse, st_type, 1000, 1000, request.user
            )
            matching_suggestions = sp_result.get("matching_suggestions", [])
            destination_list = sp_result.get("destination_list", [])
        else:
            destination_list = []
        #еЈ≤жОТзЇ¶
        scheduled_data = await self.sp_scheduled_data(warehouse, request.user)

        #жЬ™жОТиљ¶+еЈ≤жОТиљ¶
        fleets = await self._fl_unscheduled_data(request, warehouse)
        #жЬ™жОТиљ¶
        unschedule_fleet_data = fleets['shipment_list']
        #еЈ≤жОТиљ¶
        schedule_fleet_data = fleets['fleet_list']
        #е§ЦйЕНзЪДзЇ¶
        external_distribution_data = await self.sp_external_distribution_data(warehouse, request.user)
        # иОЈеПЦеПѓзФ®йҐДзЇ¶
        available_shipments = await self.sp_available_shipments(warehouse, st_type)
        
        # иЃ°зЃЧзїЯиЃ°жХ∞жНЃ
        summary = await self._sp_calculate_summary(matching_suggestions, scheduled_data, schedule_fleet_data, unschedule_fleet_data, external_distribution_data)       

        if not context:
            context = {}
        else:
            # йШ≤ж≠ҐдЉ†еЕ•зЪД context иҐЂжДПе§ЦдњЃжФє
            context = context.copy()

        context.update({
            'warehouse': warehouse,
            'st_type': st_type,
            'matching_suggestions': matching_suggestions,
            'destination_list': destination_list,
            'scheduled_data': scheduled_data,
            'unschedule_fleet': unschedule_fleet_data,
            'fleet_list': schedule_fleet_data,   #еЈ≤жОТиљ¶
            'external_distribution_data': external_distribution_data,   #е§ЦйЕНзЪДзЇ¶
            'unscheduled_fl_count': len(unschedule_fleet_data),
            'available_shipments': available_shipments,
            'summary': summary,
            'max_cbm': max_cbm,
            'max_pallet': max_pallet,
            'warehouse_options': self.warehouse_options,
            "account_options": self.account_options,
            "load_type_options": self.load_type_options,
            "shipment_type_options": self.shipment_type_options,
            "carrier_options": self.carrier_options,
            'active_tab': request.POST.get('active_tab')
        }) 
        context["matching_suggestions_json"] = json.dumps(matching_suggestions, cls=DjangoJSONEncoder)
        context["warehouse_json"] = json.dumps(warehouse, cls=DjangoJSONEncoder)
        return self.template_td_shipment, context
    
    async def sp_unscheduled_data(self, warehouse: str, st_type: str, max_cbm, max_pallet, user) -> dict:
        """иОЈеПЦжЬ™жОТзЇ¶жХ∞жНЃ"""
        delivery_method_filter = (
            ~models.Q(delivery_method__icontains='жЪВжЙ£') &
            ~models.Q(delivery_method__icontains='UPS') &
            ~models.Q(delivery_method__icontains='FEDEX') &
            ~models.Q(delivery_method__icontains='иЗ™жПР')
        )
        has_any_timestamp = (
            models.Q(container_number__orders__retrieval_id__target_retrieval_timestamp__isnull=False) |
            models.Q(container_number__orders__retrieval_id__target_retrieval_timestamp_lower__isnull=False) |
            (
                models.Q(container_number__orders__retrieval_id__actual_retrieval_timestamp__isnull=False) &
                models.Q(container_number__orders__retrieval_id__actual_retrieval_timestamp__gt=datetime(2025, 2, 1))
            )
        )
        unshipment_pos = await self._get_packing_list(
            user,
            has_any_timestamp
            & delivery_method_filter
            & models.Q(
                container_number__orders__offload_id__offload_at__isnull=True,
                shipment_batch_number__shipment_batch_number__isnull=True,
                container_number__orders__retrieval_id__retrieval_destination_precise=warehouse,
                delivery_type='public',
            ),
            delivery_method_filter
            & models.Q(
                shipment_batch_number__shipment_batch_number__isnull=True,
                container_number__orders__offload_id__offload_at__gt=datetime(2025, 1, 1),
                location=warehouse,
                delivery_type='public',
            ), True
        )
             
        # иОЈеПЦеПѓзФ®зЪДshipmentиЃ∞ељХпЉИshipment_batch_numberдЄЇз©ЇзЪДпЉЙ
        shipments = await self._get_available_shipments(warehouse)
        # зФЯжИРжЩЇиГљеМєйЕНеїЇиЃЃ
        matching_suggestions = await self._generate_matching_suggestions(unshipment_pos, shipments, warehouse, max_cbm, max_pallet,st_type, user)
        
        destination_totals = {}
        for suggestion in matching_suggestions:
            primary_group = suggestion.get("primary_group", {})
            dest = (primary_group.get("destination") or "").strip()
            if not dest:
                continue
            cbm_value = primary_group.get("total_cbm", 0) or 0
            try:
                cbm_float = float(cbm_value)
            except Exception:
                cbm_float = 0.0
            destination_totals[dest] = destination_totals.get(dest, 0.0) + cbm_float

        destination_list = [
            {"destination": dest, "total_cbm": round(total, 2)}
            for dest, total in destination_totals.items()
        ]
        destination_list.sort(key=lambda x: x.get("total_cbm", 0), reverse=True)

        return {
            "matching_suggestions": matching_suggestions,
            "destination_list": destination_list,
        }

    async def _get_available_shipments(self, warehouse: str):
        """иОЈеПЦеПѓзФ®зЪДshipmentиЃ∞ељХ"""
        now = timezone.now()
        # ињЩйЗМйЬАи¶Бж†єжНЃжВ®зЪДеЃЮйЩЕж®°еЮЛи∞ГжХіжЯ•иѓҐжЭ°дїґ
        shipments = await sync_to_async(list)(
            Shipment.objects.filter(
                models.Q(shipment_batch_number__isnull=True) | models.Q(shipment_batch_number=''),
                in_use=False,
                is_canceled=False,
                shipment_appointment__gt=now  
            ).order_by('shipment_appointment') 
        )
        return shipments

    async def _generate_matching_suggestions(self, unshipment_pos, shipments, warehouse, max_cbm, max_pallet,st_type, user):
        """зФЯжИРжЩЇиГљеМєйЕНеїЇиЃЃ - еЯЇдЇОеКЯиГљAзЪДйАїиЊСдљЖйАВйЕНshipmentеМєйЕН"""
        suggestions = []

        # дЄЇдЇЖжЦєдЊњй©ђе£ЂеЯЇиѓҐдїЈпЉМеҐЮеК†еПВжХ∞pallet_items_jsonпЉМиЃ∞ељХжѓПдЄ™жЭње≠РзЪДйХњеЃљйЂШгАБдїґжХ∞гАБйЗНйЗП
        plt_id_set = set()
        for cargo in unshipment_pos:
            if cargo.get("data_source") == "PALLET" and cargo.get("plt_ids"):
                for pid in str(cargo.get("plt_ids") or "").split(","):
                    pid = pid.strip()
                    if pid.isdigit():
                        plt_id_set.add(int(pid))

        pallet_map = {}
        if plt_id_set:
            pallets = await sync_to_async(list)(
                Pallet.objects.select_related("container_number")
                .filter(id__in=list(plt_id_set))
                .values(
                    "id",
                    "length",
                    "width",
                    "height",
                    "pcs",
                    "weight_lbs",
                    "container_number__container_number",
                )
            )
            pallet_map = {int(p["id"]): p for p in pallets if p.get("id") is not None}

        for cargo in unshipment_pos:
            if cargo.get("data_source") == "PALLET" and cargo.get("plt_ids"):
                pallet_items = []
                for pid in str(cargo.get("plt_ids") or "").split(","):
                    pid = pid.strip()
                    if not pid.isdigit():
                        continue
                    p = pallet_map.get(int(pid))
                    if not p:
                        continue
                    pallet_items.append(
                        {
                            "plt_id": str(pid),
                            "container_number": p.get("container_number__container_number") or "",
                            "length": p.get("length"),
                            "width": p.get("width"),
                            "height": p.get("height"),
                            "pieces": p.get("pcs"),
                            "weight": p.get("weight_lbs"),
                            "description": "",
                        }
                    )
                cargo["pallet_items_json"] = json.dumps(pallet_items, ensure_ascii=False)
            else:
                cargo["pallet_items_json"] = ""

        # зђђдЄАзЇІеИЖзїДпЉЪжМЙзЫЃзЪДеЬ∞еТМжіЊйАБжЦєеЉПйҐДеИЖзїД
        pre_groups = {}
        for cargo in unshipment_pos:
            if not await self.has_appointment(cargo):
                raw_dest = (cargo.get('destination') or '').strip()
                dest = raw_dest.split('-')[-1].strip().upper() if raw_dest else ''

                delivery_method = cargo.get('custom_delivery_method')
                if not dest or not delivery_method:
                    continue
                    
                group_key = f"{dest}_{delivery_method}"
                if group_key not in pre_groups:
                    pre_groups[group_key] = {
                        'destination': dest,
                        'delivery_method': delivery_method,
                        'cargos': []
                    }
                pre_groups[group_key]['cargos'].append(cargo)

        # еЬ®йҐДеИЖзїДеЊ™зОѓдєЛеЙНеИЭеІЛеМЦеЈ≤дљњзФ®зЪДshipmentйЫЖеРИ
        # еѓєжѓПдЄ™йҐДеИЖзїДжМЙеЃєйЗПйЩРеИґеИЫеїЇе§ІзїД
        for group_key, pre_group in pre_groups.items():
            cargos = pre_group['cargos']
            
            # жМЙETAжОТеЇПпЉМдЉШеЕИеЃЙжОТжЧ©зЪДиіІзЙ©
            sorted_cargos = sorted(cargos, key=lambda x: x.get('container_number__orders__vessel_id__vessel_eta') or '')
            
            # жМЙеЃєйЗПйЩРеИґеИЫеїЇе§ІзїД
            primary_groups = []
            current_primary_group = {
                'destination': pre_group['destination'],
                'delivery_method': pre_group['delivery_method'],
                'cargos': [],
                'total_pallets': 0,
                'total_cbm': 0,
            }
            
            for cargo in sorted_cargos:
                cargo_pallets = 0
                if cargo.get('label') == 'ACT':
                    cargo_pallets = cargo.get('total_n_pallet_act', 0) or 0
                else:
                    cargo_pallets = cargo.get('total_n_pallet_est', 0) or 0
                
                cargo_cbm = cargo.get('total_cbm', 0) or 0
                # ж£АжЯ•ељУеЙНе§ІзїДжШѓеР¶ињШиГљеЃєзЇ≥ињЩдЄ™иіІзЙ©
                if (current_primary_group['total_pallets'] + cargo_pallets <= max_pallet and 
                    current_primary_group['total_cbm'] + cargo_cbm <= max_cbm):
                    # еПѓдї•еК†еЕ•ељУеЙНе§ІзїД
                    current_primary_group['cargos'].append(cargo)
                    current_primary_group['total_pallets'] += cargo_pallets
                    current_primary_group['total_cbm'] += cargo_cbm
                else:
                    # ељУеЙНе§ІзїДеЈ≤жї°пЉМдњЭе≠ШеєґеИЫеїЇжЦ∞зЪДе§ІзїД
                    if current_primary_group['cargos']:
                        primary_groups.append(current_primary_group)
                    
                    # еИЫеїЇжЦ∞зЪДе§ІзїД
                    current_primary_group = {
                        'destination': pre_group['destination'],
                        'delivery_method': pre_group['delivery_method'],
                        'cargos': [cargo],
                        'total_pallets': cargo_pallets,
                        'total_cbm': cargo_cbm,
                    }
            
            # жЈїеК†жЬАеРОдЄАдЄ™е§ІзїД
            if current_primary_group['cargos']:
                primary_groups.append(current_primary_group)
            
            # дЄЇжѓПдЄ™е§ІзїДеѓїжЙЊеМєйЕНзЪДshipment
            for primary_group_index, primary_group in enumerate(primary_groups):
                # иЃ°зЃЧе§ІзїДзЪДеМєйЕНеЇ¶зЩЊеИЖжѓФ
                pallets_percentage = min(100, (primary_group['total_pallets'] / max_pallet) * 100) if max_pallet > 0 else 0
                cbm_percentage = min(100, (primary_group['total_cbm'] / max_cbm) * 100) if max_cbm > 0 else 0
                
                # еѓїжЙЊеМєйЕНзЪДshipmentпЉМињЩйЗМжФєжИРињФеЫЮеИЧи°®
                matched_shipment = await self._find_matching_shipment(primary_group, shipments, warehouse)
                
                # е¶ВжЮЬеМєйЕНеИ∞shipmentпЉМе∞ЖеЕґж†ЗиЃ∞дЄЇеЈ≤дљњзФ®
                result_intel = await self._find_intelligent_po_for_group(
                    primary_group, warehouse, user
                )
                
                intelligent_pos = result_intel['intelligent_pos']
                intelligent_pos_stats = result_intel['intelligent_pos_stats']
                
                intelligent_cargos = [{
                    'unique_id': pos.get('unique_id', ''),
                    'ids': pos.get('ids', ''),
                    'plt_ids': pos.get('plt_ids', ''),
                    'ref_ids': pos.get('ref_ids', ''),
                    'fba_ids': pos.get('fba_ids', ''),
                    'container_numbers': pos.get('container_numbers', ''),
                    'cns': pos.get('cns', ''),
                    'offload_time': pos.get('offload_time',''),
                    'delivery_window_start': pos.get('delivery_window_start'),
                    'delivery_window_end': pos.get('delivery_window_end'),
                    'total_n_pallet_act': pos.get('total_n_pallet_act') or pos.get('total_n_pallet_est', 0),
                    'total_cbm': pos.get('total_cbm', 0),
                    'total_weight': pos.get('total_weight', 0),
                    'label': pos.get('label', ''),
                    'destination': pos.get('destination', ''),
                    'location': pos.get('location') if pos.get('location') else pos.get('warehouse', ''),
                    'custom_delivery_method': pos.get('custom_delivery_method', ''),
                } for pos in intelligent_pos]
                
                # жЧ†иЃЇжШѓеР¶еМєйЕНеИ∞shipmentпЉМйГљеИЫеїЇеїЇиЃЃеИЖзїД
                suggestion = {
                    'suggestion_id': f"{group_key}_{primary_group_index}",
                    'primary_group': {
                        'destination': primary_group['destination'],
                        'delivery_method': primary_group['delivery_method'],
                        'total_pallets': primary_group['total_pallets'],
                        'total_cbm': primary_group['total_cbm'],
                        'pallets_percentage': pallets_percentage,
                        'cbm_percentage': cbm_percentage,
                        'matched_shipment': matched_shipment,  # еПѓиГљдЄЇNone
                        'suggestion_id': f"{group_key}_{primary_group_index}"
                    },
                    'cargos': [{
                        'ids': cargo.get('ids', ''),
                        'plt_ids': cargo.get('plt_ids', ''),
                        'ref_ids': cargo.get('ref_ids', ''),
                        'fba_ids': cargo.get('fba_ids', ''),
                        'container_numbers': cargo.get('container_numbers', ''),
                        'cns': cargo.get('cns', ''),
                        'customer_name': cargo.get('customer_name', ''),
                        'destination': cargo.get('destination', ''),
                        'offload_time': cargo.get('offload_time', ''),
                        'delivery_window_start': cargo.get('delivery_window_start'),
                        'delivery_window_end': cargo.get('delivery_window_end'),
                        'total_n_pallet_act': cargo.get('total_n_pallet_act', 0),
                        'total_n_pallet_est': cargo.get('total_n_pallet_est', 0),
                        'total_cbm': cargo.get('total_cbm', 0),
                        'total_weight': cargo.get('total_weight_lbs', 0),
                        'label': cargo.get('label', ''),
                        'is_dropped_pallet': cargo.get('is_dropped_pallet'),
                        'rebuilt_is_dropped_pallet': cargo.get('rebuilt_is_dropped_pallet'),
                        'shipment_note': cargo.get('shipment_note', ''),
                        'pallet_items_json': cargo.get('pallet_items_json', ''),
                    } for cargo in primary_group['cargos']],
                    'intelligent_cargos': intelligent_cargos,
                    'intelligent_pos_stats': intelligent_pos_stats,
                    'virtual_fleet': []
                }
                suggestions.append(suggestion)
        #жЯ•жЙЊеПѓдї•дЄАжПРе§ЪеНЄзЪДеПѓиГљ
        await self.calculate_virtual_fleet(suggestions, max_cbm, max_pallet)
        def to_float(val):
            try:
                return float(val)
            except (TypeError, ValueError):
                return 0.0       
        if st_type == "еН°жЭњ":
            suggestions.sort(
                key=lambda s: (
                    to_float(s['primary_group'].get('cbm_percentage')),
                    to_float(s['primary_group'].get('pallets_percentage'))
                ),
                reverse=True
            )
        else:
            suggestions.sort(
                key=lambda x: x['primary_group']['cbm_percentage'], 
                reverse=True
            )
        return suggestions

    async def calculate_virtual_fleet(self, suggestions, max_cbm, max_pallet):
        """иЃ°зЃЧжѓПдЄ™е§ІзїДеПѓдї•еРИеєґи£Еиљ¶зЪДеЕґдїЦе§ІзїД"""
        for i, current_suggestion in enumerate(suggestions):
            current_group = current_suggestion['primary_group']
            current_cbm = current_group['total_cbm']
            current_pallets = current_group['total_pallets']
            
            # иЃ°зЃЧељУеЙНе§ІзїДзЪДеЙ©дљЩеЃєйЗП
            remaining_cbm = 1000 - current_cbm
            remaining_pallets = 1000 - current_pallets
            
            # еѓїжЙЊеПѓдї•еРИеєґзЪДеЕґдїЦе§ІзїД
            compatible_groups = []
            
            for j, other_suggestion in enumerate(suggestions):
                if i == j:  # иЈ≥ињЗиЗ™еЈ±
                    continue
                    
                other_group = other_suggestion['primary_group']
                other_cbm = other_group['total_cbm']
                other_pallets = other_group['total_pallets']
                
                # ж£АжЯ•жШѓеР¶еПѓдї•еРИеєґпЉИдЄНиґЕињЗеЙ©дљЩеЃєйЗПпЉЙ
                if other_cbm <= remaining_cbm and other_pallets <= remaining_pallets:
                    # иЃ°зЃЧеМєйЕНеЇ¶еИЖжХ∞ - иґКжО•ињСеЙ©дљЩеЃєйЗПзЪДдЉШеЕИзЇІиґКйЂШ
                    cbm_match_score = other_cbm / remaining_cbm if remaining_cbm > 0 else 0
                    pallets_match_score = other_pallets / remaining_pallets if remaining_pallets > 0 else 0
                    total_match_score = (cbm_match_score + pallets_match_score) / 2
                    
                    compatible_groups.append({
                        'suggestion_id': other_group['suggestion_id'],
                        'cbm_percentage': other_group['cbm_percentage'],
                        'pallets_percentage': other_group['pallets_percentage'],
                        'total_cbm': other_cbm,
                        'total_pallets': other_pallets,
                        'match_score': total_match_score
                    })
            
            # жМЙеМєйЕНеЇ¶еИЖжХ∞жОТеЇПпЉМеМєйЕНеЇ¶йЂШзЪДжОТеЬ®еЙНйЭҐ
            compatible_groups.sort(key=lambda x: x['match_score'], reverse=True)         
            # еП™е≠ШеВ®suggestion_idеИЧи°®
            current_suggestion['virtual_fleet'] = [group['suggestion_id'] for group in compatible_groups]

    async def _find_intelligent_po_for_group(self, primary_group, warehouse, user) -> Any:
        existing_pl_ids = []
        existing_plt_ids = []
        destination = None

        for cargo in primary_group.get("cargos", []):
            destination = cargo.get("destination")
            id_str = cargo.get("ids")
            if id_str:
                existing_pl_ids.extend([r.strip() for r in id_str.split(",") if r.strip()])

            plt_str = cargo.get("plt_ids")
            if plt_str:
                existing_plt_ids.extend([int(p.strip()) for p in plt_str.split(",") if p.strip().isdigit()])

        #йҐДжГ≥пЉМNJеТМSAVзЪДеПѓдї•иАГиЩСиљђдїУпЉМLAзЪДе∞±дЄНиАГиЩСдЇЖпЉМжЙАдї•жПРдЊЫжЩЇиГљеМєйЕНжДПиІБжЧґпЉМLAзЪДдЄНиАГиЩСеИЂзЪДдїУ
        if "LA" in warehouse:
            location_condition = models.Q(location=warehouse)
            retrieval_condition = (
                models.Q(container_number__orders__retrieval_id__retrieval_destination_precise=warehouse) |
                models.Q(container_number__orders__warehouse__name=warehouse)
            )
        else:
            location_condition = models.Q(location__in=["NJ-07001", "SAV-31326"])
            retrieval_condition = (
                models.Q(container_number__orders__retrieval_id__retrieval_destination_precise__in=["NJ-07001", "SAV-31326"]) |
                models.Q(container_number__orders__warehouse__name__in=["NJ-07001", "SAV-31326"])
            )
        # жЩЇиГљжЙЊPOзЪДз≠ЫйАЙжЭ°дїґ
        target_date = datetime(2025, 10, 10)
                
        intelligent_pos = await self._get_packing_list(
            user,
            models.Q(
                shipment_batch_number__shipment_batch_number__isnull=True,
                container_number__orders__offload_id__offload_at__isnull=True,
                container_number__orders__add_to_t49=True,
                container_number__orders__vessel_id__vessel_eta__gte=target_date,
                destination=destination,
                delivery_type='public',
                
            ) & retrieval_condition
            & ~models.Q(id__in=existing_pl_ids),
            models.Q(
                shipment_batch_number__shipment_batch_number__isnull=True,
                container_number__orders__offload_id__offload_at__isnull=False,
                destination=destination,
                container_number__orders__offload_id__offload_at__gt=datetime(2025, 1, 1),
                delivery_type='public',
            ) & location_condition
            & ~models.Q(id__in=existing_plt_ids),
        )
        
        sorted_intelligent_pos = sorted(intelligent_pos, key=lambda x: (
            0 if (x.get('location') or x.get('warehouse', '')) == warehouse else 1,
        ))
        
        intelligent_cargos = [{
            'unique_id': str(uuid.uuid4()), 
            'ids': pos.get('ids', ''),
            'plt_ids': pos.get('plt_ids', ''),
            'ref_ids': pos.get('ref_ids', ''),
            'fba_ids': pos.get('fba_ids', ''),
            'container_numbers': pos.get('container_numbers', ''),
            'cns': pos.get('cns', ''),
            'offload_time': pos.get('offload_time',''),
            'delivery_window_start': pos.get('delivery_window_start'),
            'delivery_window_end': pos.get('delivery_window_end'),
            'total_n_pallet_act': pos.get('total_n_pallet_act', 0),
            'total_n_pallet_est': pos.get('total_n_pallet_est', 0),
            'total_cbm': pos.get('total_cbm', 0),
            'total_weight': pos.get('total_weight_lbs', 0),
            'label': pos.get('label', ''),
            'destination': pos.get('destination', ''),
            'custom_delivery_method': pos.get('custom_delivery_method', ''),
            'location': pos.get('location') if pos.get('location') else pos.get('warehouse', '')
        } for pos in sorted_intelligent_pos]

        organized = {
            'ACT': {'normal': [], 'hold': []},
            'EST': {'normal': [], 'hold': []}
        }
        
        for cargo in intelligent_cargos:
            label = cargo.get('label', 'EST')
            delivery_method = cargo.get('custom_delivery_method', '')
            is_hold = False
            if delivery_method:
                is_hold = 'жЪВжЙ£' in delivery_method
            else:
                continue
            if label == 'ACT':
                if is_hold:
                    organized['ACT']['hold'].append(cargo)
                else:
                    organized['ACT']['normal'].append(cargo)
            else: 
                if is_hold:
                    organized['EST']['hold'].append(cargo)
                else:
                    organized['EST']['normal'].append(cargo)
        intelligent_pos_stats = {
            'ACT_normal_count': len(organized['ACT']['normal']),
            'ACT_hold_count': len(organized['ACT']['hold']),
            'EST_normal_count': len(organized['EST']['normal']),
            'EST_hold_count': len(organized['EST']['hold']),
            'total_count': len(organized['ACT']['normal']) + len(organized['ACT']['hold']) + 
                        len(organized['EST']['normal']) + len(organized['EST']['hold'])
        }
        return {
            'intelligent_pos': intelligent_cargos,
            'intelligent_pos_stats':intelligent_pos_stats
            }
    
    async def _find_matching_shipment(self, primary_group, shipments, warehouse):
        """дЄЇиіІзЙ©е§ІзїДеѓїжЙЊеМєйЕНзЪДshipment"""
        destination = primary_group['destination']
        matched_shipments = []
        
        for shipment in shipments:
            # ж£АжЯ•ињЩдЄ™shipmentжШѓеР¶еЈ≤зїПиҐЂеЕґдїЦзїДдљњзФ®дЇЖ
            # ж£АжЯ•зЫЃзЪДеЬ∞жШѓеР¶еМєйЕН
            shipment_destination = (shipment.destination or '').strip().upper()
            if not self._is_destination_match(destination, shipment_destination):
                continue

            if shipment.origin != warehouse:
                continue
            # ж£АжЯ•жЧґйЧіз™ЧеП£жЭ°дїґ
            if not await self.check_time_window_match(primary_group, shipment):
                continue         
            # еМєйЕНжИРеКЯпЉМжЈїеК†еИ∞еМєйЕНеИЧи°®
            matched_shipment = {
                'shipment_id': shipment.id,
                'appointment_id': shipment.appointment_id,
                'shipment_cargo_id': shipment.shipment_cargo_id,
                'shipment_type': shipment.shipment_type,
                'shipment_appointment': shipment.shipment_appointment,
                'pickup_time': shipment.pickup_time,
                'pickup_number': shipment.pickup_number,
                'origin': shipment.origin,
                'load_type': shipment.load_type,
                'shipment_account': shipment.shipment_account,
                'address': shipment.address,
                'carrier': shipment.carrier,
                'note': shipment.note,
                'ARM_BOL': shipment.ARM_BOL,
                'ARM_PRO': shipment.ARM_PRO,
                'express_number': shipment.express_number,
                'address_detail': await self.get_address(destination),            
                'destination': shipment.destination
            }
            matched_shipments.append(matched_shipment)
        if matched_shipments:
            matched_shipments.sort(key=lambda x: x.get('shipment_appointment') or datetime.max)
            # еЬ®зїУе∞ЊжЈїеК†з©ЇзЪДshipmentйАЙй°є
            empty_shipment = {
                'shipment_id': None,
                'appointment_id': "дЄНйАЙжЛ©йҐДзЇ¶еПЈ",
                'shipment_cargo_id': None,
                'shipment_type': None,
                'shipment_appointment': None,
                'pickup_time': None,
                'pickup_number': None,
                'origin': None,
                'load_type': None,
                'shipment_account': None,
                'address': None,
                'carrier': None,
                'note': "",
                'ARM_BOL': None,
                'ARM_PRO': None,
                'express_number': None,
                'address_detail': await self.get_address(destination),
                'destination': destination,
                'is_empty_option': True
            }
            
            matched_shipments.append(empty_shipment)
        return matched_shipments
    
    def _is_destination_match(self, group_destination, shipment_destination):
        """ж£АжЯ•зЫЃзЪДеЬ∞жШѓеР¶еМєйЕН"""
        if not shipment_destination:
            return False
        
        # зЃАеНХзЪДзЫЃзЪДеЬ∞еМєйЕНйАїиЊСпЉМжВ®еПѓдї•ж†єжНЃеЃЮйЩЕйЬАж±Ви∞ГжХі
        group_dest_clean = group_destination.split('-')[-1].strip().upper()
        shipment_dest_clean = shipment_destination.split('-')[-1].strip().upper()
        
        return group_dest_clean == shipment_dest_clean

    async def get_address(self,destination):
        if destination in amazon_fba_locations:
            fba = amazon_fba_locations[destination]
            address = f"{fba['location']}, {fba['city']} {fba['state']}, {fba['zipcode']}"
            return address
        
        # е¶ВжЮЬзЫіжО•жЯ•жЙЊдЄНеИ∞пЉМе∞ЭиѓХжЈїеК†Walmart-еЙНзЉАжЯ•жЙЊ
        walmart_destination = f"Walmart-{destination}"
        if walmart_destination in amazon_fba_locations:
            fba = amazon_fba_locations[walmart_destination]
            address = f"{fba['location']}, {fba['city']} {fba['state']}, {fba['zipcode']}"
            return address
        return None
        # е¶ВжЮЬдЄ§зІНжЦєеЉПйГљжЙЊдЄНеИ∞пЉМжК•йФЩ
        raise ValueError(f'жЙЊдЄНеИ∞ињЩдЄ™зЫЃзЪДеЬ∞зЪДеЬ∞еЭАпЉМиѓЈж†ЄеЃЮ{destination}пЉИеЈ≤е∞ЭиѓХ{walmart_destination}пЉЙ')
        
    async def check_time_window_match(self, primary_group, shipment):
        """ж£АжЯ•жЧґйЧіз™ЧеП£жШѓеР¶еМєйЕН"""
        shipment_appointment = shipment.shipment_appointment
        if not shipment_appointment:
            return False
        
        shipment_date = shipment_appointment.date()
        # ж£АжЯ•е∞ПзїДдЄ≠зЪДжѓПдЄ™иіІзЙ©
        for cargo in primary_group['cargos']:
            window_start = cargo.get('delivery_window_start')
            window_end = cargo.get('delivery_window_end')
            
            # е¶ВжЮЬиіІзЙ©жЬЙжЧґйЧіз™ЧеП£пЉМж£АжЯ•shipmentжЧґйЧіжШѓеР¶еЬ®з™ЧеП£еЖЕ
            if window_start and window_end:
                if not (window_start <= shipment_date <= window_end):
                    return False
            # е¶ВжЮЬиіІзЙ©ж≤°жЬЙжЧґйЧіз™ЧеП£пЉМиЈ≥ињЗжЧґйЧіж£АжЯ•пЉИеП™и¶Бж±ВзЫЃзЪДеЬ∞еМєйЕНпЉЙ
        
        return True

    async def sp_scheduled_data(self, warehouse: str, user, four_major_whs: str | None = None, group: str | None = None) -> list:
        """иОЈеПЦеЈ≤жОТзЇ¶жХ∞жНЃ - жМЙshipment_batch_numberеИЖзїД"""
        # иОЈеПЦжЬЙshipment_batch_numberдљЖfleet_numberдЄЇз©ЇзЪДиіІзЙ©
        target_date = datetime(2025, 10, 10)

        pl_criteria = models.Q(
                container_number__orders__warehouse__name=warehouse,
                shipment_batch_number__isnull=False,             
                container_number__orders__offload_id__offload_at__isnull=True,
                shipment_batch_number__shipment_appointment__gt=target_date,
                shipment_batch_number__fleet_number__isnull=True,
            )
        if group and 'ltl' in group.lower():  # е¶ВжЮЬgroupеМЕеРЂltlпЉИдЄНеМЇеИЖе§Іе∞ПеЖЩпЉЙ
            pl_criteria = pl_criteria & models.Q(delivery_type='other')
        else:
            pl_criteria = pl_criteria & models.Q(delivery_type='public')

        plt_criteria = models.Q(
                shipment_batch_number__isnull=False,
                shipment_batch_number__shipment_appointment__gt=target_date,
                container_number__orders__offload_id__offload_at__isnull=False,
                shipment_batch_number__fleet_number__isnull=True,
                location=warehouse,
            )
        if group and 'ltl' in group.lower():  # е¶ВжЮЬgroupеМЕеРЂltlпЉИдЄНеМЇеИЖе§Іе∞ПеЖЩпЉЙ
            plt_criteria = plt_criteria & models.Q(delivery_type='other')
        else:
            plt_criteria = plt_criteria & models.Q(delivery_type='public')

        if four_major_whs == "four_major_whs":
            pl_criteria &= models.Q(destination__in=FOUR_MAJOR_WAREHOUSES)
            plt_criteria &= models.Q(destination__in=FOUR_MAJOR_WAREHOUSES)

        raw_data = await self._get_packing_list(
            user,
            pl_criteria,
            plt_criteria,
        )
        
        # жМЙshipment_batch_numberеИЖзїД
        grouped_data = {}
        processed_batch_numbers = set()

        for item in raw_data:           
            batch_number = item.get('shipment_batch_number__shipment_batch_number')
            if "еЇУе≠ШзЫШзВє" in batch_number:
                continue
            if batch_number not in grouped_data:
                # иОЈеПЦйҐДзЇ¶дњ°жБѓ
                try:
                    shipment = await sync_to_async(Shipment.objects.get)(
                        shipment_batch_number=batch_number,
                        shipment_appointment__gte=datetime(2025, 1, 1)
                    )
                except Shipment.DoesNotExist:
                    continue
                except MultipleObjectsReturned:
                    raise ValueError(f"shipment_batch_number={batch_number} жЯ•иѓҐеИ∞е§ЪжЭ°иЃ∞ељХпЉМиѓЈж£АжЯ•жХ∞жНЃ")
                if shipment.status == "Exception": #еЉВеЄЄзЪДзЇ¶дЄНе±Хз§Ї
                    continue
                address = await self.get_address(shipment.destination)
                grouped_data[batch_number] = {
                    'appointment_id': shipment.appointment_id,
                    'shipment_cargo_id': shipment.shipment_cargo_id,
                    'shipment_batch_number': shipment.shipment_batch_number,
                    'shipment_type': shipment.shipment_type,
                    'destination': shipment.destination,
                    'shipment_appointment': shipment.shipment_appointment,
                    'load_type': shipment.load_type,
                    'shipment_account': shipment.shipment_account,
                    'address': shipment.address,
                    'address_detail': address,
                    'cargos': [],
                    'pickup_time': shipment.pickup_time,
                    'pickup_number': shipment.pickup_number,
                    'is_notified_customer': shipment.is_notified_customer,
                    'carrier': shipment.carrier,
                    'note': shipment.note,
                    'ARM_BOL':shipment.ARM_BOL,
                }
                processed_batch_numbers.add(batch_number)
            grouped_data[batch_number]['cargos'].append(item)
        
        # жЯ•иѓҐж≤°жЬЙиіІзЙ©зЪДshipmentиЃ∞ељХ
        if four_major_whs == "four_major_whs":
            base_q = Q(
                shipped_at__isnull=True,
                shipment_appointment__gt=target_date,
                shipment_type='FTL',
                is_canceled=False,
                in_use=True,
                is_virtual_sp=False
            )
        else:
            base_q = Q(
                Q(fleet_number__isnull=False, fleet_number__is_virtual=True) |
                Q(fleet_number__isnull=True),
                shipment_appointment__gt=target_date,
                shipment_type='FTL',
                is_canceled=False,
                in_use=True,
                is_virtual_sp=False
            )
        if four_major_whs == "four_major_whs":
            base_q &= models.Q(destination__in=FOUR_MAJOR_WAREHOUSES)
        exclude_q = ~Q(shipment_batch_number__in=processed_batch_numbers)
        empty_shipments = await sync_to_async(list)(
            Shipment.objects.filter(base_q).exclude(exclude_q)
        )
        
        # жЈїеК†ж≤°жЬЙиіІзЙ©зЪДshipmentиЃ∞ељХ
        for shipment in empty_shipments:
            has_packinglist = await sync_to_async(PackingList.objects.filter(shipment_batch_number=shipment).exists)()
            has_pallet = await sync_to_async(Pallet.objects.filter(shipment_batch_number=shipment).exists)()
            
            if has_packinglist or has_pallet:
                continue  # еЈ≤зїПзїСеЃЪиіІзЙ©пЉМиЈ≥ињЗ
            batch_number = shipment.shipment_batch_number
            if batch_number not in grouped_data:
                address = await self.get_address(shipment.destination)
                grouped_data[batch_number] = {
                    'appointment_id': shipment.appointment_id,
                    'shipment_cargo_id': shipment.shipment_cargo_id,
                    'shipment_batch_number': shipment.shipment_batch_number,
                    'shipment_type': shipment.shipment_type,
                    'destination': shipment.destination,
                    'shipment_appointment': shipment.shipment_appointment,
                    'load_type': shipment.load_type,
                    'shipment_account': shipment.shipment_account,
                    'address': shipment.address,
                    'address_detail': address,
                    'cargos': [],  # з©ЇеИЧи°®и°®з§Їж≤°жЬЙиіІзЙ©
                    'pickup_time': shipment.pickup_time,
                    'pickup_number': shipment.pickup_number,
                    'is_notified_customer': shipment.is_notified_customer,
                }
        return list(grouped_data.values())

    async def sp_external_distribution_data(self, warehouse: str, user) -> list:
        """иОЈеПЦе§ЦйЕНзЪДзЇ¶жХ∞жНЃ - shipment_typeдЄЇе§ЦйЕНдЄФpod_linkдЄЇз©Ї"""
        target_date = datetime(2025, 10, 10)

        pl_criteria = models.Q(
            container_number__orders__warehouse__name=warehouse,
            shipment_batch_number__isnull=False,
            container_number__orders__offload_id__offload_at__isnull=True,
            shipment_batch_number__pickup_time__gt=target_date,
            shipment_batch_number__shipment_type="е§ЦйЕН",
            shipment_batch_number__pod_link__isnull=True,
        )

        plt_criteria = models.Q(
            location=warehouse,
            shipment_batch_number__isnull=False,           
            container_number__orders__offload_id__offload_at__isnull=False,
            shipment_batch_number__pickup_time__gt=target_date,
            shipment_batch_number__shipment_type="е§ЦйЕН",
            shipment_batch_number__pod_link__isnull=True,            
        )

        raw_data = await self._get_packing_list(
            user,
            pl_criteria,
            plt_criteria,
        )
        
        grouped_data = {}
        processed_batch_numbers = set()

        for item in raw_data:
            batch_number = item.get('shipment_batch_number__shipment_batch_number')
            if "еЇУе≠ШзЫШзВє" in batch_number:
                continue
            if batch_number not in grouped_data:
                try:
                    shipment = await sync_to_async(Shipment.objects.get)(
                        shipment_batch_number=batch_number,
                        shipment_appointment__gte=datetime(2025, 1, 1)
                    )
                except Shipment.DoesNotExist:
                    continue
                except MultipleObjectsReturned:
                    raise ValueError(f"shipment_batch_number={batch_number} жЯ•иѓҐеИ∞е§ЪжЭ°иЃ∞ељХпЉМиѓЈж£АжЯ•жХ∞жНЃ")
                if shipment.status == "Exception":
                    continue
                address = await self.get_address(shipment.destination)
                grouped_data[batch_number] = {
                    'id': shipment.id,
                    'appointment_id': shipment.appointment_id,
                    'shipment_cargo_id': shipment.shipment_cargo_id,
                    'shipment_batch_number': shipment.shipment_batch_number,
                    'shipment_type': shipment.shipment_type,
                    'destination': shipment.destination,
                    'shipment_appointment': shipment.shipment_appointment,
                    'load_type': shipment.load_type,
                    'shipment_account': shipment.shipment_account,
                    'address': shipment.address,
                    'address_detail': address,
                    'cargos': [],
                    'pickup_time': shipment.pickup_time,
                    'pickup_number': shipment.pickup_number,
                    'carrier': shipment.carrier,
                    'note': shipment.note,
                    'ARM_BOL': shipment.ARM_BOL,
                }
                processed_batch_numbers.add(batch_number)
            grouped_data[batch_number]['cargos'].append(item)
        
        base_q = Q(
            shipment_appointment__gt=target_date,
            shipment_type="е§ЦйЕН",
            pod_link__isnull=True,
            is_canceled=False,
            in_use=True,
            is_virtual_sp=False
        )
        
        exclude_q = ~Q(shipment_batch_number__in=processed_batch_numbers)
        empty_shipments = await sync_to_async(list)(
            Shipment.objects.filter(base_q).exclude(exclude_q)
        )
        
        for shipment in empty_shipments:
            has_packinglists = await sync_to_async(PackingList.objects.filter(shipment_batch_number=shipment).exists)()
            has_pallets = await sync_to_async(Pallet.objects.filter(shipment_batch_number=shipment).exists)()
            
            if has_packinglists or has_pallets:
                continue
            batch_number = shipment.shipment_batch_number
            if batch_number not in grouped_data:
                address = await self.get_address(shipment.destination)
                grouped_data[batch_number] = {
                    'id': shipment.id,
                    'appointment_id': shipment.appointment_id,
                    'shipment_cargo_id': shipment.shipment_cargo_id,
                    'shipment_batch_number': shipment.shipment_batch_number,
                    'shipment_type': shipment.shipment_type,
                    'destination': shipment.destination,
                    'shipment_appointment': shipment.shipment_appointment,
                    'load_type': shipment.load_type,
                    'shipment_account': shipment.shipment_account,
                    'address': shipment.address,
                    'address_detail': address,
                    'cargos': [],
                    'pickup_time': shipment.pickup_time,
                    'pickup_number': shipment.pickup_number,
                }
        return list(grouped_data.values())

    async def _sp_ready_to_ship_data(self, warehouse: str, user, four_major_whs: str | None = None, group: str | None = None) -> list:
        """иОЈеПЦеЊЕеЗЇеЇУжХ∞жНЃ - жМЙfleet_numberеИЖзїД"""
        # иОЈеПЦжМЗеЃЪдїУеЇУзЪДжЬ™еЗЇеПСдЄФжЬ™еПЦжґИзЪДfleet
        base_bq = models.Q(
            origin=warehouse,
            departured_at__isnull=True,
            is_canceled=False,
        )
        if group and 'ltl' in group.lower():  # е¶ВжЮЬgroupеМЕеРЂltlпЉИдЄНеМЇеИЖе§Іе∞ПеЖЩпЉЙ
            base_q = base_bq & models.Q(fleet_type__in=['LTL', 'еЃҐжИЈиЗ™жПР'])
        else:
            base_q = base_bq & models.Q(fleet_type="FTL")

        if four_major_whs == "four_major_whs":
            base_q &= models.Q(shipment__destination__in=FOUR_MAJOR_WAREHOUSES)
        fleets = await sync_to_async(list)(
            Fleet.objects.filter(base_q).prefetch_related(
                Prefetch(
                    'shipment',
                    queryset=Shipment.objects.prefetch_related(
                        Prefetch(
                            'packinglist',
                            queryset=PackingList.objects.select_related('container_number')
                        ),
                        Prefetch(
                            'pallet', 
                            queryset=Pallet.objects.select_related('packing_list', 'container_number')
                        )
                    )
                )
            )
        )
        
        grouped_data = []
        
        for fleet in fleets:
            fleet_group = {
                'fleet_number': fleet.fleet_number,
                'fleet_type': fleet.fleet_type,
                'fleet_cost': fleet.fleet_cost or 0,
                'third_party_address': fleet.third_party_address,
                'pickup_number': fleet.pickup_number,
                'motor_carrier_number': fleet.motor_carrier_number,
                'license_plate': fleet.license_plate,
                'dot_number': fleet.dot_number,
                'appointment_datetime': fleet.appointment_datetime,
                'carrier': fleet.carrier,
                'is_virtual': fleet.is_virtual,
                'shipments': {},  # жФєеЫЮе≠ЧеЕЄзїУжЮДпЉМдњЭжМБдЄОеЙНзЂѓеЕЉеЃє
                'pl_ids': [],
                'plt_ids': [],
                'total_cargos': 0,  # жАїиіІзЙ©и°МжХ∞
            }
            
            shipments = await sync_to_async(list)(
                Shipment.objects.filter(fleet_number__fleet_number=fleet.fleet_number)
            )
            
            for shipment in shipments:
                if not shipment.shipment_batch_number:
                    continue

                batch_number = shipment.shipment_batch_number
                
                # еИЭеІЛеМЦshipmentжХ∞жНЃ
                if batch_number not in fleet_group['shipments']:
                    fleet_group['shipments'][batch_number] = {
                        'shipment_batch_number': shipment.shipment_batch_number or '-',
                        'appointment_id': shipment.appointment_id or '-',
                        'destination': shipment.destination or '-',
                        'shipment_appointment': shipment.shipment_appointment,
                        'cargos': []
                    }
                
                # е§ДзРЖpackinglists
                raw_data = await self._get_packing_list(
                    user,
                    models.Q(
                        shipment_batch_number__shipment_batch_number=batch_number,
                        container_number__orders__offload_id__offload_at__isnull=True,
                    ),
                    models.Q(
                        shipment_batch_number__shipment_batch_number=batch_number,
                        container_number__orders__offload_id__offload_at__isnull=False,
                    ),
                )
                fleet_group['shipments'][batch_number]['cargos'].extend(raw_data)
            
            # жОТеЇП shipmentsпЉМcargos дЄЇз©ЇзЪДжФЊеРОйЭҐ
            fleet_group['shipments'] = dict(
                sorted(
                    fleet_group['shipments'].items(),
                    key=lambda item: not item[1]['cargos']
                )
            )
            fleet_group['total_cargos'] = sum(
                len(s['cargos']) if s['cargos'] else 1
                for s in fleet_group['shipments'].values()
            )
            if fleet_group["total_cargos"] <= 0:
                fleet_group["total_cargos"] = 1
            # еП™жЬЙжЬЙжХ∞жНЃзЪДfleetжЙНињФеЫЮ
            if fleet_group['shipments']:
                grouped_data.append(fleet_group)
            else:
                # е¶ВжЮЬж≤°жЬЙеѓєеЇФзЪДshipmentпЉМдїОжХ∞жНЃеЇУдЄ≠еИ†йЩ§иѓ•fleet
                await sync_to_async(fleet.delete)()
        # жМЙ appointment_datetime жОТеЇПпЉМжЧґйЧіжЧ©зЪДжОТеЬ®еЙНйЭҐ
        grouped_data.sort(
            key=lambda x: (
                x['appointment_datetime'].replace(tzinfo=None)
                if x['appointment_datetime'] else datetime.max
            )
        )
        return grouped_data

    async def sp_available_shipments(self, warehouse: str, st_type: str) -> list:
        """иОЈеПЦеПѓзФ®йҐДзЇ¶"""
        now = timezone.now()
        shipments = await sync_to_async(list)(
            Shipment.objects.filter(
                Q(origin__isnull=True) | Q(origin="") | Q(origin=warehouse),
                appointment_id__isnull=False,
                in_use=False,
                is_canceled=False,
                load_type=st_type.upper()
            ).order_by("shipment_appointment")
        )
        
        # жЈїеК†зКґжАБдњ°жБѓ
        for shipment in shipments:
            is_expired = shipment.shipment_appointment_utc and shipment.shipment_appointment_utc < now
            is_urgent = (
                shipment.shipment_appointment_utc and 
                (shipment.shipment_appointment_utc - now).days < 7 and
                shipment.shipment_appointment_utc >= now and
                not is_expired
            )
            
            if is_expired:
                shipment.status = 'expired'
            elif is_urgent:
                shipment.status = 'urgent'
            else:
                shipment.status = 'available'
        
        return shipments

    def _create_primary_groups(self, cargos: list, max_cbm: float, max_pallet: int) -> list:
        """жМЙеЃєйЗПйЩРеИґеИЫеїЇе§ІзїД"""
        primary_groups = []
        current_group = {
            'cargos': [],
            'total_pallets': 0,
            'total_cbm': 0,
            'destination': '',
            'delivery_method': ''
        }
        
        # зЫіжО•йБНеОЖпЉМдЄНжОТеЇП
        for cargo in cargos:
            cargo_pallets = cargo.get('total_n_pallet_act', 0) or cargo.get('total_n_pallet_est', 0) or 0
            cargo_cbm = cargo.get('total_cbm', 0) or 0
            
            if not current_group['destination']:
                current_group['destination'] = cargo.get('destination')
                current_group['delivery_method'] = cargo.get('custom_delivery_method')
            
            # ж£АжЯ•еЃєйЗП
            if (current_group['total_pallets'] + cargo_pallets <= max_pallet and 
                current_group['total_cbm'] + cargo_cbm <= max_cbm):
                current_group['cargos'].append(cargo)
                current_group['total_pallets'] += cargo_pallets
                current_group['total_cbm'] += cargo_cbm
            else:
                if current_group['cargos']:
                    primary_groups.append(current_group.copy())
                current_group = {
                    'cargos': [cargo],
                    'total_pallets': cargo_pallets,
                    'total_cbm': cargo_cbm,
                    'destination': cargo.get('destination'),
                    'delivery_method': cargo.get('custom_delivery_method')
                }
        
        if current_group['cargos']:
            primary_groups.append(current_group)
        
        return primary_groups

    async def _check_time_window_match(self, window_start, window_end, shipment) -> bool:
        """ж£АжЯ•жЧґйЧіз™ЧеП£еМєйЕН"""
        # зЃАеМЦеЃЮзО∞пЉМеЃЮйЩЕеЇФж†єжНЃдЄЪеК°йАїиЊСеЃМеЦД
        if not window_start and not window_end:
            return True
        
        shipment_time = shipment.shipment_appointment
        
        # е¶ВжЮЬеП™жЬЙеЉАеІЛжЧґйЧіпЉМж£АжЯ•йҐДзЇ¶жЧґйЧіжШѓеР¶еЬ®еЉАеІЛжЧґйЧідєЛеРО
        if window_start and not window_end:
            return shipment_time >= window_start
        
        # е¶ВжЮЬеП™жЬЙзїУжЭЯжЧґйЧіпЉМж£АжЯ•йҐДзЇ¶жЧґйЧіжШѓеР¶еЬ®зїУжЭЯжЧґйЧідєЛеЙН
        if not window_start and window_end:
            return shipment_time <= window_end
        
        # е¶ВжЮЬжЧҐжЬЙеЉАеІЛжЧґйЧіеПИжЬЙзїУжЭЯжЧґйЧіпЉМж£АжЯ•йҐДзЇ¶жЧґйЧіжШѓеР¶еЬ®жЧґйЧіз™ЧеП£еЖЕ
        if window_start and window_end:
            return window_start <= shipment_time <= window_end
        
        return False

    async def _sp_calculate_summary(self, unscheduled: list, scheduled: list, schedule_fleet_data: list, unscheduled_fl, external_distribution_data: list = None) -> dict:
        """иЃ°зЃЧзїЯиЃ°жХ∞жНЃ"""
        # иЃ°зЃЧеРДз±їжХ∞йЗП
        unscheduled_sp_count = len(unscheduled)
        scheduled_sp_count = len(scheduled)
        schedule_fl_count = len(schedule_fleet_data)
        unscheduled_fl_count = len(unscheduled_fl)
        external_distribution_count = len(external_distribution_data) if external_distribution_data else 0
        # иЃ°зЃЧжАїжЭњжХ∞
        total_pallets = 0
        for cargo in unscheduled:
            total_pallets += cargo.get('total_n_pallet_act', 0) or cargo.get('total_n_pallet_est', 0) or 0
        
        return {
            'unscheduled_sp_count': unscheduled_sp_count,
            'scheduled_sp_count': scheduled_sp_count,
            'schedule_fl_count': schedule_fl_count,
            'unscheduled_fl_count': unscheduled_fl_count,
            'external_distribution_count': external_distribution_count,
            'total_pallets': int(total_pallets),
        }

    async def get_capacity_limits(self, st_type: str) -> tuple:
        """иОЈеПЦеЃєйЗПйЩРеИґ"""
        if st_type == "pallet":
            return 80, 35
        elif st_type == "floor":
            return 80, 38
        return 80, 35
    
    async def _ltl_packing_list(
        self,
        pl_criteria: models.Q | None = None,
        plt_criteria: models.Q | None = None,
    ) -> list[Any]:
        '''LTLжЄѓеРОзЪДжХ∞жНЃжЯ•иѓҐ'''
        pl_criteria &= models.Q(container_number__orders__cancel_notification=False)& ~models.Q(container_number__orders__order_type='зЫійАБ')
        plt_criteria &= models.Q(container_number__orders__cancel_notification=False)& ~models.Q(container_number__orders__order_type='зЫійАБ')
        
        data = []
        pallet_data_to_sort = []
        packinglist_data_to_sort = []
        if plt_criteria:
            pk_subquery = PackingList.objects.filter(
                shipping_mark=OuterRef("shipping_mark"),
                PO_ID=OuterRef("PO_ID")
            )
            pal_list = await sync_to_async(list)(
                Pallet.objects.prefetch_related(
                    "container_number",
                    "container_number__orders",
                    "container_number__orders__warehouse",
                    "shipment_batch_number",
                    "shipment_batch_number__fleet_number",
                    "container_number__orders__offload_id",
                    "container_number__orders__customer_name",
                    "container_number__orders__retrieval_id",
                    "container_number__orders__vessel_id",
                )
                .filter(plt_criteria)
                .annotate(
                    str_id=Cast("id", CharField()),
                    str_container_number=Cast("container_number__container_number", CharField()),
                    data_source=Value("PALLET", output_field=CharField()),
                    is_pass=Value(True, output_field=BooleanField()),
                    # йЗНйЗПжНҐзЃЧпЉЪlbs иљђ kg
                    weight_kg=ExpressionWrapper(
                        F("weight_lbs") * 0.453592,
                        output_field=FloatField()
                    ),
                    offload_at=Func(
                        F("container_number__orders__offload_id__offload_at"),
                        Value("MM-DD"),
                        function="to_char",
                        output_field=CharField(),
                    )
                )
                .values(
                    "destination",
                    "delivery_method",
                    "shipping_mark",
                    "abnormal_palletization",
                    "delivery_window_start",
                    "delivery_window_end",
                    "note",
                    "carrier_company",
                    "ltl_bol_num",
                    "ltl_pro_num",
                    'ltl_supplier',
                    "PickupAddr",
                    "container_number",
                    "address", 
                    "is_dropped_pallet",
                    "shipment_batch_number__shipment_batch_number",
                    "shipment_batch_number__shipped_at",
                    "shipment_batch_number__arrived_at",
                    "data_source",
                    "shipment_batch_number__fleet_number__fleet_number",
                    "shipment_batch_number__destination",
                    "shipment_batch_number__address",
                    "shipment_batch_number__carrier",
                    "shipment_batch_number__shipment_appointment",
                    "shipment_batch_number__ARM_BOL",
                    "shipment_batch_number__ARM_PRO",
                    "shipment_batch_number__is_print_label",
                    "shipment_batch_number__shipment_type",
                    "shipment_batch_number__note",
                    "shipment_batch_number__fleet_number__Supplier",
                    "shipment_batch_number__fleet_number__fleet_cost",
                    "location",
                    "is_pass",
                    "est_pickup_time",
                    "ltl_cost",
                    "ltl_quote",
                    "ltl_unit_quote",
                    "offload_at",
                    "ltl_follow_status",
                    "ltl_release_command",
                    "ltl_cost_note",
                    "ltl_quote_note",
                    "ltl_contact_method",
                    "ltl_plt_size_note",
                    "PO_ID",
                    "del_qty",
                    "ltl_correlation_id",
                    "ltl_address",
                    "ltl_city",
                    "ltl_state",
                    "ltl_zipcode",
                    "ltl_address_type",
                    warehouse=F("container_number__orders__retrieval_id__retrieval_destination_precise"),
                    retrieval_destination_precise=F("container_number__orders__retrieval_id__retrieval_destination_precise"),
                    customer_name=F("container_number__orders__customer_name__zem_name"),
                    vessel_name=F("container_number__orders__vessel_id__vessel"),
                    vessel_eta=F("container_number__orders__vessel_id__vessel_eta"),                   
                )
                .annotate(
                    # е¶ВжЮЬ ltl_follow_status еМЕеРЂ 'pickup' (дЄНеМЇеИЖе§Іе∞ПеЖЩзФ® icontains)пЉМиЃЊдЄЇ 0пЉМеР¶еИЩиЃЊдЄЇ 1
                    is_pickup_priority=Case(
                        When(ltl_follow_status__icontains="pickup", then=Value(0)),
                        default=Value(1),
                        output_field=IntegerField(),
                    ),
                    # еИЖзїДдЊЭжНЃпЉЪdestination + shipping_mark
                    custom_delivery_method=F("delivery_method"),
                    shipping_marks=F("shipping_mark"),  # дњЭжМБеОЯжЬЙе≠ЧжЃµеРН
                    # зІїйЩ§ fba_ids еТМ ref_ids
                    plt_ids=StringAgg(
                        "str_id", delimiter=",", distinct=True, ordering="str_id"
                    ),
                    # жЯЬеПЈеИЧи°®
                    container_numbers=StringAgg(
                        "str_container_number", delimiter="\n", distinct=True, ordering="str_container_number"
                    ),
                    cns=StringAgg(
                        "str_container_number", delimiter="\n", distinct=True, ordering="str_container_number"
                    ),
                    total_pcs=Sum("pcs", output_field=IntegerField()),
                    total_cbm=Round(Sum("cbm", output_field=FloatField()), 3),
                    total_weight_lbs=Subquery(pk_subquery.values('total_weight_lbs')[:1]),
                    # жЦ∞еҐЮпЉЪжАїйЗНйЗПkg
                    total_weight_kg=Subquery(pk_subquery.values('total_weight_kg')[:1]),
                    total_n_pallet_act=Count("pallet_id", distinct=True),
                    label=Value("ACT"),
                    note_sp=StringAgg("note_sp", delimiter=",", distinct=True),
                )
                .order_by("is_pickup_priority","offload_at","destination", "shipping_mark")
            )

            # е§ДзРЖжЙШзЫШе∞ЇеѓЄдњ°жБѓ
            
            # жФґйЫЖжґЙеПКзЪДжЙАжЬЙ pallet ID
            all_pallet_ids = []
            for cargo in pal_list:
                if cargo.get('plt_ids'):
                    # plt_ids is a comma separated string of IDs
                    ids = str(cargo['plt_ids']).split(',')
                    all_pallet_ids.extend(ids)
            
            # жЙєйЗПжЯ•иѓҐжЙШзЫШиѓ¶жГЕ
            pallet_details_map = {}
            if all_pallet_ids:
                # еОїйЗН
                unique_ids = list(set(all_pallet_ids))
                
                # еЉВж≠•жЯ•иѓҐ Pallet
                pallets_qs = Pallet.objects.filter(id__in=unique_ids).values(
                    'id', 'length', 'width', 'height', 'pcs', 'weight_lbs'
                )
                pallets_data = await sync_to_async(list)(pallets_qs)
                
                for p in pallets_data:
                    pallet_details_map[str(p['id'])] = p

            processed_pal_list = []

            for cargo in pal_list:
                cargo['pallet_size_formatted'] = cargo['ltl_plt_size_note']
                
                # жЮДеїЇ pallet_items еИЧи°®
                pallet_items = []
                if cargo.get('plt_ids'):
                    p_ids = str(cargo['plt_ids']).split(',')
                    for p_id in p_ids:
                        detail = pallet_details_map.get(p_id)
                        if detail:
                            pallet_items.append({
                                'length': detail['length'],
                                'width': detail['width'],
                                'height': detail['height'],
                                'pieces': detail['pcs'],
                                'weight': detail['weight_lbs'],
                                'description': ''  # йїШиЃ§дЄЇз©ЇпЉМеЙНзЂѓеПѓзЉЦиЊС
                            })
                
                # иљђдЄЇ JSON е≠Чзђ¶дЄ≤пЉМдЊЫеЙНзЂѓзЫіжО•дљњзФ®
                cargo['pallet_items_json'] = json.dumps(pallet_items)
                
                processed_pal_list.append(cargo)

            pallet_data_to_sort = processed_pal_list
        # жЯ•иѓҐ PackingList жХ∞жНЃ
        if pl_criteria:
            pl_list = await sync_to_async(list)(
                PackingList.objects.prefetch_related(
                    "container_number",
                    "container_number__orders",
                    "container_number__orders__warehouse",
                    "shipment_batch_number",
                    "shipment_batch_number__fleet_number",
                    "container_number__orders__offload_id",
                    "container_number__orders__customer_name",
                    "container_number__orders__retrieval_id",
                    "container_number__orders__vessel_id",
                )
                .filter(pl_criteria)
                .annotate(
                    str_container_number=Cast("container_number__container_number", CharField()),
                    str_id=Cast("id", CharField()),
                    str_shipping_mark=Cast("shipping_mark", CharField()),
                    data_source=Value("PACKINGLIST", output_field=CharField()),
                    is_pass=Case(
                        When(
                            container_number__orders__retrieval_id__planned_release_time__isnull=False,
                            then=Value(True)
                        ),
                        When(
                            container_number__orders__retrieval_id__temp_t49_available_for_pickup=True,
                            then=Value(True)
                        ),
                        default=Value(False),
                        output_field=BooleanField()
                    ),
                )
                .values(
                    "destination",
                    "delivery_method",
                    "delivery_window_start",
                    "delivery_window_end",
                    "note",
                    "container_number",
                    "shipping_mark",
                    "address", 
                    "data_source",
                    "ltl_verify",
                    "carrier_company",
                    "ltl_bol_num",
                    "ltl_pro_num",
                    'ltl_supplier',
                    "PickupAddr",
                    "est_pickup_time",
                    "ltl_follow_status",
                    "ltl_release_command",
                    "ltl_contact_method",
                    "shipment_batch_number__shipment_batch_number",
                    "shipment_batch_number__shipped_at",
                    "shipment_batch_number__arrived_at",
                    "shipment_batch_number__fleet_number__fleet_number",
                    "shipment_batch_number__destination",
                    "shipment_batch_number__address",
                    "shipment_batch_number__carrier",
                    "shipment_batch_number__shipment_appointment",
                    "shipment_batch_number__ARM_BOL",
                    "shipment_batch_number__ARM_PRO",
                    "shipment_batch_number__is_print_label",
                    "shipment_batch_number__shipment_type",
                    "shipment_batch_number__note",
                    "shipment_batch_number__fleet_number__Supplier",
                    "shipment_batch_number__fleet_number__fleet_cost",
                    "ltl_correlation_id",
                    "ltl_address",
                    "ltl_city",
                    "ltl_state",
                    "ltl_zipcode",
                    "ltl_address_type",
                    warehouse=F("container_number__orders__retrieval_id__retrieval_destination_precise"),
                    vessel_eta=F("container_number__orders__vessel_id__vessel_eta"),
                    is_pass=F("is_pass"),
                    customer_name=F("container_number__orders__customer_name__zem_name"),
                    vessel_name=F("container_number__orders__vessel_id__vessel"),
                    actual_retrieval_time=F("container_number__orders__retrieval_id__actual_retrieval_timestamp"),
                    arm_time=F("container_number__orders__retrieval_id__generous_and_wide_target_retrieval_timestamp"),
                    estimated_time=F("container_number__orders__retrieval_id__planned_release_time"),
                    offload_at=Case(
                        When(
                            container_number__orders__retrieval_id__actual_retrieval_timestamp__isnull=False,
                            then=Func(
                                F("container_number__orders__retrieval_id__actual_retrieval_timestamp"),
                                Value("MM-DD"),
                                function="to_char",
                                output_field=CharField(),
                            )
                        ),
                        When(
                            container_number__orders__retrieval_id__planned_release_time__isnull=False,
                            then=Func(
                                F("container_number__orders__retrieval_id__planned_release_time"),
                                Value("MM-DD"),
                                function="to_char",
                                output_field=CharField(),
                            )
                        ),
                        default=Func(
                            F("container_number__orders__vessel_id__vessel_eta"),
                            Value("MM-DD"),
                            function="to_char",
                            output_field=CharField(),
                        ),
                        output_field=CharField(),
                    ),
                    offload_tag=Case(
                        When(
                            container_number__orders__retrieval_id__actual_retrieval_timestamp__isnull=False,
                            then=Value("еЃЮйЩЕжПРжЯЬ")
                        ),
                        When(
                            container_number__orders__retrieval_id__planned_release_time__isnull=False,
                            then=Value("еЃЮйЩЕжФЊи°М")
                        ),
                        default=Value("ETA")
                    )
                )
                .annotate(
                    # е¶ВжЮЬ ltl_follow_status еМЕеРЂ 'pickup' (дЄНеМЇеИЖе§Іе∞ПеЖЩзФ® icontains)пЉМиЃЊдЄЇ 0пЉМеР¶еИЩиЃЊдЄЇ 1
                    is_pickup_priority=Case(
                        When(ltl_follow_status__icontains="pickup", then=Value(0)),
                        default=Value(1),
                        output_field=IntegerField(),
                    ),
                    # еИЖзїДдЊЭжНЃпЉЪdestination + shipping_mark
                    shipping_marks=StringAgg(
                        "str_shipping_mark",
                        delimiter=",",
                        distinct=True,
                        ordering="str_shipping_mark",
                    ),
                    # зІїйЩ§ fba_ids еТМ ref_ids
                    ids=StringAgg(
                        "str_id", delimiter=",", distinct=True, ordering="str_id"
                    ),
                    # жЯЬеПЈеИЧи°®
                    container_numbers=StringAgg(
                        "str_container_number", delimiter="\n", distinct=True, ordering="str_container_number"
                    ),
                    cns=StringAgg(
                        "str_container_number", delimiter="\n", distinct=True, ordering="str_container_number"
                    ),
                    total_pcs=Sum("pcs", output_field=FloatField()),
                    total_cbm=Round(Sum("cbm", output_field=FloatField()), 3),
                    total_weight_lbs=Sum("total_weight_lbs", output_field=FloatField()),
                    total_weight_kg=Sum("total_weight_kg", output_field=FloatField()),
                    total_n_pallet_est=Ceil(Sum("cbm", output_field=FloatField()) / 2),
                    label=Value("EST"),
                    note_sp=StringAgg("note_sp", delimiter=",", distinct=True)
                )
                .distinct()
                .order_by("is_pickup_priority","actual_retrieval_time")
            )

            packinglist_data_to_sort = pl_list
        
        # 1. еЃЪдєЙжЧ•жЬЯжПРеПЦиЊЕеК©еЗљжХ∞ (еЖЕйГ®еЗљжХ∞)
        def get_pickup_date_key(status):
            if not status:
                return (99, 99)
            # жРЬзіҐжХ∞е≠Ч/жХ∞е≠Чж†ЉеЉП (е¶В 2/1 жИЦ 02/18)
            match = re.search(r'(\d+)/(\d+)', str(status))
            if match:
                try:
                    month = int(match.group(1))
                    day = int(match.group(2))
                    return (month, day)
                except ValueError:
                    pass
            # жРЬзіҐжХ∞е≠Ч.жХ∞е≠Чж†ЉеЉП (е¶В 2.1 жИЦ 02.18)
            match = re.search(r'(\d+)\.(\d+)', str(status))
            if match:
                try:
                    month = int(match.group(1))
                    day = int(match.group(2))
                    return (month, day)
                except ValueError:
                    pass
            return (99, 99) # ж≤°еМєйЕНеИ∞жЧ•жЬЯзЪДжОТеЬ®жЬАеРО

        # 2. ж£АжЯ• ltl_follow_status жШѓеР¶жЬЙжЧ•жЬЯ
        def has_date_in_status(status):
            if not status:
                return False
            if re.search(r'\d+/\d+', str(status)):
                return True
            if re.search(r'\d+\.\d+', str(status)):
                return True
            return False

        # 3. еИЖз¶їеЗЇ ltl_follow_status жЬЙжЧ•жЬЯзЪДжХ∞жНЃжЈЈжОТпЉМеЕґдїЦзЪДдњЭжМБеОЯжЭ•зЪДй°ЇеЇП
        date_group_data = []  # е≠ШжФЊ ltl_follow_status жЬЙжЧ•жЬЯзЪДжЙАжЬЙжХ∞жНЃ
        remaining_pallet_data = []  # е≠ШжФЊ pallet_data дЄ≠ж≤°жЬЙжЧ•жЬЯзЪДжХ∞жНЃ
        remaining_packinglist_data = []  # е≠ШжФЊ packinglist_data дЄ≠ж≤°жЬЙжЧ•жЬЯзЪДжХ∞жНЃ

        if pallet_data_to_sort:
            for item in pallet_data_to_sort:
                if has_date_in_status(item.get('ltl_follow_status', '')):
                    date_group_data.append(item)
                else:
                    remaining_pallet_data.append(item)
        
        if packinglist_data_to_sort:
            for item in packinglist_data_to_sort:
                if has_date_in_status(item.get('ltl_follow_status', '')):
                    date_group_data.append(item)
                else:
                    remaining_packinglist_data.append(item)

        # еѓєжЬЙжЧ•жЬЯзЪДзїДжМЙжЧ•жЬЯдїОе∞ПеИ∞е§ІжОТеЇП
        if date_group_data:
            date_group_data.sort(key=lambda x: get_pickup_date_key(x.get('ltl_follow_status', '')))
            data += date_group_data

        # еѓєеЙ©дљЩзЪД pallet_data жМЙеОЯжЭ•зЪДиІДеИЩжОТеЇП
        if remaining_pallet_data:
            remaining_pallet_data.sort(key=lambda x: (
                # зђђдЇМзїДпЉЪltl_correlation_id жЬЙеАЉзЪД
                0 if x.get('ltl_correlation_id') else 1,
                # зЫЄеРМзЪД ltl_correlation_id жОТеЬ®дЄАиµЈ
                x.get('ltl_correlation_id') or '',
                # зђђдЄЙзїДпЉЪltl_follow_status жЬЙеАЉдљЖж≤°жЬЙжЧ•жЬЯзЪД
                0 if (x.get('ltl_follow_status') and not has_date_in_status(x.get('ltl_follow_status'))) else 1,
                # зђђеЫЫзїДпЉЪdelivery_method еМЕеРЂжЪВжЙ£зЪД
                0 if (x.get('delivery_method') and 'жЪВжЙ£' in str(x.get('delivery_method'))) else 1,
                # еРОзї≠жОТеЇП
                x.get('offload_at') or '',
                x.get('destination') or '',
                x.get('shipping_marks') or x.get('shipping_mark') or ''
            ))
            data += remaining_pallet_data

        # еѓєеЙ©дљЩзЪД packinglist_data жМЙеОЯжЭ•зЪДиІДеИЩжОТеЇП
        if remaining_packinglist_data:
            remaining_packinglist_data.sort(key=lambda x: (
                # зђђдЄАзїДпЉЪltl_follow_status жЬЙеАЉзЪДпЉИж≤°жЬЙжЧ•жЬЯзЪДпЉМеЫ†дЄЇжЬЙжЧ•жЬЯзЪДеЈ≤зїПеЬ®еЙНйЭҐдЇЖпЉЙ
                0 if x.get('ltl_follow_status') else 1,
                # зђђдЇМзїДпЉЪoffload_tag жЬЙеЃЮйЩЕжПРжЯЬзЪД
                0 if x.get('offload_tag') == 'еЃЮйЩЕжПРжЯЬ' else 1,
                # зђђдЄЙзїДпЉЪoffload_tag жЬЙеЃЮйЩЕжФЊи°МзЪД
                0 if x.get('offload_tag') == 'еЃЮйЩЕжФЊи°М' else 1,
                # еРМдЄАжЯЬеПЈзЪДжОТеЬ®дЄАиµЈ
                x.get('container_numbers') or x.get('cns') or '',
                # еРОзї≠жОТеЇП
                x.get('offload_at') or '',
                x.get('destination') or '',
                x.get('shipping_marks') or x.get('shipping_mark') or ''
            ))
            data += remaining_packinglist_data
        # еЈ≤жЛЖжЯЬзЪДпЉМжМЙзЕІдЄЛйЭҐжОТеЇПпЉЪ
            # зђђдЄАдЉШеЕИзЇІпЉЪpickup жЬЙжЧ•жЬЯзЪДпЉМ
            # зђђдЇМдЉШеЕИзЇІпЉЪжЧ•жЬЯжМЙдїОе∞ПеИ∞е§ІжОТеИЧпЉМ
            # зђђдЄЙдЉШеЕИзЇІпЉЪжЬЙеЕ≥иБФIDзЪДпЉМ
            # зђђеЫЫдЉШеЕИзЇІпЉЪpickup жЬЙжЦЗе≠ЧзЪД
            # зђђдЇФдЉШеЕИзЇІпЉЪ жЪВжЙ£зЪД
        # жЬ™жЛЖжЯЬзЪДпЉЪ
            # жЬЙеЃЮйЩЕжПРжЯЬжЧґйЧізЪДпЉМеЃЮйЩЕдЄ§дЄ™е≠ЧжЫњжНҐжИРеЃЮйЩЕжПРжЯЬ
            # жЬЙеЃЮйЩЕжФЊи°МжЧґйЧіпЉМжФЊи°МдЄ§дЄ™е≠ЧжЫњжНҐжИРеЃЮйЩЕжФЊи°М

        # 3. дЄЇвАЬдЄАжПРе§ЪеНЄвАЭзїДеИЖйЕНиГМжЩѓйҐЬиЙ≤
        # еЃЪдєЙдЄАзїДжµЕиЙ≤и∞ГиГМжЩѓпЉМйБњеЉАзЇҐиЙ≤/йїДиЙ≤з≠Йи≠¶еСКиЙ≤
        color_palette = ["#e8f4fd", "#eafaf1", "#fef9e7", "#f4ecf7", "#ebf5fb", "#fdf2e9", "#e8f8f5"]
        id_to_color = {}
        color_idx = 0
        for item in data:
            corr_id = item.get('ltl_correlation_id')
            if corr_id:
                if corr_id not in id_to_color:
                    # еИЖйЕНйҐЬиЙ≤
                    id_to_color[corr_id] = color_palette[color_idx % len(color_palette)]
                    color_idx += 1
                item['correlation_bg_color'] = id_to_color[corr_id]
            else:
                item['correlation_bg_color'] = "" # жЧ†йҐЬиЙ≤
        return data
    
    async def _ltl_unscheduled_cargo(self, pl_criteria, plt_criteria) -> Dict[str, Any]:
        """иОЈеПЦжЬ™жФЊи°МиіІзЙ© - Tab 1"""
        plt_criteria = models.Q(pk__isnull=True) & models.Q(pk__isnull=False)
        # жЬ™жФЊи°МжЭ°дїґпЉЪж≤°жЬЙжЙєжђ°еПЈпЉМж≤°жЬЙеНЄиіІеЃМжИРпЉМETAеЬ®дЄ§еС®еЖЕпЉМдЄНжШѓеЉВеЄЄзКґжАБ      
        raw_cargos = await self._ltl_packing_list(
            pl_criteria,
            plt_criteria
        )
        cargos = []
        for cargo in raw_cargos:
            if not cargo["is_pass"]:
                cargos.append(cargo)
        sorted_cargos = sorted(cargos, key=lambda x: (x.get('ltl_verify', False),))
        return sorted_cargos
    
    async def _ltl_scheduled_self_pickup(self, pl_criteria, plt_criteria) -> Dict[str, Any]:
        """иОЈеПЦеЈ≤жФЊи°МеЃҐжПРиіІзЙ© - Tab 2"""
        
        # еЈ≤жФЊи°МеЃҐжПРжЭ°дїґпЉЪжЬЙжЙєжђ°еПЈпЉМdelivery_typeдЄЇself_pickup
        pl_criteria = pl_criteria&Q(
            delivery_method__contains="иЗ™жПР"
        )
        
        plt_criteria = plt_criteria&Q(
            delivery_method__contains="иЗ™жПР"
        )
        
        raw_cargos = await self._ltl_packing_list(
            pl_criteria,
            plt_criteria
        )
        cargos = []
        for cargo in raw_cargos:
            if cargo["is_pass"]:
                cargos.append(cargo)
        return cargos
    
    async def _ltl_self_delivery(self, pl_criteria, plt_criteria) -> Dict[str, Any]:
        """иОЈеПЦеЈ≤жФЊи°МиЗ™еПСиіІзЙ© - Tab 3"""
        pl_criteria = pl_criteria & ~Q(delivery_method__contains="иЗ™жПР")
        
        plt_criteria = plt_criteria & ~Q(delivery_method__contains="иЗ™жПР")
        raw_cargos = await self._ltl_packing_list(
            pl_criteria,
            plt_criteria
        )
        cargos = []
        for cargo in raw_cargos:
            if cargo["is_pass"]:
                cargos.append(cargo)
        return cargos
    
    async def export_ltl_unscheduled(
        self, request: HttpRequest
    ) -> HttpResponse:
        
        cargo_ids = request.POST.get('cargo_ids', '')
        
        # жЮДеїЇз≠ЫйАЙжЭ°дїґ
        pl_criteria = Q()
        plt_criteria = models.Q(pk__isnull=True) & models.Q(pk__isnull=False)
        
        # е¶ВжЮЬжМЗеЃЪдЇЖ IDпЉМеИЩеП™еѓЉеЗЇйАЙдЄ≠зЪДиіІзЙ©
        if cargo_ids:
            cargo_id_list = [int(id.strip()) for id in cargo_ids.split(',') if id.strip()]
            pl_criteria &= Q(id__in=cargo_id_list)
        
        # иОЈеПЦжХ∞жНЃ
        release_cargos,_ ,_ = await self._ltl_unscheduled_cargo(pl_criteria, plt_criteria)
        
        # еЗЖе§З Excel жХ∞жНЃ
        excel_data = []
        for cargo in release_cargos:
            # иОЈеПЦжХ∞жНЃеєґж†ЉеЉПеМЦ
            customer_name = cargo.get('customer_name', '-')
            container_numbers = cargo.get('container_numbers', '-')
            destination = cargo.get('destination', '-')
            shipping_marks = cargo.get('shipping_marks', '-')
            address = cargo.get('address', '-')
            note = cargo.get('note', '-')
            
            # ж†ЉеЉПеМЦжХ∞е≠Ч
            try:
                total_cbm = float(cargo.get('total_cbm', 0))
                total_cbm = round(total_cbm, 3)
            except (ValueError, TypeError):
                total_cbm = 0
            
            try:
                total_pcs = int(cargo.get('total_pcs', 0))
            except (ValueError, TypeError):
                total_pcs = 0
            
            try:
                weight_lbs = float(cargo.get('total_weight_lbs', 0))
                weight_lbs = round(weight_lbs, 2)
            except (ValueError, TypeError):
                weight_lbs = 0
            
            try:
                weight_kg = float(cargo.get('total_weight_kg', 0))
                weight_kg = round(weight_kg, 2)
            except (ValueError, TypeError):
                weight_kg = 0
            
            # ж†ЄеЃЮзКґжАБ
            ltl_verify = cargo.get('ltl_verify', False)
            verify_status = 'еЈ≤ж†ЄеЃЮ' if ltl_verify else 'жЬ™ж†ЄеЃЮ'
            
            row = {
                'еЃҐжИЈ': customer_name,
                'жЯЬеПЈ': container_numbers,
                'зЫЃзЪДеЬ∞': destination,
                'еФЫе§і': shipping_marks,
                'иѓ¶зїЖеЬ∞еЭА': address,
                'е§Зж≥®': note,
                'CBM': total_cbm,
                'дїґжХ∞': total_pcs,
                'йЗНйЗП(lbs)': weight_lbs,
                'йЗНйЗП(kg)': weight_kg,
                'ж†ЄеЃЮзКґжАБ': verify_status,
            }
            
            excel_data.append(row)
        
        # еИЫеїЇ DataFrame
        df = pd.DataFrame(excel_data)
        
        # е¶ВжЮЬж≤°жЬЙжХ∞жНЃпЉМеИЫеїЇдЄАдЄ™з©ЇзЪДDataFrame
        if df.empty:
            df = pd.DataFrame(columns=[
                'еЃҐжИЈ', 'жЯЬеПЈ', 'зЫЃзЪДеЬ∞', 'еФЫе§і', 'иѓ¶зїЖеЬ∞еЭА', 'е§Зж≥®',
                'CBM', 'дїґжХ∞', 'йЗНйЗП(lbs)', 'йЗНйЗП(kg)', 'ж†ЄеЃЮзКґжАБ'
            ])
        
        # еИЫеїЇ Excel жЦЗдїґ
        output = BytesIO()
        
        # дљњзФ® ExcelWriter
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # дЄїжХ∞жНЃ sheet
            df.to_excel(writer, sheet_name='жЬ™жФЊи°МиіІзЙ©', index=False)
            
            # иОЈеПЦ worksheet еѓєи±°
            worksheet = writer.sheets['жЬ™жФЊи°МиіІзЙ©']
            
            # иЃЊзљЃеИЧеЃљ
            column_widths = {
                'еЃҐжИЈ': 20,
                'жЯЬеПЈ': 25,
                'зЫЃзЪДеЬ∞': 15,
                'еФЫе§і': 25,
                'иѓ¶зїЖеЬ∞еЭА': 40,
                'е§Зж≥®': 40,
                'CBM': 10,
                'дїґжХ∞': 10,
                'йЗНйЗП(lbs)': 12,
                'йЗНйЗП(kg)': 12,
                'ж†ЄеЃЮзКґжАБ': 12,
            }
            
            # иЃЊзљЃеИЧеЃљ
            from openpyxl.utils import get_column_letter
            
            for i, column in enumerate(df.columns, 1):
                col_letter = get_column_letter(i)
                width = column_widths.get(column, 15)
                worksheet.column_dimensions[col_letter].width = width
            
            # иЃЊзљЃжХ∞е≠Чж†ЉеЉП
            from openpyxl.styles import numbers
            
            # иЃЊзљЃCBMеИЧдЄЇ3дљНе∞ПжХ∞ж†ЉеЉП
            if 'CBM' in df.columns:
                cbm_col_idx = df.columns.get_loc('CBM') + 1
                cbm_col_letter = get_column_letter(cbm_col_idx)
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"{cbm_col_letter}{row}"]
                    cell.number_format = numbers.FORMAT_NUMBER_00  # йїШиЃ§2дљНпЉМExcelдЉЪиЗ™еК®жШЊз§ЇеЃЮйЩЕе∞ПжХ∞дљНжХ∞
                    # е¶ВжЮЬйЬАи¶БеЬ®ExcelдЄ≠еЉЇеИґжШЊз§Ї3дљНе∞ПжХ∞пЉМдљњзФ®пЉЪ
                    # cell.number_format = '0.000'
            
            # иЃЊзљЃйЗНйЗПеИЧдЄЇ2дљНе∞ПжХ∞ж†ЉеЉП
            if 'йЗНйЗП(lbs)' in df.columns:
                lbs_col_idx = df.columns.get_loc('йЗНйЗП(lbs)') + 1
                lbs_col_letter = get_column_letter(lbs_col_idx)
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"{lbs_col_letter}{row}"]
                    cell.number_format = numbers.FORMAT_NUMBER_00
            
            if 'йЗНйЗП(kg)' in df.columns:
                kg_col_idx = df.columns.get_loc('йЗНйЗП(kg)') + 1
                kg_col_letter = get_column_letter(kg_col_idx)
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"{kg_col_letter}{row}"]
                    cell.number_format = numbers.FORMAT_NUMBER_00
            
            # иЃЊзљЃж†ЈеЉПпЉЪж†ЗйҐШи°МеК†з≤Ч
            for cell in worksheet[1]:
                cell.font = cell.font.copy(bold=True)
            
            # иЗ™еК®жНҐи°МиЃЊзљЃ
            from openpyxl.styles import Alignment
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            
            # еѓєеПѓиГљжЬЙе§Ъи°МеЖЕеЃєзЪДеИЧиЃЊзљЃиЗ™еК®жНҐи°М
            wrap_columns = ['жЯЬеПЈ', 'иѓ¶зїЖеЬ∞еЭА', 'е§Зж≥®', 'еФЫе§і']
            for col_name in wrap_columns:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    for row in range(1, len(df) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        cell.alignment = wrap_alignment
            
            # жЈїеК†з≠ЫйАЙеЩ®
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"
            
            # еПѓйАЙпЉЪеЖїзїУж†ЗйҐШи°М
            worksheet.freeze_panes = 'A2'
        
        output.seek(0)
        
        # еИЫеїЇ HTTP еУНеЇФ
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # зФЯжИРжЦЗдїґеРН
        timestamp = timezone.now().strftime('_%m%d')
        filename = f'жЬ™жФЊи°МиіІзЙ©_{timestamp}.xlsx'
        
        # еѓєжЦЗдїґеРНињЫи°М URL зЉЦз†БпЉМз°ЃдњЭдЄ≠жЦЗж≠£з°Ѓе§ДзРЖ
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename)
        
        # еИЫеїЇ HTTP еУНеЇФ
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # дљњзФ® RFC 6266 ж†ЗеЗЖиЃЊзљЃ Content-Disposition
        # ињЩж†ЈиГљз°ЃдњЭжЙАжЬЙжµПиІИеЩ®йГљиГљж≠£з°ЃжШЊз§ЇдЄ≠жЦЗжЦЗдїґеРН
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
        
        # е§ЗзФ®жЦєж°ИпЉЪеѓєдЇОдЄНжФѓжМБ RFC 6266 зЪДжЧІжµПиІИеЩ®
        response['Content-Disposition'] = f"attachment; filename={encoded_filename}"
        
        return response

    async def handle_save_releaseCommand(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        '''еНХжЭ°жИЦжЙєйЗПдњЭе≠ШжЬ™жФЊи°МзЪДжМЗдї§жХ∞жНЃ'''
        # 1. е∞ЭиѓХиОЈеПЦжЙєйЗПжХ∞жНЃ
        batch_commands_raw = request.POST.get('batch_commands')
        tasks = []

        if batch_commands_raw:
            try:
                commands_list = json.loads(batch_commands_raw)
            except json.JSONDecodeError:
                commands_list = []
        else:
            cargo_id = request.POST.get('cargo_id')
            release_command = request.POST.get('release_command')
            if cargo_id:
                commands_list = [{'cargo_id': cargo_id, 'command': release_command}]
            else:
                commands_list = []
        num = 0
        for item in commands_list:
            c_id = item.get('cargo_id')
            command_text = item.get('command')
            if not c_id or not command_text: continue

            if c_id.startswith('plt_'):
                target_ids = c_id.replace('plt_', '').split(',')
                model = Pallet
                
            else:
                target_ids = c_id.split(',')
                model = PackingList
            update_data = {'ltl_release_command': command_text}
            await sync_to_async(model.objects.filter(id__in=target_ids).update)(**update_data)
            num+=1
        context = {'success_messages': f'дњЭе≠ШжИРеКЯ{num}зїДжХ∞жНЃ!'}
        return await self.handle_ltl_unscheduled_pos_post(request,context)

    async def handle_save_shipment_note(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        cargo_ids_raw = (request.POST.get("cargo_ids") or "").strip()
        plt_ids_raw = (request.POST.get("plt_ids") or "").strip()
        note = (request.POST.get("shipment_note") or "").strip()

        if cargo_ids_raw and plt_ids_raw:
            context = {"error_messages": "cargo_ids еТМ plt_ids йГљдЄЇз©Ї"}
            return await self.handle_td_shipment_post(request, context)

        def parse_ids(raw: str) -> list[int]:
            ids: list[int] = []
            for part in raw.split(","):
                part_str = part.strip()
                if part_str.isdigit():
                    ids.append(int(part_str))
            return ids

        updated = 0
        if cargo_ids_raw:
            cargo_ids = parse_ids(cargo_ids_raw)
            if not cargo_ids:
                context = {"error_messages": "cargo_ids дЄЇз©Ї"}
                return await self.handle_td_shipment_post(request, context)
            updated = await PackingList.objects.filter(id__in=cargo_ids).aupdate(
                shipment_note=note or None
            )
        elif plt_ids_raw:
            plt_ids = parse_ids(plt_ids_raw)
            if not plt_ids:
                context = {"error_messages": "plt_ids дЄЇз©Ї"}
                return await self.handle_td_shipment_post(request, context)
            updated = await Pallet.objects.filter(id__in=plt_ids).aupdate(
                shipment_note=note or None
            )
        else:
            context = {"error_messages": "жЬ™жПРдЊЫ cargo_ids жИЦ plt_ids"}
            return await self.handle_td_shipment_post(request, context)

        context = {"success_messages": f"жОТзЇ¶е§Зж≥®дњЭе≠ШжИРеКЯ"}
        return await self.handle_td_shipment_post(request, context)
    
    async def handle_save_selfdel_cargo(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        '''LTLдњЭе≠ШиЗ™и°МзЉЦиЊСзЪДиіІзЙ©дњ°жБѓ'''
        # 1. еИ§жЦ≠жШѓжЙєйЗПдњЭе≠ШињШжШѓеНХи°МдњЭе≠Ш
        batch_data_raw = request.POST.get('batch_data')
        print(request.POST)
        if batch_data_raw:
            try:
                update_items = json.loads(batch_data_raw)
            except json.JSONDecodeError:
                update_items = []
        else:
            update_items = [{
                'cargo_id': request.POST.get('cargo_id'),
                'address': request.POST.get('address', '').strip(),
                'note': request.POST.get('note', '').strip(),
                'pallet_size': request.POST.get('pallet_size', '').strip(),
                'carrier_company': request.POST.get('carrier_company', '').strip(),
                'bol_number': request.POST.get('bol_number', '').strip(),
                'pro_number': request.POST.get('pro_number', '').strip(),
                'follow_status': request.POST.get('follow_status', '').strip(),
                'ltl_cost_note': request.POST.get('ltl_cost_note', '').strip(),
                'ltl_quote_note': request.POST.get('ltl_quote_note', '').strip(),
                'contact_method': request.POST.get('contact_method', '').strip(),
                'ltl_cost': request.POST.get('ltl_cost', '').strip(),
                'ltl_quote': request.POST.get('ltl_quote', '').strip(),
                'delivery_method': request.POST.get('delivery_method', '').strip(),
                'ltl_release_command': request.POST.get('ltl_release_command', '').strip(),
                'ltl_supplier': request.POST.get('ltl_supplier', '').strip(),
            }]

        total_status_messages = []
        username = request.user.username
        # 2. еЊ™зОѓе§ДзРЖжѓПдЄАдЄ™жЫіжЦ∞й°є
        for item in update_items:
            cargo_id = item.get('cargo_id')
            if not cargo_id: continue

            # --- дї•дЄЛдЄЇдљ†еОЯжЬЙзЪДеНХи°Ме§ДзРЖйАїиЊСпЉМдїЕе∞Ж request.POST.get жЫњжНҐдЄЇ item.get ---
            address = item.get('address', '')
            note = item.get('note', '')
            pallet_size = item.get('pallet_size', '')
            carrier_company = item.get('carrier_company', '')
            bol_number = item.get('bol_number', '')
            pro_number = item.get('pro_number', '')
            follow_status = item.get('follow_status', '')
            ltl_cost_note = item.get('ltl_cost_note', '')
            ltl_quote_note = item.get('ltl_quote_note', '')
            contact_method = item.get('contact_method', '')
            delivery_method = item.get('delivery_method')
            ltl_release_command = item.get('ltl_release_command')
            ltl_supplier = item.get('ltl_supplier', '')

            ltl_cost_raw = item.get('ltl_cost', '')
            has_ltl_cost_param = bool(ltl_cost_raw)
            ltl_quote_raw = item.get('ltl_quote', '')
            has_ltl_quote_param = bool(ltl_quote_raw)

            ltl_cost = float(ltl_cost_raw) if has_ltl_cost_param else None
            ltl_quote = float(ltl_quote_raw) if has_ltl_quote_param else None

            # ж†єжНЃ ID з±їеЮЛз°ЃеЃЪе≠ЧжЃµеРН
            if cargo_id.startswith('plt_'):
                ids = cargo_id.replace('plt_', '').split(',')
                model = Pallet
            else:
                ids = cargo_id.split(',')
                model = PackingList
        
            # жЮДеїЇжЫіжЦ∞е≠ЧеЕЄ
            update_data = {}
            
            if carrier_company or carrier_company == '': update_data['carrier_company'] = carrier_company
            if address or address == '': update_data['address'] = address
            if bol_number or bol_number == '': update_data['ltl_bol_num'] = bol_number
            if pro_number or pro_number == '': update_data['ltl_pro_num'] = pro_number
            if note or note == '': update_data['note'] = note
            if follow_status or follow_status == '': update_data['ltl_follow_status'] = follow_status
            if has_ltl_cost_param: update_data['ltl_cost'] = ltl_cost
            if has_ltl_quote_param: update_data['ltl_quote'] = ltl_quote
            if ltl_cost_note: update_data["ltl_cost_note"] = ltl_cost_note
            if ltl_quote_note: update_data["ltl_quote_note"] = ltl_quote_note
            if contact_method: update_data["ltl_contact_method"] = contact_method
            if pallet_size: update_data["ltl_plt_size_note"] = pallet_size
            if delivery_method is not None: update_data['delivery_method'] = delivery_method
            if ltl_release_command is not None: update_data['ltl_release_command'] = ltl_release_command
            if ltl_supplier: update_data['ltl_supplier'] = ltl_supplier

            # жЙєйЗПжЫіжЦ∞йАЪзФ®е≠ЧжЃµ
            if update_data:
                await sync_to_async(model.objects.filter(id__in=ids).update)(**update_data)
        
            if cargo_id.startswith('plt_') and pallet_size:
                await self._save_pallet_sizes(ids, pallet_size)
            
            if has_ltl_quote_param or 'зїДеРИжЯЬ' in ltl_quote_note:
                if not ltl_quote_note: ltl_quote_note = 'жіЊйАБиіє'
                msg = await self._delivery_account_entry(ids, ltl_quote, ltl_quote_note, username)
                if msg: total_status_messages.append(msg)

        success_message = f'жИРеКЯдњЭе≠Ш {len(update_items)} зїДжХ∞жНЃпЉБ'
        if total_status_messages:
            success_message = mark_safe(f"{success_message}<br>" + "<br>".join(set(total_status_messages)))

        context = {'success_messages': success_message}
        page = request.POST.get('page')
        if page == "history":
            return await self.handle_ltl_history_pos_post(request, context)
        return await self.handle_ltl_unscheduled_pos_post(request, context)

    async def _delivery_account_entry(self, ids, ltl_quote, ltl_quote_note, username):
        pallets = await sync_to_async(list)(
            Pallet.objects.filter(id__in=ids)
            .select_related('container_number')
        )

        # жМЙ PO_ID-shipping_marks-container_number еИЖзїД
        pallet_index = defaultdict(list)
        for pallet in pallets:
            po_id = getattr(pallet, "PO_ID", None) or "жЧ†PO_ID"
            if not po_id:
                raise ValueError('idдЄЇ{pallet.id}зЪДpalletж≤°жЬЙPO_ID')
            shipping_mark = getattr(pallet, "shipping_mark")
            if not shipping_mark:
                raise ValueError('idдЄЇ{pallet.id}зЪДpalletж≤°жЬЙеФЫе§і')
            container_num = pallet.container_number
            index_key = f"{po_id}-{shipping_mark}-{container_num.id}"
            pallet_index[index_key].append(pallet)

        # йБНеОЖжѓПзїД
        for index_key, group_pallets in pallet_index.items():
            if len(group_pallets) <= 0:
                continue
            first_pallet = group_pallets[0]
            po_id = getattr(first_pallet, "PO_ID")
            shipping_mark = getattr(first_pallet, "shipping_mark")
            container = first_pallet.container_number
            qty = len(group_pallets)
            total_cbm = sum(getattr(p, "cbm", 0) or 0 for p in group_pallets)
            total_weight = sum(getattr(p, "weight_lbs", 0) or 0 for p in group_pallets)

            # ж£АжЯ• InvoiceItemv2 жШѓеР¶еЈ≤жЬЙиЃ∞ељХ
            existing_item = await sync_to_async(
                lambda: InvoiceItemv2.objects.filter(
                    PO_ID=po_id,
                    shipping_marks=shipping_mark,
                    container_number=container
                ).first()
            )()

            if existing_item:
                receivable_statuses = await sync_to_async(
                    lambda: InvoiceStatusv2.objects.filter(
                        invoice=existing_item.invoice_number,
                        invoice_type="receivable"
                    ).first()
                )()
                if receivable_statuses and receivable_statuses.finance_status == "completed":
                    return 'иі¶еНХеЈ≤иҐЂиіҐеК°з°ЃиЃ§дЄНеПѓдњЃжФєеЗЇеЇУиієпЉБ'
                if "зїДеРИжЯЬ" in ltl_quote_note:
                    await self._single_po_match_combina(container, group_pallets, False, username, qty, existing_item)
                else:
                    # жЫіжЦ∞еОЯиЃ∞ељХ
                    existing_item.qty = qty
                    existing_item.rate = ltl_quote
                    existing_item.cbm = total_cbm
                    existing_item.weight = total_weight
                    existing_item.amount = ltl_quote
                    existing_item.note = ltl_quote_note
                    existing_item.description = ltl_quote_note
                    existing_item.item_category = "delivery_other"
                    existing_item.warehouse_code = getattr(first_pallet, "destination", "")
                    await sync_to_async(existing_item.save)()
            else:
                # жЯ• invoice_number
                invoice_record = await sync_to_async(
                    lambda: Invoicev2.objects.filter(container_number=container).first()
                )()

                if not invoice_record:
                    # и∞ГзФ®иЗ™еЃЪдєЙжЦєж≥ХеИЫеїЇ invoice
                    invoice_record, invoice_status = await self._create_invoice_and_status(container)

                if "зїДеРИжЯЬ" in ltl_quote_note:
                    #жМЙзїДеРИжЯЬжЦєеЉПиЃ°зЃЧ
                    await self._single_po_match_combina(container, group_pallets, invoice_record, username, qty, False)
                else:
                    # еИЫеїЇжЦ∞иЃ∞ељХ
                    item = InvoiceItemv2(
                        container_number=container,
                        invoice_number=invoice_record,
                        invoice_type="receivable",
                        item_category="delivery_other",
                        description=ltl_quote_note,
                        warehouse_code=getattr(first_pallet, "destination", ""),
                        shipping_marks=shipping_mark,
                        rate=ltl_quote,
                        amount=ltl_quote,
                        qty=qty,
                        cbm=total_cbm,
                        weight=total_weight,
                        delivery_type="selfdelivery",
                        PO_ID=po_id,
                        note=ltl_quote_note,
                        registered_user=username 
                    )
                    await sync_to_async(item.save)()

        # зЬЛдЄЛињЩдЄ™жЯЬе≠РзІБдїУжіЊйАБжШѓдЄНжШѓйГљељХеЃМдЇЖпЉМељХеЃМдЇЖе∞±жФєзКґжАБ
        container = pallets[0].container_number
        status_message = await self._try_complete_delivery_other_status(container)
        return status_message
    
    async def _single_po_match_combina(self, container, group, invoice_record, username, qty, existing_item):
        '''зІБдїУеМєйЕНзїДеРИжЯЬиЃ°иіє'''
        order = await sync_to_async(
            Order.objects.select_related(
                'retrieval_id',
                'vessel_id',
                'customer_name'
            ).filter(container_number=container).first
        )()
        
        quotations = await self._get_fee_details(order, order.retrieval_id.retrieval_destination_area,order.customer_name.zem_name)
        if isinstance(quotations, dict) and quotations.get("error_messages"):
            return {"error_messages": quotations["error_messages"]}
        fee_details = quotations['fees']

        warehouse = order.retrieval_id.retrieval_destination_area
        container_type_temp = 0 if "40" in container.container_type else 1
        combina_key = f"{warehouse}_COMBINA"
        if combina_key not in fee_details:
            context = {
                "error_messages": f"жЬ™жЙЊеИ∞зїДеРИжЯЬжК•дїЈи°®иІДеИЩ {combina_key},жК•дїЈи°®жШѓ{quotations['filename']}"
            }
            return (context, [])  # ињФеЫЮйФЩиѓѓпЉМз©ЇеИЧи°®
        
        rules = fee_details.get(combina_key).details
        first_pallet = group[0]
        po_id = getattr(first_pallet, "PO_ID", "")
        destination_str = getattr(first_pallet, "destination", "")
        destination_origin, destination = self._process_destination(destination_str)
        is_combina_region = False

        # еОїйҐДжК•йЗМжЙЊжАїcbmеТМйЗНйЗП
        total_pl_details = await sync_to_async(
            lambda: PackingList.objects.filter(
                PO_ID=po_id,
                container_number=container
            ).aggregate(
                total_cbm=Coalesce(Sum('cbm'), 0.0),
                total_weight=Coalesce(Sum('total_weight_lbs'), 0.0)
            )
        )()
        THRESHOLD = 0.001
        if total_pl_details['total_cbm'] > THRESHOLD:
            total_cbm = total_pl_details['total_cbm']
        else:
            total_cbm = sum(float(getattr(p, "cbm", 0.0) or 0.0) for p in group)

        if total_pl_details['total_weight'] > THRESHOLD:
            total_weight = total_pl_details['total_weight']
        else:
            total_weight = sum(float(getattr(p, "weight_lbs", 0.0) or 0.0) for p in group)
        
        for region, region_data in rules.items():
            for item in region_data:
                rule_locations = item.get("location", [])
                if isinstance(rule_locations, str):
                    rule_locations = [rule_locations] # зїЯдЄАиљђжИРеИЧи°®е§ДзРЖ
                if any(destination == loc.replace(" ", "").upper() for loc in rule_locations):
                    is_combina_region = True
                    price = item["prices"][container_type_temp]
                    match_region = region
                    break
            if is_combina_region:
                break
        if destination == "UPS":
            is_combina_region = False

        if is_combina_region:
            # иЃ°зЃЧжАїCBM
            total_container_cbm_result = await sync_to_async(
                lambda: PackingList.objects.filter(
                    container_number=container
                ).aggregate(
                    total_cbm=Coalesce(Sum('cbm'), 0.0)
                )
            )()
            cbm_ratio = round(total_cbm / total_container_cbm_result.get('total_cbm', 0), 4)
            ltl_quote = price * cbm_ratio
            shipping_marks=getattr(first_pallet, "shipping_mark", "")
            if existing_item:
                existing_item.qty = qty
                existing_item.rate = price
                existing_item.cbm = total_cbm
                existing_item.weight = total_weight
                existing_item.amount = ltl_quote
                existing_item.warehouse_code = destination_str
                existing_item.cbm_ratio = cbm_ratio
                existing_item.registered_user = username
                existing_item.delivery_type = "combine"
                existing_item.description = "жіЊйАБиіє"
                existing_item.region = match_region
                existing_item.regionPrice = price
                existing_item.item_category = "delivery_other"
                existing_item.shipping_marks = shipping_marks
                await sync_to_async(existing_item.save)()
            else:
                item = InvoiceItemv2(
                    container_number=container,
                    invoice_number=invoice_record,
                    invoice_type="receivable",
                    item_category="delivery_other",
                    description="жіЊйАБиіє",
                    warehouse_code=destination_str,
                    shipping_marks=shipping_marks,
                    rate=price,
                    amount=ltl_quote,
                    qty=qty,
                    cbm=total_cbm,
                    cbm_ratio=cbm_ratio,
                    weight=total_weight,
                    delivery_type="combine",
                    PO_ID=po_id,
                    region=match_region,
                    regionPrice=price,
                    registered_user=username 
                )
                await sync_to_async(item.save)() 
        else:
            raise ValueError(f"жЬ™еЬ®жК•дїЈи°®зЪДзїДеРИжЯЬиМГеЫіеЖЕжЙЊеИ∞ињЩдЄ™еМЇ{quotations['filename']}")    

    async def _try_complete_delivery_other_status(self, container):
        """
        еИ§жЦ≠иѓ• container дЄЛжЙАжЬЙеЇФељХеЕ•зЪД delivery_other жШѓеР¶еЈ≤еЃМжИР
        еЃМжИРеИЩжЫіжЦ∞ InvoiceStatusv2.delivery_other_status = completed
        """

        # 1пЄПвГ£ жЯ•иѓ• container дЄЛеЇФиЃ°жіЊйАБиієзЪД pallet
        delivery_pallets = await sync_to_async(list)(
            Pallet.objects.filter(
                container_number=container,
                delivery_type="other"
            ).exclude(
                delivery_method__icontains="жЪВжЙ£"
            )
        )

        if not delivery_pallets:
            return

        # 2пЄПвГ£ жЮДйА†вАЬеЇФе≠ШеЬ®вАЭзЪДзіҐеЉХйЫЖеРИ
        expected_keys = set()
        for pallet in delivery_pallets:
            expected_keys.add(
                (
                    pallet.PO_ID,
                    pallet.shipping_mark,
                    container.id
                )
            )

        # 3пЄПвГ£ жЯ•еЃЮйЩЕеЈ≤е≠ШеЬ®зЪД InvoiceItemv2
        existing_keys = set(
            await sync_to_async(list)(
                InvoiceItemv2.objects.filter(
                    container_number=container,
                    item_category="delivery_other",
                    invoice_type="receivable"
                ).values_list(
                    "PO_ID",
                    "shipping_marks",
                    "container_number_id"
                )
            )
        )
        # 4пЄПвГ£ еЕ®йГ®еЈ≤ељХ вЖТ жЫіжЦ∞зКґжАБ
        if expected_keys.issubset(existing_keys):
            invoice_status = await sync_to_async(
                lambda: InvoiceStatusv2.objects.filter(
                    container_number=container,
                    invoice_type="receivable"
                ).first()
            )()

            if invoice_status and invoice_status.delivery_other_status != "completed":
                invoice_status.delivery_other_status = "completed"
                await sync_to_async(invoice_status.save)()
            return f"иѓ•жЯЬе≠РжЙАжЬЙзІБдїУжіЊйАБиі¶еНХеЈ≤ељХеЃМ"
        return None

    async def _create_invoice_and_status(
        self,
        container: Container
    ) -> tuple[Invoicev2, InvoiceStatusv2]:
        """еЉВж≠•еИЫеїЇиі¶еНХеТМзКґжАБиЃ∞ељХ"""

        # 1пЄПвГ£ жЯ• OrderпЉИеРМж≠• ORM вЖТ async еМЕи£ЕпЉЙ
        order = await sync_to_async(
            lambda: Order.objects.select_related(
                "customer_name", "container_number"
            ).get(container_number=container)
        )()

        current_date = datetime.now().date()
        order_id = str(order.id)
        customer_id = order.customer_name.id

        # 2пЄПвГ£ жЯ•жШѓеР¶еЈ≤жЬЙ Invoice
        existing_invoice = await sync_to_async(
            lambda: Invoicev2.objects.filter(
                container_number=container
            ).first()
        )()

        if existing_invoice:
            # 3пЄПвГ£ жЯ•жШѓеР¶еЈ≤жЬЙ Status
            existing_status = await sync_to_async(
                lambda: InvoiceStatusv2.objects.filter(
                    invoice=existing_invoice,
                    invoice_type="receivable"
                ).first()
            )()

            if existing_status:
                return existing_invoice, existing_status

        # 4пЄПвГ£ еИЫеїЇ Invoice
        invoice = await sync_to_async(Invoicev2.objects.create)(
            container_number=container,
            invoice_number=(
                f"{current_date.strftime('%Y%m%d')}C{customer_id}{order_id}"
            ),
            created_at=current_date,
            is_master_bill=True,
        )

        # 5пЄПвГ£ еИЫеїЇ InvoiceStatus
        invoice_status = await sync_to_async(InvoiceStatusv2.objects.create)(
            container_number=container,
            invoice=invoice,
            invoice_type="receivable",
        )

        return invoice, invoice_status
    
    async def _delivery_account_selfpick_entry(self, ids, ltl_quote, ltl_quote_note, ltl_unit_quote, del_qty, username):
        '''иЗ™жПРзЪДеЗЇеЇУиієељХеЕ•'''
        pallets = await sync_to_async(list)(
            Pallet.objects.filter(id__in=ids)
            .select_related('container_number')
        )

        # жМЙ PO_ID-shipping_marks-container_number еИЖзїД
        pallet_index = defaultdict(list)
        for pallet in pallets:
            po_id = getattr(pallet, "PO_ID", None) or "жЧ†PO_ID"
            if not po_id:
                raise ValueError('idдЄЇ{pallet.id}зЪДpalletж≤°жЬЙPO_ID')
            shipping_mark = getattr(pallet, "shipping_mark")
            if not shipping_mark:
                raise ValueError('idдЄЇ{pallet.id}зЪДpalletж≤°жЬЙеФЫе§і')
            container_num = pallet.container_number
            index_key = f"{po_id}-{shipping_mark}-{container_num.id}"
            pallet_index[index_key].append(pallet)

        # йБНеОЖжѓПзїД
        for index_key, group_pallets in pallet_index.items():
            first_pallet = group_pallets[0]
            po_id = getattr(first_pallet, "PO_ID")
            shipping_mark = getattr(first_pallet, "shipping_mark")
            container = first_pallet.container_number
            total_cbm = sum(getattr(p, "cbm", 0) or 0 for p in group_pallets)
            total_weight = sum(getattr(p, "weight_lbs", 0) or 0 for p in group_pallets)

            # ж£АжЯ• InvoiceItemv2 жШѓеР¶еЈ≤жЬЙиЃ∞ељХ
            existing_item = await sync_to_async(
                lambda: InvoiceItemv2.objects.filter(
                    PO_ID=po_id,
                    shipping_marks=shipping_mark,
                    container_number=container
                ).first()
            )()

            if existing_item:              
                receivable_statuses = await sync_to_async(
                    lambda: InvoiceStatusv2.objects.filter(
                        invoice=existing_item.invoice_number,
                        invoice_type="receivable"
                    ).first()
                )()
                if receivable_statuses and receivable_statuses.finance_status == "completed":
                    return 'иі¶еНХеЈ≤иҐЂиіҐеК°з°ЃиЃ§дЄНеПѓдњЃжФєеЗЇеЇУиієпЉБ'
                # жЫіжЦ∞еОЯиЃ∞ељХ
                existing_item.qty = del_qty
                existing_item.rate = ltl_unit_quote
                existing_item.cbm = total_cbm
                existing_item.weight = total_weight
                existing_item.amount = ltl_quote
                existing_item.description = 'еЗЇеЇУиіє'
                existing_item.note = ltl_quote_note
                existing_item.warehouse_code = getattr(first_pallet, "destination", "")
                await sync_to_async(existing_item.save)()
            else:
                # жЯ• invoice_number
                invoice_record = await sync_to_async(
                    lambda: Invoicev2.objects.filter(container_number=container).first()
                )()

                if not invoice_record:
                    # и∞ГзФ®иЗ™еЃЪдєЙжЦєж≥ХеИЫеїЇ invoice
                    invoice_record, invoice_status = await self._create_invoice_and_status(container)

                # еИЫеїЇжЦ∞иЃ∞ељХ
                item = InvoiceItemv2(
                    container_number=container,
                    invoice_number=invoice_record,
                    invoice_type="receivable",
                    item_category="delivery_other",
                    description="еЗЇеЇУиіє",
                    warehouse_code=getattr(first_pallet, "destination", ""),
                    shipping_marks=shipping_mark,
                    rate=ltl_unit_quote,
                    amount=ltl_quote,
                    note=ltl_quote_note,
                    qty=del_qty,
                    cbm=total_cbm,
                    weight=total_weight,
                    delivery_type="selfdelivery",
                    PO_ID=po_id,
                    registered_user=username 
                )
                await sync_to_async(item.save)()

        # зЬЛдЄЛињЩдЄ™жЯЬе≠РзІБдїУжіЊйАБжШѓдЄНжШѓйГљељХеЃМдЇЖпЉМељХеЃМдЇЖе∞±жФєзКґжАБ
        container = pallets[0].container_number
        status_message = await self._try_complete_delivery_other_status(container)
        return status_message
    
    async def handle_save_selfpick_cargo(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        '''LTLеЃҐжЬНиЗ™жПРдњ°жБѓдњЭе≠Ш'''
        batch_data_raw = request.POST.get('batch_data')
        if batch_data_raw:
            try:
                update_items = json.loads(batch_data_raw)
            except json.JSONDecodeError:
                update_items = []
        else:
            ltl_quote_note = request.POST.get('ltl_quote_note', '').strip()
            update_items = [{
                'cargo_id': request.POST.get('cargo_id'),
                'address': request.POST.get('address', '').strip(),
                'pallet_size': request.POST.get('pallet_size', '').strip(),
                'follow_status': request.POST.get('follow_status', '').strip(),
                'bol_number': request.POST.get('bol_number', '').strip(),
                'carrier_company': request.POST.get('carrier_company', '').strip(),
                'pickup_date': request.POST.get('pickup_date', '').strip(),
                'del_qty': request.POST.get('del_qty', '').strip(),
                'ltl_unit_quote': request.POST.get('ltl_unit_quote', '').strip(),
                'ltl_quote': request.POST.get('ltl_quote', '').strip(),
                'ltl_quote_note': request.POST.get('ltl_quote_note', '').strip(),
                'delivery_method': request.POST.get('delivery_method', '').strip(),
                'note': request.POST.get('note', '').strip(),
                'ltl_release_command': request.POST.get('ltl_release_command', '').strip(),
            }]

        total_updated = 0
        username = request.user.username
        billing_messages = []
        # 2. еЊ™зОѓе§ДзРЖжѓПдЄАдЄ™жЫіжЦ∞й°є
        for item in update_items:
            cargo_id = item.get('cargo_id')
            if not cargo_id:
                continue
            follow_status = item.get('follow_status', '')
            pallet_size = item.get('pallet_size', '')
            del_qty_raw = item.get('del_qty', '')
            unit_quote_raw = item.get('ltl_unit_quote', '')
            ltl_quote_raw = item.get('ltl_quote', '')
            ltl_quote_note = item.get('ltl_quote_note', '')
            bol_number = item.get('bol_number', '')
            carrier_company = item.get('carrier_company', '')
            address = item.get('address', '')
            pickup_date = item.get('pickup_date', '')
            delivery_method = item.get('delivery_method')
            note = item.get('note')
            ltl_release_command = item.get('ltl_release_command')
            est_pickup_time = None

            if pickup_date:
                try:
                    # иІ£жЮРжЧ•жЬЯе≠Чзђ¶дЄ≤пЉИж†ЉеЉПпЉЪYYYY-MM-DDпЉЙ
                    pickup_date = datetime.strptime(pickup_date, '%Y-%m-%d').date()
                    # е∞ЖжЧ•жЬЯиљђжНҐдЄЇеЄ¶жЧґйЧізЪДdatetimeпЉМйїШиЃ§жЧґйЧідЄЇ00:00
                    pickup_date = timezone.make_aware(
                        datetime.combine(pickup_date, time.min)
                    )
                except ValueError as e:
                    # е∞ЭиѓХеЕґдїЦеПѓиГљзЪДж†ЉеЉП
                    try:
                        # е¶ВжЮЬдЉ†ињЗжЭ•зЪДжШѓеЃМжХізЪДжЧґйЧіж†ЉеЉП
                        pickup_date = timezone.datetime.fromisoformat(pickup_date.replace('Z', '+00:00'))
                    except ValueError:
                        pickup_date = None
            is_pallet = False
            if cargo_id.startswith('plt_'):
                is_pallet = True
                ids = cargo_id.replace('plt_', '').split(',')
                model = Pallet
            else:
                ids = cargo_id.split(',')
                model = PackingList
            update_data = {}
            if follow_status or follow_status == '': update_data['ltl_follow_status'] = follow_status
            if carrier_company or carrier_company == '': update_data['carrier_company'] = carrier_company
            if address or address == '': update_data['address'] = address
            if bol_number or bol_number == '': update_data['ltl_bol_num'] = bol_number
            if (pickup_date is not None) and pickup_date or pickup_date != '': update_data['est_pickup_time'] = pickup_date
            if delivery_method is not None: update_data['delivery_method'] = delivery_method
            if note is not None: update_data['note'] = note
            if ltl_release_command is not None: update_data['ltl_release_command'] = ltl_release_command
            
            # иіҐеК°/е∞ЇеѓЄзЫЄеЕ≥е≠ЧжЃµ
            if pallet_size: update_data['ltl_plt_size_note'] = pallet_size
            
            if ltl_quote_raw and not del_qty_raw:
                del_qty_raw = len(ids)
            if ltl_quote_raw and not unit_quote_raw:
                unit_quote_raw = float(ltl_quote_raw)
            # е§ДзРЖжХ∞еАЉе≠ЧжЃµ
            if is_pallet:
                if del_qty_raw: update_data['del_qty'] = float(del_qty_raw)
                if unit_quote_raw: update_data['ltl_unit_quote'] = float(unit_quote_raw)
                if ltl_quote_raw: update_data['ltl_quote'] = float(ltl_quote_raw)
                if ltl_quote_note: update_data['ltl_quote_note'] = ltl_quote_note
                if ltl_quote_raw and not del_qty_raw:
                    update_data['del_qty'] = len(ids)
            # жЙІи°МжХ∞жНЃеЇУжЫіжЦ∞
            if update_data:
                await sync_to_async(model.objects.filter(id__in=ids).update)(**update_data)
                total_updated += 1

            # е¶ВжЮЬжШѓжЙШзЫШжХ∞жНЃдЄФжЬЙе∞ЇеѓЄпЉМи∞ГзФ®зЙєжЃКзЪДдњЭе≠ШжЦєж≥Х
            if cargo_id.startswith('plt_') and pallet_size:
                await self._save_pallet_sizes(ids, pallet_size)

            # еРМж≠•иі¶еНХйАїиЊС (е¶ВжЮЬељХеЕ•дЇЖжК•дїЈ)
            if ltl_quote_raw:
                msg = await self._delivery_account_selfpick_entry(ids, float(ltl_quote_raw), ltl_quote_note, unit_quote_raw, float(del_qty_raw), username)
                if msg: billing_messages.append(msg)
        
        success_msg = f"жИРеКЯжЫіжЦ∞ {total_updated} жЭ°иЃ∞ељХгАВ"
        if billing_messages:
            success_msg += "<br>" + "<br>".join(set(billing_messages))
        context = {'success_messages': mark_safe(success_msg)}

        page = request.POST.get('page')
        if page == "history":
            return await self.handle_ltl_history_pos_post(request, context)
        return await self.handle_ltl_unscheduled_pos_post(request, context)

    async def _save_pallet_sizes(self, plt_ids: List[str], pallet_size: str) -> Tuple[bool, str]:
        """
        еНХзЛђдњЭе≠ШжЙШзЫШе∞ЇеѓЄ
        ж†ЉеЉПпЉЪйХњ*еЃљ*йЂШ*жХ∞йЗПжЭњ дїґжХ∞дїґ йЗНйЗПkgпЉИжНҐи°МеИЖйЪФдЄНеРМе∞ЇеѓЄпЉЙ
        з§ЇдЊЛпЉЪ32*35*60*3жЭњ\n30*30*50*2жЭњ
        """
        # 1. иОЈеПЦжЙАжЬЙжЙШзЫШ
        pallets = await sync_to_async(list)(Pallet.objects.filter(id__in=plt_ids))
        total_pallets = len(pallets)
        
        # 2. иІ£жЮРжЙШзЫШе∞ЇеѓЄ
        lines = [line.strip() for line in pallet_size.split('\n') if line.strip()]
        if not lines:
            return True, ""  # з©Їе∞ЇеѓЄпЉМзЫіжО•ињФеЫЮжИРеКЯ
        
        #е¶ВжЮЬе∞±зїЩдЇЖдЄАзїДе∞ЇеѓЄпЉМжЯ•еИ∞зЪДжЭње≠РйГљжМЙињЩдЄ™иµЛеАЉ
        if len(lines) == 1:
            line = lines[0]
            # ж†ЉеЉПпЉЪйХњ*еЃљ*йЂШ xдїґ xkg
            if 'дїґ' in line and 'kg' in line:
                parts = line.split()
                clean_parts = [part for part in parts if part != '']
                for part in clean_parts:
                    if 'дїґ' in part:
                        pcs = part.replace('дїґ', '')
                    if 'kg' in part:
                        weight = part.replace('kg', '')
                    else:
                        parts = line.split(' ')[0].replace('жЭњ', '').split('*')
            else:
                parts = line.split('*')
                pcs = None
                weight = None
            if len(parts) == 3:
                try:
                    length = float(parts[0]) if parts[0] else None
                    width = float(parts[1]) if parts[1] else None
                    height = float(parts[2]) if parts[2] else None
                    
                    # жЙАжЬЙpalletйГљжМЙињЩдЄ™е∞ЇеѓЄиµЛеАЉ
                    for pallet in pallets:
                        pallet.length = length
                        pallet.width = width
                        pallet.height = height
                        if pcs:
                            pallet.pcs = pcs
                        if weight:
                            pallet.weight_lbs = float(weight) * 2.20462
                        await sync_to_async(pallet.save)()
                    
                    return True, ""
                except ValueError:
                    return False, f"жХ∞еАЉйФЩиѓѓпЉЪ'{line}'дЄ≠зЪДйХњеЃљйЂШењЕй°їжШѓжХ∞е≠Ч"
        
        # зїЩдЇЖе§ЪзїДе∞ЇеѓЄпЉМиІ£жЮРе∞ЇеѓЄжХ∞жНЃ
        size_assignments = []
        total_specified = 0
        
        valid_lines = [line for line in lines if line.strip()]
        has_piece_kg_line = any(('дїґ' in line and 'kg' in line) for line in valid_lines)
        if has_piece_kg_line:
            for idx, line in enumerate(valid_lines, start=1):
                if not ('дїґ' in line and 'kg' in line):
                    raise ValueError(f"зђђ {idx} и°МзЉЇе∞С дїґ жИЦ kg")
                
        for line in lines:
            if not line:
                continue
            
            if 'дїґ' in line and 'kg' in line:
                parts = line.split()
                clean_parts = [part for part in parts if part != '']
                for part in clean_parts:
                    if 'дїґ' in part:
                        pcs = part.replace('дїґ', '')
                    if 'kg' in part:
                        weight = part.replace('kg', '')
                    else:
                        parts = line.split(' ')[0].replace('жЭњ', '').split('*')
            else:
                line_clean = line.replace('жЭњ', '')
                parts = line_clean.split('*')
                pcs = None
                weight = None
            
            if len(parts) == 3:
                # ж†ЉеЉПпЉЪйХњ*еЃљ*йЂШпЉИйїШиЃ§1жЭњпЉЙ
                length, width, height = parts
                count = 1
            elif len(parts) == 4:
                # ж†ЉеЉПпЉЪйХњ*еЃљ*йЂШ*жЭњжХ∞
                length, width, height, count_str = parts
                count = int(count_str) if count_str.isdigit() else 1
            else:
                return False, f"жЙШзЫШе∞ЇеѓЄзЪДж†ЉеЉПйФЩиѓѓпЉЪ'{line}'"
            
            # иљђжНҐдЄЇжХ∞еАЉ
            try:
                length_val = float(length) if length else None
                width_val = float(width) if width else None
                height_val = float(height) if height else None
                pcs_val = float(pcs) if pcs else None
                weight_val = float(weight) if weight else None
            except ValueError:
                return False, f"жЙШзЫШе∞ЇеѓЄжХ∞еАЉйФЩиѓѓпЉЪ'{line}'дЄ≠зЪДйХњеЃљйЂШењЕй°їжШѓжХ∞е≠Ч,{length_val},{width_val},{height_val},{pcs_val},{weight_val}"
            
            size_assignments.append({
                'length': length_val,
                'width': width_val,
                'height': height_val,
                'count': count,
                'pcs': pcs_val,
                'weight': weight_val,
            })
            total_specified += count
        
        # 3. й™МиѓБжАїжХ∞
        if total_specified != total_pallets:
            return False, f"жЙШзЫШе∞ЇеѓЄиµЛеАЉжЧґпЉМжЭњжХ∞дЄНеМєйЕНпЉЪе∞ЇеѓЄзїЩеЗЇ{total_specified}жЭњпЉМеЃЮйЩЕз≥їзїЯжЬЙ{total_pallets}жЭњ"
        
        # 4. еИЖйЕНе∞ЇеѓЄ
        idx = 0
        for size_info in size_assignments:
            for _ in range(size_info['count']):
                if idx >= total_pallets:
                    break
                    
                pallet = pallets[idx]
                pallet.length = size_info['length']
                pallet.width = size_info['width']
                pallet.height = size_info['height']
                if pcs:
                    pallet.pcs = size_info['pcs']
                if weight:
                    pallet.weight_lbs = size_info['weight'] * 2.20462
                await sync_to_async(pallet.save)()
                idx += 1
        
        # 5. й™МиѓБжШѓеР¶жЙАжЬЙжЙШзЫШйГљеЈ≤е§ДзРЖ
        if idx != total_pallets:
            return False, f"жЙШзЫШе∞ЇеѓЄеИЖйЕНйФЩиѓѓпЉЪеП™еИЖйЕНдЇЖ{idx}дЄ™жЙШзЫШ"
        
        return True, ""
    
    async def handle_update_pod_status(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        """LTLеѓєpoжЫіжФєж†ЄеЃЮзКґжАБ"""
        pod_to_customer_str = request.POST.get("pod_to_customer")
        new_status = True if pod_to_customer_str == "True" else False
        
        target_ids = []
        # еНХи°МдњЃжФє
        single_id = request.POST.get('shipment_batch_number')
        if single_id:
            target_ids.append(single_id)
        # е§Ъи°МдњЃжФє
        batch_ids_json = request.POST.get('batch_ids_json')
        if batch_ids_json:
            try:
                ids_list = json.loads(batch_ids_json)
                if isinstance(ids_list, list):
                    target_ids.extend(ids_list)
            except json.JSONDecodeError:
                pass
        
        target_ids = list(set(target_ids))

        if target_ids:
            await sync_to_async(Shipment.objects.filter(shipment_batch_number__in=target_ids).update)(
                pod_to_customer=new_status
            )
        succes_len = len(target_ids)
        context = {'success_messages': f'жИРеКЯжЫіжЦ∞{succes_len}жЭ°PODеЫЮдЉ†зКґжАБпЉБ'}
        return await self.handle_ltl_unscheduled_pos_post(request,context)
    
    async def handle_save_shipping_tracking(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        """LTLеѓєpoжЫіжФєж†ЄеЃЮзКґжАБ"""
        fleet_number = request.POST.get('fleet_number')
        ltl_shipping_tracking = request.POST.get('ltl_shipping_tracking')
        
        if fleet_number and ltl_shipping_tracking:
            fleet = await sync_to_async(Fleet.objects.get)(fleet_number=fleet_number)
            fleet.ltl_shipping_tracking = ltl_shipping_tracking
            await sync_to_async(fleet.save)()
        
        return await self.handle_ltl_unscheduled_pos_post(request)
    
    async def handle_save_fleet_cost(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        """LTLеѓєpoжЫіжФєж†ЄеЃЮзКґжАБ"""
        fleet_number = request.POST.get('fleet_number')
        fleet_cost = request.POST.get('fleet_cost')
        
        if fleet_number and fleet_cost:
            fleet = await sync_to_async(Fleet.objects.get)(fleet_number=fleet_number)
            fleet.fleet_cost = float(fleet_cost)
            await sync_to_async(fleet.save)()
            #еИЖжСКжИРжЬђ
            fm = FleetManagement()
            await fm.insert_fleet_shipment_pallet_fleet_cost(
                request, fleet_number, fleet_cost
            )
        
        return await self.handle_ltl_unscheduled_pos_post(request)

    async def handle_verify_ltl_cargo(
        self, request: HttpRequest, context: dict| None = None,
    ) -> tuple[str, dict[str, Any]]:
        """LTLеѓєpoжЫіжФєж†ЄеЃЮзКґжАБ"""
        if not context:
            context = {}
        cargo_ids = request.POST.get('cargo_ids', '')
        ltl_verify = request.POST.get('ltl_verify', 'false').lower() == 'true'
        
        # е§ДзРЖ PackingList зЪДж†ЄеЃЮ
        if cargo_ids:
            cargo_id_list = [int(id.strip()) for id in cargo_ids.split(',') if id.strip()]
            packinglist_ids = cargo_id_list
            if packinglist_ids:
                # жЫіжЦ∞ PackingList зЪДж†ЄеЃЮзКґжАБ
                await sync_to_async(PackingList.objects.filter(
                    id__in=packinglist_ids
                ).update)(
                    ltl_verify=ltl_verify
                )
        return await self.handle_ltl_unscheduled_pos_post(request)
        
    async def handle_ltl_history_pos_post(
        self, request: HttpRequest, context: dict| None = None,
    ) -> tuple[str, dict[str, Any]]:
        '''LTLзїДзЪДеОЖеП≤жЄѓеРОжХ∞жНЃ'''
        warehouse = request.POST.get("warehouse")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        if not context:
            context = {}
        if warehouse:
            warehouse_name = warehouse.split('-')[0]
        else:
            context.update({
                'error_messages':"ж≤°йАЙдїУеЇУпЉБ",
                'warehouse_options': self.warehouse_options,
            })
            return self.template_ltl_history_pos, context
        
        # жЬ™зїЩеЃЪжЧґйЧіжЧґпЉМиЗ™еК®жЯ•иѓҐињЗеОїдЄЙдЄ™жЬИзЪД
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-30)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date
        
        pl_criteria = Q(
            container_number__orders__offload_id__offload_at__isnull=True,
            container_number__orders__retrieval_id__retrieval_destination_area=warehouse_name,
            container_number__orders__retrieval_id__actual_retrieval_timestamp__gte=start_date,
            container_number__orders__retrieval_id__actual_retrieval_timestamp__lte=end_date,
            delivery_type="other"
        )
        plt_criteria = Q(
            location=warehouse,            
            container_number__orders__offload_id__offload_at__gte=start_date,
            container_number__orders__offload_id__offload_at__lte=end_date, 
            delivery_type="other"
        )

        # еЈ≤жФЊи°М-еЃҐжПР
        _, selfpick_cargos, selfdel_cargos = await self._get_classified_cargos(pl_criteria, plt_criteria)
        #selfpick_cargos = await self._ltl_scheduled_self_pickup(pl_criteria, plt_criteria)
        # еЈ≤жФЊи°М-иЗ™еПС
        #selfdel_cargos = await self._ltl_self_delivery(pl_criteria, plt_criteria)
        # еОЖеП≤иљ¶жђ°
        #fleet_cargos = await self._ltl_unscheduled_data(request, warehouse, start_date, end_date)
        fleet_cargos = None
        summary = {
            'selfpick_count': len(selfpick_cargos),
            'selfdel_count': len(selfdel_cargos),
            'fleet_count': 0,
        }
        if not context:
            context = {}
        context.update({
            'warehouse': warehouse,
            'warehouse_options': self.warehouse_options,
            'account_options': self.arm_account_options,
            'load_type_options': LOAD_TYPE_OPTIONS,
            "selfpick_cargos": selfpick_cargos,
            "selfdel_cargos": selfdel_cargos,
            "fleet_cargos": fleet_cargos,
            "summary": summary,
            'shipment_type_options': self.shipment_type_options,
            "carrier_options": self.carrier_options,
            "abnormal_fleet_options": self.abnormal_fleet_options,
            "warehouse_name": warehouse_name,
            "start_date": start_date,
            "end_date": end_date,
        })
        active_tab = request.POST.get('active_tab')
        if active_tab:
            context.update({'active_tab':active_tab})
        return self.template_ltl_history_pos, context
    
    async def _get_classified_cargos(self, pl_criteria, plt_criteria):
        """дЄАжђ°жАІиОЈеПЦеєґеИЖз±їLTLзЪДжЙАжЬЙPO"""
        # иОЈеПЦеЕ®йЗПжХ∞жНЃ (ж≥®жДПпЉЪж≠§жЧґдЄНи¶БеЬ® criteria йЗМеК†вАЬиЗ™жПРвАЭйЩРеИґпЉМиОЈеПЦиѓ•дїУеЇУдЄЛзЪДжЙАжЬЙ)
        all_raw_data = await self._ltl_packing_list(pl_criteria, plt_criteria)
        
        release_cargos = []     # жЬ™жФЊи°М (Tab 1)
        selfpick_cargos = []    # еЈ≤жФЊи°М-еЃҐжПР (Tab 2)
        selfdel_cargos = []     # еЈ≤жФЊи°М-иЗ™еПС (Tab 3)

        for item in all_raw_data:
            is_pass = item.get('is_pass', False)
            # еЕЉеЃє Pallet еТМ PackingList зЪД delivery_method е≠ЧжЃµеРН
            delivery_method = item.get('delivery_method') or item.get('custom_delivery_method') or ""
            
            if not is_pass:
                # жЬ™жФЊи°МйАїиЊС
                release_cargos.append(item)
            else:
                # еЈ≤жФЊи°МйАїиЊС
                if "иЗ™жПР" in delivery_method:
                    selfpick_cargos.append(item)
                else:
                    selfdel_cargos.append(item)

        # йТИеѓєжЬ™жФЊи°МињЫи°МжОТеЇПпЉЪжЬ™ж†ЄеЃЮжОТеЬ®еЙНйЭҐпЉМзДґеРОжМЙ vessel_eta жОТеЇПпЉМжЧґйЧіжЧ©зЪДеЬ®еЙН
        release_cargos.sort(key=lambda x: (
            x.get('ltl_verify', False),  # зђђдЄАдЉШеЕИзЇІпЉЪжЬ™ж†ЄеЃЮпЉИFalseпЉЙеЬ®еЙН
            str(x.get('vessel_eta')) or '9999-12-31'  # зђђдЇМдЉШеЕИзЇІпЉЪvessel_etaпЉМжЧґйЧіжЧ©зЪДеЬ®еЙН
        ))
        
        return release_cargos, selfpick_cargos, selfdel_cargos

    async def handle_ltl_cancel_shipment_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        '''еИ†йЩ§LTLйҐДзЇ¶жЙєжђ°'''
        s_id = request.POST.get("s_id")
        
        if not s_id:
            context = {"error_messages": "жЬ™жПРдЊЫйҐДзЇ¶жЙєжђ°ID"}
            return await self.handle_ltl_unscheduled_pos_post(request, context)
        
        try:
            # жЯ•жЙЊеєґеИ†йЩ§ShipmentиЃ∞ељХ
            shipment = await sync_to_async(Shipment.objects.get)(id=s_id)
            await sync_to_async(shipment.delete)()
            
            # еИ†йЩ§жИРеКЯеРОи∞ГзФ®handle_ltl_unscheduled_pos_postеИЈжЦ∞й°µйЭҐ
            context = {"success_messages": f"йҐДзЇ¶жЙєжђ° {shipment.shipment_batch_number} еИ†йЩ§жИРеКЯпЉБ"}
            return await self.handle_ltl_unscheduled_pos_post(request, context)
            
        except Shipment.DoesNotExist:
            context = {"error_messages": "жЬ™жЙЊеИ∞еѓєеЇФзЪДйҐДзЇ¶жЙєжђ°иЃ∞ељХ"}
            return await self.handle_ltl_unscheduled_pos_post(request, context)
        except Exception as e:
            context = {"error_messages": f"еИ†йЩ§йҐДзЇ¶жЙєжђ°е§±иі•: {str(e)}"}
            return await self.handle_ltl_unscheduled_pos_post(request, context)

    async def handle_shipping_order_upload(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        '''еЫЮдЉ†еЗЇеЇУеНХпЉИеНХдЄ™жИЦжЙєйЗПпЉЙ'''
        fm = FleetManagement()
        conn = await fm._get_sharepoint_auth()
        step = request.POST.get("step")
        
        if step == "batch_shipping_order_upload":
            # жЙєйЗПдЄКдЉ† - жЦ∞ж†ЉеЉП
            index = 0
            while True:
                file_key = f"file_{index}"
                batch_key = f"batch_{index}"
                
                if file_key not in request.FILES or batch_key not in request.POST:
                    break
                
                file = request.FILES[file_key]
                shipment_batch_number = request.POST[batch_key]
                await fm._upload_shipping_order_file_to_sharepoint(conn, shipment_batch_number, file)
                index += 1
        else:
            # еНХдЄ™дЄКдЉ†
            shipment_batch_number = request.POST.get("shipment_batch_number")
            if "file" in request.FILES and shipment_batch_number:
                file = request.FILES["file"]
                await fm._upload_shipping_order_file_to_sharepoint(conn, shipment_batch_number, file)
            
        template, context = await self.handle_ltl_unscheduled_pos_post(request)
        context.update({"success_messages": 'еЗЇеЇУеНХдЄКдЉ†жИРеКЯ!'})           
        return template, context

    async def handle_ltl_unscheduled_pos_post(
        self, request: HttpRequest, context: dict| None = None,
    ) -> tuple[str, dict[str, Any]]:
        '''LTLзїДзЪДжЄѓеРОеЕ®жµБз®Л'''
        warehouse = request.POST.get("warehouse")
        if not context:
            context = {}
        if warehouse:
            warehouse_name = warehouse.split('-')[0]
        else:
            context.update({'error_messages':"ж≤°йАЙдїУеЇУпЉБ"})
            return self.template_unscheduled_pos_all, context
        
        pl_criteria = Q(
            container_number__orders__offload_id__offload_at__isnull=True,
            shipment_batch_number__shipment_batch_number__isnull=True,
            container_number__orders__retrieval_id__retrieval_destination_area=warehouse_name,
            delivery_type="other"
        )
        plt_criteria = Q(
            location=warehouse,
            shipment_batch_number__shipment_batch_number__isnull=True,
            container_number__orders__offload_id__offload_at__gt=datetime(2025, 12, 1),
            delivery_type="other"
        )
        # жЬ™жФЊи°МгАБеЈ≤жФЊи°М-еЃҐжПРгАБеЈ≤жФЊи°М-иЗ™еПС
        release_cargos, selfpick_cargos, selfdel_cargos = await self._get_classified_cargos(pl_criteria, plt_criteria)
        #release_cargos = await self._ltl_unscheduled_cargo(pl_criteria, plt_criteria)

        # еЈ≤жФЊи°М-еЃҐжПР
        #selfpick_cargos = await self._ltl_scheduled_self_pickup(pl_criteria, plt_criteria)
        # еЈ≤жФЊи°М-иЗ™еПС
        #selfdel_cargos = await self._ltl_self_delivery(pl_criteria, plt_criteria)

        #жЬ™жОТиљ¶
        unschedule_fleet = await self._ltl_unscheduled_data(request, warehouse)
        #еЊЕеЗЇеЇУ
        ready_to_ship_data = await self._ltl_ready_to_ship_data(warehouse,request.user)
        # еЊЕйАБиЊЊ
        delivery_data_raw = await self._fl_delivery_get(warehouse, None, 'ltl')
        delivery_data = delivery_data_raw['shipments']
        # #еЊЕдЉ†POD
        pod_data = await self._ltl_pod_get(warehouse)
        # #еЊЕдЉ†еЗЇеЇУеНХ
        shipping_data = await self._ltl_shipping_get(warehouse)

        pod_data = sorted(
            pod_data,
            key=lambda p: p.pod_to_customer is True
        )
        summary = {
            'release_count': len(release_cargos),
            'selfpick_count': len(selfpick_cargos),
            'selfdel_count': len(selfdel_cargos),
            'ready_to_ship_count': len(ready_to_ship_data),
            'shipping_count': len(shipping_data),
            'ready_count': len(delivery_data),
            'pod_count': len(pod_data),
            'unfleet_count': len(unschedule_fleet),
        }
        if not context:
            context = {}
        context.update({
            'warehouse': warehouse,
            'warehouse_options': self.warehouse_options,
            'account_options': self.arm_account_options,
            'load_type_options': LOAD_TYPE_OPTIONS,
            "release_cargos": release_cargos,
            "selfpick_cargos": selfpick_cargos,
            "selfdel_cargos": selfdel_cargos,
            "ready_to_ship_data": ready_to_ship_data,
            "shipping_data": shipping_data,
            "delivery_data": delivery_data,
            "pod_data": pod_data,
            "summary": summary,
            'shipment_type_options': self.shipment_type_options,
            "carrier_options": self.carrier_options,
            "abnormal_fleet_options": self.abnormal_fleet_options,
            "warehouse_name": warehouse_name,
            'unschedule_fleet': unschedule_fleet,
        })
        active_tab = request.POST.get('active_tab')
        if active_tab:
            context.update({'active_tab':active_tab})
        return self.template_ltl_pos_all, context
    
    async def _ltl_unscheduled_data(
        self, request: HttpRequest, warehouse:str, start_date: str | None = None, end_date: str | None = None
    ) -> tuple[str, dict[str, Any]]:
        target_date = datetime(2025, 10, 10)
        base_q = models.Q(
            origin=warehouse,
            fleet_number__isnull=True,
            in_use=True,
            is_canceled=False,
            shipment_type__in=['LTL', 'еЃҐжИЈиЗ™жПР'],
            is_virtual_sp=False,
        )

        if start_date and end_date:
            start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
            end_datetime = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            time_filter_q = models.Q(
                shipment_appointment__range=(start_datetime, end_datetime)
            )
            
        else:
            time_filter_q = models.Q(shipment_appointment__gt=target_date)
        base_q = base_q & time_filter_q
        shipment_list = await sync_to_async(list)(
            Shipment.objects.filter(base_q).select_related("fleet_number").order_by("pickup_time", "shipment_appointment")
        )
        for shipment in shipment_list:
            #is_public = self._check_destination_delivery_type(shipment.destination)
            # if is_public:
            #     continue
            shipment.fleet_display_name = None
            if shipment.fleet_number:
                try:
                    shipment.fleet_display_name = shipment.fleet_number.fleet_number
                except ObjectDoesNotExist:
                    shipment.fleet_display_name = None
            # дїОpackinglistи°®иОЈеПЦеФЫе§і
            packinglist_marks = await sync_to_async(list)(
                PackingList.objects.filter(
                    shipment_batch_number=shipment.id
                ).values_list('shipping_mark', flat=True).distinct()
            )
            
            # дїОpalletи°®иОЈеПЦеФЫе§і
            pallet_marks = await sync_to_async(list)(
                Pallet.objects.filter(
                    shipment_batch_number=shipment.id
                ).values_list('shipping_mark', flat=True).distinct()
            )
            
            # еРИеєґеєґеОїйЗНеФЫе§і
            all_marks = set(packinglist_marks + pallet_marks)
            all_marks = [mark for mark in all_marks if mark]  # ињЗжї§з©ЇеАЉ
            
            # жЈїеК†еФЫе§іе≠ЧжЃµеИ∞shipmentеѓєи±°
            shipment.shipping_marks = all_marks  # еИЧи°®ељҐеЉП
            shipment.shipping_marks_display = "пЉМ".join(all_marks) if all_marks else "жЧ†еФЫе§і"  # жШЊз§ЇзФ®
            if shipment.pod_link:
                shipment.status_display = "еЈ≤дЄКдЉ†POD"
                shipment.status_class = "status-pod"  # зїњиЙ≤
            elif shipment.is_arrived:
                shipment.status_display = "еЈ≤йАБиЊЊ"
                shipment.status_class = "status-arrived"  # иУЭиЙ≤
            elif shipment.is_shipped:
                shipment.status_display = "еЈ≤еЗЇеЇУ"
                shipment.status_class = "status-shipped"  # йїДиЙ≤
            else:
                shipment.status_display = "еЊЕе§ДзРЖ"
                shipment.status_class = "status-pending"  # зБ∞иЙ≤
        return shipment_list
    
    async def _check_destination_delivery_type(self, destination: str) -> bool:
        """ж£АжЯ•зЫЃзЪДеЬ∞еѓєеЇФзЪДdelivery_typeжШѓеР¶еМЕеРЂиЗ™жПРжИЦиЗ™еПС"""
        
        if not destination:
            return False
        # 1. зїЯдЄАиљђдЄЇе§ІеЖЩеєґеОїйЩ§дЄ§зЂѓз©Їж†Љ
        dest_upper = str(destination).strip().upper()
        
        # 2. зІїйЩ§дЄ≠йЧіеПѓиГљзЪДз©Їж†ЉињЫи°Мж≠£еИЩж†°й™М (дЊЛе¶Ве∞Ж "ONT 8" иІЖдЄЇ "ONT8")
        dest_compact = dest_upper.replace(" ", "")

        # 3. ж≠£еИЩи°®иЊЊеЉПж†°й™М (еМєйЕНж†ЗеЗЖдїУеЇУдї£з†Б)
        if self.RE_PUBLIC_WH.match(dest_compact):
            return True

        # 4. еЕ≥йФЃе≠Чж†°й™М (еМєйЕНеМЕеРЂеЕ≥йФЃиѓНзЪДжГЕеЖµ)
        if any(kw in dest_upper for kw in self.PUBLIC_KEYWORDS):
            return True

        # 5. зЙєжЃКиІДеИЩпЉЪе¶ВжЮЬзЫЃзЪДеЬ∞еЕ®жШѓжХ∞е≠ЧпЉИжЯРдЇЫдїУзВєзЪДеЖЕйГ®зЉЦеПЈпЉЙ
        if dest_compact.isdigit() and len(dest_compact) >= 4:
            return True

        return False

    async def _ltl_pod_get(
        self, warehouse:str,
    ) -> dict[str, Any]: 

        criteria = models.Q(
            models.Q(pod_link__isnull=True) | models.Q(pod_link=""),
            pod_to_customer=False,       
            shipped_at__isnull=False,
            arrived_at__isnull=False,
            shipment_schduled_at__gte="2024-12-01",
            origin=warehouse,
        )
        criteria = criteria & models.Q(shipment_type__in=['LTL', 'еЃҐжИЈиЗ™жПР'])

        shipments = await sync_to_async(list)(
            Shipment.objects.select_related("fleet_number")
            .filter(criteria)
            .order_by("shipped_at")
        )
        for shipment in shipments:
            # иОЈеПЦдЄОиѓ•shipmentеЕ≥иБФзЪДжЙАжЬЙpallet
            pallets = await sync_to_async(list)(
                Pallet.objects.filter(shipment_batch_number=shipment)
                .select_related('container_number')
            )
             
            customer_names = set()
            details_set = set()
            for pallet in pallets:
                if pallet.container_number:
                    # иОЈеПЦдЄОиѓ•containerеЕ≥иБФзЪДжЙАжЬЙorder
                    orders = await sync_to_async(list)(
                        Order.objects.filter(container_number=pallet.container_number)
                        .select_related('customer_name')
                    )
                    
                    for order in orders:
                        if order.customer_name:
                            customer_names.add(order.customer_name.zem_name)
                # жЛЉжО• details дњ°жБѓ
                container_num = pallet.container_number.container_number if pallet.container_number else "жЧ†жЯЬеПЈ"
                destination = getattr(pallet, "destination", "") or ""
                shipping_mark = getattr(pallet, "shipping_mark", "") or ""
                detail_str = (
                    f"<span style='color:blue;'>{container_num}</span>"
                    f"<span style='color:red;'>-</span>"
                    f"<span style='color:green;'>{destination}</span>"
                    f"<span style='color:red;'>-</span>"
                    f"<span style='color:orange;'>{shipping_mark}</span>"
                )
                details_set.add(detail_str)

            
            # е∞ЖеЃҐжИЈеРНзФ®йАЧеПЈжЛЉжО•пЉМеєґжЈїеК†еИ∞shipmentеѓєи±°дЄК
            shipment.customer = ", ".join(customer_names) if customer_names else "жЧ†еЃҐжИЈдњ°жБѓ"
            shipment.details = "<br>".join(details_set) if details_set else None
        
        return shipments
    
    async def _ltl_shipping_get(
        self, warehouse:str,
    ) -> dict[str, Any]: 

        criteria = models.Q(
            models.Q(shipping_order_link__isnull=True) | models.Q(shipping_order_link=""),
            shipment_schduled_at__gte="2026-4-17",
            origin=warehouse,
            shipped_at__isnull=False,
        )
        criteria = criteria & models.Q(shipment_type__in=['LTL', 'еЃҐжИЈиЗ™жПР'])

        shipments = await sync_to_async(list)(
            Shipment.objects.select_related("fleet_number")
            .filter(criteria)
            .order_by("shipped_at")
        )
        for shipment in shipments:
            pallets = await sync_to_async(list)(
                Pallet.objects.filter(shipment_batch_number=shipment)
                .select_related('container_number')
            )
             
            customer_names = set()
            details_set = set()
            total_pallets = 0
            total_pcs = 0
            for pallet in pallets:
                total_pallets += 1
                if pallet.pcs:
                    total_pcs += pallet.pcs
                if pallet.container_number:
                    orders = await sync_to_async(list)(
                        Order.objects.filter(container_number=pallet.container_number)
                        .select_related('customer_name')
                    )
                    
                    for order in orders:
                        if order.customer_name:
                            customer_names.add(order.customer_name.zem_name)
                container_num = pallet.container_number.container_number if pallet.container_number else "жЧ†жЯЬеПЈ"
                destination = getattr(pallet, "destination", "") or ""
                shipping_mark = getattr(pallet, "shipping_mark", "") or ""
                detail_str = (
                    f"<span style='color:blue;'>{container_num}</span>"
                    f"<span style='color:red;'>-</span>"
                    f"<span style='color:green;'>{destination}</span>"
                    f"<span style='color:red;'>-</span>"
                    f"<span style='color:orange;'>{shipping_mark}</span>"
                )
                details_set.add(detail_str)

            
            shipment.customer = ", ".join(customer_names) if customer_names else "жЧ†еЃҐжИЈдњ°жБѓ"
            shipment.details = "<br>".join(details_set) if details_set else None
            shipment.shipped_pallet = total_pallets
            shipment.shipped_pcs = total_pcs
        
        return shipments
    
    async def _ltl_ready_to_ship_data(self, warehouse: str, user:User) -> list:
        """иОЈеПЦеЊЕеЗЇеЇУжХ∞жНЃ - жМЙfleet_numberеИЖзїД"""
        # иОЈеПЦжМЗеЃЪдїУеЇУзЪДжЬ™еЗЇеПСдЄФжЬ™еПЦжґИзЪДfleet
        base_bq = models.Q(
            origin=warehouse,
            departured_at__isnull=True,
            is_canceled=False,
            fleet_type__in=['LTL', 'еЃҐжИЈиЗ™жПР']
        )
        
        fleets = await sync_to_async(list)(
            Fleet.objects.filter(base_bq).prefetch_related(
                Prefetch(
                    'shipment',
                    queryset=Shipment.objects.prefetch_related(
                        Prefetch(
                            'packinglist',
                            queryset=PackingList.objects.select_related('container_number')
                        ),
                        Prefetch(
                            'pallet', 
                            queryset=Pallet.objects.select_related('packing_list', 'container_number')
                        )
                    )
                )
            )
        )
        
        grouped_data = []
        
        for fleet in fleets:
            arm_pickup = await self._get_fleet_arm_pickup(fleet)
            fleet_group = {
                'fleet_number': fleet.fleet_number,
                'fleet_type': fleet.fleet_type,
                'fleet_cost': fleet.fleet_cost or 0,
                'third_party_address': fleet.third_party_address,
                'pickup_number': fleet.pickup_number,
                'motor_carrier_number': fleet.motor_carrier_number,
                'license_plate': fleet.license_plate,
                'dot_number': fleet.dot_number,
                'appointment_datetime': fleet.appointment_datetime,
                'carrier': fleet.carrier,
                'is_virtual': fleet.is_virtual,
                'shipments': {},  # жФєеЫЮе≠ЧеЕЄзїУжЮДпЉМдњЭжМБдЄОеЙНзЂѓеЕЉеЃє
                'pl_ids': [],
                'plt_ids': [],
                'total_cargos': 0,  # жАїиіІзЙ©и°МжХ∞
                'arm_pickup':arm_pickup,
            }
            
            shipments = await sync_to_async(list)(
                Shipment.objects.filter(fleet_number__fleet_number=fleet.fleet_number)
            )
            
            for shipment in shipments:
                if not shipment.shipment_batch_number:
                    continue

                batch_number = shipment.shipment_batch_number
                
                # еИЭеІЛеМЦshipmentжХ∞жНЃ
                if batch_number not in fleet_group['shipments']:
                    fleet_group['shipments'][batch_number] = {
                        'shipment_batch_number': shipment.shipment_batch_number or '-',
                        'appointment_id': shipment.appointment_id or '-',
                        'destination': shipment.destination or '-',
                        'shipment_appointment': shipment.shipment_appointment,
                        'cargos': []
                    }
                
                # е§ДзРЖpackinglists
                raw_data = await self._get_packing_list(
                    user,
                    models.Q(
                        shipment_batch_number__shipment_batch_number=batch_number,
                        container_number__orders__offload_id__offload_at__isnull=True,
                    ),
                    models.Q(
                        shipment_batch_number__shipment_batch_number=batch_number,
                        container_number__orders__offload_id__offload_at__isnull=False,
                    ),
                )
                fleet_group['shipments'][batch_number]['cargos'].extend(raw_data)
            
            # жОТеЇП shipmentsпЉМcargos дЄЇз©ЇзЪДжФЊеРОйЭҐ
            fleet_group['shipments'] = dict(
                sorted(
                    fleet_group['shipments'].items(),
                    key=lambda item: not item[1]['cargos']
                )
            )
            fleet_group['total_cargos'] = sum(
                len(s['cargos']) if s['cargos'] else 1
                for s in fleet_group['shipments'].values()
            )
            # еП™жЬЙжЬЙжХ∞жНЃзЪДfleetжЙНињФеЫЮ
            #if fleet_group['shipments']:
            grouped_data.append(fleet_group)
        # жМЙ appointment_datetime жОТеЇПпЉМжЧґйЧіжЧ©зЪДжОТеЬ®еЙНйЭҐ
        grouped_data.sort(
            key=lambda x: (
                x['appointment_datetime'].replace(tzinfo=None)
                if x['appointment_datetime'] else datetime.max
            )
        )
        return grouped_data
    
    async def _get_fleet_arm_pickup(self, fleet:Fleet):
        arm_pickup = await sync_to_async(list)(
            Pallet.objects.select_related(
                "container_number__container_number",
                "shipment_batch_number__fleet_number",
            )
            .filter(
                shipment_batch_number__fleet_number=fleet
            )
            .values(
                "container_number__container_number",
                "zipcode",
                "shipping_mark",
                "destination",
                "shipment_batch_number__ARM_BOL",
                "shipment_batch_number__ARM_PRO",
                "shipment_batch_number__fleet_number__carrier",
                "shipment_batch_number__fleet_number__appointment_datetime",
                "address",
                "slot",
                "shipment_batch_number__note",
            )
            .annotate(
                total_pcs=Sum("pcs"),
                total_pallet=Count("pallet_id", distinct=True),
                total_weight=Sum("weight_lbs"),
                total_cbm=Sum("cbm"),
            )
        )
        for arm in arm_pickup:
            marks = arm["shipping_mark"]
            if marks:
                array = marks.split(",")
                if len(array) > 2:
                    parts = []
                    for i in range(0, len(array), 2):
                        part = ",".join(array[i : i + 2])
                        parts.append(part)
                    new_marks = "\n".join(parts)
                else:
                    new_marks = marks
                arm["shipping_mark"] = new_marks
            else:
                arm["shipping_mark"] = ""
            arm["address_parts"] = {
                "company": "",
                "road": "",
                "city": "",
                "name": "",
                "phone": ""
            }
        arm_json = []
        for item in arm_pickup:
            arm_json.append({
                'container_number': item.get('container_number__container_number', ''),
                'zipcode': item.get('zipcode', ''),
                'shipping_mark': item.get('shipping_mark', ''),
                'destination': item.get('destination', ''),
                'address': item.get('address', ''),
                'slot': item.get('slot', ''),
                'arm_bol': item.get('shipment_batch_number__ARM_BOL', ''),
                'arm_pro': item.get('shipment_batch_number__ARM_PRO', ''),
                'carrier': item.get('shipment_batch_number__fleet_number__carrier', ''),
                'appointment_datetime': (
                    item.get('shipment_batch_number__fleet_number__appointment_datetime').isoformat()
                    if item.get('shipment_batch_number__fleet_number__appointment_datetime')
                    else ''
                ),
                'note': item.get('shipment_batch_number__note', ''),
                'total_pcs': int(item.get('total_pcs', 0)),
                'total_pallet': int(item.get('total_pallet', 0)),
                'total_weight': float(item.get('total_weight', 0)),
                'total_cbm': float(item.get('total_cbm', 0)),
                'address_parts': dict(item.get('address_parts') or {
                    'company': '', 'road': '', 'city': '', 'name': '', 'phone': ''
                })
            })

        arm_json_str = json.dumps(arm_json, cls=DjangoJSONEncoder)
        return arm_json_str
    
    async def handle_unscheduled_pos_post(
        self, request: HttpRequest, context: dict| None = None,
    ) -> tuple[str, dict[str, Any]]:
        #warehouse = request.POST.get("warehouse")
        warehouse = 'LA-91761'
        if not context:
            context = {}
        
        nowtime = timezone.now()
        two_weeks_later = nowtime + timezone.timedelta(weeks=2)   
        pl_criteria = (models.Q(
                shipment_batch_number__shipment_batch_number__isnull=True,
                container_number__orders__offload_id__offload_at__isnull=True,
                container_number__orders__vessel_id__vessel_eta__lte=two_weeks_later, 
                container_number__orders__retrieval_id__retrieval_destination_precise=warehouse,
                container_number__is_abnormal_state=False,
                destination__in=FOUR_MAJOR_WAREHOUSES
            )&
            ~(
                models.Q(delivery_method__icontains='жЪВжЙ£') |
                models.Q(delivery_method__icontains='иЗ™жПР') |
                models.Q(delivery_method__icontains='UPS') |
                models.Q(delivery_method__icontains='FEDEX')
            )
        )

        plt_criteria = (models.Q(
                location=warehouse,
                shipment_batch_number__shipment_batch_number__isnull=True,
                container_number__orders__offload_id__offload_at__gt=datetime(2025, 1, 1),
                destination__in=FOUR_MAJOR_WAREHOUSES
            )&
            ~(
                models.Q(delivery_method__icontains='жЪВжЙ£') |
                models.Q(delivery_method__icontains='иЗ™жПР') |
                models.Q(delivery_method__icontains='UPS') |
                models.Q(delivery_method__icontains='FEDEX')
            )
        )
        unshipment_pos = await self._get_packing_list(
            request.user,
            pl_criteria,
            plt_criteria,
        )
        if len(unshipment_pos) == 0:
            context.update({'error_messages':"ж≤°жЬЙжЯ•еИ∞зЫЄеЕ≥еЇУе≠ШпЉБ"})
            return self.template_unscheduled_pos_all, context
        #жЬ™дљњзФ®зЪДзЇ¶еТМеЉВеЄЄзЪДзЇ¶
        shipments = await self.get_shipments_by_warehouse(warehouse,request,"four_major_whs")
        #еЈ≤жОТзЇ¶
        scheduled_data = await self.sp_scheduled_data(warehouse, request.user, "four_major_whs")

        #жЬ™жОТиљ¶+еЈ≤жОТиљ¶
        fleets = await self._fl_unscheduled_data(request, warehouse, "four_major_whs")
        #еЈ≤жОТиљ¶
        schedule_fleet_data = fleets['fleet_list']

        #еЊЕеЗЇеЇУ
        ready_to_ship_data = await self._sp_ready_to_ship_data(warehouse,request.user, "four_major_whs")
        # еЊЕйАБиЊЊ
        delivery_data_raw = await self._fl_delivery_get(warehouse, "four_major_whs")
        delivery_data = delivery_data_raw['shipments']
        #еЊЕдЉ†POD
        pod_data_raw = await self._fl_pod_get(warehouse, "four_major_whs")
        pod_data = pod_data_raw['fleet']

        summary = await self._four_major_calculate_summary(unshipment_pos, shipments, scheduled_data, schedule_fleet_data, ready_to_ship_data, delivery_data, pod_data, warehouse)

        #еЫЫе§ІдїУзЪДдЄНзЬЛиИєеИЧи°®    
        destination_list = []
        for item in unshipment_pos:
            destination = item.get('destination')
            destination_list.append(destination)
            
        #if await self._validate_user_four_major_whs(request.user):
        vessel_dict = {}       
        destination_list = sorted(list(set(destination_list)))
        if not context:
            context = {}
        context.update({
            'warehouse': warehouse,
            'warehouse_options': self.warehouse_options,
            'cargos': unshipment_pos,
            'shipments': shipments,
            'summary': summary,
            'cargo_count': len(unshipment_pos),
            'appointment_count': len(shipments),
            "vessel_dict": vessel_dict,
            "destination_list": destination_list,
            'account_options': self.arm_account_options,
            'load_type_options': LOAD_TYPE_OPTIONS,
            "scheduled_data": scheduled_data,
            "schedule_fleet_data": schedule_fleet_data,
            "ready_to_ship_data": ready_to_ship_data,
            "delivery_shipments": delivery_data,
            "pod_shipments": pod_data,
            'shipment_type_options': self.shipment_type_options,
            "carrier_options": self.carrier_options,
            "abnormal_fleet_options": self.abnormal_fleet_options,
        })
        active_tab = request.POST.get('active_tab')
        
        if active_tab:
            context.update({'active_tab':active_tab})
        return self.template_unscheduled_pos_all, context


    async def handle_appointment_management_post(
        self, request: HttpRequest, context: dict| None = None,
    ) -> tuple[str, dict[str, Any]]:
        '''дЇЪй©ђйАКе§ЗзЇ¶зЃ°зРЖ'''
        warehouse = request.POST.get("warehouse")
        if warehouse:
            warehouse_name = warehouse.split('-')[0]
        else:
            raise ValueError('жЬ™йАЙжЛ©дїУеЇУпЉБ')
        
        nowtime = timezone.now()
        two_weeks_later = nowtime + timezone.timedelta(weeks=2)
        three_months_ago = nowtime - timezone.timedelta(days=90)
        # 1гАБPOзЃ°зРЖвАФвАФжЙАжЬЙж≤°зЇ¶дЄФдЄ§еС®еЖЕеИ∞жЄѓзЪДиіІзЙ©
        unshipment_pos = await self._get_packing_list(
            request.user,
            models.Q(
                shipment_batch_number__shipment_batch_number__isnull=True,
                container_number__orders__offload_id__offload_at__isnull=True,
                container_number__orders__vessel_id__vessel_eta__gte=three_months_ago,
                container_number__orders__vessel_id__vessel_eta__lte=two_weeks_later, 
                container_number__orders__retrieval_id__retrieval_destination_area=warehouse_name,
                delivery_type='public',
                container_number__is_abnormal_state=False,
                #container_number__orders__warehouse__name=warehouse,
            )&
            ~(
                models.Q(delivery_method__icontains='жЪВжЙ£') |
                models.Q(delivery_method__icontains='иЗ™жПР') |
                models.Q(delivery_method__icontains='UPS') |
                models.Q(delivery_method__icontains='FEDEX')
            ),
            models.Q(
                container_number__orders__offload_id__offload_at__isnull=False,
            )& models.Q(pk=0),
        )
        
        # 2гАБйҐДзЇ¶зЃ°зРЖвАФвАФжЬ™дљњзФ®зЪДзЇ¶еТМеЉВеЄЄзЪДзЇ¶
        shipments = await self.get_shipments_by_warehouse(warehouse,request)
        
        summary = await self.calculate_summary(unshipment_pos, shipments, warehouse)

        # 3гАБжЩЇиГљеМєйЕНеЖЕеЃєвАФвАФжЪВдЄНдљњзФ®
        # st_type = request.POST.get('st_type')
        # max_cbm, max_pallet = await self.get_capacity_limits(st_type)
        # matching_suggestions = await self.get_matching_suggestions(unshipment_pos, shipments,max_cbm,max_pallet)
        # primary_group_keys = set()
        # for suggestion in matching_suggestions:
        #     group_key = f"{suggestion['primary_group']['destination']}_{suggestion['primary_group']['delivery_method']}"
        #     primary_group_keys.add(group_key)
        # auto_matches = await self.get_auto_matches(unshipment_pos, shipments)
        
        # 4гАБе§±жХИзЪДPOзЃ°зРЖ
        invalid_pos = await self._invalid_po_check(warehouse)

        vessel_names = []
        vessel_dict = OrderedDict()
        destination_list = []
        vessel_eta_dict = {}
        for item in unshipment_pos:
            destination = item.get('destination')
            destination_list.append(destination)
            vessel_name = item.get('vessel_name')
            vessel_voyage = item.get('vessel_voyage')
            vessel_eta = item.get('vessel_eta')

            if vessel_name and vessel_name not in vessel_names:
                eta_date = vessel_eta
                vessel_eta_dict[vessel_name] = eta_date
                display_text = f"{vessel_name} / {vessel_voyage} вЖТ {str(vessel_eta).split()[0] if vessel_eta else 'жЬ™зЯ•'}"
                vessel_dict[vessel_name] = display_text
                vessel_names.append(vessel_name)
        vessel_dict = OrderedDict(
            sorted(vessel_dict.items(), key=lambda x: vessel_eta_dict[x[0]])
        )        
        destination_list = sorted(list(set(destination_list)))
        if not context:
            context = {}
        context.update({
            'warehouse': warehouse,
            'warehouse_options': self.warehouse_options,
            'cargos': unshipment_pos,
            'shipments': shipments,
            'summary': summary,
            'cargo_count': len(unshipment_pos),
            'appointment_count': len(shipments),
            'invalid_po_count': len(invalid_pos),
            #'matching_count': len(primary_group_keys),
            #'matching_suggestions': matching_suggestions,
            #'auto_matches': auto_matches,
            #'st_type': st_type,
            #'max_cbm': max_cbm,
            #'max_pallet': max_pallet,
            "invalid_pos": invalid_pos,
            "vessel_names": vessel_names,
            "vessel_dict": vessel_dict,
            "destination_list": destination_list,
            'account_options': self.account_options,
            'load_type_options': LOAD_TYPE_OPTIONS,
        })
        return self.template_main_dash, context
    
    async def handle_po_invalid_save(self, request):
        '''POж†°й™МдњЭе≠Ш'''
        json_data = request.POST.get('updated_data_json')
        if json_data:
            items_to_update = json.loads(json_data)
            
            # дљњзФ®еИЧи°®жЭ•жЙєйЗПжЫіжЦ∞жИЦеЊ™зОѓжЫіжЦ∞
            # е¶ВжЮЬжХ∞жНЃйЗПдЄНе§ІпЉМеЊ™зОѓжЫіжЦ∞жЬАзЃАеНХеЃЙеЕ®
            count = 0
            for item in items_to_update:
                po_id = item.get('id')
                # жЯ•жЙЊеѓєеЇФзЪДеѓєи±° (з°ЃдњЭжШѓPoCheckEtaSevenж®°еЮЛ)
                po_obj = await sync_to_async(
                    lambda pid=po_id: PoCheckEtaSeven.objects.filter(id=pid).first()
                )()
                
                if po_obj:
                    po_obj.fba_id = item.get('fba_id')
                    po_obj.ref_id = item.get('ref_id')
                    po_obj.handling_method = item.get('handling_method')
                    
                    # е§ДзРЖ Boolean е≠ЧжЃµ
                    new_notify_status = item.get('is_notified')
                    po_obj.is_notified = new_notify_status
                    
                    # е¶ВжЮЬзКґжАБеПШдЄЇеЈ≤йАЪзЯ•дЄФдєЛеЙНж≤°жЬЙжЧґйЧіпЉМиЃ∞ељХжЧґйЧі (еПѓйАЙйАїиЊС)
                    if new_notify_status and not po_obj.notified_time:
                        po_obj.notified_time = datetime.now()
                    elif not new_notify_status:
                        po_obj.notified_time = None
                        
                    po_obj.is_active = item.get('is_active')
                    
                    await sync_to_async(po_obj.save)()
                    count += 1
            context = {"success_messages": f'жИРеКЯжЫіжЦ∞ {count} жЭ° PO иЃ∞ељХ!'}
        
        return await self.handle_appointment_management_post(request,context)

    async def _invalid_po_check(
        self, warehouse
    ) -> dict[str, dict]:
        '''дЇЪй©ђйАКе§ЗзЇ¶дєЛжЯ•иѓҐе§±жХИзЪДpo'''
        warehouse = warehouse.split('-')[0]
        # е¶ВжЮЬжПРжЯЬеЙНдЄАе§©зКґжАБдЄЇе§±жХИпЉМжИЦиАЕжПРжЯЬеЙНдЄАе§©ж≤°жЬЙжЯ•пЉМеИ∞жЄѓеЙНдЄАеС®жЯ•дЇЖжШѓе§±жХИ
        # --- 1. жЮДеїЇжЯ•иѓҐжЭ°дїґ (дњЭжМБдЄНеПШ) ---
        query1 = models.Q(last_retrieval_checktime__isnull=False) & models.Q(
            last_retrieval_status=False
        )
        query2 = (
            models.Q(last_retrieval_checktime__isnull=True)
            & models.Q(last_eta_checktime__isnull=False)
            & models.Q(last_eta_status=False)
        )
        query = query1 | query2
        query &= models.Q(
            container_number__orders__retrieval_id__retrieval_destination_area=warehouse
        )
        query &= models.Q(ref_id__isnull=False)
        query &= ~models.Q(ref_id="")
        query &= models.Q(vessel_eta__gt=date(2026, 3, 31))

        # --- 2. еЃЪдєЙеРМж≠•жЯ•иѓҐеЗљжХ∞ (еЕ≥йФЃдњЃжФє) ---
        def get_po_data():
            # дљњзФ® select_related йҐДеК†иљље§ЦйФЃе≠ЧжЃµ
            # иѓЈз°ЃиЃ§ 'customer_name' еТМ 'container_number' жШѓдљ†зЪДе§ЦйФЃе≠ЧжЃµеРН
            # е¶ВжЮЬињШжЬЙеЕґдїЦе§ЦйФЃеЬ®ж®°жЭњдЄ≠жШЊз§ЇпЉИе¶В destination дєЯжШѓе§ЦйФЃпЉЙпЉМдєЯйЬАи¶БеК†ињЫеОї
            qs = PoCheckEtaSeven.objects.filter(query).select_related(
                'customer_name', 
                'container_number'
            ).distinct() 
            return list(qs)
        # --- 3. еЉВж≠•и∞ГзФ® ---
        po_checks_list = await sync_to_async(get_po_data)()
        # жОТеЇП
        po_checks_list.sort(key=lambda po: po.is_notified)

        return po_checks_list
    
    async def _get_packing_list(
        self,user,
        pl_criteria: models.Q | None = None,
        plt_criteria: models.Q | None = None,
        name: str | None = None
    ) -> list[Any]:
        pl_criteria &= models.Q(container_number__orders__cancel_notification=False)& ~models.Q(container_number__orders__order_type='зЫійАБ')
        plt_criteria &= models.Q(container_number__orders__cancel_notification=False)& ~models.Q(container_number__orders__order_type='зЫійАБ')
        if await self._validate_user_four_major_whs(user):
            pl_criteria &= models.Q(destination__in=FOUR_MAJOR_WAREHOUSES)
            plt_criteria &= models.Q(destination__in=FOUR_MAJOR_WAREHOUSES)

        def sort_key(item):
            custom_method = item.get("custom_delivery_method")
            if custom_method is None:
                custom_method = ""
            keywords = ["жЪВжЙ£", "HOLD", "зХЩдїУ"]
            return (any(k in custom_method for k in keywords),)
        
        def sort_key_pl(item):
            # зђђдЄАдЉШеЕИзЇІпЉЪжМЙеЫЫзїДеИЖз±ї + жЧґйЧіжОТеЇП
            if item.get('has_actual_retrieval'):
                # еЃЮйЩЕжПРжЯЬ
                actual_time = item.get('actual_retrieval_time')
                group = 0
                sort_time = actual_time or datetime.min

            elif item.get('has_appointment_retrieval'):
                # з†Бе§ійҐДзЇ¶
                arm_time = item.get('arm_time')
                group = 1
                sort_time = arm_time or datetime.min

            elif item.get('has_estimated_retrieval'):
                # йҐДиЃ°жПРжЯЬ
                estimated_time = item.get('estimated_time')
                group = 2
                sort_time = estimated_time or datetime.min

            else:
                # жЧ†иЃ°еИТ
                group = 3
                sort_time = datetime.min
            
            # дЉШеЕИзЇІ2: жККеМЕеРЂжЪВжЙ£зЪДжФЊжЬАеРОйЭҐ
            custom_method = item.get("custom_delivery_method", "") or ""
            keywords = ["жЪВжЙ£", "HOLD", "зХЩдїУ"]
            has_hold = any(k in custom_method.upper() for k in keywords)
            hold_flag = 1 if has_hold else 0
            return (group, sort_time, hold_flag)
        
        data = []
        if plt_criteria:
            pal_list = await sync_to_async(list)(
                Pallet.objects.prefetch_related(
                    "container_number",
                    "container_number__orders",
                    "container_number__orders__warehouse",
                    "shipment_batch_number",
                    "shipment_batch_number__fleet_number",
                    "container_number__orders__offload_id",
                    "container_number__orders__customer_name",
                    "container_number__orders__retrieval_id",
                    "container_number__orders__vessel_id",
                )
                .filter(plt_criteria)
                .annotate(
                    str_id=Cast("id", CharField()),
                    str_container_number=Cast("container_number__container_number", CharField()),                  
                    # ж†ЉеЉПеМЦvessel_etaдЄЇжЬИжЧ•
                    formatted_offload_at=Func(
                        F('container_number__orders__offload_id__offload_at'),
                        Value('MM-DD'),
                        function='to_char',
                        output_field=CharField()
                    ),
                    is_zhunshida=Case(
                        When(container_number__orders__customer_name__zem_name__icontains='еЗЖжЧґиЊЊ', then=Value(True)),
                        default=Value(False),
                        output_field=BooleanField()
                    ),
                    # еИЫеїЇеЃМжХізЪДзїДеРИе≠ЧжЃµпЉМйАЪињЗеЙНзЉАеМЇеИЖзКґжАБ
                    container_with_eta_retrieval=Concat(
                        Value("[еЈ≤еЕ•дїУ]"),
                        "container_number__container_number",
                        Value(" еЕ•дїУ:"),
                        "formatted_offload_at",
                        output_field=CharField()
                    ),
                    data_source=Value("PALLET", output_field=CharField()),  # жЈїеК†жХ∞жНЃжЇРж†ЗиѓЖ
                    is_pass=Value(True, output_field=BooleanField()),
                    rebuilt_is_dropped_pallet=Case(
                        When(
                            is_dropped_pallet=True,
                            then=Func(
                                F('master_shipment_batch_number__shipment_appointment'),
                                Value('YYYY-MM-DD'),
                                function='to_char',
                                output_field=CharField()
                            )
                        ),
                        default=Value(""),
                        output_field=CharField()
                    )                   
                )
                .values(
                    "destination",
                    "delivery_method",
                    "abnormal_palletization",
                    "delivery_window_start",
                    "delivery_window_end",
                    "note",
                    "container_number",
                    "is_dropped_pallet",
                    "rebuilt_is_dropped_pallet",
                    "shipment_batch_number__shipment_batch_number",
                    "data_source",  # еМЕеРЂжХ∞жНЃжЇРж†ЗиѓЖ
                    "shipment_batch_number__fleet_number__fleet_number",
                    "location",  # жЈїеК†locationзФ®дЇОжѓФиЊГ
                    "is_pass",
                    "is_zhunshida",
                    "shipment_note",
                    customer_name=F("container_number__orders__customer_name__zem_name"),
                    warehouse=F(
                        "container_number__orders__retrieval_id__retrieval_destination_precise"
                    ),               
                    retrieval_destination_precise=F("container_number__orders__retrieval_id__retrieval_destination_precise"),
                )
                .annotate(
                    custom_delivery_method=F("delivery_method"),
                    fba_ids=F("fba_id"),
                    ref_ids=F("ref_id"),
                    shipping_marks=F("shipping_mark"),
                    plt_ids=StringAgg(
                        "str_id", delimiter=",", distinct=True, ordering="str_id"
                    ),
                    container_numbers=StringAgg(  # иБЪеРИеЃМжХізЪДзїДеРИе≠ЧжЃµ
                        "container_with_eta_retrieval", delimiter="\n", distinct=True, ordering="container_with_eta_retrieval"
                    ),
                    cns=StringAgg(
                        "str_container_number", delimiter="\n", distinct=True, ordering="str_container_number"
                    ),
                    offload_time=StringAgg(
                        "formatted_offload_at", delimiter="\n", distinct=True, ordering="formatted_offload_at"
                    ),
                    total_pcs=Sum("pcs", output_field=IntegerField()),
                    total_cbm = Round(Sum("cbm", output_field=FloatField()), 3),
                    total_weight_lbs=Round(Sum("weight_lbs", output_field=FloatField()),3),
                    total_n_pallet_act=Count("pallet_id", distinct=True),
                    label=Value("ACT"),
                    note_sp=StringAgg("note_sp", delimiter=",", distinct=True),
                )
                .order_by("container_number__orders__offload_id__offload_at")
            )
            #еОїжОТжЯ•жШѓеР¶жЬЙиљђдїУзЪДпЉМжЬЙиљђдїУзЪДи¶БзЙєжЃКе§ДзРЖ
            pal_list_trans = await self._find_transfer(pal_list)
            pal_list_sorted = sorted(pal_list_trans, key=sort_key)          
            
            data += pal_list_sorted
        
        # PackingList жЯ•иѓҐ - жЈїеК†жХ∞жНЃжЇРж†ЗиѓЖ
        if pl_criteria:
            pl_list = await sync_to_async(list)(
                PackingList.objects.prefetch_related(
                    "container_number",
                    "container_number__orders",
                    "container_number__orders__warehouse",
                    "shipment_batch_number",
                    "shipment_batch_number__fleet_number",
                    "container_number__orders__offload_id",
                    "container_number__orders__customer_name",
                    "container_number__orders__retrieval_id",
                    "container_number__orders__vessel_id",
                )
                .filter(pl_criteria)
                .annotate(
                    #жЦєдЊњеРОзї≠жОТеЇП
                    has_actual_retrieval=Case(
                        When(container_number__orders__retrieval_id__actual_retrieval_timestamp__isnull=False, then=Value(1)),
                        default=Value(0),
                        output_field=IntegerField()
                    ),
                    has_appointment_retrieval=Case(
                        When(container_number__orders__retrieval_id__generous_and_wide_target_retrieval_timestamp__isnull=False, then=Value(1)),
                        default=Value(0),
                        output_field=IntegerField()
                    ),
                    has_estimated_retrieval=Case(
                        When(
                            Q(container_number__orders__retrieval_id__target_retrieval_timestamp_lower__isnull=False) |
                            Q(container_number__orders__retrieval_id__target_retrieval_timestamp__isnull=False),
                            then=Value(1)
                        ),
                        default=Value(0),
                        output_field=IntegerField()
                    ),
                    is_zhunshida=Case(
                        When(container_number__orders__customer_name__zem_name__icontains='еЗЖжЧґиЊЊ', then=Value(True)),
                        default=Value(False),
                        output_field=BooleanField()
                    ),
                    custom_delivery_method=Case(
                        When(
                            Q(delivery_method="жЪВжЙ£зХЩдїУ(HOLD)")
                            | Q(delivery_method="жЪВжЙ£зХЩдїУ"),
                            then=Concat(
                                "delivery_method",
                                Value("-"),
                                "fba_id",
                                Value("-"),
                                "id",
                            ),
                        ),
                        default=F("delivery_method"),
                        output_field=CharField(),
                    ),     
                    str_container_number=Cast("container_number__container_number", CharField()),    
                    # ж†ЉеЉПеМЦvessel_etaдЄЇжЬИжЧ•
                    formatted_vessel_eta=Func(
                        F('container_number__orders__vessel_id__vessel_eta'),
                        Value('MM-DD'),
                        function='to_char',
                        output_field=CharField()
                    ),
                    
                    # ж†ЉеЉПеМЦеЃЮйЩЕжПРжЯЬжЧґйЧідЄЇжЬИжЧ•
                    formatted_actual_retrieval=Func(
                        F('container_number__orders__retrieval_id__actual_retrieval_timestamp'),
                        Value('MM-DD'),
                        function='to_char',
                        output_field=CharField()
                    ),
                    # ж†ЉеЉПеМЦз†Бе§ійҐДзЇ¶жЧґйЧідЄЇжЬИжЧ•
                    formatted_appointment_retrieval=Func(
                        F('container_number__orders__retrieval_id__generous_and_wide_target_retrieval_timestamp'),
                        Value('MM-DD'),
                        function='to_char',
                        output_field=CharField()
                    ),
                    # ж†ЉеЉПеМЦйҐДиЃ°жПРжЯЬжЧґйЧідЄЇжЬИжЧ•
                    formatted_target_low=Func(
                        F('container_number__orders__retrieval_id__target_retrieval_timestamp_lower'),
                        Value('MM-DD'),
                        function='to_char',
                        output_field=CharField()
                    ),
                    
                    formatted_target=Func(
                        F('container_number__orders__retrieval_id__target_retrieval_timestamp'),
                        Value('MM-DD'),
                        function='to_char',
                        output_field=CharField()
                    ),
                    
                    # еИЫеїЇеЃМжХізЪДзїДеРИе≠ЧжЃµпЉМйАЪињЗеЙНзЉАеМЇеИЖзКґжАБ
                    container_with_eta_retrieval=Case(
                        # жЬЙеЃЮйЩЕжПРжЯЬжЧґйЧі - дљњзФ®еЙНзЉА [еЃЮйЩЕ]
                        When(container_number__orders__retrieval_id__actual_retrieval_timestamp__isnull=False,
                            then=Concat(
                                # Value(" "),
                                # "container_number__orders__vessel_id__vessel", 
                                Value("[еЈ≤жПРжЯЬ]"),
                                "container_number__container_number",                          
                                # Value(" ETA:"),
                                # "formatted_vessel_eta",
                                Value(" жПРжЯЬ:"),
                                "formatted_actual_retrieval",
                                output_field=CharField()
                            )),
                        # жЬЙз†Бе§ійҐДзЇ¶жЧґйЧі - дљњзФ®еЙНзЉА [з†Бе§ійҐДзЇ¶]
                        When(container_number__orders__retrieval_id__generous_and_wide_target_retrieval_timestamp__isnull=False,
                            then=Concat(
                                Value("[з†Бе§ійҐДзЇ¶]"),
                                "container_number__container_number",
                                Value(" йҐДиЃ°жПРжЯЬ:"),
                                "formatted_appointment_retrieval",
                                output_field=CharField()
                            )),
                        # жЬЙйҐДиЃ°жПРжЯЬжЧґйЧіиМГеЫі - дљњзФ®еЙНзЉА [йҐДиЃ°]
                        When(container_number__orders__retrieval_id__target_retrieval_timestamp_lower__isnull=False,
                            then=Concat( 
                                # Value(" "),
                                # "container_number__orders__vessel_id__vessel", 
                                Value("[йҐДиЃ°]"),
                                "container_number__container_number",
                                # Value(" ETA:"),
                                # "formatted_vessel_eta",
                                Value(" жПРжЯЬ:"),
                                "formatted_target_low",
                                Value("~"),
                                Coalesce("formatted_target", "formatted_target_low"),
                                output_field=CharField()
                            )),
                        # ж≤°жЬЙжПРжЯЬиЃ°еИТ - дљњзФ®еЙНзЉА [жЬ™еЃЙжОТ]
                        default=Concat(
                            # Value(" "),
                            # "container_number__orders__vessel_id__vessel", 
                            Value("[жЬ™еЃЙжОТжПРжЯЬ]"),
                            "container_number__container_number",
                            Value(" ETA:"),
                            "formatted_vessel_eta",
                            output_field=CharField()
                        ),
                        output_field=CharField()
                    ),
                    str_id=Cast("id", CharField()),
                    str_fba_id=Cast("fba_id", CharField()),
                    str_ref_id=Cast("ref_id", CharField()),
                    str_shipping_mark=Cast("shipping_mark", CharField()),
                    data_source=Value("PACKINGLIST", output_field=CharField()),  # жЈїеК†жХ∞жНЃжЇРж†ЗиѓЖ
                    is_pass=Case(
                        # 1. еЕИзЬЛ planned_release_time жШѓеР¶жЬЙеАЉ
                        When(
                            container_number__orders__retrieval_id__planned_release_time__isnull=False,
                            then=Value(True)
                        ),
                        # 2. йГљдЄНжї°иґ≥еИЩдЄЇ False
                        default=Value(False),
                        output_field=BooleanField()
                    ),
                )
                .values(
                    "destination",
                    "custom_delivery_method",
                    "delivery_window_start",
                    "delivery_window_end",
                    "note",
                    "container_number",
                    "data_source",  # еМЕеРЂжХ∞жНЃжЇРж†ЗиѓЖ
                    "shipment_batch_number__shipment_batch_number",
                    "shipment_batch_number__fleet_number__fleet_number",
                    # жОТеЇПе≠ЧжЃµ
                    "has_actual_retrieval",
                    "has_appointment_retrieval", 
                    "has_estimated_retrieval",
                    "is_zhunshida", 
                    "shipment_note",
                    customer_name=F("container_number__orders__customer_name__zem_name"),
                    warehouse=F(
                        "container_number__orders__retrieval_id__retrieval_destination_precise"
                    ),
                    vessel_name=F("container_number__orders__vessel_id__vessel"),
                    vessel_voyage=F("container_number__orders__vessel_id__voyage"),
                    vessel_eta=F("container_number__orders__vessel_id__vessel_eta"),
                    is_pass=F("is_pass"),
                    # жЈїеК†жЧґйЧіе≠ЧжЃµзФ®дЇОжОТеЇП
                    actual_retrieval_time=F("container_number__orders__retrieval_id__actual_retrieval_timestamp"),
                    arm_time=F("container_number__orders__retrieval_id__generous_and_wide_target_retrieval_timestamp"),
                    estimated_time=F("container_number__orders__retrieval_id__target_retrieval_timestamp"),                  
                )
                .annotate(
                    fba_ids=StringAgg(
                        "str_fba_id",
                        delimiter=",",
                        distinct=True,
                        ordering="str_fba_id",
                    ),
                    ref_ids=StringAgg(
                        "str_ref_id",
                        delimiter=",",
                        distinct=True,
                        ordering="str_ref_id",
                    ),
                    shipping_marks=StringAgg(
                        "str_shipping_mark",
                        delimiter=",",
                        distinct=True,
                        ordering="str_shipping_mark",
                    ),
                    ids=StringAgg(
                        "str_id", delimiter=",", distinct=True, ordering="str_id"
                    ),
                    container_numbers=StringAgg(  # иБЪеРИеЃМжХізЪДзїДеРИе≠ЧжЃµ
                        "container_with_eta_retrieval", delimiter="\n", distinct=True, ordering="container_with_eta_retrieval"
                    ),
                    cns=StringAgg(
                        "str_container_number", delimiter="\n", distinct=True, ordering="str_container_number"
                    ),
                    offload_time = Case(
                        # жЬЙеЃЮйЩЕжПРжЯЬжЧґйЧі
                        When(
                            container_number__orders__retrieval_id__actual_retrieval_timestamp__isnull=False,
                            then=Concat(
                                Value("еЃЮйЩЕжПРжЯЬпЉЪ"),
                                Func(
                                    F('container_number__orders__retrieval_id__actual_retrieval_timestamp'),
                                    Value('YYYY-MM-DD'),
                                    function='to_char'
                                ),
                                output_field=CharField()
                            )
                        ),

                        # еРМжЧґжЬЙдЄКдЄЛйЩР вЖТ иМГеЫі
                        When(
                            Q(container_number__orders__retrieval_id__target_retrieval_timestamp_lower__isnull=False)
                            & Q(container_number__orders__retrieval_id__target_retrieval_timestamp__isnull=False),
                            then=Concat(
                                Value("йҐДиЃ°жПРжЯЬпЉЪ"),
                                Func(
                                    F('container_number__orders__retrieval_id__target_retrieval_timestamp_lower'),
                                    Value('YYYY-MM-DD'),
                                    function='to_char'
                                ),
                                Value("~"),
                                Func(
                                    F('container_number__orders__retrieval_id__target_retrieval_timestamp'),
                                    Value('YYYY-MM-DD'),
                                    function='to_char'
                                ),
                                output_field=CharField()
                            )
                        ),

                        # еП™жЬЙдЄЛйЩР
                        When(
                            container_number__orders__retrieval_id__target_retrieval_timestamp_lower__isnull=False,
                            then=Concat(
                                Value("йҐДиЃ°жПРжЯЬпЉЪ"),
                                Func(
                                    F('container_number__orders__retrieval_id__target_retrieval_timestamp_lower'),
                                    Value('YYYY-MM-DD'),
                                    function='to_char'
                                ),
                                output_field=CharField()
                            )
                        ),

                        # еП™жЬЙдЄКйЩР
                        When(
                            container_number__orders__retrieval_id__target_retrieval_timestamp__isnull=False,
                            then=Concat(
                                Value("йҐДиЃ°жПРжЯЬпЉЪ"),
                                Func(
                                    F('container_number__orders__retrieval_id__target_retrieval_timestamp'),
                                    Value('YYYY-MM-DD'),
                                    function='to_char'
                                ),
                                output_field=CharField()
                            )
                        ),

                        # йГљж≤°жЬЙ
                        default=Value("жЧ†йҐДиЃ°жПРжЯЬ"),
                        output_field=CharField()
                    ),
                    total_pcs=Sum("pcs", output_field=FloatField()),
                    total_cbm = Round(Sum("cbm", output_field=FloatField()), 3),
                    total_weight_lbs=Round(Sum("total_weight_lbs", output_field=FloatField()),3),
                    total_n_pallet_est= Ceil(Sum("cbm", output_field=FloatField()) / 2),
                    total_n_pallet_act= Ceil(Sum("cbm", output_field=FloatField()) / 2),
                    label=Value("EST"),
                    note_sp=StringAgg("note_sp", delimiter=",", distinct=True)
                )
                .distinct()
            )
            pl_list_sorted = sorted(pl_list, key=sort_key_pl)
            data += pl_list_sorted      
        return data

    async def _find_transfer(self, pal_list:list):
        # зђђдЄАж≠•пЉЪеЕИз≠ЫйАЙеЗЇйЬАи¶БдњЃжФєзЪДиЃ∞ељХ
        need_update_pallets = []
        for pallet in pal_list:
            retrieval_destination = pallet.get('retrieval_destination_precise')
            current_location = pallet.get('location')
            
            # ж£АжЯ•жШѓеР¶йЬАи¶БдњЃжФє
            if retrieval_destination and current_location and retrieval_destination != current_location:
                need_update_pallets.append(pallet)

        # зђђдЇМж≠•пЉЪеП™еѓєйЬАи¶БдњЃжФєзЪДиЃ∞ељХжЯ•иѓҐTransferLocation
        if need_update_pallets:
            # иОЈеПЦйЬАи¶БжЯ•иѓҐзЪДpallet IDs
            all_need_update_ids = set()
            plt_ids_to_pallet_map = {} 
            for pallet in need_update_pallets:
                plt_ids_str = pallet.get('plt_ids', '')
                if plt_ids_str:
                    try:
                        plt_id_list = [pid.strip() for pid in plt_ids_str.split(',') if pid.strip()]
                        for plt_id in plt_id_list:
                            if plt_id.isdigit():
                                plt_id_int = int(plt_id)
                                all_need_update_ids.add(plt_id_int)
                                plt_ids_to_pallet_map[plt_id_int] = pallet
                    except (ValueError, AttributeError):
                        continue
            # жЙєйЗПжЯ•иѓҐTransferLocationиЃ∞ељХ
            transfer_locations = await sync_to_async(list)(
                TransferLocation.objects.filter(plt_ids__isnull=False)
            )
            
            # еИЫеїЇplt_idеИ∞TransferLocationзЪДжШ†е∞Д
            plt_id_transfer_map = {}
            for transfer in transfer_locations:
                if transfer.plt_ids:
                    try:
                        transfer_plt_ids = [pid.strip() for pid in transfer.plt_ids.split(',') if pid.strip()]
                        for plt_id in transfer_plt_ids:
                            if plt_id.isdigit():
                                plt_id_int = int(plt_id)
                                if plt_id_int in all_need_update_ids:
                                    plt_id_transfer_map[plt_id_int] = transfer
                    except (ValueError, AttributeError):
                        continue
            
            # зђђдЄЙж≠•пЉЪе§ДзРЖжѓПдЄ™йЬАи¶БжЫіжЦ∞зЪДpalletиЃ∞ељХ
            processed_pallets = set()  # иЃ∞ељХеЈ≤зїПе§ДзРЖињЗзЪДpalletиЃ∞ељХпЉИйБњеЕНйЗНе§Не§ДзРЖпЉЙ
            
            for plt_id, transfer_record in plt_id_transfer_map.items():
                pallet = plt_ids_to_pallet_map.get(plt_id)
                if pallet and id(pallet) not in processed_pallets:
                    retrieval_destination = pallet.get('retrieval_destination_precise')
                    
                    if transfer_record:
                        # жПРеПЦеОЯеІЛдїУеРНзІ∞пЉИretrieval_destination_preciseдї•-еИЖзїДпЉМеПЦеЙНйЭҐзЪДеАЉпЉЙ
                        original_warehouse = retrieval_destination.split('-')[0] if '-' in retrieval_destination else retrieval_destination
                        if transfer_record.arrival_time:              
                            # ж†ЉеЉПеМЦеИ∞иЊЊжЧґйЧі
                            arrival_time_str = transfer_record.arrival_time.strftime('%m-%d')
                        elif transfer_record.ETA: 
                            # ж†ЉеЉПеМЦеИ∞иЊЊжЧґйЧі
                            arrival_time_str = transfer_record.ETA.strftime('%m-%d')
                        else:
                            arrival_time_str = "иљђдїУдЄ≠"                     
                        # дњЃжФєoffload_time
                        pallet['offload_time'] = f"{original_warehouse}-{arrival_time_str}"
                    
                    processed_pallets.add(id(pallet))
            
            # зђђеЫЫж≠•пЉЪе§ДзРЖж≤°жЬЙжЙЊеИ∞TransferLocationиЃ∞ељХдљЖйЬАи¶БжЫіжЦ∞зЪДpallet
            for pallet in need_update_pallets:
                if id(pallet) not in processed_pallets:
                    retrieval_destination = pallet.get('retrieval_destination_precise')
                    original_warehouse = retrieval_destination.split('-')[0] if '-' in retrieval_destination else retrieval_destination
                    pallet['offload_time'] = f"{original_warehouse}-иљђдїУдЄ≠"
        return pal_list

    async def get_shipments_by_warehouse(self, warehouse, request, four_major_whs: str | None = None):
        """еЉВж≠•иОЈеПЦжМЗеЃЪдїУеЇУзЫЄеЕ≥зЪДйҐДзЇ¶жХ∞жНЃ"""
        criteria = (
            (
                models.Q(origin__isnull=True)
                | models.Q(origin="")
                | models.Q(origin=warehouse)
            )
            & models.Q(appointment_id__isnull=False)
            & models.Q(in_use=False, is_canceled=False)
        )
        if four_major_whs == "four_major_whs":
            criteria &= models.Q(destination__in=FOUR_MAJOR_WAREHOUSES)
        appointment = await sync_to_async(list)(
            Shipment.objects.filter(criteria).order_by("shipment_appointment","shipment_account")
        )
        
        return appointment
    
    async def get_used_pallets(self, shipment):
        """еЉВж≠•иЃ°зЃЧйҐДзЇ¶еЈ≤дљњзФ®зЪДжЭњжХ∞"""
        
        # еЉВж≠•иОЈеПЦзЫЄеЕ≥иЃҐеНХ
        related_orders = await sync_to_async(list)(shipment.order_set.all())
        total_pallets = 0
        
        async def process_order(order):
            order_pallets = 0
            packing_lists = await sync_to_async(list)(order.container_number.packinglist_set.all())
            for packing in packing_lists:
                order_pallets += packing.n_pallet or 0
            return order_pallets
        
        tasks = [process_order(order) for order in related_orders]
        order_pallets_list = await asyncio.gather(*tasks)
        total_pallets = sum(order_pallets_list)
        
        return total_pallets
    
    def _parse_tzinfo(self, s: str) -> str:
        if not s:
            return "America/New_York"
        if "NJ" in s.upper():
            return "America/New_York"
        elif "SAV" in s.upper():
            return "America/New_York"
        elif "LA" in s.upper():
            return "America/Los_Angeles"
        else:
            return "America/New_York"

    async def _now_time_get(self, warehouse):
        today = timezone.now()
        if 'LA' in warehouse:
            local_tz = pytz.timezone("America/Los_Angeles")
        else:
            local_tz = pytz.timezone("America/New_York")

        today = timezone.localtime(today, local_tz)
        return today
    
    async def _four_major_calculate_summary(self, unshipment_pos, shipments, scheduled_data, schedule_fleet_data, ready_to_ship_data, delivery_data, pod_data, warehouse):
        now = await self._now_time_get(warehouse)
        # иЃ°зЃЧйҐДзЇ¶зКґжАБзїЯиЃ°
        expired_count = 0
        urgent_count = 0
        available_count = 0
        used_count = 0  # еЈ≤дљњзФ®зЪДйҐДзЇ¶жХ∞йЗП
        
        for shipment in shipments:
            # ж£АжЯ•йҐДзЇ¶жШѓеР¶еЈ≤ињЗжЬЯ
            is_expired = (
                shipment.shipment_appointment_utc and 
                shipment.shipment_appointment_utc < now
            )
            
            # ж£АжЯ•йҐДзЇ¶жШѓеР¶еН≥е∞ЖињЗжЬЯпЉИ7е§©еЖЕпЉЙ
            is_urgent = (
                shipment.shipment_appointment_utc and 
                shipment.shipment_appointment_utc - now < timedelta(days=7) and
                not is_expired
            )
            
            # ж£АжЯ•йҐДзЇ¶жШѓеР¶еЈ≤иҐЂдљњзФ®пЉИйАЪињЗ PackingList жИЦ Pallet зїСеЃЪпЉЙ
            has_packinglist = await self.has_related_packinglist(shipment)
            has_pallet = await self.has_related_pallet(shipment)
            is_used = has_packinglist or has_pallet
            
            if is_used:
                shipment.status = "used"
                used_count += 1
            elif is_expired:
                shipment.status = "expired"
                expired_count += 1
            elif is_urgent:
                shipment.status = "urgent"
                urgent_count += 1
            else:
                shipment.status = "available"
                available_count += 1
        
        # иЃ°зЃЧиіІзЙ©зїЯиЃ°
        pending_cargos_count = len(unshipment_pos)
        
        # иЃ°зЃЧжАїжЭњжХ∞
        total_pallets = 0
        for cargo in unshipment_pos:
            if cargo.get('label') == 'ACT':  # еЃЮйЩЕжЭњжХ∞
                total_pallets += cargo.get('total_n_pallet_act', 0) or 0
            else:  # йҐДдЉ∞жЭњжХ∞
                total_pallets += cargo.get('total_n_pallet_est', 0) or 0
        
        return {
            'expired_count': expired_count,
            'urgent_count': urgent_count,
            'available_count': available_count,
            'used_count': used_count,  # еЈ≤дљњзФ®зЪДйҐДзЇ¶жХ∞йЗП
            'pending_cargo_count': pending_cargos_count,
            'total_pallets': int(total_pallets),
            'scheduled_sp_count': len(scheduled_data),
            'schedule_fl_count': len(schedule_fleet_data),
            'ready_to_ship_count': len(ready_to_ship_data),
            'ready_count': len(delivery_data),
            'pod_count': len(pod_data),
        }

    async def calculate_summary(self, unshipment_pos, shipments, warehouse):
        """еЉВж≠•иЃ°зЃЧзїЯиЃ°жХ∞жНЃ - йАВйЕНжЦ∞зЪДжХ∞жНЃзїУжЮД"""
        now = await self._now_time_get(warehouse)
        # иЃ°зЃЧйҐДзЇ¶зКґжАБзїЯиЃ°
        expired_count = 0
        urgent_count = 0
        available_count = 0
        used_count = 0  # еЈ≤дљњзФ®зЪДйҐДзЇ¶жХ∞йЗП
        
        for shipment in shipments:
            # ж£АжЯ•йҐДзЇ¶жШѓеР¶еЈ≤ињЗжЬЯ
            is_expired = (
                shipment.shipment_appointment_utc and 
                shipment.shipment_appointment_utc < now
            )
            
            # ж£АжЯ•йҐДзЇ¶жШѓеР¶еН≥е∞ЖињЗжЬЯпЉИ7е§©еЖЕпЉЙ
            is_urgent = (
                shipment.shipment_appointment_utc and 
                shipment.shipment_appointment_utc - now < timedelta(days=7) and
                not is_expired
            )
            
            # ж£АжЯ•йҐДзЇ¶жШѓеР¶еЈ≤иҐЂдљњзФ®пЉИйАЪињЗ PackingList жИЦ Pallet зїСеЃЪпЉЙ
            has_packinglist = await self.has_related_packinglist(shipment)
            has_pallet = await self.has_related_pallet(shipment)
            is_used = has_packinglist or has_pallet
            
            if is_used:
                shipment.status = "used"
                used_count += 1
            elif is_expired:
                shipment.status = "expired"
                expired_count += 1
            elif is_urgent:
                shipment.status = "urgent"
                urgent_count += 1
            else:
                shipment.status = "available"
                available_count += 1
        
        # иЃ°зЃЧиіІзЙ©зїЯиЃ°
        pending_cargos_count = len(unshipment_pos)
        
        # иЃ°зЃЧжАїжЭњжХ∞
        total_pallets = 0
        for cargo in unshipment_pos:
            if cargo.get('label') == 'ACT':  # еЃЮйЩЕжЭњжХ∞
                total_pallets += cargo.get('total_n_pallet_act', 0) or 0
            else:  # йҐДдЉ∞жЭњжХ∞
                total_pallets += cargo.get('total_n_pallet_est', 0) or 0
        
        return {
            'expired_count': expired_count,
            'urgent_count': urgent_count,
            'available_count': available_count,
            'used_count': used_count,  # еЈ≤дљњзФ®зЪДйҐДзЇ¶жХ∞йЗП
            'pending_cargo_count': pending_cargos_count,
            'total_pallets': int(total_pallets),
        }

    async def has_related_packinglist(self, shipment):
        """ж£АжЯ•йҐДзЇ¶жШѓеР¶жЬЙзЫЄеЕ≥зЪД PackingList иЃ∞ељХ"""
        
        try:
            # дљњзФ® sync_to_async еМЕи£ЕжХ∞жНЃеЇУжЯ•иѓҐ
            packinglist_exists = await sync_to_async(
                PackingList.objects.filter(shipment_batch_number=shipment).exists
            )()
            return packinglist_exists
        except Exception:
            return False

    async def has_related_pallet(self, shipment):
        """ж£АжЯ•йҐДзЇ¶жШѓеР¶жЬЙзЫЄеЕ≥зЪД Pallet иЃ∞ељХ"""
        
        try:
            # дљњзФ® sync_to_async еМЕи£ЕжХ∞жНЃеЇУжЯ•иѓҐ
            pallet_exists = await sync_to_async(
                Pallet.objects.filter(shipment_batch_number=shipment).exists
            )()
            return pallet_exists
        except Exception:
            return False
    
    async def has_appointment(self, cargo):
        """еЉВж≠•еИ§жЦ≠иіІзЙ©жШѓеР¶еЈ≤жЬЙйҐДзЇ¶ - йАВйЕНжЦ∞зЪДжХ∞жНЃзїУжЮД"""
        # ж†єжНЃдљ†зЪДжХ∞жНЃзїУжЮДпЉМеИ§жЦ≠жШѓеР¶жЬЙйҐДзЇ¶еПЈ
        return cargo.get('shipment_batch_number__shipment_batch_number') is not None
    
    async def get_matching_suggestions(self, unshipment_pos, shipments, max_cbm,max_pallet):
        """еЉВж≠•зФЯжИРжЩЇиГљеМєйЕНеїЇиЃЃ - йАВйЕНжЦ∞зЪДжХ∞жНЃзїУжЮД"""
        
        suggestions = []

        # зђђдЄАзЇІеИЖзїДпЉЪжМЙзЫЃзЪДеЬ∞еТМжіЊйАБжЦєеЉПйҐДеИЖзїД
        pre_groups = {}
        for cargo in unshipment_pos:
            if not await self.has_appointment(cargo):
                dest = (cargo.get('destination') or '').strip().upper()
                raw_method = (cargo.get('custom_delivery_method') or '').strip()
                if 'еН°иљ¶жіЊйАБ' in raw_method:
                    delivery_method = 'еН°жіЊ'
                elif 'жЪВжЙ£' in raw_method:
                    delivery_method = 'жЪВжЙ£'
                elif 'еЃҐжИЈиЗ™жПР' in raw_method:
                    delivery_method = 'еЃҐжПР'
                else:
                    delivery_method = raw_method.upper()
                if not dest or not delivery_method:
                    continue
                    
                group_key = f"{dest}_{delivery_method}"
                if group_key not in pre_groups:
                    pre_groups[group_key] = {
                        'destination': dest,
                        'delivery_method': delivery_method,
                        'cargos': []
                    }
                pre_groups[group_key]['cargos'].append(cargo)
        
        # еѓєжѓПдЄ™йҐДеИЖзїДжМЙеЃєйЗПйЩРеИґеИЫеїЇе§ІзїД
        for group_key, pre_group in pre_groups.items():
            cargos = pre_group['cargos']
            
            # жМЙETAжОТеЇПпЉМдЉШеЕИеЃЙжОТжЧ©зЪДиіІзЙ©
            sorted_cargos = sorted(cargos, key=lambda x: x.get('container_number__orders__vessel_id__vessel_eta') or '')
            
            # жМЙеЃєйЗПйЩРеИґеИЫеїЇе§ІзїД
            primary_groups = []
            current_primary_group = {
                'destination': pre_group['destination'],
                'delivery_method': pre_group['delivery_method'],
                'cargos': [],  # ињЩдЄ™е§ІзїДеМЕеРЂзЪДжЙАжЬЙиіІзЙ©пЉИжѓПдЄ™иіІзЙ©е∞±жШѓдЄАдЄ™е∞ПзїДпЉЙ
                'total_pallets': 0,
                'total_cbm': 0,
            }
            
            for cargo in sorted_cargos:
                cargo_pallets = 0
                if cargo.get('label') == 'ACT':
                    cargo_pallets = cargo.get('total_n_pallet_act', 0) or 0
                else:
                    cargo_pallets = cargo.get('total_n_pallet_est', 0) or 0
                
                cargo_cbm = cargo.get('total_cbm', 0) or 0
                
                # ж£АжЯ•ељУеЙНе§ІзїДжШѓеР¶ињШиГљеЃєзЇ≥ињЩдЄ™иіІзЙ©
                if (current_primary_group['total_pallets'] + cargo_pallets <= max_pallet and 
                    current_primary_group['total_cbm'] + cargo_cbm <= max_cbm):
                    # еПѓдї•еК†еЕ•ељУеЙНе§ІзїД
                    current_primary_group['cargos'].append(cargo)
                    current_primary_group['total_pallets'] += cargo_pallets
                    current_primary_group['total_cbm'] += cargo_cbm
                else:
                    # ељУеЙНе§ІзїДеЈ≤жї°пЉМдњЭе≠ШеєґеИЫеїЇжЦ∞зЪДе§ІзїД
                    if current_primary_group['cargos']:
                        primary_groups.append(current_primary_group)
                    
                    # еИЫеїЇжЦ∞зЪДе§ІзїД
                    current_primary_group = {
                        'destination': pre_group['destination'],
                        'delivery_method': pre_group['delivery_method'],
                        'cargos': [cargo],
                        'total_pallets': cargo_pallets,
                        'total_cbm': cargo_cbm,
                    }
            
            # жЈїеК†жЬАеРОдЄАдЄ™е§ІзїД
            if current_primary_group['cargos']:
                primary_groups.append(current_primary_group)
            
            # дЄЇжѓПдЄ™е§ІзїДеИЫеїЇеїЇиЃЃпЉМе§ІзїДдЄ≠зЪДжѓПдЄ™иіІзЙ©йГљжШѓдЄАдЄ™е∞ПзїДпЉИдЄАи°МпЉЙ
            for primary_group_index, primary_group in enumerate(primary_groups):
                # иЃ°зЃЧе§ІзїДзЪДеМєйЕНеЇ¶зЩЊеИЖжѓФ
                pallets_percentage = min(100, (primary_group['total_pallets'] / max_pallet) * 100) if max_pallet > 0 else 0
                cbm_percentage = min(100, (primary_group['total_cbm'] / max_cbm) * 100) if max_cbm > 0 else 0
                
                # е§ІзїДдЄ≠зЪДжѓПдЄ™иіІзЙ©йГљжШѓдЄАдЄ™е∞ПзїДпЉИдЄАи°МпЉЙ
                for subgroup_index, cargo in enumerate(primary_group['cargos']):
                    # иЃ°зЃЧињЩдЄ™иіІзЙ©зЪДжЭњжХ∞еТМCBM
                    cargo_pallets = 0
                    if cargo.get('label') == 'ACT':
                        cargo_pallets = cargo.get('total_n_pallet_act', 0) or 0
                    else:
                        cargo_pallets = cargo.get('total_n_pallet_est', 0) or 0
                    
                    cargo_cbm = cargo.get('total_cbm', 0) or 0
                    
                    suggestion = {
                        'id': f"{group_key}_{primary_group_index}_{subgroup_index}",
                        'primary_group': {
                            'destination': primary_group['destination'],
                            'delivery_method': primary_group['delivery_method'],
                            'total_pallets': primary_group['total_pallets'],
                            'total_cbm': primary_group['total_cbm'],
                            'pallets_percentage': pallets_percentage,
                            'cbm_percentage': cbm_percentage,
                        },
                        'subgroup': {
                            'cargos': [{
                                'ids': cargo.get('ids', ''),  # з°ЃдњЭеМЕеРЂids
                                'plt_ids': cargo.get('plt_ids', ''),  # з°ЃдњЭеМЕеРЂplt_ids
                                'container_numbers': cargo.get('container_numbers', ''),
                                'cns': cargo.get('cns', ''),
                                'offload_time': cargo.get('offload_time',''),
                                'total_n_pallet_act': cargo.get('total_n_pallet_act', 0),
                                'total_n_pallet_est': cargo.get('total_n_pallet_est', 0),
                                'total_cbm': cargo.get('total_cbm', 0),
                                'label': cargo.get('label', ''),
                            }],
                            'total_pallets': cargo_pallets,
                            'total_cbm': cargo_cbm,
                            'container_numbers': cargo.get('container_numbers', ''),
                            'cns': cargo.get('cns', ''),
                            'offload_time': cargo.get('offload_time',''),
                            'cargo_count': 1
                        },
                        'subgroup_index': subgroup_index + 1,
                    }
                    suggestions.append(suggestion)
        return suggestions
    
    async def is_shipment_available(self, shipment):
        """еИ§жЦ≠йҐДзЇ¶жШѓеР¶еПѓзФ®"""     
        now = timezone.now()
        
        # еЈ≤еПСиіІзЪДдЄНеПѓзФ®
        if shipment.shipped_at:
            return False
        
        # еЈ≤ињЗжЬЯзЪДдЄНеПѓзФ®
        if (shipment.shipment_appointment and 
            shipment.shipment_appointment < now):
            return False
        
        return True
    
    async def get_auto_matches(self, unshipment_pos, shipments):
        """еЉВж≠•иОЈеПЦиЗ™еК®еМєйЕНзїУжЮЬ - йАВйЕНжЦ∞зЪДжХ∞жНЃзїУжЮД"""
        matches = []
        
        # жМЙзЫЃзЪДеЬ∞еИЖзїДиіІзЙ©
        destination_groups = {}
        for cargo in unshipment_pos:
            if not await self.has_appointment(cargo):
                dest = cargo.get('destination')
                if dest not in destination_groups:
                    destination_groups[dest] = []
                destination_groups[dest].append(cargo)
        
        # дЄЇжѓПдЄ™зЫЃзЪДеЬ∞зФЯжИРеМєйЕНзїДеРИ
        match_id = 1
        for destination, cargo_list in destination_groups.items():
            # жМЙжЭњжХ∞жОТеЇПпЉМдЉШеЕИеМєйЕНе§ІжЭњжХ∞зЪДиіІзЙ©
            sorted_cargos = sorted(cargo_list, 
                                 key=lambda x: x.get('total_n_pallet_act', 0) or x.get('total_n_pallet_est', 0) or 0, 
                                 reverse=True)
            
            # зФЯжИРеМєйЕНзїДеРИпЉИе∞љйЗПжО•ињС35жЭњпЉЙ
            current_group = []
            current_pallets = 0
            
            for cargo in sorted_cargos:
                cargo_pallets = (cargo.get('total_n_pallet_act', 0) or 
                               cargo.get('total_n_pallet_est', 0) or 0)
                
                if current_pallets + cargo_pallets <= 35:
                    current_group.append(cargo)
                    current_pallets += cargo_pallets
                else:
                    # ељУеЙНзїДеЈ≤жї°пЉМеИЫеїЇеМєйЕН
                    if current_group:
                        match_percentage = min(int((current_pallets / 35) * 100), 100)
                        
                        # жЯ•жЙЊжЬАдљ≥йҐДзЇ¶
                        best_shipment = await self.find_best_shipment_for_match(destination, current_pallets, shipments)
                        
                        matches.append({
                            'id': match_id,
                            'destination': destination,
                            'cargo_count': len(current_group),
                            'total_pallets': int(current_pallets),
                            'recommended_appointment': best_shipment,
                            'match_percentage': match_percentage,
                            'cargos': current_group[:5]  # еП™жШЊз§ЇеЙН5дЄ™иіІзЙ©иѓ¶жГЕ
                        })
                        match_id += 1
                    
                    # еЉАеІЛжЦ∞зїД
                    current_group = [cargo]
                    current_pallets = cargo_pallets
            
            # е§ДзРЖжЬАеРОдЄАзїД
            if current_group:
                match_percentage = min(int((current_pallets / 35) * 100), 100)
                best_shipment = await self.find_best_shipment_for_match(destination, current_pallets, shipments)
                
                matches.append({
                    'id': match_id,
                    'destination': destination,
                    'cargo_count': len(current_group),
                    'total_pallets': int(current_pallets),
                    'recommended_appointment': best_shipment,
                    'match_percentage': match_percentage,
                    'cargos': current_group[:5]
                })
                match_id += 1
        
        return matches[:10]  # йЩРеИґињФеЫЮжХ∞йЗП
    
    async def find_best_shipment_for_match(self, destination, total_pallets, shipments):
        """дЄЇеМєйЕНзїДеРИжЯ•жЙЊжЬАдљ≥йҐДзЇ¶"""
        best_shipment = None
        best_capacity_diff = float('inf')
        
        for shipment in shipments:
            if (shipment.destination == destination and 
                await self.is_shipment_available(shipment)):
                
                used_pallets = await self.get_used_pallets(shipment)
                remaining_capacity = 35 - used_pallets
                capacity_diff = abs(remaining_capacity - total_pallets)
                
                if capacity_diff < best_capacity_diff:
                    best_capacity_diff = capacity_diff
                    best_shipment = shipment
        
        return best_shipment

    async def handle_fleet_po_search_post(self, request: HttpRequest, context=None) -> tuple[str, dict[str, Any]]:
        """иљ¶йШЯPOж†ЄеѓєжРЬзіҐеКЯиГљ"""
        if context is None:
            context = {}
        
        # иОЈеПЦжРЬзіҐжЭ°дїґ
        fleet_number = request.POST.get('fleet_number', '').strip()
        pickup_number = request.POST.get('pickup_number', '').strip()
        container_number = request.POST.get('container_number', '').strip()
        destination = request.POST.get('destination', '').strip()
        shipment_batch_number = request.POST.get('shipment_batch_number', '').strip()
        isa = request.POST.get('isa', '').strip()
        
        # е∞ЖжРЬзіҐжЭ°дїґдЉ†йАТеИ∞еЙНзЂѓ
        context.update({
            'fleet_number': fleet_number,
            'pickup_number': pickup_number,
            'container_number': container_number,
            'destination': destination,
            'shipment_batch_number': shipment_batch_number,
            'isa': isa
        })
        
        # жЮДеїЇжЯ•иѓҐжЭ°дїґ
        criteria = Q()
        
        # ж†єжНЃдЄНеРМзЪДжРЬзіҐжЭ°дїґжЮДеїЇfilter
        if fleet_number:
            criteria &= Q(shipment_batch_number__fleet_number__fleet_number__icontains=fleet_number)
        elif pickup_number:
            criteria &= Q(shipment_batch_number__pickup_number__icontains=pickup_number)
        elif shipment_batch_number:
            criteria &= Q(shipment_batch_number__shipment_batch_number__icontains=shipment_batch_number)
        elif isa:
            criteria &= Q(shipment_batch_number__appointment_id__icontains=isa)
        elif container_number:
            criteria &= Q(container_number__container_number__icontains=container_number)
        elif destination:
            criteria &= Q(destination__icontains=destination)

        # е¶ВжЮЬж≤°жЬЙдїїдљХжРЬзіҐжЭ°дїґпЉМињФеЫЮз©ЇзїУжЮЬ
        if not any([fleet_number, pickup_number, container_number, destination, shipment_batch_number, isa]):
            context['search_results'] = []
            return self.template_fleet_po_check, context
        
        # жЯ•иѓҐPalletжХ∞жНЃ
        pallets = await sync_to_async(list)(
            Pallet.objects.select_related(
                'container_number',
                'shipment_batch_number',
                'shipment_batch_number__fleet_number'
            )
            .filter(criteria)
            .values(
                'id',
                'PO_ID',
                'container_number__container_number',
                'destination',
                'shipment_batch_number__shipment_batch_number',
                'shipment_batch_number__fleet_number__fleet_number',
                'shipment_batch_number__pickup_number',
                'shipment_batch_number__appointment_id',
                'shipment_batch_number_id'
            )
        )
        
        # жМЙPO_IDеТМshipment_batch_numberињЫи°МеИЖзїД
        grouped_results = {}
        for pallet in pallets:
            po_id = pallet.get('PO_ID', 'жЧ†PO_ID')
            batch_number = pallet.get('shipment_batch_number__shipment_batch_number', 'жЧ†жЙєжђ°еПЈ')
            group_key = f"{po_id}_{batch_number}"
            
            if group_key not in grouped_results:
                grouped_results[group_key] = {
                    'po_id': po_id,
                    'shipment_batch_number': batch_number,
                    'fleet_number': pallet.get('shipment_batch_number__fleet_number__fleet_number', ''),
                    'items': [],
                    'pallet_ids': [],
                    'total_pallets': 0
                }
            
            # жЈїеК†palletдњ°жБѓеИ∞зїДдЄ≠
            grouped_results[group_key]['items'].append({
                'fleet_number': pallet.get('shipment_batch_number__fleet_number__fleet_number', ''),
                'fleet_number_detail': pallet.get('shipment_batch_number__fleet_number__fleet_number', ''),
                'pickup_number': pallet.get('shipment_batch_number__pickup_number', ''),
                'shipment_batch_number': batch_number,
                'appointment_id': pallet.get('shipment_batch_number__appointment_id', ''),
                'container_number': pallet.get('container_number__container_number', ''),
                'destination': pallet.get('destination', ''),
                'pallet_count': 1,  # жѓПдЄ™palletзЃЧ1жЭњ
                'pallet_ids': [pallet['id']],
                'shipment_id': pallet.get('shipment_batch_number_id')
            })
            
            grouped_results[group_key]['pallet_ids'].append(pallet['id'])
            grouped_results[group_key]['total_pallets'] += 1
        
        # дЄЇжѓПдЄ™зїДеЖЕзЪДitemиЃ°зЃЧpallet_count
        for group in grouped_results.values():
            # жМЙcontainer_numberеТМdestinationињЫдЄАж≠•еИЖзїД
            sub_grouped = {}
            for item in group['items']:
                sub_key = f"{item['container_number']}_{item['destination']}"
                if sub_key not in sub_grouped:
                    sub_grouped[sub_key] = {
                        'fleet_number': item['fleet_number'],
                        'fleet_number_detail': item['fleet_number_detail'],
                        'shipment_batch_number': item['shipment_batch_number'],
                        'appointment_id': item['appointment_id'],
                        'container_number': item['container_number'],
                        'destination': item['destination'],
                        'pallet_count': 0,
                        'pallet_ids': [],
                        'shipment_id': item['shipment_id']
                    }
                
                sub_grouped[sub_key]['pallet_count'] += 1
                sub_grouped[sub_key]['pallet_ids'].extend(item['pallet_ids'])
            
            # жЫіжЦ∞зїДзЪДitemsдЄЇе≠РеИЖзїДзїУжЮЬ
            group['items'] = list(sub_grouped.values())
        
        context['search_results'] = list(grouped_results.values())
        return self.template_fleet_po_check, context

    async def handle_fleet_po_delete_post(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        """иљ¶йШЯPOж†ЄеѓєеИ†йЩ§еКЯиГљ"""
        context = {}
        
        # иОЈеПЦеИ†йЩ§еПВжХ∞
        pallet_ids_str = request.POST.get('pallet_ids', '').strip()
        shipment_id = request.POST.get('shipment_id', '').strip()
        actual_pallets = int(request.POST.get('actual_pallets', 0))
        
        # дњЭе≠ШеОЯжЬЙзЪДжРЬзіҐжЭ°дїґ
        fleet_number = request.POST.get('fleet_number', '').strip()
        pickup_number = request.POST.get('pickup_number', '').strip()
        container_number = request.POST.get('container_number', '').strip()
        destination = request.POST.get('destination', '').strip()
        shipment_batch_number = request.POST.get('shipment_batch_number', '').strip()
        isa = request.POST.get('isa', '').strip()
        
        if not pallet_ids_str or not shipment_id:
            context['error_messages'] = 'зЉЇе∞СењЕи¶БзЪДеПВжХ∞'
            return self.template_fleet_po_check, context
        
        # иІ£жЮРpallet_ids
        pallet_ids = [int(pid.strip()) for pid in pallet_ids_str.split(',') if pid.strip()]
        
        if actual_pallets < 0 or actual_pallets > len(pallet_ids):
            context['error_messages'] = f'еЃЮйЩЕжЭњжХ∞ењЕй°їеЬ®0-{len(pallet_ids)}дєЛйЧі'
            return self.template_fleet_po_check, context
        
        # иЃ°зЃЧйЬАи¶БеИ†йЩ§зЪДpalletжХ∞йЗП
        pallets_to_delete = len(pallet_ids) - actual_pallets
        
        if pallets_to_delete > 0:
            # иОЈеПЦеЙНpallets_to_deleteдЄ™palletзЪДID
            pallet_ids_to_delete = pallet_ids[:pallets_to_delete]
            # е∞ЖињЩдЇЫpalletзЪДshipment_batch_numberиЃЊзљЃдЄЇз©Ї
            await sync_to_async(Pallet.objects.filter(id__in=pallet_ids_to_delete).update)(
                shipment_batch_number=None,
                is_dropped_pallet=True
            )
            
            context['success_messages'] = f'жИРеКЯеИ†йЩ§{pallets_to_delete}дЄ™жЙШзЫШиЃ∞ељХ'
        else:
            context['success_messages'] = 'жЧ†йЬАеИ†йЩ§дїїдљХжЙШзЫШиЃ∞ељХ'
        
        # йЗНжЦ∞и∞ГзФ®жРЬзіҐеКЯиГљдї•еИЈжЦ∞й°µйЭҐпЉМеєґдЉ†йАТеОЯжЬЙзЪДжРЬзіҐжЭ°дїґ
        # еИЫеїЇдЄАдЄ™жЦ∞зЪДHttpRequestеѓєи±°жЭ•ж®°жЛЯжРЬзіҐиѓЈж±В
        from django.http import QueryDict
        
        # жЮДеїЇеМЕеРЂеОЯжЬЙжРЬзіҐжЭ°дїґзЪДPOSTжХ∞жНЃ
        post_data = QueryDict(mutable=True)
        post_data.update({
            'step': 'fleet_po_search',
            'fleet_number': fleet_number,
            'pickup_number': pickup_number,
            'container_number': container_number,
            'destination': destination,
            'shipment_batch_number': shipment_batch_number,
            'isa': isa
        })
        
        # еИЫеїЇжЦ∞зЪДиѓЈж±Веѓєи±°
        new_request = HttpRequest()
        new_request.method = 'POST'
        new_request.POST = post_data
        new_request.FILES = request.FILES
        new_request.user = request.user
        # жПТеЕ•жХ∞жНЃ йЗНжЦ∞еИЖжСКиљ¶жђ°жЭње≠РжХ∞дї•еПКжИРжЬђ
        fm = FleetManagement()
        await fm.insert_fleet_shipment_pallet(request, fleet_number)
        # и∞ГзФ®жРЬзіҐеКЯиГљ
        return await self.handle_fleet_po_search_post(new_request)

    async def get_used_pallets(self, shipment):
        """еЉВж≠•иЃ°зЃЧйҐДзЇ¶еЈ≤дљњзФ®зЪДжЭњжХ∞ - йАВйЕНжЦ∞зЪДжХ∞жНЃзїУжЮД"""
        # ж†єжНЃдљ†зЪДжХ∞жНЃзїУжЮДпЉМињЩйЗМйЬАи¶БиЃ°зЃЧиѓ•йҐДзЇ¶еЈ≤зїПдљњзФ®зЪДжЭњжХ∞
        # зФ±дЇОдљ†зЪДжХ∞жНЃзїУжЮДдЄ≠ж≤°жЬЙзЫіжО•еЕ≥иБФпЉМињЩйЗМеПѓиГљйЬАи¶Бж†єжНЃеЃЮйЩЕжГЕеЖµи∞ГжХі
        
        # дЄіжЧґеЃЮзО∞пЉЪеБЗиЃЊдїОеЕ≥иБФзЪДиЃҐеНХдЄ≠иЃ°зЃЧ
        try:
            # иОЈеПЦиѓ•йҐДзЇ¶еЕ≥иБФзЪДжЙАжЬЙиіІзЙ©
            related_packing_lists = await sync_to_async(list)(
                PackingList.objects.filter(
                    shipment_batch_number=shipment
                )
            )
            
            total_pallets = 0
            for pl in related_packing_lists:
                total_pallets += pl.n_pallet or 0
            
            return total_pallets
        except:
            return 0
    
    def _parse_ts(self, ts: str, tzinfo: str) -> str:
        if ts:
            if isinstance(ts, str):
                ts_naive = datetime.fromisoformat(ts)
            else:
                ts_naive = ts.replace(tzinfo=None)
            tz = pytz.timezone(tzinfo)
            ts = tz.localize(ts_naive).astimezone(timezone.utc)
            return ts.strftime("%Y-%m-%d %H:%M:%S")
        else:
            return None
        
    async def _user_authenticate(self, request: HttpRequest):
        if await sync_to_async(lambda: request.user.is_authenticated)():
            return True
        return False
    
    async def _validate_user_four_major_whs(self, user: User) -> bool:       
        return await sync_to_async(
            lambda: user.groups.filter(name="four_major_whs").exists()
        )()

    async def handle_add_po_query_plt(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        """жЯ•иѓҐжЯЬеПЈдїУзВєеѓєеЇФзЪДpalletиЃ∞ељХ"""
        context = {}
        
        query_container_number = request.POST.get('query_container_number', '').strip().upper()
        query_destination = request.POST.get('query_destination', '').strip().upper()
        
        # дњЭе≠ШжЯ•иѓҐжЭ°дїґеИ∞еЙНзЂѓ
        context.update({
            'query_container_number': query_container_number,
            'query_destination': query_destination
        })
        
        # жЮДеїЇжЯ•иѓҐжЭ°дїґ
        criteria = Q()
        if query_container_number:
            criteria &= Q(container_number__container_number__icontains=query_container_number)
        if query_destination:
            criteria &= Q(destination__icontains=query_destination)
        
        try:
            # жЯ•иѓҐpalletжХ∞жНЃпЉМжМЙidеИЖзїД
            pallets = await sync_to_async(list)(
                Pallet.objects.select_related('container_number', 'shipment_batch_number')
                .filter(criteria)
                .values('id', 'container_number__container_number', 'destination', 'shipment_batch_number__shipment_batch_number', 'shipment_batch_number__appointment_id')
                .order_by('id')
            )
            
            # е§ДзРЖжХ∞жНЃпЉМзІїйЩ§statusзКґжАБ
            processed_pallets = []
            for pallet in pallets:
                processed_pallets.append({
                    'id': pallet['id'],
                    'container_number': pallet.get('container_number__container_number', ''),
                    'destination': pallet.get('destination', ''),
                    'shipment_batch_number': pallet.get('shipment_batch_number__shipment_batch_number', ''),
                    'appointment_id': pallet.get('shipment_batch_number__appointment_id', '')
                })
            
            context['query_pallets'] = processed_pallets
            context['show_pallet_modal'] = True
            
        except Exception as e:
            context['error_messages'] = f'жЯ•иѓҐе§±иі•: {str(e)}'
        request.POST._mutable = True
        shipment_value = request.POST.pop('search_shipment_batch_number')[0]
        request.POST['shipment_batch_number'] = shipment_value
        isa_value = request.POST.pop('search_isa')[0]
        request.POST['isa'] = isa_value
        request.POST._mutable = False

        return await self.handle_fleet_po_search_post(request,context)

    async def handle_add_pallets_to_shipment(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        """е∞ЖйАЙдЄ≠зЪДpalletиЃ∞ељХеЕ≥иБФеИ∞shipment"""
        context = {}
        # иОЈеПЦеПВжХ∞
        selected_pallet_ids_str = request.POST.get('selected_pallet_ids', '').strip()
        shipment_batch_number = request.POST.get('shipment_batch_number', '').strip()
        isa = request.POST.get('isa', '').strip()
        
        if not selected_pallet_ids_str:
            context['error_messages'] = 'зЉЇе∞СйАЙдЄ≠зЪДpalletиЃ∞ељХ'
            return self.template_fleet_po_check, context
        
        # иІ£жЮРpallet IDs
        selected_pallet_ids = [int(pid.strip()) for pid in selected_pallet_ids_str.split(',') if pid.strip()]
        
        # ж†єжНЃжРЬзіҐжЭ°дїґжЯ•жЙЊеѓєеЇФзЪДshipment
        shipment_criteria = Q()
        if shipment_batch_number:
            shipment_criteria &= Q(shipment_batch_number__icontains=shipment_batch_number)
        elif isa:
            shipment_criteria &= Q(appointment_id__icontains=isa)
        # жЯ•жЙЊshipment
        target_shipment = await sync_to_async(Shipment.objects.get)(shipment_criteria)
        
        if not target_shipment:
            context['error_messages'] = 'жЬ™жЙЊеИ∞еѓєеЇФзЪДйҐДзЇ¶жЙєжђ°'
            return self.template_fleet_po_check, context
        
        # жЫіжЦ∞йАЙдЄ≠зЪДpalletиЃ∞ељХзЪДshipment_batch_number
        updated_count = await sync_to_async(Pallet.objects.filter(id__in=selected_pallet_ids).update)(
            shipment_batch_number=target_shipment
        )
        
        context['success_messages'] = f'жИРеКЯеЕ≥иБФ{updated_count}дЄ™жЭње≠РиЃ∞ељХеИ∞йҐДзЇ¶жЙєжђ°'
        
        # йЗНжЦ∞и∞ГзФ®жРЬзіҐеКЯиГљдї•еИЈжЦ∞й°µйЭҐпЉМеєґдЉ†йАТеОЯжЬЙзЪДжРЬзіҐжЭ°дїґ
        from django.http import QueryDict
        
        # жЮДеїЇеМЕеРЂеОЯжЬЙжРЬзіҐжЭ°дїґзЪДPOSTжХ∞жНЃ
        post_data = QueryDict(mutable=True)
        post_data.update({
            'step': 'fleet_po_search',
            'shipment_batch_number': shipment_batch_number,
            'isa': isa,
        })
        
        # еИЫеїЇжЦ∞зЪДиѓЈж±Веѓєи±°
        new_request = HttpRequest()
        new_request.method = 'POST'
        new_request.POST = post_data
        new_request.FILES = request.FILES
        new_request.user = request.user

        # жЙЊеИ∞иљ¶жђ°дњ°жБѓпЉМйЗНжЦ∞еИЖжСК
        fleet = target_shipment.fleet_number
        fm = FleetManagement()
        await fm.insert_fleet_shipment_pallet_fleet_cost(
                request, fleet.fleet_number, fleet.fleet_cost
            )
        # и∞ГзФ®жРЬзіҐеКЯиГљ
        return await self.handle_fleet_po_search_post(new_request, context)

    async def handle_client_exception_search_post(self, request: HttpRequest, context: str | None = None) -> tuple[str, dict[str, Any]]:
        """еЃҐжИЈзЂѓеЉВеЄЄиѓ¶жГЕжРЬзіҐеКЯиГљ"""  
        if not context:
            context = {}
        
        # иОЈеПЦжРЬзіҐжЭ°дїґ
        start_date = request.POST.get('start_date', '').strip()
        end_date = request.POST.get('end_date', '').strip()
        container_number = request.POST.get('container_number', '').strip()
        destination = request.POST.get('destination', '').strip()
        warehouse = request.POST.get('warehouse', '').strip()
        delivery_type = request.POST.get('delivery_type', 'public').strip()

        # е∞ЖжРЬзіҐжЭ°дїґдЉ†йАТеИ∞еЙНзЂѓ
        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'container_number': container_number,
            'destination': destination,
            'warehouse': warehouse,
            'delivery_type': delivery_type
        })
        
        # е¶ВжЮЬж≤°жЬЙиЊУеЕ•дїїдљХеАЉпЉМйїШиЃ§жЬАињСдЄЙдЄ™жЬИ
        if not any([start_date, end_date, container_number, destination, warehouse]):
            today = datetime.now().date()
            three_months_ago = today - timedelta(days=90)
            start_date = three_months_ago.strftime('%Y-%m-%d')
            end_date = today.strftime('%Y-%m-%d')
            context['start_date'] = start_date
            context['end_date'] = end_date
        
        # жЮДеїЇзїЯдЄАзЪДжЯ•иѓҐжЭ°дїґ
        base_criteria = Q()
        
        # жЧґйЧіжЯ•иѓҐ - йАЪињЗcontainer_number->orders->offload_id->offload_at
        if start_date:
            base_criteria &= Q(container_number__orders__offload_id__offload_at__gte=start_date)
        if end_date:
            base_criteria &= Q(container_number__orders__offload_id__offload_at__lte=end_date)
        
        # жЯЬеПЈжЯ•иѓҐ
        if container_number:
            base_criteria &= Q(container_number__container_number__icontains=container_number)
        
        # дїУзВєжЯ•иѓҐ
        if destination:
            base_criteria &= Q(destination__icontains=destination)
        
        # дїУеЇУжЯ•иѓҐ - йАЪињЗcontainer_number->orders->retrieval_id->retrieval_destination_area
        if warehouse:
            base_criteria &= Q(container_number__orders__retrieval_id__retrieval_destination_area=warehouse)
        
        base_criteria &= Q(delivery_type=delivery_type)

        # еИЖеИЂжЮДеїЇPalletеТМPackingListзЪДжЯ•иѓҐжЭ°дїґ
        pallet_criteria = base_criteria & Q(container_number__orders__offload_id__offload_at__isnull=False)
        packinglist_criteria = base_criteria & Q(container_number__orders__offload_id__offload_at__isnull=True)
        
        results = []
        combined_dict = {}
        
        # дљњзФ® values() еТМ annotate() еЬ®жХ∞жНЃеЇУе±ВйЭҐеИЖзїДжЯ•иѓҐ Pallet жХ∞жНЃ
        pallet_groups = await sync_to_async(list)(
            Pallet.objects
            .filter(pallet_criteria)
            .select_related('container_number', 'master_shipment_batch_number')
            .values(
                'PO_ID',
                'container_number__container_number',
                'destination',
                'master_shipment_batch_number_id',
                'master_shipment_batch_number__shipment_batch_number',
                'master_shipment_batch_number__appointment_id',
                'master_shipment_batch_number__pod_link',
                'master_shipment_batch_number__arrived_at',
                'master_shipment_batch_number__shipped_at',
                'master_shipment_batch_number__shipment_appointment',
                'master_shipment_batch_number__is_virtual_sp',
            )
            .annotate(pallet_count=Count('id'))
            .order_by('PO_ID', 'container_number__container_number', 'destination')
        )
        
        # е§ДзРЖ Pallet еИЖзїДзїУжЮЬ
        for group in pallet_groups:
            po_id = group['PO_ID'] or ''
            container_num = group['container_number__container_number'] or ''
            dest = group['destination'] or ''
            
            if po_id and container_num and dest:
                key = f"{po_id}_{container_num}_{dest}"
            elif container_num and dest:
                key = f"{container_num}_{dest}"
            else:
                key = f"pallet_{po_id}_{container_num}_{dest}"
            
            if key not in combined_dict:
                combined_dict[key] = self._process_master_shipment_item(group, 'pallet')
                combined_dict[key]['ids'] = []
        
        # жЯ•иѓҐжЙАжЬЙ Pallet зЪД idпЉМзФ®дЇОжФґйЫЖ ids
        pallet_ids = await sync_to_async(list)(
            Pallet.objects
            .filter(pallet_criteria)
            .values(
                'id',
                'PO_ID',
                'container_number__container_number',
                'destination'
            )
        )
        
        for pallet in pallet_ids:
            po_id = pallet['PO_ID'] or ''
            container_num = pallet['container_number__container_number'] or ''
            dest = pallet['destination'] or ''
            
            if po_id and container_num and dest:
                key = f"{po_id}_{container_num}_{dest}"
            elif container_num and dest:
                key = f"{container_num}_{dest}"
            else:
                key = f"pallet_{po_id}_{container_num}_{dest}"
            
            if key in combined_dict:
                combined_dict[key]['ids'].append(f"plt_{pallet['id']}")
        
        # дљњзФ® values() еТМ annotate() еЬ®жХ∞жНЃеЇУе±ВйЭҐеИЖзїДжЯ•иѓҐ PackingList жХ∞жНЃ
        packinglist_groups = await sync_to_async(list)(
            PackingList.objects
            .filter(packinglist_criteria)
            .select_related('container_number', 'master_shipment_batch_number')
            .values(
                'PO_ID',
                'container_number__container_number',
                'destination',
                'master_shipment_batch_number_id',
                'master_shipment_batch_number__shipment_batch_number',
                'master_shipment_batch_number__appointment_id',
                'master_shipment_batch_number__pod_link',
                'master_shipment_batch_number__arrived_at',
                'master_shipment_batch_number__shipped_at',
                'master_shipment_batch_number__shipment_appointment',
                'master_shipment_batch_number__is_virtual_sp',
            )
            .order_by('PO_ID', 'container_number__container_number', 'destination')
        )
        
        # жЯ•иѓҐжЙАжЬЙ PackingList зЪД id еТМ cbmпЉМзФ®дЇОиЃ°зЃЧжЭњжХ∞еТМжФґйЫЖ ids
        packinglist_details = await sync_to_async(list)(
            PackingList.objects
            .filter(packinglist_criteria)
            .values(
                'id',
                'PO_ID',
                'container_number__container_number',
                'destination',
                'cbm'
            )
        )
        
        # е§ДзРЖ PackingList еИЖзїДзїУжЮЬ
        for pl in packinglist_details:
            po_id = pl['PO_ID'] or ''
            container_num = pl['container_number__container_number'] or ''
            dest = pl['destination'] or ''
            
            if po_id and container_num and dest:
                key = f"{po_id}_{container_num}_{dest}"
            elif container_num and dest:
                key = f"{container_num}_{dest}"
            else:
                key = f"packinglist_{po_id}_{container_num}_{dest}"
            
            if key not in combined_dict:
                # дїОеИЖзїДдЄ≠жЯ•жЙЊеЃМжХізЪДдњ°жБѓ
                item_dict = None
                for group in packinglist_groups:
                    if (group['PO_ID'] == pl['PO_ID'] and
                        group['container_number__container_number'] == container_num and
                        group['destination'] == dest):
                        item_dict = group
                        break
                if not item_dict:
                    item_dict = {
                        'PO_ID': pl['PO_ID'],
                        'container_number__container_number': container_num,
                        'destination': dest,
                        'master_shipment_batch_number_id': None,
                    }
                combined_dict[key] = self._process_master_shipment_item(item_dict, 'packinglist')
                combined_dict[key]['ids'] = []
            
            combined_dict[key]['ids'].append(f"pl_{pl['id']}")
        
        # иљђжНҐidsдЄЇе≠Чзђ¶дЄ≤пЉМеєґжЈїеК†еЕґдїЦе≠ЧжЃµ
        for key, value in combined_dict.items():
            value['ids_string'] = ','.join(value['ids'])
            value['has_master_shipment'] = bool(value.get('shipment_batch_number'))
        
        # жФґйЫЖжЙАжЬЙ pallet_idпЉМзФ®дЇОжЯ•иѓҐеЉВеЄЄ
        all_pallet_ids = []
        for key, value in combined_dict.items():
            for item_id in value['ids']:
                if item_id.startswith('plt_'):
                    all_pallet_ids.append(int(item_id.replace('plt_', '')))
        
        # жЯ•иѓҐжЙАжЬЙзЫЄеЕ≥зЪДеЉВеЄЄ
        if all_pallet_ids:
            from django.apps import apps
            try:
                PalletException = apps.get_model('warehouse', 'PalletException')
                exceptions = await sync_to_async(list)(
                    PalletException.objects
                    .filter(pallet_id__in=all_pallet_ids)
                    .select_related('pallet')
                    .order_by('-created_at')
                )
                # еИЫеїЇдЄАдЄ™е≠ЧеЕЄпЉМжМЙ pallet id еИЖзїД
                exception_dict = {}
                for exc in exceptions:
                    if exc.pallet_id not in exception_dict:
                        exception_dict[exc.pallet_id] = []
                    exception_dict[exc.pallet_id].append(exc)
                
                # е∞ЖеЉВеЄЄжЈїеК†еИ∞еѓєеЇФзЪДжХ∞жНЃй°єдЄ≠
                for key, value in combined_dict.items():
                    item_exceptions = []
                    for item_id in value['ids']:
                        if item_id.startswith('plt_'):
                            pallet_id = int(item_id.replace('plt_', ''))
                            if pallet_id in exception_dict:
                                item_exceptions.extend(exception_dict[pallet_id])
                    value['exceptions'] = item_exceptions
            except Exception:
                pass
        
        context['results'] = list(combined_dict.values())
        return self.template_client_exception, context

    async def handle_load_exceptions_post(self, request: HttpRequest) -> JsonResponse:
        """еК†иљљеЉВеЄЄеИЧи°®"""
        try:
            ids_string = request.POST.get('ids_string', '').strip()
            if not ids_string:
                return JsonResponse({'success': False, 'message': 'ж≤°жЬЙйАЙжЛ©дїїдљХиЃ∞ељХ'})
            
            # иІ£жЮР idsпЉМжПРеПЦ pallet_id
            pallet_ids = []
            for item_id in ids_string.split(','):
                if item_id.startswith('plt_'):
                    pallet_ids.append(int(item_id.replace('plt_', '')))
            
            if not pallet_ids:
                return JsonResponse({'success': True, 'exceptions': []})
            
            # жЯ•иѓҐеЉВеЄЄ
            from django.apps import apps
            PalletException = apps.get_model('warehouse', 'PalletException')
            exceptions = await sync_to_async(list)(
                PalletException.objects
                .filter(pallet_id__in=pallet_ids)
                .select_related('pallet')
                .order_by('-created_at')
            )
            
            # еЇПеИЧеМЦеЉВеЄЄжХ∞жНЃ
            exception_list = []
            for exc in exceptions:
                exception_list.append({
                    'id': exc.id,
                    'exception_type': exc.exception_type,
                    'exception_type_display': exc.get_exception_type_display(),
                    'exception_reason': exc.exception_reason,
                    'created_at': exc.created_at.strftime('%Y-%m-%d %H:%M:%S') if exc.created_at else ''
                })
            
            return JsonResponse({'success': True, 'exceptions': exception_list})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    async def handle_add_exception_post(self, request: HttpRequest) -> JsonResponse:
        """жЈїеК†еЉВеЄЄ"""
        try:
            ids_string = request.POST.get('ids_string', '').strip()
            exception_type = request.POST.get('exception_type', '').strip()
            exception_reason = request.POST.get('exception_reason', '').strip()
            
            if not ids_string:
                return JsonResponse({'success': False, 'message': 'ж≤°жЬЙйАЙжЛ©дїїдљХиЃ∞ељХ'})
            if not exception_type:
                return JsonResponse({'success': False, 'message': 'иѓЈйАЙжЛ©еЉВеЄЄз±їеЮЛ'})
            if not exception_reason:
                return JsonResponse({'success': False, 'message': 'иѓЈиЊУеЕ•еЉВеЄЄеОЯеЫ†'})
            
            # иІ£жЮР idsпЉМжПРеПЦ pallet_id
            pallet_ids = []
            for item_id in ids_string.split(','):
                if item_id.startswith('plt_'):
                    pallet_ids.append(int(item_id.replace('plt_', '')))
            
            if not pallet_ids:
                return JsonResponse({'success': False, 'message': 'ж≤°жЬЙеПѓжЈїеК†еЉВеЄЄзЪДиЃ∞ељХ'})
            
            # еѓЉеЕ•ж®°еЮЛ
            from django.apps import apps
            PalletException = apps.get_model('warehouse', 'PalletException')
            
            # дЄЇжѓПдЄ™ pallet жЈїеК†еЉВеЄЄ
            for pallet_id in pallet_ids:
                await sync_to_async(PalletException.objects.create)(
                    pallet_id=pallet_id,
                    exception_type=exception_type,
                    exception_reason=exception_reason
                )
            
            return JsonResponse({'success': True, 'message': 'жЈїеК†жИРеКЯ'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    async def handle_delete_exception_post(self, request: HttpRequest) -> JsonResponse:
        """еИ†йЩ§еЉВеЄЄ"""
        try:
            exception_id = request.POST.get('exception_id', '').strip()
            if not exception_id:
                return JsonResponse({'success': False, 'message': 'иѓЈйАЙжЛ©и¶БеИ†йЩ§зЪДеЉВеЄЄ'})
            
            from django.apps import apps
            PalletException = apps.get_model('warehouse', 'PalletException')
            await sync_to_async(PalletException.objects.filter(id=int(exception_id)).delete)()
            
            return JsonResponse({'success': True, 'message': 'еИ†йЩ§жИРеКЯ'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})


    async def handle_master_shipment_check_post(self, request: HttpRequest, context: str | None = None) -> tuple[str, dict[str, Any]]:
        """еЃҐжИЈзЂѓзЇ¶иѓ¶жГЕжРЬзіҐеКЯиГљ"""  
        if not context:
            context = {}
        
        # иОЈеПЦжРЬзіҐжЭ°дїґ
        start_date = request.POST.get('start_date', '').strip()
        end_date = request.POST.get('end_date', '').strip()
        container_number = request.POST.get('container_number', '').strip()
        destination = request.POST.get('destination', '').strip()
        warehouse = request.POST.get('warehouse', '').strip()
        delivery_type = request.POST.get('delivery_type', 'public').strip()
        pending_shipment = request.POST.get('pending_shipment', 'True').strip()
        
        # е∞ЖжРЬзіҐжЭ°дїґдЉ†йАТеИ∞еЙНзЂѓ
        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'container_number': container_number,
            'destination': destination,
            'warehouse': warehouse,
            'delivery_type': delivery_type,
            'pending_shipment': pending_shipment
        })
        
        # е¶ВжЮЬж≤°жЬЙиЊУеЕ•дїїдљХеАЉпЉМйїШиЃ§жЬАињСдЄЙдЄ™жЬИ
        if not any([start_date, end_date, container_number, destination, warehouse]):
            today = datetime.now().date()
            three_months_ago = today - timedelta(days=90)
            start_date = three_months_ago.strftime('%Y-%m-%d')
            end_date = today.strftime('%Y-%m-%d')
            context['start_date'] = start_date
            context['end_date'] = end_date
        
        # жЮДеїЇзїЯдЄАзЪДжЯ•иѓҐжЭ°дїґ
        base_criteria = Q()
        
        # жЧґйЧіжЯ•иѓҐ - йАЪињЗcontainer_number->orders->offload_id->offload_at
        if start_date:
            base_criteria &= Q(container_number__orders__offload_id__offload_at__gte=start_date)
        if end_date:
            base_criteria &= Q(container_number__orders__offload_id__offload_at__lte=end_date)
        
        # жЯЬеПЈжЯ•иѓҐ
        if container_number:
            base_criteria &= Q(container_number__container_number__icontains=container_number)
        
        # дїУзВєжЯ•иѓҐ
        if destination:
            base_criteria &= Q(destination__icontains=destination)
        
        # дїУеЇУжЯ•иѓҐ - йАЪињЗcontainer_number->orders->retrieval_id->retrieval_destination_area
        if warehouse:
            base_criteria &= Q(container_number__orders__retrieval_id__retrieval_destination_area=warehouse)
        
        base_criteria &= Q(delivery_type=delivery_type)

        # еЊЕе§ДзРЖзЇ¶з≠ЫйАЙжЭ°дїґ
        if pending_shipment == 'True':
            # shipment_batch_numberдЄНдЄЇз©Ї
            base_criteria &= Q(shipment_batch_number__isnull=False)
            
            # еєґдЄФ (master_shipment_batch_numberдЄЇз©Ї жИЦиАЕ (is_virtual_spдЄЇзЬЯ еєґдЄФ master_shipment_batch_numberе§ЦйФЃжМЗеРСзЪДshipmentзЪДpod_linkдЄЇз©Ї))
            base_criteria &= (
                Q(master_shipment_batch_number__isnull=True) |
                (
                    Q(master_shipment_batch_number__is_virtual_sp=True) &
                    Q(master_shipment_batch_number__pod_link__isnull=True)
                )
            )

        # еИЖеИЂжЮДеїЇPalletеТМPackingListзЪДжЯ•иѓҐжЭ°дїґ
        pallet_criteria = base_criteria & Q(container_number__orders__offload_id__offload_at__isnull=False)
        packinglist_criteria = base_criteria & Q(container_number__orders__offload_id__offload_at__isnull=True)
        
        results = []
        
        try:
            # жЯ•иѓҐPalletжХ∞жНЃ
            pallets = await sync_to_async(list)(
                Pallet.objects.select_related(
                    'container_number',
                    'master_shipment_batch_number'
                )
                .filter(pallet_criteria)
            )
            
            # жЯ•иѓҐPackingListжХ∞жНЃ
            packinglists = await sync_to_async(list)(
                PackingList.objects.select_related(
                    'container_number',
                    'master_shipment_batch_number'
                )
                .filter(packinglist_criteria)
            )
            
            # е∞ЖдЄ§дЄ™еИЧи°®еРИеєґпЉМж†ЗиЃ∞жЇРз±їеЮЛ
            all_items = []
            for pallet in pallets:
                all_items.append({
                    'item': pallet,
                    'type': 'pallet'
                })
            for packinglist in packinglists:
                all_items.append({
                    'item': packinglist,
                    'type': 'packinglist'
                })
            
            # зїЯдЄАе§ДзРЖеИЖзїД
            combined_dict = {}
            for item_data in all_items:
                item = item_data['item']
                item_type = item_data['type']
                
                # иОЈеПЦеИЖзїДkey
                po_id = item.PO_ID or ''
                container_number = item.container_number.container_number if item.container_number else ''
                destination = item.destination or ''
                master_id = item.master_shipment_batch_number_id
                
                if master_id:
                    key = f"master_{master_id}"
                elif po_id and container_number and destination:
                    key = f"{po_id}_{container_number}_{destination}"
                elif container_number and destination:
                    key = f"{container_number}_{destination}"
                else:
                    key = f"{item_type}_{item.id}"
                
                if key not in combined_dict:
                    # иљђжНҐдЄЇе≠ЧеЕЄж†ЉеЉП
                    item_dict = {
                        'PO_ID': item.PO_ID,
                        'master_shipment_batch_number_id': item.master_shipment_batch_number_id,
                        'container_number__container_number': container_number,
                        'destination': destination,
                        'master_shipment_batch_number__shipment_batch_number': item.master_shipment_batch_number.shipment_batch_number if item.master_shipment_batch_number else None,
                        'master_shipment_batch_number__appointment_id': item.master_shipment_batch_number.appointment_id if item.master_shipment_batch_number else None,
                        'master_shipment_batch_number__pod_link': item.master_shipment_batch_number.pod_link if item.master_shipment_batch_number else None,
                        'master_shipment_batch_number__pod_uploaded_at': item.master_shipment_batch_number.pod_uploaded_at if item.master_shipment_batch_number else None,
                        'master_shipment_batch_number__arrived_at': item.master_shipment_batch_number.arrived_at if item.master_shipment_batch_number else None,
                        'master_shipment_batch_number__shipped_at': item.master_shipment_batch_number.shipped_at if item.master_shipment_batch_number else None,
                        'master_shipment_batch_number__shipment_appointment': item.master_shipment_batch_number.shipment_appointment if item.master_shipment_batch_number else None,
                        'master_shipment_batch_number__is_virtual_sp': item.master_shipment_batch_number.is_virtual_sp if item.master_shipment_batch_number else False
                    }
                    combined_dict[key] = self._process_master_shipment_item(item_dict, item_type)
                    combined_dict[key]['pallet_count'] = 0
                    combined_dict[key]['ids'] = []
                
                # иЃ°зЃЧжЭњжХ∞
                if item_type == 'pallet':
                    combined_dict[key]['pallet_count'] += 1
                    combined_dict[key]['ids'].append(f"plt_{item.id}")
                else:
                    total_cbm = item.cbm or 0
                    pl_count = math.ceil(total_cbm / 1.8) if total_cbm else 0
                    if pl_count > combined_dict[key]['pallet_count']:
                        combined_dict[key]['pallet_count'] = pl_count
                    combined_dict[key]['ids'].append(f"pl_{item.id}")
            
            # иљђжНҐidsдЄЇе≠Чзђ¶дЄ≤пЉМеєґжЈїеК†еЕґдїЦе≠ЧжЃµ
            for key, value in combined_dict.items():
                value['ids_string'] = ','.join(value['ids'])
                value['has_master_shipment'] = bool(value.get('shipment_batch_number'))
            
            results = list(combined_dict.values())
            
        except Exception as e:
            context['error_messages'] = f'жЯ•иѓҐе§±иі•: {str(e)}'
        
        context['results'] = results
        return self.template_master_shipment_check, context
    
    async def handle_create_fictional_master_post(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        """еИЫеїЇиЩЪжЮДдЄїзЇ¶"""      
        context = {}
        # иОЈеПЦи°®еНХжХ∞жНЃ
        ids_string = request.POST.get('ids_string', '')
        appointment_type = request.POST.get('appointment_type', '')
        appointment_account = request.POST.get('appointment_account', '')
        appointment_number = request.POST.get('appointment_number', '')
        scheduled_time = request.POST.get('scheduled_time', '')
        pickup_time = request.POST.get('pickup_time', '')
        shipping_warehouse = request.POST.get('shipping_warehouse', '')
        loading_type = request.POST.get('loading_type', '')
        pickup_number = request.POST.get('pickup_number', '')
        destination = request.POST.get('sp_destination', '')
        address = request.POST.get('address', '')
        note = request.POST.get('note', '')
        shipment_id = request.POST.get('shipment_id', '')
        
        # еЕИж£АжЯ•дЄАдЄЛињЩдЄ™ISAжЬЙж≤°жЬЙзФ®ињЗ
        if not appointment_number and appointment_type == "FTL":
            # ж≤°жЬЙеАЉпЉМеИЫеїЇйЪПжЬЇзЪД11дљНжХ∞
            while True:
                # зФЯжИР11дљНйЪПжЬЇжХ∞пЉИзђђдЄАдљНдЄНиГљжШѓ0пЉЙ
                appointment_number = str(random.randint(10000000000, 99999999999))
                # ж£АжЯ•жШѓеР¶еЈ≤е≠ШеЬ®
                if not await sync_to_async(Shipment.objects.filter(appointment_id=appointment_number).exists)():
                    break
        else:
            # жЬЙеАЉпЉМж£АжЯ•жШѓеР¶еЈ≤е≠ШеЬ®
            if await sync_to_async(Shipment.objects.filter(appointment_id=appointment_number).exists)():
                template, search_context = await self.handle_master_shipment_check_post(request)
                context.update(search_context)
                return template, context

        # иЗ™еК®зФЯжИРжЙєжђ°еПЈ
        shipment_batch_number = await self.generate_unique_batch_number(destination)
        # еИЫеїЇжЦ∞зЪДShipmentиЃ∞ељХ
        new_shipment = Shipment()
        
        new_shipment.shipment_type = appointment_type
        new_shipment.shipment_account = appointment_account
        new_shipment.appointment_id = appointment_number
        new_shipment.shipment_schduled_at = timezone.now()
        new_shipment.shipment_appointment = scheduled_time
        if pickup_time:
            new_shipment.pickup_time = pickup_time
        new_shipment.origin = shipping_warehouse
        new_shipment.load_type = loading_type
        new_shipment.pickup_number = pickup_number
        new_shipment.destination = destination
        new_shipment.address = address
        new_shipment.note = note
        new_shipment.shipment_batch_number = shipment_batch_number
        new_shipment.shipment_cargo_id = shipment_id
        new_shipment.is_virtual_sp = True
        
        await sync_to_async(new_shipment.save)()
        
        # иІ£жЮРids_string
        ids_list = ids_string.split(',') if ids_string else []
        
        pallet_ids = []
        packinglist_ids = []
        
        for id_str in ids_list:
            if id_str.startswith('plt_'):
                pallet_ids.append(id_str.replace('plt_', ''))
            elif id_str.startswith('pl_'):
                packinglist_ids.append(id_str.replace('pl_', ''))

        def _update():
            updated_pallet = 0
            updated_packing = 0

            if pallet_ids and isinstance(pallet_ids, list):
                updated_pallet = Pallet.objects.filter(id__in=pallet_ids).update(
                    master_shipment_batch_number=new_shipment
                )

            if packinglist_ids and isinstance(packinglist_ids, list):
                updated_packing = PackingList.objects.filter(id__in=packinglist_ids).update(
                    master_shipment_batch_number=new_shipment
                )

            return updated_pallet, updated_packing

        updated_pallet, updated_packing = await sync_to_async(_update)()
        
        
        # йЗНжЦ∞и∞ГзФ®жРЬзіҐеКЯиГљпЉМдњЭзХЩеОЯжЬЙзЪДжРЬзіҐжЭ°дїґ
        context = {
            'success_messages': 'иЩЪжЮДдЄїзЇ¶еИЫеїЇжИРеКЯ!'
        }
                  
        # йЗНжЦ∞и∞ГзФ®жРЬзіҐеКЯиГљ
        template, search_context = await self.handle_master_shipment_check_post(request)
        context.update(search_context)
        
        return template, context
    
    async def handle_bind_existing_shipment_post(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        """poзїСеЃЪеЈ≤жЬЙзЪДзЇ¶дЄЇдЄїзЇ¶"""       
        context = {}
        # иОЈеПЦи°®еНХжХ∞жНЃ
        ids_string = request.POST.get('bind_ids_string', '')
        bind_shipment_batch_number = request.POST.get('bind_shipment_batch_number', '')
        bind_appointment_id = request.POST.get('bind_appointment_id', '')
        
        # жЯ•жЙЊзЫЃж†Зshipment
        target_shipment = None
        if bind_shipment_batch_number:
            target_shipment = await sync_to_async(Shipment.objects.get)(
                shipment_batch_number=bind_shipment_batch_number
            )
        elif bind_appointment_id:
            target_shipment = await sync_to_async(Shipment.objects.get)(
                appointment_id=bind_appointment_id
            )
        
        if not target_shipment:
            context['error_messages'] = 'жЬ™жЙЊеИ∞еѓєеЇФзЪДзЇ¶'
            template, search_context = await self.handle_master_shipment_check_post(request)
            context.update(search_context)
            return template, context
        
        # иІ£жЮРids_string
        ids_list = ids_string.split(',') if ids_string else []
        
        pallet_ids = []
        packinglist_ids = []
        po_ids = set()
        
        for id_str in ids_list:
            if id_str.startswith('plt_'):
                pallet_ids.append(id_str.replace('plt_', ''))
            elif id_str.startswith('pl_'):
                packinglist_ids.append(id_str.replace('pl_', ''))
        
        # жЫіжЦ∞PalletиЃ∞ељХ
        if pallet_ids:
            await sync_to_async(Pallet.objects.filter(id__in=pallet_ids).update)(
                master_shipment_batch_number=target_shipment
            )
            
            # иОЈеПЦињЩдЇЫpalletзЪДPO_ID
            pallets = await sync_to_async(list)(
                Pallet.objects.filter(id__in=pallet_ids)
            )
            for pallet in pallets:
                if pallet.PO_ID:
                    po_ids.add(pallet.PO_ID)
        
        # жЫіжЦ∞PackingListиЃ∞ељХ
        if packinglist_ids:
            await sync_to_async(PackingList.objects.filter(id__in=packinglist_ids).update)(
                master_shipment_batch_number=target_shipment
            )
        
        # е¶ВжЮЬжЬЙpalletзїСеЃЪдЇЖпЉМдєЯжККзЫЄеРМPO_IDзЪДpackinglistзїСеЃЪдЄК
        if po_ids:
            await sync_to_async(PackingList.objects.filter(PO_ID__in=po_ids).update)(
                master_shipment_batch_number=target_shipment
            )
        
        context['success_messages'] = 'зїСеЃЪжИРеКЯ!'

        
        # йЗНжЦ∞и∞ГзФ®жРЬзіҐеКЯиГљ
        template, search_context = await self.handle_master_shipment_check_post(request)
        context.update(search_context)
        
        return template, context
    
    async def handle_save_virtual_shipment_time_post(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        """дњЭе≠ШиЩЪжЮДзЇ¶зЪДеЗЇеЇУжЧґйЧіжИЦйАБиЊЊжЧґйЧі"""
        context = {}
        
        shipment_batch_number = request.POST.get('shipment_batch_number', '')
        time_type = request.POST.get('time_type', '')
        time_value = request.POST.get('time_value', '')
        
        try:
            # жЯ•жЙЊshipment
            shipment = await sync_to_async(Shipment.objects.get)(
                shipment_batch_number=shipment_batch_number
            )
            
            if not shipment:
                context['error_messages'] = 'жЬ™жЙЊеИ∞еѓєеЇФзЪДзЇ¶'
                template, search_context = await self.handle_master_shipment_check_post(request)
                context.update(search_context)
                return template, context
            
            # иІ£жЮРжЧґйЧі
            time_obj = datetime.strptime(time_value, '%Y-%m-%d')
            
            # жЫіжЦ∞еѓєеЇФе≠ЧжЃµ
            if time_type == 'shipped':
                shipment.shipped_at = time_obj
                shipment.is_shipped = True
            elif time_type == 'arrived':
                shipment.arrived_at = time_obj
                shipment.is_arrived = True
            
            await sync_to_async(shipment.save, thread_sensitive=True)()
            
            context['success_messages'] = 'дњЭе≠ШжИРеКЯ!'
            
        except Exception as e:
            context['error_messages'] = f'дњЭе≠Ше§±иі•: {str(e)}'
        
        # йЗНжЦ∞и∞ГзФ®жРЬзіҐеКЯиГљ
        template, search_context = await self.handle_master_shipment_check_post(request)
        context.update(search_context)
        
        return template, context
    
    async def handle_unbind_master_shipment_post(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        """иІ£зїСдЄїзЇ¶"""
        context = {}
        
        try:
            container_number = request.POST.get('unbind_container_number', '').strip()
            destination = request.POST.get('unbind_destination', '').strip()
            shipment_batch_number = request.POST.get('unbind_shipment_batch_number', '').strip()
            
            if not all([container_number, destination, shipment_batch_number]):
                context['error_messages'] = 'зЉЇе∞СењЕи¶БзЪДеПВжХ∞'
                template, search_context = await self.handle_master_shipment_check_post(request)
                context.update(search_context)
                return template, context
            
            # жЯ•жЙЊзЫЃж†З shipment
            target_shipment = await sync_to_async(Shipment.objects.get)(
                shipment_batch_number=shipment_batch_number
            )
            
            # жЯ•жЙЊзЫЄеЕ≥зЪД container
            container = await sync_to_async(Container.objects.get)(
                container_number=container_number
            )
            
            # жЫіжЦ∞ Pallet иЃ∞ељХ
            pallet_count = await sync_to_async(Pallet.objects.filter(
                container_number=container,
                destination=destination,
                master_shipment_batch_number=target_shipment
            ).update)(
                master_shipment_batch_number=None
            )
            
            # жЫіжЦ∞ PackingList иЃ∞ељХ
            packinglist_count = await sync_to_async(PackingList.objects.filter(
                container_number=container,
                destination=destination,
                master_shipment_batch_number=target_shipment
            ).update)(
                master_shipment_batch_number=None
            )
            
            # ж£АжЯ•жШѓеР¶ињШжЬЙеЕґдїЦиЃ∞ељХзїСеЃЪињЩдЄ™ shipment
            # ж£АжЯ• Pallet и°®пЉИеМЕжЛђ shipment_batch_number еТМ master_shipment_batch_numberпЉЙ
            pallet_has_binding = await sync_to_async(
                Pallet.objects.filter(
                    Q(shipment_batch_number=target_shipment) | Q(master_shipment_batch_number=target_shipment)
                ).exists
            )()
            
            # ж£АжЯ• PackingList и°®пЉИеМЕжЛђ shipment_batch_number еТМ master_shipment_batch_numberпЉЙ
            packinglist_has_binding = await sync_to_async(
                PackingList.objects.filter(
                    Q(shipment_batch_number=target_shipment) | Q(master_shipment_batch_number=target_shipment)
                ).exists
            )()
            
            # е¶ВжЮЬдЄ§дЄ™и°®йГљж≤°жЬЙиЃ∞ељХзїСеЃЪдЇЖ
            if not pallet_has_binding and not packinglist_has_binding:
                # еИЈжЦ∞дЄАдЄЛ target_shipment еѓєи±°
                target_shipment = await sync_to_async(Shipment.objects.get)(
                    shipment_batch_number=shipment_batch_number
                )
                
                if target_shipment.is_virtual_sp:
                    # иЩЪжЛЯиіІзЇ¶пЉМзЫіжО•еИ†йЩ§
                    await sync_to_async(target_shipment.delete)()
                    context['success_messages'] = f'иІ£зїСжИРеКЯпЉБеЈ≤иІ£зїС {pallet_count} дЄ™жЙШзЫШеТМ {packinglist_count} дЄ™и£ЕзЃ±еНХпЉМиѓ•иЩЪжЛЯиіІзЇ¶еЈ≤еИ†йЩ§'
                else:
                    # зЬЯеЃЮиіІзЇ¶пЉМиЃЊзљЃдЄЇеПЦжґИ
                    target_shipment.is_canceled = True
                    await sync_to_async(target_shipment.save)()
                    context['success_messages'] = f'иІ£зїСжИРеКЯпЉБеЈ≤иІ£зїС {pallet_count} дЄ™жЙШзЫШеТМ {packinglist_count} дЄ™и£ЕзЃ±еНХпЉМиѓ•иіІзЇ¶еЈ≤ж†ЗиЃ∞дЄЇеПЦжґИ'
            else:
                context['success_messages'] = f'иІ£зїСжИРеКЯпЉБеЈ≤иІ£зїС {pallet_count} дЄ™жЙШзЫШеТМ {packinglist_count} дЄ™и£ЕзЃ±еНХ'
            
        except Exception as e:
            context['error_messages'] = f'иІ£зїСе§±иі•: {str(e)}'
        
        # йЗНжЦ∞и∞ГзФ®жРЬзіҐеКЯиГљ
        template, search_context = await self.handle_master_shipment_check_post(request)
        context.update(search_context)
        
        return template, context
    
    async def handle_pod_reupload_post(self, request: HttpRequest, context: str | None = None) -> tuple[str, dict[str, Any]]:
        """PODйЗНжЦ∞дЄКдЉ†жРЬзіҐ"""
        start_date = request.POST.get('start_date', '').strip()
        end_date = request.POST.get('end_date', '').strip()
        shipment_batch_number = request.POST.get('shipment_batch_number', '').strip()
        appointment_id = request.POST.get('appointment_id', '').strip()

        if not any([start_date, end_date, shipment_batch_number, appointment_id]):
            # иЃЊзљЃжЧґйЧіиМГеЫідЄЇеЙНдЄАдЄ™жЬИ
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        
        context = {
            'start_date': start_date,
            'end_date': end_date,
            'shipment_batch_number': shipment_batch_number,
            'appointment_id': appointment_id,
            'search_performed': True,
            'results': [],
        }
        
        query = Q()
        
        if start_date:
            start_datetime = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            query &= Q(shipment_appointment__gte=start_datetime)
        
        if end_date:
            end_datetime = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
            query &= Q(shipment_appointment__lt=end_datetime)
        
        if shipment_batch_number:
            query &= Q(shipment_batch_number__icontains=shipment_batch_number)
        
        if appointment_id:
            query &= Q(appointment_id__icontains=appointment_id)
        
        if query:
            shipments = await sync_to_async(list)(
                Shipment.objects.filter(query).values(
                    'shipment_batch_number',
                    'appointment_id',
                    'shipment_appointment',
                    'pod_link',
                ).order_by('-shipment_appointment')
            )
            
            context['results'] = shipments
        else:
            context['error_messages'] = 'иѓЈиЗ≥е∞СиЊУеЕ•дЄАдЄ™жРЬзіҐжЭ°дїґ'
        
        return self.template_pod_reupload, context
    
    async def _delete_old_pod_file(self, shipment_batch_number: str) -> None:
        """еИ†йЩ§жЧІзЪДPODжЦЗдїґ - йҐДзХЩеЗљжХ∞пЉМеЖЕеЃєеЊЕи°•еЕЕ"""
        # TODO: и°•еЕЕеИ†йЩ§жЧІжЦЗдїґзЪДйАїиЊС
        pass
    
    async def handle_pod_reupload_upload_post(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        """PODйЗНжЦ∞дЄКдЉ†е§ДзРЖ"""
        context = {}
        
        shipment_batch_number = request.POST.get('shipment_batch_number', '').strip()
        
        if not shipment_batch_number:
            context['error_messages'] = 'зЉЇе∞СжЙєжђ°еПЈ'
            return self.template_pod_reupload, context
        
        if "file" not in request.FILES:
            context['error_messages'] = 'иѓЈйАЙжЛ©и¶БдЄКдЉ†зЪДжЦЗдїґ'
            return self.template_pod_reupload, context
        
        # еЕИеИ†йЩ§жЧІжЦЗдїґ
        await self._delete_old_pod_file(shipment_batch_number)
        
        # и∞ГзФ® FleetManagement з±їзЪДдЄКдЉ†жЦєж≥Х
        fm = FleetManagement()
        await fm.handle_pod_upload_post(request, 'post_nsop')
        
        context['success_messages'] = 'PODдњЃжФєжИРеКЯпЉБ'
        
        return await self.handle_pod_reupload_post(request, context)
    
    def _process_master_shipment_item(self, item, source_type):
        """е§ДзРЖеНХдЄ™master_shipmentй°єзЫЃпЉМиЃ°зЃЧзКґжАБеТМж†ЉеЉПеМЦжЧ•жЬЯ"""
        
        # иОЈеПЦshipmentзЫЄеЕ≥е≠ЧжЃµ
        shipment_batch_number = item.get('master_shipment_batch_number__shipment_batch_number', '')
        appointment_id = item.get('master_shipment_batch_number__appointment_id', '')
        pod_link = item.get('master_shipment_batch_number__pod_link', '')
        pod_uploaded_at = item.get('master_shipment_batch_number__pod_uploaded_at')
        arrived_at = item.get('master_shipment_batch_number__arrived_at')
        shipped_at = item.get('master_shipment_batch_number__shipped_at')
        shipment_appointment = item.get('master_shipment_batch_number__shipment_appointment')
        is_virtual_sp = item.get('master_shipment_batch_number__is_virtual_sp', False)
        
        # иЃ°зЃЧзКґжАБ
        status = 'жЬ™йҐДзЇ¶'
        status_class = 'unappointed'
        
        if pod_link and pod_uploaded_at:
            status = 'еЈ≤еЃМжИР'
            status_class = 'completed'
        elif arrived_at:
            status = 'еЈ≤йАБиЊЊ'
            status_class = 'arrived'
        elif shipped_at:
            status = 'еЈ≤еЗЇеЇУ'
            status_class = 'shipped'
        elif shipment_appointment:
            status = 'еЈ≤йҐДзЇ¶'
            status_class = 'appointed'
        
        # ж†ЉеЉПеМЦжЧ•жЬЯпЉИеЃМжХіж†ЉеЉПпЉЙ
        def format_date(dt):
            if not dt:
                return ''
            if hasattr(dt, 'strftime'):
                return dt.strftime('%Y-%m-%d %H:%M:%S')
            return str(dt)
        
        # еП™ж†ЉеЉПеМЦеєіжЬИжЧ•
        def format_date_only(dt):
            if not dt:
                return ''
            if hasattr(dt, 'strftime'):
                return dt.strftime('%Y-%m-%d')
            return str(dt)
        
        return {
            'container_number': item.get('container_number__container_number', ''),
            'destination': item.get('destination', ''),
            'shipment_batch_number': shipment_batch_number,
            'appointment_id': appointment_id,
            'status': status,
            'status_class': status_class,
            'shipment_appointment': format_date(shipment_appointment),
            'shipment_appointment_date': format_date_only(shipment_appointment),
            'shipped_at': format_date(shipped_at),
            'shipped_at_date': format_date_only(shipped_at),
            'arrived_at': format_date(arrived_at),
            'arrived_at_date': format_date_only(arrived_at),
            'pod_link': pod_link,
            'source_type': source_type,
            'is_virtual_sp': is_virtual_sp
        }
