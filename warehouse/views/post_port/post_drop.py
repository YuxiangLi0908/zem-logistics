import json
import os
import string
import uuid
import math
from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from itertools import zip_longest
from pathlib import Path
import random
from typing import Any, Coroutine
import re

import chardet
import numpy as np
import pandas as pd
from asgiref.sync import sync_to_async
from django.core.exceptions import ObjectDoesNotExist
from django.db import models
from django.db.models import Sum
from django.http import HttpRequest, HttpResponse, Http404
from django.shortcuts import render
from django.utils.dateparse import parse_date
from django.views import View
from simple_history.utils import bulk_update_with_history, bulk_create_with_history

from warehouse.forms.upload_file import UploadFileForm
from django.utils import timezone

from warehouse.models.container import Container
from warehouse.models.container_pickup_carrier import ContainerPickupCarrier
from warehouse.models.fee_detail import FeeDetail
from warehouse.models.fleet import Fleet
from warehouse.models.invoicev2 import Invoicev2, InvoiceStatusv2, InvoiceItemv2
from warehouse.models.quotation_master import QuotationMaster
from warehouse.models.offload import Offload
from warehouse.models.order import Order
from warehouse.models.packing_list import PackingList
from warehouse.models.pallet import Pallet
from warehouse.models.po_check_eta import PoCheckEtaSeven
from warehouse.models.quotation_master import QuotationMaster
from warehouse.models.retrieval import Retrieval
from warehouse.models.shipment import Shipment
from warehouse.models.vessel import Vessel
from warehouse.models.warehouse import ZemWarehouse
from warehouse.models.system_parameter import SystemParameter
from warehouse.models.dropship_cargo import DropshipCargo
from warehouse.models.dropship_inventory import DropshipInventory
from warehouse.models.dropship_shipment import DropshipShipment
from warehouse.models.dropship_shipment_detail import DropshipShipmentDetail
from warehouse.utils.shipment_binding_utils import ShipmentBindingLogger
from warehouse.utils.constants import SHIPPING_LINE_OPTIONS, DELIVERY_METHOD_OPTIONS, ADDITIONAL_CONTAINER, \
    PACKING_LIST_TEMP_COL_MAPPING, DROPSHIPPING_PACKING_LIST_TEMP_COL_MAPPING, DELIVERY_METHOD_CODE
from warehouse.views.export_file import export_do

import json
from django.utils.safestring import mark_safe
from django.shortcuts import redirect, render
from django.db.models import Prefetch, F, Subquery, OuterRef, Exists, Min
from typing import Any, Dict, List, Tuple
from django.contrib.postgres.aggregates import StringAgg
from django.core.exceptions import ObjectDoesNotExist
from django.db import models
from django.db.models import (
    Case,
    CharField,
    BooleanField,
    Count,
    F,
    Func,
    FloatField,
    IntegerField,
    Max,
    Q,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Coalesce
from datetime import datetime, timedelta, time
from warehouse.utils.constants import (
    LOAD_TYPE_OPTIONS,
    DELIVERY_METHOD_OPTIONS, DELIVERY_METHOD_CODE
)
from warehouse.views.post_port.post_nsop import PostNsop
from warehouse.views.post_port.shipment.fleet_management import FleetManagement
from warehouse.views.post_port.shipment.shipping_management import ShippingManagement
from warehouse.views.receivable_accounting import ReceivableAccounting

    

class PostDrop(View):
    template_ltl_pos_all = "post_port/new_sop/07_drop_shipping/07_ltl_main.html"
    template_account_rec = "post_port/new_sop/08_drop_ship_account/09_account_main.html"
    template_account_edit = "post_port/new_sop/08_drop_ship_account/account_detail.html"
    template_ltl_inventory = "post_port/new_sop/09_drop_ship_inventory/09_ltl_inventory.html"


    container_type = {
        "": "",
        "40HQ/GP": "40HQ/GP",
        "45HQ/GP": "45HQ/GP",
        "20GP": "20GP",
        "53HQ": "53HQ",
    }
    shipment_type_options = {
        "FTL": "FTL",
        "外配": "外配",
        # "LTL": "LTL",     
        # "快递": "快递",
        # "客户自提": "客户自提",
    }
    arm_account_options = {
        "": "",
        "Carrier Central ADWG": "Carrier Central ADWG",
        "Carrier Central1": "Carrier Central1",
        "Carrier Central2": "Carrier Central2",
        "ZEM-AMF": "ZEM-AMF",
        "ARM-AMF": "ARM-AMF",
        "walmart": "walmart",
    }
    order_type = {"一件代发": "一件代发"}
    area = {"NJ": "NJ", "SAV": "SAV", "LA": "LA",}

    async def get(self, request: HttpRequest, **kwargs) -> Any | None:
        step = request.GET.get("step")
        pk = kwargs.get("pk", None)
        if step == "postport_delivery":
            context = {
                "warehouse_options": await sync_to_async(list)(
                    ZemWarehouse.objects
                    .order_by("name")
                    .values_list("name", "name")
                )
            }
            return await sync_to_async(render)(request, self.template_ltl_pos_all, context)
        elif step =="account_rec":
            # 应收账单界面
            context = {
                "warehouse_options": await sync_to_async(list)(
                    ZemWarehouse.objects
                    .order_by("name")
                    .values_list("name", "name")
                )
            }
            return await sync_to_async(render)(request, self.template_account_rec, context)
        elif step == "dropship_invetory":
            warehouse_name = request.GET.get("warehouse", "")
            context = await self.handle_dropship_inventory(request, warehouse_name)
            return await sync_to_async(render)(request, self.template_ltl_inventory, context)
        else:
            raise ValueError('wrong step',step)


    async def post(self, request: HttpRequest, **kwargs) -> None | HttpResponse | tuple[Any, Any] | Any:
        step = request.POST.get("step")
        if step == "ltl_post_warehouse":
            template, context = await self.handle_ltl_unscheduled_pos_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "verify_ltl_cargo":
            template, context = await self.handle_verify_ltl_cargo(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "export_ltl_unscheduled":
            return await self.export_ltl_unscheduled(request)
        elif step == "save_releaseCommand":
            template, context = await self.handle_save_releaseCommand(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "account_search":
            template, context = await self.handle_account_search(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "account_edit_fee":
            template, context = await self.handle_account_edit_fee(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "account_save_fee":
            template, context = await self.handle_account_save_fee(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "save_selfdel_cargo":
            template, context = await self.handle_save_selfdel_cargo(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "ltl_bind_group_shipment":
            template, context = await self.handle_ltl_bind_group_shipment(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "dropship_invetory":
            warehouse_name = request.POST.get("warehouse", "")
            context = await self.handle_dropship_inventory(request, warehouse_name)
            return await sync_to_async(render)(request, self.template_ltl_inventory, context)
        elif step == "cancel_fleet":
            # 删除车次后返回一件代发待出库界面
            fm = FleetManagement()
            await fm.handle_cancel_fleet_post(request, 'post_nsop')
            template, context = await self.handle_ltl_unscheduled_pos_post(request)
            context.update({"success_messages": '取消批次成功!'})
            return await sync_to_async(render)(request, template, context)
        elif step == "upload_self_pickup_file":
            # 客户自提BOL文件下载
            pn = PostNsop()
            return await pn.handle_bol_upload_post(request)
        elif step == "export_ltl_label":
            # 导出LTL Label
            pn = PostNsop()
            return await pn.export_ltl_label(request)
        elif step == "export_ltl_bol":
            # 导出LTL BOL
            pn = PostNsop()
            return await pn.export_ltl_bol(request)
        elif step == "export_maersk_label":
            # 导出Maersk Label
            pn = PostNsop()
            return await pn.handle_export_maersk_label(request)
        elif step == "export_maersk_bol":
            # 导出Maersk BOL
            pn = PostNsop()
            return await pn.handle_export_maersk_bol(request)
        elif step == "confirm_delivery":
            # 确认送达
            fm = FleetManagement()
            context = await fm.handle_confirm_delivery_post(request, 'post_nsop')
            template, context = await self.handle_ltl_unscheduled_pos_post(request)
            context.update({"success_messages": '确认送达成功!'})
            return await sync_to_async(render)(request, template, context)
        elif step == "batch_confirm_delivery":
            # 批量确认送达
            fm = FleetManagement()
            count = await fm.handle_batch_confirm_delivery_post(request)
            template, context = await self.handle_ltl_unscheduled_pos_post(request)
            context.update({"success_messages": f'已成功确认 {count} 个批次送达!'})
            return await sync_to_async(render)(request, template, context)
        elif step == "set_shipping_no_link" or step == "batch_set_shipping_no_link":
            # 设置出库单不回传（单个或批量）
            pn = PostNsop()
            _, pn_ctx = await pn.handle_set_shipping_no_link(request)
            template, context = await self.handle_ltl_unscheduled_pos_post(request)
            context.update({"success_messages": pn_ctx.get('success_messages', '操作成功!')})
            return await sync_to_async(render)(request, template, context)
        elif step == "shipping_order_upload" or step == "batch_shipping_order_upload":
            # 回传出库单（单个或批量）
            pn = PostNsop()
            _, pn_ctx = await pn.handle_shipping_order_upload(request)
            template, context = await self.handle_ltl_unscheduled_pos_post(request)
            context.update({"success_messages": pn_ctx.get('success_messages', '出库单上传成功!')})
            return await sync_to_async(render)(request, template, context)
        elif step == "batch_pod_upload":
            # 批量POD上传
            fm = FleetManagement()
            await fm.handle_pod_upload_post(request, 'post_nsop')
            template, context = await self.handle_ltl_unscheduled_pos_post(request)
            context.update({"success_messages": '批量POD上传成功!'})
            return await sync_to_async(render)(request, template, context)
        elif step == "save_shipping_tracking":
            # 保存物流跟踪信息
            pn = PostNsop()
            await pn.handle_save_shipping_tracking(request)
            template, context = await self.handle_ltl_unscheduled_pos_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "update_pod_status":
            # 更新POD回传状态
            pn = PostNsop()
            _, pn_ctx = await pn.handle_update_pod_status(request)
            template, context = await self.handle_ltl_unscheduled_pos_post(request)
            context.update({"success_messages": pn_ctx.get('success_messages', 'POD状态更新成功!')})
            return await sync_to_async(render)(request, template, context)
        elif step == "confirm_shipment":
            # 确认出库
            template, context = await self.handle_confirm_shipment(request)
            return await sync_to_async(render)(request, template, context)
        else:
            raise ValueError('wrong step',step)
        
    async def handle_ltl_unscheduled_pos_post(
            self, request: HttpRequest, context: dict | None = None,
    ) -> tuple[str, dict[str, Any]]:
        '''LTL组的港后全流程'''
        warehouse = request.POST.get("warehouse")
        if not context:
            context = {}
        if warehouse:
            warehouse_name = warehouse
        else:
            context.update({'error_messages': "没选仓库！"})
            return self.template_ltl_pos_all, context

        release_cargos = await self._get_release_cargos(warehouse)
        selfdel_cargos = await self._get_selfdel_cargos(warehouse)
        # 待出库批次
        ready_to_ship_data = await self._get_ready_to_ship_data(warehouse)
        pod_data = await self._get_pod_data(warehouse)

        summary = {
            'release_count': len(release_cargos),
            'selfdel_count': len(selfdel_cargos),
            'ready_to_ship_count': len(ready_to_ship_data),
            'pod_count': len(pod_data),
        }

        supplier_mapping = await sync_to_async(SystemParameter.get_active_by_category)("私仓供应商")

        context.update({
            'warehouse': warehouse,
            'warehouse_options': await sync_to_async(list)(
                ZemWarehouse.objects
                .order_by("name")
                .values_list("name", "name")
            ),
            'account_options': self.arm_account_options,
            'load_type_options': LOAD_TYPE_OPTIONS,
            "release_cargos": release_cargos,
            "selfdel_cargos": selfdel_cargos,
            "ready_to_ship_data": ready_to_ship_data,
            "pod_data": pod_data,
            "summary": summary,
            'shipment_type_options': self.shipment_type_options,
            "warehouse_name": warehouse_name,
            "supplier_mapping_json": mark_safe(json.dumps(supplier_mapping)),
        })
        active_tab = request.POST.get('active_tab')
        if active_tab:
            context.update({'active_tab': active_tab})
        return self.template_ltl_pos_all, context

    async def _get_release_cargos(self, warehouse: str) -> list:
        '''获取未放行货物数据'''
        base_criteria = Q(warehouse__name=warehouse) & Q(status='not_in_stock') & (
            Q(order__retrieval_id__planned_release_time__isnull=False) |
            Q(order__retrieval_id__temp_t49_available_for_pickup=True)
        )

        cargos_qs = DropshipCargo.objects.select_related(
            'order',
            'order__customer_name',
            'order__container_number',
            'order__retrieval_id',
            'order__vessel_id',
            'warehouse',
        ).filter(base_criteria)
        release_cargos_raw = await sync_to_async(list)(cargos_qs)

        release_cargos = []
        for cargo in release_cargos_raw:
            release_cargos.append({
                'ids': cargo.id,
                'plt_ids': '',
                'customer_name': cargo.order.customer_name.zem_name if cargo.order and cargo.order.customer_name else '-',
                'vessel_eta': cargo.order.vessel_id.vessel_eta if cargo.order and cargo.order.vessel_id else None,
                'container_numbers': cargo.order.container_number.container_number if cargo.order and cargo.order.container_number else '-',
                'dropshipping_item_name': cargo.product_name or '-',
                'dropshipping_item_model_number': cargo.model or '-',
                'shipping_marks': cargo.shipping_mark or '-',
                'address': cargo.address,
                'note': cargo.note or '-',
                'total_cbm': cargo.cbm or 0,
                'total_pcs': cargo.pcs or 0,
                'total_weight_lbs': cargo.total_weight_lbs or 0,
                'total_weight_kg': cargo.total_weight_kg or 0,
            })
        return release_cargos

    async def _get_selfdel_cargos(self, warehouse: str) -> list:
        '''获取已放行货物数据'''       
        base_criteria = Q(warehouse__name=warehouse) & ~Q(status='all_out') & (
            Q(order__retrieval_id__planned_release_time__isnull=False) |
            Q(order__retrieval_id__temp_t49_available_for_pickup=True)
        )

        cargos_qs = DropshipCargo.objects.select_related(
            'order',
            'order__customer_name',
            'order__container_number',
            'order__retrieval_id',
            'order__vessel_id',
            'warehouse',
        ).filter(base_criteria)
        selfdel_cargos_raw = await sync_to_async(list)(cargos_qs)

        selfdel_cargos = []
        for cargo in selfdel_cargos_raw:
            selfdel_cargos.append({
                'ids': cargo.id,
                'plt_ids': '',
                'customer_name': cargo.order.customer_name.zem_name if cargo.order and cargo.order.customer_name else '-',
                'vessel_eta': cargo.order.vessel_id.vessel_eta if cargo.order and cargo.order.vessel_id else None,
                'container_numbers': cargo.order.container_number.container_number if cargo.order and cargo.order.container_number else '-',
                'dropshipping_item_name': cargo.product_name or '-',
                'dropshipping_item_model_number': cargo.model or '-',
                'shipping_marks': cargo.shipping_mark or '-',
                'address': cargo.address,
                'note': cargo.note or '-',
                'delivery_method': cargo.delivery_method,
                'pallets': cargo.pallets,
                'total_cbm': cargo.cbm or 0,
                'total_pcs': cargo.pcs or 0,
                'total_weight_lbs': cargo.total_weight_lbs or 0,
                'total_weight_kg': cargo.total_weight_kg or 0,
            })
        return selfdel_cargos

    async def _get_ready_to_ship_data(self, warehouse: str) -> list:
        '''获取待出库批次数据'''
        warehouse_obj = await sync_to_async(ZemWarehouse.objects.filter(name=warehouse).first)()
        if not warehouse_obj:
            return []

        dropship_shipments = await sync_to_async(list)(
            DropshipShipment.objects
            .filter(warehouse=warehouse_obj, shipped_at__isnull=True)
            .select_related('warehouse')
            .order_by('-created_at')
        )

        ready_to_ship_data = []
        for shipment in dropship_shipments:
            details = await sync_to_async(list)(
                DropshipShipmentDetail.objects
                .filter(shipment=shipment)
                .select_related('cargo')
            )

            shipment_data = {
                'shipment_batch_number': shipment.shipment_batch_number,
                'total_pcs': shipment.total_pcs,
                'pickup_time': shipment.pickup_time,
                'details': [],
            }

            models_set = set()
            marks_set = set()

            for detail in details:
                cargo = detail.cargo
                models_set.add(cargo.model or '')
                marks_set.add(cargo.shipping_mark or '')

                inventory = await sync_to_async(
                    DropshipInventory.objects
                    .filter(cargo=cargo, transaction_type='pick', shipment_detail=detail)
                    .first
                )()

                shipment_data['details'].append({
                    'cargo_id': cargo.id,
                    'detail_id': detail.id,
                    'model': cargo.model or '',
                    'shipping_mark': cargo.shipping_mark or '',
                    'pcs': detail.pcs,
                    'original_pcs': detail.pcs,
                    'current_cargo_pcs': cargo.pcs or 0,
                    'inventory_id': inventory.id if inventory else None,
                    'inventory_after_pcs': inventory.after_pcs if inventory else 0,
                })

            shipment_data['models'] = ', '.join([m for m in models_set if m])
            shipment_data['shipping_marks'] = ', '.join([m for m in marks_set if m])

            ready_to_ship_data.append(shipment_data)
        return ready_to_ship_data

    async def _get_pod_data(self, warehouse: str) -> list:
        '''获取待传POD批次数据'''
        pn = PostNsop()
        pod_data = await pn._ltl_pod_get(warehouse)
        pod_data = sorted(
            pod_data,
            key=lambda p: p.pod_to_customer is True
        )
        return pod_data

    async def handle_confirm_shipment(
            self, request: HttpRequest, context: dict | None = None,
    ) -> tuple[str, dict[str, Any]]:
        '''确认出库处理'''
        if not context:
            context = {}

        batch_number = request.POST.get('shipment_batch_number')
        ship_time_str = request.POST.get('ship_time')
        cargo_details_raw = request.POST.get('cargo_details')

        if not batch_number or not ship_time_str or not cargo_details_raw:
            context.update({'error_messages': '参数不完整!'})
            return await self.handle_ltl_unscheduled_pos_post(request, context)

        try:
            cargo_details = json.loads(cargo_details_raw)
        except json.JSONDecodeError:
            context.update({'error_messages': '数据格式错误!'})
            return await self.handle_ltl_unscheduled_pos_post(request, context)

        from datetime import datetime
        from zoneinfo import ZoneInfo
        ship_time = datetime.fromisoformat(ship_time_str).replace(tzinfo=ZoneInfo('Asia/Shanghai'))

        shipment = await sync_to_async(DropshipShipment.objects.filter(shipment_batch_number=batch_number).first)()
        if not shipment:
            context.update({'error_messages': '未找到该预约批次!'})
            return await self.handle_ltl_unscheduled_pos_post(request, context)

        total_pcs = 0
        for detail_data in cargo_details:
            cargo_id = detail_data.get('cargo_id')
            pcs = detail_data.get('pcs', 0)
            original_pcs = detail_data.get('original_pcs', 0)
            inventory_id = detail_data.get('inventory_id')

            cargo = await sync_to_async(DropshipCargo.objects.filter(id=cargo_id).first)()
            if pcs != original_pcs:
                delta = original_pcs - pcs
                verify_pcs = max(0, cargo.pcs + delta)
                cargo.pcs = verify_pcs
                await sync_to_async(cargo.save)()

            inventory = await sync_to_async(DropshipInventory.objects.filter(id=inventory_id).first)()
            inventory.is_verify = True
            if pcs == original_pcs:
                inventory.verify_pcs = inventory.after_pcs
            else:
                inventory.verfiy_pcs_change = -pcs
                inventory.verify_pcs = verify_pcs
            await sync_to_async(inventory.save)()

            total_pcs += pcs

        shipment.total_pcs = total_pcs
        shipment.shipped_at = ship_time
        await sync_to_async(shipment.save)()

        template, context = await self.handle_ltl_unscheduled_pos_post(request)
        context.update({"success_messages": f'批次 {batch_number} 出库确认成功!'})
        return template, context

    async def handle_verify_ltl_cargo(
            self, request: HttpRequest, context: dict | None = None,
    ) -> tuple[str, dict[str, Any]]:
        """LTL对po更改核实状态"""
        if not context:
            context = {}
        cargo_ids = request.POST.get('cargo_ids', '')
        ltl_verify = request.POST.get('ltl_verify', 'false').lower() == 'true'

        
        return await self.handle_ltl_unscheduled_pos_post(request)

    async def handle_save_releaseCommand(
            self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        '''单条或批量保存未放行的指令数据'''
        # 1. 尝试获取批量数据
        batch_commands_raw = request.POST.get('batch_commands')
        tasks = []

        if batch_commands_raw:
            try:
                commands_list = json.loads(batch_commands_raw)
            except json.JSONDecodeError:
                commands_list = []
        else:
            cargo_id = request.POST.get('cargo_id')
            release_command = request.POST.get('release_command')
            if cargo_id:
                commands_list = [{'cargo_id': cargo_id, 'command': release_command}]
            else:
                commands_list = []
        num = 0
        for item in commands_list:
            c_id = item.get('cargo_id')
            command_text = item.get('command')
            if not c_id or not command_text: continue

            if c_id.startswith('plt_'):
                continue

            num += 1
        context = {'success_messages': f'保存成功{num}组数据!'}
        return await self.handle_ltl_unscheduled_pos_post(request, context)

    async def handle_account_search(
            self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        """一件代发应收账单查询：根据筛选条件查询订单(主查order表)，
        按柜号分组，并查询 Invoicev2 / InvoiceStatusv2，
        将柜子分为 待录入 / 已录入 两类。"""
        # 1. 读取筛选条件
        warehouse_filter = request.POST.get("warehouse_filter", "").strip()
        start_date = request.POST.get("start_date", "").strip()
        end_date = request.POST.get("end_date", "").strip()
        container_number_filter = request.POST.get("container_number_filter", "").strip()

        # --- 1. 日期处理 ---
        if container_number_filter:
            start_date = None
            end_date = None
        else:
            current_date = datetime.now().date()
            start_date = (
                (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
                if not start_date
                else start_date
            )
            end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date

        # --- 2. 构建查询条件 ---
        if container_number_filter:
            criteria = Q(container_number__container_number=container_number_filter)
        else:
            criteria = (
                Q(cancel_notification=False)
                & Q(order_type="一件代发") 
                & Q(vessel_id__vessel_etd__gte=start_date)
                & Q(vessel_id__vessel_etd__lte=end_date)
                & Q(offload_id__offload_at__isnull=False)
            )
            if warehouse_filter and warehouse_filter != 'None':
                if "LA" in warehouse_filter:
                    criteria &= Q(retrieval_id__retrieval_destination_precise__contains='LA')
                else:
                    criteria &= Q(retrieval_id__retrieval_destination_precise=warehouse_filter)

        # 3. 查询符合条件的订单（按柜号去重），主查 order 表
        base_orders = (
            Order.objects
            .select_related(
                'retrieval_id', 
                'offload_id', 
                'container_number',
                'customer_name',
                'warehouse',
            )
            .annotate(
                retrieval_time=F("retrieval_id__actual_retrieval_timestamp"),
                empty_returned_time=F("retrieval_id__empty_returned_at"),
                offload_time=F("offload_id__offload_at"),
            )
            .filter(criteria)
            .distinct()
        )

        orders_list = await sync_to_async(list)(base_orders)
        # 提取 Container IDs
        container_ids = set()
        for order in orders_list:
            if order.container_number_id:
                container_ids.add(order.container_number_id)
        container_ids = list(container_ids)

        # --- 4. 批量获取 Invoice 和 InvoiceStatus ---
        status_prefetch = Prefetch(
            'invoicestatusv2_set',
            queryset=InvoiceStatusv2.objects.filter(invoice_type="receivable"),
            to_attr='receivable_status_list'
        )

        all_invoices_qs = Invoicev2.objects.filter(
            container_number_id__in=container_ids
        ).prefetch_related(status_prefetch)
        all_invoices = await sync_to_async(list)(all_invoices_qs)

         # --- 5. [安全机制] 批量创建缺失的 InvoiceStatus ---
        missing_statuses = []
        invoices_needing_update = []

        for inv in all_invoices:
            # 如果预查询列表为空，说明缺数据
            if not (hasattr(inv, 'receivable_status_list') and inv.receivable_status_list):
                new_status = InvoiceStatusv2(
                    invoice=inv,
                    container_number_id=inv.container_number_id, # 使用ID赋值更轻量
                    invoice_type="receivable",
                    # 默认状态
                    warehouse_public_status="completed",
                    warehouse_other_status="unstarted",
                    preport_status="unstarted",
                    delivery_public_status="completed",
                    delivery_other_status="completed",
                    finance_status="unstarted"
                )
                missing_statuses.append(new_status)
                invoices_needing_update.append(inv)

        if missing_statuses:
            await sync_to_async(bulk_create_with_history)(missing_statuses, InvoiceStatusv2)
            # 手动回填内存，避免重新查询
            for i, inv in enumerate(invoices_needing_update):
                inv.receivable_status_list = [missing_statuses[i]]
        
        # --- 6. 内存分组 & 统计预计算 ---
        container_invoice_map = defaultdict(list)
        for inv in all_invoices:
            container_invoice_map[inv.container_number_id].append(inv)

        # 状态中文展示映射
        preport_status_map = {
            "unstarted": "未录入",
            "in_progress": "录入中",
            "save": "暂存",
            "pending_review": "待组长审核",
            "completed": "已完成",
            "rejected": "已驳回",
        }
        warehouse_status_map = {
            "unstarted": "未录入",
            "in_progress": "录入中",
            "completed": "已完成",
            "rejected": "已驳回",
            "financials_completed": "财务已完成",
        }

        # --- 7. 主循环 ---
        pending_orders = [] #待录入
        recorded_orders = []  #已录入
        for order in orders_list:
            container = order.container_number
            if not container:
                continue
            
            c_id = container.id

            container_invoices = container_invoice_map.get(c_id, [])

            # 定义构建函数
            def build_order_data(inv=None, status_obj=None):
                created_at = None
                # 只有多账单才去拿时间，且只拿 created_at，不碰 history 以免 N+1
                if inv and len(container_invoices) > 1 and not inv.is_master_bill:
                    created_at = inv.created_at 

                preport_status = status_obj.preport_status if status_obj else None
                warehouse_status = status_obj.warehouse_other_status if status_obj else None
                finance_status = status_obj.finance_status if status_obj else None

                return {
                    'order': order,
                    'container_number': order.container_number,
                    'invoice_number': inv.invoice_number if inv else None,
                    'invoice_id': inv.id if inv else None,
                    'invoice_created_at': created_at,
                    # 提拆费 / 库内费（来自账单 Invoicev2）
                    'preport_amount': inv.receivable_preport_amount if inv else None,
                    'wh_other_amount': inv.receivable_wh_other_amount if inv else None,
                    # 状态原始值
                    'preport_status': preport_status,
                    'warehouse_status': warehouse_status,
                    'finance_status': finance_status,
                    # 状态中文展示
                    'preport_status_display': preport_status_map.get(preport_status, preport_status or '未录入'),
                    'warehouse_status_display': warehouse_status_map.get(warehouse_status, warehouse_status or '未录入'),
                    'finance_status_value': 'completed' if finance_status == 'completed' else 'unconfirmed',
                    'finance_status_display': '已完成' if finance_status == 'completed' else '未确认',
                    'has_invoice': bool(inv),
                    'offload_time': order.offload_time,
                    'actual_retrieval_timestamp': order.retrieval_id.actual_retrieval_timestamp if order.retrieval_id else None,
                    'customer_name': order.customer_name.zem_name if order.customer_name else None,
                    'warehouse': order.warehouse.name if order.warehouse else None,
                }

            if not container_invoices:
                # === 场景 A: 无账单 ===
                base_data = build_order_data(None, None)
                pending_orders.append(base_data)
                
            else:
                # === 场景 B: 有账单 ===
                for invoice in container_invoices:
                    status_obj = None
                    if hasattr(invoice, 'receivable_status_list') and invoice.receivable_status_list:
                        for status in invoice.receivable_status_list:
                            if status.invoice_id == invoice.id:
                                status_obj = status
                                break
                        # 如果没有找到匹配的，才取第一个
                        if not status_obj and invoice.receivable_status_list:
                            raise ValueError('账单没有状态表')
                    
                    base_data = build_order_data(invoice, status_obj)
                    p_item = base_data.copy()
                    preport_status = p_item['preport_status']
                    warehouse_status = p_item['warehouse_status']

                    if preport_status == "completed" and warehouse_status == "completed":
                        recorded_orders.append(base_data)
                    else:
                        pending_orders.append(base_data)

        context = {
            "warehouse_options": await sync_to_async(list)(
                ZemWarehouse.objects.order_by("name").values_list("name", "name")
            ),
            "warehouse_filter": warehouse_filter,
            "start_date": start_date,
            "end_date": end_date,
            "pending_orders": pending_orders,
            "recorded_orders": recorded_orders,
        }
        return self.template_account_rec, context

    async def handle_account_edit_fee(
            self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        """一件代发应收账单：提拆费/库内费录入页面。
        前端通过表单提交 invoice_number 和 fee_type(preport/warehouse)。
        根据 invoice_number 查询账单和已录费用，显示到 template_account_edit 页面。"""
        invoice_number = request.POST.get("invoice_number", "").strip()
        container_number = request.POST.get("container_number", "").strip()
        fee_type = request.POST.get("fee_type", "").strip()

        if not invoice_number or invoice_number.lower() == "none":
            ra = ReceivableAccounting()
            invoice, invoice_status, invoice_status_payable = await sync_to_async(
                ra._create_invoice_and_status
            )(container_number)
            invoice_number = invoice.invoice_number
            

        # 查询账单
        invoice = await sync_to_async(
            Invoicev2.objects.select_related("container_number")
            .filter(invoice_number=invoice_number).first
        )()

        if not invoice:
            template, context = await self.handle_account_search(request)
            context['error_messages'] = f"未找到账单号 {invoice_number} 对应的账单！"
            return template, context

        # 确定 item_category
        if fee_type == "preport":
            item_category = "preport"
            fee_type_display = "提拆费"
        elif fee_type == "warehouse":
            item_category = "warehouse_other"
            fee_type_display = "库内费"
        else:
            template, context = await self.handle_account_search(request)
            context['error_messages'] = f"未知的费用类型：{fee_type}！"
            return template, context

        # 自动初始化费用项：如果账单状态未录入，自动创建默认费用
        status_obj = await sync_to_async(InvoiceStatusv2.objects.filter(
            invoice=invoice, invoice_type="receivable"
        ).first)()
        
        if fee_type == "preport" and status_obj and status_obj.preport_status != "completed":
            existing = await sync_to_async(InvoiceItemv2.objects.filter(
                invoice_number=invoice, item_category="preport"
            ).exists)()
            if not existing:
                dropship_quote = await sync_to_async(
                    QuotationMaster.objects.filter(
                        quote_type="receivable",
                        effective_date__isnull=False,
                    ).order_by("-effective_date").first
                )()
                
                if dropship_quote:
                    preport_fee_detail = await sync_to_async(
                        FeeDetail.objects.filter(
                            quotation_id=dropship_quote,
                            fee_type="preport"
                        ).first
                    )()
                    
                    if preport_fee_detail and preport_fee_detail.details:
                        for fee_code, fee_data in preport_fee_detail.details.items():
                            await sync_to_async(InvoiceItemv2.objects.create)(
                                container_number=invoice.container_number,
                                invoice_number=invoice,
                                invoice_type="receivable",
                                item_category="preport",
                                description=fee_data.get("description", fee_code),
                                amount=fee_data.get("rate", 0),
                                rate=fee_data.get("rate", 0),
                                qty=1,
                            )
                    else:
                        await sync_to_async(InvoiceItemv2.objects.create)(
                            container_number=invoice.container_number,
                            invoice_number=invoice,
                            invoice_type="receivable",
                            item_category="preport",
                            description="提拆费",
                            amount=950,
                            rate=950,
                            qty=1,
                        )
                else:
                    await sync_to_async(InvoiceItemv2.objects.create)(
                        container_number=invoice.container_number,
                        invoice_number=invoice,
                        invoice_type="receivable",
                        item_category="preport",
                        description="提拆费",
                        amount=950,
                        rate=950,
                        qty=1,
                    )
                    
        elif fee_type == "warehouse" and status_obj and status_obj.warehouse_other_status != "completed":
            existing = await sync_to_async(InvoiceItemv2.objects.filter(
                invoice_number=invoice, item_category="warehouse_other"
            ).exists)()
            if not existing:
                total_pcs = 0
                if invoice.container_number:
                    pallets = await sync_to_async(list)(
                        Pallet.objects.filter(
                            container_number=invoice.container_number
                        ).values("pcs")
                    )
                    if pallets:
                        total_pcs = sum(p["pcs"] or 0 for p in pallets)
                    else:
                        packing_items = await sync_to_async(list)(
                            PackingList.objects.filter(
                                container_number=invoice.container_number
                            ).values("pcs")
                        )
                        if packing_items:
                            total_pcs = sum(p["pcs"] or 0 for p in packing_items)
                
                dropship_quote = await sync_to_async(
                    QuotationMaster.objects.filter(
                        quote_type="receivable",
                        effective_date__isnull=False,
                    ).order_by("-effective_date").first
                )()
                
                if dropship_quote:
                    warehouse_fee_detail = await sync_to_async(
                        FeeDetail.objects.filter(
                            quotation_id=dropship_quote,
                            fee_type="warehouse"
                        ).first
                    )()
                    
                    if warehouse_fee_detail and warehouse_fee_detail.details:
                        for fee_code, fee_data in warehouse_fee_detail.details.items():
                            rate = fee_data.get("rate", 0)
                            description = fee_data.get("description", fee_code)
                            if "出库费" in description:
                                amount = 0
                            elif "仓储费" in description:
                                amount = 0
                            else:
                                amount = 0
                            
                            await sync_to_async(InvoiceItemv2.objects.create)(
                                container_number=invoice.container_number,
                                invoice_number=invoice,
                                invoice_type="receivable",
                                item_category="warehouse_other",
                                description=description,
                                amount=amount,
                                rate=rate,
                            )
                    else:
                        warehouse_fees = [
                            {"description": "出库费TBR", "rate": 3, "amount": 0},
                            {"description": "出库费PCR", "rate": 1, "amount": 0},
                            {"description": "仓储费TBR", "rate": 3.5, "amount": 0},
                            {"description": "仓储费PCR", "rate": 3.5, "amount": 0},
                        ]
                        for fee in warehouse_fees:
                            await sync_to_async(InvoiceItemv2.objects.create)(
                                container_number=invoice.container_number,
                                invoice_number=invoice,
                                invoice_type="receivable",
                                item_category="warehouse_other",
                                description=fee["description"],
                                amount=fee["amount"],
                                rate=fee["rate"],
                            )
                else:
                    warehouse_fees = [
                        {"description": "出库费TBR", "rate": 3, "amount": 0},
                        {"description": "出库费PCR", "rate": 1, "amount": 0},
                        {"description": "仓储费TBR", "rate": 3.5, "amount": 0},
                        {"description": "仓储费PCR", "rate": 3.5, "amount": 0},
                    ]
                    for fee in warehouse_fees:
                        await sync_to_async(InvoiceItemv2.objects.create)(
                            container_number=invoice.container_number,
                            invoice_number=invoice,
                            invoice_type="receivable",
                            item_category="warehouse_other",
                            description=fee["description"],
                            amount=fee["amount"],
                            rate=fee["rate"],
                        )

        # 查询已录费用
        items = await sync_to_async(list)(
            InvoiceItemv2.objects.filter(
                invoice_number=invoice,
                item_category=item_category,
            ).order_by("id")
        )

        # 查询柜号
        container_number_str = (
            invoice.container_number.container_number
            if invoice.container_number else None
        )

        # 查询仓库（订单 retrieval_id.retrieval_destination_precise）
        warehouse_name = None
        warehouse_precise = None
        if invoice.container_number:
            order = await sync_to_async(Order.objects.select_related(
                "retrieval_id", "warehouse"
            ).filter(container_number=invoice.container_number).first)()
            if order:
                if order.retrieval_id:
                    warehouse_precise = order.retrieval_id.retrieval_destination_precise
                if order.warehouse:
                    warehouse_name = order.warehouse.name

        # 查询财务状态
        finance_status_value = None
        finance_status_display = "财务未确认，可编辑"
        is_editable = True
        status_obj = await sync_to_async(InvoiceStatusv2.objects.filter(
            invoice=invoice, invoice_type="receivable"
        ).first)()
        if status_obj:
            finance_status_value = status_obj.finance_status
            if finance_status_value == "completed":
                finance_status_display = "财务已确认，不可编辑"
                is_editable = False
            else:
                finance_status_display = "财务未确认，可编辑"
                is_editable = True

        # 计算当前总价
        total_amount = sum((item.amount or 0) for item in items)

        # 件数详情：库内费时读取 Pallet/PackingList，按唛头分组统计件数
        piece_details = []
        if fee_type == "warehouse" and invoice.container_number:
            pallets = await sync_to_async(list)(
                Pallet.objects.select_related("shipment_batch_number")
                .filter(container_number=invoice.container_number)
                .values("shipping_mark", "pcs", "PO_ID", "shipment_batch_number__shipped_at")
            )
            if pallets:
                grouped = {}
                for p in pallets:
                    mark = p["shipping_mark"] or "无唛头"
                    if mark not in grouped:
                        grouped[mark] = {"pcs": 0, "shipped_at": p.get("shipment_batch_number__shipped_at")}
                    grouped[mark]["pcs"] += (p["pcs"] or 0)
                    if p.get("shipment_batch_number__shipped_at"):
                        if not grouped[mark]["shipped_at"] or p["shipment_batch_number__shipped_at"] > grouped[mark]["shipped_at"]:
                            grouped[mark]["shipped_at"] = p["shipment_batch_number__shipped_at"]
                piece_details = [
                    {"shipping_mark": mark, "pcs": data["pcs"], "shipped_at": data["shipped_at"]}
                    for mark, data in grouped.items()
                ]
            else:
                packing_items = await sync_to_async(list)(
                    PackingList.objects.select_related("shipment_batch_number")
                    .filter(container_number=invoice.container_number)
                    .values("shipping_mark", "pcs", "shipment_batch_number__shipped_at")
                )
                if packing_items:
                    grouped = {}
                    for p in packing_items:
                        mark = p["shipping_mark"] or "无唛头"
                        if mark not in grouped:
                            grouped[mark] = {"pcs": 0, "shipped_at": p.get("shipment_batch_number__shipped_at")}
                        grouped[mark]["pcs"] += (p["pcs"] or 0)
                        if p.get("shipment_batch_number__shipped_at"):
                            if not grouped[mark]["shipped_at"] or p["shipment_batch_number__shipped_at"] > grouped[mark]["shipped_at"]:
                                grouped[mark]["shipped_at"] = p["shipment_batch_number__shipped_at"]
                    piece_details = [
                        {"shipping_mark": mark, "pcs": data["pcs"], "shipped_at": data["shipped_at"]}
                        for mark, data in grouped.items()
                    ]

            if piece_details:
                orders = await sync_to_async(list)(
                    Order.objects.select_related("offload_id")
                    .filter(container_number=invoice.container_number)
                    .values("offload_id__offload_other_at")
                )
                offload_at = None
                if orders:
                    for o in orders:
                        if o.get("offload_id__offload_other_at"):
                            offload_at = o["offload_id__offload_other_at"]
                            break

                for detail in piece_details:
                    detail["offload_at"] = offload_at
                    if offload_at and detail["shipped_at"]:
                        storage_days = (detail["shipped_at"] - offload_at).days
                        detail["storage_days"] = max(0, storage_days)
                        detail["storage_months"] = max(1, (storage_days + 29) // 30)
                    else:
                        detail["storage_days"] = None
                        detail["storage_months"] = None
        piece_details_json = json.dumps(piece_details, ensure_ascii=False)

        context = {
            "invoice_number": invoice_number,
            "invoice_id": invoice.id,
            "fee_type": fee_type,
            "item_category": item_category,
            "fee_type_display": fee_type_display,
            "container_number": container_number_str,
            "warehouse_name": warehouse_name,
            "warehouse_precise": warehouse_precise,
            "finance_status_value": finance_status_value,
            "finance_status_display": finance_status_display,
            "is_editable": is_editable,
            "items": items,
            "total_amount": total_amount,
            "delivery_type_display": "一件代发",
            "piece_details": piece_details,
            "piece_details_json": piece_details_json,
        }
        return self.template_account_edit, context

    async def handle_account_save_fee(
            self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        """一件代发应收账单：保存提拆费/库内费明细到 InvoiceItemv2 表。
        保存后返回 template_account_rec 界面。"""
        invoice_number_str = request.POST.get("invoice_number", "").strip()
        fee_type = request.POST.get("fee_type", "").strip()

        messages: dict[str, Any] = {}

        if not invoice_number_str:
            messages['error_messages'] = "缺少账单号，保存失败！"
        else:
            # 查询账单（预加载 container_number 关联，避免异步上下文里懒加载报错）
            invoice = await sync_to_async(
                Invoicev2.objects.select_related("container_number")
                .filter(invoice_number=invoice_number_str).first
            )()
            if not invoice:
                messages['error_messages'] = f"未找到账单号 {invoice_number_str} 对应的账单，保存失败！"
            else:
                # 确定 item_category
                if fee_type == "preport":
                    item_category = "preport"
                    fee_type_display = "提拆费"
                elif fee_type == "warehouse":
                    item_category = "warehouse_other"
                    fee_type_display = "库内费"
                else:
                    item_category = None
                    fee_type_display = fee_type

                if item_category:
                    # 校验财务状态，completed 不可保存
                    status_obj = await sync_to_async(InvoiceStatusv2.objects.filter(
                        invoice=invoice, invoice_type="receivable"
                    ).first)()
                    if status_obj and status_obj.finance_status == "completed":
                        messages['error_messages'] = "财务已确认，不可编辑！"
                    else:
                        # 收集表单数据
                        total_amount = 0
                        item_ids = request.POST.getlist("item_id[]")
                        descriptions = request.POST.getlist("description[]")
                        warehouse_codes = request.POST.getlist("warehouse_code[]")
                        rates = request.POST.getlist("rate[]")
                        qtys = request.POST.getlist("qty[]")
                        surcharges_list = request.POST.getlist("surcharges[]")
                        amounts = request.POST.getlist("amount[]")
                        notes = request.POST.getlist("note[]")

                        # 删除现有该分类下的所有明细，再重新创建
                        await sync_to_async(InvoiceItemv2.objects.filter(
                            invoice_number=invoice,
                            item_category=item_category,
                        ).delete)()

                        # 获取关联的 container
                        container = invoice.container_number

                        # 创建新的明细
                        new_items = []
                        for i in range(len(descriptions)):
                            if not descriptions[i].strip():
                                continue
                            try:
                                rate_val = float(rates[i]) if rates[i] else 0
                            except (ValueError, TypeError):
                                rate_val = 0
                            try:
                                qty_val = float(qtys[i]) if qtys[i] else 0
                            except (ValueError, TypeError):
                                qty_val = 0
                            try:
                                surcharge_val = float(surcharges_list[i]) if surcharges_list[i] else 0
                            except (ValueError, TypeError):
                                surcharge_val = 0
                            try:
                                amount_val = float(amounts[i]) if amounts[i] else 0
                            except (ValueError, TypeError):
                                amount_val = 0
                            
                            total_amount += amount_val

                            new_items.append(InvoiceItemv2(
                                container_number=container,
                                invoice_number=invoice,
                                invoice_type="receivable",
                                item_category=item_category,
                                description=descriptions[i].strip(),
                                warehouse_code=warehouse_codes[i].strip() if warehouse_codes[i] else None,
                                rate=rate_val,
                                qty=qty_val,
                                surcharges=surcharge_val,
                                amount=amount_val,
                                note=notes[i].strip() if notes[i] else None,
                            ))

                        if new_items:
                            await sync_to_async(bulk_create_with_history)(
                                new_items, InvoiceItemv2
                            )

                        # 保存成功后，更新对应状态表字段为 completed
                        if fee_type == "preport":
                            await sync_to_async(InvoiceStatusv2.objects.filter(
                                invoice=invoice, invoice_type="receivable"
                            ).update)(preport_status="completed")
                            invoice.receivable_preport_amount = total_amount
                        elif fee_type == "warehouse":
                            await sync_to_async(InvoiceStatusv2.objects.filter(
                                invoice=invoice, invoice_type="receivable"
                            ).update)(warehouse_other_status="completed")
                            invoice.receivable_wh_other_amount = total_amount
                        await sync_to_async(invoice.save)()

                        messages['success_messages'] = (
                            f"账单号 {invoice_number_str}<br>"
                            f"{fee_type_display}<br>"
                            f"已保存 {len(new_items)} 条明细！"
                        )

        # 返回账单主页面
        template, context = await self.handle_account_search(request)
        context.update(messages)
        return template, context

    async def export_ltl_unscheduled(
            self, request: HttpRequest
    ) -> HttpResponse:
        """导出未放行货物到Excel"""
        cargo_ids = request.POST.get('cargo_ids', '')
        warehouse = request.POST.get('warehouse')

        # 构建筛选条件
        base_criteria = Q(
            warehouse__name=warehouse,
            status='not_in_stock',
        ) & (
            Q(order__retrieval_id__planned_release_time__isnull=True) |
            Q(order__retrieval_id__temp_t49_available_for_pickup=False)
        )

        # 如果指定了 ID，则只导出选中的货物
        if cargo_ids:
            cargo_id_list = []
            for id_str in cargo_ids.split(','):
                id_str = id_str.strip()
                if id_str and not id_str.startswith('plt_'):
                    try:
                        cargo_id_list.append(int(id_str))
                    except ValueError:
                        pass
            if cargo_id_list:
                base_criteria &= Q(id__in=cargo_id_list)

        # 获取数据
        release_cargos_qs = DropshipCargo.objects.select_related(
            'order',
            'order__customer_name',
            'order__container_number',
            'order__retrieval_id',
        ).filter(base_criteria)
        release_cargos_raw = await sync_to_async(list)(release_cargos_qs)

        # 转换为前端需要的格式
        release_cargos = []
        for cargo in release_cargos_raw:
            release_cargos.append({
                'customer_name': cargo.order.customer_name.zem_name if cargo.order and cargo.order.customer_name else '-',
                'container_numbers': cargo.order.container_number.container_number if cargo.order and cargo.order.container_number else '-',
                'dropshipping_item_name': cargo.product_name or '-',
                'dropshipping_item_model_number': cargo.model or '-',
                'shipping_marks': cargo.shipping_mark or '-',
                'address': cargo.address or cargo.order.retrieval_id.retrieval_address if cargo.order and cargo.order.retrieval_id else '-',
                'note': cargo.note or '-',
                'total_cbm': cargo.cbm or 0,
                'total_pcs': cargo.pcs or 0,
                'total_weight_lbs': cargo.total_weight_lbs or 0,
                'total_weight_kg': cargo.total_weight_kg or 0,
                'ltl_verify': False,
            })

        # 准备 Excel 数据
        excel_data = []
        for cargo in release_cargos:
            customer_name = cargo.get('customer_name', '-')
            container_numbers = cargo.get('container_numbers', '-')
            item_name = cargo.get('dropshipping_item_name', '-')
            item_model = cargo.get('dropshipping_item_model_number', '-')
            shipping_marks = cargo.get('shipping_marks', '-')
            address = cargo.get('address', '-')
            note = cargo.get('note', '-')

            # 格式化数字
            try:
                total_cbm = float(cargo.get('total_cbm', 0))
                total_cbm = round(total_cbm, 3)
            except (ValueError, TypeError):
                total_cbm = 0

            try:
                total_pcs = int(cargo.get('total_pcs', 0))
            except (ValueError, TypeError):
                total_pcs = 0

            try:
                weight_lbs = float(cargo.get('total_weight_lbs', 0))
                weight_lbs = round(weight_lbs, 2)
            except (ValueError, TypeError):
                weight_lbs = 0

            try:
                weight_kg = float(cargo.get('total_weight_kg', 0))
                weight_kg = round(weight_kg, 2)
            except (ValueError, TypeError):
                weight_kg = 0

            # 核实状态
            ltl_verify = cargo.get('ltl_verify', False)
            verify_status = '已核实' if ltl_verify else '未核实'

            row = {
                '客户': customer_name,
                '柜号': container_numbers,
                '商品型号': item_name,
                '商品类型': item_model,
                '唛头': shipping_marks,
                '详细地址': address,
                '备注': note,
                'CBM': total_cbm,
                '件数': total_pcs,
                '重量(lbs)': weight_lbs,
                '重量(kg)': weight_kg,
                '核实状态': verify_status,
            }

            excel_data.append(row)

        # 创建 DataFrame
        df = pd.DataFrame(excel_data)

        # 如果没有数据，创建一个空的DataFrame
        if df.empty:
            df = pd.DataFrame(columns=[
                '客户', '柜号', '商品型号', '商品类型', '唛头', '详细地址', '备注',
                'CBM', '件数', '重量(lbs)', '重量(kg)', '核实状态'
            ])

        # 创建 Excel 文件
        output = BytesIO()

        # 使用 ExcelWriter
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 主数据 sheet
            df.to_excel(writer, sheet_name='未放行货物', index=False)

            # 获取 worksheet 对象
            worksheet = writer.sheets['未放行货物']

            # 设置列宽
            column_widths = {
                '客户': 20,
                '柜号': 25,
                '商品型号': 20,
                '商品类型': 20,
                '唛头': 25,
                '详细地址': 40,
                '备注': 40,
                'CBM': 10,
                '件数': 10,
                '重量(lbs)': 12,
                '重量(kg)': 12,
                '核实状态': 12,
            }

            # 设置列宽
            from openpyxl.utils import get_column_letter

            for i, column in enumerate(df.columns, 1):
                col_letter = get_column_letter(i)
                width = column_widths.get(column, 15)
                worksheet.column_dimensions[col_letter].width = width

            # 设置数字格式
            from openpyxl.styles import numbers

            # 设置CBM列为3位小数格式
            if 'CBM' in df.columns:
                cbm_col_idx = df.columns.get_loc('CBM') + 1
                cbm_col_letter = get_column_letter(cbm_col_idx)
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"{cbm_col_letter}{row}"]
                    cell.number_format = numbers.FORMAT_NUMBER_00

            # 设置重量列为2位小数格式
            if '重量(lbs)' in df.columns:
                lbs_col_idx = df.columns.get_loc('重量(lbs)') + 1
                lbs_col_letter = get_column_letter(lbs_col_idx)
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"{lbs_col_letter}{row}"]
                    cell.number_format = numbers.FORMAT_NUMBER_00

            if '重量(kg)' in df.columns:
                kg_col_idx = df.columns.get_loc('重量(kg)') + 1
                kg_col_letter = get_column_letter(kg_col_idx)
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"{kg_col_letter}{row}"]
                    cell.number_format = numbers.FORMAT_NUMBER_00

            # 设置样式：标题行加粗
            for cell in worksheet[1]:
                cell.font = cell.font.copy(bold=True)

            # 自动换行设置
            from openpyxl.styles import Alignment
            wrap_alignment = Alignment(wrap_text=True, vertical='top')

            # 对可能有多行内容的列设置自动换行
            wrap_columns = ['柜号', '详细地址', '备注', '唛头']
            for col_name in wrap_columns:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    for row in range(1, len(df) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        cell.alignment = wrap_alignment

            # 添加筛选器
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"

            # 冻结标题行
            worksheet.freeze_panes = 'A2'

        output.seek(0)

        # 生成文件名
        timestamp = timezone.now().strftime('_%m%d')
        filename = f'未放行货物_{timestamp}.xlsx'

        # 对文件名进行 URL 编码，确保中文正确处理
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename)

        # 创建 HTTP 响应
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # 使用 RFC 6266 标准设置 Content-Disposition
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"

        # 备用方案：对于不支持 RFC 6266 的旧浏览器
        response['Content-Disposition'] = f"attachment; filename={encoded_filename}"

        return response

    async def handle_save_selfdel_cargo(
            self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        '''一件代发自发保存自行编辑的货物信息'''
        batch_data_raw = request.POST.get('batch_data')
        if batch_data_raw:
            try:
                update_items = json.loads(batch_data_raw)
            except json.JSONDecodeError:
                update_items = []
        else:
            update_items = [{
                'cargo_id': request.POST.get('cargo_id'),
                'address': request.POST.get('address', '').strip(),
                'note': request.POST.get('note', '').strip(),
                'delivery_method': request.POST.get('delivery_method', '').strip(),
            }]

        total_status_messages = []
        saved_count = 0
        for item in update_items:
            cargo_id = item.get('cargo_id')
            if not cargo_id:
                continue

            address = item.get('address', '')
            note = item.get('note', '')
            delivery_method = item.get('delivery_method')

            if cargo_id.startswith('plt_'):
                ids = cargo_id.replace('plt_', '').split(',')
                model = Pallet
            else:
                ids = cargo_id.split(',')
                model = DropshipCargo

            update_data = {}
            if address is not None:
                update_data['address'] = address
            if note is not None:
                update_data['note'] = note
            if delivery_method is not None:
                update_data['delivery_method'] = delivery_method

            if update_data:
                objs = await sync_to_async(list)(model.objects.filter(id__in=ids))
                for obj in objs:
                    for key, value in update_data.items():
                        setattr(obj, key, value)
                    await sync_to_async(obj.save)()
                saved_count += len(objs)

        success_message = f'成功保存 {saved_count} 条记录！'
        if total_status_messages:
            success_message = mark_safe(f"{success_message}<br>" + "<br>".join(set(total_status_messages)))

        context = {'success_messages': success_message}
        return await self.handle_ltl_unscheduled_pos_post(request, context)

    async def handle_ltl_bind_group_shipment(
            self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        '''一件代发 预约出库 - 生成DropshipInventory和DropshipShipment记录'''
        context = {}
        
        cargo_pcs_data_json = request.POST.get('cargo_pcs_data', '[]').strip()
        shipment_appointment = request.POST.get('shipment_appointment', '').strip()
        carrier = request.POST.get('carrier', '').strip()
        arm_bol = request.POST.get('arm_bol', '').strip()
        shipment_type = request.POST.get('shipment_type', '客户自提').strip()
        warehouse = request.POST.get('warehouse')

        if not shipment_appointment:
            context = {'error_messages': '请填写提货时间！'}
            return await self.handle_ltl_unscheduled_pos_post(request, context)

        try:
            cargo_pcs_data = json.loads(cargo_pcs_data_json)
        except (json.JSONDecodeError, TypeError):
            context = {'error_messages': '货物数据格式错误！'}
            return await self.handle_ltl_unscheduled_pos_post(request, context)

        if not cargo_pcs_data:
            context = {'error_messages': '请选择要预约出库的货物！'}
            return await self.handle_ltl_unscheduled_pos_post(request, context)

        try:
            pickup_time = timezone.make_aware(datetime.fromisoformat(shipment_appointment.replace('Z', '')))
        except (ValueError, TypeError):
            context = {'error_messages': '提货时间格式错误！'}
            return await self.handle_ltl_unscheduled_pos_post(request, context)

        try:
            import pytz
            beijing_tz = pytz.timezone('Asia/Shanghai')
            current_time_beijing = datetime.now(beijing_tz)
        except ImportError:
            current_time_beijing = timezone.now()

        cargo_ids = []
        for item in cargo_pcs_data:
            if 'cargo_id' in item:
                cargo_id = item['cargo_id']
                try:
                    cargo_ids.append(int(cargo_id))
                except (ValueError, TypeError):
                    continue

        if not cargo_ids:
            context = {'error_messages': '请选择要预约出库的货物！'}
            return await self.handle_ltl_unscheduled_pos_post(request, context)

        cargos = await sync_to_async(list)(
            DropshipCargo.objects.filter(id__in=cargo_ids).select_related('warehouse')
        )
        cargo_map = {c.id: c for c in cargos}

        warehouse_obj = None
        if warehouse:
            try:
                warehouse_obj = await sync_to_async(ZemWarehouse.objects.filter(id=warehouse).first)()
            except (ValueError, TypeError):
                pass
        
        if not warehouse_obj and cargos:
            warehouse_obj = cargos[0].warehouse

        item_models = []
        for item in cargo_pcs_data:
            if 'cargo_id' in item:
                try:
                    cargo_id = int(item['cargo_id'])
                    cargo = cargo_map.get(cargo_id)
                    if cargo and cargo.model:
                        item_models.append(cargo.model)
                except (ValueError, TypeError):
                    continue
        destination_for_batch = '-'.join(item_models[:3]) if item_models else 'DROPSHIP'

        pn = PostNsop()
        batch_number = await pn.generate_unique_batch_number(destination_for_batch)

        total_pcs = sum(item.get('pcs', 0) for item in cargo_pcs_data)

        dropship_shipment = await sync_to_async(DropshipShipment.objects.create)(
            shipment_batch_number=batch_number,
            warehouse=warehouse_obj,
            created_at=current_time_beijing,
            pickup_time=pickup_time.date(),
            total_pcs=total_pcs,
            total_pallets=0,
            shipping_address='',
            contact_person='',
            contact_phone='',
            operator=await sync_to_async(lambda: request.user.username)()
        )

        inventory_records = []
        shipment_details = []
        updated_cargos = []

        for item in cargo_pcs_data:
            if 'cargo_id' not in item:
                continue
            try:
                cargo_id = int(item['cargo_id'])
            except (ValueError, TypeError):
                continue
            pcs = item.get('pcs', 0)
            
            cargo = cargo_map.get(cargo_id)
            if not cargo:
                continue
            
            if pcs <= 0 or pcs > cargo.pcs:
                continue

            after_pcs = cargo.pcs - pcs

            shipment_detail = DropshipShipmentDetail(
                shipment=dropship_shipment,
                cargo=cargo,
                pcs=pcs,
                pallets=0
            )
            shipment_details.append(shipment_detail)

            inventory = DropshipInventory(
                cargo=cargo,
                transaction_type='pick',
                pcs_change=-pcs,
                after_pcs=after_pcs,
                shipment_detail=shipment_detail,
                transaction_date=current_time_beijing,
                operator=await sync_to_async(lambda: request.user.username)(),
                is_verify=False
            )
            inventory_records.append(inventory)

            cargo.pcs = after_pcs
            updated_cargos.append(cargo)

        if shipment_details:
            await sync_to_async(DropshipShipmentDetail.objects.bulk_create)(shipment_details)

        if inventory_records:
            await sync_to_async(DropshipInventory.objects.bulk_create)(inventory_records)

        for cargo in updated_cargos:
            await sync_to_async(cargo.save)()

        success_msg = f'预约出库成功!<br>共出库 {total_pcs} 件<br>批次号: {batch_number}'
        context = {'success_messages': mark_safe(success_msg)}
        return await self.handle_ltl_unscheduled_pos_post(request, context)

    async def handle_dropship_inventory(
            self, request: HttpRequest, warehouse_name: str = ""
    ) -> dict[str, Any]:
        '''库存流水查询 - 根据仓库筛选DropshipInventory记录'''
        context = {
            "warehouse_options": await sync_to_async(list)(
                ZemWarehouse.objects
                .order_by("name")
                .values_list("name", "name")
            ),
            "selected_warehouse": warehouse_name,
            "inventory_records": [],
        }

        if warehouse_name:
            warehouse_obj = await sync_to_async(ZemWarehouse.objects.filter(name=warehouse_name).first)()
            if warehouse_obj:
                inventory_records = await sync_to_async(list)(
                    DropshipInventory.objects
                    .filter(cargo__warehouse=warehouse_obj)
                    .select_related('cargo', 'shipment_detail', 'shipment_detail__shipment')
                    .order_by('-transaction_date')
                )
                
                inventory_data = []
                for record in inventory_records:
                    shipment_batch = ""
                    if record.shipment_detail and record.shipment_detail.shipment:
                        shipment_batch = record.shipment_detail.shipment.shipment_batch_number
                    
                    transaction_type_display = dict(record.TRANSACTION_TYPES).get(record.transaction_type, record.transaction_type)
                    
                    inventory_data.append({
                        'id': record.id,
                        'transaction_type': transaction_type_display,
                        'transaction_type_code': record.transaction_type,
                        'transaction_date': record.transaction_date,
                        'shipping_mark': record.cargo.shipping_mark,
                        'model': record.cargo.model,
                        'pcs_change': record.pcs_change,
                        'after_pcs': record.after_pcs,
                        'verify_pcs': record.verify_pcs,
                        'operator': record.operator,
                        'is_verify': record.is_verify,
                        'shipment_batch': shipment_batch,
                    })
                
                context["inventory_records"] = inventory_data

        return context
