import asyncio
import base64
import io
import json
import os
import string
import uuid
from collections import defaultdict
from datetime import datetime, timedelta, date
from decimal import Decimal, InvalidOperation
from itertools import zip_longest
from pathlib import Path
import random
from typing import Any, Coroutine
import re

import barcode
import chardet
import numpy as np
import openpyxl
import pandas as pd
import pytz
from asgiref.sync import sync_to_async
from barcode.writer import ImageWriter
from django.contrib.postgres.aggregates import StringAgg
from django.core.exceptions import ObjectDoesNotExist
from django.db import models, transaction
from django.db.models import Sum, Count, FloatField, IntegerField, Min, Case, When, Q, Value, F, CharField, Prefetch, \
    Subquery, OuterRef
from django.db.models.functions import Coalesce, Round, Cast, Concat
from django.http import HttpRequest, HttpResponse, Http404
from django.shortcuts import render
from django.template.loader import get_template
from django.utils.dateparse import parse_date
from django.views import View
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from simple_history.utils import bulk_update_with_history, bulk_create_with_history
from xhtml2pdf import pisa

from warehouse.forms.packling_list_form import PackingListForm
from warehouse.forms.upload_file import UploadFileForm
from django.utils import timezone

from warehouse.forms.warehouse_form import ZemWarehouseForm
from warehouse.models.container import Container
from warehouse.models.container_pickup_carrier import ContainerPickupCarrier
from warehouse.models.customer import Customer
from warehouse.models.dropship_cargo import DropshipCargo
from warehouse.models.dropship_inventory import DropshipInventory
from warehouse.models.fee_detail import FeeDetail
from warehouse.models.fleet import Fleet
from warehouse.models.offload import Offload
from warehouse.models.offload_status import AbnormalOffloadStatus
from warehouse.models.order import Order
from warehouse.models.packing_list import PackingList
from warehouse.models.pallet import Pallet
from warehouse.models.pallet_destroyed import PalletDestroyed
from warehouse.models.quotation_master import QuotationMaster
from warehouse.models.retrieval import Retrieval
from warehouse.models.shipment import Shipment
from warehouse.models.transfer_location import TransferLocation
from warehouse.models.vessel import Vessel
from warehouse.models.warehouse import ZemWarehouse
from warehouse.utils.constants import SHIPPING_LINE_OPTIONS, ADDITIONAL_CONTAINER, \
    DROPSHIPPING_PACKING_LIST_TEMP_COL_MAPPING, DELIVERY_METHOD_CODE, DROPSHIPPING_DELIVERY_METHOD_OPTIONS
from warehouse.utils.shipment_binding_utils import ShipmentBindingLogger
from warehouse.views.export_file import export_do_dropshipping
from warehouse.views.pre_port.pre_port_dash import PrePortDash
from io import BytesIO
from openpyxl.worksheet.page import PageMargins

class Dropshipping(View):
    template_order_create_supplement = "dropshipping/03_order_creation.html"
    template_order_create_base = "dropshipping/02_base_order_creation_status.html"
    template_order_details = "dropshipping/order_details.html"
    template_order_details_pl = "dropshipping/order_details_pl_tab.html"
    template_order_create_supplement_pl_tab = "dropshipping/03_order_creation_packing_list_tab.html"
    template_order_list = "dropshipping/order_list.html"
    template_repeat_t49_all = "dropshipping/repeat_t49_all.html"
    template_terminal_dispatch = "dropshipping/02_terminal_dispatch.html"
    template_update_container_pickup_schedule = "dropshipping/04_update_container_pickup_schedule.html"
    template_schedule_container_pickup = "dropshipping/03_schedule_container_pickup.html"
    template_batch_update_container_pickup_schedule = "dropshipping/04_batch_update_container_pickup_schedule.html"
    template_batch_schedule_container = "dropshipping/03_batch_schedule_container_pickup.html"
    template_status_summary = "dropshipping/01_container_status_summary.html"
    template_palletization_main = "dropshipping/palletization.html"
    template_pallet_label = "export_file/pallet_label_template.html"
    template_palletize = "dropshipping/palletization_packing_list.html"
    template_inventory_management_main = "dropshipping/01_inventory_management_main.html"
    template_inventory_po_update = "dropshipping/02_inventory_po_update.html"
    template_counting_main = "dropshipping/01_inventory_count_main.html"
    template_pallet_abnormal_records_display = "dropshipping/palletization_abnormal_records_display.html"
    template_pallet_abnormal = "dropshipping/palletization_abnormal.html"
    template_pallet_abnormal_records_search = "dropshipping/palletization_abnormal_records_search.html"
    template_pallet_daily_operation = "dropshipping/daily_operation.html"

    container_type = {
        "": "",
        "40HQ/GP": "40HQ/GP",
        "45HQ/GP": "45HQ/GP",
        "20GP": "20GP",
        "53HQ": "53HQ",
    }

    order_type = {"一件代发": "一件代发"}
    area = {"NJ": "NJ", "SAV": "SAV", "LA": "LA",}
    is_shipment_batch_numbers = {
        "无论是否有约": "True",
        "否": "False"
    }

    async def get(self, request: HttpRequest, **kwargs) -> Any | None:
        step = request.GET.get("step", None)
        pk = kwargs.get("pk", None)
        if step == "all":
            template, context = await self.handle_order_basic_info_get()
            return await sync_to_async(render)(request, template, context)
        elif step == "container_info_supplement":
            template, context = await self.handle_order_supplemental_info_get(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "order_management_list":
            template, context = await self.handle_order_management_list_get()
            return await sync_to_async(render)(request, template, context)
        elif step == "repeat_t49_all":
            template, context = await self.repeat_t49_all()
            return await sync_to_async(render)(request, template, context)
        elif step == "order_management_container":
            template, context = await self.handle_order_management_container_get(
                request
            )
            return await sync_to_async(render)(request, template, context)
        elif step == "terminal_all":
            template, context = await self.handle_all_get_terminal()
            return await sync_to_async(render)(request, template, context)
        elif step == "update_pickup_schedule":
            template, context = await self.hanlde_update_pickup_schedule_get(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "schedule_container_pickup":
            template, context = await self.handle_schedule_container_pickup_get(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "contaier_pickup_status_all":
            template, context = await self.handle_all_get_contaier_pickup_status()
            return await sync_to_async(render)(request, template, context)
        elif step == "palletize_all":
            template, context = await self.handle_all_get_palletize(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "container_palletization":
            template, context = await self.handle_container_palletization_get(
                request, pk
            )
            await sync_to_async(lambda: request.user.is_authenticated)()
            return await sync_to_async(render)(request, template, context)
        elif step == "inventory_all":
            template, context = await self.handle_inventory_management_get()
            return await sync_to_async(render)(request,template,context)
        elif step == "abnormal":
            template, context = await self.handle_palletization_abnormal_get()
            return await sync_to_async(render)(request, template, context)
        elif step == "abnormal_records":
            return await sync_to_async(render)(request, self.template_pallet_abnormal_records_search, {"page_title": "异常拆柜记录",})



    async def post(self, request: HttpRequest, **kwargs) -> None | HttpResponse | tuple[Any, Any] | Any:
        step = request.POST.get("step")
        if step == "create_order_basic":
            template, context = await self.handle_create_order_basic_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "update_order_basic_info":
            template, context = await self.handle_update_order_basic_info_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "update_order_shipping_info":
            template, context = await self.handle_update_order_shipping_info_post(
                request
            )
            return await sync_to_async(render)(request, template, context)
        elif step == "upload_template":
            template, context = await self.handle_upload_template_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "download_template":
            return await self.handle_download_template_post()
        elif step == "update_order_packing_list_info_v1":
            template, context = await self.handle_update_order_packing_list_info_post_v1(
                request
            )
            return await sync_to_async(render)(request, template, context)
        elif step == "update_order_packing_list_info":
            template, context = await self.handle_update_order_packing_list_info_post(
                request
            )
            return await sync_to_async(render)(request, template, context)
        elif step == "export_forecast":
            return await self.handle_export_forecast(request)
        elif step == "order_management_search":
            start_date_eta = request.POST.get("start_date_eta")
            end_date_eta = request.POST.get("end_date_eta")
            start_date_etd = request.POST.get("start_date_etd")
            end_date_etd = request.POST.get("end_date_etd")
            template, context = await self.handle_order_management_list_get(
                start_date_eta, end_date_eta, start_date_etd, end_date_etd
            )
            return await sync_to_async(render)(request, template, context)
        elif step == "export_do":
            return await sync_to_async(export_doexport_do_dropshipping)(request)
        elif step == "export_details_by_destination":
            return await self.handle_export_details_by_destination(request)
        elif step == "delete_order":
            template, context = await self.handle_delete_order_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "add_t49_order":
            template, context = await self.add_t49_order(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "update_order_retrieval_info":
            template, context = await self.handle_update_order_retrieval_info_post(
                request
            )
            return await sync_to_async(render)(request, template, context)
        elif step == "cancel_notification":
            template, context = await self.handle_cancel_notification(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "check_destination":
            template, context = await self.handle_check_destination(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "check_order_status":
            template, context = await self.check_order_status(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "check_order_type_destination":
            template, context = await self.handle_check_order_type_destination(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "update_delivery_type_all":
            template, context = await self.handle_update_delivery_type(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "update_container_unpacking_priority":
            template, context = await self.handle_update_container_unpacking_priority(
                request
            )
            return await sync_to_async(render)(request, template, context)
        elif step == "batch_confirm_pickup":
            template, context = await self.handle_batch_confirm_pickup_get(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "batch_schedule_container":
            template, context = await self.handle_batch_schedule_container_get(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "pickup_schedule_confirmation":
            template, context = await self.handle_pickup_schedule_confirmation_post(
                request
            )
            return await sync_to_async(render)(request, template, context)
        elif step == "confirm_pickup":
            template, context = await self.handle_confirm_pickup_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "confirm_pickup_appointment_time":
            template, context = await self.handle_confirm_pickup_post_appointment_time(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "pickup_schedule_update":
            template, context = await self.handle_pickup_schedule_confirmation_post(
                request
            )
            return await sync_to_async(render)(request, template, context)
        elif step == "batch_confirm_pickup_submit":
            template, context = await self.handle_batch_confirm_pickup_submit_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "batch_confirm_pickup_submit_appointment_time":
            template, context = await self.handle_batch_confirm_pickup_submit_post_appointment_time(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "batch_pickup_schedule_confirmation":
            template, context = await self.handle_batch_pickup_schedule_confirmation(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "arrive_at_destination":
            template, context = await self.handle_arrive_at_destination_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "batch_arrive_at_destination":
            template, context = await self.handle_batch_arrive_at_destination_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "empty_return":
            template, context = await self.handle_empty_return_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "batch_empty_return":
            template, context = await self.handle_batch_empty_return_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "warehouse_palletize":
            template, context = await self.handle_warehouse_post_palletize(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "trans_arrival":
            template, context = await self.handle_trans_arrival_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "export_pallet_label":
            return await self._export_pallet_label(request)
        elif step == "export_palletization_list":
            return await self.export_palletization_list(request)
        elif step == "new_export_palletization_list":
            return await self.export_palletization_list_v2(request)
        elif step == "back":
            template, context = await self.handle_warehouse_post_palletize(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "palletize":
            pk = kwargs.get("pk")
            template, context = await self.handle_packing_list_post(request, pk)
            return await sync_to_async(render)(request, template, context)
        elif step == "warehouse_inventory":
            template, context = await self.handle_warehouse_post_inventory(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "export_inventory":
            return await self.handle_export_inventory(request)
        elif step == "update_po_page":
            template, context = await self.handle_update_po_page_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "repalletize":
            template, context = await self.handle_repalletize_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "update_po":
            template, context = await self.handle_update_po_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "warehouse_abnormal":
            template, context = await self.handle_warehosue_abnormal_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "amend_abnormal":
            template, context = await self.handle_amend_abnormal_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "warehouse_daily":
            template, context = await self.handle_warehouse_daily_post(request)
            return await sync_to_async(render)(request, template, context)
        elif step == "counting":
            template, context = await self.handle_counting_post(request)
            return await sync_to_async(render)(request, template, context)

    async def handle_amend_abnormal_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("warehouse")
        selected = request.POST.getlist("is_case_selected")
        abnormal_pl_ids = request.POST.getlist("ids")
        abnormal_reasons = request.POST.getlist("abnormal_reason")
        notes = request.POST.getlist("note")
        confirmed_by_warehouse = request.POST.getlist("confirmed_by_warehouse")

        abnormal_pl_ids = [
            abnormal_pl_ids[i] for i in range(len(selected)) if selected[i] == "on"
        ]
        abnormal_records = await sync_to_async(list)(
            AbnormalOffloadStatus.objects.select_related("container_number").filter(
                id__in=abnormal_pl_ids, delivery_type="一件代发"
            )
        )
        updated_records = []
        if confirmed_by_warehouse:
            shipment = set()
            for record in abnormal_records:
                pallet = await sync_to_async(list)(
                    Pallet.objects.select_related(
                        "container_number", "shipment_batch_number"
                    ).filter(
                        model=record.model,
                        container_number__container_number=record.container_number.container_number,
                        delivery_method=record.delivery_method,
                        delivery_type="一件代发"
                    )
                )
                shipment.update([p.shipment_batch_number for p in pallet])
                for p in pallet:
                    p.abnormal_palletization = False
                record.confirmed_by_warehouse = True
                updated_records.append(record)
                await sync_to_async(bulk_update_with_history)(
                    pallet,
                    Pallet,
                    fields=["abnormal_palletization"],
                )
            await sync_to_async(bulk_update_with_history)(
                updated_records,
                AbnormalOffloadStatus,
                fields=["confirmed_by_warehouse"],
            )
            await self._update_shipment_abnormal_palletization(shipment)
            return await self.handle_daily_operation_get()
        else:
            abnormal_reasons = [
                abnormal_reasons[i] for i in range(len(selected)) if selected[i] == "on"
            ]
            notes = [notes[i] for i in range(len(selected)) if selected[i] == "on"]
            for record, reason, note in zip(abnormal_records, abnormal_reasons, notes):
                record.note = note
                record.abnormal_reason = reason
                record.is_resolved = True
                updated_records.append(record)
            await sync_to_async(bulk_update_with_history)(
                updated_records,
                AbnormalOffloadStatus,
                fields=["is_resolved", "abnormal_reason", "note"],
            )
            if warehouse == "None":
                warehouse = ""
            return await self.handle_palletization_abnormal_get(warehouse)

    async def handle_daily_operation_get(
        self, warehouse: str = None, include_all: bool = False
    ) -> tuple[str, dict[str, Any]]:
        # 拆柜异常，客服已解决货物
        retrieval_precise_subquery = Subquery(
            Retrieval.objects.filter(
                id=OuterRef("container_number__orders__retrieval_id")
            ).values("retrieval_destination_precise")[:1],
            output_field=CharField(),
        )
        query = (
            AbnormalOffloadStatus.objects.select_related("container_number")
            .annotate(
                retrieval_destination_precise=Subquery(retrieval_precise_subquery)
            )
            .filter(is_resolved=True, confirmed_by_warehouse=False, delivery_type="一件代发")
            .order_by("created_at")
        )
        if warehouse:
            query = query.filter(retrieval_destination_precise=warehouse)
        all_status = await sync_to_async(list)(query)
        abnormal = []
        for status in all_status:
            status_dict = {
                "id": status.id,
                "container_number": status.container_number.container_number,
                "created_at": (
                    status.created_at.strftime("%b-%d") if status.created_at else None
                ),
                "resolved_at": status.resolved_at,
                "confirmed_by_warehouse": (
                    True if status.confirmed_by_warehouse else False
                ),
                "model": status.model,
                "delivery_method": status.delivery_method,
                "pcs_reported": status.pcs_reported,
                "pcs_actual": status.pcs_actual,
                "abnormal_reason": status.abnormal_reason,
                "note": status.note,
                "ddl_status": status.abnormal_status,
            }
            abnormal.append(status_dict)

        cn = pytz.timezone("Asia/Shanghai")
        current_time_cn = datetime.now(cn)
        today = current_time_cn.date()

        # 当日+下一天的预约信息
        query = Shipment.objects.prefetch_related(
            "packinglist",
            "packinglist__container_number",
            "packinglist__container_number__orders",
            "packinglist__container_number__orders__warehouse",
            "order",
        ).filter(shipment_schduled_at__date=today, packinglist__delivery_type="一件代发")
        if warehouse:
            query = query.filter(
                packinglist__container_number__orders__retrieval_id__retrieval_destination_precise=warehouse
            ).distinct()
        shipment = await sync_to_async(list)(query)

        # 当日+下一天的预约信息
        query = Fleet.objects.prefetch_related(
            "shipment",
            "shipment__packinglist",
            "shipment__packinglist__container_number",
            "shipment__packinglist__container_number__orders",
            "shipment__packinglist__container_number__orders__retrieval_id",
        ).filter(scheduled_at__date=today, shipment__packinglist__delivery_type="一件代发")
        if warehouse:
            query = query.filter(
                shipment__packinglist__container_number__orders__retrieval_id__retrieval_destination_precise=warehouse
            ).distinct()
        fleet = await sync_to_async(list)(query)

        # 当日到港货柜
        query = Order.objects.select_related(
            "vessel_id", "container_number", "customer_name", "retrieval_id"
        ).filter(
            (
                models.Q(retrieval_id__target_retrieval_timestamp__date=today)
                & models.Q(cancel_notification=False)
                & models.Q(delivery_type="一件代发")
            )
        )
        if warehouse:
            query = query.filter(retrieval_id__retrieval_destination_precise=warehouse)
        containers = await sync_to_async(list)(query)
        arrived_containers = []
        for o in containers:
            con_dict = {
                "id": o.id,
                "container_number": o.container_number.container_number,
                "order_type": o.order_type,
                "vessel_id": o.vessel_id,
                "temp_t49_pod_arrive_at": o.retrieval_id.temp_t49_pod_arrive_at,
            }
            arrived_containers.append(con_dict)
        context = {
            "abnormal": abnormal,
            "shipment": shipment,
            "fleet": fleet,
            "arrived_containers": arrived_containers,
            "warehouse_options": [("", "")] + await sync_to_async(list)(
                ZemWarehouse.objects
                .order_by("name")
                .values_list("name", "name")
            ),
            "warehouse_filter": warehouse,
        }
        return self.template_pallet_daily_operation, context

    async def handle_warehouse_daily_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("warehouse_filter")
        template, context = await self.handle_daily_operation_get(warehouse)
        return template, context

    async def handle_warehosue_abnormal_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("warehouse")
        template, context = await self.handle_palletization_abnormal_get(warehouse)
        return template, context

    async def handle_palletization_abnormal_get(
        self, warehouse: str = None, include_all: bool = False
    ) -> tuple[str, dict[str, Any]]:
        retrieval_precise_subquery = Subquery(
            Retrieval.objects.filter(
                id=OuterRef("container_number__orders__retrieval_id")
            ).values("retrieval_destination_precise")[:1],
            output_field=CharField(),
        )
        if include_all:
            all_status = await sync_to_async(list)(
                AbnormalOffloadStatus.objects.select_related(
                    "container_number", "offload"
                )
                .filter(delivery_type="一件代发")
                .annotate(
                    retrieval_destination_precise=Subquery(retrieval_precise_subquery)
                )
                .all()
                .order_by("created_at")
            )
        else:
            all_status = await sync_to_async(list)(
                AbnormalOffloadStatus.objects.select_related(
                    "container_number", "offload"
                )
                .filter(is_resolved=False, delivery_type="一件代发")
                .annotate(
                    retrieval_destination_precise=Subquery(retrieval_precise_subquery)
                )
                .all()
                .order_by("created_at")
            )
        abnormal = []
        for status in all_status:
            if warehouse:
                if warehouse in status.retrieval_destination_precise:
                    status_dict = {
                        "id": status.id,
                        "offload": status.offload.offload_id,
                        "container_number": status.container_number.container_number,
                        "created_at": (
                            status.created_at.strftime("%b-%d")
                            if status.created_at
                            else None
                        ),
                        "resolved_at": status.resolved_at,
                        "is_resolved": True if status.is_resolved else False,
                        "model": status.model,
                        "delivery_method": status.delivery_method,
                        "pcs_reported": status.pcs_reported,
                        "pcs_actual": status.pcs_actual,
                        "abnormal_reason": status.abnormal_reason,
                        "note": status.note,
                        "retrieval_destination_precise": status.retrieval_destination_precise,
                        "ddl_status": status.abnormal_status,
                    }
                    abnormal.append(status_dict)
            else:
                status_dict = {
                    "id": status.id,
                    "offload": status.offload.offload_id,
                    "container_number": status.container_number.container_number,
                    "created_at": (
                        status.created_at.strftime("%b-%d")
                        if status.created_at
                        else None
                    ),
                    "resolved_at": status.resolved_at,
                    "is_resolved": True if status.is_resolved else False,
                    "model": status.model,
                    "delivery_method": status.delivery_method,
                    "pcs_reported": status.pcs_reported,
                    "pcs_actual": status.pcs_actual,
                    "abnormal_reason": status.abnormal_reason,
                    "note": status.note,
                    "retrieval_destination_precise": status.retrieval_destination_precise,
                    "ddl_status": status.abnormal_status,
                }
                abnormal.append(status_dict)
        context = {
            "abnormal": abnormal,
            "warehouse": warehouse,
            "warehouse_options": [("", "")] + await sync_to_async(list)(
                ZemWarehouse.objects
                .order_by("name")
                .values_list("name", "name")
            ),
            "page_title": "待处理异常",
        }
        if include_all:
            return self.template_pallet_abnormal_records_display, context
        else:
            return self.template_pallet_abnormal, context

    @staticmethod
    def _is_hold_delivery_method(delivery_method: str | None) -> bool:
        return "暂扣留仓(HOLD)" in str(delivery_method)

    async def handle_update_po_post(
            self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        """批量更新货物PO信息"""
        plt_ids_raw = request.POST.get("plt_ids")
        plt_ids = [int(i) for i in plt_ids_raw.split(",")] if plt_ids_raw else []

        # 无选中货物直接返回
        if not plt_ids:
            return await self.handle_warehouse_post_inventory(request)

        # 判断是否销毁
        is_destroyed = request.POST.get("is_destroyed") == "True"
        if is_destroyed:
            await sync_to_async(DropshipCargo.objects.filter(id__in=plt_ids).delete)()
            return await self.handle_warehouse_post_inventory(request)

        # 表单字段清洗
        product_name = request.POST.get("product_name", "").strip()
        model_new = request.POST.get("model", "").strip()
        address_new = request.POST.get("address", "").strip()
        delivery_method_new = request.POST.get("delivery_method", "")
        delivery_type_new = request.POST.get("delivery_type", "")
        total_weight_new = round(float(request.POST.get("weight", 0)), 4)
        total_pcs_new = int(request.POST.get("pcs", 0))
        total_cbm_new = round(float(request.POST.get("cbm", 0)), 4)
        n_pallet_new = int(request.POST.get("n_pallet", 1))
        note_new = request.POST.get("note", "").strip()
        shipping_mark_list = request.POST.getlist("shipping_mark")
        shipping_mark_new = ",".join([s.strip() for s in shipping_mark_list if s.strip()])

        dropship_cargo = await sync_to_async(
            lambda: list(DropshipCargo.objects.filter(id__in=plt_ids, delivery_type="一件代发"))
        )()
        if not dropship_cargo:
            return await self.handle_warehouse_post_inventory(request)

        update_list = []
        for cargo in dropship_cargo:
            cargo.product_name = product_name
            cargo.model = model_new
            cargo.address = address_new
            cargo.delivery_method = delivery_method_new
            cargo.delivery_type = delivery_type_new
            cargo.total_weight_lbs = total_weight_new
            cargo.pcs = total_pcs_new
            cargo.cbm = total_cbm_new
            cargo.pallets = n_pallet_new
            cargo.note = note_new
            cargo.shipping_mark = shipping_mark_new
            update_list.append(cargo)

        await sync_to_async(bulk_update_with_history)(
            update_list,
            DropshipCargo,
            fields=[
                "product_name", "model", "address", "delivery_method", "delivery_type",
                "total_weight_lbs", "pcs", "cbm", "pallets", "note", "shipping_mark"
            ]
        )

        return await self.handle_warehouse_post_inventory(request)

    async def handle_counting_post(
            self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        plt_ids = request.POST.getlist("plt_ids")
        n_pallet = [int(i) for i in request.POST.getlist("n_pallet")]
        counted_n_pallet = [int(i) for i in request.POST.getlist("counted_n_pallet")]
        updated_pallets = []
        for ids, n, n_counted in zip(plt_ids, n_pallet, counted_n_pallet):
            if n > n_counted:
                pallet_ids = [int(i) for i in ids.split(",")]
                dropship_cargo = await sync_to_async(list)(
                    DropshipCargo.objects.filter(id__in=pallet_ids)
                )
                diff = n - n_counted
                for p in dropship_cargo[:diff]:
                    updated_pallets.append(p)
        if updated_pallets:
            await sync_to_async(bulk_update_with_history)(
                updated_pallets,
                DropshipCargo,
            )
        return await self.handle_warehouse_post_inventory(request)

    async def handle_repalletize_post(
            self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        """分拣/重打板操作"""
        plt_ids = request.POST.get("plt_ids")
        if not plt_ids:
            raise ValueError("未选择任何需要重打板的货物")
        plt_ids = [int(i) for i in plt_ids.split(",")]

        old_pallet = await sync_to_async(lambda: list(DropshipCargo.objects.filter(id__in=plt_ids)))()
        if not old_pallet:
            raise ValueError("选中的货物在系统中不存在")
        old_po_id = old_pallet[0].PO_ID

        cargo_list = await sync_to_async(
            lambda: list(DropshipCargo.objects.select_related("order", "container", "warehouse").filter(id__in=plt_ids))
        )()
        if not cargo_list:
            raise ValueError(f"DropshipCargo {plt_ids} 下无任何货物，无法执行重打板")
        # 取第一条的公共关联对象
        base_cargo = cargo_list[0]
        container = base_cargo.container
        order = base_cargo.order
        warehouse = base_cargo.warehouse

        total_weight = float(request.POST.get("weight"))
        total_cbm = float(request.POST.get("cbm"))
        total_pcs = int(request.POST.get("pcs"))
        # 前端表单数组
        models = request.POST.getlist("model_repalletize")
        product_names = request.POST.getlist("product_name_repalletize")
        delivery_methods = request.POST.getlist("delivery_method_repalletize")
        addresses = request.POST.getlist("address_repalletize")
        shipping_marks = request.POST.getlist("shipping_mark_repalletize")
        pcses = request.POST.getlist("pcs_repalletize")
        n_pallets = request.POST.getlist("n_pallet_repalletize")
        notes = request.POST.getlist("note_repalletize")


        pcses = [int(i) for i in pcses]
        n_pallets = [int(i) for i in n_pallets]
        n_pallets_total = sum(n_pallets)

        if n_pallets_total == 0:
            raise ValueError("总板子数不能为0，无法分摊CBM和重量！")

        new_pallets = []
        seq_num = 1

        for dest, product_name, dm, addr, sm, pcs, n_pallet, note in zip(
                models,
                product_names,
                delivery_methods,
                addresses,
                shipping_marks,
                pcses,
                n_pallets,
                notes,
        ):
            delivery_type = "一件代发"
            if n_pallet == 0:
                continue

            base_pcs = pcs // n_pallet
            remainder = pcs % n_pallet
            current_po_id = f"{old_po_id}_{seq_num}"

            # 拆分每一块板的数据
            for i in range(n_pallet):
                single_pcs = base_pcs + (1 if i < remainder else 0)
                single_cbm = round(total_cbm * (pcs / total_pcs) * (1 / n_pallet), 4)
                single_weight = round(total_weight * (pcs / total_pcs) * (1 / n_pallet), 2)

                new_pallets.append(DropshipCargo(
                    container=container,
                    order=order,
                    warehouse=warehouse,
                    model=dest,
                    product_name=product_name,
                    pallets=n_pallet,
                    status="in_stock",
                    address=addr,
                    delivery_method=dm,
                    pcs=single_pcs,
                    cbm=single_cbm,
                    total_weight_lbs=single_weight,
                    note=note,
                    shipping_mark=sm.strip() if sm else "",
                    PO_ID=current_po_id,
                    delivery_type=delivery_type,
                ))
            seq_num += 1

        if new_pallets:
            instances = new_pallets
            await sync_to_async(bulk_create_with_history)(instances, DropshipCargo)

        await sync_to_async(DropshipCargo.objects.filter(id__in=plt_ids).delete)()

        return await self.handle_warehouse_post_inventory(request)

    async def handle_update_po_page_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("warehouse")
        plt_ids = request.POST.get("plt_ids")
        plt_ids = [int(i) for i in plt_ids.split(",")]
        criteria_pallet = models.Q(
            id__in=plt_ids,
            delivery_type="一件代发"
        )
        dropship_cargo = await self._get_inventory_pallet(warehouse, criteria_pallet)
        context = {
            "dropship_cargo": dropship_cargo[0],
            "warehouse": warehouse,
            "delivery_type": [
            ("一件代发", "一件代发")],
            "delivery_method_options": DROPSHIPPING_DELIVERY_METHOD_OPTIONS,
            "plt_ids": ",".join([str(i) for i in plt_ids]),
        }
        return self.template_inventory_po_update, context

    async def handle_export_inventory(self, request: HttpRequest) -> HttpResponse:
        warehouse = request.POST.get("warehouse")
        pallet = await self._get_inventory_pallet(warehouse, models.Q(delivery_type="一件代发"))
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{warehouse}__报表"
        # 添加固定表头
        headers = [
            "客户名称",
            "柜号",
            "货物型号",
            "拆柜时间",
            "派送方式",
            "重量(kg)",
            "件数",
            "体积(CBM)",
            "托盘数",
            "备注",
        ]
        ws.append(headers)

        # 批量写入数据
        for p in pallet:
            # 处理运输方式特殊逻辑
            delivery_method = p.get("delivery_method", "")
            if "pickup" in delivery_method:
                delivery_method = f"{delivery_method} - {p.get('shipping_mark', '')}"

            # 按固定顺序构建行数据
            row = [
                p.get("customer_name", ""),
                p.get("container_number", ""),
                p.get("model", ""),
                p.get("offload_at", ""),
                delivery_method,
                round(float(p.get("weight", 0)), 2),
                int(float(p.get("pcs", 0))),
                round(float(p.get("cbm", 0)), 2),
                int(float(p.get("n_pallet", 0))),
                p.get("note", ""),
            ]
            ws.append(row)

        total_cbm = sum(float(p.get("cbm", 0)) for p in pallet)
        total_pallet = sum(int(float(p.get("n_pallet", 0))) for p in pallet)
        ws.append(
            ["总计", "", "", "", "", "", round(total_cbm, 2), total_pallet, "", "", ""]
        )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{warehouse}_库存报表.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"

        wb.save(response)
        return response

    async def handle_warehouse_post_inventory(
            self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("warehouse")
        dropship_cargo = await self._get_inventory_pallet(warehouse, models.Q(delivery_type="一件代发"))
        pallet_json = {
            p.get("plt_ids"): {
                k: (
                    # 1. 数字类型：保留并保留两位小数
                    round(v, 2) if isinstance(v, (float, int)) else
                    # 2. 时间类型：转为字符串
                    v.strftime("%Y-%m-%d %H:%M:%S") if isinstance(v, datetime) else
                    # 3. 字符串类型：清洗特殊字符
                    re.sub(r'[\x00-\x1F\x7F\t"\']', " ", str(v))
                    if v not in (None, "None") else ""
                )
                for k, v in p.items()
            }
            for p in dropship_cargo
        }

        total_cbm = sum([p.get("cbm", 0) for p in dropship_cargo])
        total_pallet = sum([p.get("n_pallet", 0) for p in dropship_cargo])

        context = {
            "warehouse": warehouse,
            "warehouse_options": [("", "")] + await sync_to_async(list)(
                ZemWarehouse.objects
                .order_by("name")
                .values_list("name", "name")
            ),
            "delivery_method_options": DROPSHIPPING_DELIVERY_METHOD_OPTIONS,
            "dropship_cargo": dropship_cargo,
            "total_cbm": round(total_cbm, 2),
            "total_pallet": total_pallet,
            "pallet_json": json.dumps(pallet_json, ensure_ascii=False),
        }
        return self.template_inventory_management_main, context


    async def _get_inventory_pallet(
        self, warehouse: str, criteria: models.Q | None = None
    ) -> list[Pallet]:
        # 固定基础条件：拆柜入库的状态
        base_q = models.Q(status="in_stock")

        if criteria:
            criteria &= base_q
        else:
            criteria = base_q
        criteria &= models.Q(warehouse__name=warehouse)

        return await sync_to_async(list)(
            DropshipCargo.objects.prefetch_related(
                "cargo_set", "warehouse", "container", "order", "order__customer_name"
            )
            .filter(criteria)
            .annotate(str_id=Cast("id", CharField()))

            # 格式化时间为 年月日
            .annotate(
                offload_date=Cast(
                    F("order__offload_id__offload_at"),
                    output_field=models.DateField()
                )
            )

            .values(
                "model",
                "delivery_method",
                "delivery_type",
                "shipping_mark",
                "note",
                "PO_ID",
                "address",
                "status",
                "product_name",
                customer_name=F("order__customer_name__zem_name"),
                warehouse_name=F("order__warehouse__name"),
                container_number=F("container__container_number"),
                offload_at=F("offload_date"),
                retrieval_destination_precise=F(
                    "order__retrieval_id__retrieval_destination_precise")
            )
            .annotate(
                plt_ids=StringAgg(
                    "str_id", delimiter=",", distinct=True, ordering="str_id"
                ),
                pcs=Sum("pcs", output_field=IntegerField()),
                cbm=Sum("cbm", output_field=FloatField()),
                weight=Sum("total_weight_lbs", output_field=FloatField()),
                n_pallet=Count("pallets", output_field=IntegerField()),
            )
            .order_by("-n_pallet")
        )

    async def handle_inventory_management_get(self) -> tuple[str, dict[str, Any]]:
        context = {
            "warehouse_options": [("", "")] + await sync_to_async(list)(
                ZemWarehouse.objects
                .order_by("name")
                .values_list("name", "name")
            ),
            "page_title": "库存管理",
        }
        return self.template_inventory_management_main, context

    async def handle_packing_list_post(
            self, request: HttpRequest, pk: int
    ) -> tuple[str, dict[str, Any]]:
        '''拆柜录入的确认'''
        order_selected = await sync_to_async(
            Order.objects.select_related(
                "offload_id", "warehouse", "container_number"
            ).prefetch_related("container_number__pallet_set").get
        )(pk=pk)
        offload = order_selected.offload_id
        container = order_selected.container_number
        additional_pallets = request.POST.getlist("new_models")
        if not offload.offload_at:
            offload_time = request.POST.get("offload_time")
            if not offload_time:
                offload_time = datetime.now()
            ids = request.POST.getlist("ids")
            ids = [i.split(",") for i in ids]
            n_pallet = [int(n) for n in request.POST.getlist("n_pallet")]
            pcs_actual = [int(n) for n in request.POST.getlist("pcs_actul")]
            pcs_reported = [int(d) for d in request.POST.getlist("pcs_reported")]
            cbm = [float(c) for c in request.POST.getlist("cbms")]
            weight = [float(c) for c in request.POST.getlist("weights")]
            product_names = [c for c in request.POST.getlist("product_names")]
            models = [d for d in request.POST.getlist("models")]
            addresses = [d for d in request.POST.getlist("address")]
            delivery_method = [d for d in request.POST.getlist("delivery_method")]
            delivery_type = [d for d in request.POST.getlist("delivery_type")]
            shipping_marks = request.POST.getlist("shipping_marks")
            # 因为库位只有LA仓库有，所以前端没传过来值，就构建一个空的
            slots = (
                request.POST.getlist("slots")
                if "slots" in request.POST
                else [None] * len(n_pallet)
            )
            notes = [d for d in request.POST.getlist("notes")]
            po_ids = request.POST.getlist("po_ids")
            total_pallet = sum(n_pallet)
            abnormal_offloads = []
            for (
                    n,
                    p_a,
                    p_r,
                    c,
                    w,
                    product_name,
                    model,
                    d_m,
                    d_t,
                    note,
                    shipping_mark,
                    addr,
                    po_id,
                    slot,
            ) in zip(
                n_pallet,
                pcs_actual,
                pcs_reported,
                cbm,
                weight,
                product_names,
                models,
                delivery_method,
                delivery_type,
                notes,
                shipping_marks,
                addresses,
                po_ids,
                slots,
            ):
                if p_a > 0:  # 如果实际箱数大于0，才构建板子的信息
                    created_at = offload_time
                if p_a != p_r:
                    abnormal_offloads.append(
                        {
                            "offload": offload,
                            "container_number": container,
                            "created_at": offload_time,
                            "is_resolved": False,
                            "model": model,
                            "delivery_method": d_m,
                            "pcs_reported": p_r,
                            "pcs_actual": p_a,
                        }
                    )
            if additional_pallets:
                # 如果有多货的情况，因为前端目前新增行的时候通过clone id="palletization-row-empty"的行，所以会增加input，值为空，所以下面就进行了去重工作
                # 计划是把多货的打板和正常预报的货一起做，但是因为多的input比较乱的插入在input中，不太好去重，所以就把新增的新命名了，然后直接去重
                new_models = request.POST.getlist("new_models")
                new_product_names = request.POST.getlist("new_product_names")
                new_delivery_method = request.POST.getlist("new_delivery_method")
                new_pcs_actul = [
                    int(value) for value in request.POST.getlist("new_pcs_actul")
                ]
                new_pallets = [
                    int(value) for value in request.POST.getlist("new_pallets")
                ]
                shipping_marks = request.POST.getlist("new_shipping_marks")
                new_slots = request.POST.getlist("new_slots")
                new_notes = request.POST.getlist("new_notes")
                new_cbm = [
                    float(value) if value else 0
                    for value in request.POST.getlist("new_cbms")
                ]
                # 生成新的PO_ID
                new_po_ids = []
                seq_num = 0
                for dm, dest in zip(new_delivery_method, new_models):
                    if dm == "pickup" or dest == "pickup":
                        po_id_seg = f"S{''.join(random.choices(string.ascii_letters.upper() + string.digits, k=4))}"
                    else:
                        po_id_seg = f"{DELIVERY_METHOD_CODE.get(dm, 'UN')}{''.join(random.choices(string.ascii_letters.upper() + string.digits, k=4))}"
                    random.seed(container.container_number[-4:])
                    random_code = "".join(
                        random.choices(string.ascii_uppercase + string.digits, k=6)
                    )
                    new_po_ids.append(f"A{random_code}{po_id_seg}{seq_num}")
                    seq_num += 1

                for (
                        n,
                        p_a,
                        c,
                        product_name,
                        model,
                        d_m,
                        note,
                        shipping_mark,
                        po_id,
                        slot,
                ) in zip(
                    new_pallets,
                    new_pcs_actul,
                    new_cbm,
                    new_product_names,
                    new_models,
                    new_delivery_method,
                    new_notes,
                    shipping_marks,
                    new_po_ids,
                    new_slots,
                ):
                    created_at = offload_time
                    # 记录异常拆柜
                    abnormal_offloads.append(
                        {
                            "offload": offload,
                            "container_number": container,
                            "created_at": offload_time,
                            "is_resolved": False,
                            "model": model,
                            "delivery_method": d_m,
                            "pcs_reported": 0,
                            "pcs_actual": p_a,
                        }
                    )
            offload.total_pallet = total_pallet
            offload.offload_at = offload_time
            await sync_to_async(offload.save)()


            ids_flat = []
            for item in ids:
                if isinstance(item, list):
                    ids_flat.extend(item)
                else:
                    ids_flat.append(item)
            ids_flat = [int(i) for i in ids_flat if str(i).strip().isdigit()]
            if not ids_flat:
                return

            n_pallet_flat = []
            for item in n_pallet:
                if isinstance(item, list):
                    n_pallet_flat.extend(item)
                else:
                    n_pallet_flat.append(item)
            id_to_n = dict(zip(ids_flat, n_pallet_flat))

            cargo_list = await sync_to_async(lambda: list(DropshipCargo.objects.filter(id__in=ids_flat)))()
            update_list = []
            inventory_create_list = []

            for cargo in cargo_list:
                pallet_num = id_to_n.get(cargo.id, 0)
                cargo.pallets = pallet_num
                cargo.status = "in_stock"
                update_list.append(cargo)
                inventory_create_list.append(DropshipInventory(
                    cargo=cargo,
                    transaction_type="unpack",
                    pcs_change=cargo.pcs,
                    transaction_date=created_at,
                ))

            # 批量更新&批量创建
            if update_list:
                await sync_to_async(DropshipCargo.objects.bulk_update)(update_list, ["pallets"])
            if inventory_create_list:
                await sync_to_async(DropshipInventory.objects.bulk_create)(inventory_create_list)

            abnormal_offload_instances = [
                AbnormalOffloadStatus(**d) for d in abnormal_offloads
            ]
            await sync_to_async(bulk_create_with_history)(
                abnormal_offload_instances, AbnormalOffloadStatus
            )

        # 更新柜子的delivery_type
        co = await sync_to_async(Container.objects.get, thread_sensitive=True)(
            container_number=container.container_number
        )
        co.delivery_type = "一件代发"
        await sync_to_async(co.save, thread_sensitive=True)()

        mutable_post = request.POST.copy()
        mutable_post["name"] = order_selected.warehouse.name
        request.POST = mutable_post
        return await self.handle_warehouse_post_palletize(request)

    async def _split_pallet(
            self,
            order: Order,
            n: int,
            p_a: int,
            p_r: int,
            c: float,
            w: float,
            product_name: str,
            model: str,
            delivery_method: str,
            delivery_type: str,
            note: str,
            shipping_mark: str,
            po_id: str,
            pk: int,
            address: str | None = None,
            slot: str | None = None,
            seed: int = 0,
            created_at=None,
            released_at=None,
    ) -> tuple[list[Any], None] | list[Any]:
        if n == 0 or n is None:
            return [], None
        pallet_ids = [
            str(
                uuid.uuid3(
                    uuid.NAMESPACE_DNS, str(uuid.uuid4()) + str(pk) + str(i) + str(seed)
                )
            )
            for i in range(n)
        ]
        if p_r == 0:  # 多货的货物
            cbm_actual = c
            weight_actual = 0
        else:
            cbm_actual = c * p_a / p_r
            weight_actual = w * p_a / p_r
        pallet_data = []
        pallet_pcs = [p_a // n for _ in range(n)]
        for i in range(p_a % n):
            pallet_pcs[i] += 1
        for i in range(n):
            cbm_loaded = cbm_actual * pallet_pcs[i] / p_a
            weight_loaded = weight_actual * pallet_pcs[i] / p_a
            pallet_data.append(
                {
                    "container_number": order.container_number,
                    "product_name": product_name,
                    "model": model,
                    "address": address,
                    "delivery_method": delivery_method,
                    "delivery_type": delivery_type,
                    "pallet_id": pallet_ids[i],
                    "pcs": pallet_pcs[i],
                    "cbm": cbm_loaded,
                    "weight_lbs": weight_loaded,
                    "note": None if note == "None" else note,
                    "shipping_mark": shipping_mark if shipping_mark else "",
                    "abnormal_palletization": p_a != p_r,
                    "location": order.warehouse.name,
                    "PO_ID": po_id,
                    "slot": slot,
                    "created_at": created_at,
                    "released_at": released_at,
                }
            )
        return pallet_data


    async def _update_fleet_stats(self, shipment_batch_number: list[str]) -> None:
        fleet = await sync_to_async(list)(
            Fleet.objects.filter(
                shipment__shipment_batch_number__in=shipment_batch_number
            )
        )
        if fleet:
            fleet_number = [f.fleet_number for f in fleet]
            fleet_stats = await sync_to_async(list)(
                Pallet.objects.select_related(
                    "shipment_batch_number", "shipment_batch_number__fleet_number"
                )
                .filter(
                    shipment_batch_number__fleet_number__fleet_number__in=fleet_number
                )
                .values("shipment_batch_number__fleet_number__fleet_number")
                .annotate(
                    total_pcs=Sum("pcs", output_field=IntegerField()),
                    total_cbm=Sum("cbm", output_field=FloatField()),
                    weight_lbs=Sum("weight_lbs", output_field=FloatField()),
                    total_n_pallet=Count(
                        "pallet_id", distinct=True, output_field=IntegerField()
                    ),
                )
            )
            fleet_stats = {
                f["shipment_batch_number__fleet_number__fleet_number"]: {
                    "total_pcs": f["total_pcs"],
                    "total_cbm": f["total_cbm"],
                    "weight_lbs": f["weight_lbs"],
                    "total_n_pallet": f["total_n_pallet"],
                }
                for f in fleet_stats
            }
            for f in fleet:
                f.total_cbm = fleet_stats[f.fleet_number]["total_cbm"]
                f.total_pallet = fleet_stats[f.fleet_number]["total_n_pallet"]
                f.total_weight = fleet_stats[f.fleet_number]["weight_lbs"]
                f.total_pcs = fleet_stats[f.fleet_number]["total_pcs"]
            await sync_to_async(bulk_update_with_history)(
                fleet,
                Fleet,
                fields=["total_cbm", "total_pallet", "total_weight", "total_pcs"],
            )


    async def _update_shipment_abnormal_palletization(
        self, shipment: set[Shipment]
    ) -> None:
        abnormal_shipment = await sync_to_async(list)(
            Shipment.objects.filter(
                shipment_batch_number__in=[
                    p.shipment_batch_number for p in shipment if p
                ],
                pallet__abnormal_palletization=True,
            )
        )
        abnormal_shipment = set(s.shipment_batch_number for s in abnormal_shipment)
        updated_shipment = []
        for s in list(shipment - abnormal_shipment):
            if s:
                s.abnormal_palletization = False
                updated_shipment.append(s)
        await sync_to_async(bulk_update_with_history)(
            updated_shipment,
            Shipment,
            fields=["abnormal_palletization"],
        )

    async def export_palletization_list_v2(self, request: HttpRequest) -> HttpResponse:
        """
        (新)拆柜单导出
        """
        status = request.POST.get("status")
        warehouse = request.POST.get("warehouse").split("-")[0].upper() if request.POST.get("warehouse") else ""
        container_number = request.POST.get("container_number")
        offload_id = request.POST.get("offload_id")

        WAREHOUSE_TIMEZONE = {
            "NJ": "America/New_York",
            "SAV": "America/New_York",
            "LA": "America/Los_Angeles",
        }

        warehouse_unpacking_time = None

        try:
            tz_name = WAREHOUSE_TIMEZONE.get(warehouse, "UTC")
            tz = pytz.timezone(tz_name)

            warehouse_unpacking_time = (
                datetime.now(tz)
                .strftime("%Y-%m-%d")
            )
        except:
            pass

        if warehouse_unpacking_time and offload_id:
            try:
                # 异步获取单条记录
                offload = await sync_to_async(
                    Offload.objects.get, thread_sensitive=True
                )(id=offload_id)

                # 只有原来没有时间，才覆盖（避免重复更新）
                if not offload.warehouse_unpacking_time:
                    offload.warehouse_unpacking_time = warehouse_unpacking_time
                    await sync_to_async(offload.save, thread_sensitive=True)()

            except Offload.DoesNotExist:
                pass

        TARGET_WAREHOUSES = {"GEU3", "GYR2", "GYR3", "LAX9", "LGB8", "SBD1"}
        UTC_TZ = pytz.UTC
        BASE_ETA = UTC_TZ.localize(datetime(2026, 1, 19))

        vessel_prefetch_queryset = Vessel.objects.all()
        retrieval_prefetch_queryset = Retrieval.objects.all()
        dropship_cargo = await sync_to_async(list)(
            DropshipCargo.objects.select_related("container")
            .prefetch_related(
                Prefetch(
                    "order__vessel_id",
                    queryset=vessel_prefetch_queryset
                )
            )
            .prefetch_related(
                Prefetch(
                    "order__retrieval_id",
                    queryset=retrieval_prefetch_queryset
                )
            )
            .filter(container__container_number=container_number)
            .annotate(
                custom_delivery_method=Case(
                    When(
                        Q(delivery_method="pickup") & ~Q(model="pickup"),
                        then=Concat(
                            "delivery_method",
                            Value("-"),
                            "model",
                        ),
                    ),
                    When(
                        Q(delivery_method="pickup") | Q(model="pickup"),
                        then=Concat(
                            "delivery_method",
                            Value("-"),
                            "model",
                            Value("-"),
                            "shipping_mark",
                        ),
                    ),
                    default=F("delivery_method"),
                    output_field=CharField(),
                ),
                str_id=Cast("id", CharField()),
                str_shipping_mark=Cast("shipping_mark", CharField()),
                vessel_eta=F("order__vessel_id__vessel_eta"),
                retrieval_destination_area=F(
                    "order__retrieval_id__retrieval_destination_area"
                ),
            )
            .values(
                "container__container_number",
                "model",
                "address",
                "custom_delivery_method",
                "note",
                "PO_ID",
                "delivery_type",
                "vessel_eta",
                "retrieval_destination_area",
            )
            .annotate(
                shipping_marks=StringAgg(
                    "str_shipping_mark", delimiter=",", distinct=True
                ),
                ids=StringAgg("str_id", delimiter=",", distinct=True),
                pcs=Sum("pcs", output_field=IntegerField()),
                cbm=Sum("cbm", output_field=FloatField()),
                n_pallet=Sum("pallets", output_field=IntegerField()),
                weight_lbs=Sum("total_weight_lbs", output_field=FloatField()),
                plt_ids=StringAgg(
                    "str_id", delimiter=",", distinct=True, ordering="str_id"
                ),
            )
            .order_by("-cbm")
        )

        data = [i for i in dropship_cargo]
        df = pd.DataFrame.from_records(data)
        df["拆柜备注"] = ""

        if not df.empty:
            df["vessel_eta_dt"] = pd.to_datetime(df["vessel_eta"], errors="coerce", utc=True)
            mask = (
                    (df["vessel_eta_dt"] >= BASE_ETA)
                    & (df["model"].isin(TARGET_WAREHOUSES))
            )
            df.loc[mask, "拆柜备注"] = "100 height"
            df["note"] = df["note"].fillna("").astype(str)

            def merge_note_to_remark(row):
                remark = str(row["拆柜备注"] or "").strip()
                note = str(row["note"] or "").strip()

                # 原有逻辑：拼接note
                if note.strip():
                    if remark == "100 height":
                        remark = f"{remark}, {note.strip()}"
                    else:
                        remark = note.strip()
                return remark

            df["拆柜备注"] = df.apply(merge_note_to_remark, axis=1)
            df["note"] = ""

            df = df.rename(
                {
                    "container__container_number": "container_number",
                    "custom_delivery_method": "delivery_method",
                    "shipping_marks": "shipping_mark",
                    "n_pallet": "pl"
                },
                axis=1,
            )
            df["delivery_method"] = df["delivery_method"].apply(
                lambda x: x.split("-")[0] if isinstance(x, str) else x
            )
            df["pcs_original"] = df["pcs"].astype(str)
            df["pcs"] = df["pcs_original"].copy()

            def extract_original_note(remark):
                if not isinstance(remark, str) or remark.strip() == "":
                    return ""
                note_part = remark.replace("100 height, ", "").replace("100 height", "")
                return note_part.strip()

            df["original_note_from_remark"] = df["拆柜备注"].apply(extract_original_note)

            df = df.drop("original_note_from_remark", axis=1)

            df["pl"] = ""  # 清空打板数字段

            df = df[["model", "delivery_method", "shipping_mark", "pcs", "pl", "note", "拆柜备注"]]

        else:
            df = pd.DataFrame(columns=["model", "delivery_method", "shipping_mark", "pcs", "pl", "note", "拆柜备注"])

        df = df[["model", "delivery_method", "shipping_mark", "拆柜备注", "pcs", "pl", "note"]]
        buffer = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "拆柜单"

        # 纸张设为 Letter
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        # 纵向打印（已修改）
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        # 缩放到 1页宽，保证不超出
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # 缩小边距，给内容更多空间
        ws.page_margins = PageMargins(
            left=0.4, right=0.4, top=0.4, bottom=0.4,
            header=0.2, footer=0.2
        )

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header1_font = Font(size=15, bold=True)
        header2_font = Font(size=11, bold=True)
        data_font = Font(size=12, bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        UNIFIED_ROW_HEIGHT = 30

        # 第一行合并单元格（柜号）
        ws.merge_cells('A1:C1')
        for col_idx in range(1, 4):
            ws.cell(1, col_idx).border = thin_border
        ws['A1'] = container_number or "未指定柜号"
        ws['A1'].font = header1_font
        ws['A1'].alignment = left_alignment
        ws.row_dimensions[1].height = UNIFIED_ROW_HEIGHT

        # 第一行合并单元格（拆柜时间）
        ws.merge_cells('D1:E1')
        for col_idx in range(4, 6):
            ws.cell(1, col_idx).border = thin_border
        ws['D1'] = warehouse_unpacking_time
        ws['D1'].font = header1_font
        ws['D1'].alignment = center_alignment

        # 第一行合并单元格（dock）
        ws.merge_cells('F1:G1')
        for col_idx in range(6, 8):
            ws.cell(1, col_idx).border = thin_border
        ws['F1'] = 'dock'
        ws['F1'].font = header1_font
        ws['F1'].alignment = center_alignment

        # 表头行（第二行）
        column_names = ["model", "delivery_method", "shipping_mark", "拆柜备注", "pcs", "pl", "note"]
        for col_idx, name in enumerate(column_names, 1):
            cell = ws.cell(2, col_idx, name)
            cell.font = header2_font
            cell.border = thin_border
            cell.alignment = center_alignment
        ws.row_dimensions[2].height = UNIFIED_ROW_HEIGHT

        # 数据行
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 3):
            ws.row_dimensions[row_idx].height = UNIFIED_ROW_HEIGHT
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value if pd.notna(value) else "")
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = left_alignment if col_idx == 6 else center_alignment

        for col_idx in range(1, len(column_names) + 1):
            letter = get_column_letter(col_idx)
            max_len = max(len(str(ws.cell(row, col_idx).value or "")) for row in range(1, ws.max_row + 1))

            if col_idx == 5:  # pcs
                ws.column_dimensions[letter].width = 15
            elif col_idx == 6:  # pl
                ws.column_dimensions[letter].width = 10
            elif col_idx == 7:  # note
                ws.column_dimensions[letter].width = 30
            else:
                ws.column_dimensions[letter].width = min(max_len + 2, 22)

        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{container_number if container_number else '拆柜单'}.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    async def export_palletization_list(self, request: HttpRequest) -> HttpResponse:
        container_number = request.POST.get("container_number")
        warehouse = request.POST.get("warehouse").split("-")[0].upper() if request.POST.get("warehouse") else ""
        offload_id = request.POST.get("offload_id")
        WAREHOUSE_TIMEZONE = {
            "NJ": "America/New_York",
            "SAV": "America/New_York",
            "LA": "America/Los_Angeles",
        }

        warehouse_unpacking_time = None

        try:
            tz_name = WAREHOUSE_TIMEZONE.get(warehouse, "UTC")
            tz = pytz.timezone(tz_name)

            warehouse_unpacking_time = (
                datetime.now(tz)
                .strftime("%Y-%m-%d")
            )
        except:
            pass

        if warehouse_unpacking_time and offload_id:
            try:
                # 异步获取单条记录
                offload = await sync_to_async(
                    Offload.objects.get, thread_sensitive=True
                )(id=offload_id)

                # 只有原来没有时间，才覆盖（避免重复更新）
                if not offload.warehouse_unpacking_time:
                    offload.warehouse_unpacking_time = warehouse_unpacking_time
                    await sync_to_async(offload.save, thread_sensitive=True)()

            except Offload.DoesNotExist:
                pass

        dropship_cargo = await sync_to_async(list)(
            DropshipCargo.objects.select_related("container")
            .filter(container__container_number=container_number)
            .annotate(
                custom_delivery_method=Case(
                    When(
                        Q(delivery_method="pickup") | Q(model="pickup"),
                        then=Concat(
                            "delivery_method",
                            Value("-"),
                            "model",
                            Value("-"),
                            "shipping_mark",
                        ),
                    ),
                    default=F("delivery_method"),
                    output_field=CharField(),
                ),
                str_id=Cast("id", CharField()),
                str_shipping_mark=Cast("shipping_mark", CharField()),
            )
            .values(
                "container__container_number",
                "model",
                "address",
                "custom_delivery_method",
                "note",
                "PO_ID",
            )
            .annotate(
                shipping_marks=StringAgg(
                    "str_shipping_mark", delimiter=",", distinct=True
                ),
                ids=StringAgg("str_id", delimiter=",", distinct=True),
                pcs=Sum("pcs", output_field=IntegerField()),
                cbm=Round(Sum("cbm"), 2, output_field=FloatField()),
                n_pallet=Sum("pallets", output_field=IntegerField()),
                weight_lbs=Sum("total_weight_lbs", output_field=FloatField()),
                plt_ids=StringAgg(
                    "str_id", delimiter=",", distinct=True, ordering="str_id"
                ),
            )
            .order_by("-cbm")
        )

        data = [i for i in dropship_cargo]
        df = pd.DataFrame.from_records(data)
        df = df.rename(
            {
                "container__container_number": "container_number",
                "custom_delivery_method": "delivery_method",
                "shipping_marks": "shipping_mark",
            },
            axis=1,
        )
        df["delivery_method"] = df["delivery_method"].apply(lambda x: x.split("-")[0])
        df = df[
            [
                "container_number",
                "model",
                "delivery_method",
                "shipping_mark",
                "pcs",
                "cbm",
                "n_pallet",
                "PO_ID",
                "note",
            ]
        ]
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={container_number}.xlsx"
        df.to_excel(excel_writer=response, index=False, columns=df.columns)
        return response

    async def _export_pallet_label(self, request: HttpRequest) -> HttpResponse:
        data = []
        container_number = request.POST.get("container_number")
        customerInfo = request.POST.get("customerInfo")

        # ======================
        # 内置：唛头自动截断（最多 3 行，多余省略）
        # ======================
        def truncate_marks(marks_str, max_lines=3, max_chars=80):
            if not marks_str:
                return marks_str
            # 分离 TTT 后缀（保留不删）
            ttt_part = ""
            if "TTT" in marks_str:
                idx = marks_str.index("TTT")
                ttt_part = marks_str[idx:]
                marks_str = marks_str[:idx]
            # 按行切割
            lines = [line.strip() for line in marks_str.split("\n") if line.strip()]
            # 保留前 N 行
            keep_lines = lines[:max_lines]
            # 每行超长截断
            final = []
            for line in keep_lines:
                if len(line) > max_chars:
                    final.append(line[:max_chars] + "...")
                else:
                    final.append(line)
            # 拼接回去
            return "\n".join(final) + ttt_part

        if customerInfo:
            customer_info = json.loads(customerInfo)
            for row in customer_info:
                if len(row) > 10:
                    is_hold = row[10].strip()
                else:
                    is_hold = row[7].strip()
                date_str = row[6].strip()
                parts = date_str.split("-")
                month_day = f"{parts[1]}-{parts[2]}"

                model = f"{row[3].strip()}"
                shipping_marks = row[1].strip()
                new_marks = None

                if "pickup" in model:
                    model = "S/P"
                    marks = row[1].strip()
                    if marks:
                        array = marks.split(",")
                        if len(array) > 2:
                            parts = []
                            for i in range(0, len(array), 2):
                                part = ",".join(array[i: i + 2])
                                parts.append(part)
                            new_marks = "\n".join(parts)
                            newline_count = new_marks.count("\n") + 1
                            new_marks = new_marks + "TTT" + str(newline_count)
                        else:
                            new_marks = shipping_marks + "TTT1"
                elif is_hold == "是":
                    new_marks = shipping_marks
                else:
                    model = model.replace("Walmart", "WMT-").replace("沃尔玛", "WMT-").replace("WALMART",
                                                                                                           "WMT-")
                    new_marks = None

                # 生成后 强制截断唛头
                if new_marks:
                    new_marks = truncate_marks(new_marks)

                for num in range(int(row[5])):
                    num += 1
                    # 生成条形码
                    barcode_type = "code128"
                    barcode_class = barcode.get_barcode_class(barcode_type)
                    barcode_content = f"{row[0].strip()}|{model}-{num}"
                    my_barcode = barcode_class(barcode_content, writer=ImageWriter())
                    buffer = io.BytesIO()
                    my_barcode.write(buffer, options={"dpi": 300})
                    buffer.seek(0)
                    barcode_base64 = base64.b64encode(buffer.read()).decode("utf-8")

                    new_data = {
                        "container_number": row[0].strip(),
                        "model": f"{model}-{num}",
                        "date": month_day,
                        "customer": row[1].strip(),
                        "hold": (is_hold == "是"),
                        "barcode": barcode_base64,
                        "shipping_marks": new_marks,
                    }
                    if len(row) > 10:
                        new_data.update({
                            "has_delivery_window": True,
                        })
                    else:
                        new_data.update({
                            "has_delivery_window": False,
                        })
                    for i in range(4):
                        data.append(new_data)
        else:
            customer_name = request.POST.get("customer_name")
            status = request.POST.get("status")
            n_label = int(request.POST.get("n_label", 4))
            retrieval = await sync_to_async(Retrieval.objects.get)(
                order__container_number__container_number=container_number
            )
            retrieval_date = retrieval.target_retrieval_timestamp
            retrieval_destination_area = retrieval.retrieval_destination_area
            # 转换成当地时间
            preportdash = PrePortDash()
            retrieval_date = preportdash._convert_utc_to_local(retrieval_date, retrieval_destination_area)
            if retrieval_date:
                # 如果是字符串，先转成 datetime
                if isinstance(retrieval_date, str):
                    retrieval_date = datetime.strptime(retrieval_date, "%Y-%m-%d %H:%M:%S")
                # 现在一定是 datetime，直接格式化
                retrieval_date = retrieval_date.strftime("%m/%d")
            else:
                # 空值，取当前时间
                retrieval_date = datetime.now().strftime("%m/%d")
            dropship_cargo = await self._get_packing_list(
                container_number=container_number, status=status
            )
            for pl in dropship_cargo:
                delivery_method = pl.get("custom_delivery_method") or pl.get("delivery_method", "")
                pcs = pl.get("pcs") or 0

                remainder = pcs % 1
                pcs = int(pcs)
                if pcs % 2:
                    pcs += pcs % 2
                elif remainder:
                    pcs += 2

                pcs /= 25
                pcs *= n_label
                label_count = int(pcs)
                if "pickup" in pl.get("model"):
                    model = "S/P"
                else:
                    model = pl.get("model")
                marks = pl.get("shipping_marks") or pl.get("shipping_mark")
                new_marks = None
                if marks:
                    array = marks.split(",")
                    if len(array) > 2:
                        parts = []
                        for i in range(0, len(array), 2):
                            part = ",".join(array[i: i + 2])
                            parts.append(part)
                        new_marks = "\n".join(parts)
                        newline_count = new_marks.count("\n") + 1
                        new_marks = new_marks + "TTT" + str(newline_count)
                    else:
                        new_marks = marks + "TTT1"


                # 生成后 强制截断唛头
                if new_marks:
                    new_marks = truncate_marks(new_marks)


                for num in range(label_count):
                    i = num // n_label + 1
                    barcode_type = "code128"
                    barcode_class = barcode.get_barcode_class(barcode_type)
                    model = model.replace('\xa0', ' ')
                    barcode_content = f"{pl.get('container_number__container_number')}|{model}-{i}"
                    try:
                        my_barcode = barcode_class(barcode_content, writer=ImageWriter())
                    except:
                        barcode_content = barcode_content.encode("ascii", "ignore").decode()
                        my_barcode = barcode_class(barcode_content, writer=ImageWriter())
                    buffer = io.BytesIO()
                    my_barcode.write(buffer, options={"dpi": 300})
                    buffer.seek(0)
                    barcode_base64 = base64.b64encode(buffer.read()).decode("utf-8")

                    new_data = {
                        "container_number": pl.get("container__container_number"),
                        "model": f"{model}-{i}",
                        "date": retrieval_date,
                        "customer": customer_name,
                        "hold": ("暂扣留仓" in delivery_method.split("-")[0]),
                        "barcode": barcode_base64,
                        "shipping_marks": new_marks,
                        "pcs": pl.get("pcs"),
                    }
                    data.append(new_data)

        context = {"data": data}
        template = get_template(self.template_pallet_label)
        html = template.render(context)
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="pallet_label_{container_number}.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            raise ValueError(
                "Error during PDF generation: %s" % pisa_status.err,
                content_type="text/plain",
            )
        return response

    async def handle_container_palletization_get(
        self, request: HttpRequest, pk: int
    ) -> tuple[str, dict[str, Any]]:

        order_selected = await sync_to_async(
            Order.objects.select_related(
                "container_number", "warehouse", "offload_id"
            ).get
        )(pk=pk)

        container = order_selected.container_number
        offload = order_selected.offload_id
        order_packing_list = []
        step = request.GET.get("step", "")

        # 默认值，防止未进入条件导致 packing_list 未定义
        packing_list = []
        context = {"status": "palletized"}

        if step == "container_palletization":
            # 未拆柜
            if offload.offload_at is None:
                packing_list = await self._get_packing_list(
                    container_number=container.container_number, status="non_palletized"
                )
                context = {"status": "non_palletized"}

            # 已拆柜
            else:
                packing_list = await self._get_packing_list(
                    container_number=container.container_number, status="palletized"
                )
                context = {"status": "palletized"}

        # 组装列表
        for pl in packing_list:
            pl["PO_ID"] = pl.get("PO_ID", "")
            pl_form = PackingListForm(initial={"n_pallet": pl.get("n_pallet", 0)})
            order_packing_list.append((pl, pl_form))

        # 上下文
        context.update({
            "warehouse": request.GET.get("warehouse"),
            "order_packing_list": order_packing_list,
            "delivery_method_options": DROPSHIPPING_DELIVERY_METHOD_OPTIONS,
            "container_number": container.container_number,
            "pk": pk
        })

        return self.template_palletize, context

    async def _get_packing_list(
        self, container_number: str, status: str
    ) -> DropshipCargo:
        if status == "non_palletized":
            return await sync_to_async(list)(
                DropshipCargo.objects.select_related("container", "order", "order__offload_id")
                .filter(container__container_number=container_number, order__offload_id__offload_at__isnull=True)
                .annotate(
                    custom_delivery_method=Case(
                        When(
                            Q(delivery_method="pickup") | Q(model="pickup"),
                            then=Concat(
                                "delivery_method",
                                Value("-"),
                                "model",
                                Value("-"),
                                "shipping_mark",
                            ),
                        ),
                        default=F("delivery_method"),
                        output_field=CharField(),
                    ),
                    str_id=Cast("id", CharField()),
                    str_shipping_mark=Cast("shipping_mark", CharField()),
                )
                .values(
                    "container__container_number",
                    "product_name",
                    "model",
                    "address",
                    "custom_delivery_method",
                    "note",
                    "PO_ID",
                    "delivery_type",
                )
                .annotate(
                    shipping_marks=StringAgg(
                        "str_shipping_mark", delimiter=",", distinct=True
                    ),
                    ids=StringAgg("str_id", delimiter=",", distinct=True),
                    pcs=Sum("pcs", output_field=IntegerField()),
                    cbm=Sum("cbm", output_field=FloatField()),
                    n_pallet=Sum("pallets", output_field=IntegerField()),
                    weight_lbs=Sum("total_weight_lbs", output_field=FloatField()),
                    plt_ids=StringAgg(
                        "str_id", delimiter=",", distinct=True, ordering="str_id"
                    ),
                )
                .order_by("-cbm")
            )
        elif status == "palletized":
            return await sync_to_async(list)(
                DropshipCargo.objects.select_related("container", "order", "order__offload_id")
                .filter(container__container_number=container_number, order__offload_id__offload_at__isnull=False)
                .annotate(
                    custom_delivery_method=Case(
                        When(
                            Q(delivery_method="pickup") | Q(model="pickup"),
                            then=Concat(
                                "delivery_method",
                                Value("-"),
                                "model",
                                Value("-"),
                                "shipping_mark",
                            ),
                        ),
                        default=F("delivery_method"),
                        output_field=CharField(),
                    ),
                    str_id=Cast("id", CharField()),
                    str_shipping_mark=Cast("shipping_mark", CharField()),
                )
                .values(
                    "container__container_number",
                    "product_name",
                    "model",
                    "address",
                    "custom_delivery_method",
                    "note",
                    "PO_ID",
                    "delivery_type",
                )
                .annotate(
                    shipping_marks=StringAgg(
                        "str_shipping_mark", delimiter=",", distinct=True
                    ),
                    ids=StringAgg("str_id", delimiter=",", distinct=True),
                    pcs=Sum("pcs", output_field=IntegerField()),
                    cbm=Sum("cbm", output_field=FloatField()),
                    n_pallet=Sum("pallets", output_field=IntegerField()),
                    weight_lbs=Sum("total_weight_lbs", output_field=FloatField()),
                    plt_ids=StringAgg(
                        "str_id", delimiter=",", distinct=True, ordering="str_id"
                    ),
                )
                .order_by("-cbm")
            )

    async def handle_trans_arrival_post(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        batch_id = request.POST.get("batch_id")
        trans = await sync_to_async(TransferLocation.objects.get)(id=batch_id)
        trans.arrival_time = timezone.now()
        await sync_to_async(trans.save)()
        warehouse = request.POST.get("warehouse")
        mutable_post = request.POST.copy()
        mutable_post["name"] = warehouse
        request.POST = mutable_post
        return await self.handle_warehouse_post_palletize(request)

    async def handle_warehouse_post_palletize(
        self, request: HttpRequest
    ) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("name", "").strip()
        template, context = await self.handle_all_get_palletize(request, warehouse)
        return template, context

    async def handle_all_get_palletize(self, request, warehouse: str = None, error_msg: str = []) -> tuple[str, dict[str, Any]]:
        @sync_to_async
        def check_perm():
            if not request.user.is_authenticated:
                return False
            return request.user.groups.filter(name="unpacking_personnel").exists()

        unpacking_personnel = await check_perm()

        if warehouse:
            warehouse = None if warehouse == "Empty" else warehouse
            order_not_palletized, order_palletized = (
                await asyncio.gather(
                    self._get_order_not_palletized(warehouse),
                    self._get_order_palletized(warehouse)
                )
            )

            order_not_palletized = (
                [o for o in order_not_palletized if isinstance(o, dict)]
                + [
                    o
                    for o in order_not_palletized
                    if not isinstance(o, dict)
                ]
            )
            context = {
                "order_not_palletized": order_not_palletized,
                "order_palletized": order_palletized,
                "warehouse_form": ZemWarehouseForm(initial={"name": warehouse}),
                "warehouse": warehouse,
                "unpacking_personnel": unpacking_personnel,
                "error_msg": error_msg,
                "page_title": "拆柜入库",
            }
        else:
            context = {
                "warehouse_form": ZemWarehouseForm(),
                "unpacking_personnel": unpacking_personnel,
                "error_msg": error_msg,
                "page_title": "拆柜入库",
            }
        return self.template_palletization_main, context


    async def _get_order_palletized(self, warehouse: str) -> Order:
        return await sync_to_async(list)(
            Order.objects.select_related(
                "customer_name",
                "container_number",
                "retrieval_id",
                "offload_id",
                "warehouse",
            )
            .filter(
                models.Q(
                    warehouse__name=warehouse,
                    order_type="一件代发",
                    offload_id__offload_required=True,
                    offload_id__offload_at__isnull=False,
                    cancel_notification=False,
                    created_at__gte=timezone.now() - timedelta(days=120),  #最近3个月的柜子
                )
            )
            .order_by("offload_id__offload_at")
        )

    async def _get_order_not_palletized(self, warehouse: str) -> Order:
        packinglist = await sync_to_async(list)(
            Order.objects.select_related(
                "customer_name",
                "container_number",
                "retrieval_id",
                "offload_id",
                "warehouse",
            )
            .filter(
                models.Q(
                    warehouse__name=warehouse,
                    order_type="一件代发",
                    offload_id__offload_required=True,
                    offload_id__offload_at__isnull=True,
                    cancel_notification=False,
                )
            )
            .order_by("retrieval_id__arrive_at")
        )
        return packinglist

    async def handle_empty_return_post(self, request: HttpRequest) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        empty_returned_at = request.POST.get("empty_returned_at")
        retrieval = await sync_to_async(Retrieval.objects.get)(
            order__container_number__container_number=container_number
        )
        tzinfo = self._parse_tzinfo(retrieval.retrieval_destination_precise)
        empty_returned_at = self._parse_ts(empty_returned_at, tzinfo)
        retrieval.empty_returned = True
        retrieval.empty_returned_at = empty_returned_at
        await sync_to_async(retrieval.save)()
        return await self.handle_all_get_contaier_pickup_status()

    async def handle_batch_empty_return_post(self, request: HttpRequest) -> tuple[Any, Any]:
        """
        批量修改还空时间
        """
        container_numbers = request.POST.get("batch_container_numbers", "").split(',')
        empty_returned_at = request.POST.get("batch_empty_returned_at")
        for container_number in container_numbers:
            retrieval = await sync_to_async(Retrieval.objects.get)(
                order__container_number__container_number=container_number.strip()
            )
            tzinfo = self._parse_tzinfo(retrieval.retrieval_destination_precise)
            parsed_empty_returned_at = self._parse_ts(empty_returned_at, tzinfo)
            retrieval.empty_returned = True
            retrieval.empty_returned_at = parsed_empty_returned_at
            await sync_to_async(retrieval.save)()
        return await self.handle_all_get_contaier_pickup_status()

    async def handle_batch_arrive_at_destination_post(
            self, request: HttpRequest
    ) -> tuple[Any, Any]:
        """
        批量修改到仓时间
        """
        container_numbers = request.POST.get("batch_container_numbers", "").split(',')
        arrive_at = request.POST.get("batch_arrive_at")
        for container_number in container_numbers:
            order = await sync_to_async(
                Order.objects.select_related("retrieval_id", "offload_id").get
            )(container_number__container_number=container_number.strip())
            retrieval = order.retrieval_id
            tzinfo = self._parse_tzinfo(retrieval.retrieval_destination_precise)
            parsed_arrive_at = self._parse_ts(arrive_at, tzinfo)
            retrieval.arrive_at = parsed_arrive_at
            retrieval.arrive_at_destination = True
            await sync_to_async(retrieval.save)()
        return await self.handle_all_get_contaier_pickup_status()

    async def handle_arrive_at_destination_post(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        arrive_at = request.POST.get("arrive_at")
        order = await sync_to_async(
            Order.objects.select_related("retrieval_id", "offload_id").get
        )(container_number__container_number=container_number)
        retrieval = order.retrieval_id
        tzinfo = self._parse_tzinfo(retrieval.retrieval_destination_precise)
        arrive_at = self._parse_ts(arrive_at, tzinfo)
        retrieval.arrive_at = arrive_at
        retrieval.arrive_at_destination = True
        await sync_to_async(retrieval.save)()
        return await self.handle_all_get_contaier_pickup_status()

    async def handle_all_get_contaier_pickup_status(self) -> tuple[Any, Any]:
        current_date = datetime.now().date()
        orders_pickup_scheduled = await sync_to_async(list)(
            Order.objects.select_related(
                "vessel_id", "container_number", "customer_name", "retrieval_id"
            )
            .filter(
                (
                    models.Q(add_to_t49=True)
                    & models.Q(retrieval_id__actual_retrieval_timestamp__isnull=False)
                    & models.Q(retrieval_id__arrive_at_destination=False)
                    & models.Q(cancel_notification=False)
                    & models.Q(order_type="一件代发")
                )
                & (
                    models.Q(created_at__gte="2024-08-19")
                    | models.Q(
                        container_number__container_number__in=ADDITIONAL_CONTAINER
                    )
                )
            )
            .order_by("retrieval_id__actual_retrieval_timestamp")
        )
        orders_at_warehouse = await sync_to_async(list)(
            Order.objects.select_related(
                "vessel_id",
                "container_number",
                "customer_name",
                "retrieval_id",
                "offload_id",
            )
            .filter(
                (
                    models.Q(container_number__orders__order_type="一件代发")
                    & models.Q(add_to_t49=True)
                    & models.Q(retrieval_id__actual_retrieval_timestamp__isnull=False)
                    & models.Q(retrieval_id__arrive_at_destination=True)
                    & models.Q(offload_id__offload_at__isnull=True)
                    & models.Q(cancel_notification=False)
                )
                & (
                    models.Q(created_at__gte="2024-08-19")
                    | models.Q(
                        container_number__container_number__in=ADDITIONAL_CONTAINER
                    )
                )
            )
            .order_by("retrieval_id__arrive_at")
        )
        orders_palletized = await sync_to_async(list)(
            Order.objects.select_related(
                "vessel_id",
                "container_number",
                "customer_name",
                "retrieval_id",
                "offload_id",
            )
            .filter(
                (
                    models.Q(add_to_t49=True)
                    & models.Q(retrieval_id__actual_retrieval_timestamp__isnull=False)
                    & models.Q(retrieval_id__arrive_at_destination=True)
                    & models.Q(offload_id__offload_at__isnull=False)
                    & models.Q(retrieval_id__empty_returned=False)
                    & models.Q(cancel_notification=False)
                    & models.Q(order_type="一件代发")
                )
                & (
                    models.Q(created_at__gte="2024-08-19")
                    | models.Q(
                        container_number__container_number__in=ADDITIONAL_CONTAINER
                    )
                )
            )
            .order_by("offload_id__offload_at")
        )
        context = {
            "orders_pickup_scheduled": orders_pickup_scheduled,
            "orders_at_warehouse": orders_at_warehouse,
            "orders_palletized": orders_palletized,
            "current_date": current_date,
            "page_title": "货柜追踪提醒",
        }
        return self.template_status_summary, context

    async def handle_batch_pickup_schedule_confirmation(
            self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_numbers = request.POST.getlist("container_numbers")
        retrieval_destination = request.POST.get("retrieval_destination")
        retrieval_carrier = request.POST.get("retrieval_carrier")
        tzinfo = self._parse_tzinfo(retrieval_destination)
        note = request.POST.get("note", "").strip()

        for cn in container_numbers:
            order = await sync_to_async(Order.objects.select_related('retrieval_id').get)(
                container_number__container_number=cn
            )
            dropship_cargo = await sync_to_async(list)(
                DropshipCargo.objects.filter(
                    models.Q(container__container_number=cn)
                )
            )
            if order.order_type == "一件代发":
                warehouse = await sync_to_async(ZemWarehouse.objects.get)(name=retrieval_destination)
                order.warehouse = warehouse
                await sync_to_async(order.save)()
                for d_c in dropship_cargo:
                    d_c.warehouse = warehouse
                    await sync_to_async(d_c.save)()
            # 更新retrieval记录
            order.retrieval_id.retrieval_destination_precise = retrieval_destination
            order.retrieval_id.retrieval_carrier = retrieval_carrier
            if retrieval_destination and retrieval_carrier:
                order.retrieval_id.retrieval_delegation_status = True
            if request.POST.get("target_retrieval_timestamp"):
                ts = request.POST.get("target_retrieval_timestamp")
                order.retrieval_id.target_retrieval_timestamp = self._parse_ts(ts, tzinfo)
            else:
                order.retrieval_id.target_retrieval_timestamp = None
            if request.POST.get("target_retrieval_timestamp_lower"):
                ts = request.POST.get("target_retrieval_timestamp_lower")
                order.retrieval_id.target_retrieval_timestamp_lower = self._parse_ts(ts, tzinfo)
            else:
                order.retrieval_id.target_retrieval_timestamp_lower = None
            order.retrieval_id.note = note
            order.retrieval_id.scheduled_at = datetime.now()

            await sync_to_async(order.retrieval_id.save)()
        return await self.handle_all_get_terminal()

    async def handle_batch_confirm_pickup_submit_post_appointment_time(
            self, request: HttpRequest
    ) -> tuple[Any, Any]:
        """待确认提柜-处理批量确认预约时间的请求"""
        container_numbers = request.POST.getlist("container_numbers")
        appointment_time_start = request.POST.get("appointment_time_start")
        appointment_time_end = request.POST.get("appointment_time_end")
        planned_release_time = request.POST.get("planned_release_time")
        for cn in container_numbers:
            # 获取retrieval记录
            retrieval = await sync_to_async(Retrieval.objects.get)(
                order__container_number__container_number=cn
            )
            # 解析时区信息
            tzinfo = self._parse_tzinfo(retrieval.retrieval_destination_precise)
            appointment_time_start_ts = self._parse_ts(appointment_time_start, tzinfo)
            appointment_time_end_ts = self._parse_ts(appointment_time_end, tzinfo)
            planned_release_time_ts = self._parse_ts(planned_release_time, tzinfo)

            # 更新retrieval记录
            retrieval.target_retrieval_timestamp_lower = appointment_time_start_ts
            retrieval.target_retrieval_timestamp = appointment_time_end_ts
            retrieval.planned_release_time = planned_release_time_ts
            await sync_to_async(retrieval.save)()
        return await self.handle_all_get_terminal()

    async def handle_batch_confirm_pickup_submit_post(
            self, request: HttpRequest
    ) -> tuple[Any, Any]:
        """处理批量确认提柜的请求"""
        container_numbers = request.POST.getlist("container_numbers")
        actual_retrieval_timestamp = request.POST.get("actual_retrieval_timestamp")

        for cn in container_numbers:
            # 获取retrieval记录
            retrieval = await sync_to_async(Retrieval.objects.get)(
                order__container_number__container_number=cn
            )

            # 解析时区信息
            tzinfo = self._parse_tzinfo(retrieval.retrieval_destination_precise)
            actual_retrieval_ts = self._parse_ts(actual_retrieval_timestamp, tzinfo)

            # 更新retrieval记录
            retrieval.actual_retrieval_timestamp = actual_retrieval_ts

            # 填了实际提柜但是没有写预计提柜的，就默认预计提柜时间为实际提柜时间
            if not retrieval.target_retrieval_timestamp:
                retrieval.target_retrieval_timestamp = actual_retrieval_ts
            if not retrieval.target_retrieval_timestamp_lower:
                retrieval.target_retrieval_timestamp_lower = actual_retrieval_ts
            await sync_to_async(retrieval.save)()
        return await self.handle_all_get_terminal()

    async def handle_confirm_pickup_post_appointment_time(self, request: HttpRequest) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        retrieval = await sync_to_async(Retrieval.objects.get)(
            order__container_number__container_number=container_number
        )
        tzinfo = self._parse_tzinfo(retrieval.retrieval_destination_precise)
        appointment_time_start = request.POST.get("appointment_time_start")
        appointment_time_end = request.POST.get("appointment_time_end")
        appointment_time_start_ts = self._parse_ts(appointment_time_start, tzinfo)
        appointment_time_end_ts = self._parse_ts(appointment_time_end, tzinfo)
        retrieval.target_retrieval_timestamp_lower = appointment_time_start_ts
        retrieval.target_retrieval_timestamp = appointment_time_end_ts

        await sync_to_async(retrieval.save)()
        return await self.handle_all_get_terminal()

    async def handle_confirm_pickup_post(self, request: HttpRequest) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        retrieval = await sync_to_async(Retrieval.objects.get)(
            order__container_number__container_number=container_number
        )
        tzinfo = self._parse_tzinfo(retrieval.retrieval_destination_precise)
        ts = request.POST.get("actual_retrieval_timestamp")
        actual_retrieval_ts = self._parse_ts(ts, tzinfo)
        retrieval.actual_retrieval_timestamp = actual_retrieval_ts
        # 填了实际提柜但是没有写预计提柜的，就默认预计提柜时间为实际提柜时间
        if not retrieval.target_retrieval_timestamp:
            retrieval.target_retrieval_timestamp = actual_retrieval_ts
        if not retrieval.target_retrieval_timestamp_lower:
            retrieval.target_retrieval_timestamp_lower = actual_retrieval_ts
        await sync_to_async(retrieval.save)()
        return await self.handle_all_get_terminal()

    async def handle_pickup_schedule_confirmation_post(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        destination = request.POST.get("retrieval_destination").upper().strip()
        order_type = request.POST.get("order_type")
        planned_release_time_str = request.POST.get("planned_release_time")
        if planned_release_time_str:
            try:
                planned_release_time = datetime.strptime(
                    planned_release_time_str,
                    "%Y-%m-%dT%H:%M"
                )
                planned_release_time = timezone.make_aware(planned_release_time)
            except ValueError:
                return "error_template.html", {"error": "无效的时间格式，应为 'YYYY-MM-DDTHH:MM'"}
        else:
            planned_release_time = None
        order = await sync_to_async(Order.objects.select_related("retrieval_id").get)(
            models.Q(container_number__container_number=container_number)
        )
        dropship_cargo = await sync_to_async(list)(
            DropshipCargo.objects.filter(
                models.Q(container__container_number=container_number)
            )
        )

        if order_type == "一件代发":
            warehouse = await sync_to_async(ZemWarehouse.objects.get)(name=destination)
            order.warehouse = warehouse
            await sync_to_async(order.save)()
            for d_c in dropship_cargo:
                d_c.warehouse = warehouse
                await sync_to_async(d_c.save)()
        retrieval = order.retrieval_id
        retrieval.retrieval_destination_precise = destination
        retrieval.retrieval_carrier = request.POST.get("retrieval_carrier").strip()
        if destination and request.POST.get("retrieval_carrier").strip():
            retrieval.retrieval_delegation_status = True
        retrieval.planned_release_time = planned_release_time
        tzinfo = self._parse_tzinfo(destination)
        if request.POST.get("target_retrieval_timestamp"):
            ts = request.POST.get("target_retrieval_timestamp")
            retrieval.target_retrieval_timestamp = self._parse_ts(ts, tzinfo)
        else:
            retrieval.target_retrieval_timestamp = None
        if request.POST.get("target_retrieval_timestamp_lower"):
            ts = request.POST.get("target_retrieval_timestamp_lower")
            retrieval.target_retrieval_timestamp_lower = self._parse_ts(ts, tzinfo)
        else:
            retrieval.target_retrieval_timestamp_lower = None
        retrieval.note = request.POST.get("note", "").strip()
        retrieval.scheduled_at = timezone.now()
        if container_number and destination and planned_release_time and request.POST.get(
                "target_retrieval_timestamp") and request.POST.get("target_retrieval_timestamp_lower") and request.POST.get("retrieval_carrier").strip():
            retrieval.actual_release_status = True
        if request.POST.get("retrieval_carrier") == "pickup":
            if request.POST.get("target_retrieval_timestamp"):
                ts = request.POST.get("target_retrieval_timestamp")
                retrieval.target_retrieval_timestamp = self._parse_ts(ts, tzinfo)
            else:
                retrieval.target_retrieval_timestamp = None
            if request.POST.get("target_retrieval_timestamp_lower"):
                ts = request.POST.get("target_retrieval_timestamp_lower")
                retrieval.target_retrieval_timestamp_lower = self._parse_ts(ts, tzinfo)
            else:
                retrieval.target_retrieval_timestamp_lower = None
        await sync_to_async(retrieval.save)()
        return await self.handle_all_get_terminal()

    async def handle_batch_schedule_container_get(
            self, request: HttpRequest
    ) -> tuple[Any, Any]:
        if request.POST.getlist("containers[]"):
            container_numbers = request.POST.getlist("containers[]")
        else:
            container_numbers_json = request.POST.get("selected_containers")
            container_numbers = json.loads(container_numbers_json)

        # 获取所有选中的订单
        selected_orders = []
        for cn in container_numbers:
            order = await sync_to_async(
                Order.objects.select_related(
                    "container_number", "customer_name", "vessel_id", "retrieval_id"
                ).get
            )(container_number__container_number=cn)
            selected_orders.append(order)

        _, context = await self.handle_all_get_terminal()
        context["selected_orders"] = selected_orders
        context["warehouse_options"] = [("", "")] + await sync_to_async(list)(
            ZemWarehouse.objects
            .order_by("name")
            .values_list("name", "name")
        )
        context["carrier_options"] = [("", "")] + await sync_to_async(list)(
            ContainerPickupCarrier.objects
            .filter(is_active=True)
            .order_by("name")
            .values_list("name", "name")
        )

        return self.template_batch_schedule_container, context

    async def handle_batch_confirm_pickup_get(
            self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_numbers_json = request.POST.get("selected_containers")
        container_numbers = json.loads(container_numbers_json)

        # 获取所有选中的订单
        selected_orders = []
        for cn in container_numbers:
            try:
                order = await sync_to_async(
                    Order.objects.select_related(
                        "container_number", "customer_name", "vessel_id", "retrieval_id"
                    ).get
                )(container_number__container_number=cn)
                selected_orders.append(order)
            except Order.DoesNotExist:
                continue

        _, context = await self.handle_all_get_terminal()
        context["selected_orders"] = selected_orders
        context["warehouse_options"] = [("", "")] + await sync_to_async(list)(
            ZemWarehouse.objects
            .order_by("name")
            .values_list("name", "name")
        )
        context["carrier_options"] = [("", "")] + await sync_to_async(list)(
            ContainerPickupCarrier.objects
            .filter(is_active=True)
            .order_by("name")
            .values_list("name", "name")
        )

        return self.template_batch_update_container_pickup_schedule, context

    async def handle_schedule_container_pickup_get(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.GET.get("container_number")
        _, context = await self.handle_all_get_terminal()
        order = await sync_to_async(
            Order.objects.select_related(
                "container_number", "customer_name", "vessel_id", "retrieval_id"
            ).get
        )(container_number__container_number=container_number)
        context["container_number"] = container_number
        context["selected_order"] = order
        context["warehouse_options"] = [("", "")] + await sync_to_async(list)(
            ZemWarehouse.objects
            .order_by("name")
            .values_list("name", "name")
        )
        context["carrier_options"] = [("", "")] + await sync_to_async(list)(
            ContainerPickupCarrier.objects
            .filter(is_active=True)
            .order_by("name")
            .values_list("name", "name")
        )
        return self.template_schedule_container_pickup, context

    async def hanlde_update_pickup_schedule_get(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        _, context = await self.handle_schedule_container_pickup_get(request)
        return self.template_update_container_pickup_schedule, context

    async def handle_all_get_terminal(self) -> tuple[Any, Any]:
        current_date = datetime.now().date()
        orders_not_scheduled = await sync_to_async(list)(
            Order.objects.select_related(
                "vessel_id", "container_number", "customer_name", "retrieval_id"
            )
            .filter(
                (
                    models.Q(add_to_t49=True)
                    & models.Q(retrieval_id__actual_retrieval_timestamp__isnull=True)
                    & models.Q(retrieval_id__retrieval_carrier__isnull=True)
                    & models.Q(
                        vessel_id__vessel_eta__lte=timezone.now() + timedelta(weeks=2)
                    )
                    & models.Q(cancel_notification=False)
                    & models.Q(order_type="一件代发")
                )
                & (
                    models.Q(created_at__gte="2024-08-19")
                    | models.Q(
                        container_number__container_number__in=ADDITIONAL_CONTAINER
                    )
                )
            )
            .order_by("vessel_id__vessel_eta")
        )
        orders_not_pickup = await sync_to_async(list)(
            Order.objects.select_related(
                "vessel_id", "container_number", "customer_name", "retrieval_id"
            )
            .filter(
                (
                    models.Q(add_to_t49=True)
                    & models.Q(retrieval_id__actual_retrieval_timestamp__isnull=True)
                    & models.Q(retrieval_id__retrieval_carrier__isnull=False)
                    & models.Q(cancel_notification=False)
                    & models.Q(order_type="一件代发")
                )
                & (
                    models.Q(created_at__gte="2024-08-19")
                    | models.Q(
                        container_number__container_number__in=ADDITIONAL_CONTAINER
                    )
                )
            )
            .order_by("vessel_id__vessel_eta")
        )
        context = {
            "orders_not_scheduled": orders_not_scheduled,
            "orders_not_pickup": orders_not_pickup,
            "current_date": current_date,
            "page_title": "港口调度",
        }
        return self.template_terminal_dispatch, context

    # 更新5月1之后所有柜子的优先级
    async def handle_update_container_unpacking_priority(
            self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_numbers = await sync_to_async(
            lambda: list(
                Order.objects.filter(
                    created_at__date__gt=date(2025, 7, 1)
                ).values_list("container_number__container_number", flat=True).distinct()
            )
        )()

        return await self.handle_order_management_container_get(request)

    async def handle_update_delivery_type(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:

        # 处理PackingList
        batch_size = 10000
        queryset = PackingList.objects.filter(
            models.Q(delivery_type__isnull=True)
            | models.Q(delivery_type=None)
            | models.Q(delivery_type="None")
        ).order_by("id")
        total = await sync_to_async(queryset.count)()

        for start in range(0, total, batch_size):
            batch = await sync_to_async(list)(
                queryset[start : start + batch_size].values("id", "model")
            )

            public_ids = [
                item["id"]
                for item in batch
                if (
                    re.fullmatch(r"^[A-Za-z]{4}\s*$", str(item["model"]).strip())
                    or re.fullmatch(r"^[A-Za-z]{3}\s*\d$", str(item["model"]).strip())
                    or re.fullmatch(
                        r"^[A-Za-z]{3}\s*\d\s*[A-Za-z]$", str(item["model"]).strip()
                    )
                    or any(
                        kw.lower() in str(item["model"]).lower()
                        for kw in {"walmart", "沃尔玛"}
                    )
                )
            ]

            # 批量更新
            if public_ids:
                await sync_to_async(
                    PackingList.objects.filter(id__in=public_ids).update
                )(delivery_type="一件代发")
        plt_size = 10000
        queryset = Pallet.objects.filter(
            models.Q(delivery_type__isnull=True)
            | models.Q(delivery_type=None)
            | models.Q(delivery_type="None")
        ).order_by("id")
        total = await sync_to_async(queryset.count)()

        for start in range(0, total, plt_size):
            batch_plt = await sync_to_async(list)(
                queryset[start : start + plt_size].values("id", "model")
            )

            public_ids = [
                item["id"]
                for item in batch_plt
                if (
                    re.fullmatch(r"^[A-Za-z]{4}\s*$", str(item["model"]).strip())
                    or re.fullmatch(r"^[A-Za-z]{3}\s*\d$", str(item["model"]).strip())
                    or re.fullmatch(
                        r"^[A-Za-z]{3}\s*\d\s*[A-Za-z]$", str(item["model"]).strip()
                    )
                    or any(
                        kw.lower() in str(item["model"]).lower()
                        for kw in {"walmart", "沃尔玛"}
                    )
                )
            ]

            # 批量更新
            if public_ids:
                await sync_to_async(Pallet.objects.filter(id__in=public_ids).update)(
                    delivery_type="一件代发"
                )
        return await self.handle_order_management_container_get(request)

    async def handle_check_order_type_destination(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        orders = await sync_to_async(list)(
            Order.objects.filter(models.Q(order_type="一件代发"))
            .select_related("container_number")  # 优化查询性能
            .values_list("container_number__container_number", flat=True)
            .distinct()  # 确保柜号唯一
        )

        matched_containers = []

        for container_number in orders:
            model = await sync_to_async(
                lambda: list(
                    PackingList.objects.filter(
                        container_number__container_number=container_number
                    )
                    .values_list("model", flat=True)
                    .distinct()
                )
            )()
            if len(model) == 1:
                matched_containers.append(container_number)
        request.abnormal_container = matched_containers
        return await self.handle_order_management_container_get(request)

    async def check_order_status(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        order_id = request.POST.get("order_id")
        await Order.objects.filter(id=order_id).aupdate(status="checked")
        return await self.handle_order_management_container_get(request)

    async def handle_check_destination(self, request: HttpRequest) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        # 准备参数
        # 找组合柜报价
        matched_regions = await self._is_combina(container_number)
        # 非组合柜区域
        request.non_combina_region = matched_regions["non_combina_dests"]
        # 组合柜区域
        request.combina_region = matched_regions["combina_dests"]
        request.is_combina = matched_regions["is_combina"]
        request.quotation_file = matched_regions["quotation_file"]
        return await self.handle_order_management_container_get(request)

    async def _is_combina(self, container_number: str) -> bool:
        container = await sync_to_async(Container.objects.get)(container_number=container_number)
        order = await sync_to_async(
            lambda: Order.objects.select_related("retrieval_id", "vessel_id", "customer_name")
            .get(container_number__container_number=container_number)
        )()
        customer = order.customer_name
        customer_name = customer.zem_name
        # 从报价表找+客服录的数据
        warehouse = order.retrieval_id.retrieval_destination_area
        vessel_etd = order.vessel_id.vessel_etd

        container_type = container.container_type
        is_combina = True
        has_pallet = True
        #  基础数据统计
        plts = await sync_to_async(
            lambda: Pallet.objects.filter(
                container_number__container_number=container_number
            ).aggregate(
                unique_models=Count("model", distinct=True),
                total_weight=Sum("weight_lbs"),
                total_cbm=Sum("cbm"),
                total_pallets=Count("id"),
            )
        )()
        if plts['total_pallets'] == 0:
            has_pallet = False
            plts = await sync_to_async(
                lambda: PackingList.objects.filter(
                    container_number__container_number=container_number
                ).aggregate(
                    unique_models=Count("model", distinct=True),
                    total_weight=Sum("total_weight_lbs"),
                    total_cbm=Sum("cbm"),
                    total_pallets=Coalesce(
                        Round(
                            Cast(Sum("cbm"), output_field=FloatField()) / 1.8,
                            output_field=IntegerField()
                        ),
                        0  # 默认值，当Sum("cbm")为None时设为0
                    )
                )
            )()
        plts["total_cbm"] = round(plts["total_cbm"], 2)
        plts["total_weight"] = round(plts["total_weight"], 2)
        # 获取匹配的报价表
        matching_quotation = await sync_to_async(
            lambda: QuotationMaster.objects.filter(
                effective_date__lte=vessel_etd,
                is_user_exclusive=True,
                exclusive_user=customer_name,
                quote_type='receivable',
            ).order_by("-effective_date").first()
        )()

        if not matching_quotation:
            matching_quotation = await sync_to_async(
                lambda: QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=False,
                    quote_type='receivable',
                ).order_by("-effective_date").first()
            )()
        if matching_quotation:
            quotation_file = matching_quotation.filename
        else:
            quotation_file = '未找到对应报价表'

        # 获取费用详情
        stipulate = await sync_to_async(
            lambda: FeeDetail.objects.get(
                quotation_id=matching_quotation.id, fee_type="COMBINA_STIPULATE"
            ).details
        )()

        combina_fee = await sync_to_async(
            lambda: FeeDetail.objects.get(
                quotation_id=matching_quotation.id, fee_type=f"{warehouse}_COMBINA"
            ).details
        )()
        if isinstance(combina_fee, str):
            combina_fee = json.loads(combina_fee)

        # 看是否超出组合柜限定仓点,NJ/SAV是14个
        warehouse_specific_key = f'{warehouse}_max_mixed'
        if warehouse_specific_key in stipulate.get("global_rules", {}):
            combina_threshold = stipulate["global_rules"][warehouse_specific_key]["default"]
        else:
            combina_threshold = stipulate["global_rules"]["max_mixed"]["default"]

        warehouse_specific_key1 = f'{warehouse}_bulk_threshold'
        if warehouse_specific_key1 in stipulate.get("global_rules", {}):
            uncombina_threshold = stipulate["global_rules"][warehouse_specific_key1]["default"]
        else:
            uncombina_threshold = stipulate["global_rules"]["bulk_threshold"]["default"]

        if plts["unique_models"] > uncombina_threshold:
            container.account_order_type = "一件代发"
            container.non_combina_reason = (
                f"总仓点超过{uncombina_threshold}个"
            )
            await sync_to_async(container.save)()
            is_combina = False  # 不是组合柜

        # 按区域统计
        if has_pallet:
            plts_by_destination = await sync_to_async(
                lambda: list(Pallet.objects.filter(
                    container_number__container_number=container_number
                ).values("model").annotate(total_cbm=Sum("cbm")))
            )()
        else:
            plts_by_destination = await sync_to_async(
                lambda: list(PackingList.objects.filter(
                    container_number__container_number=container_number
                ).values("model").annotate(total_cbm=Sum("cbm")))
            )()
        total_cbm_sum = sum(item["total_cbm"] for item in plts_by_destination)
        # 区分组合柜区域和非组合柜区域
        container_type_temp = 0 if container_type == "40HQ/GP" else 1
        matched_regions = self._find_compre_matching_regions(
            plts_by_destination, combina_fee, container_type_temp, total_cbm_sum, combina_threshold
        )

        # 判断是否混区，False表示满足混区条件
        is_mix = await self.is_mixed_region(
            matched_regions["matching_regions"], warehouse, vessel_etd
        )
        if is_mix:
            container.account_order_type = "一件代发"
            container.non_combina_reason = "混区不符合标准"
            await sync_to_async(container.save)()
            is_combina = False
        # 非组合柜区域
        non_combina_region_count = matched_regions["non_combina_dests"]
        # 组合柜区域
        combina_region_count = matched_regions["combina_dests"]

        if len(non_combina_region_count) + len(combina_region_count) > uncombina_threshold:
            # 当非组合柜的区域数量超出时，不能按转运组合
            container.account_order_type = "一件代发"
            container.non_combina_reason = "总仓点的数量不符合标准"
            await sync_to_async(container.save)()
            is_combina = False
        if is_combina:
            container.account_order_type = "一件代发"
            container.non_combina_reason = ""
            await sync_to_async(container.save)()

        return {
            "combina_dests": combina_region_count,
            "non_combina_dests": non_combina_region_count,
            "is_combina": is_combina,
            "quotation_file": quotation_file,
        }

    def _find_compre_matching_regions(
            self,
            plts_by_destination: dict,
            combina_fee: dict,
            container_type,
            total_cbm_sum: float,
            combina_threshold: int,
    ) -> dict:
        matching_regions = defaultdict(float)  # 各区的cbm总和
        destination_matches = set()  # 组合柜的仓点
        non_combina_dests = set()  # 非组合柜的仓点
        dest_cbm_list = []  # 临时存储初筛组合柜内的cbm和匹配信息
        sum_des = set()
        price_display = defaultdict(set)
        for plts in plts_by_destination:
            sum_des.add(plts["model"])
            if "UPS" in plts["model"]:
                non_combina_dests.add("UPS")
                continue

            destination_str = plts["model"]
            destination_origin, destination = self._process_destination(destination_str)
            dest = destination.replace("沃尔玛", "").split("-")[-1].strip()

            cbm = plts["total_cbm"]
            if cbm == 0:
                non_combina_dests.add(dest)
                continue
            matched = False

            # 遍历所有区域和location
            for region, fee_data_list in combina_fee.items():
                for fee_data in fee_data_list:
                    if dest in fee_data["location"]:
                        price_display[region].add(dest)
                        matching_regions[region] += cbm
                        matched = True
                        # 记录下来，方便后续排序
                        dest_cbm_list.append({"dest": dest, "cbm": cbm})
                        destination_matches.add(dest)

            if not matched:
                non_combina_dests.add(dest)
        # 阈值处理：如果组合柜仓点超过限制，就只保留前 N 个 cbm 最大的
        if len(destination_matches) > combina_threshold:
            sorted_dests = sorted(dest_cbm_list, key=lambda x: x["cbm"], reverse=True)
            destination_matches = {item["dest"] for item in sorted_dests[:combina_threshold]}
            for item in sorted_dests[combina_threshold:]:
                non_combina_dests.add(item["dest"])
        combina_dests = {}
        for region, dests_in_region in price_display.items():
            # 取交集：只保留既在该区域又在 destination_matches 中的仓点
            matched_dests = dests_in_region.intersection(destination_matches)
            if matched_dests:  # 只添加非空的区域
                combina_dests[region] = list(matched_dests)
        # 返回精简后的结构
        return {
            "matching_regions": matching_regions,  # 各区的CBM总和
            "combina_dests": combina_dests,  # 组合柜仓点 set
            "non_combina_dests": non_combina_dests,  # 非组合柜仓点 set
        }

    def _process_destination(self, destination_origin):
        """处理目的地字符串"""

        def clean_all_spaces(s):
            if not s:  # 处理None/空字符串
                return ""
            # 匹配所有空格类型：
            # \xa0 非中断空格 | \u3000 中文全角空格 | \s 普通空格/制表符/换行等
            import re
            cleaned = re.sub(r'[\xa0\u3000\s]+', '', str(s))
            return cleaned

        destination_origin = str(destination_origin)

        # 匹配模式：按"改"或"送"分割，分割符放在第一组的末尾
        if "改" in destination_origin or "送" in destination_origin:
            # 找到第一个"改"或"送"的位置
            first_change_pos = min(
                (destination_origin.find(char) for char in ["改", "送"]
                 if destination_origin.find(char) != -1),
                default=-1
            )

            if first_change_pos != -1:
                # 第一部分：到第一个"改"或"送"（包含分隔符）
                first_part = destination_origin[:first_change_pos + 1]
                # 第二部分：剩下的部分
                second_part = destination_origin[first_change_pos + 1:]

                # 处理第一部分：按"-"分割取后面的部分
                if "-" in first_part:
                    first_result = first_part.split("-", 1)[1]
                else:
                    first_result = first_part

                # 处理第二部分：按"-"分割取后面的部分
                if "-" in second_part:
                    second_result = second_part.split("-", 1)[1]
                else:
                    second_result = second_part

                return clean_all_spaces(first_result), clean_all_spaces(second_result)
            else:
                raise ValueError(first_change_pos)

        # 如果不包含"改"或"送"或者没有找到
        # 只处理第二部分（假设第一部分为空）
        if "-" in destination_origin:
            second_result = destination_origin.split("-", 1)[1]
        else:
            second_result = destination_origin

        return None, clean_all_spaces(second_result)

    async def is_mixed_region(self, matched_regions, warehouse, vessel_etd) -> bool:
        regions = list(matched_regions.keys())
        # LA仓库的特殊规则：CDEF区不能混
        if warehouse == "LA":
            if vessel_etd.year > 2025:
                return False
            if vessel_etd.month > 7 or (
                vessel_etd.month == 7 and vessel_etd.day >= 15
            ):  # 715之后没有混区限制
                return False
            if len(regions) <= 1:  # 只有一个区，就没有混区的情况
                return False
            if set(regions) == {"A区", "B区"}:  # 如果只有A区和B区，也满足混区规则
                return False
            return True
        # 其他仓库无限制
        return False

    async def add_t49_order(
            self, request: HttpRequest
    ) -> tuple[Any, Any]:
        order_id = request.POST.get("order_id")
        await sync_to_async(
            Order.objects.filter(id=order_id).update,
            thread_sensitive=False  # 非线程敏感操作，提升执行效率
        )(add_to_t49=True)
        return await self.repeat_t49_all()

    async def handle_cancel_notification(self, request: HttpRequest) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        # 查询order表的contain_number
        order = await sync_to_async(Order.objects.get)(
            models.Q(container_number__container_number=container_number)
        )
        order.cancel_notification = True
        order.cancel_time = datetime.now()
        await sync_to_async(order.save)()
        mutable_get = request.GET.copy()
        mutable_get["container_number"] = container_number
        mutable_get["step"] = "cancel_notification"
        request.GET = mutable_get
        return await self.handle_order_management_container_get(request)


    async def handle_update_order_retrieval_info_post(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        order = await sync_to_async(Order.objects.select_related("retrieval_id").get)(
            container_number__container_number=container_number
        )
        dropship_cargo = await sync_to_async(list)(
            DropshipCargo.objects.filter(
                models.Q(container__container_number=container_number)
            )
        )
        retrieval_destination_precise = request.POST.get("retrieval_destination_precise")
        warehouse = await sync_to_async(ZemWarehouse.objects.get)(name=retrieval_destination_precise)
        order.warehouse = warehouse
        await sync_to_async(order.save)()
        for d_c in dropship_cargo:
            d_c.warehouse = warehouse
            await sync_to_async(d_c.save)()
        retrieval = await sync_to_async(Retrieval.objects.get)(
            models.Q(retrieval_id=order.retrieval_id)
        )
        tzinfo = self._parse_tzinfo(retrieval.retrieval_destination_precise)
        retrieval.retrieval_carrier = request.POST.get("retrieval_carrier")
        retrieval.retrieval_destination_precise = retrieval_destination_precise
        target_retrieval_timestamp = request.POST.get("target_retrieval_timestamp")
        retrieval.target_retrieval_timestamp = (
            self._parse_ts(target_retrieval_timestamp, tzinfo)
            if target_retrieval_timestamp
            else None
        )
        actual_retrieval_timestamp = request.POST.get("actual_retrieval_timestamp")
        retrieval.actual_retrieval_timestamp = (
            self._parse_ts(actual_retrieval_timestamp, tzinfo)
            if actual_retrieval_timestamp
            else None
        )

        arrive_at = request.POST.get("arrive_at")
        retrieval.arrive_at = self._parse_ts(arrive_at, tzinfo) if arrive_at else None
        empty_returned_at = request.POST.get("empty_returned_at")
        retrieval.empty_returned_at = (
            self._parse_ts(empty_returned_at, tzinfo) if empty_returned_at else None
        )
        if not empty_returned_at:
            retrieval.empty_returned = False
        retrieval.note = request.POST.get("retrieval_note").strip()
        await sync_to_async(retrieval.save)()
        mutable_get = request.GET.copy()
        mutable_get["container_number"] = container_number
        mutable_get["step"] = "container_info_supplement"
        request.GET = mutable_get
        return await self.handle_order_management_container_get(request)

    def _parse_tzinfo(self, s: str | None) -> str:
        if not s:
            return "America/New_York"
        elif "NJ" in s.upper():
            return "America/New_York"
        elif "SAV" in s.upper():
            return "America/New_York"
        elif "LA" in s.upper():
            return "America/Los_Angeles"
        else:
            return "America/New_York"

    def _parse_ts(self, ts: str, tzinfo: str) -> str:
        ts_naive = datetime.fromisoformat(ts)
        tz = pytz.timezone(tzinfo)
        ts = tz.localize(ts_naive).astimezone(timezone.utc)
        return ts.strftime("%Y-%m-%d %H:%M:%S")

    async def handle_export_forecast(self, request: HttpRequest) -> tuple[Any, Any]:
        selected_orders = json.loads(request.POST.get("selectedOrders", "[]"))
        selected_orders = list(set(selected_orders))
        orders = await sync_to_async(list)(
            Order.objects.select_related(
                "vessel_id",
                "container_number",
                "customer_name",
                "retrieval_id",
                "warehouse",
            )
            .values(
                "container_number__container_number",
                "customer_name__zem_name",
                "order_type",
                "vessel_id__vessel_eta",
                "vessel_id__vessel_etd",
                "cancel_time",
                "created_at",
                "retrieval_id__retrieval_carrier",
                "vessel_id__destination_port",
                "vessel_id__master_bill_of_lading",
                "warehouse__name",
                "container_number__container_type",
            )
            .filter(models.Q(container_number__container_number__in=selected_orders))
        )
        for order in orders:
            # 由于carrier的内容为中文，导出的文件中为乱码，所以修改编码，但是这段代码并没有解决编码问题，依旧是乱码，没有找到解决方案
            if order.get("retrieval_id__retrieval_carrier"):
                raw_data = order["retrieval_id__retrieval_carrier"]
                raw_data = raw_data.encode("utf-8")
                encoding = chardet.detect(raw_data)["encoding"]
                order["retrieval_id__retrieval_carrier"] = raw_data.decode(encoding)

        df = pd.DataFrame(orders)
        df = df.rename(
            {
                "container_number__container_number": "container",
                "customer_name__zem_name": "customer",
                "vessel_id__master_bill_of_lading": "MBL",
                "vessel_id__destination_port": "destination_port",
                "vessel_id__vessel_eta": "ETA",
                "vessel_id__vessel_etd": "ETD",
                "retrieval_id__retrieval_carrier": "carrier",
                "container_number__container_type": "container_type",
                "order_type": "order_type",
            },
            axis=1,
        )

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f"attachment; filename=forecast.csv"
        df.to_csv(path_or_buf=response, index=False, encoding="utf-8-sig")
        return response

    async def handle_order_basic_info_get(self) -> tuple[Any, Any]:
        customers = await sync_to_async(list)(Customer.objects.all())
        original_dict = {c.zem_name: c.id for c in customers}
        customers = {"": ""}
        customers.update(original_dict)
        orders = await sync_to_async(list)(
            Order.objects.select_related(
                "vessel_id",
                "container_number",
                "customer_name",
                "retrieval_id",
                "offload_id"
            ).prefetch_related(
                "dropshipcargo_set")
            .values(
                "container_number__container_number",
                "container_number__weight_lbs",
                "container_number__container_type",
                "customer_name__zem_name",
                "vessel_id",
                "vessel_id__vessel_eta",
                "vessel_id__vessel_etd",
                "vessel_id__vessel",
                "vessel_id__shipping_line",
                "vessel_id__destination_port",
                "vessel_id__master_bill_of_lading",
                "order_type",
                "created_at",
                "retrieval_id__retrieval_destination_area",
                "packing_list_updloaded",
                "cancel_notification",
                "id",
            ).filter(
                models.Q(offload_id__offload_at__isnull=True)
                & models.Q(cancel_notification=False)
                & models.Q(order_type='一件代发')
            )
        )

        unfinished_orders = [
            o for o in orders
            if
            #基础信息不完整
            not o.get("customer_name__zem_name") or
            not o.get("order_type") or
            not o.get("created_at") or
            not o.get("container_number__container_number") or
            not o.get("vessel_id__master_bill_of_lading") or
            not o.get("vessel_id__vessel_eta") or
            # 航运信息不完整的情况
            not o.get("vessel_id") or
            not o.get("vessel_id__vessel") or
            not o.get("vessel_id__shipping_line") or
            not o.get("vessel_id__destination_port") or
            not o.get("container_number__container_type") or
            not o.get("vessel_id__vessel_etd") or
            # 未上传装箱单的情况
            not o.get("packing_list_updloaded")
        ]
        context = {
            "customers": customers,
            "order_type": self.order_type,
            "area": self.area,
            "container_type": self.container_type,
            "unfinished_orders": unfinished_orders,
            "page_title": "创建订单",
        }
        return self.template_order_create_base, context

    async def handle_order_supplemental_info_get(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        _, context = await self.handle_order_basic_info_get()
        container_number = request.GET.get("container_number")
        order = await sync_to_async(
            Order.objects.select_related(
                "customer_name",
                "container_number",
                "retrieval_id",
                "vessel_id",
            ).get
        )(container_number__container_number=container_number)
        dropship_cargo = await sync_to_async(list)(
            DropshipCargo.objects.filter(
                models.Q(container__container_number=container_number)
            )
        )
        try:
            vessel = await sync_to_async(Vessel.objects.get)(
                order__container_number__container_number=container_number
            )
        except:
            vessel = []
        if order.customer_name.zem_name and order.order_type and order.created_at and order.container_number.container_number and vessel.master_bill_of_lading and vessel.vessel_etd and vessel.id and vessel.vessel and vessel.shipping_line and vessel.destination_port and order.packing_list_updloaded and order.container_number.container_type and vessel.vessel_eta:
            order.status = "completed"
            await sync_to_async(order.save)()
            return await self.handle_order_basic_info_get()
        context["selected_order"] = order
        context["dropship_cargo"] = dropship_cargo
        context["vessel"] = vessel
        context["shipping_lines"] = SHIPPING_LINE_OPTIONS
        context["delivery_options"] = DROPSHIPPING_DELIVERY_METHOD_OPTIONS
        context["delivery_types"] = [
            ("一件代发", "一件代发"),
        ]
        context["container_type"] = self.container_type
        context["packing_list_upload_form"] = UploadFileForm()
        return self.template_order_create_supplement, context

    async def handle_create_order_basic_post(
            self, request: HttpRequest
    ) -> tuple[Any, Any]:
        customer_id = request.POST.get("customer")
        customer = await sync_to_async(Customer.objects.get)(id=customer_id)
        order_type = request.POST.get("order_type")
        area = request.POST.get("area")
        container_number = request.POST.get("container_number")
        mbl = request.POST.get("mbl")
        etd_str = request.POST.get("etd")  # 原始字符串
        eta_str = request.POST.get("eta")  # 原始字符串
        is_expiry_guaranteed = (
            True if request.POST.get("is_expiry_guaranteed", None) else False
        )
        created_at_str = request.POST.get("created_at")

        # 处理日期格式：将纯日期转换为带时间的datetime，空值设为None
        def parse_date(date_str):
            if not date_str:  # 为空时返回None
                return None
            try:
                # 前端传的是日期（YYYY-MM-DD），转换为datetime（YYYY-MM-DD 00:00:00）
                return datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.get_current_timezone())
            except ValueError:
                raise RuntimeError(f"无效的日期格式: {date_str}，请使用YYYY-MM-DD")

        # 解析日期（etd可为空，eta和created_at必传）
        etd = parse_date(etd_str)
        eta = parse_date(eta_str)
        created_at = parse_date(created_at_str)

        # 处理已存在的柜号
        try:
            existing_order = await sync_to_async(Order.objects.get)(
                container_number__container_number=container_number
            )
            if existing_order:
                old_created_at = await sync_to_async(lambda: existing_order.created_at)()
                year_month = old_created_at.strftime("%Y%m")
                old_container = await sync_to_async(lambda: existing_order.container_number)()
                old_container.container_number = f"{old_container.container_number}_{year_month}"
                await sync_to_async(old_container.save)()
        except ObjectDoesNotExist:
            pass

        if await sync_to_async(list)(
                Order.objects.filter(container_number__container_number=container_number)
        ):
            raise RuntimeError(f"Container {container_number} exists!")

        # 生成UUID
        order_id = str(
            uuid.uuid3(
                uuid.NAMESPACE_DNS,
                str(uuid.uuid4()) + customer.zem_name + created_at_str,
            )
        )
        retrieval_id = str(uuid.uuid3(uuid.NAMESPACE_DNS, order_id + container_number))
        offload_id = str(uuid.uuid3(uuid.NAMESPACE_DNS, order_id + order_type))
        vessel_id = str(
            uuid.uuid3(uuid.NAMESPACE_DNS, container_number + (mbl or ""))
        )

        # 保存集装箱
        container_data = {
            "container_number": request.POST.get("container_number").upper().strip(),
            "is_expiry_guaranteed": is_expiry_guaranteed,
            "note": request.POST.get("note"),
        }
        container = Container(**container_data)
        await sync_to_async(container.save)()

        # 保存船舶信息（允许etd为空）
        vessel_data = {
            "vessel_id": vessel_id,
            "master_bill_of_lading": mbl,
            "vessel_etd": etd,  # 可为None
            "vessel_eta": eta,  # 必传，已通过前端验证
        }
        vessel = Vessel(**vessel_data)
        await sync_to_async(vessel.save)()

        # 保存其他关联表
        retrieval_data = {
            "retrieval_id": retrieval_id,
            "retrieval_destination_area": area
        }
        retrieval = Retrieval(**retrieval_data)
        await sync_to_async(retrieval.save)()

        offload_data = {
            "offload_id": offload_id,
            "offload_required": True if order_type in ("一件代发") else False,
        }
        offload = Offload(**offload_data)
        await sync_to_async(offload.save)()

        # 保存订单
        order_data = {
            "order_id": order_id,
            "customer_name": customer,
            "created_at": created_at,
            "order_type": order_type,
            "container_number": container,
            "retrieval_id": retrieval,
            "offload_id": offload,
            "vessel_id": vessel,
            "packing_list_updloaded": False,
        }
        order = Order(**order_data)
        await sync_to_async(order.save)()

        return await self.handle_order_basic_info_get()

    async def handle_update_order_basic_info_post(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        # check if container number is changed
        input_container_number = request.POST.get("container_number")
        original_container_number = request.POST.get("original_container_number")
        order = await sync_to_async(
            Order.objects.select_related(
                "customer_name",
                "container_number",
                "retrieval_id",
                "vessel_id",
                "offload_id",
            ).get
        )(container_number__container_number=original_container_number)
        container = order.container_number
        retrieval = order.retrieval_id
        offload = order.offload_id
        vessel = order.vessel_id
        if input_container_number != original_container_number:
            # check if the input container exists
            new_container = await sync_to_async(list)(
                Container.objects.filter(container_number=input_container_number)
            )
            if new_container:
                raise ValueError(f"container {input_container_number} exists!")
            else:
                container.container_number = input_container_number
        container.container_type = request.POST.get("container_type")
        vessel.master_bill_of_lading = request.POST.getlist('mbl')[0]
        weight = request.POST.get("weight")
        if weight:
            weight=float(weight)
            weight *= 2.20462
            container.weight_lbs = weight
        weight_lbs = request.POST.get("weight_lbs", "").strip()
        if weight_lbs in ("", "None"):
            container.weight_lbs = None
        else:
            try:
                container.weight_lbs = float(weight_lbs)
            except ValueError:
                container.weight_lbs = None
        container.is_special_container = (
            True if request.POST.get("is_special_container", None) else False
        )
        is_expiry_guaranteed = (
            True if request.POST.get("is_expiry_guaranteed", None) else False
        )
        container.is_expiry_guaranteed = is_expiry_guaranteed
        if not request.POST.get("is_special_container", None):
            container.note = ""
        else:
            container.note = request.POST.get("note")

        # check cunstomer
        input_customer_id = request.POST.get("customer")
        original_customer_id = request.POST.get("original_customer")
        if input_customer_id != original_customer_id:
            order.customer_name = await sync_to_async(Customer.objects.get)(
                id=input_customer_id
            )

        # check order_type
        retrieval.retrieval_destination_area = request.POST.get("area")
        await sync_to_async(offload.save)()
        await sync_to_async(retrieval.save)()
        await sync_to_async(container.save)()
        await sync_to_async(vessel.save)()
        await sync_to_async(order.save)()
        if is_expiry_guaranteed:
            #如果是保时效的，就重新判断一下优先级
            await self._update_container_unpacking_priority(input_container_number)
        mutable_get = request.GET.copy()
        mutable_get["container_number"] = container.container_number
        mutable_get["step"] = "container_info_supplement"
        request.GET = mutable_get
        source = request.POST.get("source")
        if source == "order_management":
            return await self.handle_order_management_container_get(request)
        else:
            return await self.handle_order_supplemental_info_get(request)

    #给定一个柜号，判定这个柜子的优先级
    async def _update_container_unpacking_priority(
        self, container_number:str
    ) -> None:
        is_expiry_guaranteed = await sync_to_async(
            lambda: Container.objects.filter(
                container_number=container_number,
                is_expiry_guaranteed=True
            ).exists()
        )()
        if is_expiry_guaranteed:
            priority = "P2"
        else:
            has_shipment = await sync_to_async(
                lambda: (
                    DropshipCargo.objects.filter(
                        container__container_number=container_number,
                        shipment_batch_number__isnull=False
                    ).exists()
                    or
                    Pallet.objects.filter(
                        container_number__container_number=container_number,
                        shipment_batch_number__isnull=False
                    ).exists()
                )
            )()
            priority = "P3" if has_shipment else "P4"
        await sync_to_async(
            lambda: Order.objects.filter(
                container_number__container_number=container_number
            ).update(unpacking_priority=priority)
        )()

    async def handle_order_management_container_get(
            self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.GET.get("container_number")
        customers = await sync_to_async(list)(Customer.objects.all())
        customers = {c.zem_name: c.id for c in customers}
        order = await sync_to_async(
            Order.objects.select_related(
                "customer_name",
                "container_number",
                "retrieval_id",
                "vessel_id",
                "warehouse",
                "offload_id",
                "shipment_id",
            ).get
        )(container_number__container_number=container_number)
        dropship_cargo = await sync_to_async(list)(
            DropshipCargo.objects.filter(
                models.Q(container__container_number=container_number)
            )
        )
        offload = order.offload_id
        cancel_access = await sync_to_async(
            lambda: (
                    request.user.is_authenticated
                    and request.user.groups.filter(
                name="create_order_dropshipping"
            ).exists()
            )
        )()
        context = {
            "selected_order": order,
            "dropship_cargo": dropship_cargo,
            "vessel": order.vessel_id,
            "retrieval": order.retrieval_id,
            "shipping_lines": SHIPPING_LINE_OPTIONS,
            "delivery_options": DROPSHIPPING_DELIVERY_METHOD_OPTIONS,
            "packing_list_upload_form": UploadFileForm(),
            "order_type": self.order_type,
            "container_type": self.container_type,
            "customers": customers,
            "area": self.area,
            "offload_at": offload.offload_at,
            "cancel_access": cancel_access
        }
        context["carrier_options"] = await sync_to_async(list)(
            ContainerPickupCarrier.objects
            .filter(is_active=True)
            .order_by("name")
            .values_list("name", "name")
        )
        context["warehouse_options"] = await sync_to_async(list)(
            ZemWarehouse.objects
            .order_by("name")
            .values_list("name", "name")
        )
        context["delivery_types"] = [
            ("一件代发", "一件代发"),
        ]
        non_combina_region = getattr(request, "non_combina_region", 0)
        combina_region = getattr(request, "combina_region", 0)

        abnormal_container = getattr(request, "abnormal_container", 0)
        if combina_region or non_combina_region:
            container = await sync_to_async(Container.objects.get)(container_number=container_number)
            is_combina = getattr(request, "is_combina", 0)
            context["is_combina"] = is_combina
            context["non_combina_reason"] = container.non_combina_reason
            context["check_destination"] = True
            context["non_combina_region"] = non_combina_region
            context["combina_region"] = combina_region
            context["quotation_file"] = getattr(request, "quotation_file")
        else:
            context["check_destination"] = False
            context["non_combina_region"] = non_combina_region
            context["combina_region"] = combina_region
        context["abnormal_container"] = abnormal_container
        return self.template_order_details, context

    async def handle_update_order_shipping_info_post(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        container_type = request.POST.get("container_type")
        if request.POST.get("is_vessel_created").upper().strip() == "YES":
            vessel = await sync_to_async(Vessel.objects.get)(
                models.Q(order__container_number__container_number=container_number)
            )
            vessel.destination_port = request.POST.get("pod").upper().strip()
            vessel.shipping_line = request.POST.get("shipping_line").strip()
            vessel.vessel = request.POST.get("vessel").upper().strip()
            vessel.vessel_eta = request.POST.get("eta")
            vessel.vessel_etd = request.POST.get("etd")
            await sync_to_async(vessel.save)()
            await self.update_container_by_number(container_number, container_type)
        else:
            await self.update_container_by_number(container_number, container_type)
            order = await sync_to_async(Order.objects.get)(
                models.Q(container_number__container_number=container_number)
            )
            vessel_id = str(
                uuid.uuid3(
                    uuid.NAMESPACE_DNS, container_number + request.POST.get("mbl")
                )
            )
            vessel = Vessel(
                vessel_id=vessel_id,
                destination_port=request.POST.get("pod").upper().strip(),
                shipping_line=request.POST.get("shipping_line"),
                vessel=request.POST.get("vessel").upper().strip(),
                vessel_eta=request.POST.get("eta"),
                vessel_etd=request.POST.get("etd"),
            )
            await sync_to_async(vessel.save)()
            order.vessel_id = vessel
            await sync_to_async(order.save)()
        mutable_get = request.GET.copy()
        mutable_get["container_number"] = container_number
        mutable_get["step"] = "container_info_supplement"
        request.GET = mutable_get
        source = request.POST.get("source")
        if source == "order_management":
            return await self.handle_order_management_container_get(request)
        else:
            return await self.handle_order_supplemental_info_get(request)

    @sync_to_async
    def update_container_by_number(self, container_number, container_type):
        # 所有同步代码在这里执行，与异步上下文完全隔离
        Container.objects.filter(container_number=container_number).update(
            container_type=container_type
        )

    async def handle_upload_template_post(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES["file"]
            df = pd.read_excel(file)
            df = df.rename(columns=DROPSHIPPING_PACKING_LIST_TEMP_COL_MAPPING)
            df = df.dropna(
                how="all",
                subset=[c for c in df.columns if c not in ["delivery_method", "note"]],
            )  # 除了delivery_method和note，删除其他列为空的
            df = df.replace(np.nan, "")
            df = df.reset_index(drop=True)
            if df["cbm"].isna().sum():  # 检查这几个字段是否为空，为空就报错
                raise ValueError(f"cbm number N/A error!")
            if (
                df["total_weight_lbs"].isna().sum()
                and df["total_weight_kg"].isna().sum()
            ):
                raise ValueError(f"weight number N/A error!")
            if df["pcs"].isna().sum():
                raise ValueError(f"boxes number N/A error!")
            for idx, row in df.iterrows():  # 转换单位
                if row["total_weight_kg"] and not row["total_weight_lbs"]:
                    df.loc[idx, "total_weight_lbs"] = round(
                        df.loc[idx, "total_weight_kg"] * 2.20462, 2
                    )
            df["product_name"] = df["product_name"].str.strip()
            df["model"] = df["model"].apply(lambda x: x.strip() if isinstance(x, str) else x)
            df["delivery_type"] = "一件代发"
            # model_fields获取pl模型的所有字段名
            model_fields = [field.name for field in DropshipCargo._meta.fields]
            col = [c for c in df.columns if c in model_fields]
            pl_data = df[col].to_dict("records")
            dropship_cargo = [DropshipCargo(**data) for data in pl_data]
        else:
            raise ValueError(f"invalid file format!")
        source = request.POST.get("source")
        if source == "order_management":
            container_number = request.POST.get("container_number")
            mutable_get = request.GET.copy()
            mutable_get["container_number"] = container_number
            request.GET = mutable_get
            _, context = await self.handle_order_management_container_get(request)
            context["dropship_cargo"] = dropship_cargo
            return self.template_order_details_pl, context
        else:
            _, context = await self.handle_order_supplemental_info_get(request)
            context["dropship_cargo"] = dropship_cargo
            return self.template_order_create_supplement_pl_tab, context

    async def handle_download_template_post(self) -> HttpResponse:
        file_path = (
            Path(__file__)
            .parent.parent.resolve()
            .joinpath("templates/export_file/dropshipping_packing_list_template.xlsx")
        )
        if not os.path.exists(file_path):
            raise Http404("File does not exist")
        with open(file_path, "rb") as file:
            response = HttpResponse(
                file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = (
                f'attachment; filename="zem_packing_list_template.xlsx"'
            )
            return response

    async def handle_update_order_packing_list_info_post_v1(
            self, request: HttpRequest
    ) -> tuple[Any, Any]:
        '''更新Packing List信息'''
        container_number = request.POST.get("container_number")
        order = await sync_to_async(
            Order.objects.select_related(
                "container_number", "offload_id", "vessel_id", "warehouse"
            ).get
        )(container_number__container_number=container_number)
        if order.warehouse_id:
            warehouse = await sync_to_async(lambda: ZemWarehouse.objects.get(id=order.warehouse_id))()
        else:
            warehouse = None
        offload = order.offload_id
        if (
                offload.offload_at and "pl_id" in request.POST
        ):  # 打板后走更新逻辑
            updated_pl = []
            pl_ids = request.POST.getlist("pl_id")
            pl_id_idx_mapping = {int(pl_ids[i]): i for i in range(len(pl_ids))}
            dropship_cargo = await sync_to_async(list)(
                DropshipCargo.objects.filter(
                    container__container_number=container_number
                )
            )
            for pl in dropship_cargo:
                idx = pl_id_idx_mapping[pl.id]
                pl.delivery_method = request.POST.getlist("delivery_method")[idx]
                pl.delivery_type = request.POST.getlist("delivery_type")[idx]

                # 【这段你写的是正确的，保留】
                sm_val = request.POST.getlist("shipping_mark")[idx].strip()
                model_val = request.POST.getlist("model")[idx].strip()
                pl.shipping_mark = sm_val if sm_val else model_val
                pl.product_name = request.POST.getlist("product_name")[idx].strip()
                pl.model = model_val

                pl.address = request.POST.getlist("address")[idx]
                pl.note = request.POST.getlist("note")[idx]
                long = request.POST.getlist("long")[idx]
                pl.long = Decimal(long) if long else None
                width = request.POST.getlist("width")[idx]
                pl.width = Decimal(width) if width else None
                height = request.POST.getlist("height")[idx]
                pl.height = Decimal(height) if height else None

                updated_pl.append(pl)
            await sync_to_async(bulk_update_with_history)(
                updated_pl,
                DropshipCargo,
                fields=[
                    "delivery_method",
                    "delivery_type",
                    "shipping_mark",
                    "product_name",
                    "model",
                    "address",
                    "note",
                    "long",
                    "width",
                    "height",
                ],
            )
        else:
            # 没打板：删除旧数据，批量新建
            await sync_to_async(
                DropshipCargo.objects.filter(
                    container__container_number=container_number
                ).delete
            )()
            # Generate PO_ID
            po_ids = []
            po_id_hash = {}
            seq_num = 1
            # 修复：这里不要替换shipping_mark数组，保持原生下标对齐
            for dm, sm, dest in zip_longest(
                    request.POST.getlist("delivery_method"),
                    request.POST.getlist("shipping_mark"),
                    request.POST.getlist("model"),
                    fillvalue='',
            ):
                po_id: str = ""
                po_id_seg: str = ""
                po_id_hkey: str = ""
                if dm == "pickup" or dest == "pickup":
                    po_id_hkey = f"{container_number}-{dm}-{dest}"
                    po_id_seg = (
                        f"S{sm[-4:]}"
                        if sm.strip()
                        else f"S{''.join(random.choices(string.ascii_letters.upper() + string.digits, k=6))}"
                    )
                else:
                    po_id_hkey = f"{container_number}-{dm}-{dest}"
                    po_id_seg = f"{DELIVERY_METHOD_CODE.get(dm, 'UN')}{dest.replace(' ', '').split('-')[-1]}"
                if po_id_hkey in po_id_hash:
                    po_id = po_id_hash.get(po_id_hkey)
                else:
                    container_tag = f"{container_number[:2].upper()}{container_number[-4:]}"

                    random.seed(container_number[-4:])
                    po_id = (
                        f"{container_tag}"
                        f"{''.join(random.choices(string.ascii_uppercase + string.digits, k=6))}"
                        f"{po_id_seg}"
                        f"{seq_num}"
                    )
                    po_id = re.sub(r"[\u4e00-\u9fff]", "", po_id)
                    po_id_hash[po_id_hkey] = po_id
                    seq_num += 1
                po_ids.append(po_id)
            del po_id_hash, po_id, po_id_seg, po_id_hkey

            total_weight_lbs_list = []
            for lbs_str in request.POST.getlist("total_weight_lbs"):
                if lbs_str.strip():
                    try:
                        total_weight_lbs_list.append(float(lbs_str))
                    except ValueError:
                        raise RuntimeError(f"无效的重量值: {lbs_str}，请输入数字")

            total_weight_lbs_sum = sum(total_weight_lbs_list)
            container = await sync_to_async(Container.objects.get)(
                container_number=container_number
            )
            container.weight_lbs = total_weight_lbs_sum
            await sync_to_async(container.save)()

            # 修复：zip_longest参数恢复原生，不替换数组
            pl_data = zip_longest(
                request.POST.getlist("delivery_method"),
                request.POST.getlist("shipping_mark"),
                request.POST.getlist("product_name"),
                request.POST.getlist("model"),
                request.POST.getlist("address"),
                request.POST.getlist("pcs"),
                request.POST.getlist("total_weight_kg"),
                request.POST.getlist("total_weight_lbs"),
                request.POST.getlist("cbm"),
                request.POST.getlist("note"),
                request.POST.getlist("long"),
                request.POST.getlist("width"),
                request.POST.getlist("height"),
                po_ids,
                request.POST.getlist("delivery_type"),
                fillvalue=""
            )

            def parse_decimal(value):
                if not value or str(value).strip() == "":
                    return None
                try:
                    return Decimal(str(value).strip())
                except InvalidOperation:
                    raise ValueError(f"无效的数值格式: {value}")

            pl_to_create = [
                DropshipCargo(
                    container=container,
                    order=order,
                    warehouse=warehouse,
                    delivery_method=d[0],
                    # 空唛头自动取model(d[3])
                    shipping_mark=d[1].strip() if d[1].strip() else d[3].strip(),
                    product_name=d[2],
                    model=d[3].strip(),
                    address=d[4],
                    pcs=int(float(d[5])),
                    total_weight_kg=d[6],
                    total_weight_lbs=d[7],
                    cbm=d[8],
                    note=d[9],
                    long=parse_decimal(d[10]),
                    width=parse_decimal(d[11]),
                    height=parse_decimal(d[12]),
                    PO_ID=d[13],
                    delivery_type=d[14],
                )
                for d in pl_data
            ]

            await sync_to_async(bulk_create_with_history)(pl_to_create, DropshipCargo)
            order.packing_list_updloaded = True
            await sync_to_async(order.save)()

        # 统一更新container配送类型
        container = await sync_to_async(Container.objects.get, thread_sensitive=True)(
            container_number=container_number
        )
        container.delivery_type = "一件代发"
        await sync_to_async(container.save, thread_sensitive=True)()

        source = request.POST.get("source")
        if source == "order_management":
            mutable_get = request.GET.copy()
            mutable_get["container_number"] = container_number
            mutable_get["step"] = "container_info_supplement"
            request.GET = mutable_get
            return await self.handle_order_management_container_get(request)
        else:
            return await self.handle_order_basic_info_get()

    async def handle_update_order_packing_list_info_post(
            self, request: HttpRequest
    ) -> tuple[Any, Any]:

        container_number = request.POST.get("container_number")

        selected_ids = request.POST.get("selected_pl_ids", "")

        selected_ids = [int(i) for i in selected_ids.split(",") if i]

        delete_ids = request.POST.get("delete_pl_ids")


        # 删除
        if delete_ids:
            ids = [int(i) for i in delete_ids.split(",") if i]
            await sync_to_async(
                DropshipCargo.objects.filter(id__in=ids).delete
            )()
            source = request.POST.get("source")

            if source == "order_management":
                mutable_get = request.GET.copy()
                mutable_get["container_number"] = container_number
                mutable_get["step"] = "container_info_supplement"
                request.GET = mutable_get
                return await self.handle_order_management_container_get(request)

            return await self.handle_order_basic_info_get()

        # 查询已有DropshipCargo
        cargo_dict = {
            cargo.id: cargo
            for cargo in await sync_to_async(list)(
                DropshipCargo.objects.filter(id__in=selected_ids)
            )
        }

        order = await sync_to_async(
            Order.objects.select_related(
                "container_number",
                "warehouse",
            ).get
        )(container_number__container_number=container_number)
        container = await sync_to_async(Container.objects.get)(
            container_number=container_number
        )
        warehouse = order.warehouse if order.warehouse_id else None

        update_list = []
        po_id_hash = {}
        seq_num = 1
        for pl_id in selected_ids:
            d_m = request.POST.get(f"delivery_method_{pl_id}")
            s_m = request.POST.get(f"shipping_mark_{pl_id}")
            p_name = request.POST.get(f"product_name_{pl_id}")
            m = request.POST.get(f"model_{pl_id}")
            addr = request.POST.get(f"address_{pl_id}")
            pcs = request.POST.get(f"pcs_{pl_id}")
            kg = request.POST.get(f"total_weight_kg_{pl_id}")
            lbs = request.POST.get(f"total_weight_lbs_{pl_id}")
            cbm = request.POST.get(f"cbm_{pl_id}")
            note = request.POST.get(f"note_{pl_id}")
            long_value = request.POST.get(f"long_{pl_id}")
            width_value = request.POST.get(f"width_{pl_id}")
            height_value = request.POST.get(f"height_{pl_id}")

            # Generate PO_ID
            if d_m == "pickup" or m == "pickup":
                po_id_hkey = f"{container_number}-{d_m}-{m}"
                po_id_seg = (
                    f"S{s_m[-4:]}"
                    if s_m.strip()
                    else f"S{''.join(random.choices(string.ascii_letters.upper() + string.digits, k=6))}"
                )
            else:
                po_id_hkey = f"{container_number}-{d_m}-{m}"
                po_id_seg = f"{DELIVERY_METHOD_CODE.get(d_m, 'UN')}{m.replace(' ', '').split('-')[-1]}"
            if po_id_hkey in po_id_hash:
                po_id = po_id_hash.get(po_id_hkey)
            else:
                container_tag = f"{container_number[:2].upper()}{container_number[-4:]}"

                random.seed(container_number[-4:])
                po_id = (
                    f"{container_tag}"
                    f"{''.join(random.choices(string.ascii_uppercase + string.digits, k=6))}"
                    f"{po_id_seg}"
                    f"{seq_num}"
                )
                po_id = re.sub(r"[\u4e00-\u9fff]", "", po_id)
                po_id_hash[po_id_hkey] = po_id
                seq_num += 1

            if pl_id:
                # 更新
                cargo = cargo_dict[int(pl_id)]

            else:
                # 新增
                cargo = DropshipCargo(
                    container=container,
                    order=order,
                    warehouse=warehouse,
                )

            cargo.delivery_method = d_m
            cargo.delivery_type = "一件代发"
            cargo.shipping_mark = (
                s_m.strip()
                if s_m and s_m.strip()
                else m.strip()
            )
            cargo.product_name = p_name
            cargo.model = m
            cargo.address = addr
            cargo.pcs = int(float(pcs))
            cargo.total_weight_kg = kg
            cargo.total_weight_lbs = lbs
            cargo.cbm = cbm
            cargo.note = note

            cargo.long = Decimal(long_value) if long_value else None
            cargo.width = Decimal(width_value) if width_value else None
            cargo.height = Decimal(height_value) if height_value else None

            cargo.PO_ID = po_id

            if pl_id:
                update_list.append(cargo)

        # 6. 批量更新
        if update_list:
            await sync_to_async(
                bulk_update_with_history
            )(
                update_list,
                DropshipCargo,
                fields=[
                    "delivery_method",
                    "delivery_type",
                    "shipping_mark",
                    "product_name",
                    "model",
                    "address",
                    "pcs",
                    "total_weight_kg",
                    "total_weight_lbs",
                    "cbm",
                    "note",
                    "long",
                    "width",
                    "height",
                    "PO_ID",
                ],
            )

        # 新增 Packing List
        if request.POST.get("is_add") == "1":

            cargo = DropshipCargo(
                container=container,
                order=order,
                warehouse=warehouse,
            )

            cargo.delivery_method = request.POST.get("delivery_method")
            cargo.delivery_type = "一件代发"

            shipping_mark = request.POST.get("shipping_mark")
            model = request.POST.get("model")

            cargo.shipping_mark = (
                shipping_mark.strip()
                if shipping_mark and shipping_mark.strip()
                else model.strip()
            )

            cargo.product_name = request.POST.get("product_name")
            cargo.model = model
            cargo.address = request.POST.get("address")
            cargo.pcs = int(float(request.POST.get("pcs") or 0))
            cargo.total_weight_kg = request.POST.get("total_weight_kg")
            cargo.total_weight_lbs = request.POST.get("total_weight_lbs")
            cargo.cbm = request.POST.get("cbm")
            cargo.note = request.POST.get("note")

            cargo.long = Decimal(request.POST.get("long")) if request.POST.get("long") else None
            cargo.width = Decimal(request.POST.get("width")) if request.POST.get("width") else None
            cargo.height = Decimal(request.POST.get("height")) if request.POST.get("height") else None

            # 生成 PO_ID
            d_m = cargo.delivery_method
            s_m = cargo.shipping_mark
            m = cargo.model

            if d_m == "pickup" or m == "pickup":
                po_id_seg = (
                    f"S{s_m[-4:]}"
                    if s_m
                    else f"S{''.join(random.choices(string.ascii_letters.upper() + string.digits, k=6))}"
                )
            else:
                po_id_seg = f"{DELIVERY_METHOD_CODE.get(d_m, 'UN')}{m.replace(' ', '').split('-')[-1]}"

            container_tag = f"{container_number[:2].upper()}{container_number[-4:]}"
            random.seed(container_number[-4:])
            cargo.PO_ID = re.sub(
                r"[\u4e00-\u9fff]",
                "",
                f"{container_tag}"
                f"{''.join(random.choices(string.ascii_uppercase + string.digits, k=6))}"
                f"{po_id_seg}"
                "1",
            )

            await sync_to_async(cargo.save)()

        total_weight_lbs_sum = await sync_to_async(
            lambda: DropshipCargo.objects.filter(
                container__container_number=container_number,
                delivery_type="一件代发",
            ).aggregate(
                total=Sum("total_weight_lbs")
            )["total"] or 0
        )()
        container.weight_lbs = total_weight_lbs_sum
        container.delivery_type = "一件代发"
        await sync_to_async(container.save)()

        # 8. 返回页面
        source = request.POST.get("source")

        if source == "order_management":
            mutable_get = request.GET.copy()
            mutable_get["container_number"] = container_number
            mutable_get["step"] = "container_info_supplement"
            request.GET = mutable_get
            return await self.handle_order_management_container_get(request)

        return await self.handle_order_basic_info_get()

    async def handle_order_management_list_get(
            self,
            start_date_eta: str = None,
            end_date_eta: str = None,
            start_date_etd: str = None,
            end_date_etd: str = None,
    ) -> tuple[Any, Any]:
        # 默认时间：ETA 最近30天
        if not start_date_eta and not start_date_etd:
            start_date_eta = (datetime.now().date() + timedelta(days=-30)).strftime("%Y-%m-%d")
        if not end_date_eta and not end_date_etd:
            end_date_eta = (datetime.now().date() + timedelta(days=30)).strftime("%Y-%m-%d")

        criteria = None
        if start_date_eta and end_date_eta:
            criteria = (
                    models.Q(vessel_id__vessel_eta__gte=start_date_eta, vessel_id__vessel_eta__lte=end_date_eta, order_type="一件代发") |
                    models.Q(created_at__gte=start_date_eta, created_at__lte=end_date_eta, order_type="一件代发")
            )

        if start_date_etd and end_date_etd:
            etd_q = models.Q(vessel_id__vessel_etd__gte=start_date_etd, vessel_id__vessel_etd__lte=end_date_etd, order_type="一件代发")
            criteria = criteria & etd_q if criteria else etd_q

        # 查询订单
        orders = await sync_to_async(list)(
            Order.objects.select_related(
                "vessel_id",
                "container_number",
                "customer_name",
                "retrieval_id",
                "offload_id",
                "warehouse",
            )
            .filter(criteria if criteria else models.Q())
        )

        # 遍历设置状态
        for order in orders:
            status = "未知"

            # 1. 已取消
            if order.cancel_notification:
                status = "已取消"

            # 2. T49 待追踪
            elif not order.add_to_t49:
                status = "T49待追踪"

            else:
                ret = order.retrieval_id
                offload = order.offload_id

                # 安全判断：防止 None 报错
                if not ret:
                    status = "未设置提柜信息"
                    order.status = status
                    continue

                # 3. 未设置实际提柜时间
                if not ret.actual_retrieval_timestamp:
                    if not ret.retrieval_carrier:
                        status = "待预约提柜"
                    else:
                        status = "待确认提柜"

                # 4. 已提柜，未到仓
                elif not ret.arrive_at_destination:
                    status = "待确认到仓"

                # 5. 已到仓 → 拆柜逻辑
                else:
                    if not offload.offload_at:
                        status = "待确认拆柜"
                    else:
                        status = "已拆柜"

            order.status = status

        context = {
            "orders": orders,
            "start_date_eta": start_date_eta,
            "end_date_eta": end_date_eta,
            "start_date_etd": start_date_etd,
            "end_date_etd": end_date_etd,
            "page_title": "订单列表",
        }

        return self.template_order_list, context

    async def handle_export_details_by_destination(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        selected_orders = json.loads(request.POST.get("selectedOrders", "[]"))
        selected_orders = list(set(selected_orders))
        orders = await sync_to_async(list)(
            Order.objects.select_related(
                "vessel_id",
                "container_number",
                "retrieval_id",
                "warehouse",
                "customer_name",
            )
            .values(
                "container_number__container_number",
                "vessel_id__vessel_etd",
                "retrieval_id__retrieval_destination_area",
                "customer_name__zem_name",
            )
            .filter(models.Q(container_number__container_number__in=selected_orders))
        )
        results = []
        for order in orders:
            container_number = order.get("container_number__container_number")
            vessel_etd = order.get("vessel_id__vessel_etd")
            warehouse = order.get("retrieval_id__retrieval_destination_area")

            # 找报价表
            customer_name = order.get("customer_name__zem_name")
            matching_quotation = await (
                QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=True,
                    exclusive_user=customer_name,
                    quote_type='receivable',
                )
                .order_by("-effective_date")
                .afirst()
            )

            if not matching_quotation:
                matching_quotation = await (
                    QuotationMaster.objects.filter(
                        effective_date__lte=vessel_etd,
                        is_user_exclusive=False,  # 非用户专属的通用报价单
                        quote_type='receivable',
                    )
                    .order_by("-effective_date")
                    .afirst()
                )
            if not matching_quotation:
                raise ValueError("找不到报价表")

            # 找组合柜报价
            combina_fee_detail = await FeeDetail.objects.aget(
                quotation_id=matching_quotation.id, fee_type=f"{warehouse}_COMBINA"
            )
            combina_fee = combina_fee_detail.details
            plts_by_destination = await sync_to_async(
                lambda: list(
                    PackingList.objects.filter(
                        container_number__container_number=container_number
                    ).values("model")
                )
            )()
            matched_regions = self.find_matching_regions(
                plts_by_destination, combina_fee
            )

            combina_keys = "+".join(matched_regions["combina_dests"].keys())
            non_combina_vals = ",".join(matched_regions["non_combina_dests"])
            ups_total_pcs = await sync_to_async(
                lambda: PackingList.objects.filter(
                    container_number__container_number=container_number,
                    model__icontains="UPS",
                ).aggregate(total=Sum("pcs"))["total"]
            )()
            fxdex_total_pcs = await sync_to_async(
                lambda: PackingList.objects.filter(
                    container_number__container_number=container_number,
                    model__icontains="FEDEX",
                ).aggregate(total=Sum("pcs"))["total"]
            )()
            results.append(
                {
                    "container_number": container_number,
                    "combina_dests": combina_keys,
                    "non_combina_dests": non_combina_vals,
                    "UPS": ups_total_pcs,
                    "FEDEX": fxdex_total_pcs,
                }
            )

        df = pd.DataFrame(
            results,
            columns=[
                "container_number",
                "combina_dests",
                "non_combina_dests",
                "UPS",
                "FEDEX",
            ],
        )
        df = df.rename(
            {
                "container_number": "柜号",
                "combina_dests": "组合柜的区",
                "non_combina_dests": "非组合柜仓点",
            },
            axis=1,
        )
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f"attachment; filename=destination_details.csv"
        )
        df.to_csv(path_or_buf=response, index=False, encoding="utf-8-sig")
        return response

    def find_matching_regions(
        self, plts_by_destination: list, combina_fee: dict
    ) -> dict:
        non_combina_dests = set()
        price_display = defaultdict(set)

        for plts in plts_by_destination:
            if "UPS" in plts["model"]:
                # 如果包含UPS，不需要显示细节，就显示UPS就可以了，张楠提
                non_combina_dests.add("UPS")
                continue
            dest = plts["model"]
            dest = dest.replace("沃尔玛", "").split("-")[-1].strip()
            matched = False
            # 遍历所有区域和location
            for region, fee_data_list in combina_fee.items():
                for fee_data in fee_data_list:
                    if dest in fee_data["location"]:
                        price_display[region].add(dest)
                        matched = True

            # 记录匹配结果
            if not matched:
                non_combina_dests.add(dest)  # 未匹配的仓点
        combina_dests = {k: list(v) for k, v in price_display.items()}
        return  {
            "combina_dests": combina_dests,
            "non_combina_dests": non_combina_dests,
        }

    async def handle_delete_order_post(self, request: HttpRequest) -> tuple[Any, Any]:
        selected_orders = json.loads(request.POST.get("selectedOrders", "[]"))
        selected_orders = list(set(selected_orders))
        # 在这里进行订单删除操作，例如：
        for order_number in selected_orders:
            await sync_to_async(
                Order.objects.filter(
                    container_number__container_number=order_number
                ).delete
            )()
        start_date_eta = request.POST.get("start_date_eta")
        end_date_eta = request.POST.get("end_date_eta")
        start_date_etd = (
            request.POST.get("start_date_etd")
            if request.POST.get("start_date_etd")
            and request.POST.get("start_date_etd") != "None"
            else ""
        )
        end_date_etd = (
            request.POST.get("end_date_etd")
            if request.POST.get("end_date_etd")
            and request.POST.get("end_date_etd") != "None"
            else ""
        )
        return await self.handle_order_management_list_get(
            start_date_eta, end_date_eta, start_date_etd, end_date_etd
        )

    async def repeat_t49_all(self) -> tuple[Any, Any]:
        """
        T49待追踪订单（add_to_t49=False且满足过滤条件）
        """
        # 公共过滤条件（抽离复用，避免重复代码）
        common_filter = (
                (models.Q(created_at__gte=timezone.make_aware(datetime(2024, 8, 19)))
                 | models.Q(container_number__container_number__in=ADDITIONAL_CONTAINER))
                & models.Q(offload_id__offload_at__isnull=True)
                & models.Q(cancel_notification=False)
        )
        t49_pending_orders = await sync_to_async(list)(
            Order.objects.select_related("container_number")
            .values(
                "id",
                "created_at",
                "container_number__container_number",
                "add_to_t49"
            )
            .filter(
                common_filter,
                add_to_t49=False,
                order_type="一件代发"
            )
            .order_by("-created_at")
        )

        context = {
            "t49_pending_orders": t49_pending_orders,
            "t49_pending_count": len(t49_pending_orders),
            "page_title": "T49待追踪",
        }

        return self.template_repeat_t49_all, context

