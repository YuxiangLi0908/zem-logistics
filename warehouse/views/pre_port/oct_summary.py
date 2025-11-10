import json
import zipfile
from datetime import datetime
from io import BytesIO
from typing import Any

from asgiref.sync import sync_to_async
from dateutil.relativedelta import relativedelta
from django.db.models import F, Sum, Case, When, IntegerField, Max, CharField, Value, Q, DateTimeField
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views import View
from openpyxl import Workbook
from sqlalchemy import values

from warehouse.models.order import Order
from warehouse.models.pallet import Pallet
from warehouse.models.retrieval import Retrieval
from warehouse.utils.constants import PORT_TO_WAREHOUSE_AREA
from warehouse.views.export_file import export_do, export_do_branch


class OctSummaryView(View):
    main_template="pre_port/oct_summary/01_pre_port_oct_summary.html"
    async def get(self, request: HttpRequest) -> HttpResponse:
        if not await self._user_authenticate(request):
            return redirect("login")
        step = request.GET.get("step", None)
        if step == "all":
            template, context = await self.check_disnation()
            return render(request, template, context)

    async def post(self, request: HttpRequest) -> HttpResponse:
        if not await self._user_authenticate(request):
            return redirect("login")
        step = request.POST.get("step", None)
        if step == "warehouse":
            template, context = await self.oct_handle_warehouse_post(request)
            return render(request, template, context)
        elif step == "save_do_sent":
            template, context = await self.save_do_sent(request)
            return render(request, template, context)
        elif step == "batch_export":
            # 直接返回 batch_export 生成的文件响应（无需 render）
            return await self.batch_export(request)
        elif step == "batch_export_v1":
            # 直接返回 batch_export 生成的文件响应（无需 render）
            return await self.batch_export_v1(request)
        elif step == "save_retrieval_carrier_planned":
            template, context = await self.save_retrieval_carrier_planned(request)
            return render(request, template, context)
        elif step == "save_planned_release_time":
            template, context = await self.save_planned_release_time(request)
            return render(request, template, context)
        elif step == "batch_save_retrieval_carrier_planned":
            template, context = await self.batch_save_retrieval_carrier_planned(request)
            return render(request, template, context)

    async def check_disnation(self) -> tuple[Any, Any]:
        context = {"warehouse_options": self.warehouse_options}
        return self.main_template, context

    async def _user_authenticate(self, request: HttpRequest) -> bool:
        if await sync_to_async(lambda: request.user.is_authenticated)():
            return True
        return False

    async def oct_handle_warehouse_post(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        warehouse = request.POST.get("warehouse")
        eta_start = request.POST.get("eta_start")
        eta_end = request.POST.get("eta_end")
        order_start = request.POST.get("order_start")
        order_end = request.POST.get("order_end")
        time_range_type = request.POST.get("timeRangeType", "")
        pallets_list = []
        if eta_start and eta_end:
            if warehouse == "NJ,SAV,LA,MO,TX":
                warehouse_list = warehouse.split(",")
                for w in warehouse_list:
                    pallets = await self._get_pallet(w, eta_start, eta_end)
                    pallets_list.extend(pallets)
            else:
                pallets_list = await self._get_pallet(warehouse, eta_start, eta_end)
        elif order_start and order_end:
            if warehouse == "NJ,SAV,LA,MO,TX":
                warehouse_list = warehouse.split(",")
                for w in warehouse_list:
                    pallets = await self._get_pallet_by_order_at(w, order_start, order_end)
                    pallets_list.extend(pallets)
            else:
                pallets_list = await self._get_pallet_by_order_at(warehouse, order_start, order_end)

        # 3. 构建上下文（包含更新提示）
        context = {
            "warehouse_options": self.warehouse_options,
            "warehouse": warehouse,
            "datas": pallets_list,
            "eta_start": eta_start,
            "eta_end": eta_end,
            "order_start": order_start,
            "order_end": order_end,
            "time_range_type": time_range_type,
        }

        # 4. 返回模板路径和上下文（与原有调用逻辑兼容）
        return self.main_template, context

    async def batch_export(self, request: HttpRequest):
        try:
            container_numbers = request.POST.getlist("container_numbers[]")
            def sync_process_data():
                # 1. 同步查询订单数据
                orders = list(
                    Order.objects.select_related("container_number", "retrieval_id").filter(
                        container_number__container_number__in=container_numbers
                    ).all()
                )

                if not orders:
                    return None

                port_stats = {}
                customer_stats = {}
                for item in orders:
                    port = item.retrieval_id.retrieval_destination_area or "未知港口"
                    customer = getattr(item.customer_name, "zem_name", "未知客户")

                    # 统计港口数据
                    if port not in port_stats:
                        port_stats[port] = [0, 0]
                    port_stats[port][0] += 1
                    if item.status == "unfinished":
                        port_stats[port][1] += 1

                    # 统计客户数据
                    key = (port, customer)
                    customer_stats[key] = customer_stats.get(key, 0) + 1

                # 港口统计：[(港口1, 总数量1, 未建单数量1), (港口2, 总数量2, 未建单数量2), ...]
                port_list = [(port, stats[0], stats[1]) for port, stats in port_stats.items()]
                # 客户统计：[(港口1, 客户1, 数量1), (港口1, 客户2, 数量2), ...]
                customer_list = [(k[0], k[1], v) for k, v in customer_stats.items()]

                return (port_list, customer_list)

            result = await sync_to_async(sync_process_data)()
            if not result:
                return HttpResponse("未查询到数据", status=404)
            port_list, customer_list = result
            wb = Workbook()

            # 1. 港口统计表（按港口分组显示）
            ws1 = wb.active
            ws1.title = "港口统计"
            ws1.append(["港口", "预报数量", "未建单数量"])
            for port_data in port_list:
                ws1.append(port_data)

            # 2. 客户统计表（按港口+客户分组显示）
            ws2 = wb.create_sheet(title="客户统计")
            ws2.append(["港口", "客户", "数量"])
            for customer_data in customer_list:
                ws2.append(customer_data)

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="港口客户统计.xlsx"'
            return response

        except Exception as e:
            return HttpResponse(f"导出失败：{str(e)}", status=500)

    async def batch_export_v1(self, request: HttpRequest):
        try:
            container_numbers = request.POST.getlist("container_numbers[]")
            if not container_numbers:
                return HttpResponse("未选择任何数据", status=400)

            def sync_process_data():
                orders = list(
                    Order.objects.select_related("container_number", "retrieval_id", "vessel_id").filter(
                        container_number__container_number__in=container_numbers
                    ).all()
                )
                if not orders:
                    return None

                export_data = []
                for item in orders:
                    container = item.container_number.container_number or ""
                    mbl = getattr(item.vessel_id, "master_bill_of_lading", "")
                    container_type = item.container_number.container_type or ""
                    eta = item.vessel_id.vessel_eta.strftime("%Y-%m-%d") if (
                            item.vessel_id and item.vessel_id.vessel_eta) else ""
                    terminal = ""
                    remark = ""
                    export_data.append([container, mbl, terminal, container_type, eta, remark])
                return export_data

            export_data = await sync_to_async(sync_process_data)()
            if not export_data:
                return HttpResponse("未查询到数据", status=404)

            excel_buffer = BytesIO()
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "选中MBL数据统计"
            ws1.append(["柜号", "MBL", "码头", "柜型", "ETA", "备注"])
            for row in export_data:
                ws1.append(row)
            wb.save(excel_buffer)
            excel_buffer.seek(0)  # 重置指针
            #DO的PDF
            async def get_do_pdf(container_number):
                return await sync_to_async(export_do_branch)(container_number)

            # 创建ZIP压缩包（核心：把Excel和PDF都放进ZIP）
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                # 添加Excel文件到ZIP（文件名保持.xlsx后缀）
                zip_file.writestr("批量导出选中MBL数据.xlsx", excel_buffer.getvalue())

                for container_number in container_numbers:
                    pdf_response = await get_do_pdf(container_number)
                    zip_file.writestr(f"DO_{container_number}.pdf", pdf_response.content)

            zip_buffer.seek(0)
            response = HttpResponse(
                zip_buffer.getvalue(),
                content_type="application/zip"
            )
            response["Content-Disposition"] = 'attachment; filename="批量导出数据（含Excel和DO）.zip"'

            excel_buffer.close()
            zip_buffer.close()

            return response

        except Exception as e:
            return HttpResponse(f"导出失败：{str(e)}", status=500)

    async def save_do_sent(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        order_id = request.POST.get("order_id")
        do_sent = request.POST.get("do_sent")
        await sync_to_async(lambda:Order.objects.filter(id=order_id).update(do_sent=do_sent))()
        template, context = await self.oct_handle_warehouse_post(request)
        return template, context

    async def save_retrieval_carrier_planned(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        retrieval_carrier_planned = request.POST.get("retrieval_carrier_planned")
        retrieval_id = request.POST.get("retrieval_id__retrieval_id")
        await sync_to_async(lambda:Retrieval.objects.filter(retrieval_id=retrieval_id).update(retrieval_carrier_planned=retrieval_carrier_planned))()
        template, context = await self.oct_handle_warehouse_post(request)
        return template, context

    async def save_planned_release_time(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        planned_release_time_str = request.POST.get("planned_release_time")
        retrieval_id = request.POST.get("retrieval_id__retrieval_id")
        if planned_release_time_str:
            try:
                planned_release_time = datetime.strptime(
                    planned_release_time_str,
                    "%Y-%m-%dT%H:%M"
                )
                planned_release_time = timezone.make_aware(planned_release_time)
            except ValueError:
                return "error_template.html", {"error": "无效的时间格式，应为 'YYYY-MM-DDTHH:MM'"}
        else:
            planned_release_time = None
        await sync_to_async(lambda:Retrieval.objects.filter(retrieval_id=retrieval_id).update(
                                planned_release_time=planned_release_time))()

        template, context = await self.oct_handle_warehouse_post(request)
        return template, context

    async def batch_save_retrieval_carrier_planned(self, request: HttpRequest) -> tuple[str, dict[str, Any]]:
        batch_carrier = request.POST.get("batch_carrier")
        if not batch_carrier:
            return "错误页面", {"error": "请输入计划提柜供应商"}
        retrieval_ids = request.POST.getlist("retrieval_ids[]")
        if not retrieval_ids:
            return "错误页面", {"error": "未选择任何记录"}

        await sync_to_async(
            lambda: Retrieval.objects.filter(retrieval_id__in=retrieval_ids).update(
                retrieval_carrier_planned=batch_carrier
            )
        )()
        template, context = await self.oct_handle_warehouse_post(request)
        return template, context


    async def _get_pallet(self, warehouse: str, eta_start: str, eta_end: str) -> list[dict]:
        # 时间处理：将年月转换为完整的时间范围（月初到月末最后一秒）
        eta_start_datetime = timezone.datetime.strptime(eta_start, "%Y-%m")
        eta_end_datetime = timezone.datetime.strptime(eta_end, "%Y-%m")
        eta_end_plus_1month = eta_end_datetime + relativedelta(months=1)
        eta_end_last_second = eta_end_plus_1month - relativedelta(seconds=1)

        eta_start_datetime = timezone.make_aware(eta_start_datetime)
        eta_end_datetime = timezone.make_aware(eta_end_last_second)

        filter_conditions = (Q(
            container_number__order__vessel_id__vessel_eta__gte=eta_start_datetime,
            container_number__order__vessel_id__vessel_eta__lte=eta_end_datetime,
        )& ~Q(cancel_notification=True)) | (Q(
            container_number__order__vessel_id__isnull=True,
        )& ~Q(cancel_notification=True))

        # 组合条件：直送订单通过港口映射获取仓点，其他订单直接匹配仓点
        filter_conditions &= Q(
            # 直送订单：港口映射到仓点
            Q(
                container_number__order__order_type="直送",
                container_number__order__vessel_id__destination_port__in=[
                    port for port, wh in PORT_TO_WAREHOUSE_AREA.items() if wh == warehouse
                ]
            ) |
            # 其他订单：直接匹配仓点
            Q(
                container_number__order__order_type__in=["转运组合", "转运"],
                container_number__order__retrieval_id__retrieval_destination_area=warehouse
            )
        )

        return await sync_to_async(list)(
            Order.objects.prefetch_related(
                "container_number",
                "shipment_id",
                "customer_name",
                "retrieval_id",
                "vessel_id",
                "shipment_id__packinglist",
            )
            .filter(filter_conditions)
            .values("container_number__container_number")
            .annotate(
                created_at=Max("created_at"),
                customer_name=F("customer_name__zem_name"),
                retrieval_destination_area=Case(
                    When(order_type="直送",
                        # 直送订单从港口映射获取仓点
                        then=Value(warehouse)
                    ),
                    When(order_type="转运组合",
                        then=F("retrieval_id__retrieval_destination_area")
                    ),
                    When(order_type="转运",
                        then=F("retrieval_id__retrieval_destination_area")
                    ),
                    default=Value(""),
                    output_field=CharField()
                ),
                retrieval_note=F("retrieval_id__note"),
                retrieval_delegation_status = F("retrieval_id__retrieval_delegation_status"),
                actual_release_status = F("retrieval_id__actual_release_status"),
                planned_release_time = F("retrieval_id__planned_release_time"),
                retrieval_carrier_planned=F("retrieval_id__retrieval_carrier_planned"),
                mbl=F("vessel_id__master_bill_of_lading"),
                shipping_line=F("vessel_id__shipping_line"),
                vessel=F("vessel_id__vessel"),
                container_type=F("container_number__container_type"),
                vessel_date=F("vessel_id__vessel_eta") - F("vessel_id__vessel_etd"),
                vessel_etd=F("vessel_id__vessel_etd"),
                vessel_eta=F("vessel_id__vessel_eta"),
                destination_port=F("vessel_id__destination_port"),
                t49_empty_returned_at=F("retrieval_id__t49_empty_returned_at"),
                t49_pod_full_out_at=F("retrieval_id__t49_pod_full_out_at"),
                temp_t49_available_for_pickup=F("retrieval_id__temp_t49_available_for_pickup"),
                retrieval_carrier=F("retrieval_id__retrieval_carrier"),
                ke_destination_num=Sum(
                    Case(
                        When(
                            shipment_id__packinglist__delivery_type="other",
                            shipment_id__packinglist__destination="客户自提",
                            then=1
                        ),
                        default=0,
                        output_field=IntegerField()
                    )
                ),
                # 其他配送类型数量（聚合统计：delivery_type=other 的所有记录数）
                total_other_delivery=Sum(
                    Case(
                        When(
                            shipment_id__packinglist__delivery_type="other",
                            then=1
                        ),
                        default=0,
                        output_field=IntegerField()
                    )
                ),
                # 私仓数量 = 其他配送总数 - 客户自提数量
                si_destination_num=F("total_other_delivery") - F("ke_destination_num"),
                # 重命名柜号字段（前端直接使用 container 字段，无需处理长关联路径）
                container=F("container_number__container_number"),
                # 保留订单类型（便于前端后续扩展逻辑，如按类型筛选）
                order_type=F("order_type"),
                retrieval_destination_precise=F("retrieval_id__retrieval_destination_precise")
            )
            # 最终返回字段（与前端模板的字段严格对应，避免冗余）
            .values(
                "container", "created_at", "customer_name", "retrieval_destination_area",
                "retrieval_note", "mbl", "shipping_line", "vessel", "container_type",
                "vessel_date", "vessel_etd", "vessel_eta", "temp_t49_available_for_pickup",
                "retrieval_carrier", "ke_destination_num", "si_destination_num", "order_type",
                "do_sent", "id", "status", "retrieval_carrier_planned", "retrieval_id__retrieval_id",
                "retrieval_delegation_status", "planned_release_time", "actual_release_status", "destination_port",
                "t49_empty_returned_at", "t49_pod_full_out_at"
            )
            .order_by(
                # 第一优先级：区分"有/无实际放行时间"（有则排前面）
                Case(
                    When(planned_release_time__isnull=False, then=0),  # 有实际放行时间
                    When(planned_release_time__isnull=True, then=1),  # 无实际放行时间
                    default=2,
                    output_field=IntegerField()
                ),
                # 第二优先级：有实际放行时间的按放行时间升序
                Case(
                    When(planned_release_time__isnull=False, then=F('planned_release_time')),
                    default=None
                ),
                # 第三优先级：无实际放行时间的按ETA升序
                Case(
                    When(planned_release_time__isnull=True, then=F('vessel_eta')),
                    default=None
                ),
                # 兜底排序
                'vessel_eta'
            )
        )

    async def _get_pallet_by_order_at(self, warehouse: str, order_start: str, order_end: str) -> list[dict]:
        # 时间处理：将年月转换为完整的时间范围（月初到月末最后一秒）
        order_start_datetime = timezone.datetime.strptime(order_start, "%Y-%m")
        order_end_datetime = timezone.datetime.strptime(order_end, "%Y-%m")
        order_end_plus_1month = order_end_datetime + relativedelta(months=1)
        order_end_last_second = order_end_plus_1month - relativedelta(seconds=1)

        order_start_datetime = timezone.make_aware(order_start_datetime)
        order_end_datetime = timezone.make_aware(order_end_last_second)

        # 构建Q对象用于条件过滤，区分不同订单类型的仓点过滤逻辑
        filter_conditions = Q(
            created_at__gte=order_start_datetime,
            created_at__lte=order_end_datetime,
            cancel_notification__isnull=False,
        )& ~Q(cancel_notification=True)

        # 组合条件：直送订单通过港口映射获取仓点，其他订单直接匹配仓点
        filter_conditions &= Q(
            # 直送订单：港口映射到仓点
            Q(
                container_number__order__order_type="直送",
                container_number__order__vessel_id__destination_port__in=[
                    port for port, wh in PORT_TO_WAREHOUSE_AREA.items() if wh == warehouse
                ]
            ) |
            # 其他订单：直接匹配仓点
            Q(
                container_number__order__order_type__in=["转运组合", "转运"],
                container_number__order__retrieval_id__retrieval_destination_area=warehouse
            )
        )

        return await sync_to_async(list)(
            Order.objects.prefetch_related(
                "container_number",
                "shipment_id",
                "customer_name",
                "retrieval_id",
                "vessel_id",
                "shipment_id__packinglist",
            )
            .filter(filter_conditions)
            .values("container_number__container_number")
            .annotate(
                created_at=Max("created_at"),
                customer_name=F("customer_name__zem_name"),
                retrieval_destination_area=Case(
                    When(order_type="直送",
                        # 直送订单从港口映射获取仓点
                        then=Value(warehouse)
                    ),
                    When(order_type="转运组合",
                        then=F("retrieval_id__retrieval_destination_area")
                    ),
                    When(order_type="转运",
                        then=F("retrieval_id__retrieval_destination_area")
                    ),
                    default=Value(""),
                    output_field=CharField()
                ),
                retrieval_note=F("retrieval_id__note"),
                retrieval_delegation_status=F("retrieval_id__retrieval_delegation_status"),
                actual_release_status=F("retrieval_id__actual_release_status"),
                planned_release_time=F("retrieval_id__planned_release_time"),
                retrieval_carrier_planned=F("retrieval_id__retrieval_carrier_planned"),
                mbl=F("vessel_id__master_bill_of_lading"),
                shipping_line=F("vessel_id__shipping_line"),
                vessel=F("vessel_id__vessel"),
                container_type=F("container_number__container_type"),
                vessel_date=F("vessel_id__vessel_eta") -
                            F("vessel_id__vessel_etd"),
                vessel_etd=F("vessel_id__vessel_etd"),
                vessel_eta=F("vessel_id__vessel_eta"),
                destination_port=F("vessel_id__destination_port"),
                t49_empty_returned_at=F("retrieval_id__t49_empty_returned_at"),
                t49_pod_full_out_at=F("retrieval_id__t49_pod_full_out_at"),
                temp_t49_available_for_pickup=F("retrieval_id__temp_t49_available_for_pickup"),
                retrieval_carrier=F("retrieval_id__retrieval_carrier"),
                ke_destination_num=Sum(
                    Case(
                        When(
                            shipment_id__packinglist__delivery_type="other",
                            shipment_id__packinglist__destination="客户自提",
                            then=1
                        ),
                        default=0,
                        output_field=IntegerField()
                    )
                ),
                # 其他配送类型数量（聚合统计：delivery_type=other 的所有记录数）
                total_other_delivery=Sum(
                    Case(
                        When(
                            shipment_id__packinglist__delivery_type="other",
                            then=1
                        ),
                        default=0,
                        output_field=IntegerField()
                    )
                ),
                # 私仓数量 = 其他配送总数 - 客户自提数量
                si_destination_num=F("total_other_delivery") - F("ke_destination_num"),
                # 重命名柜号字段（前端直接使用 container 字段，无需处理长关联路径）
                container=F("container_number__container_number"),
                # 保留订单类型（便于前端后续扩展逻辑，如按类型筛选）
                order_type=F("order_type"),
                retrieval_destination_precise=F("retrieval_id__retrieval_destination_precise")
            )
            # 最终返回字段（与前端模板的字段严格对应，避免冗余）
            .values(
                "container", "created_at", "customer_name", "retrieval_destination_area",
                "retrieval_note", "mbl", "shipping_line", "vessel", "container_type",
                "vessel_date", "vessel_etd", "vessel_eta", "temp_t49_available_for_pickup",
                "retrieval_carrier", "ke_destination_num", "si_destination_num", "order_type",
                "do_sent", "id", "status", "retrieval_carrier_planned", "retrieval_id__retrieval_id",
                "retrieval_delegation_status", "planned_release_time", "actual_release_status", "destination_port",
                "t49_empty_returned_at", "t49_pod_full_out_at"
            )
            .order_by(
                # 第一优先级：区分"有/无实际放行时间"（有则排前面）
                Case(
                    When(planned_release_time__isnull=False, then=0),  # 有实际放行时间
                    When(planned_release_time__isnull=True, then=1),  # 无实际放行时间
                    default=2,
                    output_field=IntegerField()
                ),
                # 第二优先级：有实际放行时间的按放行时间升序
                Case(
                    When(planned_release_time__isnull=False, then=F('planned_release_time')),
                    default=None
                ),
                # 第三优先级：无实际放行时间的按ETA升序
                Case(
                    When(planned_release_time__isnull=True, then=F('vessel_eta')),
                    default=None
                ),
                # 兜底排序
                'vessel_eta'
            )
        )

    warehouse_options = {
        "所有仓库": "NJ,SAV,LA,MO,TX",
        "NJ": "NJ",
        "SAV": "SAV",
        "LA": "LA",
        "MO": "MO",
        "TX": "TX",
        "CA": "CA"
    }