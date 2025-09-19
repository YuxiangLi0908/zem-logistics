import os
import random
import re
import string
import uuid
from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any

import pandas as pd
from django.contrib.auth.decorators import login_required
from django.db import models
from django.forms import formset_factory
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render
from django.utils.decorators import method_decorator
from django.views import View
from openpyxl import load_workbook
from simple_history.utils import bulk_create_with_history

from warehouse.forms.order_form import OrderForm
from warehouse.forms.quote_form import QuoteForm
from warehouse.forms.upload_file import UploadFileForm
from warehouse.models.fee_detail import FeeDetail
from warehouse.models.order import Order
from warehouse.models.quotation_master import QuotationMaster
from warehouse.models.quote import Quote


@method_decorator(login_required(login_url="login"), name="dispatch")
class QuoteManagement(View):
    template_create = "quote/quote_creation.html"
    template_update = "quote/quote_list.html"
    template_edit = "quote/quote_edit.html"
    template_quote_master = "accounting/quote_master.html"
    template_payable_quote_master = "accounting/payable_quote_master.html"

    def get(self, request: HttpRequest) -> HttpResponse:
        step = request.GET.get("step")
        if step == "new":
            return render(
                request, self.template_create, self.handle_new_quote_get(request)
            )
        elif step == "history":
            return render(
                request, self.template_update, self.handle_history_quote_get(request)
            )
        elif step == "edit":
            return render(request, self.template_edit, self.handle_edit_get(request))
        elif step == "quote_master":
            template, context = self.handle_quote_master_get(request)
            return render(request, template, context)
        elif step == "payable_quote_master":
            template, context = self.handle_payable_quote_master_get(request)
            return render(request, template, context)
        context = {}
        return render(request, self.template_create, context)

    def post(self, request: HttpRequest) -> HttpResponse:
        step = request.POST.get("step")
        if step == "create":
            return render(
                request, self.template_create, self.handle_create_post(request)
            )
        elif step == "export_single_quote_excel":
            return self.handle_single_excel_export(request)
        elif step == "search":
            return render(
                request, self.template_update, self.handle_quote_search_post(request)
            )
        elif step == "update":
            return render(
                request, self.template_update, self.handle_update_post(request)
            )
        elif step == "upload_quote_excel":
            template, context = self.handle_upload_quote_post(request)
            return render(request, template, context)
        elif step == "upload_payable_quote_excel":
            template, context = self.handle_upload_payable_quote_post(request)
            return render(request, template, context)

    def handle_payable_quote_master_get(self, request: HttpRequest) -> dict[str, Any]:
        # 查询历史版本
        quotes = QuotationMaster.objects.filter(quote_type="payable")
        context = {"order_form": OrderForm(), "quotes": quotes}
        return self.template_payable_quote_master, context

    def handle_quote_master_get(self, request: HttpRequest) -> dict[str, Any]:
        # 查询历史版本
        quotes = QuotationMaster.objects.filter(quote_type="receivable").order_by(
            "-effective_date"
        )
        context = {"order_form": OrderForm(), "quotes": quotes}
        return self.template_quote_master, context

    def process_preport_sheet(self, df, file, quote):
        # 从A列，先看下哪些单元格满足：是合并单元格且包含‘提拆’两个字，这种都是仓库的提拆费，因为不确定未来会增减仓库，所以这个是动态的
        file.seek(0)
        file_content = file.read()
        wb = load_workbook(filename=BytesIO(file_content))
        ws = wb["码头费用说明"]
        merge_range = []
        merged_cells = ws.merged_cells.ranges
        for merged_cell in merged_cells:
            if (
                merged_cell.min_col == 1 and merged_cell.max_col == 1
            ):  # 只处理 A 列的合并单元格
                cell_value = ws.cell(row=merged_cell.min_row, column=1).value
                if cell_value and "提拆" in str(cell_value):
                    # 记录合并单元格的范围
                    merge_range.append(
                        (
                            f"A{merged_cell.min_row}-A{merged_cell.max_row}",
                            merged_cell.min_row,
                            merged_cell.max_row,
                        )
                    )
        # 之后，就开始将‘码头费用说明’sheet表存到数据库，其中仓库的提拆打托费的存储方式是'warehouse':{'20':x,'40':x,'45':x}，剩下的都是字段和费用一一对应的
        result = {}
        max_row = 0  # 这个是记一下表示仓库的提拆费用的最后一行是多少
        for cell_range, start_row, end_row in merge_range:
            max_row = max(max_row, end_row)
            A_fee = df.iloc[start_row - 2, 0]  # 获取A列包含仓库的单元格值
            match = re.match(r"[a-zA-Z]+", A_fee)
            warehouse = match.group()  # 只提取其中的仓库名
            container_dict = {}
            for row in range(start_row - 2, end_row - 1):
                b_value = df.iloc[row, 1]
                b_value = re.findall(r"\d+", b_value)[0]
                c_value = df.iloc[row, 2]
                container_dict[b_value] = c_value
            result[warehouse] = container_dict
        chassis = df.iloc[max_row - 1, 1]
        if pd.isna(chassis) or chassis == '' or chassis is None:
            max_row = max_row + 1
            if max_row < len(df):  
                chassis = df.iloc[max_row - 1, 1]
            else:
                raise ValueError('缺少拆柜费')
        variable = re.findall(r"[\u4e00-\u9fff]+", chassis)[0]
        result[variable] = (
            "$" + str(df.iloc[max_row - 1, 2]) + "/" + df.iloc[max_row - 1, 3]
        )
        remaining_data = df.iloc[max_row:]
        for row in remaining_data.itertuples(index=False):
            if pd.isnull(row[0]):  # 表示是提柜杂费，因为提柜杂费之外的，没有价格提示
                if not pd.isna(row[2]) and not pd.isna(
                    row[3]
                ):  # 如果C列D列有费用，再加入
                    if (
                        "/" in row[1]
                    ):  # 没有/表示这个收费项目没有中文名和英文名，只有一个名，就不需要提取其中的中文内容
                        variable = re.findall(r"[\u4e00-\u9fff]+", row[1])[0]  # 键
                        if "/" in str(row[2]):
                            result[variable] = row[
                                4
                            ]  # C列有/说明，这个价格不是固定的，要看E列的说明
                        else:
                            result[variable] = "$" + str(row[2]) + "/" + row[3]
                    else:
                        if "/" in str(row[2]):
                            result[variable] = row[4]
                        else:
                            result[row[1]] = "$" + str(row[2]) + "/" + row[3]
        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "preport",
            "details": result,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    # def process_warehouse_sheet(self, df, file, quote):  #这是库内费用ABC为第一列，D为第二列，E-H为第三列时的方法
    #     # 库内操作表没有合并的问题，一行一行读就行
    #     result = {}
    #     for index, row in df.iterrows():
    #         if index == 0:
    #             continue
    #         else:
    #             print(row.iloc[0],row.iloc[3],row.iloc[4])
    #             if "激活" in row.iloc[0]:  # 激活那条报价和其他条位置不一样
    #                 result[row.iloc[0]] = row.iloc[4]
    #             else:
    #                 result[row.iloc[0]] = row.iloc[3]
    #     # 创建 FeeDetail 记录
    #     fee_detail_data = {
    #         "quotation_id": quote,
    #         "fee_detail_id": str(uuid.uuid4())[:4].upper(),
    #         "fee_type": "warehouse",
    #         "details": result,
    #     }

    #     fee_detail = FeeDetail(**fee_detail_data)
    #     print(fee_detail)
    #     fee_detail.save()
    def process_warehouse_sheet(self, df, file, quote):
        result = {}

        # 检查列数是否足够（现在只有ABC三列，索引0-2）
        if len(df.columns) < 3:
            raise ValueError("上传的表格需要至少包含ABC三列数据")

        # 预先计算列索引（ABC三列）
        col_a_idx = 0  # A列索引（项目名称/标识）
        col_b_idx = 1  # B列索引（常规价格）
        col_c_idx = 2  # C列索引（激活价格）

        for index, row in df.iterrows():
            if index == 0:  # 跳过表头
                continue

            # 获取当前行数据（只取ABC三列）
            project_name = str(row.iloc[col_a_idx]).strip()  # A列
            normal_price = row.iloc[col_b_idx]  # B列
            active_price = row.iloc[col_c_idx]  # C列

            # 空值处理（如果单元格可能为空）
            normal_price = 0 if pd.isna(normal_price) else normal_price
            active_price = 0 if pd.isna(active_price) else active_price

            if "激活" in project_name:
                result[project_name] = active_price  # 使用C列值
            else:
                result[project_name] = normal_price  # 使用B列值

        # 创建并保存记录
        fee_detail = FeeDetail(
            quotation_id=quote,
            fee_detail_id=str(uuid.uuid4())[:4].upper(),
            fee_type="warehouse",
            details=result,
        )
        fee_detail.save()

    def process_nj_local_sheet(self, df, file, quote):
        # NJ本地派送也没有合并的问题，一行一行读
        result = {}
        # 从第一行的标题判断是哪一列开始出现邮编的
        zipcode_col_index = None
        for col_idx, col_name in enumerate(df.columns[:10]):
            if "zipcode" in col_name.lower():  # 检查列名
                zipcode_col_index = col_idx
                break
        for index, row in df.iloc[1:].iterrows():
            if (
                not pd.isnull(row.iloc[1])
                and not pd.isnull(row.iloc[2])
                and not pd.isnull(row.iloc[3])
            ):
                # zipcodes = row.iloc[5:].dropna().tolist()
                zipcodes = list(
                    map(int, row.iloc[zipcode_col_index:].dropna().tolist())
                )
                result[index] = {
                    "zipcodes": zipcodes,
                    "prices": [row.iloc[1], row.iloc[2], row.iloc[3]],
                }
        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "NJ_LOCAL",
            "details": result,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    def process_nj_amazon_sheet(self, df, file, quote):
        result_amazon_dict = defaultdict(set)
        result_walmart_dict = defaultdict(set)
        niche_warehouse = set()
        flag_walmart = 0
        for index, row in df.iloc[1:].iterrows():
            cell_value = str(row.iloc[3])
            if pd.notna(row.iloc[3]) and "冷门仓点" in cell_value:
                niche_warehouse.add(row.iloc[0])
            if not flag_walmart and "walmart" in str(row.iloc[0]).strip().lower():
                flag_walmart = True
                continue
            if pd.notna(row.iloc[0]) and pd.notna(row.iloc[2]):
                if flag_walmart:
                    result_walmart_dict[row.iloc[2]].add(row.iloc[0])
                else:
                    result_amazon_dict[row.iloc[2]].add(row.iloc[0])
        result = {
            "NJ_WALMART": {
                k: self.extract_locations(list(v))
                for k, v in result_walmart_dict.items()
            },
            "NJ_AMAZON": {
                k: self.extract_locations(list(v))
                for k, v in result_amazon_dict.items()
            },
        }
        niche_warehouse = self.extract_locations(niche_warehouse)

        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "NJ_PUBLIC",
            "details": result,
            "niche_warehouse": niche_warehouse,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    # 这个表格要求两个价格列不能出现别的内容，只有价格，每个价格组之间有至少一个空行
    def process_nj_combina_sheet(self, df, file, quote):
        result = {}
        group = []
        current_region = None
        current_prices = None
        niche_warehouse = set()
        for index, row in df.iterrows():
            cell_value = str(row.iloc[5])
            if pd.notna(row.iloc[5]) and "冷门仓点" in cell_value:
                niche_warehouse.add(row.iloc[1])
            if pd.notna(row.iloc[3]) and pd.notna(
                row.iloc[4]
            ):  # 如果值不为空，说明是一个新的价格组
                if group and current_region:
                    if current_region not in result:
                        result[current_region] = []
                    result[current_region].append(
                        {
                            "prices": current_prices,
                            "location": self.extract_locations(group),
                        }
                    )
                    group = []
                current_region = row.iloc[0]
                current_prices = [row.iloc[3], row.iloc[4]]
                if pd.notna(row.iloc[1]):
                    group.append(row.iloc[1])
            elif pd.notna(row.iloc[1]):  # 没有价格但是有仓库，就只记录仓库
                group.append(row.iloc[1])
            elif all(pd.isna(row.iloc[i]) for i in range(0, 5)):  # 空行，就记录上一组
                if group and current_region:
                    if current_region not in result:
                        result[current_region] = []
                    result[current_region].append(
                        {
                            "prices": current_prices,
                            "location": self.extract_locations(group),
                        }
                    )
                    group = []  # 重置仓库组
                    current_region = None  # 重置 region
        if current_region and current_prices and group:  # 记录最后一组
            if current_region not in result:
                result[current_region] = []
            result[current_region].append(
                {"prices": current_prices, "location": self.extract_locations(group)}
            )
        niche_warehouse = self.extract_locations(niche_warehouse)
        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "NJ_COMBINA",
            "details": result,
            "niche_warehouse": niche_warehouse,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    def process_la_amazon_sheet(self, df, file, quote):
        result_amazon_dict = defaultdict(set)
        niche_warehouse = set()
        for index, row in df.iloc[1:].iterrows():
            if pd.notna(row.iloc[1]) and pd.notna(row.iloc[3]):
                if len(row) >= 6 and pd.notna(row.iloc[5]):
                    if "冷门仓点" in row.iloc[5]:
                        niche_warehouse.add(row.iloc[1])
                result_amazon_dict[row.iloc[3]].add(row.iloc[1])
        result = {
            key: self.extract_locations(list(values))
            for key, values in result_amazon_dict.items()
        }
        niche_warehouse = self.extract_locations(niche_warehouse)
        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "LA_PUBLIC",
            "details": result,
            "niche_warehouse": niche_warehouse,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    def process_la_combina_new_sheet(self, df, file, quote):
        result = {}
        group = []
        region = None
        current_prices = None
        for index, row in df.iloc[1:].iterrows():
            if pd.notna(row.iloc[1]):
                if (
                    pd.notna(row.iloc[2])
                    and pd.notna(row.iloc[3])
                    and pd.notna(row.iloc[4])
                ):  # 如果值不为空，说明是一个新的价格组
                    if (
                        len(group) > 0
                    ):  # 首先要判断下，如果这是一个新价格组，且已经记录了上一个价格组，先存储上一组价格组
                        if region not in result:
                            result[region] = []
                        new_group = [
                            item.split("-")[-1] if "Walmart" in item else item
                            for item in group
                        ]
                        result[region].append(
                            {
                                "prices": current_prices,
                                "location": new_group,
                            }
                        )
                        group = []
                    # 然后开始存新的价格组，直接存价格和第一行仓库代码
                    region = (
                        region if pd.isna(row.iloc[0]) else row.iloc[0]
                    )  # A区B区这种是外键
                    current_prices = [row.iloc[2], row.iloc[3], row.iloc[4]]
                    group_partial = row.iloc[1].replace("\n", "/").split("/")
                    group.extend(group_partial)
                else:  # 如果这一行的价格已记录，直接加仓库代码就可以
                    group_partial = row.iloc[1].replace("\n", "/").split("/")
                    group.extend(group_partial)

        if region and current_prices and group:
            if region not in result:
                result[region] = []
            result[region].append(
                {"prices": current_prices, "location": self.extract_locations(group)}
            )

        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "LA_COMBINA",
            "details": result,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    def process_la_combina_sheet(self, df, file, quote):
        result = {}
        group = []
        niche_warehouse = set()
        region = None
        current_prices = None
        for index, row in df.iloc[1:].iterrows():
            if pd.notna(row.iloc[5]) and pd.notna(
                row.iloc[6]
            ):  # 如果值不为空，说明是一个新的价格组
                if (
                    len(group) > 0
                ):  # 首先要判断下，如果这是一个新价格组，且已经记录了上一个价格组，先存储上一组价格组
                    if region not in result:
                        result[region] = []
                    result[region].append(
                        {
                            "prices": current_prices,
                            "location": self.extract_locations(group),
                        }
                    )
                    group = []
                # 然后开始存新的价格组，直接存价格和第一行仓库代码
                region = (
                    region if pd.isna(row.iloc[0]) else row.iloc[0]
                )  # A区B区这种是外键
                current_prices = [row.iloc[5], row.iloc[6]]
                group.extend(row.iloc[i] for i in range(2, 5) if pd.notna(row.iloc[i]))
            else:  # 如果这一行的价格已记录，直接加仓库代码就可以
                group.extend(row.iloc[i] for i in range(2, 5) if pd.notna(row.iloc[i]))
                if pd.notna(row.iloc[1]):
                    group.append(row.iloc[1])
            # 查看每一行是否有写冷门仓点
            cell_value = str(row.iloc[7])
            if pd.notna(row.iloc[7]) and "冷门仓点" in cell_value:
                warehouses = cell_value.split("冷门仓点：")[-1].split("、")
                niche_warehouse.update([w.strip() for w in warehouses if w.strip()])
        if region and current_prices and group:
            if region not in result:
                result[region] = []
            result[region].append(
                {"prices": current_prices, "location": self.extract_locations(group)}
            )
        niche_warehouse = self.extract_locations(niche_warehouse)

        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "LA_COMBINA",
            "details": result,
            "niche_warehouse": niche_warehouse,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    def process_sav_amazon_sheet(self, df, file, quote):
        result_amazon_dict = defaultdict(set)
        result_walmart_dict = defaultdict(set)
        niche_warehouse = set()
        flag_walmart = 0
        for index, row in df.iloc[1:].iterrows():
            cell_value = str(row.iloc[3])
            if pd.notna(row.iloc[3]) and "冷门仓点" in cell_value:
                niche_warehouse.add(row.iloc[0])
            if not flag_walmart and "walmart" in str(row.iloc[0]).strip().lower():
                flag_walmart = True  # 如果某一行第一列写了walmart，就切换result_walmart_dict处理后续工作
                continue
            if pd.notna(row.iloc[0]) and pd.notna(
                row.iloc[2]
            ):  # 第一列是仓点，第三列是价格
                if flag_walmart:
                    result_walmart_dict[row.iloc[2]].add(row.iloc[0])
                else:
                    result_amazon_dict[row.iloc[2]].add(row.iloc[0])
        result = {
            "SAV_WALMART": {
                k: self.extract_locations(list(v))
                for k, v in result_walmart_dict.items()
            },
            "SAV_AMAZON": {
                k: self.extract_locations(list(v))
                for k, v in result_amazon_dict.items()
            },
        }
        niche_warehouse = self.extract_locations(niche_warehouse)
        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "SAV_PUBLIC",
            "details": result,
            "niche_warehouse": niche_warehouse,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    def process_sav_combina_sheet(self, df, file, quote):
        result = {}
        group = []
        current_region = None
        current_prices = None
        niche_warehouse = set()
        for index, row in df.iterrows():
            cell_value = str(row.iloc[5])
            if pd.notna(row.iloc[5]) and "冷门仓点" in cell_value:
                niche_warehouse.add(row.iloc[1])
            if pd.notna(row.iloc[3]) and pd.notna(
                row.iloc[4]
            ):  # 如果值不为空，说明是一个新的价格组
                if group and current_region:
                    if current_region not in result:
                        result[current_region] = []
                    result[current_region].append(
                        {
                            "prices": current_prices,
                            "location": self.extract_locations(group),
                        }
                    )
                    group = []
                current_region = row.iloc[0]
                current_prices = [row.iloc[3], row.iloc[4]]
                if pd.notna(row.iloc[1]):
                    group.append(row.iloc[1])
            elif pd.notna(row.iloc[1]):  # 没有价格但是有仓库，就只记录仓库
                group.append(row.iloc[1])
            elif all(pd.isna(row.iloc[i]) for i in range(0, 5)):  # 空行，就记录上一组
                if group and current_region:
                    if current_region not in result:
                        result[current_region] = []
                    result[current_region].append(
                        {
                            "prices": current_prices,
                            "location": self.extract_locations(group),
                        }
                    )
                    group = []  # 重置仓库组
                    current_region = None  # 重置 region
        if current_region and current_prices and group:  # 记录最后一组
            if current_region not in result:
                result[current_region] = []
            result[current_region].append(
                {"prices": current_prices, "location": self.extract_locations(group)}
            )
        niche_warehouse = self.extract_locations(niche_warehouse)
        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "SAV_COMBINA",
            "details": result,
            "niche_warehouse": niche_warehouse,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    def process_payable_direct_sheet(self, df, file, quote):
        header_row_idx = None
        for idx, row in df.iterrows():
            if "目的港" in str(row.iloc[0]):
                header_row_idx = idx
                break

        if header_row_idx is None:
            raise ValueError("未找到包含'目的港'的标题行")
        # 找到供应商价格部分的最大列值，因为供应商和后面ZEM的有重复列名，是以找到的第一个单元格为价格的前一个单元格的列值为准
        max_col = len(df.columns)
        max_col = 0
        for idx, row in df.iterrows():
            for col_idx, cell in enumerate(row.iloc[0:]):
                if isinstance(cell, str) and "价格" in cell:
                    max_col = col_idx - 1
        # 获取各列的位置映射
        col_mapping = {
            "warehouse": None,  # 目的港
            "warehouse_precise": None,  # 仓点
            "carrier": None,  # 供应商
            "price": None,  # 报价
            "chassis": None,  # 等候费
        }
        for col_idx, cell in enumerate(df.iloc[header_row_idx]):
            if col_idx > max_col:
                break
            cell_value = str(cell).strip()
            if cell_value == "目的港":
                col_mapping["warehouse"] = col_idx
            elif cell_value == "仓点":
                col_mapping["warehouse_precise"] = col_idx
            elif cell_value == "供应商":
                col_mapping["carrier"] = col_idx
            elif cell_value == "报价":
                col_mapping["price"] = col_idx
            elif cell_value == "托架费/等候费/存储费":
                col_mapping["chassis"] = col_idx
        result = defaultdict(lambda: defaultdict(dict))

        # 处理数据行
        for idx, row in df.iterrows():
            # 跳过标题行和空行
            if idx <= header_row_idx or pd.isna(row.iloc[col_mapping["warehouse"]]):
                continue

            warehouse = str(row.iloc[col_mapping["warehouse"]]).strip()
            warehouse_precise = str(row.iloc[col_mapping["warehouse_precise"]]).strip()
            carrier = str(row.iloc[col_mapping["carrier"]]).strip()
            price = str(row.iloc[col_mapping["price"]]).strip()
            chassis = str(row.iloc[col_mapping["chassis"]]).strip()

            # 构建记录
            if "NJ" in warehouse or "NY" in warehouse:
                warehouse = "NJ"
            warehouse_precise = warehouse_precise.replace(" ", "")
            if "/" in warehouse_precise:
                warehouse_list = []
                parts = warehouse_precise.split("/")
                prefix = "".join(
                    [char for char in parts[0] if not char.isdigit()]
                )  # 提取前缀（去掉数字）
                nums = []
                for part in parts:
                    if part.isdigit():  # 如果是纯数字
                        nums.append(part)
                    else:  # 如果不是纯数字（如 ONT8），提取末尾的数字
                        nums.append("".join([char for char in part if char.isdigit()]))
                # 组合前缀和数字
                warehouse_list.extend([f"{prefix}{num}" for num in nums])
                for wh in warehouse_list:
                    result[warehouse][wh][carrier] = {
                        "price": float(price),
                        "chassis": chassis,
                    }
            else:
                result[warehouse][warehouse_precise][carrier] = {
                    "price": float(price),
                    "chassis": chassis,
                }
        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "PAYABLE_DIRECT",
            "details": result,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    def process_payable_sheet(self, df, file, quote):
        header_row_idx = None
        for idx, row in df.iterrows():
            if "目的港" in str(row.iloc[0]):
                header_row_idx = idx
                break

        if header_row_idx is None:
            raise ValueError("未找到包含'目的港'的标题行")

        # 找到供应商价格部分的最大列值，因为供应商和后面ZEM的有重复列名，是以找到的第二个单元格为提柜费时的前一个单元格的列值为准
        max_col = len(df.columns)
        flag = False
        max_col = 0
        for idx, row in df.iterrows():
            for col_idx, cell in enumerate(row.iloc[0:]):
                if isinstance(cell, str) and not flag and "提柜费" in cell:
                    flag = True
                elif isinstance(cell, str) and flag and "提柜费" in cell:
                    max_col = col_idx - 1
            if flag:
                break

        # 获取各列的位置映射
        col_mapping = {
            "warehouse": None,  # 目的港
            "warehouse_precise": None,  # 仓库
            "carrier": None,  # 供应商
            "basic_40": None,  # 40尺
            "basic_45": None,  # 45尺
            "chassis": None,  # 车架费
            "overweight": None,  # 提柜超重费
            "arrive_warehouse": None,  # 入库
            "palletization": None,  # 拆柜
        }
        # 根据标题名确定标题的列
        for col_idx, cell in enumerate(df.iloc[header_row_idx]):
            if col_idx > max_col:
                break
            cell_value = str(cell).strip()
            if cell_value == "目的港":
                col_mapping["warehouse"] = col_idx
            elif cell_value == "仓库":
                col_mapping["warehouse_precise"] = col_idx
            elif cell_value == "供应商":
                col_mapping["carrier"] = col_idx
            elif cell_value == "40尺":
                col_mapping["basic_40"] = col_idx
            elif cell_value == "45尺":
                col_mapping["basic_45"] = col_idx
            elif cell_value == "车架费价格":
                col_mapping["chassis"] = col_idx
            elif cell_value == "车架费免费天数":
                col_mapping["chassis_free_day"] = col_idx
            elif "提柜超重费" in cell_value:
                col_mapping["overweight"] = col_idx

        # 处理入库和拆柜列
        prev_row = df.iloc[header_row_idx - 1] if header_row_idx > 0 else None
        if prev_row is not None:
            for col_idx, cell in enumerate(prev_row):
                if col_idx > max_col:
                    break
                cell_value = str(cell).strip()
                if "入库" in cell_value:
                    col_mapping["arrive_warehouse"] = col_idx
                elif "拆柜" in cell_value:
                    col_mapping["palletization"] = col_idx
        # 初始化结果数据结构（使用defaultdict自动创建嵌套结构）
        result = defaultdict(lambda: defaultdict(dict))

        # 处理数据行
        for idx, row in df.iterrows():
            # 跳过标题行和空行
            if idx <= header_row_idx or pd.isna(row.iloc[col_mapping["warehouse"]]):
                continue

            warehouse = str(row.iloc[col_mapping["warehouse"]]).strip()
            warehouse_precise = str(row.iloc[col_mapping["warehouse_precise"]]).strip()
            carrier = str(row.iloc[col_mapping["carrier"]]).strip()

            # 处理合并单元格情况（入库和拆柜）
            arrive_warehouse = None
            palletization = None
            if col_mapping["arrive_warehouse"] is not None:
                if col_mapping["palletization"] is None:  # 合并单元格情况
                    arrive_warehouse = row.iloc[col_mapping["arrive_warehouse"]]
                else:
                    arrive_warehouse = row.iloc[col_mapping["arrive_warehouse"]]
                    palletization = row.iloc[col_mapping["palletization"]]

            # 构建记录
            record = {
                "basic_40": (
                    row.iloc[col_mapping["basic_40"]]
                    if col_mapping["basic_40"] is not None
                    else None
                ),
                "basic_45": (
                    row.iloc[col_mapping["basic_45"]]
                    if col_mapping["basic_45"] is not None
                    else None
                ),
                "chassis": (
                    row.iloc[col_mapping["chassis"]]
                    if col_mapping["chassis"] is not None
                    else None
                ),
                "chassis_free_day": (
                    row.iloc[col_mapping["chassis_free_day"]]
                    if col_mapping["chassis_free_day"] is not None
                    else None
                ),
                "overweight": (
                    row.iloc[col_mapping["overweight"]]
                    if col_mapping["overweight"] is not None
                    else None
                ),
                "arrive_warehouse": arrive_warehouse,
                "palletization": palletization,
            }

            # 清理空值
            record = {k: v for k, v in record.items() if pd.notna(v)}

            # 添加到结果
            result[warehouse][warehouse_precise][carrier] = record
        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "PAYABLE",
            "details": result,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    def process_combine_stipulate(self, df, file, quote):
        result = {
            "global_rules": {},
            "warehouse_pricing": {},
            "special_warehouse": {},
            "tiered_pricing": {},
        }
        current_section = None
        for index, row in df.iterrows():
            if pd.notna(row.iloc[0]) and "▶" in str(row.iloc[0]):
                current_section = str(row.iloc[0]).strip("▶ ").strip()
                continue
            if (
                str(row.iloc[0]) == "参数名称"  # 检查是否标题行
                or str(row.iloc[0]) == "仓库"  # 检查是否包含warehouse(不区分大小写)
                or str(row.iloc[0]) == "warehouse"
                or not pd.notna(row.iloc[0])  # 检查是否空值
            ):
                continue
            if current_section == "global_rules":
                rule_name = str(row.iloc[0])
                rule_value = row.iloc[1]
                exception_zone = row.iloc[2] if pd.notna(row.iloc[2]) else None
                exception_value = row.iloc[3] if pd.notna(row.iloc[3]) else None

                result["global_rules"][rule_name] = {
                    "default": rule_value,
                    "exceptions": (
                        {exception_zone: exception_value} if exception_zone else {}
                    ),
                }
            elif current_section == "warehouse_pricing":
                warehouse = str(row.iloc[0])
                result["warehouse_pricing"][warehouse] = {
                    "base_40ft": row.iloc[1],
                    "base_45ft": row.iloc[2],
                    "nonmix_40ft": row.iloc[3],
                    "nonmix_45ft": row.iloc[4],
                    "pickup_min": row.iloc[5],
                    "pickup_max": row.iloc[6],
                    "palletizing": row.iloc[7] if pd.notna(row.iloc[7]) else None,
                    "nonmix_53ft": row.iloc[8] if pd.notna(row.iloc[8]) else None,
                }
            elif current_section == "special_warehouse":
                warehouse = str(row.iloc[0])
                destination = str(row.iloc[1])
                multiplier = str(row.iloc[2])
                if warehouse in result["special_warehouse"]:
                    if (
                        result["special_warehouse"][warehouse]["multiplier"]
                        == multiplier
                    ):
                        # 确保destination是列表形式
                        if not isinstance(
                            result["special_warehouse"][warehouse]["destination"], list
                        ):
                            result["special_warehouse"][warehouse]["destination"] = [
                                result["special_warehouse"][warehouse]["destination"]
                            ]
                        result["special_warehouse"][warehouse]["destination"].append(
                            destination
                        )
                        continue
                result["special_warehouse"][warehouse] = {
                    "destination": [destination],  # 始终存储为列表
                    "multiplier": multiplier,
                }
            elif current_section == "tiered_pricing":  # LA附加的超出仓点数量加收费用
                warehouse = str(row.iloc[0])
                min_points = str(row.iloc[1])
                max_points = str(row.iloc[2])
                fee = int(row.iloc[3])
                note = str(row.iloc[4].strip())
                addition_rule = {
                    "min_points": min_points,
                    "max_points": max_points,
                    "fee": fee,
                    "note": note,
                }
                if warehouse not in result["tiered_pricing"]:
                    result["tiered_pricing"][warehouse] = []
                result["tiered_pricing"][warehouse].append(addition_rule)
        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "COMBINA_STIPULATE",
            "details": result,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    def process_direct_sheet(self, df, file, quote):
        # 从A列，先看下哪些单元格满足：是合并单元格且包含‘提拆’两个字，这种都是仓库的提拆费，因为不确定未来会增减仓库，所以这个是动态的
        file.seek(0)
        file_content = file.read()
        wb = load_workbook(filename=BytesIO(file_content))
        ws = wb["整柜直送"]
        merge_range = []
        merged_cells = ws.merged_cells.ranges
        for merged_cell in merged_cells:
            if (
                merged_cell.min_col == 1 and merged_cell.max_col == 1
            ):  # 只处理 A 列的合并单元格
                cell_value = ws.cell(row=merged_cell.min_row, column=1).value
                if cell_value:
                    # 记录合并单元格的范围
                    merge_range.append(
                        (
                            f"A{merged_cell.min_row}-A{merged_cell.max_row}",
                            merged_cell.min_row,
                            merged_cell.max_row,
                        )
                    )
        # 之后，就开始将‘整柜直送’sheet表存到数据库，其中仓库的提拆打托费的存储方式是'warehouse':{'20':x,'40':x,'45':x}，剩下的都是字段和费用一一对应的
        result = {"pickup": {}}
        max_row = 0  # 这个是记一下表示仓库的提拆费用的最后一行是多少
        for cell_range, start_row, end_row in merge_range:
            location = []
            max_row = max(max_row, end_row)
            fee_dict = {}
            location = []
            for row in range(start_row - 2, end_row - 1):
                if pd.notna(df.iloc[row, 3]):  # 如果价格列不是空，说明是一个新的价格组
                    if location:
                        if price in fee_dict:
                            fee_dict[price].extend(location)
                        else:
                            fee_dict[price] = location  # 先把上一组记录上
                    price = df.iloc[row, 3]
                    location = [
                        re.sub(
                            r"^([A-Za-z]{3})\s(\d)$",
                            r"\1\2",  # Remove space between 3 letters and 1 digit
                            f"{prefix}{part}" if part.isdigit() else part,
                        )
                        for i in range(1, 3)
                        if pd.notna(df.iloc[row, i])
                        and not re.search("[\u4e00-\u9fff]", str(df.iloc[row, i]))
                        for part in str(df.iloc[row, i]).split("/")
                        for prefix in [
                            "".join(
                                [
                                    char
                                    for char in str(df.iloc[row, i]).split("/")[0]
                                    if not char.isdigit()
                                ]
                            )
                        ]
                    ]
                else:  # 说明价格键已经有了，这里只加上仓点就行
                    location.extend(
                        re.sub(
                            r"^([A-Za-z]{3})\s(\d)$",
                            r"\1\2",  # Remove space between 3 letters and 1 digit
                            f"{prefix}{part}" if part.isdigit() else part,
                        )
                        for i in range(1, 3)
                        if pd.notna(df.iloc[row, i])
                        and not re.search("[\u4e00-\u9fff]", str(df.iloc[row, i]))
                        for part in str(df.iloc[row, i]).split("/")
                        for prefix in [
                            "".join(
                                [
                                    char
                                    for char in str(df.iloc[row, i]).split("/")[0]
                                    if not char.isdigit()
                                ]
                            )
                        ]
                    )
            if price in fee_dict:
                fee_dict[price].extend(location)
            else:
                fee_dict[price] = location
            for price, locations in fee_dict.items():
                cleaned_locations = [
                    re.sub(r"^([A-Za-z]{3})\s(\d)$", r"\1\2", loc) for loc in locations
                ]
                if price in result["pickup"]:
                    result["pickup"][price].extend(cleaned_locations)
                else:
                    result["pickup"][price] = cleaned_locations
        remaining_data = df.iloc[max_row + 1 :]
        result["二次派送"] = {}
        for row in remaining_data.itertuples(index=False):
            if pd.notna(row[1]) and "报价" not in str(row[1]):
                if "二次派送" in row[0]:  # 因为有两条二次派送记录，涉及的仓点不一样
                    pattern = r"[A-Z]{3}\d(?:/\d)?"
                    locations = re.findall(pattern, row[0])
                    locations = self.extract_locations(locations)
                    result["二次派送"][row[1]] = locations
                else:
                    if "实报实销" in str(row[1]):
                        cell_1 = "实报实销"
                    else:
                        cell_1 = row[1]
                    cell = re.sub(
                        r"[【】（）(){}[\]]+.*?[【】（）(){}[\]]+", "", row[0]
                    ).strip()
                    if "/" in cell:
                        variable = re.findall(r"[\u4e00-\u9fff]+", cell)[0]
                        result[variable] = cell_1
                    else:  # 没有/的表示只有一个名字，可能没有中文名
                        temp = cell.lstrip("0123456789、")
                        result[temp] = cell_1
        # 创建 FeeDetail 记录
        fee_detail_data = {
            "quotation_id": quote,
            "fee_detail_id": str(uuid.uuid4())[:4].upper(),
            "fee_type": "direct",
            "details": result,
        }
        fee_detail = FeeDetail(**fee_detail_data)
        fee_detail.save()

    def extract_locations(self, locations) -> dict[list]:
        result = []
        for loc in locations:
            loc = str(loc)
            loc = (
                loc.split("（")[0].split("(")[0].strip()
            )  # 有的仓点后面会写上（新增），这个不要
            parts = loc.split("-")
            loc = (
                parts[1] if len(parts) > 1 else parts[0]
            ).strip()  # 有的仓点是沃尔玛-xx，只要后面的仓点名
            loc = loc.replace("新增", "")
            if "私人地址" in loc:
                result.append(loc)
            else:
                if "/" in str(loc):
                    # 拆分前缀和数字部分
                    parts = loc.split("/")
                    prefix = "".join(
                        [char for char in parts[0] if not char.isdigit()]
                    )  # 提取前缀（去掉数字）
                    nums = []
                    for part in parts:
                        if part.isdigit():  # 如果是纯数字
                            nums.append(part)
                        else:  # 如果不是纯数字（如 ONT8），提取末尾的数字
                            nums.append(
                                "".join([char for char in part if char.isdigit()])
                            )
                    # 组合前缀和数字
                    result.extend([f"{prefix}{num}" for num in nums])
                else:
                    result.append(loc)
        return result

    def handle_upload_payable_quote_post(self, request: HttpRequest) -> dict[str, Any]:
        order_form = OrderForm(request.POST)
        effective_date = request.POST.get("effective_date")
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data["file"]
            base_name, extension = os.path.splitext(uploaded_file.name)
            updated_count = QuotationMaster.objects.filter(quote_type="payable").count()
            count = updated_count + 1
            # 创建 FeeDetail 记录
            upload_date = datetime.now()
            quotation_id = upload_date.strftime("%m%d") + str(uuid.uuid4())[:4].upper()
            quotation_id = quotation_id.replace(" ", "").upper()
            quote_data = {
                "quotation_id": quotation_id,
                "upload_date": upload_date,
                "version": count,
                "filename": base_name,
                "is_user_exclusive": False,
                "exclusive_user": None,
                "effective_date": effective_date,
                "quote_type": "payable",
            }
            quote = QuotationMaster(**quote_data)
            quote.save()
            # 读表内容
            file = request.FILES["file"]
            excel_file = pd.ExcelFile(file)
            SHEET_HANDLERS = {
                "提拆": self.process_payable_sheet,  # 已验证
                "直送": self.process_payable_direct_sheet,
            }
            for sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name)
                if sheet_name in SHEET_HANDLERS:
                    handler = SHEET_HANDLERS[sheet_name]
                    handler(df, file, quote)
        quotes = QuotationMaster.objects.filter(quote_type="payable")
        context = {"quotes": quotes}
        return self.template_payable_quote_master, context

    def handle_upload_quote_post(self, request: HttpRequest) -> dict[str, Any]:
        order_form = OrderForm(request.POST)
        if order_form.is_valid():
            customer = order_form.cleaned_data.get("customer_name")
        is_user_exclusive = request.POST.get("is_user_exclusive")
        effective_date = request.POST.get("effective_date")
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data["file"]
            base_name, extension = os.path.splitext(uploaded_file.name)
            updated_count = QuotationMaster.objects.filter(
                quote_type="receivable"
            ).count()
            count = updated_count + 1
            # 创建 FeeDetail 记录
            upload_date = datetime.now()
            quotation_id = upload_date.strftime("%m%d") + str(uuid.uuid4())[:4].upper()
            quotation_id = quotation_id.replace(" ", "").upper()
            quote_data = {
                "quotation_id": quotation_id,
                "upload_date": upload_date,
                "version": count,
                "filename": base_name,
                "is_user_exclusive": (
                    (is_user_exclusive.lower() == "on") if is_user_exclusive else False
                ),
                "exclusive_user": customer,
                "effective_date": effective_date,
                "quote_type": "receivable",
            }
            quote = QuotationMaster(**quote_data)
            quote.save()

            file = request.FILES["file"]
            excel_file = pd.ExcelFile(file)

            # 因为7/15之后的组合柜报价格式变了，所以需要判断一下
            la_combina_handler = self.process_la_combina_sheet
            effective_date_obj = datetime.strptime(effective_date, "%Y-%m-%d").date()
            cutoff_date = datetime.strptime("2025-07-15", "%Y-%m-%d").date()
            if effective_date_obj >= cutoff_date:
                la_combina_handler = self.process_la_combina_new_sheet
            SHEET_HANDLERS = {
                "码头费用说明": self.process_preport_sheet,  # 已验证
                "仓库库内操作费": self.process_warehouse_sheet,  # 已验证
                "NJ本地派送": self.process_nj_local_sheet,  # 没有冷门仓点
                "NJ亚马逊派送费": self.process_nj_amazon_sheet,  # 已验证，有仓点分类，记录冷门仓点
                "NJ组合柜": self.process_nj_combina_sheet,  # 已验证，有仓点分类，记录冷门仓点
                "LA亚马逊派送费": self.process_la_amazon_sheet,  # 已验证，有仓点分类，记录冷门仓点
                "LA组合柜": la_combina_handler,  # 已验证，有仓点分类，记录冷门仓点
                "SAV亚马逊派送表": self.process_sav_amazon_sheet,  # 已验证，有仓点分类，记录冷门仓点
                "SAV组合柜": self.process_sav_combina_sheet,  # 已验证，有仓点分类，记录冷门仓点
                "整柜直送": self.process_direct_sheet,  # 已验证
                "组合柜规则": self.process_combine_stipulate,
                # "LA本地派送": self.process_la_local_sheet,
            }
            for sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name)
                if sheet_name in SHEET_HANDLERS:
                    handler = SHEET_HANDLERS[sheet_name]
                    handler(df, file, quote)
        quotes = QuotationMaster.objects.filter(quote_type="receivable")
        context = {"quotes": quotes}
        return self.template_quote_master, context

    def handle_new_quote_get(self, request: HttpRequest) -> dict[str, Any]:
        quote_form = QuoteForm()
        quote_formset = formset_factory(QuoteForm, extra=1)
        context = {
            "quote_form": quote_form,
            "quote_formset": quote_formset,
            "step": "create",
        }
        return context

    def handle_history_quote_get(self, request: HttpRequest) -> dict[str, Any]:
        current_date = datetime.now().date()
        start_date = current_date + timedelta(days=-30)
        end_date = current_date
        context = {
            "start_date": start_date.strftime("%Y-%m-%d"),
            "end_date": end_date.strftime("%Y-%m-%d"),
        }
        return context

    def handle_edit_get(self, request: HttpRequest) -> dict[str, Any]:
        quote_id = request.GET.get("qid")
        quote = Quote.objects.select_related("warehouse", "customer_name").get(
            quote_id=quote_id
        )
        quote_form = QuoteForm(instance=quote)
        context = {
            "quote": quote,
            "quote_form": quote_form,
        }
        return context

    def handle_create_post(self, request: HttpRequest) -> dict[str, Any]:
        quote_form_p1 = QuoteForm(request.POST)
        timestamp = datetime.now()
        if quote_form_p1.is_valid():
            cleaned_data_p1 = quote_form_p1.cleaned_data
            try:
                cleaned_data_p1["note"] = cleaned_data_p1["note"].strip()
            except:
                pass
            try:
                cleaned_data_p1["zipcode"] = cleaned_data_p1["zipcode"].strip().upper()
            except:
                pass
            try:
                cleaned_data_p1["address"] = cleaned_data_p1["address"].strip()
            except:
                pass
            customer = cleaned_data_p1.get("customer_name")
            customer_id = customer.id if customer else 999
            id_prefix = f"QT{timestamp.strftime('%Y%m%d')}{''.join(random.choices(string.ascii_uppercase, k=2))}{customer_id}"
            parent_id = f"{id_prefix}{timestamp.strftime('%H%M')}"
            quote_formset = formset_factory(QuoteForm, extra=1)
            quote_form = quote_formset(request.POST)
            q_valid = all([q.is_valid() for q in quote_form])
            if q_valid:
                cleaned_data_p2 = [q.cleaned_data for q in quote_form]
                i = 1
                quote_data = []
                for d in cleaned_data_p2:
                    data = {}
                    quote_id = f"{parent_id}{''.join(random.choices(string.ascii_uppercase, k=2))}{i}"
                    data.update(cleaned_data_p1)
                    data["created_at"] = timestamp.date()
                    data["parent_id"] = parent_id
                    data["quote_id"] = quote_id
                    data["warehouse"] = d["warehouse"]
                    data["load_type"] = d["load_type"]
                    data["is_lift_gate"] = d["is_lift_gate"]
                    data["cost"] = d["cost"]
                    data["price"] = d["price"]
                    data["distance_mile"] = d["distance_mile"]
                    try:
                        data["comment"] = d["comment"].strip()
                    except:
                        data["comment"] = d["comment"]
                    quote_data.append(data)
                    i += 1
                quote_instances = [Quote(**d) for d in quote_data]
                all_quotes = bulk_create_with_history(quote_instances, Quote)
            else:
                raise RuntimeError(f"invalid 报价!")
        else:
            raise RuntimeError(f"invalid 询价信息!")
        context = {}
        context["parent_id"] = parent_id
        context["all_quotes"] = all_quotes
        context["step"] = "review"
        return context

    def handle_quote_search_post(self, request: HttpRequest) -> dict[str, Any]:
        start_date = request.POST.get("start_date", None)
        end_date = request.POST.get("end_date", None)
        if start_date and end_date:
            criteria = models.Q(created_at__gte=start_date) & models.Q(
                created_at__lte=end_date
            )
        elif start_date:
            criteria = models.Q(created_at__gte=start_date)
        elif end_date:
            criteria = models.Q(created_at__lte=end_date)
        else:
            default_date = datetime.now().date() + timedelta(days=-30)
            criteria = models.Q(created_at__gte=default_date)
        quote = Quote.objects.select_related("warehouse", "customer_name").filter(
            criteria
        )
        context = {"start_date": start_date, "end_date": end_date, "quote": quote}
        return context

    def handle_single_excel_export(self, request: HttpRequest) -> HttpResponse:
        parent_id = request.POST.get("parent_id")
        quote = (
            Quote.objects.select_related("warehouse", "customer_name")
            .filter(parent_id=parent_id)
            .values(
                "quote_id",
                "customer_name__full_name",
                "created_at",
                "warehouse__name",
                "zipcode",
                "address",
                "load_type",
                "is_lift_gate",
                "price",
            )
            .order_by("-price")
        )
        data = [q for q in quote]
        df = pd.DataFrame.from_records(data)
        df = df.rename(
            {
                "quote_id": "询盘号",
                "customer_name__full_name": "客户",
                "created_at": "询盘日期",
                "warehouse__name": "发货仓库",
                "zipcode": "目的地",
                "address": "详细地址",
                "load_type": "FTL/LTL",
                "is_lift_gate": "是否带尾板",
                "price": "报价($)",
            },
            axis=1,
        )
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={parent_id}.xlsx"
        df.to_excel(excel_writer=response, index=False, columns=df.columns)
        return response

    def handle_update_post(self, request: HttpRequest) -> dict[str, Any]:
        quote_id = request.POST.get("quote_id")
        quote = Quote.objects.select_related("warehouse", "customer_name").get(
            quote_id=quote_id
        )
        quote_form = QuoteForm(request.POST)
        if quote_form.is_valid():
            data = quote_form.cleaned_data
            for k, v in data.items():
                if v:
                    setattr(quote, k, v)
            quote.save()
        else:
            raise ValueError(f"invalid quote data: {quote_form}")
        context = self.handle_history_quote_get(request)
        mutable_post = request.POST.copy()
        mutable_post.update(context)
        request.POST = mutable_post
        return self.handle_quote_search_post(request)
