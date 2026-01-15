import io
import json
import math
import re
import json
from collections import defaultdict
from django.utils.timezone import make_aware
from datetime import date, datetime, timedelta, time as datetime_time


from django.db import transaction

from typing import Any, Dict, List

import openpyxl
import openpyxl.workbook
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
from django.contrib.auth.decorators import login_required
from django.contrib.postgres.aggregates import ArrayAgg, StringAgg
from django.db import models, transaction
from django.db.models import (
    BooleanField,
    Case,
    Count,
    F,
    FloatField,
    Sum,
    Value,
    When,
    Q,
    Prefetch,
)
from decimal import Decimal, InvalidOperation
from django.utils.safestring import mark_safe

from django.db import transaction
from django.http import JsonResponse, HttpRequest, HttpResponse, HttpResponseForbidden, QueryDict
from django.shortcuts import render
from django.utils import timezone
from django.views import View
from office365.runtime.auth.user_credential import UserCredential
from office365.runtime.client_request_exception import ClientRequestException
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.sharing.links.kind import SharingLinkKind
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from simple_history.utils import bulk_create_with_history, bulk_update_with_history
from sqlalchemy.util import await_only

from warehouse.forms.order_form import OrderForm
from warehouse.models.container import Container
from warehouse.models.shipment import Shipment
from warehouse.models.customer import Customer
from warehouse.models.fee_detail import FeeDetail
from warehouse.models.fleet_shipment_pallet import FleetShipmentPallet
from warehouse.models.invoicev2 import (
    Invoicev2,
    InvoiceItemv2,
    InvoiceStatusv2,
)
from warehouse.models.order import Order
from warehouse.models.packing_list import PackingList
from warehouse.models.pallet import Pallet
from warehouse.models.quotation_master import QuotationMaster
from warehouse.models.retrieval import Retrieval
from warehouse.models.transaction import Transaction
from warehouse.views.quote_management import QuoteManagement
from warehouse.forms.order_form import OrderForm

from warehouse.views.accounting import Accounting
from warehouse.utils.constants import (
    ACCT_ACH_ROUTING_NUMBER,
    ACCT_BANK_NAME,
    ACCT_BENEFICIARY_ACCOUNT,
    ACCT_BENEFICIARY_ADDRESS,
    ACCT_BENEFICIARY_NAME,
    ACCT_SWIFT_CODE,
    APP_ENV,
    CARRIER_FLEET,
    CONTAINER_PICKUP_CARRIER,
    SP_DOC_LIB,
    SP_CLIENT_ID,
    SP_DOC_LIB,
    SP_PRIVATE_KEY,
    SP_SCOPE,
    SP_TENANT,
    SP_THUMBPRINT,
    SP_URL,
    SYSTEM_FOLDER,
    DELIVERY_METHOD_OPTIONS
)
from warehouse.views.export_file import export_invoice


class ReceivableAccounting(View):
    template_invoice_search = "receivable_accounting/invoice_search.html"
    template_alert_monitoring = "receivable_accounting/alert_monitoring.html"

    template_preport_entry = "receivable_accounting/preport_entry.html"
    template_preport_edit= "receivable_accounting/preport_edit.html"
    template_warehouse_entry = "receivable_accounting/warehouse_entry.html"
    template_warehouse_edit = "receivable_accounting/warehouse_edit.html"

    template_delivery_entry = "receivable_accounting/delivery_entry.html"
    template_delivery_public_edit = "receivable_accounting/delivery_public__edit.html"
    template_delivery_other_edit = "receivable_accounting/delivery_other_edit.html"

    template_confirm_entry = "receivable_accounting/confirm_entry.html"
    template_confirm_transfer_edit = "receivable_accounting/confirm_transfer_entry.html"

    template_invoice_combina_edit = "receivable_accounting/invoice_combina_edit.html"
    template_invoice_statement = "receivable_accounting/invoice_statement.html"
    template_invoice_items_edit = "receivable_accounting/invoice_items_edit.html"

    template_supplementary_entry = "receivable_accounting/supplementary_entry.html"
    template_invoice_items_all = "receivable_accounting/invoice_items_all.html"

    template_completed_bills = "receivable_accounting/completed_bills.html"

    template_financial_statistics = "receivable_accounting/financial_statistics.html"
    template_quotation_management = "receivable_accounting/quotation_management.html"
    
    allowed_group = "accounting"
    warehouse_options = {
        "": "",
        "NJ-07001": "NJ-07001",
        "NJ-08817": "NJ-08817",
        "SAV-31326": "SAV-31326",
        "LA-91761": "LA-91761",
        "LA-91748": "LA-91748",
        "MO-62025": "MO-62025",
        "TX-77503": "TX-77503",
        "LA-91789": "LA-91789",
        "LA-91766": "LA-91766",
        "直送": "直送",
    }
    CATEGORY_STATUS_FIELD = {
        "preport": "preport_status",
        "warehouse_public": "warehouse_public_status",
        "warehouse_other": "warehouse_other_status",
        "delivery_public": "delivery_public_status",
        "delivery_other": "delivery_other_status",
    }

    def get(self, request: HttpRequest) -> HttpResponse:
        # if not self._validate_user_group(request.user):
        #     return HttpResponseForbidden("You are not authenticated to access this page!")
        step = request.GET.get("step", None)
        if step == "invoice_search":  #账单进度
            template, context = self.handle_invoice_search_get(request)
            return render(request, template, context)
        elif step == "alert":  #预警监控
            template, context = self.handle_alert_monitoring_get()
            return render(request, template, context)
        elif step == "preport":  #港前
            context = {"warehouse_options": self.warehouse_options,"order_form": OrderForm()}
            return render(request, self.template_preport_entry, context)
        elif step == "warehouse":  #库内
            context = {"warehouse_options": self.warehouse_options,"order_form": OrderForm()}
            return render(request, self.template_warehouse_entry, context)    
        elif step == "delivery":  # 派送
            context = {"warehouse_options": self.warehouse_options,"order_form": OrderForm()}
            return render(request, self.template_delivery_entry, context)
        elif step == "confirm":
            context = {"warehouse_options": self.warehouse_options,"order_form": OrderForm()}
            return render(request, self.template_confirm_entry, context)
        elif step == "container_confirm":
            template, context = self.handle_container_invoice_confirm_get(request)
            return render(request, template, context)       
        elif step == "supplementary": #补开账单
            context = {"warehouse_options": self.warehouse_options,"order_form": OrderForm()}
            return render(request, self.template_supplementary_entry, context) 

            return render(request, template, context)
        elif step == "finance_stats": #财务统计分析
            template, context = self.handle_financial_statistics_get(request)
            return render(request, template, context)
        elif step == "quotation_management": #报价表管理
            quotes = QuotationMaster.objects.filter(quote_type="receivable")
            context = {"order_form": OrderForm(), "quotes": quotes}
            return render(request, self.template_quotation_management, context)  
        elif step == "container_preport":
            tempalte, context = self.handle_container_preport_post(request)
            return render(request, tempalte, context)
        elif step == "container_warehouse":
            context = self.handle_container_warehouse_post(request)
            return render(request, self.template_warehouse_edit, context)       
        elif step == "container_delivery":
            template, context = self.handle_container_delivery_post(request)
            return render(request, template, context)   
        elif step == "invoice_manual":
            template, context = self.handle_invoice_item_search(request)
            return render(request, template, context)
        else:
            raise ValueError(f"unknow request {step}")

    def post(self, request: HttpRequest) -> HttpResponse:
        # if not self._validate_user_group(request.user):
        #     return HttpResponseForbidden("You are not authenticated to access this page!")
        step = request.POST.get("step", None)
        print('step',step)
        if step == "preport_search":  #港前
            context = self.handle_preport_entry_post(request)
            return render(request, self.template_preport_entry, context)
        elif step == "warehouse_search":  #库内
            context = self.handle_warehouse_entry_post(request)
            return render(request, self.template_warehouse_entry, context)
        elif step == "delivery_search":
            context = self.handle_delivery_entry_post(request)
            return render(request, self.template_delivery_entry, context)
        elif step == "confirm_search":
            template, context = self.handle_confirm_entry_post(request)
            return render(request, template, context)
        elif step == "release_hold":
            template, context = self.handle_release_hold_post(request)
            return render(request, template, context) 
        elif step == "save_single":
            template, context = self.handle_save_single_post(request)
            return render(request, template, context)
        elif step == "save_all":
            context = self.handle_save_all_post(request)
            return render(request, self.template_delivery_entry, context)
        elif step == "save_all_combina":
            context = self.handle_save_all_combina_post(request)
            return render(request, self.template_delivery_entry, context)
        elif step == "save_activation_fees":
            context = self.handle_save_activation_fees(request)
            return render(request, self.template_delivery_entry, context)
        elif step == "convert_type":
            template, context = self.handle_convert_type_post(request)
            return render(request, template, context) 
        elif step == "preport_save":
            context = self.handle_invoice_preport_save(request)
            return render(request, self.template_preport_entry, context)
        elif step == "modify_order_type":
            template, context = self.handle_modify_order_type(request)
            return render(request, template, context)
        elif step == "warehouse_save":
            context = self.handle_invoice_warehouse_save(request)
            return render(request, self.template_warehouse_entry, context)
        elif step == "reject_category":
            template, context = self.handle_reject_category(request)
            return render(request, template, context)
        elif step == "dismiss":
            template, context = self.handle_dismiss_category(request)
            return render(request, template, context)
        elif step == "confirm_save_all":
            # 转运，财务保存
            template, context = self.handle_confirm_save_all(request)
            return render(request, template, context)
        elif step == "manual_process_search":
            template, context = self.handle_manual_process_search(request)
            return render(request, template, context)
        elif step == "confirm_combina_save":
            template, context = self.handle_invoice_confirm_combina_save(request)
            return render(request, template, context)
        elif step == "supplement_order":
            template, context = self.handle_supplement_order_post(request)
            return render(request, template, context)
        elif step == "adjustBalance":
            template, context = self.handle_adjust_balance_save(request)
            return render(request, template, context)
        elif step == "invoice_order_select":
            return self.handle_invoice_order_select_post(request)
        elif step == "invoice_order_batch_export":
            return self.handle_invoice_order_batch_export(request)
        elif step == "invoice_order_delivered":
            template, context = self.handle_invoice_order_batch_delivered(request)
            return render(request, template, context)
        elif step == "invoice_order_reject":
            template, context = self.handle_invoice_order_batch_reject(request)
            return render(request, template, context)
        elif step == "invoice_search":
            template, context = self.handle_invoice_search_get(request)
            return render(request, template, context)
        elif step =="save_manual_invoice_items":
            template, context = self.handle_save_manual_invoice_items(request)
            return render(request, template, context)
        elif step == "generate_manual_excel":
            template, context = self.handle_generate_manual_excel(request)
            return render(request, template, context)
        elif step == "export_invoice":
            return self.handle_export_invoice_post(request)
        else:
            raise ValueError(f"unknow request {step}")

    def handle_save_manual_invoice_items(self, request: HttpRequest):
        """手动保存所有账单记录"""
        context = {}
        container_number = request.POST.get("container_number")
        invoice_number = request.POST.get("invoice_number")
        items_data = request.POST.get("items_data")

        invoice = Invoicev2.objects.get(invoice_number=invoice_number)
        container = Container.objects.get(container_number=container_number)
        items_data_json = request.POST.get("items_data")
        if not items_data_json:
            context.update({"error_messages": "没有接收到数据"})
        else:
            items_data = json.loads(items_data_json)
            for item in items_data:
                item_id = item.get('item_id')
                item_category = item.get('item_category','')
                if item_category in ['delivery_public','delivery_other']:
                    self._save_delivery_items(item, item_id, invoice, container)
                else:
                    self._save_other_items(item, item_id, invoice, container)
        self._update_invoice_total(invoice,container)
        context = {'success_messages':'保存账单明细成功！'}
        return self.handle_invoice_item_search(request,context)

    def _save_delivery_items(self, item_data, item_id, invoice, container):
        """保存派送费用项"""
        delivery_data = {
            'invoice_number': invoice,
            'invoice_type': "receivable",
            'container_number': container,
            'item_category': item_data.get('item_category'),
            'delivery_type': item_data.get('delivery_type'),
            'description': item_data.get('description', ''),
            'warehouse_code': item_data.get('destination', ''),  
            'cbm': float(item_data.get('cbm', 0)),
            'weight': float(item_data.get('weight', 0)),
            'rate': float(item_data.get('rate', 0)),
            'surcharges': float(item_data.get('surcharges', 0)),
            'amount': float(item_data.get('amount', 0)),
            'note': item_data.get('note', ''),
        }
        if item_data.get('delivery_type') == "combine":
            delivery_data['cbm_ratio'] = item_data.get('cbm_ratio')
        else:
            delivery_data['qty'] = item_data.get('qty')
        if item_data.get('po_id'):
            delivery_data['PO_ID'] = item_data.get('po_id')
        
        if item_id:
            # 更新现有记录
            InvoiceItemv2.objects.filter(id=item_id).update(**delivery_data)
        else:
            # 创建新记录
            InvoiceItemv2.objects.create(**delivery_data)

    def _save_other_items(self, item_data, item_id, invoice, container):
        """保存其他费用项"""
        other_data = {
            'invoice_number': invoice,
            'container_number': container,
            'invoice_type': "receivable",
            'item_category': item_data.get('item_category'),
            'description': item_data.get('description', '派送费'),
            'qty': float(item_data.get('qty', 0)),
            'rate': float(item_data.get('rate', 0)),
            'surcharges': float(item_data.get('surcharges', 0)),
            'amount': float(item_data.get('amount', 0)),
            'note': item_data.get('note', ''),
        }
        
        if item_id:
            InvoiceItemv2.objects.filter(id=item_id).update(**other_data)
        else:
            InvoiceItemv2.objects.create(**other_data)
            

    def _update_invoice_total(self, invoice, container):
        """更新发票总金额"""
        # 计算所有费用项的总金额
        total_amount = 0
        
        # 计算其他费用
        receivable_preport_amount = InvoiceItemv2.objects.filter(
            invoice_number=invoice,
            container_number=container,
            invoice_type="receivable",
            item_category="preport",
        ).aggregate(total=models.Sum('amount'))['total'] or 0
        total_amount += float(receivable_preport_amount)

        receivable_wh_public_amount = InvoiceItemv2.objects.filter(
            invoice_number=invoice,
            container_number=container,
            invoice_type="receivable",
            item_category="warehouse_public",
        ).aggregate(total=models.Sum('amount'))['total'] or 0
        total_amount += float(receivable_wh_public_amount)

        receivable_wh_other_amount = InvoiceItemv2.objects.filter(
            invoice_number=invoice,
            container_number=container,
            invoice_type="receivable",
            item_category="warehouse_other",
        ).aggregate(total=models.Sum('amount'))['total'] or 0
        total_amount += float(receivable_wh_other_amount)

        receivable_delivery_public_amount = InvoiceItemv2.objects.filter(
            invoice_number=invoice,
            container_number=container,
            invoice_type="receivable",
            item_category="delivery_public",
        ).aggregate(total=models.Sum('amount'))['total'] or 0
        total_amount += float(receivable_delivery_public_amount)

        receivable_delivery_other_amount = InvoiceItemv2.objects.filter(
            invoice_number=invoice,
            container_number=container,
            invoice_type="receivable",
            item_category="delivery_other",
        ).aggregate(total=models.Sum('amount'))['total'] or 0
        total_amount += float(receivable_delivery_other_amount)

        activation_items_total = InvoiceItemv2.objects.filter(
            invoice_number=invoice,
            container_number=container,
            invoice_type="receivable",
            item_category="activation_fee",
        ).aggregate(total=models.Sum('amount'))['total'] or 0
        total_amount += float(activation_items_total)
        
        # 更新发票
        invoice.receivable_total_amount = total_amount
        invoice.receivable_preport_amount = receivable_preport_amount
        invoice.receivable_wh_public_amount = receivable_wh_public_amount
        invoice.receivable_wh_other_amount = receivable_wh_other_amount
        invoice.receivable_delivery_public_amount = receivable_delivery_public_amount + activation_items_total
        invoice.receivable_delivery_other_amount = receivable_delivery_other_amount
        invoice.save()

    def handle_invoice_item_search(self, request:HttpRequest, context: dict| None = None) -> Dict[str, Any]:
        '''查询全部的账单详情'''
        if not context:
            context = {}
        container_number = request.GET.get("container_number")
        invoice_id = request.GET.get("invoice_id")
        if not container_number:
            container_number = request.POST.get("container_number")
            invoice_id = request.POST.get("invoice_id")

        order = Order.objects.select_related(
            'container_number',
            'customer_name',
            'warehouse',
            'vessel_id',
            'retrieval_id'
        ).get(container_number__container_number=container_number)

        if invoice_id and invoice_id != "None":
            #找到要修改的那份账单
            invoice = Invoicev2.objects.get(id=invoice_id)
            invoice_status, created = InvoiceStatusv2.objects.get_or_create(
                invoice=invoice,
                invoice_type="receivable",
                defaults={
                    "container_number": order.container_number,
                    "invoice": invoice,
                }
            )
        else:
            #说明这个柜子没有创建过账单，需要创建
            invoice, invoice_status = self._create_invoice_and_status(container_number)
            invoice_id = invoice.id

        items = InvoiceItemv2.objects.filter(
            invoice_number=invoice,
            invoice_type="receivable"
        )
        other_items = []
        delivery_items = []
        for item in items:
            if item.item_category == "delivery_public" or item.item_category == "delivery_other":
                delivery_items.append(item)
            else:
                other_items.append(item)
        context.update({
            'other_items': other_items,
            'delivery_items': delivery_items,
            'container_number': container_number,
            'invoice_number': invoice.invoice_number,
            'start_date': request.POST.get("start_date"),
            'end_date': request.POST.get("end_date"),
            'order_type': order.order_type,
        })
        return self.template_invoice_items_all, context

    def handle_manual_process_search(self, request: HttpRequest, context: dict | None = None, ) -> Dict[str, Any]:
        if not context:
            context = {}
        container_number = request.POST.get("container_number")
        warehouse = request.POST.get("warehouse_filter")
        customer = request.POST.get("customer")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        if not container_number:
            current_date = datetime.now().date()
            start_date = (
                (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
                if not start_date
                else start_date
            )
            end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date
            criteria = (
                    Q(cancel_notification=False)
                    & (Q(order_type="转运") | Q(order_type="转运组合"))
                    & Q(vessel_id__vessel_etd__gte=start_date)
                    & Q(vessel_id__vessel_etd__lte=end_date)
                    & Q(offload_id__offload_at__isnull=False)
            )
            if warehouse:
                criteria &= Q(retrieval_id__retrieval_destination_precise=warehouse)
            if customer:
                criteria &= Q(customer_name__zem_name=customer)
        else:
            criteria = (
                Q(container_number__container_number=container_number)
            )

            # 获取基础订单数据
        base_orders = (
            Order.objects
            .select_related(
                "retrieval_id",
                "offload_id",
                "container_number",
                "customer_name",
            )
            .annotate(
                retrieval_time=F("retrieval_id__actual_retrieval_timestamp"),
                empty_returned_time=F("retrieval_id__empty_returned_at"),
                offload_time=F("offload_id__offload_at"),
            )
            .filter(criteria)
            .distinct()
        )
        rows = []

        for o in base_orders:
            container = o.container_number
            if not container:
                continue

            invoices = (
                Invoicev2.objects
                .filter(container_number=container)
            )

            if not invoices.exists():
                continue

            # 一次性取 status，避免 N+1
            status_map = {
                s.invoice_id: s
                for s in InvoiceStatusv2.objects.filter(
                    invoice__in=invoices,
                    invoice_type="receivable",
                )
            }
            # 状态映射字典
            status_mapping = {
                'unstarted': '未录入',
                'in_progress': '录入中',
                'pending_review': '待组长审核',
                'completed': '已完成',
                'rejected': '已驳回',
                'tobeconfirmed': '待确认',
            }
            for invoice in invoices:
                invoice_status = status_map.get(invoice.id)
                if not invoice_status:
                    continue

                rows.append({
                    # ===== Order =====
                    "order_id": o.id,
                    "order_type": o.order_type,
                    "created_at": o.created_at,
                    "offload_time": o.offload_time,

                    "container_id": container.id,
                    "container_number": container.container_number,

                    "customer_name": o.customer_name.zem_name if o.customer_name else None,
                    "warehouse": (
                        o.retrieval_id.retrieval_destination_precise
                        if o.retrieval_id else None
                    ),

                    # ===== Invoice =====
                    "invoice_id": invoice.id,
                    "invoice_number": invoice.invoice_number,
                    "invoice_date": invoice.invoice_date,
                    "invoice_created_at": invoice.created_at,

                    # ===== Status =====
                    "preport_status": status_mapping.get(invoice_status.preport_status, invoice_status.preport_status),
                    "warehouse_public_status": status_mapping.get(invoice_status.warehouse_public_status,
                                                                  invoice_status.warehouse_public_status),
                    "warehouse_other_status": status_mapping.get(invoice_status.warehouse_other_status,
                                                                 invoice_status.warehouse_other_status),
                    "delivery_public_status": status_mapping.get(invoice_status.delivery_public_status,
                                                                 invoice_status.delivery_public_status),
                    "delivery_other_status": status_mapping.get(invoice_status.delivery_other_status,
                                                                invoice_status.delivery_other_status),
                    "finance_status": status_mapping.get(invoice_status.finance_status, invoice_status.finance_status),

                    # ===== Amounts =====
                    "receivable_total_amount": invoice.receivable_total_amount,
                    "receivable_preport_amount": invoice.receivable_preport_amount,
                    "receivable_wh_public_amount": invoice.receivable_wh_public_amount,
                    "receivable_wh_other_amount": invoice.receivable_wh_other_amount,
                    "receivable_delivery_public_amount": invoice.receivable_delivery_public_amount,
                    "receivable_delivery_other_amount": invoice.receivable_delivery_other_amount,
                    "receivable_direct_amount": invoice.receivable_direct_amount,
                })

        context.update({
            "rows": rows,
            "order_form": OrderForm(),
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
            "start_date": start_date,
            "end_date": end_date,
        })

        return self.template_supplementary_entry, context

    def handle_generate_manual_excel(self,request: HttpRequest) -> tuple[Any, Any]:
        '''手动编辑账单时生成新的excel'''
        container_number = request.POST.get("container_number")
        invoice_number = request.POST.get("invoice_number")

        order = Order.objects.get(container_number__container_number=container_number)
        invoice = Invoicev2.objects.get(invoice_number=invoice_number)

        ctx = self._parse_invoice_excel_data(order, invoice)
        ac = Accounting()
        workbook, invoice_data = ac._generate_invoice_excel(ctx)
        invoice.invoice_date = invoice_data["invoice_date"]
        invoice.invoice_link = invoice_data["invoice_link"]
        invoice.save()

        # 返回成功消息
        success_message = f"成功生成新的excel!"
        context = {'success_message': success_message}
        return self.template_invoice_items_all ,context

    def handle_export_invoice_post(self, request: HttpRequest) -> HttpResponse:
        resp, file_name, pdf_file, context = export_invoice(request)
        return resp

    def handle_invoice_search_get(
        self,
        request: HttpRequest,
    ) -> tuple[Any, Any]:
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date
        criteria = models.Q(
            vessel_id__vessel_etd__gte=start_date, vessel_id__vessel_etd__lte=end_date
        )
        warehouse = request.POST.get("warehouse")
        customer = request.POST.get("customer")
        if warehouse:
            criteria &= models.Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= models.Q(customer_name__zem_name=customer)
        orders = Order.objects.select_related(
            "customer_name", "container_number", "receivable_status", "payable_status"
        ).filter(criteria)

        processed_order_items = self.process_orders_display_status_v1(orders)
        selected_customer_id = request.POST.get("customer_name", "")  # 重点：字段名是"customer_name"（Form字段名）
        order_form = OrderForm(
            initial={
                "customer_name": selected_customer_id  # 关键：让表单默认选中该客户
            }
        )
        context = {
            "order_items": processed_order_items,
            "order_form": order_form,
            "selected_customer_id": selected_customer_id,
            "start_date": start_date,
            "end_date": end_date,
            "customer": customer,
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
        }
        return self.template_invoice_search, context

    def process_orders_display_status_v1(self, orders):
        # 应收状态映射
        RECEIVABLE_STATUS_MAPPING = {
            "unstarted": "未开始",
            "in_progress": "录入中",
            "pending_review": "待审核",
            "completed": "已完成",
            "rejected": "已驳回",
            "tobeconfirmed": "待确认",
            "rejected": "已驳回",
            "confirmed": "已确认",
        }
        
        # 应付状态映射（保持原来的）
        STAGE_MAPPING = {
            "unstarted": "未录入",
            "preport": "提拆柜录入阶段",
            "warehouse": "仓库录入阶段",
            "delivery": "派送录入阶段",
            "tobeconfirmed": "待财务确认",
            "confirmed": "财务已确认",
        }

        SUB_STAGE_MAPPING = {
            "pending": "仓库待处理",
            "warehouse_completed": "仓库已完成",
            "delivery_completed": "派送已完成",
            "warehouse_rejected": "仓库已驳回",
            "delivery_rejected": "派送已驳回",
        }

        # 批量查询应收状态
        container_numbers = [order.container_number.container_number for order in orders]
        receivable_statuses = InvoiceStatusv2.objects.filter(
            container_number__container_number__in=container_numbers,
            invoice_type="receivable",
        ).select_related('invoice')

        # 创建映射字典
        status_dict = {}
        for status in receivable_statuses:
            cn = status.container_number.container_number
            if cn not in status_dict:
                status_dict[cn] = []
            status_dict[cn].append(status)

        processed_items = []

        for order in orders:         
            container_number = order.container_number.container_number
            status_list = status_dict.get(container_number, [])
            if not status_list:
                order_copy = self._copy_order_object(order)
                order_copy.display_key_status = "没有状态表"
                order_copy.invoice_number = None
                processed_items.append(order_copy)
            else:
                 # 为每个InvoiceStatusv2创建一行
                for receivable_status in status_list:
                    # 创建order的副本（浅拷贝）
                    order_copy = self._copy_order_object(order)
                    
                    # 设置invoice_number
                    order_copy.invoice_number = getattr(receivable_status.invoice, 'invoice_number', None) if hasattr(receivable_status, 'invoice') else None
                    order_copy.invoice_id = receivable_status.invoice_id
                    
                    # 处理应收状态各个阶段
                    order_copy.display_preport_status = RECEIVABLE_STATUS_MAPPING.get(
                        receivable_status.preport_status, receivable_status.preport_status
                    ) or "未开始"
                    
                    order_copy.display_warehouse_public_status = RECEIVABLE_STATUS_MAPPING.get(
                        receivable_status.warehouse_public_status, receivable_status.warehouse_public_status
                    ) or "未开始"
                    
                    order_copy.display_warehouse_other_status = RECEIVABLE_STATUS_MAPPING.get(
                        receivable_status.warehouse_other_status, receivable_status.warehouse_other_status
                    ) or "未开始"
                    
                    order_copy.display_delivery_public_status = RECEIVABLE_STATUS_MAPPING.get(
                        receivable_status.delivery_public_status, receivable_status.delivery_public_status
                    ) or "未开始"
                    
                    order_copy.display_delivery_other_status = RECEIVABLE_STATUS_MAPPING.get(
                        receivable_status.delivery_other_status, receivable_status.delivery_other_status
                    ) or "未开始"
                    
                    order_copy.display_finance_status = RECEIVABLE_STATUS_MAPPING.get(
                        receivable_status.finance_status, receivable_status.finance_status
                    ) or "未开始"
                    
                    # 状态对应的CSS类
                    order_copy.display_preport_status_class = self._get_status_class(receivable_status.preport_status)
                    order_copy.display_warehouse_public_status_class = self._get_status_class(receivable_status.warehouse_public_status)
                    order_copy.display_warehouse_other_status_class = self._get_status_class(receivable_status.warehouse_other_status)
                    order_copy.display_delivery_public_status_class = self._get_status_class(receivable_status.delivery_public_status)
                    order_copy.display_delivery_other_status_class = self._get_status_class(receivable_status.delivery_other_status)
                    order_copy.display_finance_status_class = self._get_status_class(receivable_status.finance_status)
                    
                    processed_items.append(order_copy)
            
            # 处理应付状态（每个order只处理一次，因为应付状态是order级别的）
            # 注意：应付状态信息会在每个order_copy中相同
            payable_status = getattr(order, "payable_status", None)
            if payable_status:
                raw_stage = payable_status.stage
                raw_public_stage = payable_status.stage_public
                raw_other_stage = payable_status.stage_other
                
                if raw_stage in ["warehouse", "delivery"]:
                    stage1 = SUB_STAGE_MAPPING.get(raw_public_stage, str(raw_public_stage))
                    stage2 = SUB_STAGE_MAPPING.get(raw_other_stage, str(raw_other_stage))
                    display_stage = f"公仓: {stage1} \n私仓: {stage2}"
                else:
                    base_stage = STAGE_MAPPING.get(raw_stage, str(raw_stage)) if raw_stage else "未录入任何费用"
                    display_stage = base_stage
            else:
                display_stage = "未录入"
            
            # 为当前order的所有副本设置相同的应付状态
            for item in processed_items[-len(status_list or [1]):]:  # 获取刚添加的items
                item.display_stage = display_stage

        return processed_items
    
    def _copy_order_object(self, order):
        """创建order对象的浅拷贝"""
        # 简单方法：创建一个新的对象，复制所有属性
        class OrderCopy:
            pass
        
        order_copy = OrderCopy()
        
        # 复制所有属性
        for attr_name in dir(order):
            if not attr_name.startswith('_'):
                try:
                    attr_value = getattr(order, attr_name)
                    # 排除callable方法
                    if not callable(attr_value):
                        setattr(order_copy, attr_name, attr_value)
                except:
                    pass
        
        return order_copy

    def _get_status_class(self, status):
        """根据状态返回对应的CSS类"""
        status_class_map = {
            "completed": "completed",
            "confirmed": "completed",
            "pending_review": "pending",
            "tobeconfirmed": "pending",
            "in_progress": "inprogress",
            "rejected": "rejected",
            "unstarted": "unstarted",
        }
        return status_class_map.get(status, "unstarted")
    
    def handle_invoice_order_batch_delivered(
        self, request: HttpRequest
    ) -> HttpResponse:
        raw = request.POST.get("selectedInvoiceIds", "[]")
        try:
            invoice_id_list = [int(i) for i in json.loads(raw)]
        except (ValueError, TypeError, json.JSONDecodeError):
            invoice_id_list = []

        if not invoice_id_list:
            context = {'error_messages':'未选择任何账单！'}
            return self.handle_confirm_entry_post(request,context)

        with transaction.atomic():
            invoice_updated = Invoicev2.objects.filter(
                id__in=invoice_id_list
            ).update(
                is_invoice_delivered=True
            )
        context = {'success_messages':'账单通知客户成功！'}
        return self.handle_confirm_entry_post(request,context)
    
    def handle_invoice_order_batch_export(self, request: HttpRequest) -> HttpResponse:
        raw = request.POST.get("selectedInvoiceIds", "[]")
        try:
            invoice_id_list = [int(i) for i in json.loads(raw)]
        except (ValueError, TypeError, json.JSONDecodeError):
            invoice_id_list = []

        if not invoice_id_list:
            context = {'error_messages':'未选择任何账单！'}
            return self.handle_confirm_entry_post(request,context)

        invoices = (
            Invoicev2.objects
            .select_related("container_number", "customer")
            .filter(id__in=invoice_id_list)
        )
        container_ids = [
            inv.container_number_id for inv in invoices if inv.container_number_id
        ]
        order_map = {
            order.container_number_id: order
            for order in Order.objects.select_related(
                "retrieval_id", "container_number", "customer_name"
            ).filter(container_number_id__in=container_ids)
        }

        contexts = []
        invoice_numbers = []
        current_date = datetime.now().date()  # 统一使用当前日期生成invoice_number

        for invoice in invoices:
            container_id = invoice.container_number_id
            if not container_id:
                continue

            order = order_map.get(container_id)
            if not order:
                continue
            context = self._parse_invoice_excel_data(
                order=order,
                invoice=invoice,
            )
            contexts.append(context)

            order_id = str(order.id)
            customer_id = order.customer_name.id
            inv_num = f"{current_date.strftime('%Y-%m-%d').replace('-', '')}C{customer_id}{order_id}"
            invoice_numbers.append(inv_num)

        workbook, _ = self._generate_combined_invoice_excel_v1(contexts, invoice_numbers)

        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        response = HttpResponse(
            excel_file.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="combined_invoices.xlsx"'
        return response
    
    def _generate_combined_invoice_excel_v1(
            self,
            contexts: list[dict[Any, Any]],
            invoice_numbers: list[str],
            save_to_sharepoint: bool = False,
    ) -> tuple[openpyxl.workbook.Workbook, dict[Any, Any]]:
        current_date = datetime.now().date()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Combined Invoices"

        # 合并单元格
        cells_to_merge = [
            "A1:E1", "A3:A4", "B3:D3", "B4:D4", "E3:E4", "F3:I4",
            "A5:A6", "B5:D5", "B6:D6", "E5:E6", "F5:I6",
            "A9:B9", "A10:B10", "F1:I1", "C1:E1", "A2:I2",
            "A7:I7", "A8:I8", "C9:I9", "C10:I10", "A11:I11",
        ]
        self._merge_ws_cells(worksheet, cells_to_merge)

        worksheet.column_dimensions["A"].width = 18
        worksheet.column_dimensions["B"].width = 18
        worksheet.column_dimensions["C"].width = 15
        worksheet.column_dimensions["D"].width = 7
        worksheet.column_dimensions["E"].width = 8
        worksheet.column_dimensions["F"].width = 7
        worksheet.column_dimensions["G"].width = 11
        worksheet.column_dimensions["H"].width = 11
        worksheet.column_dimensions["I"].width = 11
        worksheet.row_dimensions[1].height = 40

        worksheet["A1"] = "Zem Elitelink Logistics Inc"
        worksheet["A3"] = "NJ"
        worksheet["B3"] = "27 Engelhard Ave. Avenel NJ 07001"
        worksheet["B4"] = "Contact: Marstin Ma 929-810-9968"
        worksheet["A5"] = "SAV"
        worksheet["B5"] = "1001 Trade Center Pkwy, Rincon, GA 31326, USA"
        worksheet["B6"] = "Contact: Ken 929-329-4323"
        worksheet["A7"] = "E-mail: OFFICE@ZEMLOGISTICS.COM"
        worksheet["A9"] = "BILL TO"
        worksheet["A10"] = contexts[0]["order"].customer_name.zem_name if contexts else ""
        worksheet["F1"] = "Invoice"
        worksheet["E3"] = "Date"
        worksheet["F3"] = current_date.strftime("%Y-%m-%d")
        worksheet["E5"] = "Invoice #"
        worksheet["F5"] = ", ".join(invoice_numbers)

        worksheet["A1"].font = Font(size=20)
        worksheet["F1"].font = Font(size=28)
        worksheet["A3"].alignment = Alignment(vertical="center")
        worksheet["A5"].alignment = Alignment(vertical="center")
        worksheet["E3"].alignment = Alignment(vertical="center")
        worksheet["E5"].alignment = Alignment(vertical="center")
        worksheet["F3"].alignment = Alignment(vertical="center")
        worksheet["F5"].alignment = Alignment(vertical="center")

        worksheet.append([
            "CONTAINER #", "DESCRIPTION", "WAREHOUSE CODE", "CBM", "WEIGHT",
            "QTY", "RATE", "AMOUNT", "NOTE",
        ])
        invoice_item_starting_row = 12
        current_row = 13
        # 总合计（累计所有context）
        grand_total_amount = 0.0
        grand_total_cbm = 0.0
        grand_total_weight = 0.0
        # 记录所有数据行范围（用于设置边框）
        all_data_rows = []

        for context in contexts:
            context_subtotal_amount = 0.0
            context_subtotal_cbm = 0.0
            context_subtotal_weight = 0.0
            context_start_row = current_row

            for d, wc, cbm, weight, qty, r, amt, n in context["data"]:
                if r == {}:
                    continue
                worksheet.append([
                    context["container_number"],
                    d, wc, cbm, weight, qty, r, amt, n
                ])
                # 累计当前context的小计
                context_subtotal_amount += float(amt) if amt else 0.0
                context_subtotal_cbm += float(cbm) if cbm else 0.0
                context_subtotal_weight += float(weight) if weight else 0.0
                current_row += 1

            # 追加当前context的小计行（可选，清晰区分每个context）
            worksheet.append([
                f"Subtotal ({context['container_number']})",  # 显示柜号小计
                None, None, context_subtotal_cbm, context_subtotal_weight,
                None, None, context_subtotal_amount, None
            ])
            current_row += 1

            # 累计到总合计
            grand_total_amount += context_subtotal_amount
            grand_total_cbm += context_subtotal_cbm
            grand_total_weight += context_subtotal_weight

            all_data_rows.extend(range(context_start_row, current_row))

        worksheet.append([
            "Grand Total", None, None, grand_total_cbm, grand_total_weight,
            None, None, grand_total_amount, None
        ])
        grand_total_row = current_row
        current_row += 1
        all_data_rows.append(grand_total_row)

        if all_data_rows:
            min_border_row = invoice_item_starting_row
            max_border_row = grand_total_row

            for row in worksheet.iter_rows(
                    min_row=min_border_row,
                    max_row=max_border_row,
                    min_col=1,
                    max_col=9,
            ):
                for cell in row:
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

        self._merge_ws_cells(worksheet, [f"A{grand_total_row}:C{grand_total_row}"])
        self._merge_ws_cells(worksheet, [f"F{grand_total_row}:G{grand_total_row}"])
        worksheet[f"A{grand_total_row}"].alignment = Alignment(horizontal="center")
        worksheet[f"H{grand_total_row}"].number_format = numbers.FORMAT_NUMBER_00  # 金额格式化
        worksheet[f"H{grand_total_row}"].alignment = Alignment(horizontal="right")

        row_count = current_row
        self._merge_ws_cells(worksheet, [f"A{row_count}:I{row_count}"])
        row_count += 1

        bank_info = [
            f"Beneficiary Name: {ACCT_BENEFICIARY_NAME}",
            f"Bank Name: {ACCT_BANK_NAME}",
            f"SWIFT Code: {ACCT_SWIFT_CODE}",
            f"ACH/Wire Transfer Routing Number: {ACCT_ACH_ROUTING_NUMBER}",
            f"Beneficiary Account #: {ACCT_BENEFICIARY_ACCOUNT}",
            f"Beneficiary Address: {ACCT_BENEFICIARY_ADDRESS}",
            f"Business Beneficiary Address: 215 Durham Park Way, Pooler,GA31322",
            f"Email:FINANCE@ZEMLOGISTICS.COM",
            f"phone: 929-810-9968",
        ]
        for c in bank_info:
            worksheet.append([c])
            self._merge_ws_cells(worksheet, [f"A{row_count}:I{row_count}"])
            row_count += 1
        self._merge_ws_cells(worksheet, [f"A{row_count}:I{row_count}"])

        worksheet["A9"].font = Font(color="00FFFFFF")
        worksheet["A9"].fill = PatternFill(
            start_color="00000000", end_color="00000000", fill_type="solid"
        )

        invoice_data = {
            "invoice_numbers": invoice_numbers,
            "invoice_date": current_date.strftime("%Y-%m-%d"),
            "total_amount": grand_total_amount,
        }
        if save_to_sharepoint:
            pass
        return workbook, invoice_data
    
    def _merge_ws_cells(
        self, ws: openpyxl.worksheet.worksheet, cells: list[str]
    ) -> None:
        for c in cells:
            ws.merge_cells(c)

    def handle_invoice_order_batch_reject(self, request: HttpRequest) -> tuple[Any, Any]:
        raw = request.POST.get("selectedInvoiceIds", "[]")
        try:
            invoice_id_list = [int(i) for i in json.loads(raw)]
        except (ValueError, TypeError, json.JSONDecodeError):
            invoice_id_list = []

        if not invoice_id_list:
            context = {'error_messages':'未选择任何账单！'}
            return self.handle_confirm_entry_post(request,context)

        with transaction.atomic():
            # 重开时，把是组合柜额外费用类型的记录删除，因为再开的时候会重新生成
            InvoiceItemv2.objects.filter(
                invoice_number_id__in=invoice_id_list,
                item_category='combina_extra_fee'
            ).delete()

            status_updated = InvoiceStatusv2.objects.filter(
                invoice_id__in=invoice_id_list,
                invoice_type="receivable",
            ).update(
                finance_status="tobeconfirmed"
            )

            invoice_updated = Invoicev2.objects.filter(
                id__in=invoice_id_list
            ).update(
                is_invoice_delivered=False,
                receivable_is_locked=False
            )


        context = {'success_messages':'账单退回状态成功！'}
        return self.handle_confirm_entry_post(request,context)

    def handle_invoice_order_select_post(self, request: HttpRequest) -> HttpResponse:
        '''生成STMT'''
        raw = request.POST.get("selectedInvoiceIds", "[]")
        try:
            invoice_id_list = [int(i) for i in json.loads(raw)]
        except (ValueError, TypeError, json.JSONDecodeError):
            invoice_id_list = []

        if not invoice_id_list:
            context = {'error_messages':'未选择任何账单！'}
            return self.handle_confirm_entry_post(request,context)

        selected_orders = list(
            Invoicev2.objects.filter(id__in=invoice_id_list)
            .values_list("container_number__container_number", flat=True)
        )

        if selected_orders:
            order = Order.objects.select_related(
                "customer_name", "container_number", "invoice_id"
            ).filter(container_number__container_number__in=selected_orders)
            order_id = [o.id for o in order]
            customer = order[0].customer_name
            current_date = datetime.now().date().strftime("%Y-%m-%d")
            invoice_statement_id = (
                f"{current_date.replace('-', '')}S{customer.id}{max(order_id)}"
            )
            context = {
                "order": order,
                "customer": customer,
                "invoice_statement_id": invoice_statement_id,
                "current_date": current_date,
                "invoice_type": "receivable",
            }
            return render(request, self.template_invoice_statement, context)
        else:
            template, context = self.handle_confirm_entry_post(request)
            return render(request, template, context)


    def handle_adjust_balance_save(self, request: HttpRequest) -> tuple[Any, Any]:
        customer_id = request.POST.get("customerId")
        customer = Customer.objects.get(id=customer_id)
        amount = float(request.POST.get("usdamount"))
        note = request.POST.get("note")
        user = request.user if request.user.is_authenticated else None

        selected_orders = json.loads(request.POST.get("selectedOrders", "[]"))
        selected_orders = list(set(selected_orders))
        # 查账单，按待核销金额从小到大排序
        invoices = Invoicev2.objects.filter(
            container_number__container_number__in=selected_orders
        ).order_by("remain_offset")
        sum_offset = 0.0
        for invoice in invoices:
            if amount <= 0:
                break
            offset_amount = min(amount, invoice.remain_offset)
            sum_offset += offset_amount
            invoice.remain_offset -= offset_amount
            invoice.save()
            amount -= offset_amount

        transaction = Transaction.objects.create(
            customer=customer,
            amount=sum_offset,
            transaction_type="write_off",
            note=note,
            created_by=user,
            created_at=timezone.now(),
        )

        customer.balance = customer.balance - sum_offset
        customer.save()
        return self.handle_confirm_entry_post(
            request
        )
    
    def handle_supplement_order_post(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        '''补开账单'''   
        container_number = request.POST.get("container_number")
        container_number = container_number.strip()

        status = request.POST.get("status")
        
        invoices = Invoicev2.objects.filter(container_number__container_number=container_number)
        if not invoices.exists():
            context = {'error_messages': f'{container_number}当前一份账单都没有录入，不可重开！'}
            if "delivery" in status:
                context = self.handle_delivery_entry_post(request, context)
                return self.template_delivery_entry, context
            elif "warehouse" in status:
                context = self.handle_warehouse_entry_post(request, context)
                return self.template_warehouse_entry, context
        else:
            for invoice in invoices:
                # 查询这个账单对应的状态
                try:
                    status_obj = InvoiceStatusv2.objects.get(
                        invoice=invoice,
                        invoice_type='receivable'
                    )
                    finance_status = status_obj.finance_status
                    if finance_status != "completed":
                        context = {'error_messages': f'{container_number}还存在未开完的账单，不可重开！'}
                        if "delivery" in status:
                            context = self.handle_delivery_entry_post(request, context)
                            return self.template_delivery_entry, context
                        elif "warehouse" in status:
                            context = self.handle_warehouse_entry_post(request, context)
                            return self.template_warehouse_entry, context
                except InvoiceStatusv2.DoesNotExist:
                    context = {'error_messages': f'{container_number}还存在未开完的账单，不可重开！'}
                    if "delivery" in status:
                        context = self.handle_delivery_entry_post(request, context)
                        return self.template_delivery_entry, context
                    elif "warehouse" in status:
                        context = self.handle_warehouse_entry_post(request, context)
                        return self.template_warehouse_entry, context
        #创建一份新的账单
        invoice, invoice_status = self._create_new_invoice_and_status(container_number)
        self._update_invoice_status(invoice_status, status)

        context = {'success_messages': f'{container_number}补开成功，编号为{invoice.invoice_number}！'}
        if "delivery" in status:
            context = self.handle_delivery_entry_post(request, context)
            return self.template_delivery_entry, context
        elif "warehouse" in status:
            context = self.handle_warehouse_entry_post(request, context)
            return self.template_warehouse_entry, context
    
    def _create_new_invoice_and_status(self, container_number: str) -> tuple[Invoicev2, InvoiceStatusv2]:
        """创建账单和状态记录"""
        order = Order.objects.select_related(
            "customer_name", "container_number"
        ).get(container_number__container_number=container_number)
        # 创建 Invoicev2
        current_date = datetime.now().date()
        order_id = str(order.id)
        customer_id = order.customer_name.id

        # 先检查是否已经存在对应柜号的发票       
        invoice = Invoicev2.objects.create(
            container_number=order.container_number,
            invoice_number=f"{current_date.strftime('%Y-%m-%d').replace('-', '')}C{customer_id}{order_id}",
            created_at=current_date,
        )
        
        # 创建 InvoiceStatusv2
        invoice_status = InvoiceStatusv2.objects.create(
            container_number=order.container_number,
            invoice=invoice,
            invoice_type="receivable"
        )
        invoice.save()
        invoice_status.save()
        return invoice, invoice_status
    
    def _update_invoice_status(self, invoice_status, status_field):
        # 定义有效的状态字段
        VALID_STATUS_FIELDS = [
            'preport_status',
            'warehouse_public_status', 
            'warehouse_other_status',
            'delivery_public_status',
            'delivery_other_status'
        ]
        
        # 验证status_field是否有效
        if status_field not in VALID_STATUS_FIELDS:
            raise ValueError(f"无效的状态字段: {status_field}，必须是: {VALID_STATUS_FIELDS}")
        
        # 获取所有状态字段
        ALL_STATUS_FIELDS = [
            'preport_status',
            'warehouse_public_status', 
            'warehouse_other_status',
            'delivery_public_status',
            'delivery_other_status'
        ]
        
        # 更新状态：指定的状态设为unstarted，其他设为completed
        update_data = {}
        for field in ALL_STATUS_FIELDS:
            if field == status_field:
                update_data[field] = 'unstarted'
            else:
                update_data[field] = 'completed'
        
        # 批量更新字段
        for field, value in update_data.items():
            setattr(invoice_status, field, value)
        
        invoice_status.save()
        return invoice_status

    def handle_confirm_save_all(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        context = {}
        container_number = request.POST.get("container_number")
        invoice_number = request.POST.get("invoice_number")
        items_data_json = request.POST.get('items_data')      
        
        #有错误时，要重新加载页面而准备的数据
        get_params = QueryDict(mutable=True)
        get_params["container_number"] = container_number
        get_params["invoice_number"] = request.POST.get("invoice_number")
        
        request.GET = get_params
        if not container_number or not invoice_number or not items_data_json:       
            context.update({'error_messages': '缺少必要参数！'})
            return self.handle_container_invoice_confirm_get(request)

        try:
            items_data = json.loads(items_data_json)
            if isinstance(items_data, str):  # 双重JSON编码的情况
                items_data = json.loads(items_data)
        except json.JSONDecodeError as e:
            context.update({'error_messages': '表格数据解析错误！'})
            return self.handle_container_invoice_confirm_get(request)
        
        # 获取容器和发票对象
        try:
            container = Container.objects.get(container_number=container_number)
            invoice = Invoicev2.objects.get(invoice_number=invoice_number)
        except (Container.DoesNotExist, Invoicev2.DoesNotExist) as e:
            context.update({'error_messages': '查不到柜号或者账单记录'})
            return self.handle_container_invoice_confirm_get(request)
        
        # 存储统计信息
        saved_items = []
        skipped_items = []
        total_amount = 0.0
        for item in items_data:
            try:
                # 处理item_id
                item_id = item.get('item_id')
                
                item_category = item.get('category', 'preport')
                
                # 处理数量、单价、金额等字段
                qty = item.get('qty', '1')
                rate = item.get('rate', '0')
                surcharges = item.get('surcharges', '')
                amount = item.get('amount', '0')
                
                # 转换为浮点数，处理可能的字符串格式
                def to_decimal(v):
                    try:
                        return Decimal(str(v)) if v not in (None, '') else Decimal('0')
                    except (InvalidOperation, ValueError):
                        return Decimal('0')

                qty = to_decimal(item.get('qty', 0))
                rate = to_decimal(item.get('rate', 0))
                surcharges = to_decimal(item.get('surcharges', 0))
                amount = to_decimal(item.get('amount', 0))
        
                # 处理PO_ID和warehouse_code
                po_id = item.get('po_id', '').strip()
                warehouse_code = item.get('warehouse_code', '').strip()
                
                # 如果item_id是'new'，创建新记录
                if item_id == 'new':
                    invoice_item = InvoiceItemv2(
                        container_number=container,
                        invoice_number=invoice,
                        invoice_type='receivable', 
                        item_category=item_category,
                        description=item.get('description', '').strip(),
                        qty=qty,
                        rate=rate,
                        surcharges=surcharges,
                        amount=amount,
                        PO_ID=po_id if po_id else None,
                        warehouse_code=warehouse_code if warehouse_code else None,
                        note=item.get('note', '').strip(),
                        # 以下字段根据业务需求设置
                        cbm=None,
                        weight=None,
                        delivery_type=None
                    )
                    invoice_item.save()
                    saved_items.append({
                        'id': invoice_item.id,
                        'description': invoice_item.description,
                        'amount': invoice_item.amount
                    })
                
                # 如果item_id是数字，更新现有记录
                elif item_id and item_id.isdigit():
                    try:
                        invoice_item = InvoiceItemv2.objects.get(
                            id=int(item_id),
                            invoice_number=invoice
                        )
                        
                        # 更新字段
                        update_fields = [
                            'description', 'qty', 'rate', 'surcharges', 'amount',
                            'PO_ID', 'warehouse_code', 'note', 'item_category'
                        ]
                        
                        invoice_item.description = item.get('description', '').strip()
                        invoice_item.item_category = item_category
                        invoice_item.qty = qty
                        invoice_item.rate = rate
                        invoice_item.surcharges = surcharges
                        invoice_item.amount = amount
                        invoice_item.PO_ID = po_id if po_id else None
                        invoice_item.warehouse_code = warehouse_code if warehouse_code else None
                        invoice_item.note = item.get('note', '').strip()
                        
                        invoice_item.save(update_fields=update_fields)
                        saved_items.append({
                            'id': invoice_item.id,
                            'description': invoice_item.description,
                            'amount': invoice_item.amount
                        })
                        
                    except InvoiceItemv2.DoesNotExist:
                        skipped_items.append({
                            'item_id': item_id,
                            'description': item.get('description', ''),
                            'reason': '找不到对应记录'
                        })
                    
            except Exception as e:
                skipped_items.append({
                    'item_id': item.get('item_id', '未知'),
                    'description': item.get('description', '未知'),
                    'reason': str(e)
                })
                error_messages = f'保存项目失败: {str(e)}, 数据: {item}'
                context.update({'error_messages': error_messages})
                return self.handle_container_invoice_confirm_get(request)
        
        # 更新发票总额
        try:
            invoice.receivable_total_amount = total_amount
            invoice.receivable_is_locked = True
            invoice.save()
        except Exception as e:
            error_messages = f'更新发票总额失败: {str(e)}'
            context.update({'error_messages': error_messages})
            return self.handle_container_invoice_confirm_get(request)
        
        status_obj = InvoiceStatusv2.objects.get(
            invoice=invoice,
            invoice_type='receivable'
        )
        status_obj.finance_status = "completed"
        status_obj.save()
        #生成excel账单
        order = Order.objects.select_related("retrieval_id", "container_number").get(
            container_number__container_number=container_number
        )
        ctx = self._parse_invoice_excel_data(order, invoice)
        
        ac = Accounting()
        workbook, invoice_data = ac._generate_invoice_excel(ctx)
        invoice.invoice_date = invoice_data["invoice_date"]
        invoice.invoice_link = invoice_data["invoice_link"]
        invoice.save()

        # 返回成功消息
        success_message = f"成功保存 {len(saved_items)} 条记录，总额: {total_amount:.2f} USD"
        context.update({'success_message': success_message})
        return self.handle_confirm_entry_post(request)

    def _parse_invoice_excel_data(
        self, order: Order, invoice: Invoicev2
    ) -> dict[str, Any]:
        description = []
        warehouse_code = []
        cbm = []
        weight = []
        qty = []
        rate = []
        amount = []
        note = []

        invoice_item = InvoiceItemv2.objects.filter(
            container_number=order.container_number,
            invoice_number=invoice,
            invoice_type="receivable"
        )

        for item in invoice_item:
            if item.delivery_type == "combine":
                qty.append(item.cbm_ratio)
                note.append(item.region)
            else:
                qty.append(item.qty)
                note.append(item.note)
            description.append(item.description)
            warehouse_code.append(item.warehouse_code)
            cbm.append(item.cbm)
            weight.append(item.weight)           
            rate.append(item.rate)
            amount.append(item.amount)
            
           
        context = {
            "order": order,
            "container_number": order.container_number.container_number,
            "data": zip(
                description, warehouse_code, cbm, weight, qty, rate, amount, note
            ),
        }
        return context
    
    def handle_dismiss_category(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.POST.get("container_number")
        status = request.POST.get("status")
        delivery_type = request.POST.get("delivery_type")
        reject_reason = request.POST.get("reject_reason")
        invoice_number = request.POST.get("invoice_number")

        invoice = Invoicev2.objects.get(invoice_number=invoice_number)
        status_obj = InvoiceStatusv2.objects.get(
            invoice=invoice,
            invoice_type='receivable'
        )
        if status == "delivery":
            if delivery_type == "public":
                status_obj.delivery_public_status = "rejected"
                status_obj.delivery_public_reason = reject_reason
                reject_status = "公仓派送"
            else:
                status_obj.delivery_other_status = "rejected"
                status_obj.delivery_other_reason = reject_reason
                reject_status = "私仓派送"
        elif status == "warehouse":
            if delivery_type == "public":
                status_obj.warehouse_public_status = "rejected"
                status_obj.warehouse_public_reason = reject_reason
                reject_status = "公仓库内"
            else:
                status_obj.warehouse_other_status = "rejected"
                status_obj.warehouse_self_reason = reject_reason
                reject_status = "私仓库内"
        else:
            status_obj.preport_status = "rejected"
            status_obj.preport_reason = reject_reason
            reject_status = "港前"
        status_obj.save()
        container_number = request.POST.get("container_number")
        success_messages = f'{container_number}已驳回到{reject_status}阶段！'
        context = {'success_messages':success_messages}
        return self.handle_confirm_entry_post(request,context)
    
    def handle_reject_category(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        invoice_number = request.POST.get("invoice_number")
        category = request.POST.get("category")
        reject_reason = request.POST.get("reject_reason")
       
        invoice = Invoicev2.objects.get(invoice_number=invoice_number)
        status_obj = InvoiceStatusv2.objects.get(
            invoice=invoice,
            invoice_type='receivable'
        )
        reject_status = None
        if category == "preport":
            status_obj.preport_status = "rejected"
            status_obj.preport_reason = reject_reason
            reject_status = "港前"
        elif category == "warehouse_public":
            status_obj.warehouse_public_status = "rejected"
            status_obj.warehouse_public_reason = reject_reason
            reject_status = "公仓库内"
        elif category == "warehouse_other":
            status_obj.warehouse_other_status = "rejected"
            status_obj.warehouse_self_reason = reject_reason
            reject_status = "私仓库内"
        elif category == "delivery_public":
            status_obj.delivery_public_status = "rejected"
            status_obj.delivery_public_reason = reject_reason
            reject_status = "公仓派送"
        elif category == "delivery_other":
            status_obj.delivery_other_status = "rejected"
            status_obj.delivery_other_reason = reject_reason
            reject_status = "私仓派送"
        elif category == "activation_fee":
            status_obj.delivery_public_status = "rejected"
            status_obj.delivery_public_reason = reject_reason
            reject_status = "私仓派送"
        status_obj.save()

        container_number = request.POST.get("container_number")
        success_messages = f'{container_number}已驳回到{reject_status}阶段！'
        context = {'success_messages':success_messages}
        return self.handle_confirm_entry_post(request,context)
    
    def handle_container_invoice_confirm_get(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.GET.get("container_number")
        invoice_id = request.GET.get("invoice_id")
        status = request.GET.get("status")
        
        invoice = Invoicev2.objects.get(id=invoice_id)
        order = Order.objects.select_related("container_number").get(
            container_number__container_number=container_number
        )
        
        if status == "confirmed":
            #直接读取所有的invoiceitem
            ctx = self._all_invoice_items_get(invoice,container_number)
            ctx['category_totals'] = {
                'preport': invoice.receivable_preport_amount or 0,
                'warehouse_public': invoice.receivable_wh_public_amount or 0,
                'warehouse_other': invoice.receivable_wh_other_amount or 0,
                'delivery_public': invoice.receivable_delivery_public_amount or 0,
                'delivery_other': invoice.receivable_delivery_other_amount or 0,
                'combina_extra': ctx.get('combina_extra_fee'),
            }
            return self.template_invoice_items_edit, ctx
        
        if order.order_type == "直送":
            is_combina = False
        else:
            # 看下是不是补开的账单，如果是补开的按转运方式展示记录就醒了，不用组合柜形式展示了
            other_invoices = Invoicev2.objects.filter(
                container_number=order.container_number 
            ).exclude(id=invoice_id)
            
            if other_invoices.exists(): 
                is_combina = False
            else:
                # 这里要区分一下，如果是组合柜的柜子，跳转就直接跳转到组合柜计算界面
                ctx, is_combina, non_combina_reason = self._is_combina(order.container_number.container_number)
                if ctx.get('error_messages'):
                    return self.template_invoice_combina_edit, ctx
        
        if is_combina:       
            # 这里表示是组合柜的方式计算
            new_get = request.GET.copy()
            new_get['is_new_version'] = True
            request.GET = new_get
            setattr(request, "is_from_account_confirmation", True)
            ctx = self.handle_container_invoice_combina_get(request)
            return self.template_invoice_combina_edit, ctx
        else:
            items = InvoiceItemv2.objects.filter(
                invoice_number=invoice,
                container_number__container_number=container_number,
                invoice_type="receivable"
            ).order_by("item_category", "id")

            # 分组（按 5 大类）
            grouped = {
                "preport": [],
                "warehouse_public": [],
                "warehouse_other": [],
                "delivery_public": [],
                "delivery_other": [],
                "activation_fee": [],
            }

                # 计算每个类别的总金额
            category_totals = {}
            total_amount = 0
            
            for it in items:
                grouped.setdefault(it.item_category, []).append(it)
                if it.amount:
                    total_amount += float(it.amount)
            
            # 计算每个类别的金额
            for category, items_list in grouped.items():
                category_total = sum(float(item.amount or 0) for item in items_list)
                category_totals[category] = category_total
            
            groups_order = [
                ("preport", "📌 港前", grouped.get("preport", [])),
                ("warehouse_public", "🏬 公仓库内", grouped.get("warehouse_public", [])),
                ("warehouse_other", "🏭 私仓库内", grouped.get("warehouse_other", [])),
                ("delivery_public", "🚚 公仓派送", grouped.get("delivery_public", [])),
                ("delivery_other", "🚚 私仓派送", grouped.get("delivery_other", [])),
                ("activation_fee", "⚡ PO激活费", grouped.get("activation_fee", [])),
            ]
            
            context = {
                "invoice_number": invoice.invoice_number,
                "invoice": invoice,  # 添加invoice对象，用于获取状态等信息
                "container_number": container_number,
                "groups_order": groups_order,
                "category_totals": category_totals,
                "total_amount": total_amount,
                "start_date": request.GET.get("start_date"),
                "end_date": request.GET.get("end_date"),
                "is_combina": is_combina,
            }
            return self.template_confirm_transfer_edit, context

    def _all_invoice_items_get(self, invoice: Invoicev2, container_number):
        try:
            # 2. 获取所有InvoiceItemv2记录
            items = list(
                InvoiceItemv2.objects.filter(
                    invoice_number=invoice,
                    container_number__container_number=container_number,
                    invoice_type="receivable"  # 只显示应收
                ).order_by(
                    Case(
                        When(item_category="preport", then=Value(1)),
                        When(item_category="warehouse_public", then=Value(2)),
                        When(item_category="warehouse_other", then=Value(3)),
                        When(item_category="delivery_public", then=Value(4)),
                        When(item_category="delivery_other", then=Value(5)),
                        When(item_category="combina_extra_fee", then=Value(6)),
                        default=Value(7),
                    ),
                    "item_category",
                    "-amount"  # 金额从大到小
                )
            )
            
            # 3. 按item_category分组统计数据
            categories_data = {}
            for item in items:
                category = item.item_category
                if category not in categories_data:
                    categories_data[category] = {
                        'count': 0,
                        'total_amount': 0,
                        'items': []
                    }
                categories_data[category]['count'] += 1
                categories_data[category]['total_amount'] += item.amount or 0
                categories_data[category]['items'].append(item)
            
            # 4. 分离组合柜数据和其他数据
            combina_items = []  # 组合柜数据
            other_items = []    # 其他数据
            combina_extra_fee = 0.0
            # 处理所有类别
            for category, data in categories_data.items():
                if category == 'delivery_public':
                    # 处理delivery_public类别
                    delivery_public_items = data['items']
                    
                    # 找出所有组合柜的记录
                    combina_delivery_items = [item for item in delivery_public_items 
                                            if item.delivery_type == 'combine']
                    
                    # 按region和rate分组组合柜记录
                    combina_groups = {}
                    for item in combina_delivery_items:
                        key = (item.region, item.rate, item.regionPrice)
                        if key not in combina_groups:
                            combina_groups[key] = {
                                'region': item.region,
                                'rate': item.rate,
                                'regionPrice': item.regionPrice,
                                'total_cbm': 0,
                                'total_amount': 0,
                                'items': [],
                                'rowspan': 0
                            }
                        
                        combina_groups[key]['total_cbm'] += item.cbm or 0
                        combina_groups[key]['total_amount'] += item.amount or 0
                        combina_groups[key]['items'].append(item)
                        combina_groups[key]['rowspan'] += 1
                    

                    # 处理组合柜数据
                    for group_key, group_data in combina_groups.items():
                        group_items = []
                        for idx, item in enumerate(group_data['items']):
                            # 为每个组合柜项目添加分组信息
                            item.combina_group_id = f"combina_{group_key}"
                            item.combina_is_first = (idx == 0)
                            item.combina_total_cbm = group_data['total_cbm']
                            item.combina_total_amount = group_data['total_amount']
                            item.combina_rowspan = group_data['rowspan']
                            group_items.append(item)
                        
                        combina_items.extend(group_items)
                    
                    # 非组合柜的delivery_public记录添加到其他数据
                    non_combina_items = [item for item in delivery_public_items 
                                        if item.delivery_type != 'combine']
                    other_items.extend(non_combina_items)
                else:
                    # 其他类别的所有项目都添加到其他数据
                    other_items.extend(data['items'])
                    if data['items'].item_category == "combina_extra_fee":
                        combina_extra_fee += item.amount or 0
            
            # 5. 分别计算统计数据
            # 组合柜统计
            combina_stats = {
                'total_items': len(combina_items),
                'total_cbm': sum(item.cbm or 0 for item in combina_items),
                'total_amount': sum(item.amount or 0 for item in combina_items),
                'unique_groups': len(set(item.combina_group_id for item in combina_items if hasattr(item, 'combina_group_id')))
            }
            
            # 其他项目统计
            other_stats = {
                'total_items': len(other_items),
                'total_amount': sum(item.amount or 0 for item in other_items),
                'category_counts': {
                    category: data['count'] 
                    for category, data in categories_data.items()
                }
            }
            
            # 总体统计
            total_stats = {
                'total_items': len(items),
                'total_amount': sum(item.amount or 0 for item in items),
            }
            
            # 6. 获取item_category的中文显示名称
            category_display_names = {
                'preport': '港前费用',
                'warehouse_public': '公仓库内费用',
                'warehouse_other': '私仓库内费用',
                'delivery_public': '公仓派送费用',
                'delivery_other': '私仓派送费用',
                'combina_extra': '组合柜额外费用',
            }
            
            # 7. 获取delivery_type的中文显示
            delivery_type_display = {
                'combine': '组合柜',
                'amazon': '亚马逊',
                'upsdelivery': 'UPS',
                'walmart': '沃尔玛',
                'hold': '暂扣',
                'other': '其他',
            }
            customer =  invoice.customer
            if not customer:
                order = Order.objects.get(container_number__container_number=container_number)
                customer = order.customer_name
                invoice.customer = customer
                invoice.save()
            context = {
                'invoice_number': invoice.invoice_number,
                'container_number': container_number,
                'customer_name': invoice.customer.zem_name,
                'invoice_date': invoice.invoice_date,
                
                # 分离的数据
                'combina_items': combina_items,  # 组合柜数据
                'other_items': other_items,      # 其他数据
                
                # 统计数据
                'categories_data': categories_data,
                'category_display_names': category_display_names,
                'delivery_type_display': delivery_type_display,
                'combina_extra_fee': combina_extra_fee,
                # 三种统计数据
                'combina_stats': combina_stats,
                'other_stats': other_stats,
                'total_stats': total_stats,
                
                'success': True,
            }
            
            return context
            
        except Invoicev2.DoesNotExist:
            context = {
                'error_messages': f'找不到发票号: {invoice.invoice_number if hasattr(invoice, "invoice_number") else "未知"}',
                'success': False,
            }
            return context
        
    def handle_convert_type_post(self, request: HttpRequest):
        context = {}
        container_number = request.POST.get("container_number")
        to_delivery_type = request.POST.get("to_delivery_type")
        template_delivery_type = request.POST.get("template_delivery_type")
        item_id = request.POST.get("item_id")
        po_id = request.POST.get("po_id")
        qs = Pallet.objects.filter(
            container_number__container_number=container_number,
            PO_ID=po_id
        )
        
        updated = qs.update(delivery_type=to_delivery_type)
        unique_destinations = list(qs.values_list('destination', flat=True).distinct())
        destinations = ', '.join(unique_destinations)
        operation_messages = []

        operation_messages.append(f"{destinations}转类型成功！共更新 {updated} 个板子派送方式为{to_delivery_type}")
        if item_id and item_id != 'None':
            try:
                # 删除指定的记录
                InvoiceItemv2.objects.get(id=item_id).delete()
                operation_messages.append(f"成功删除原账单记录")
            except InvoiceItemv2.DoesNotExist:
                operation_messages.append(f"原账单记录不存在")
            except Exception as e:
                operation_messages.append(f"删除原账单记录出错，id: {item_id}, 错误: {e}")
        

        # 构造新的 GET 查询参数
        get_params = QueryDict(mutable=True)
        get_params["container_number"] = container_number
        get_params["delivery_type"] = template_delivery_type
        get_params["invoice_id"] = request.POST.get("invoice_id")
        
        request.GET = get_params
        context = {"success_messages":operation_messages}
        return self.handle_container_delivery_post(request,context)
    
    def handle_save_activation_fees(self, request: HttpRequest):
        """处理保存所有激活费操作"""
        context = {}
        activation_fee_data_str = request.POST.get('activation_fee_data', '[]')
        try:
            activation_fee_items = json.loads(activation_fee_data_str)
        except json.JSONDecodeError:
            activation_fee_items = []
        container_number = request.POST.get("container_number")
        invoice_id = request.POST.get("invoice_id")
        current_user = request.user
        username = current_user.username 

        try:
            container = Container.objects.get(container_number=container_number)
            invoice = Invoicev2.objects.get(id=invoice_id)
        except Container.DoesNotExist:
            context.update({"error_messages": f"柜号 {container_number} 不存在"})
            return self.handle_delivery_entry_post(request, context)
        except Invoicev2.DoesNotExist:
            context.update({"error_messages": f"账单ID {invoice_id} 不存在"})
            return self.handle_delivery_entry_post(request, context)
        
        item_category = "activation_fee"
        items_data = []
        for item_data in activation_fee_items:
            item_data = {
                "item_id": item_data.get("item_id", ""),
                "container": container,  # container对象
                "invoice": invoice,      # invoice对象
                "po_id": item_data.get("po_id", ""),
                "description": "PO激活费",  # 固定描述
                "destination": item_data.get("destination", ""),
                "delivery_category": "activation",  # 空字符串，因为是激活费
                "rate": item_data.get("amount", 0),  # 激活费的rate等于amount
                "qty": item_data.get("pallet", 0),  # 激活费没有板数，设为0
                "surcharges": 0,  # 激活费没有附加费，设为0
                "amount": item_data.get("amount", 0),
                "note": item_data.get("note", ""),
                "cbm": item_data.get("cbm", 0),
                "cbm_ratio": '',  # 激活费固定为1
                "weight": item_data.get("weight", 0),
                "registered_user": username,  # 当前用户
                "delivery_type": "public",  # 固定为公仓
                "invoice_type": "receivable",  # 应收账单
                "item_category": item_category,  # 专门分类
            }
            items_data.append(item_data)
        
        context = self.batch_save_delivery_item(container, invoice, items_data, item_category, context, username)
        # 构造新的 GET 查询参数
        get_params = QueryDict(mutable=True)
        get_params["container_number"] = container_number
        get_params["delivery_type"] = "public"
        get_params["invoice_id"] = invoice_id

        request.GET = get_params
        return self.handle_delivery_entry_post(request)
    
    def handle_save_all_combina_post(self, request: HttpRequest):
        """处理保存所有组合柜操作"""
        context = {}
        container_number = request.POST.get("container_number")
        invoice_id = request.POST.get("invoice_id")
        current_user = request.user
        username = current_user.username 

        item_category = "delivery_public"
        combina_items_json = request.POST.get("combina_items")
        combina_items = json.loads(combina_items_json)
        if not combina_items:
            context.update({"error_messages": "没有接收到数据"})
            return self.handle_delivery_entry_post(request, context)
        
        try:
            container = Container.objects.get(container_number=container_number)
            invoice = Invoicev2.objects.get(id=invoice_id)
        except Container.DoesNotExist:
            context.update({"error_messages": f"柜号 {container_number} 不存在"})
            return self.handle_delivery_entry_post(request, context)
        except Invoicev2.DoesNotExist:
            context.update({"error_messages": f"账单ID {invoice_id} 不存在"})
            return self.handle_delivery_entry_post(request, context)
        
        context = self.batch_save_delivery_item(container, invoice, combina_items, item_category, context, username)

        #计算派送总费用
        self._calculate_delivery_total_amount("public",invoice,container_number)
        # 构造新的 GET 查询参数
        get_params = QueryDict(mutable=True)
        get_params["container_number"] = container_number
        get_params["delivery_type"] = "public"
        get_params["invoice_id"] = invoice_id

        request.GET = get_params
        return self.handle_delivery_entry_post(request,context)
    
    def handle_save_all_post(self, request: HttpRequest):
        """处理保存所有账单记录的操作"""
        context = {}
        container_number = request.POST.getlist("container_number")[0]
        invoice_id = request.POST.getlist("invoice_id")[0]

        current_user = request.user
        username = current_user.username 

        delivery_type = request.POST.get("delivery_type")
        if delivery_type == "other":
            item_category = "delivery_other"
        else:
            item_category = "delivery_public"
       
        items_data_json = request.POST.get("items_data")
        if not items_data_json:
            context.update({"error_messages": "没有接收到数据"})
            return self.handle_delivery_entry_post(request, context)
        
        try:
            container = Container.objects.get(container_number=container_number)
            invoice = Invoicev2.objects.get(id=invoice_id)
        except Container.DoesNotExist:
            context.update({"error_messages": f"柜号 {container_number} 不存在"})
            return self.handle_delivery_entry_post(request, context)
        except Invoicev2.DoesNotExist:
            context.update({"error_messages": f"账单ID {invoice_id} 不存在"})
            return self.handle_delivery_entry_post(request, context)
        
        # 处理派送费
        try:
            items_data = json.loads(items_data_json)
        except json.JSONDecodeError as e:
            context.update({"error_messages": f"数据格式错误: {str(e)}"})
            return self.handle_delivery_entry_post(request, context)
    
        context = self.batch_save_delivery_item(container, invoice, items_data, item_category, context, username)

        # 处理激活费
        activation_fee_data_str = request.POST.get('activation_fee_data', '[]')
        try:
            activation_fee_items = json.loads(activation_fee_data_str)
        except json.JSONDecodeError:
            activation_fee_items = []

        item_category = "activation_fee"
        items_data = []
        for item_data in activation_fee_items:
            item_data = {
                "item_id": item_data.get("item_id", ""),
                "container": container,  # container对象
                "invoice": invoice,      # invoice对象
                "po_id": item_data.get("po_id", ""),
                "description": "PO激活费",  # 固定描述
                "destination": item_data.get("destination", ""),
                "delivery_category": "activation",  # 空字符串，因为是激活费
                "rate": item_data.get("amount", 0),  # 激活费的rate等于amount
                "qty": item_data.get("pallet", 0),  # 激活费没有板数，设为0
                "surcharges": 0,  # 激活费没有附加费，设为0
                "amount": item_data.get("amount", 0),
                "note": item_data.get("note", ""),
                "cbm": item_data.get("cbm", 0),
                "cbm_ratio": '',  # 激活费固定为1
                "weight": item_data.get("weight", 0),
                "registered_user": username,  # 当前用户
                "delivery_type": "public",  # 固定为公仓
                "invoice_type": "receivable",  # 应收账单
                "item_category": item_category,  # 专门分类
            }
            items_data.append(item_data)
        
        context = self.batch_save_delivery_item(container, invoice, items_data, item_category, context, username)

        container_delivery_type = getattr(container, 'delivery_type', 'mixed')

        status_obj = InvoiceStatusv2.objects.get(
                invoice=invoice,
                invoice_type='receivable'
            )
        
        # 根据柜子类型自动更新另一边的状态
        if delivery_type == "public":
            status_obj.delivery_public_status = "completed"
            if container_delivery_type == "public":
                status_obj.delivery_other_status = "completed"
        else:
            status_obj.delivery_other_status = "completed"
            if container_delivery_type == "other":
                status_obj.delivery_public_status = "completed"
        status_obj.save()
        #计算派送总费用
        self._calculate_delivery_total_amount(delivery_type,invoice,container_number)

        return self.handle_delivery_entry_post(request, context)
    
    def _calculate_delivery_total_amount(self, delivery_type: str, invoice: Invoicev2, container_number: str):
        if delivery_type == "public":
            items = InvoiceItemv2.objects.filter(
                invoice_number=invoice,
                container_number__container_number=container_number,
                invoice_type="receivable",
                item_category="delivery_public"
            )
            total_amount = items.aggregate(
                total=Sum('amount')
            )['total'] or 0
            invoice.receivable_delivery_public_amount = total_amount
            self._calculate_invoice_total_amount(invoice)
        else:
            items = InvoiceItemv2.objects.filter(
                invoice_number=invoice,
                container_number__container_number=container_number,
                invoice_type="receivable",
                item_category="delivery_other"
            )
            total_amount = items.aggregate(
                total=Sum('amount')
            )['total'] or 0
            invoice.receivable_delivery_other_amount = total_amount
            self._calculate_invoice_total_amount(invoice)


    def _search_region(self, items_data, container:Container):
        # 查找报价表
        order = Order.objects.select_related("container_number").get(
            container_number=container
        )
        container_type_temp = 0 if "40" in container.container_type else 1

        quotations = self._get_fee_details(order, order.retrieval_id.retrieval_destination_area,order.customer_name.zem_name)
        if isinstance(quotations, dict) and quotations.get("error_messages"):
            raise ValueError(quotations["error_messages"])
        
        fee_details = quotations.get('fees', {})
        warehouse = order.retrieval_id.retrieval_destination_area
        combina_key = f"{warehouse}_COMBINA"
        combina_fee = fee_details.get(combina_key, {})
        if not combina_fee:
            return items_data
        rules = fee_details.get(combina_key).details
        
        # 检查是否属于组合区域
        for item_data in items_data:
            destination_str = item_data.get("destination")
            if not destination_str:
                continue
            destination_origin, destination = self._process_destination(destination_str)
            is_combina_region = False

            for region, region_data in rules.items():
                for item in region_data:
                    locations = item.get("location", [])
                    if destination in locations:
                        prices = item.get("prices", [])
                        item_data["combina_region"] = region
                        item_data["rate"] = prices[container_type_temp]
                        is_combina_region = True
                        break
                if is_combina_region:
                    break
            
        return items_data

    def batch_save_delivery_item(self, container, invoice, items_data, item_category, context, username: str| None=None):
        if not context:
            context = {}
        success_count = 0
        error_messages = []

        if item_category != "activation_fee":
            # 检查一遍是否都有仓点和价格
            need_search = False
            for item_data in items_data:
                region = item_data.get("combina_region", "")
                rate = item_data.get("rate")
                delivery_category = item_data.get("delivery_category")
                if delivery_category == "combine":
                    if not region or not rate:
                        need_search = True
                        break  # 只要发现有一条缺少，就跳出循环
            
            # 如果需要查找，一次性处理所有数据
            if need_search:
                items_data = self._search_region(items_data, container)

        # 遍历每条数据
        for item_data in items_data:
            
            po_id = item_data.get("po_id", "")
            if not po_id:
                raise ValueError('缺少PO_ID')
            row_index = item_data.get("rowIndex")

            # 提取数据
            delivery_category = item_data.get("delivery_category", "")
            if not delivery_category:
                error_messages.append(f"第{row_index + 1}行: 派送类型不能为空")
                continue
            
            item_id = item_data.get("item_id")
            
            destination = item_data.get("destination", "")
            
            rate = item_data.get("rate")
            pallets = item_data.get("pallets")
            surcharges = item_data.get("surcharges")
            amount = item_data.get("amount")
            description = item_data.get("description", "")
            if not description:
                description = "派送费"
            region = item_data.get("combina_region", "")
            cbm = item_data.get("cbm", "")
            cbm_ratio = item_data.get("cbmRatio", 0)
            weight = item_data.get("weight", "")
            note = item_data.get("note", "")
            shipping_marks = item_data.get("shipping_marks", "")

            registered_user = item_data.get("registered_user") or username
            if delivery_category == "hold":
                note = f"暂扣, {note}"
            elif delivery_category == "combine":
                note = region
            if not po_id:
                error_messages.append(f"第{row_index + 1}行: PO号不能为空")
                continue                 
        
            # 转换数据类型
            def to_float(val):
                if val is None or val == "":
                    return None
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return None
            
            rate_float = to_float(rate)
            pallets_float = to_float(pallets)
            surcharges_float = to_float(surcharges)
            amount_float = to_float(amount)
            
            # 更新或创建记录
            if item_id:
                # 更新现有记录
                try:
                    item = InvoiceItemv2.objects.get(id=item_id)
                except InvoiceItemv2.DoesNotExist:
                    error_messages.append(f"第{row_index + 1}行: 未查询到ID为 {item_id} 的记录")
                    continue
            else:
                # 新建记录
                item = InvoiceItemv2(
                    container_number=container,
                    invoice_number=invoice,
                    invoice_type="receivable",
                    item_category=item_category,
                    PO_ID=po_id,
                )
            
            # 更新字段
            item.delivery_type = delivery_category
            item.invoice_number = invoice
            item.container_number = container
            item.PO_ID = po_id
            item.rate = rate_float
            item.qty = pallets_float
            item.surcharges = surcharges_float
            item.amount = amount_float
            item.description = description
            item.warehouse_code = destination
            item.region = region
            item.regionPrice = rate_float
            item.cbm = cbm
            item.weight = weight
            item.cbm_ratio = cbm_ratio
            item.registered_user = registered_user
            item.shipping_marks = shipping_marks
            item.note = note
            
            # 保存
            item.save()
            success_count += 1
        
        # 准备返回消息
        success_messages = []
        if success_count > 0:
            success_messages.append(f"{container.container_number}成功保存 {success_count} 条记录")
        
        # 更新上下文
        if success_messages:
            context.update({"success_messages": success_messages})
        if error_messages:
            context.update({"error_messages": error_messages})
        return context

    def handle_save_single_post(self, request: HttpRequest):
        """处理单条派送账单保存操作"""
        context = {}
        container_number = request.POST.get("container_number")
        invoice_id = request.POST.get("invoice_id")

        current_user = request.user
        username = current_user.username 

        delivery_type = request.POST.get("delivery_type")
        if delivery_type == "other":
            item_category = "delivery_other"
        else:
            item_category = "delivery_public"

        container = Container.objects.get(container_number=container_number)
        invoice = Invoicev2.objects.get(id=invoice_id)

        item_data = [{
            "item_id": request.POST.get("item_id"),
            "container": container,
            "invoice": invoice,
            "po_id": request.POST.get("po_id"),
            "description": "派送费",
            "destination": request.POST.get("destination"),
            "delivery_category": request.POST.get("delivery_category"),
            "rate": request.POST.get("rate"),
            "pallets": request.POST.get("pallets"),
            "surcharges": request.POST.get("surcharges"),
            "amount": request.POST.get("amount"),
            "note": request.POST.get("note"),
            "cbm": request.POST.get("cbm"),
            "cbm_ratio": 1,
            "weight": request.POST.get("weight"),
            "registered_user": username,
            "shipping_marks": request.POST.get("shipping_marks"),
        }]
        
        context = self.batch_save_delivery_item(container, invoice, item_data, item_category, context)

        #计算派送总费用
        self._calculate_delivery_total_amount(delivery_type,invoice,container_number)

        # 构造新的 GET 查询参数
        get_params = QueryDict(mutable=True)
        get_params["container_number"] = container_number
        get_params["delivery_type"] = delivery_type
        get_params["invoice_id"] = invoice_id

        request.GET = get_params
        return self.handle_container_delivery_post(request,context)
    
    def handle_release_hold_post(self, request: HttpRequest):
        """处理解扣操作"""
        po_id = request.POST.get("po_id")
        container_number = request.POST.get("container_number")
        delivery_method = request.POST.get("delivery_method")
        # 更新托盘状态，移除暂扣标记
        qs = Pallet.objects.filter(
            container_number__container_number=container_number,
            PO_ID=po_id
        )

        updated = qs.update(delivery_method=delivery_method)

        delivery_type = request.POST.get("delivery_type")
        invoice_id = request.POST.get("invoice_id")

        # 构造新的 GET 查询参数
        get_params = QueryDict(mutable=True)
        get_params["container_number"] = container_number
        get_params["delivery_type"] = delivery_type
        get_params["invoice_id"] = invoice_id

        request.GET = get_params
        success_messages = f"解扣成功！共更新 {updated} 个板子派送方式为{delivery_method}"
        context = {"success_messages":success_messages}
        return self.handle_container_delivery_post(request,context)

    def old_handle_preport_entry_post(self, request:HttpRequest, context: dict| None = None,) -> Dict[str, Any]:
        '''港前账单查询'''
        warehouse = request.POST.get("warehouse_filter")
        customer = request.POST.get("customer")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date
        #按提柜时间改
        criteria = (
            Q(cancel_notification=False)
            & Q(retrieval_id__actual_retrieval_timestamp__gte=start_date)
            & Q(retrieval_id__actual_retrieval_timestamp__lte=end_date)
            & Q(offload_id__offload_at__isnull=False)
        )
        
        if warehouse and warehouse != 'None':
            criteria &= Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= Q(customer_name__zem_name=customer)

        # 获取基础订单数据
        base_orders = (
            Order.objects
            .select_related(
                'retrieval_id', 
                'offload_id', 
                'container_number',
                'customer_name'
            )
            .annotate(
                retrieval_time=F("retrieval_id__actual_retrieval_timestamp"),
                empty_returned_time=F("retrieval_id__empty_returned_at"),
                offload_time=F("offload_id__offload_at"),
            )
            .filter(criteria)
            .order_by("-retrieval_time")
            .distinct()
        )
        preport_to_record_orders = [] #待录入
        preport_recorded_orders = []  #已录入
        preport_pending_review_orders = []  #待审核
        preport_completed_orders = []  #已审核

        for order in base_orders:
            container = order.container_number
            
            if not container:
                continue
                
            # 查询这个柜子的所有应收账单
            invoices = Invoicev2.objects.filter(container_number=container)
            
            if not invoices.exists():
                # 没有账单的情况 - 归到待录入
                order_data = {
                    'order': order,
                    'invoice_number': None,
                    'invoice_id': None,
                    'invoice_created_at': None,
                    'preport_status': None,
                    'finance_status': None,
                    'has_invoice': False
                }
                preport_to_record_orders.append(order_data)
            else:
                has_multiple_invoices = invoices.count() > 1 #看看是不是补开的账单
                # 有账单的情况 - 每个账单都要单独处理
                for invoice in invoices:
                    # 查询这个账单对应的状态
                    try:
                        status_obj = InvoiceStatusv2.objects.get(
                            invoice=invoice,
                            invoice_type='receivable'
                        )
                        preport_status = status_obj.preport_status
                        finance_status = status_obj.finance_status
                    except InvoiceStatusv2.DoesNotExist:
                        preport_status = None
                        finance_status = None
                    
                    # 只在有多个账单时添加 invoice_created_at
                    invoice_created_at = None
                    if has_multiple_invoices:
                        invoice_created_at = (
                            invoice.created_at if hasattr(invoice, 'created_at') 
                            else invoice.history.first().history_date if invoice.history.exists() 
                            else None
                        )

                    order_data = {
                        'order': order,
                        'invoice_number': invoice.invoice_number,
                        'invoice_id': invoice.id,
                        'invoice_created_at': invoice_created_at,
                        'preport_status': preport_status,
                        'finance_status': finance_status,
                        'has_invoice': True
                    }
                    
                    # 根据状态分组
                    if preport_status in ["unstarted", "in_progress"] or preport_status is None:
                        preport_to_record_orders.append(order_data)
                    elif preport_status == "pending_review":
                        preport_pending_review_orders.append(order_data)
                        preport_recorded_orders.append(order_data)
                    elif preport_status == "completed":
                        preport_completed_orders.append(order_data)
                        preport_recorded_orders.append(order_data)
                    elif preport_status == "rejected":
                        preport_recorded_orders.append(order_data)

        # 对已录入的订单按状态排序（rejected置顶）
        preport_recorded_orders.sort(key=lambda x: {
            "rejected": 0,
            "pending_review": 1, 
            "completed": 2
        }.get(x['preport_status'], 3))

        # 判断用户权限，决定默认标签页
        groups = [group.name for group in request.user.groups.all()]
        if not context:
            context = {}
        # 如果用户有 invoice_preport_leader 权限，默认打开审核标签页，否则打开录入标签页
        is_leader = False
        if 'invoice_preport_leader' in groups:
            is_leader = True
            default_tab = 'review'
        else:
            default_tab = 'entry'
        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'preport_to_record_orders': preport_to_record_orders,
            'preport_recorded_orders': preport_recorded_orders,
            'preport_pending_review_orders': preport_pending_review_orders,
            'preport_completed_orders': preport_completed_orders,
            "order_form": OrderForm(),
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
            "default_tab": default_tab, 
            "is_leader": is_leader,
        })
        return context
    
    def handle_preport_entry_post(self, request:HttpRequest, context: dict| None = None,) -> Dict[str, Any]:
        '''港前账单查询'''
        warehouse = request.POST.get("warehouse_filter")
        customer = request.POST.get("customer")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        # --- 1. 日期处理 ---
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date

        # --- 2. 构建查询条件 (按提柜时间) ---
        criteria = (
            Q(cancel_notification=False)
            & Q(retrieval_id__actual_retrieval_timestamp__gte=start_date)
            & Q(retrieval_id__actual_retrieval_timestamp__lte=end_date)
            & Q(offload_id__offload_at__isnull=False)
        )
        
        if warehouse and warehouse != 'None':
            criteria &= Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= Q(customer_name__zem_name=customer)

        # --- 3. 获取基础订单数据 ---
        base_orders = (
            Order.objects
            .select_related(
                'retrieval_id', 
                'offload_id', 
                'container_number',
                'customer_name'
            )
            .annotate(
                retrieval_time=F("retrieval_id__actual_retrieval_timestamp"),
                empty_returned_time=F("retrieval_id__empty_returned_at"),
                offload_time=F("offload_id__offload_at"),
            )
            .filter(criteria)
            .order_by("-retrieval_time")
            .distinct()
        )

        orders_list = list(base_orders)
        # 提取 Container IDs
        container_ids = set()
        for order in orders_list:
            if order.container_number_id:
                container_ids.add(order.container_number_id)
        container_ids = list(container_ids)

        # --- 4. 批量获取 Invoice 和 InvoiceStatus ---
        status_prefetch = Prefetch(
            'invoicestatusv2_set',
            queryset=InvoiceStatusv2.objects.filter(invoice_type="receivable"),
            to_attr='receivable_status_list'
        )

        all_invoices = Invoicev2.objects.filter(
            container_number_id__in=container_ids
        ).prefetch_related(status_prefetch)

        # --- 5. [安全机制] 批量创建缺失的 InvoiceStatus ---
        missing_statuses = []
        invoices_needing_update = []

        for inv in all_invoices:
            if not (hasattr(inv, 'receivable_status_list') and inv.receivable_status_list):
                new_status = InvoiceStatusv2(
                    invoice=inv,
                    container_number_id=inv.container_number_id,
                    invoice_type="receivable",
                    # 默认初始化所有状态
                    preport_status="unstarted",
                    finance_status="unstarted",
                    warehouse_public_status="unstarted",
                    warehouse_other_status="unstarted",
                    delivery_public_status="unstarted",
                    delivery_other_status="unstarted",
                )
                missing_statuses.append(new_status)
                invoices_needing_update.append(inv)

        if missing_statuses:
            InvoiceStatusv2.objects.bulk_create(missing_statuses)
            for i, inv in enumerate(invoices_needing_update):
                inv.receivable_status_list = [missing_statuses[i]]

        # --- 6. 内存分组 ---
        container_invoice_map = defaultdict(list)
        for inv in all_invoices:
            container_invoice_map[inv.container_number_id].append(inv)

        # --- 7. 主循环处理 ---
        preport_to_record_orders = [] #待录入
        preport_recorded_orders = []  #已录入 (包含审核中、已完成、已驳回)
        preport_pending_review_orders = []  #待审核
        preport_completed_orders = []  #已审核

        for order in base_orders:
            container = order.container_number
            
            if not container:
                continue
            
            c_id = container.id
            container_invoices = container_invoice_map.get(c_id, [])
            # 定义构建函数
            def build_order_data(inv=None, status_obj=None):
                created_at = None
                if inv and len(container_invoices) > 1:
                    created_at = inv.created_at
                
                return {
                    'order': order,
                    'invoice_number': inv.invoice_number if inv else None,
                    'invoice_id': inv.id if inv else None,
                    'invoice_created_at': created_at,
                    'preport_status': status_obj.preport_status if status_obj else None,
                    'finance_status': status_obj.finance_status if status_obj else None,
                    'has_invoice': bool(inv)
                }

            if not container_invoices:
                # === 场景 A: 无账单 -> 归期待录入 ===
                base_data = build_order_data(None, None)
                preport_to_record_orders.append(base_data)
            else:
                # === 场景 B: 有账单 ===
                for invoice in container_invoices:
                    status_obj = None
                    if hasattr(invoice, 'receivable_status_list') and invoice.receivable_status_list:
                        status_obj = invoice.receivable_status_list[0]
                    
                    base_data = build_order_data(invoice, status_obj)
                    
                    preport_status = base_data['preport_status']
                    
                    # 根据状态分组 (逻辑完全参考原代码)
                    if preport_status in ["unstarted", "in_progress", None]:
                        preport_to_record_orders.append(base_data)
                    
                    elif preport_status == "pending_review":
                        preport_pending_review_orders.append(base_data)
                        preport_recorded_orders.append(base_data)
                    
                    elif preport_status == "completed":
                        preport_completed_orders.append(base_data)
                        preport_recorded_orders.append(base_data)
                    
                    elif preport_status == "rejected":
                        preport_recorded_orders.append(base_data)

        # --- 8. 排序逻辑 ---
        preport_recorded_orders.sort(key=lambda x: {
            "rejected": 0,
            "pending_review": 1, 
            "completed": 2
        }.get(x['preport_status'], 3))

        # --- 9. 权限与上下文 ---
        groups = [group.name for group in request.user.groups.all()] if request.user else []
        if not context:
            context = {}
        # 如果用户有 invoice_preport_leader 权限，默认打开审核标签页，否则打开录入标签页
        is_leader = False
        if 'invoice_preport_leader' in groups:
            is_leader = True
            default_tab = 'review'
        else:
            default_tab = 'entry'

        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'preport_to_record_orders': preport_to_record_orders,
            'preport_recorded_orders': preport_recorded_orders,
            'preport_pending_review_orders': preport_pending_review_orders,
            'preport_completed_orders': preport_completed_orders,
            "order_form": OrderForm(),
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
            "default_tab": default_tab, 
            "is_leader": is_leader,
        })
        return context

    def old_handle_warehouse_entry_post(self, request:HttpRequest, context: dict| None = None,) -> Dict[str, Any]:
        warehouse = request.POST.get("warehouse_filter")
        customer = request.POST.get("customer")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date

        criteria = (
            Q(cancel_notification=False)
            & (Q(order_type="转运") | Q(order_type="转运组合"))
            & Q(vessel_id__vessel_etd__gte=start_date)
            & Q(vessel_id__vessel_etd__lte=end_date)
            & Q(offload_id__offload_at__isnull=False)
        )

        if warehouse and warehouse != 'None':
            if "LA" in warehouse:
                criteria &= Q(retrieval_id__retrieval_destination_precise__contains='LA')
            else:
                criteria &= Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= Q(customer_name__zem_name=customer)

        # 获取基础订单数据
        base_orders = (
            Order.objects
            .select_related(
                'retrieval_id', 
                'offload_id', 
                'container_number',
                'customer_name'
            )
            .annotate(
                retrieval_time=F("retrieval_id__actual_retrieval_timestamp"),
                empty_returned_time=F("retrieval_id__empty_returned_at"),
                offload_time=F("offload_id__offload_at"),
            )
            .filter(criteria)
            .distinct()
        )
        wh_public_to_record_orders = [] #公仓待录入
        wh_public_recorded_orders = []  #公仓已录入
        wh_self_to_record_orders = []  #私仓待录入
        wh_self_recorded_orders = []  #私仓已录入
        
        for order in base_orders:
            container = order.container_number
            
            if not container:
                continue
            
            container_delivery_type = getattr(container, 'delivery_type', 'mixed')
        
            # 判断是否应该处理公仓或私仓
            should_process_public = container_delivery_type in ['public', 'mixed']
            should_process_self = container_delivery_type in ['other', 'mixed']
            
            # 如果没有柜子类型信息，默认都处理
            if container_delivery_type not in ['public', 'other', 'mixed']:
                should_process_public = True
                should_process_self = True

            # 查询这个柜子的所有应收账单
            invoices = Invoicev2.objects.filter(container_number=container)
            if not invoices.exists():
                # 没有账单的情况 - 归到待录入
                order_data = {
                    'order': order,
                    'container_number': order.container_number,
                    'invoice_number': None,
                    'invoice_id': None,
                    'invoice_created_at': None,
                    'preport_status': None,
                    'finance_status': None,
                    'has_invoice': False,
                    'offload_time': order.offload_time,
                }
                # 根据柜子类型决定添加到哪个列表
                if should_process_public:
                    wh_public_to_record_orders.append(order_data)
                if should_process_self:
                    wh_self_to_record_orders.append(order_data)
            else:
                has_multiple_invoices = invoices.count() > 1 #看看是不是补开的账单
                # 有账单的情况 - 每个账单都要单独处理
                for invoice in invoices:
                    # 查询这个账单对应的状态
                    try:
                        status_obj = InvoiceStatusv2.objects.get(
                            invoice=invoice,
                            invoice_type='receivable'
                        )
                        public_status = status_obj.warehouse_public_status #公仓状态
                        self_status = status_obj.warehouse_other_status  #私仓状态
                        finance_status = status_obj.finance_status #财务状态
                    except InvoiceStatusv2.DoesNotExist:
                        public_status = None
                        self_status = None
                        finance_status = None
                    
                    # 只在有多个账单时添加 invoice_created_at
                    invoice_created_at = None
                    if has_multiple_invoices:
                        invoice_created_at = (
                            invoice.created_at if hasattr(invoice, 'created_at') 
                            else invoice.history.first().history_date if invoice.history.exists() 
                            else None
                        )
                    order_data = {
                        'order': order,
                        'container_number': order.container_number,
                        'invoice_number': invoice.invoice_number,
                        'invoice_id': invoice.id,
                        'invoice_created_at': invoice_created_at,
                        'public_status': public_status,
                        'self_status': self_status,
                        'finance_status': finance_status,
                        'has_invoice': True,
                        'offload_time': order.offload_time,
                    }
                    
                    # 根据状态分组，同时考虑柜子类型
                    if should_process_public:
                        if public_status in ["unstarted", "in_progress"] or public_status is None:
                            wh_public_to_record_orders.append(order_data)
                        elif public_status in ["completed", "rejected", "confirmed"]:
                            wh_public_recorded_orders.append(order_data)

                    if should_process_self:
                        if self_status in ["unstarted", "in_progress"] or self_status is None:
                            wh_self_to_record_orders.append(order_data)
                        elif self_status in ["completed", "rejected", "confirmed"]:
                            wh_self_recorded_orders.append(order_data)
        
        #按照出库比例排序
        wh_self_to_record_orders = self._add_shipment_group_stats(wh_self_to_record_orders, "other")
        #已录入中，驳回的优先显示
        wh_self_recorded_orders.sort(key=lambda x: x.get('self_status') != 'rejected')

        #公仓的先按照入库时间排序
        # 在排序前处理时区
        for order_data in wh_public_to_record_orders:
            offload_time = order_data.get('offload_time')
            if offload_time and hasattr(offload_time, 'tzinfo') and offload_time.tzinfo:
                # 去除时区信息，保留原生时间
                order_data['offload_time'] = offload_time.replace(tzinfo=None)

        # 现在排序
        wh_public_to_record_orders.sort(key=lambda x: (
            x.get('offload_time') or datetime.max
        ))
        # 判断用户权限，决定默认标签页
        groups = [group.name for group in request.user.groups.all()]
        if not context:
            context = {}

        # 根据权限，决定打开的标签页
        if 'warehouse_other' in groups:
            default_tab = 'self'
        else:
            default_tab = 'public'

        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'wh_public_to_record_orders': wh_public_to_record_orders,
            'wh_public_recorded_orders': wh_public_recorded_orders,
            'wh_self_to_record_orders': wh_self_to_record_orders,
            'wh_self_recorded_orders': wh_self_recorded_orders,
            "order_form": OrderForm(),
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
            "default_tab": default_tab, 
        })
        return context
    
    def handle_warehouse_entry_post(self, request:HttpRequest, context: dict| None = None,) -> Dict[str, Any]:
        warehouse = request.POST.get("warehouse_filter")
        customer = request.POST.get("customer")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        # --- 1. 日期处理 ---
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date

        # --- 2. 构建查询条件 ---
        criteria = (
            Q(cancel_notification=False)
            & (Q(order_type="转运") | Q(order_type="转运组合"))
            & Q(vessel_id__vessel_etd__gte=start_date)
            & Q(vessel_id__vessel_etd__lte=end_date)
            & Q(offload_id__offload_at__isnull=False)
        )

        if warehouse and warehouse != 'None':
            if "LA" in warehouse:
                criteria &= Q(retrieval_id__retrieval_destination_precise__contains='LA')
            else:
                criteria &= Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= Q(customer_name__zem_name=customer)

        # --- 3. 获取基础订单数据 ---
        base_orders = (
            Order.objects
            .select_related(
                'retrieval_id', 
                'offload_id', 
                'container_number',
                'customer_name'
            )
            .annotate(
                retrieval_time=F("retrieval_id__actual_retrieval_timestamp"),
                empty_returned_time=F("retrieval_id__empty_returned_at"),
                offload_time=F("offload_id__offload_at"),
            )
            .filter(criteria)
            .distinct()
        )

        orders_list = list(base_orders)
        # 提取 Container IDs
        container_ids = set()
        for order in orders_list:
            if order.container_number_id:
                container_ids.add(order.container_number_id)
        container_ids = list(container_ids)

        # --- 4. 批量获取 Invoice 和 InvoiceStatus ---
        status_prefetch = Prefetch(
            'invoicestatusv2_set',
            queryset=InvoiceStatusv2.objects.filter(invoice_type="receivable"),
            to_attr='receivable_status_list'
        )

        all_invoices = Invoicev2.objects.filter(
            container_number_id__in=container_ids
        ).prefetch_related(status_prefetch)

         # --- 5. [安全机制] 批量创建缺失的 InvoiceStatus ---
        missing_statuses = []
        invoices_needing_update = []

        for inv in all_invoices:
            # 如果预查询列表为空，说明缺数据
            if not (hasattr(inv, 'receivable_status_list') and inv.receivable_status_list):
                new_status = InvoiceStatusv2(
                    invoice=inv,
                    container_number_id=inv.container_number_id, # 使用ID赋值更轻量
                    invoice_type="receivable",
                    # 默认状态
                    warehouse_public_status="unstarted",
                    warehouse_other_status="unstarted",
                    preport_status="unstarted",
                    delivery_public_status="unstarted",
                    delivery_other_status="unstarted",
                    finance_status="unstarted"
                )
                missing_statuses.append(new_status)
                invoices_needing_update.append(inv)

        if missing_statuses:
            InvoiceStatusv2.objects.bulk_create(missing_statuses)
            # 手动回填内存，避免重新查询
            for i, inv in enumerate(invoices_needing_update):
                inv.receivable_status_list = [missing_statuses[i]]
        
        # --- 6. 内存分组 & 统计预计算 ---
        container_invoice_map = defaultdict(list)
        for inv in all_invoices:
            container_invoice_map[inv.container_number_id].append(inv)

        # 调用复用的统计方法
        shipment_stats_map = self._bulk_calculate_shipment_stats(container_ids)

        # --- 7. 主循环 ---
        wh_public_to_record_orders = [] #公仓待录入
        wh_public_recorded_orders = []  #公仓已录入
        wh_self_to_record_orders = []  #私仓待录入
        wh_self_recorded_orders = []  #私仓已录入
        
        for order in orders_list:
            container = order.container_number
            if not container:
                continue
            
            c_id = container.id
            container_delivery_type = container.delivery_type if container.delivery_type else 'mixed'
        
            should_process_public = container_delivery_type in ['public', 'mixed']
            should_process_self = container_delivery_type in ['other', 'mixed']
            
            if container_delivery_type not in ['public', 'other', 'mixed']:
                should_process_public = True
                should_process_self = True

            container_invoices = container_invoice_map.get(c_id, [])

            # 定义构建函数
            def build_order_data(inv=None, status_obj=None):
                created_at = None
                # 只有多账单才去拿时间，且只拿 created_at，不碰 history 以免 N+1
                if inv and len(container_invoices) > 1:
                    created_at = inv.created_at 
                
                return {
                    'order': order,
                    'container_number': order.container_number,
                    'invoice_number': inv.invoice_number if inv else None,
                    'invoice_id': inv.id if inv else None,
                    'invoice_created_at': created_at,
                    # 注意：这里取的是 warehouse 相关的状态
                    'public_status': status_obj.warehouse_public_status if status_obj else None,
                    'self_status': status_obj.warehouse_other_status if status_obj else None,
                    'finance_status': status_obj.finance_status if status_obj else None,
                    'has_invoice': bool(inv),
                    'offload_time': order.offload_time,
                    # 放入统计数据供后续计算
                    'stats_raw': shipment_stats_map.get(c_id, {})
                }

            if not container_invoices:
                # === 场景 A: 无账单 ===
                base_data = build_order_data(None, None)
                
                if should_process_public:
                    wh_public_to_record_orders.append(base_data) # 公仓不需要统计比例，无需 inject
                if should_process_self:
                    item = base_data.copy()
                    self._inject_stats_and_ratio(item, 'other') # 私仓需要统计比例
                    wh_self_to_record_orders.append(item)
            else:
                # === 场景 B: 有账单 ===
                for invoice in container_invoices:
                    status_obj = None
                    if hasattr(invoice, 'receivable_status_list') and invoice.receivable_status_list:
                        status_obj = invoice.receivable_status_list[0]
                    
                    base_data = build_order_data(invoice, status_obj)
                    
                    # --- 公仓分组 ---
                    if should_process_public:
                        # 公仓这里为了保持逻辑独立，copy一份
                        p_item = base_data.copy()
                        p_status = p_item['public_status']
                        
                        if p_status in ["unstarted", "in_progress", None]:
                            wh_public_to_record_orders.append(p_item)
                        elif p_status in ["completed", "rejected", "confirmed"]:
                            wh_public_recorded_orders.append(p_item)

                    # --- 私仓分组 ---
                    if should_process_self:
                        s_item = base_data.copy()
                        # 注入统计数据 (因为私仓排序要用)
                        self._inject_stats_and_ratio(s_item, 'other')
                        s_status = s_item['self_status']
                        
                        if s_status in ["unstarted", "in_progress", None]:
                            wh_self_to_record_orders.append(s_item)
                        elif s_status in ["completed", "rejected", "confirmed"]:
                            wh_self_recorded_orders.append(s_item)

        # 私仓待录入：按出库比例排序
        wh_self_to_record_orders.sort(key=lambda x: x.get('completion_ratio', 0), reverse=True)
        # 私仓已录入：驳回优先
        wh_self_recorded_orders.sort(key=lambda x: x.get('self_status') == 'rejected', reverse=True)

        # 公仓待录入：按入库(offload)时间排序
        # 处理时区问题：如果有时区则去掉，None则排到最后
        def get_naive_offload_time(item):
            t = item.get('offload_time')
            if not t:
                return datetime.max # None 排最后
            # 如果是 timezone aware，转 naive
            if hasattr(t, 'tzinfo') and t.tzinfo:
                return t.replace(tzinfo=None)
            return t

        wh_public_to_record_orders.sort(key=get_naive_offload_time)

        # --- 9. 权限与上下文 ---
        groups = [group.name for group in request.user.groups.all()] if request.user else []
        if not context:
            context = {}

        # 根据权限，决定打开的标签页
        if 'warehouse_other' in groups:
            default_tab = 'self'
        else:
            default_tab = 'public'

        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'wh_public_to_record_orders': wh_public_to_record_orders,
            'wh_public_recorded_orders': wh_public_recorded_orders,
            'wh_self_to_record_orders': wh_self_to_record_orders,
            'wh_self_recorded_orders': wh_self_recorded_orders,
            "order_form": OrderForm(),
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
            "default_tab": default_tab, 
        })
        return context
    
    def handle_supplement_search(self, request:HttpRequest, context: dict| None = None,) -> Dict[str, Any]:
        if not context:
            context = {}
        warehouse = request.POST.get("warehouse_filter")
        customer = request.POST.get("customer")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date

        criteria = (
            Q(cancel_notification=False)
            & (Q(order_type="转运") | Q(order_type="转运组合"))
            & Q(vessel_id__vessel_etd__gte=start_date)
            & Q(vessel_id__vessel_etd__lte=end_date)
            & Q(offload_id__offload_at__isnull=False)
        )

        if warehouse:
            criteria &= Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= Q(customer_name__zem_name=customer)
    
        # 获取基础订单数据
        base_orders = (
            Order.objects
            .select_related(
                'retrieval_id', 
                'offload_id', 
                'container_number',
                'customer_name'
            )
            .annotate(
                retrieval_time=F("retrieval_id__actual_retrieval_timestamp"),
                empty_returned_time=F("retrieval_id__empty_returned_at"),
                offload_time=F("offload_id__offload_at"),
            )
            .filter(criteria)
            .distinct()
        )
        order = [] #可补开的账单
        previous_order = []  #已补开的账单
        for o in base_orders:
            container = o.container_number
            
            if not container:
                continue

            # 查询这个柜子的所有应收账单
            invoices = Invoicev2.objects.filter(container_number=container)
            
            if invoices.exists():
                continue

            for invoice in invoices:
                try:
                    invoice_status = InvoiceStatusv2.objects.get(
                        invoice=invoice,
                        invoice_type="receivable"
                    )
                    finance_status = invoice_status.finance_status
                except InvoiceStatusv2.DoesNotExist:
                    # 如果没有状态记录，跳过处理
                    continue
                
                # 如果只有一条发票记录
                if invoices.count() == 1:
                    # finance_status不是completed的不用处理
                    if finance_status != 'completed':
                        continue

                    # 查询Pallet表的所有PO_ID去重
                    pallet_po_ids = Pallet.objects.filter(
                        container_number=container
                    ).exclude(PO_ID__isnull=True).exclude(PO_ID='').values_list(
                        'PO_ID', flat=True
                    ).distinct()
                    
                    # 查询InvoiceItemv2表已记录的PO_ID
                    recorded_po_ids = InvoiceItemv2.objects.filter(
                        container_number=container,
                        invoice_number=invoice,
                        invoice_type="receivable",
                    ).exclude(PO_ID__isnull=True).exclude(PO_ID='').values_list(
                        'PO_ID', flat=True
                    ).distinct()

                    # 找出未记录的PO_ID
                    unrecorded_po_ids = set(pallet_po_ids) - set(recorded_po_ids)

                    # 如果有没记录到的PO_ID，order_data就加入到order列表
                    if unrecorded_po_ids:
                        order_data = {
                            'order': o,
                            'order_id': o.id,
                            'container_number__container_number': o.container_number.container_number if o.container_number else None,
                            'customer_name__zem_name': o.customer_name.zem_name if o.customer_name else None,
                            'order_type': o.order_type,
                            'retrieval_id__retrieval_destination_precise': o.retrieval_id.retrieval_destination_precise if o.retrieval_id else None,
                            'retrieval_id__retrieval_carrier': getattr(o.retrieval_id, 'retrieval_carrier', None) if o.retrieval_id else None,
                            'retrieval_id__actual_retrieval_timestamp': getattr(o.retrieval_id, 'actual_retrieval_timestamp', None) if o.retrieval_id else None,
                            'created_at': o.created_at,
                            'offload_time': o.offload_time,
                            
                            # 发票信息
                            'invoice_id__invoice_number': invoice.invoice_number,
                            'invoice_number': invoice.invoice_number,
                            'invoice_id': invoice.id,
                            'invoice_created_at': invoice.created_at if hasattr(invoice, 'created_at') else invoice.history.first().history_date if invoice.history.exists() else None,
                            'finance_status': finance_status,
                            'has_invoice': True,
                            'offload_time': o.offload_time,
                            
                            # 新增字段用于前端显示
                            'unrecorded_po_ids_count': len(unrecorded_po_ids),
                            'unrecorded_po_ids': list(unrecorded_po_ids)[:10],  # 只显示前10个
                            'pallet_total_count': len(pallet_po_ids),
                            'recorded_po_ids_count': len(recorded_po_ids),
                            'container_id': container.id if container else None,
                        }
                        order.append(order_data)
                
                # 如果有多条invoices的话
                else:
                    order_data = {
                        'order': o,
                        'order_id': o.id,
                        'container_number__container_number': o.container_number.container_number if o.container_number else None,
                        'customer_name__zem_name': o.customer_name.zem_name if o.customer_name else None,
                        'order_type': o.order_type,
                        'retrieval_id__retrieval_destination_precise': o.retrieval_id.retrieval_destination_precise if o.retrieval_id else None,
                        'retrieval_id__retrieval_carrier': getattr(o.retrieval_id, 'retrieval_carrier', None) if o.retrieval_id else None,
                        'retrieval_id__actual_retrieval_timestamp': getattr(o.retrieval_id, 'actual_retrieval_timestamp', None) if o.retrieval_id else None,
                        'created_at': o.created_at,
                        'offload_time': o.offload_time,
                        
                        # 发票信息
                        'invoice_id__invoice_number': invoice.invoice_number,
                        'invoice_number': invoice.invoice_number,
                        'invoice_id': invoice.id,
                        'invoice_created_at': invoice.created_at if hasattr(invoice, 'created_at') else invoice.history.first().history_date if invoice.history.exists() else None,
                        'finance_status': finance_status,
                        'has_invoice': True,
                        'offload_time': o.offload_time,
                        
                        # 新增字段用于显示有多条发票
                        'invoice_count': invoices.count(),
                        'invoice_index': list(invoices).index(invoice) + 1,
                        'container_id': container.id if container else None,
                    }
                    previous_order.append(order_data)

        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'order': order,
            'previous_order': previous_order,
            "order_form": OrderForm(),
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
        })
        return self.template_supplementary_entry, context


    def old_handle_confirm_entry_post(self, request:HttpRequest, context: dict| None = None,) -> Dict[str, Any]:
        if not context:
            context = {}
        warehouse = request.POST.get("warehouse_filter")
        customer = request.POST.get("customer")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date

        criteria = (
            Q(cancel_notification=False)
            & Q(vessel_id__vessel_etd__gte=start_date)
            & Q(vessel_id__vessel_etd__lte=end_date)
            & Q(offload_id__offload_at__isnull=False)
        )

        if warehouse:
            criteria &= Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= Q(customer_name__zem_name=customer)
    
        # 获取基础订单数据
        base_orders = (
            Order.objects
            .select_related(
                'retrieval_id', 
                'offload_id', 
                'container_number',
                'customer_name'
            )
            .annotate(
                retrieval_time=F("retrieval_id__actual_retrieval_timestamp"),
                empty_returned_time=F("retrieval_id__empty_returned_at"),
                offload_time=F("offload_id__offload_at"),
            )
            .filter(criteria)
            .distinct()
        )
        order = [] #公仓待录入
        previous_order = []  #公仓已录入
        
        for o in base_orders:
            container = o.container_number
            
            if not container:
                continue

            # 查询这个柜子的所有应收账单
            invoices = Invoicev2.objects.filter(container_number=container)
            
            if invoices.exists():
                has_multiple_invoices = invoices.count() > 1  #看是不是多份账单
                # 有账单的情况 - 每个账单都要单独处理
                for invoice in invoices:
                    # 查询这个账单对应的状态
                    try:
                        status_obj = InvoiceStatusv2.objects.get(
                            invoice=invoice,
                            invoice_type='receivable'
                        )
                        preport_status = status_obj.preport_status
                        warehouse_public_status = status_obj.warehouse_public_status
                        warehouse_other_status = status_obj.warehouse_other_status
                        delivery_other_status = status_obj.delivery_other_status
                        delivery_public_status = status_obj.delivery_public_status #公仓状态
                        finance_status = status_obj.finance_status #财务状态
                    except InvoiceStatusv2.DoesNotExist:
                        continue
                    if preport_status == "completed" and warehouse_public_status == "completed" and warehouse_other_status == "completed" and delivery_other_status == "completed" and delivery_public_status == "completed":
                        
                        invoice_created_at = None
                        if has_multiple_invoices:
                            invoice_created_at = (
                                invoice.created_at if hasattr(invoice, 'created_at') 
                                else invoice.history.first().history_date if invoice.history.exists() 
                                else None
                            )
                       
                        order_data = {
                            'order': o,
                            'order_id': o.id,
                            'container_number__container_number': o.container_number,
                            'customer_name__zem_name': o.customer_name.zem_name,
                            'order_type': o.order_type,
                            'invoice_created_at': invoice_created_at,
                            'retrieval_id__retrieval_destination_precise': o.retrieval_id.retrieval_destination_precise if o.retrieval_id else None,
                            'retrieval_id__retrieval_carrier': getattr(o.retrieval_id, 'retrieval_carrier', None) if o.retrieval_id else None,
                            'retrieval_id__actual_retrieval_timestamp': getattr(o.retrieval_id, 'actual_retrieval_timestamp', None) if o.retrieval_id else None,
                            'created_at': o.created_at,
                            'offload_time': o.offload_time,
                            
                            # 发票信息 - 使用与原来前端模板相同的字段名
                            'invoice_id__invoice_number': invoice.invoice_number,
                            'invoice_number': invoice.invoice_number,
                            'invoice_id': invoice.id,
                            'finance_status': finance_status,
                            'has_invoice': True,
                            'offload_time': o.offload_time,
                        }

                        is_hold = False

                        if finance_status != "completed":  
                            
                            #未开账单才看是否有暂扣
                            hold_subquery = Pallet.objects.filter(
                                container_number=container,
                                delivery_method__contains="暂扣留仓",
                                delivery_type="public"
                            )
                            if hold_subquery.exists():
                                is_hold = True
                            order_data.update({'is_hold':is_hold})
                            order.append(order_data)
                        else:
                            #已开的才看剩余金额
                            remain_offset = remain_offset
                            invoice_date = invoice.invoice_date
                            invoice_link = invoice.invoice_link
                            receivable_total_amount = getattr(invoice, 'receivable_total_amount', 0)
                            payable_total_amount = getattr(invoice, 'payable_total_amount', 0)
                            invoice_statement_id = invoice.statement_id.invoice_statement_id if invoice.statement_id else None
                            statement_link =  invoice.statement_id.statement_link if invoice.statement_id else None
                            order_data.update({
                                'invoice_id__invoice_date': invoice_date,
                                'invoice_id__invoice_link': invoice_link,
                                'invoice_id__receivable_total_amount': receivable_total_amount,
                                'invoice_id__payable_total_amount': payable_total_amount,
                                'invoice_id__remain_offset': remain_offset,
                                'invoice_id__is_invoice_delivered': invoice.is_invoice_delivered,
                                'invoice_id__statement_id__invoice_statement_id': invoice_statement_id,
                                'invoice_id__statement_id__statement_link':statement_link,
                            })
                            previous_order.append(order_data)
        
        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'order': order,
            'previous_order': previous_order,
            "order_form": OrderForm(),
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
        })
        return self.template_confirm_entry, context
    
    def handle_confirm_entry_post(self, request:HttpRequest, context: dict| None = None,) -> Dict[str, Any]:
        '''财务账单确认查询'''
        if not context:
            context = {}
        warehouse = request.POST.get("warehouse_filter")
        customer = request.POST.get("customer")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        # --- 1. 日期处理 ---
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date

        # --- 2. 构建查询条件 ---
        criteria = (
            Q(cancel_notification=False)
            & Q(vessel_id__vessel_etd__gte=start_date)
            & Q(vessel_id__vessel_etd__lte=end_date)
            & Q(offload_id__offload_at__isnull=False)
        )

        if warehouse:
            criteria &= Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= Q(customer_name__zem_name=customer)
    
        # --- 3. 获取基础订单数据 ---
        base_orders = (
            Order.objects
            .select_related(
                'retrieval_id', 
                'offload_id', 
                'container_number',
                'customer_name'
            )
            .annotate(
                retrieval_time=F("retrieval_id__actual_retrieval_timestamp"),
                empty_returned_time=F("retrieval_id__empty_returned_at"),
                offload_time=F("offload_id__offload_at"),
            )
            .filter(criteria)
            .distinct()
        )

        orders_list = list(base_orders)
    
        # 提取所有涉及的 Container IDs
        container_ids = set()
        for o in orders_list:
            if o.container_number_id:
                container_ids.add(o.container_number_id)
        container_ids = list(container_ids)

        # --- 4. [优化核心] 批量获取发票及其状态 ---
        # 预加载条件：只取 invoice_type='receivable' 的状态
        status_prefetch = Prefetch(
            'invoicestatusv2_set',
            queryset=InvoiceStatusv2.objects.filter(invoice_type="receivable"),
            to_attr='receivable_status_list'
        )

        # 预加载 Statement (为了获取 link 和 ID)
        # 假设 Invoicev2 有 statement_id 字段关联到 Statement 表
        all_invoices = Invoicev2.objects.filter(
            container_number_id__in=container_ids
        ).select_related(
            'statement_id' # 避免循环中访问 statement_id 再次查询
        ).prefetch_related(status_prefetch)

        # 将发票按 container_id 分组
        container_invoice_map = defaultdict(list)
        for inv in all_invoices:
            container_invoice_map[inv.container_number_id].append(inv)

        # --- 5. [优化核心] 批量获取“暂扣”信息 ---
        # 只需要知道哪些 container_id 有暂扣记录即可
        hold_container_ids = set(
            Pallet.objects.filter(
                container_number_id__in=container_ids,
                delivery_method__contains="暂扣留仓",
                delivery_type="public"
            ).values_list('container_number_id', flat=True)
        )

        # --- 6. 循环处理 ---
        order_data_list = []      # 公仓待录入 (对应原代码 order)
        previous_order_data_list = [] # 公仓已录入 (对应原代码 previous_order)
        
        for o in base_orders:
            container = o.container_number
            
            if not container:
                continue
            
            c_id = container.id
            invoices = container_invoice_map.get(c_id, [])
            
            if not invoices:
                continue
            has_multiple_invoices = len(invoices) > 1

            for invoice in invoices:
                # 获取预加载的状态列表
                status_list = getattr(invoice, 'receivable_status_list', [])
                if not status_list:
                    continue
                
                status_obj = status_list[0] # 取第一个匹配的状态

                # 提取状态字段
                preport_status = status_obj.preport_status
                wh_public = status_obj.warehouse_public_status
                wh_other = status_obj.warehouse_other_status
                del_other = status_obj.delivery_other_status
                del_public = status_obj.delivery_public_status
                finance_status = status_obj.finance_status

                # --- 核心判断逻辑：前置节点全 Completed ---
                if (preport_status == "completed" and 
                    wh_public == "completed" and 
                    wh_other == "completed" and 
                    del_other == "completed" and 
                    del_public == "completed"):

                    # 处理发票创建时间逻辑
                    invoice_created_at = None
                    if has_multiple_invoices:
                        invoice_created_at = invoice.created_at

                    # 构建基础数据字典
                    row_data = {
                        'order': o,
                        'order_id': o.id,
                        'container_number__container_number': container.container_number,
                        'customer_name__zem_name': o.customer_name.zem_name if o.customer_name else None,
                        'order_type': o.order_type,
                        'invoice_created_at': invoice_created_at,
                        # 使用 getattr 安全获取关联属性，防止 NoneType 报错
                        'retrieval_id__retrieval_destination_precise': o.retrieval_id.retrieval_destination_precise if o.retrieval_id else None,
                        'retrieval_id__retrieval_carrier': getattr(o.retrieval_id, 'retrieval_carrier', None) if o.retrieval_id else None,
                        'retrieval_id__actual_retrieval_timestamp': getattr(o.retrieval_id, 'actual_retrieval_timestamp', None) if o.retrieval_id else None,
                        'created_at': o.created_at,
                        'offload_time': o.offload_time,
                        
                        # 发票信息
                        'invoice_id__invoice_number': invoice.invoice_number,
                        'invoice_number': invoice.invoice_number,
                        'invoice_id': invoice.id,
                        'finance_status': finance_status,
                        'has_invoice': True,
                    }

                    if finance_status != "completed":
                        # === 待录入 (Order List) ===
                        # 检查是否有暂扣 (直接查 Set，O(1)复杂度)
                        is_hold = c_id in hold_container_ids
                        
                        row_data['is_hold'] = is_hold
                        order_data_list.append(row_data)
                    
                    else:
                        # === 已录入/已完成 (Previous Order List) ===
                        # 计算剩余金额
                        rec_total = getattr(invoice, 'receivable_total_amount', 0) or 0
                        rec_offset = getattr(invoice, 'receivable_offset_amount', 0) or 0

                        # 处理 Statement 关联
                        stmt = invoice.statement_id
                        stmt_id = stmt.invoice_statement_id if stmt else None
                        stmt_link = stmt.statement_link if stmt else None

                        row_data.update({
                            'invoice_id__invoice_date': invoice.invoice_date,
                            'invoice_id__invoice_link': invoice.invoice_link,
                            'invoice_id__receivable_total_amount': rec_total,
                            'invoice_id__payable_total_amount': getattr(invoice, 'payable_total_amount', 0),
                            'invoice_id__remain_offset': rec_offset,
                            'invoice_id__is_invoice_delivered': invoice.is_invoice_delivered,
                            'invoice_id__statement_id__invoice_statement_id': stmt_id,
                            'invoice_id__statement_id__statement_link': stmt_link,
                        })
                        previous_order_data_list.append(row_data)
        
        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'order': order_data_list,
            'previous_order': previous_order_data_list,
            "order_form": OrderForm(),
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
        })
        return self.template_confirm_entry, context
    
    def old_handle_delivery_entry_post(self, request:HttpRequest, context: dict| None = None,) -> Dict[str, Any]:
        warehouse = request.POST.get("warehouse_filter")
        customer = request.POST.get("customer")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date

        criteria = (
            Q(cancel_notification=False)
            & (Q(order_type="转运") | Q(order_type="转运组合"))
            & Q(vessel_id__vessel_etd__gte=start_date)
            & Q(vessel_id__vessel_etd__lte=end_date)
            & Q(offload_id__offload_at__isnull=False)
        )

        if warehouse and warehouse != 'None':
            if "LA" in warehouse:
                criteria &= Q(retrieval_id__retrieval_destination_precise__contains='LA')
            else:
                criteria &= Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= Q(customer_name__zem_name=customer)
    
        # 获取基础订单数据
        base_orders = (
            Order.objects
            .select_related(
                'retrieval_id', 
                'offload_id', 
                'container_number',
                'customer_name'
            )
            .annotate(
                retrieval_time=F("retrieval_id__actual_retrieval_timestamp"),
                empty_returned_time=F("retrieval_id__empty_returned_at"),
                offload_time=F("offload_id__offload_at"),
            )
            .filter(criteria)
            .distinct()
        )
        dl_public_to_record_orders = [] #公仓待录入
        dl_public_recorded_orders = []  #公仓已录入
        dl_self_to_record_orders = []  #私仓待录入
        dl_self_recorded_orders = []  #私仓已录入

        containers = [
            order.container_number
            for order in base_orders
            if order.container_number
        ]

        hold_map = defaultdict(lambda: {"public": False, "other": False})
        pallets = Pallet.objects.filter(
            container_number__in=containers,
            delivery_method__contains="暂扣留仓",
            delivery_type__in=["public", "other"],
        ).values("container_number", "delivery_type")
        for p in pallets:
            hold_map[p["container_number"]][p["delivery_type"]] = True

        invoice_map = defaultdict(list)
        invoices = Invoicev2.objects.filter(container_number__in=containers)
        for inv in invoices:
            invoice_map[inv.container_number].append(inv)

        invoice_ids = [inv.id for inv in invoices]
        status_map = {}
        statuses = InvoiceStatusv2.objects.filter(
            invoice_id__in=invoice_ids,
            invoice_type="receivable",
        )
        for s in statuses:
            status_map[s.invoice_id] = s

        for order in base_orders:
            container = order.container_number
            
            if not container:
                continue

            container_delivery_type = getattr(container, 'delivery_type', 'mixed')   
            # 判断是否应该处理公仓或私仓 
            should_process_public = container_delivery_type in ['public', 'mixed']
            should_process_self = container_delivery_type in ['other', 'mixed']
            # 如果没有柜子类型信息，默认都处理
            if container_delivery_type not in ['public', 'other', 'mixed']:
                should_process_public = True
                should_process_self = True

            is_hold = False
            if should_process_public and hold_map[container]["public"]:
                is_hold = True
            if should_process_self and hold_map[container]["other"]:
                is_hold = True

            # 查询这个柜子的所有应收账单
            container_invoices = invoice_map.get(container, [])

            if not container_invoices:
                # 没有账单的情况 - 归到待录入
                order_data = {
                    'order': order,
                    'container_number': order.container_number,
                    'invoice_number': None,
                    'invoice_id': None,
                    'invoice_created_at': None,
                    'preport_status': None,
                    'finance_status': None,
                    'has_invoice': False,
                    'offload_time': order.offload_time,
                    "is_hold": is_hold,
                }
                if should_process_public:
                    dl_public_to_record_orders.append(order_data)
                if should_process_self:
                    dl_self_to_record_orders.append(order_data)
            else:
                has_multiple_invoices = len(container_invoices) > 1 #看看是不是补开的账单
                # 有账单的情况 - 每个账单都要单独处理
                for invoice in container_invoices:
                    # 查询这个账单对应的状态
                    status_obj = status_map.get(invoice.id)

                    if status_obj:
                        public_status = status_obj.delivery_public_status
                        self_status = status_obj.delivery_other_status
                        finance_status = status_obj.finance_status
                        delivery_public_reason = status_obj.delivery_public_reason
                        delivery_other_reason = status_obj.delivery_other_reason
                    else:
                        public_status = None
                        self_status = None
                        finance_status = None
                        delivery_public_reason = None
                        delivery_other_reason = None

                    # 只在有多个账单时添加 invoice_created_at                 
                    invoice_created_at = None
                    if has_multiple_invoices:
                        invoice_created_at = (
                            invoice.created_at if hasattr(invoice, 'created_at') 
                            else invoice.history.first().history_date if invoice.history.exists() 
                            else None
                        )
                    order_data = {
                        'order': order,
                        'container_number': order.container_number,
                        'invoice_number': invoice.invoice_number,
                        'invoice_id': invoice.id,
                        'invoice_created_at': invoice_created_at,
                        'public_status': public_status,
                        'self_status': self_status,
                        'finance_status': finance_status,
                        'has_invoice': True,
                        'offload_time': order.offload_time,
                        'is_hold': is_hold,
                        'delivery_public_reason': delivery_public_reason,
                        'delivery_other_reason': delivery_other_reason,
                    }
                    # 根据状态分组
                    if should_process_public:                 
                        if public_status in ["unstarted", "in_progress"] or public_status is None:
                            dl_public_to_record_orders.append(order_data)
                        elif public_status in ["completed", "rejected"]:
                            dl_public_recorded_orders.append(order_data)
                    if should_process_self:
                        if self_status in ["unstarted", "in_progress"] or self_status is None:
                            dl_self_to_record_orders.append(order_data)
                        elif self_status in ["completed", "rejected"]:
                            dl_self_recorded_orders.append(order_data)
        
        #按照出库比例排序
        dl_public_to_record_orders = self._add_shipment_group_stats(dl_public_to_record_orders, "public")
        dl_self_to_record_orders = self._add_shipment_group_stats(dl_self_to_record_orders, "other")

        #已录入的，驳回优先显示
        dl_public_recorded_orders.sort(key=lambda x: x.get('public_status') != 'rejected')
        dl_self_recorded_orders.sort(key=lambda x: x.get('self_status') != 'rejected')
        
        # 判断用户权限，决定默认标签页
        groups = [group.name for group in request.user.groups.all()]
        if not context:
            context = {}

        # 根据权限，决定打开的标签页
        if 'warehouse_other' in groups:
            default_tab = 'self'
        else:
            default_tab = 'public'
        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'wh_public_to_record_orders': dl_public_to_record_orders,
            'wh_public_recorded_orders': dl_public_recorded_orders,
            'wh_self_to_record_orders': dl_self_to_record_orders,
            'wh_self_recorded_orders': dl_self_recorded_orders,
            "order_form": OrderForm(),
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
            "default_tab": default_tab, 
        })
        return context
    
    def handle_delivery_entry_post(self, request:HttpRequest, context: dict| None = None,) -> Dict[str, Any]:
        warehouse = request.POST.get("warehouse_filter")
        customer = request.POST.get("customer")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        # --- 1. 日期处理优化 ---
        current_date = datetime.now().date()
        start_date = (
            (current_date + timedelta(days=-90)).strftime("%Y-%m-%d")
            if not start_date
            else start_date
        )
        end_date = current_date.strftime("%Y-%m-%d") if not end_date else end_date

        # --- 2. 构建基础查询条件 ---
        criteria = (
            Q(cancel_notification=False)
            & (Q(order_type="转运") | Q(order_type="转运组合"))
            & Q(vessel_id__vessel_etd__gte=start_date)
            & Q(vessel_id__vessel_etd__lte=end_date)
            & Q(offload_id__offload_at__isnull=False)
        )

        if warehouse and warehouse != 'None':
            if "LA" in warehouse:
                criteria &= Q(retrieval_id__retrieval_destination_precise__contains='LA')
            else:
                criteria &= Q(retrieval_id__retrieval_destination_precise=warehouse)
        if customer:
            criteria &= Q(customer_name__zem_name=customer)

        # --- 3. 获取基础订单数据 ---
        base_orders = (
            Order.objects
            .select_related(
                'retrieval_id', 
                'offload_id', 
                'container_number',
                'customer_name'
            )
            .annotate(
                retrieval_time=F("retrieval_id__actual_retrieval_timestamp"),
                empty_returned_time=F("retrieval_id__empty_returned_at"),
                offload_time=F("offload_id__offload_at"),
            )
            .filter(criteria)
            .distinct()
        )

        # 转换为列表，避免后续多次触发 DB 查询
        orders_list = list(base_orders)

        # 提取所有涉及的 Container ID，用于后续批量查询
        containers = set()
        container_ids = []
        for order in orders_list:
            if order.container_number:
                containers.add(order.container_number)
                container_ids.append(order.container_number_id)
        
        container_ids = list(set(container_ids)) # 去重

        # --- 4. 批量获取 Hold (暂扣) 状态 ---
        hold_map = defaultdict(lambda: {"public": False, "other": False})
        if container_ids:
            hold_pallets = Pallet.objects.filter(
                container_number_id__in=container_ids,
                delivery_method__contains="暂扣留仓",
                delivery_type__in=["public", "other"],
            ).values("container_number_id", "delivery_type")
            
            for p in hold_pallets:
                hold_map[p["container_number_id"]][p["delivery_type"]] = True

        # --- 5. 批量获取 Invoice 和 InvoiceStatus (关键优化) ---
        status_prefetch = Prefetch(
            'invoicestatusv2_set', 
            queryset=InvoiceStatusv2.objects.filter(invoice_type="receivable"),
            to_attr='receivable_status_list'
        )
        
        all_invoices = Invoicev2.objects.filter(
            container_number_id__in=container_ids
        ).prefetch_related(status_prefetch) # 执行预查询
        
        # 将 Invoice 按 Container ID 分组
        container_invoice_map = defaultdict(list)
        for inv in all_invoices:
            container_invoice_map[inv.container_number_id].append(inv)

        # --- 6. 批量预计算出库统计 ---
        shipment_stats_map = self._bulk_calculate_shipment_stats(container_ids)

        # --- 7. 数据组装 (纯内存操作) ---
        dl_public_to_record_orders = [] #公仓待录入
        dl_public_recorded_orders = []  #公仓已录入
        dl_self_to_record_orders = []  #私仓待录入
        dl_self_recorded_orders = []  #私仓已录入

        for order in orders_list:
            container = order.container_number
            
            if not container:
                continue
            
            c_id = container.id
            container_delivery_type = container.delivery_type if container.delivery_type else 'mixed' 
            # 判断是否应该处理公仓或私仓 
            should_process_public = container_delivery_type in ['public', 'mixed']
            should_process_self = container_delivery_type in ['other', 'mixed']
            # 如果没有柜子类型信息，默认都处理
            if container_delivery_type not in ['public', 'other', 'mixed']:
                should_process_public = True
                should_process_self = True

            is_hold_public = should_process_public and hold_map[c_id]["public"]
            is_hold_other = should_process_self and hold_map[c_id]["other"]

            # 查询这个柜子的所有应收账单
            container_invoices = container_invoice_map.get(c_id, [])

            # 提取公共数据构建逻辑
            def build_order_data(inv=None, status_obj=None, is_hold=False):
                created_at = None
                if inv and len(container_invoices) > 1:
                    # 只有多个账单时才需要时间区分，避免额外 getattr 开销
                    created_at = inv.created_at # 假设 created_at 存在，不做 history 复杂查询以保性能
                
                return {
                    'order': order,
                    'container_number': container,
                    'invoice_number': inv.invoice_number if inv else None,
                    'invoice_id': inv.id if inv else None,
                    'invoice_created_at': created_at,
                    'public_status': status_obj.delivery_public_status if status_obj else None,
                    'self_status': status_obj.delivery_other_status if status_obj else None,
                    'finance_status': status_obj.finance_status if status_obj else None,
                    'has_invoice': bool(inv),
                    'offload_time': order.offload_time,
                    'is_hold': is_hold,
                    'delivery_public_reason': status_obj.delivery_public_reason if status_obj else None,
                    'delivery_other_reason': status_obj.delivery_other_reason if status_obj else None,
                    # 从批量计算的 map 中获取统计信息
                    'stats_raw': shipment_stats_map.get(c_id, {}) 
                }
            
            if not container_invoices:
                # 无账单逻辑
                base_data = build_order_data(None, None, False) # is_hold 会在下面覆盖
                
                if should_process_public:
                    item = base_data.copy()
                    item['is_hold'] = is_hold_public
                    self._inject_stats_and_ratio(item, 'public') # 注入并计算比例
                    dl_public_to_record_orders.append(item)
                    
                if should_process_self:
                    item = base_data.copy()
                    item['is_hold'] = is_hold_other
                    self._inject_stats_and_ratio(item, 'other')
                    dl_self_to_record_orders.append(item)
            else:
                # 有账单逻辑
                for invoice in container_invoices:
                    # 因为做了 prefetch_related，这里访问 receivable_statuses 不会查库
                    # 且 filter(invoice_type='receivable') 已经在 prefetch 中做了，这里取第一个即可
                    # 注意：invoice_statusesv2 是 OneToMany 还是一对一？模型里是 ForeignKey，所以是 list
                    status_obj = None
                    if hasattr(invoice, 'receivable_status_list') and invoice.receivable_status_list:
                        status_obj = invoice.receivable_status_list[0]
                    
                    base_data = build_order_data(invoice, status_obj, False)

                    # 公仓分派
                    if should_process_public:
                        item = base_data.copy()
                        item['is_hold'] = is_hold_public
                        self._inject_stats_and_ratio(item, 'public')
                        p_status = item['public_status']
                        if p_status in ["unstarted", "in_progress", None]:
                            dl_public_to_record_orders.append(item)
                        elif p_status in ["completed", "rejected"]:
                            dl_public_recorded_orders.append(item)

                    # 私仓分派
                    if should_process_self:
                        item = base_data.copy()
                        item['is_hold'] = is_hold_other
                        self._inject_stats_and_ratio(item, 'other')
                        s_status = item['self_status']
                        if s_status in ["unstarted", "in_progress", None]:
                            dl_self_to_record_orders.append(item)
                        elif s_status in ["completed", "rejected"]:
                            dl_self_recorded_orders.append(item)
        
         # --- 8. 排序 ---
        dl_public_to_record_orders.sort(key=lambda x: x.get('completion_ratio', 0), reverse=True)
        dl_self_to_record_orders.sort(key=lambda x: x.get('completion_ratio', 0), reverse=True)

        #已录入的，驳回优先显示
        dl_public_recorded_orders.sort(key=lambda x: x.get('public_status') != 'rejected')
        dl_self_recorded_orders.sort(key=lambda x: x.get('self_status') != 'rejected')
        
        # 判断用户权限，决定默认标签页
        groups = [group.name for group in request.user.groups.all()] if request.user else []
        if not context:
            context = {}

        # 根据权限，决定打开的标签页
        default_tab = 'self' if 'warehouse_other' in groups else 'public'

        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'dl_public_to_record_orders': dl_public_to_record_orders,
            'dl_public_recorded_orders': dl_public_recorded_orders,
            'dl_self_to_record_orders': dl_self_to_record_orders,
            'dl_self_recorded_orders': dl_self_recorded_orders,
            "order_form": OrderForm(),
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
            "default_tab": default_tab, 
        })
        return context
    
    def _bulk_calculate_shipment_stats(self, container_ids):
        """
        一次性查询所有 Container 的 Pallet 和 PackingList 数据，并在内存中聚合
        """
        if not container_ids:
            return {}

        stats_map = defaultdict(lambda: {
            "public": {"total": 0, "shipped": 0, "unshipped": 0},
            "other": {"total": 0, "shipped": 0, "unshipped": 0},
        })

        # 1. 批量查询 Pallet
        # 注意：这里去掉了 offload_at 的复杂过滤，因为逻辑上我们已经在 Base Order 里筛选了 Offloaded 的订单
        # 如果必须严格遵循原来的逻辑（Pallet 算，PackingList 不算），可以在这里加条件
        pallets = (
            Pallet.objects
            .filter(container_number_id__in=container_ids)
            .values(
                'container_number_id',
                'delivery_type',
                'PO_ID',
                'shipment_batch_number',
            )
            .distinct()
        )

        for p in pallets:
            c_id = p['container_number_id']
                     
            d_type = p['delivery_type'] if p['delivery_type'] in ['public', 'other'] else 'public' # 默认归类
            
            stats_map[c_id][d_type]['total'] += 1

            
            if p['shipment_batch_number']:
                shipment = Shipment.objects.get(id=p['shipment_batch_number'])
                if shipment.shipped_at:
                    stats_map[c_id][d_type]['shipped'] += 1
                else:
                    stats_map[c_id][d_type]['unshipped'] += 1
            else:
                stats_map[c_id][d_type]['unshipped'] += 1
        return stats_map

    def _inject_stats_and_ratio(self, order_item, display_type):
        """
        从 order_item 临时存储的 'stats_raw' 中提取特定类型（public或other）的统计数据，
        计算完成比例，并注入到 order_item 字典中供模板使用。
        
        :param order_item: 包含订单信息和 stats_raw 的字典
        :param display_type: 'public' 或 'other'，指示当前需要显示哪种类型的进度
        """
        # 1. 获取预先计算好的原始统计数据，如果不存在则给空字典
        raw_stats = order_item.get('stats_raw', {})
        
        # 2. 根据 display_type ('public' 或 'other') 获取具体的计数值
        # 如果找不到对应类型的统计，默认全为 0
        default_stats = {"total": 0, "shipped": 0, "unshipped": 0}
        type_stats = raw_stats.get(display_type, default_stats)
        
        total = type_stats['total']
        shipped = type_stats['shipped']
        unshipped = type_stats['unshipped']
        
        # 3. 将统计结果注入到 order_item 中 (保持原有命名习惯)
        order_item['total_shipment_groups'] = total
        order_item['shipped_shipment_groups'] = shipped
        order_item['unshipped_shipment_groups'] = unshipped
        
        # 4. 计算完成比例 (用于前端进度条或排序)
        # 避免除以零错误
        if total > 0:
            order_item['completion_ratio'] = shipped / total
        else:
            order_item['completion_ratio'] = 0.0
        
        # 5. 清理临时数据 stats_raw，保持传递给前端的数据整洁
        # 注意：这里使用 copy() 或者直接修改都可以，因为每行数据是独立的字典
        if 'stats_raw' in order_item:
            del order_item['stats_raw']
            
        return order_item
    
    def get_shipment_group_stats(self, queryset, delivery_type_q):
        """
        获取分组统计信息
        """
        # 应用delivery_type筛选
        if delivery_type_q:
            queryset = queryset.filter(delivery_type_q)
        
        # 按destination和shipment_batch_number分组
        groups = queryset.values('destination', 'shipment_batch_number__shipment_batch_number').annotate(
            group_count=Count('id')
        )
        total_groups = groups.count()
        
        # 统计已出库和未出库的分组
        shipped_groups = 0
        unshipped_groups = 0
        
        for group in groups:
            shipment_batch_number = group['shipment_batch_number__shipment_batch_number']
            
            if shipment_batch_number:
                # 检查shipment是否已出库
                shipment = Shipment.objects.get(shipment_batch_number=shipment_batch_number)
                if shipment.shipped_at:
                    shipped_groups += 1
                else:
                    unshipped_groups += 1
            else:
                # 没有shipment_batch_number的视为未出库
                unshipped_groups += 1
        
        return {
            'total_groups': total_groups,
            'shipped_groups': shipped_groups,
            'unshipped_groups': unshipped_groups
        }
    
    def _add_shipment_group_stats(self, orders, display_mix):
        """
        为每个order添加分组统计信息
        """
        if not orders:
            return []
        # 获取用户权限对应的delivery_type筛选条件
        for order in orders:
            # 查找该order关联的packinglist和pallet
            packinglist_stats = self.get_shipment_group_stats(
                PackingList.objects.filter(
                    container_number__container_number=order["container_number"],
                    container_number__orders__offload_id__offload_at__isnull=True,
                ).select_related('shipment_batch_number'),
                Q(delivery_type=display_mix)
            )
            pallet_stats = self.get_shipment_group_stats(
                Pallet.objects.filter(
                    container_number__container_number=order['container_number'],
                    container_number__orders__offload_id__offload_at__isnull=False,
                ).select_related('shipment_batch_number'),
                Q(delivery_type=display_mix)
            )
            
            # 合并统计结果
            total_groups = packinglist_stats['total_groups'] + pallet_stats['total_groups']
            shipped_groups = packinglist_stats['shipped_groups'] + pallet_stats['shipped_groups']
            unshipped_groups = packinglist_stats['unshipped_groups'] + pallet_stats['unshipped_groups']
            
            # 添加到order对象（不改变原有结构）
            order['total_shipment_groups'] = total_groups
            order['shipped_shipment_groups'] = shipped_groups
            order['unshipped_shipment_groups'] = unshipped_groups
            order['completion_ratio'] = shipped_groups / total_groups if total_groups > 0 else 0
            
        sorted_orders = sorted(orders, key=lambda x: x['completion_ratio'], reverse=True)
        return sorted_orders
    
    def handle_modify_order_type(self, request:HttpRequest) -> Dict[str, Any]:
        container_number = request.POST.get("container_number")
        new_order_type = request.POST.get("new_order_type")

        container = Container.objects.get(container_number=container_number)
        if new_order_type == "转运":
            actual_non_combina_reason = request.POST.get("actual_non_combina_reason")
            container.manually_order_type = "转运"
            container.non_combina_reason = actual_non_combina_reason
        elif new_order_type == "转运组合":
            container.manually_order_type = "转运组合"
        container.save()
        context = {"success_messages": f"{container_number}修改类型成功！"}
        page =  request.POST.get("page")
        if page == "delivery_edit":
            return self.handle_container_delivery_post(request, context)
        else:
            return self.handle_container_preport_post(request, context)

    def handle_container_preport_post(self, request:HttpRequest, context: dict|None=None) -> Dict[str, Any]:
        """处理柜号点击进入港前账单编辑页面"""
        if not context:
            context = {}
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        container_number = request.GET.get("container_number")
        invoice_id = request.GET.get("invoice_id")
        
        #获取订单信息
        order = Order.objects.select_related(
            "retrieval_id", "container_number", "warehouse"
        ).get(container_number__container_number=container_number)

        if invoice_id and invoice_id != 'None':
            #找到要修改的那份账单
            invoice = Invoicev2.objects.get(id=invoice_id)
            invoice_status, created = InvoiceStatusv2.objects.get_or_create(
                invoice=invoice,
                invoice_type="receivable",
                defaults={
                    "container_number": order.container_number,
                    "invoice": invoice,
                }
            )
        else:
            #说明这个柜子没有创建过账单，需要创建
            invoice, invoice_status = self._create_invoice_and_status(container_number)

        # 查看仓库和柜型，计算提拆费
        warehouse = order.retrieval_id.retrieval_destination_area
        container_type = order.container_number.container_type
        
        #查找报价表
        quotation, quotation_error = self._get_quotation_for_order(order, 'receivable')
        if quotation_error:
            context.update({"error_messages": quotation_error})
            return self.template_preport_edit, context

        order_type = order.order_type
        non_combina_reason = None

        if order_type != "转运组合":
            iscombina = False
        else:
            container = Container.objects.get(container_number=container_number)
            if container.manually_order_type == "转运":
                iscombina = False
                non_combina_reason = container.non_combina_reason
            elif container.manually_order_type == "转运组合":
                iscombina = True
            else:
                combina_context, iscombina,non_combina_reason = self._is_combina(container_number)
                if combina_context.get("error_messages"):
                    return self.template_preport_edit, combina_context
        if order_type == "直送":
            fee_detail, fee_error = self._get_fee_details_from_quotation(quotation, "direct")
        else:
            fee_detail, fee_error = self._get_fee_details_from_quotation(quotation, "preport")
        if fee_error:
            context.update({"error_messages": fee_error})
            return self.template_preport_edit, context
        
        # 计算提拆费   
        match = re.match(r"\d+", container_type)
        pickup_fee = 0
        if match:
            pick_subkey = match.group()
            try:
                pickup_fee = fee_detail.details[warehouse][pick_subkey]
            except KeyError:
                pickup_fee = 0
                # context.update({"error_messages": f"在报价表中找不到{warehouse}仓库{pick_subkey}柜型的提拆费"})
                # return self.template_preport_edit, context
        # 构建费用提示信息
        if order_type == "直送":
            destination = order.retrieval_id.retrieval_destination_area
            new_destination = destination.replace(" ", "") if destination else ""

            # 提拆、打托缠膜费用
            pickup_fee = 0
            pickup = fee_detail.details["pickup"]
            for fee, location in pickup.items():
                if warehouse in location:
                    pickup_fee = fee
            #二次派送
            second_delivery = fee_detail.details.get("二次派送")
            second_pickup = None
            if second_delivery:
                for fee, location in second_delivery.items():
                    if new_destination in location:
                        second_pickup = fee
            FS = {
                "提+派送": pickup_fee,
                "查验柜运费": f"{fee_detail.details.get('查验柜运费', 'N/A')}",  # 查验费
                "二次派送": second_pickup,  # 二次派送
                "滞港费": f"{fee_detail.details.get('滞港费', 'N/A')}",  # 滞港费
                "滞箱费": f"{fee_detail.details.get('滞箱费', 'N/A')}",  # 滞箱费
                "港口拥堵费": f"{fee_detail.details.get('港口拥堵费', 'N/A')}",  # 港口拥堵费
                "车架费": f"{fee_detail.details.get('车架费', 'N/A')}",  # 车架费
                "预提费": f"{fee_detail.details.get('预提费', 'N/A')}",  # 预提费
                "货柜储存费": f"{fee_detail.details.get('货柜储存费', 'N/A')}",  # 货柜储存费
                "等待费": f"{fee_detail.details.get('等待费', 'N/A')}",  # 等待费
                "车架分离费": f"{fee_detail.details.get('车架分离费', 'N/A')}",  # 车架分离费
                "超重费": f"{fee_detail.details.get('超重费', 'N/A')}",  # 超重费
            }
            # 标准费用项目列表
            standard_fee_items = [
                "提+派送","查验柜运费", "二次派送", "滞港费",  
                "滞箱费", "港口拥堵费", "车架费", "预提费", 
                "货柜储存费", "等待费", "车架分离费", "超重费"
            ]
        else:
            FS = {
                "提拆/打托缠膜": f"{pickup_fee}",
                "托架费": f"{fee_detail.details.get('托架费', 'N/A')}",
                "托架提取费": f"{fee_detail.details.get('托架提取费', 'N/A')}",
                "预提费": f"{fee_detail.details.get('预提费', 'N/A')}",
                "货柜放置费": f"{fee_detail.details.get('货柜放置费', 'N/A')}",
                "操作处理费": f"{fee_detail.details.get('操作处理费', 'N/A')}",
                "码头": fee_detail.details.get("码头", "N/A"),
                "港口拥堵费": f"{fee_detail.details.get('港口拥堵费', 'N/A')}",
                "吊柜费": f"{fee_detail.details.get('火车站吊柜费', 'N/A')}",
                "空跑费": f"{fee_detail.details.get('空跑费', 'N/A')}",
                "查验费": f"{fee_detail.details.get('查验费', 'N/A')}",
                "危险品": f"{fee_detail.details.get('危险品', 'N/A')}",
                "超重费": f"{fee_detail.details.get('超重费', 'N/A')}",
                "加急费": f"{fee_detail.details.get('加急费', 'N/A')}",
                "其他服务": f"{fee_detail.details.get('其他服务', 'N/A')}",
                "港内滞期费": f"{fee_detail.details.get('港内滞期费', 'N/A')}",
                "港外滞期费": f"{fee_detail.details.get('港外滞期费', 'N/A')}",
                "二次提货": f"{fee_detail.details.get('二次提货', 'N/A')}",
            }
            # 标准费用项目列表
            standard_fee_items = [
                "提拆/打托缠膜", "托架费", "托架提取费", "预提费", "货柜放置费", 
                "操作处理费", "码头", "港口拥堵费", "吊柜费", "空跑费", 
                "查验费", "危险品", "超重费", "加急费", "其他服务", 
                "港内滞期费", "港外滞期费", "二次提货"
            ]
        # 获取现有的费用项目
        existing_items = InvoiceItemv2.objects.filter(
            invoice_number=invoice,
            item_category="preport"
        )
        # 获取已存在的费用描述列表，用于前端过滤
        existing_descriptions = [item.description for item in existing_items]

        
        # 构建费用数据
        fee_data = []
        for item in existing_items:
            fee_data.append({
                'id': item.id,
                'description': item.description,
                'qty': item.qty or 1,
                'rate': item.rate or 0,
                'surcharges': item.surcharges or 0, 
                'amount': item.amount or 0,
                'note': item.note or ''
            })

        # 如果是第一次录入且没有费用记录，添加提拆费作为默认
        if not existing_items.exists() and invoice_status.preport_status == 'unstarted':          
            for fee_name in standard_fee_items:    
                if order_type != "直送" and fee_name == '提拆/打托缠膜':
                    # 提拆费特殊处理
                    pickup_qty = 0 if iscombina else 1
                    pickup_amount = pickup_fee * pickup_qty
                    fee_data.append({
                        'id': None,
                        'description': fee_name,
                        'qty': pickup_qty,
                        'rate': pickup_fee,
                        'surcharges': 0,
                        'amount': pickup_amount,
                        'note': '',
                    })
                else:                  
                    ref_price = FS.get(fee_name, 0)                   
                    numeric_price = self._extract_number(ref_price)
                    # 其他费用默认显示，但数量和金额为0
                    fee_data.append({
                        'id': None,
                        'description': fee_name,
                        'qty': 0,
                        'rate': numeric_price,
                        'surcharges': 0,
                        'amount': 0,
                        'note': '',
                    })
        groups = [group.name for group in request.user.groups.all()]
        is_leader = False
        if "invoice_preport_leader" in groups or request.user.is_staff:
            is_leader = True
  
        COMBINA_STIPULATE = FeeDetail.objects.get(
            quotation_id=quotation.id,
            fee_type='COMBINA_STIPULATE'
        )
        rules_text = self._parse_combina_rules(COMBINA_STIPULATE.details, order.retrieval_id.retrieval_destination_area)
        context.update({
            "warehouse": warehouse,
            "warehouse_filter": request.GET.get("warehouse_filter"),
            "order_type": order_type,
            "container_type": container_type,
            "reject_reason": order.invoice_reject_reason,
            "container_number": container_number,
            "is_leader": is_leader,
            "start_date": start_date,
            "end_date": end_date,
            "FS": FS,
            "receivable_is_locked": invoice.receivable_is_locked,
            "invoice_type": "receivable",
            "non_combina_reason": non_combina_reason,
            "fee_data": fee_data,
            "invoice_number": invoice.invoice_number,
            "invoice": invoice,  # 传递整个invoice对象
            "quotation_info": {
                "quotation_id": quotation.quotation_id,
                "version": quotation.version,
                "effective_date": quotation.effective_date,
                "is_user_exclusive": quotation.is_user_exclusive,
                "exclusive_user": quotation.exclusive_user,
                "filename": quotation.filename,  # 添加文件名
            },
            "pickup_fee": pickup_fee,
            "standard_fee_items": standard_fee_items,
            "existing_descriptions": existing_descriptions,  # 用于前端过滤
            "preport_status": invoice_status.preport_status,
            "combina_rules_text": rules_text,
            "is_combina":iscombina,
        })

        return self.template_preport_edit, context
    
    def  handle_container_warehouse_post(self, request:HttpRequest, context: dict|None=None) -> Dict[str, Any]:
        if not context:
            context = {}
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        container_number = request.GET.get("container_number")
        invoice_id = request.GET.get("invoice_id")
        delivery_type = request.GET.get("delivery_type")
        # 获取订单信息
        order = Order.objects.select_related(
            "retrieval_id", "container_number", "warehouse", "customer_name", "vessel_id"
        ).get(container_number__container_number=container_number)
        
        # 获取或创建账单和状态
        if invoice_id and invoice_id != 'None':
            #找到要修改的那份账单
            invoice = Invoicev2.objects.get(id=invoice_id)
            invoice_status, created = InvoiceStatusv2.objects.get_or_create(
                invoice=invoice,
                invoice_type="receivable",
                defaults={
                    "container_number": order.container_number,
                    "invoice": invoice,
                }
            )
        else:
            #说明这个柜子没有创建过账单，需要创建
            invoice, invoice_status = self._create_invoice_and_status(container_number)
        
        # 设置item_category
        item_category = f"warehouse_{delivery_type}"
        
        # 确定当前状态
        if invoice_status.finance_status == "completed":
            current_status = "confirmed"
        else:
            status_field = f"warehouse_{delivery_type}_status"
            current_status = getattr(invoice_status, status_field, 'unstarted')
        context.update({
            "invoice_number": invoice.invoice_number,
            "warehouse": order.retrieval_id.retrieval_destination_area,
            "warehouse_filter": request.GET.get("warehouse_filter"),
            "container_number": container_number,
            "receivable_is_locked": invoice.receivable_is_locked,
            "start_date": start_date,
            "end_date": end_date,
            "invoice": invoice,
            "status": current_status,
            "warehouse_status": current_status,  # 兼容性
            "invoice_type": "receivable",
            "delivery_type": delivery_type,
            "item_category": item_category,
            "receivable_is_locked": invoice.receivable_is_locked,     
        })
        # 获取报价表的相关信息
        quotation, quotation_error = self._get_quotation_for_order(order, 'receivable')
        if quotation_error:
            context.update({"error_messages": quotation_error})
            return context
        
        # 获取报价表中仓库费用详情
        fee_detail, fee_error = self._get_fee_details_from_quotation(quotation, "warehouse")
        if fee_error:
            context.update({"error_messages": fee_error})
            return context
        
        # 仓库费用项目列表
        standard_fee_items = [
            "分拣费", "拦截费", "亚马逊PO激活", "客户自提", "重新打板",
            "货品清点费", "仓租", "指定贴标", "内外箱", "托盘标签",
            "开封箱", "销毁", "拍照", "拍视频", "重复操作费"
        ]
        
        # 构建参考费用信息
        FS = {
            "分拣费": fee_detail.details.get("分拣费", "N/A"),
            "拦截费": fee_detail.details.get("拦截费", "N/A"),
            "亚马逊PO激活": fee_detail.details.get("亚马逊PO激活", "N/A"),
            "客户自提": fee_detail.details.get("客户自提", "N/A"),
            "重新打板": fee_detail.details.get("重新打板", "N/A"),
            "货品清点费": fee_detail.details.get("货品清点费", "N/A"),
            "仓租": fee_detail.details.get("仓租", "N/A"),
            "指定贴标": fee_detail.details.get("指定贴标", "N/A"),
            "内外箱": fee_detail.details.get("内外箱", "N/A"),
            "托盘标签": fee_detail.details.get("托盘标签", "N/A"),
            "开封箱": fee_detail.details.get("开封箱", "N/A"),
            "销毁": fee_detail.details.get("销毁", "N/A"),
            "拍照": fee_detail.details.get("拍照", "N/A"),
            "拍视频": fee_detail.details.get("拍视频", "N/A"),
            "重复操作费": fee_detail.details.get("重复操作费", "N/A"),
        }
        
        # 获取现有的费用项目
        existing_items = InvoiceItemv2.objects.filter(
            invoice_number=invoice,
            item_category=item_category
        )
        
        # 构建费用数据
        fee_data = []
        for item in existing_items:
            fee_data.append({
                'id': item.id,
                'description': item.description,
                'qty': item.qty or 0,
                'rate': item.rate or 0,
                'surcharges': item.surcharges or 0,
                'amount': item.amount or 0,
                'note': item.note or ''
            })
        # 如果是第一次录入且没有费用记录，添加所有标准费用项目为默认
        if not existing_items.exists():
            status_field = f"warehouse_{delivery_type}_status"
            current_status = getattr(invoice_status, status_field, 'unstarted')
            
            if current_status == 'unstarted':
                for fee_name in standard_fee_items:
                    ref_price = FS.get(fee_name, 0)
                    numeric_price = self._extract_number(str(ref_price)) if ref_price != "N/A" else 0
                    
                    fee_data.append({
                        'id': None,
                        'description': fee_name,
                        'qty': 0,
                        'rate': numeric_price,
                        'surcharges': 0,
                        'amount': 0,
                        'note': '',
                    })
        
        # 获取已存在的费用描述列表，用于前端过滤
        existing_descriptions = [item.description for item in existing_items]
        # 计算可用的标准费用项目（还没有被添加的）
        available_standard_items = [item for item in standard_fee_items if item not in existing_descriptions]
        

        context.update({          
            "FS": FS,             
            "fee_data": fee_data,         
            "quotation_info": {
                "quotation_id": quotation.quotation_id,
                "version": quotation.version,
                "effective_date": quotation.effective_date,
                "is_user_exclusive": quotation.is_user_exclusive,
                "exclusive_user": quotation.exclusive_user,
                "filename": quotation.filename,
            },
            "standard_fee_items": standard_fee_items,
            "available_standard_items": available_standard_items,
            "existing_descriptions": existing_descriptions,
            
        })
        
        return context

    def _determine_is_combina(self, order):
        is_combina = False
        if order.order_type != "转运组合":
            is_combina = False
        else:
            if order.container_number.manually_order_type == "转运组合":
                is_combina = True
            elif order.container_number.manually_order_type == "转运":
                is_combina = False
            else:
                # 未定义，直接去判断
                if self._is_combina(order.container_number.container_number):
                    is_combina = True
        return is_combina

    def _merge_combina_info(self, info1: dict, info2: dict , info3: dict = None) -> dict:
        """合并两个 combina_info（两个都可能为空）"""
        infos = [info for info in [info1, info2, info3] if info]
    
        if not infos:
            return {}
        
        # 初始化总和
        total_base_fee = 0.0
        total_cbm = 0.0
        total_weight = 0.0
        total_pallets = 0
        total_region_count = 0
        
        # 累加所有信息
        for info in infos:
            total_base_fee += float(info.get("base_fee", 0) or 0)
            total_cbm += float(info.get("total_cbm", 0) or 0)
            total_weight += float(info.get("total_weight", 0) or 0)
            total_pallets += int(info.get("total_pallets", 0) or 0)
            total_region_count += int(info.get("region_count", 0) or 0)
        
        result = {
            "base_fee": round(total_base_fee, 2),
            "total_cbm": round(total_cbm, 4),
            "region_count": total_region_count,
        }
        
        # 只有有值的字段才包含（保持向后兼容）
        if total_weight > 0:
            result["total_weight"] = round(total_weight, 4)
        if total_pallets > 0:
            result["total_pallets"] = total_pallets
        
        return result
    
    def handle_container_delivery_post(self, request:HttpRequest, context: dict| None = None) -> Dict[str, Any]:
        if not context:
            context = {}
        container_number = request.GET.get("container_number")
        delivery_type = request.GET.get("delivery_type", "public")
        current_user = request.user
        username = current_user.username
        if delivery_type == "public":
            template = self.template_delivery_public_edit
        else:
            template = self.template_delivery_other_edit
        
        invoice_id = request.GET.get("invoice_id")

        order = Order.objects.select_related(
            'container_number',
            'customer_name',
            'warehouse',
            'vessel_id',
            'retrieval_id'
        ).get(container_number__container_number=container_number)
         
        if invoice_id and invoice_id != "None": 
            #找到要修改的那份账单
            invoice = Invoicev2.objects.get(id=invoice_id)
            invoice_status, created = InvoiceStatusv2.objects.get_or_create(
                invoice=invoice,
                invoice_type="receivable",
                defaults={
                    "container_number": order.container_number,
                    "invoice": invoice,
                }
            )
        else:
            #说明这个柜子没有创建过账单，需要创建
            invoice, invoice_status = self._create_invoice_and_status(container_number)
            invoice_id = invoice.id

        previous_item_dict = {}
        #查看下之前有没有开过账单，之前记录给费用的仓点，这次就不再计费了
        other_invoices = Invoicev2.objects.filter(
            container_number=order.container_number 
        ).exclude(id=invoice_id)
        
        if other_invoices.exists(): 
            previous_items = InvoiceItemv2.objects.filter(
                container_number=order.container_number,  
                invoice_number__in=other_invoices,
                invoice_type="receivable",
            ).exclude(item_category="hold")
            # 按PO_ID建立索引
            
            for item in previous_items:
                if item.PO_ID:
                    previous_item_dict[item.PO_ID] = item

        # 获取板子数据
        pallet_groups, other_pallet_groups, ctx = self._get_pallet_groups_by_po(container_number, delivery_type, invoice)
        if ctx.get('error_messages'):
            return template, ctx
        
        #如果是公仓的，还有激活费，所以要把pallet_groups赋值出来再作为激活费的表格
        activation_table = None

        # 查看是不是组合柜
        is_combina = False
        if delivery_type == "public":
            activation_table = pallet_groups
            is_combina = self._determine_is_combina(order)

        # 获取本次账单已录入的激活费项
        activation_fee_groups = self._get_existing_activation_items(invoice, order.container_number)
        # 获取本次账单已录入的派送费项
        existing_items = self._get_existing_invoice_items(invoice, "delivery_" + delivery_type)

        # 如果所有PO都已录入，直接返回已有数据
        if existing_items:
            if delivery_type =="other":
                result_existing = self._separate_other_existing_items(invoice, pallet_groups)
                existing_keys = set(existing_items.keys())
                # 筛选未计费的分组
                unbilled_groups = []
                for g in pallet_groups:
                    po_id = g.get("PO_ID")
                    shipping_mark = g.get("shipping_marks", "")
                    
                    # 构建与_get_existing_invoice_items中相同的组合键
                    dict_key = f"{po_id}-{shipping_mark}"
                    
                    # 判断这个组合键是否在existing_items中
                    if dict_key not in existing_keys:
                        unbilled_groups.append(g)
            else:
                result_existing = self._separate_existing_items(existing_items, pallet_groups)        
                unbilled_groups = [g for g in pallet_groups if g.get("PO_ID") not in existing_items]
        else:
            result_existing = {
                "normal_items": [],
                "combina_groups": [],
                "combina_info": {}
            }
            unbilled_groups = pallet_groups

        # 再去除过去账单录过的派送费
        if previous_item_dict:
            result_previous_existing = self._set_free_charge_des(invoice, previous_item_dict, unbilled_groups, username)
            unbilled_groups = result_previous_existing['unbilled_groups']
        else:
            result_previous_existing = {
                "normal_items": [],
                "combina_groups": [],
                "combina_info": {}
            }
        if unbilled_groups:
            has_previous_items = bool(previous_item_dict)  # 判断是否有过账单
            # 有未录入的PO，需要进一步处理
            result_new = self._process_unbilled_items(
                pallet_groups=unbilled_groups,
                container=order.container_number,
                order=order,
                delivery_type=delivery_type,
                invoice=invoice,
                is_combina=is_combina,
                has_previous_items=has_previous_items
            )
            if isinstance(result_new, dict) and result_new.get('error_messages'):
                return template, result_new
            
            final_result = {
                "normal_items": result_existing["normal_items"] + result_new["normal_items"] + result_previous_existing["normal_items"],
                "combina_groups": result_existing["combina_groups"] + result_new["combina_groups"] + result_previous_existing["combina_groups"],
                "combina_info": self._merge_combina_info(result_existing["combina_info"],
                                                        result_new["combina_info"],
                                                        result_previous_existing["combina_info"])
            }
            
        else:
            final_result = result_existing

        # 报价表相关
        quotation, quotation_error = self._get_quotation_for_order(order, 'receivable')
        if quotation_error:
            context.update({"error_messages": quotation_error})
            return template, context
        try:
            COMBINA_STIPULATE = FeeDetail.objects.get(
                quotation_id=quotation.id,
                fee_type='COMBINA_STIPULATE'
            )
        except Exception as e:
            context.update({"error_messages": f'{quotation.filename}-{quotation.version}-缺少组合柜信息'})
            return template, context
        rules_text = self._parse_combina_rules(COMBINA_STIPULATE.details, order.retrieval_id.retrieval_destination_area)

        total_container_cbm = PackingList.objects.filter(
            container_number__container_number=container_number  
        ).aggregate(
            total_cbm=Sum('cbm')
        )['total_cbm'] or 0.0
        # 构建上下文
        context.update({
            "container_number": container_number,
            "container_type": order.container_number.container_type,
            "delivery_type": delivery_type,
            "order_type": order.order_type,
            "warehouse": order.warehouse.name if order.warehouse else "",
            "customer_name": order.customer_name.zem_name if order.customer_name else "",
            "manually_order_type": order.container_number.manually_order_type,
            # 分组数据
            "normal_items": final_result.get("normal_items", []),
            "combina_info": final_result.get("combina_info", {}),
            "combina_groups": final_result.get("combina_groups", []),
            "invoice_id": invoice.id,
            "is_combina": is_combina,
            "invoice_number": invoice.invoice_number,
            "delivery_method_options": DELIVERY_METHOD_OPTIONS,
            "other_pallet_groups": other_pallet_groups,
            "quotation_info": {
                "quotation_id": quotation.quotation_id,
                "version": quotation.version,
                "effective_date": quotation.effective_date,
                "is_user_exclusive": quotation.is_user_exclusive,
                "exclusive_user": quotation.exclusive_user,
                "filename": quotation.filename,  # 添加文件名
            },
            "combina_rules_text": rules_text,
            "warehouse_filter": request.GET.get("warehouse_filter"),
            "activation_fee_groups": activation_fee_groups,
            "activation_table": activation_table,
            "start_date": request.GET.get("start_date"),
            "end_date": request.GET.get("end_date"),
            "total_container_cbm": total_container_cbm,
            "receivable_is_locked": invoice.receivable_is_locked,
        })
        
        return template, context
    
    def _parse_combina_rules(self, rules_data: dict, warehouse_code: str):
        result = {
            "global_rules": "",           
            "warehouse_rules": "",
            "special_des_rules": "",
            "tiered_pricing": ""
        }

        g = rules_data.get("global_rules", {})

        # ---------- Global Rules ----------
        global_lines = []
        def get_rule(key):
            wh_key = f"{warehouse_code}_{key}"
            if wh_key in g:
                return g[wh_key]["default"]
            return g[key]["default"]

        global_lines.append(f"- 最大组合柜数量（区域）：{get_rule('max_mixed')}")
        global_lines.append(f"- 非组合柜区域最大数量：{get_rule('bulk_threshold')}")
        global_lines.append(f"- 40 尺标准板数：{get_rule('std_40ft_plts')}")
        global_lines.append(f"- 45 尺标准板数：{get_rule('std_45ft_plts')}")
        global_lines.append(f"- 标准每板 CBM：{g['cbm_per_pl']['default']}")
        global_lines.append(f"- 单柜限重：{g['weight_limit']['default']} 磅")
        global_lines.append(f"- 超重费区间：{g['overweight_min']['default']} - {g['overweight_max']['default']}")

        result["global_rules"] = "\n".join(global_lines)

        # ---------- Warehouse Pricing ----------
        wp = rules_data.get("warehouse_pricing", {})
        warehouse_lines = []
        if warehouse_code in wp:
            w = wp[warehouse_code]
            warehouse_lines.append(f"- 40 尺非组合柜提拆费：{w['nonmix_40ft']}")
            warehouse_lines.append(f"- 45 尺非组合柜提拆费：{w['nonmix_45ft']}")
            warehouse_lines.append(f"- 自提出库费（最低）：{w['pickup_min']}")
            warehouse_lines.append(f"- 自提出库费（最高）：{w['pickup_max']}")

        result["warehouse_rules"] = "\n".join(warehouse_lines)

        # ---------- 特别仓点（你数据结构里没有，先空） ----------
        dp = rules_data.get("special_warehouse", {})
        warehouse_lines = []
        if warehouse_code in dp:
            d = dp[warehouse_code]
            destinations_str = " - ".join(d["destination"]) 
            warehouse_lines.append(f"- 特殊仓点：{destinations_str}")
            warehouse_lines.append(f"- 倍数：{d['multiplier']}")
        result["special_des_rules"] = "\n".join(warehouse_lines)
        

        # ---------- Tiered Pricing ----------
        tp = rules_data.get("tiered_pricing", {})
        tier_lines = []
        if warehouse_code in tp:
            for item in tp[warehouse_code]:
                tier_lines.append(
                    f"- 仓点 {item['min_points']}~{item['max_points']} 个：加收 {item['fee']} 美元"
                )

        result["tiered_pricing"] = "\n".join(tier_lines)
        return result
    
    def _find_pallet_group_by_po_id(self, pallet_groups, target_po_id):
        """查找匹配的 pallet_group，对每组PO_ID逐步去除下划线后缀进行匹配"""
        for group in pallet_groups:
            if group.get("PO_ID") == target_po_id:
                return group
            
         # 如果没有精确匹配，尝试逐步去除下划线匹配
        for group in pallet_groups:
            group_po = group.get("PO_ID")

            current = group_po
            # 逐步去除下划线后缀
            while "_" in current:
                current = current.rsplit("_", 1)[0]
                if current == target_po_id:
                    return group
            
        return None

    def _set_free_charge_des(self, invoice: Invoicev2, old_item_dict, unbilled_groups, username):
        """之前账单记过费的，这次记为0"""
        combina_items = []
        normal_items = []
        combina_groups = []
        combina_info = {}
        combina_items_by_region = {}
        PO_ID = []
        total_amount = 0.0
        for po_id, existing_item in old_item_dict.items():
            pallet_group = self._find_pallet_group_by_po_id(unbilled_groups, po_id)
            if pallet_group:
                PO_ID.append(pallet_group.get("PO_ID"))
                item_category = "delivery_public" if pallet_group.get("delivery_type") == "public" else "delivery_other"

                total_amount += existing_item.amount
                item_data = {
                    "id": None,
                    "PO_ID": pallet_group.get("PO_ID"),       
                    "destination": pallet_group.get("destination", ""),
                    "zipcode": "",
                    "delivery_method": "",
                    "delivery_category": existing_item.delivery_type,

                    "total_pallets": existing_item.qty,
                    "total_cbm": existing_item.cbm,
                    "cbm_ratio": existing_item.cbm_ratio,
                    "total_weight_lbs": existing_item.weight,
                    "shipping_marks": pallet_group.get("shipping_marks", ""),
                    "pallet_ids": pallet_group.get("pallet_ids", []),
                    "rate": existing_item.rate,
                    "description": existing_item.description + " (已在前单计费)" if existing_item.description else "已在前单计费",
                    "surcharges": existing_item.surcharges,
                    "note": existing_item.note,
                    "amount": 0, 
                    "is_existing": False,  
                    "is_previous_existing": True,
                    "need_manual_input": False,
                    "is_hold": False,
                    "region": existing_item.region,
                    "registered_user": username,
                    "item_category": item_category, 
                }

                
                if existing_item.delivery_type == 'combine':
                    combina_items.append(item_data)
                    region = item_data.get("region", "未知")
                    combina_items_by_region.setdefault(region, []).append(item_data)
                else:
                    normal_items.append(item_data)
                
                unbilled_groups.remove(pallet_group)
        
        if combina_items:
            for region, items in combina_items_by_region.items():
                if items:
                    price = float(items[0].get("rate", 0))
                    total_cbm = sum(item.get("total_cbm", 0) for item in items)
                    total_cbm_ratio = round(sum(float(item.get("cbm_ratio") or 0) for item in items), 4)
                    
                    combina_groups.append({
                        "region": region,
                        "price": price,
                        "total_cbm": round(total_cbm, 4),
                        "total_cbm_ratio": total_cbm_ratio,
                        "destinations": list(set(item.get("destination", "") for item in items)),
                        "items": items
                    })
            total_base_fee = total_amount
            total_cbm = sum(item.get("total_cbm", 0) for item in combina_items)
            total_weight = sum(item.get("total_weight_lbs", 0) for item in combina_items)
            total_pallets = sum(item.get("total_pallets", 0) for item in combina_items)
            combina_info = {
                "base_fee": round(total_base_fee, 2),
                "total_cbm": round(total_cbm, 4),
                "total_weight": round(total_weight, 4),
                "total_pallets": total_pallets,
                "region_count": len(combina_groups)
            }
        return {
            "normal_items": normal_items,
            "combina_groups": combina_groups,
            "combina_info": combina_info,
            "unbilled_groups":unbilled_groups,

        }

    def _separate_other_existing_items(self, invoice, pallet_groups):
        '''私仓的派送已录入数据'''
        combina_groups = []
        combina_info = {}
        normal_items_list = list(
            InvoiceItemv2.objects
            .filter(
                invoice_number=invoice,
                item_category="delivery_other",
                invoice_type="receivable",
            )
            .annotate(destination=F("warehouse_code"))
            .annotate(total_cbm=F("cbm"))
            .annotate(total_weight_lbs=F("weight"))          
            .annotate(total_pallets=F("qty"))
            .annotate(delivery_category=F("delivery_type"))
            .annotate(is_existing=Value(True, output_field=BooleanField()))
        )

        return {
            "normal_items": normal_items_list,
            "combina_groups": combina_groups,
            "combina_info": combina_info
        }


    def _separate_existing_items(self, existing_items, pallet_groups):
        """将已有数据按组合柜和非组合柜分开"""
        combina_items = []
        normal_items = []
        combina_groups = []
        combina_info = {}
        
        # 按区域分组组合柜数据
        combina_items_by_region = {}
        
        combina_total_cbm = 0.0
        for po_id, existing_item in existing_items.items():
            # 找到对应的pallet组
            pallet_group = next((g for g in pallet_groups if g.get("PO_ID") == po_id), None)
            if pallet_group:
                item_data = self._create_item_from_existing(existing_item, pallet_group)
                # 根据类型分类
                if existing_item.delivery_type == 'combine':
                    combina_total_cbm += item_data.get("total_cbm", 0.0)
                    combina_items.append(item_data)
                    
                    # 按区域分组
                    region = item_data.get("region", "未知")
                    combina_items_by_region.setdefault(region, []).append(item_data)
                else:
                    normal_items.append(item_data)
        # 构建组合柜分组数据
        for region, items in combina_items_by_region.items():
            if items:
                price = float(items[0].get("rate", 0))
                total_cbm = sum(item.get("total_cbm", 0) for item in items)
                total_cbm_ratio = round(sum(float(item.get("cbm_ratio") or 0) for item in items), 4)
                region_price = round(total_cbm * price / combina_total_cbm, 3)

                combina_groups.append({
                    "region": region,
                    "price": price,
                    "total_cbm": round(total_cbm, 4),
                    "total_cbm_ratio": total_cbm_ratio,
                    "region_price": region_price,
                    "destinations": list(set(item.get("destination", "") for item in items)),
                    "items": items
                })
        
        # 计算组合柜总信息
        if combina_items:
            total_base_fee = sum(item.get("amount", 0) for item in combina_items)
            total_cbm = sum(item.get("total_cbm", 0) for item in combina_items)
            total_cbm_ratio = round(sum((item.get("cbm_ratio") or 0) for item in combina_items), 4)
            total_weight = sum(item.get("total_weight_lbs", 0) for item in combina_items)
            total_pallets = sum(item.get("total_pallets", 0) for item in combina_items)
            combina_info = {
                "base_fee": round(total_base_fee, 2),
                "total_cbm": round(total_cbm, 4),
                "total_cbm_ratio": total_cbm_ratio,
                "total_weight": round(total_weight, 4),
                "total_pallets": total_pallets,
                "region_count": len(combina_groups)
            }
        
        return {
            "normal_items": normal_items,
            "combina_groups": combina_groups,
            "combina_info": combina_info
        }

    def _process_unbilled_items(
        self,
        pallet_groups: List[Dict],
        container,
        order,
        delivery_type: str,
        invoice,
        is_combina,
        has_previous_items=False
    ) -> List[Dict[str, Any]]:
        """处理未录入费用的PO组"""
        result = {
            "normal_items": [],
            "combina_items": [], 
            "combina_items": [],
            "combina_groups": [],
            "combina_info": {},
        }
        # 按区域分组组合柜数据
        combina_items_by_region = {}
        fee_details = {}
        quotation_info = None

        total_combina_cbm = 0.0
        if delivery_type == "public":
            # 获取报价表，私仓不用找报价表
            quotations = self._get_fee_details(order, order.retrieval_id.retrieval_destination_area,order.customer_name.zem_name)
            if isinstance(quotations, dict) and quotations.get("error_messages"):
                return {"error_messages": quotations["error_messages"]}
            
            fee_details = quotations['fees']
            quotation_info = quotations['quotation']
            if is_combina:
                combina_result = self._process_combina_items_with_grouping(
                    pallet_groups=pallet_groups,
                    container=container,
                    order=order,
                    fee_details=fee_details,
                    has_previous_items=has_previous_items
                )

                if isinstance(combina_result, dict) and combina_result.get('error_messages'):
                    return combina_result
                
                result["combina_groups"] = combina_result.get("groups", [])
                # 合并组合柜数据
                new_items = combina_result.get("items", [])
                processed_po_ids = set(combina_result.get("processed_po_ids", []))
                result["combina_items"].extend(new_items)

                for item in new_items:
                    total_combina_cbm += item.get("total_cbm")
                    if item.get("PO_ID") in processed_po_ids:
                        region = item.get("combina_region", "未知")
                        if region not in combina_items_by_region:
                            combina_items_by_region[region] = []
                        combina_items_by_region[region].append(item)
                result["combina_info"] = combina_result.get("info", {})
                # 从待处理的pallet_groups中移除已处理的组合柜记录
                pallet_groups = [g for g in pallet_groups if g.get("PO_ID") not in processed_po_ids]

        # 处理未录入的PO
        if pallet_groups:
            for group in pallet_groups:
                destination = group.get("destination", "")
                location = group.get("location")
                
                if delivery_type == "public":
                    # 公仓：尝试自动计算费用
                    item_data = self._process_public_unbilled(
                        group=group,
                        container=container,
                        order=order,
                        destination=destination,
                        location=location,
                        fee_details=fee_details
                    )
                    
                else:
                    # 私仓：只确定类型，不创建记录
                    item_data = self._process_private_unbilled(
                        group=group,
                        invoice=invoice
                    )

                if isinstance(item_data, dict) and item_data.get("error_messages"):
                    if quotation_info:
                        extra = f"（报价表：{quotation_info.filename} v{quotation_info.version}）"
                        item_data["error_messages"] += extra
                        return item_data
                if not item_data:
                    continue
                # 如果是组合柜项目，添加到对应的分组
                result["normal_items"].append(item_data)

        # 建组合柜分组           
        # for region, items in combina_items_by_region.items():
        #     price = items[0].get("combina_price", 0)
        #     total_cbm = sum(item.get("total_cbm", 0) for item in items)
        #     region_cbm_price = round(float(price) * total_cbm / total_combina_cbm,3)
            
        #     result["combina_groups"].append({
        #         "region": region,
        #         "price": price,
        #         "region_price" : region_cbm_price,
        #         "total_cbm": round(total_cbm, 2),
        #         "destinations": list({item.get("destination") for item in items}),
        #         "items": items,
        #     })
        
        # 计算组合柜总信息（如果还没有计算过）
        if not result["combina_info"] and result["combina_items"]:
            total_base_fee = sum(item.get("amount", 0) for item in result["combina_items"])
            total_cbm = sum(item.get("total_cbm", 0) for item in result["combina_items"])
            
            result["combina_info"] = {
                "base_fee": round(total_base_fee, 2),
                "total_cbm": round(total_cbm, 2),
                "region_count": len(result["combina_groups"])
            }
        return result
    
    def _process_destination(self, destination_origin):
        """处理目的地字符串"""
        destination_origin = str(destination_origin)

        # 匹配模式：按"改"或"送"分割，分割符放在第一组的末尾
        if "改" in destination_origin or "送" in destination_origin:
            # 找到第一个"改"或"送"的位置
            first_change_pos = min(
                (destination_origin.find(char) for char in ["改", "送"] 
                if destination_origin.find(char) != -1),
                default=-1
            )
            
            if first_change_pos != -1:
                # 第一部分：到第一个"改"或"送"（包含分隔符）
                first_part = destination_origin[:first_change_pos + 1]
                # 第二部分：剩下的部分
                second_part = destination_origin[first_change_pos + 1:]
                
                # 处理第一部分：按"-"分割取后面的部分
                if "-" in first_part:
                    if first_part.upper().startswith("UPS-"):
                        first_result = first_part
                    else:
                        first_result = first_part.split("-", 1)[1]
                else:
                    first_result = first_part
                
                # 处理第二部分：按"-"分割取后面的部分
                if "-" in second_part:
                    if second_part.upper().startswith("UPS-"):
                        second_result = second_part
                    else:
                        second_result = second_part.split("-", 1)[1]
                else:
                    second_result = second_part
                
                return first_result, second_result
            else:
                raise ValueError(first_change_pos)
        
        # 如果不包含"改"或"送"或者没有找到
        # 只处理第二部分（假设第一部分为空）
        if "-" in destination_origin:
            if destination_origin.upper().startswith("UPS-"):
                second_result = destination_origin
            else:
                second_result = destination_origin.split("-", 1)[1]
            
        else:
            second_result = destination_origin
        
        return None, second_result

    def _process_destination_wlm(self,destination):
        """处理目的地字段"""
        if destination and '-' in destination:
            parts = destination.split('-')
            if len(parts) > 1:
                return parts[1]
        return destination
    
    def _process_combina_items_with_grouping(
        self,
        pallet_groups: List[Dict],
        container,
        order,
        fee_details,
        has_previous_items
    ) -> Dict:
        """处理组合柜区域的计费逻辑返回: (更新后的billing_items, 已处理的pallet_groups)"""
        warehouse = order.retrieval_id.retrieval_destination_area
        container_type_temp = 0 if "40" in container.container_type else 1
        
        # 1. 获取组合柜报价规则
        combina_key = f"{warehouse}_COMBINA"
        if combina_key not in fee_details:
            context = {
                "error_messages": f"未找到组合柜报价表规则 {combina_key}"
            }
            return (context, [])  # 返回错误，空列表
        
        rules = fee_details.get(combina_key).details

        # 2. 筛选出属于组合区域的pallet_groups
        combina_pallet_groups = []
        processed_po_ids = set()
        need_Additional_des = []
        poid_list = []

        for group in pallet_groups:
            po_id = group.get("PO_ID", "")
            destination_str = group.get("destination", "")
            poid_list.append(po_id)

            #改前和改后的
            destination_origin, destination = self._process_destination(destination_str)
            
            is_combina_origin = False
            if has_previous_items and destination_origin:
                #判断改之前是不是组合柜，如果是组合->非组合，要补收一份组合柜，非组合->组合，正常按组合收，            
                for region, region_data in rules.items():
                    for item in region_data:
                        normalized_locations = [loc.strip() for loc in item["location"] if loc]
                        if destination_origin in normalized_locations:
                            is_combina_origin = True
                            break
                    if is_combina_origin:
                        break
            
            # 检查是否属于组合区域
            is_combina_region = False
            for region, region_data in rules.items():
                for item in region_data:
                    normalized_locations = [loc.strip() for loc in item["location"] if loc]
                    if destination in normalized_locations:
                        is_combina_region = True
                        break
                if is_combina_region:
                    break
            if destination.upper() == "UPS":
                is_combina_region = False
            
            if is_combina_region:
                combina_pallet_groups.append(group)
                processed_po_ids.add(po_id)
            if is_combina_origin and not is_combina_region:
                # 如果是组合->非组合，要补收一份组合柜
                need_Additional_des.append(destination_str)
                combina_pallet_groups.append(group)
        # 如果没有组合区域，直接返回原数据和空列表，都按转运算
        if not combina_pallet_groups:
            return {"items": [], "info": {}}

        # 3. 计算组合柜每目的地 CBM（保留两位小数）
        combina_destinations_cbm = {}  # 记录每个目的地的CBM
        total_combina_pallets = 0
        
        # 因为组合柜区域每个仓点的cbm占比要看在整柜子的比例，所以这俩查询一遍总的cbm
        total_container_cbm_result = PackingList.objects.filter(
            container_number=container  # 使用container对象，或者container_number字符串
        ).aggregate(
            total_cbm=Sum('cbm')
        )
        total_container_cbm = round(total_container_cbm_result['total_cbm'] or 0.0, 2)

        for group in combina_pallet_groups:
            destination_str = group.get("destination", "")     
            destination = self._process_destination_wlm(destination_str)    
            cbm = round(group.get("total_cbm", 0), 2) 
            total_combina_pallets += group.get("total_pallets", 0)     
            
            # 记录每个目的地的CBM
            if destination in combina_destinations_cbm:
                combina_destinations_cbm[destination] += cbm
            else:
                combina_destinations_cbm[destination] = cbm
        
        # 4. 计算占比（保留四位小数）
        group_cbm_ratios = {}

        for g in combina_pallet_groups:
            po_id = g.get("PO_ID")
            destination_str = g.get("destination", "")
            destination = self._process_destination_wlm(destination_str)  
            key = (po_id, destination)

            cbm = round(g.get("total_cbm", 0), 2)
            if total_container_cbm > 0:
                
                group_cbm_ratios[key] = round(cbm / total_container_cbm, 4)

            else:
                group_cbm_ratios[key] = 0.0

        # 判断下如果所有仓点都是组合柜区域内，那就要保证总和为1
        unique_poids = set(poid_list)
        prefixes = {po_id.split('_')[0] for po_id in unique_poids if '_' in po_id}
        poid_list = list(unique_poids | prefixes)

        missing_records = PackingList.objects.filter(container_number=container).exclude(PO_ID__in=poid_list)

        has_missing = missing_records.exists()
        
        if not has_missing:
            #修正比例：保证总和 = 1.0000, 现在不按组合柜占比为1了，和其他仓点不好算
            ratio_sum = round(sum(group_cbm_ratios.values()), 4)
            if ratio_sum != 1.0:
                diff = round(1.0 - ratio_sum, 4)

                # 最大 CBM 仓点承担误差
                max_key = max(
                    group_cbm_ratios,
                    key=lambda k: next(
                        (
                            round(g.get("total_cbm", 0), 2)
                            for g in combina_pallet_groups
                            if (g.get("PO_ID"), g.get("destination", "")) == k
                        ),
                        0
                    )
                )
                group_cbm_ratios[max_key] = round(group_cbm_ratios[max_key] + diff, 4)

        # 5. 计算组合柜总费用
        combina_tiers_data = {}  # 存储结构为：(区名, 梯度索引) -> 数据
        # 记录每个具体的 PO+Dest 属于哪个梯度索引
        item_to_tier_map = {}

        for group in combina_pallet_groups:
            dest_str = group.get("destination", "")
            dest_fixed = self._process_destination_wlm(dest_str)
            po_id = group.get("PO_ID")
            
            match_found = False
            for region_name, region_list in rules.items():
                for idx, tier in enumerate(region_list):
                    if dest_fixed in tier["location"]:
                        # 找到了具体的区和该区下的价格梯度
                        tier_key = f"{region_name}_{idx}"
                        price = tier["prices"][container_type_temp]
                        
                        if tier_key not in combina_tiers_data:
                            combina_tiers_data[tier_key] = {
                                "region_name": region_name,
                                "price": price,
                                "total_cbm": 0,
                                "destinations": set()
                            }
                        
                        combina_tiers_data[tier_key]["total_cbm"] += round(group.get("total_cbm", 0), 2)
                        combina_tiers_data[tier_key]["destinations"].add(dest_fixed)
                        item_to_tier_map[(po_id, dest_fixed)] = tier_key
                        
                        match_found = True
                        break
                if match_found: break
                
        # 6. 计算组合柜总费用
        combina_base_fee = 0
        for key, data in combina_tiers_data.items():
            ratio = data["total_cbm"] / total_container_cbm if total_container_cbm > 0 else 0
            combina_base_fee += data["price"] * ratio

           
        # 7. 构建组合柜项目数据（按区域分组）
        combina_items = []
        region_groups_display = {}
        for group in combina_pallet_groups:
            dest_str = group.get("destination", "")
            dest_fixed = self._process_destination_wlm(dest_str)
            po_id = group.get("PO_ID")
            
            tier_key = item_to_tier_map.get((po_id, dest_fixed))
            if not tier_key: continue
            
            tier_info = combina_tiers_data[tier_key]
            price = tier_info["price"]
            region_name = tier_info["region_name"]
            
            # 计算金额
            cbm_ratio = group_cbm_ratios.get((po_id, dest_fixed), 0.0)
            amount = round(price * cbm_ratio, 2)

            if dest_str in need_Additional_des:
                description, amount = "由于组合转非组合，补交组合费用", 0 - amount
            else:
                description = ""

            item_data = {
                "id": None,
                "PO_ID": po_id,
                "destination": destination,
                "delivery_method": group.get("delivery_method", ""),
                "delivery_category": "combine",
                "total_pallets": group.get("total_pallets", 0),
                "total_cbm": round(cbm, 2),
                "total_weight_lbs": round(group.get("total_weight_lbs", 0), 2),
                "shipping_marks": group.get("shipping_marks", ""),
                "pallet_ids": group.get("pallet_ids", []),
                "rate": price,
                "description": description,
                "surcharges": 0,
                "note": "",
                "amount": amount,
                "is_existing": False,
                "is_previous_existing": False,
                "need_manual_input": False,
                "is_combina_item": True,
                "combina_region": region_name,
                "combina_price": price,
                "cbm_ratio": cbm_ratio,
            }
            combina_items.append(item_data)

            # 组织给前端展示用的 groups
            if tier_key not in region_groups_display:
                region_groups_display[tier_key] = {
                    "region": region_name,
                    "price": price,
                    "region_price": 0,
                    "total_cbm": 0,
                    "items": []
                }
            region_groups_display[tier_key]["items"].append(item_data)
            region_groups_display[tier_key]["region_price"] += amount
            region_groups_display[tier_key]["total_cbm"] += item_data["total_cbm"]
        
        return {
            "items": combina_items,
            "groups": list(region_groups_display.values()),
            "info": {
                "base_fee": round(combina_base_fee, 2),
                "total_cbm": total_container_cbm,
                "total_pallets": total_combina_pallets,
                "region_count": len(combina_tiers_data)
            },
            "processed_po_ids": list(processed_po_ids)
        }

    def _process_private_unbilled(
        self,
        group: Dict,
        invoice
    ) -> Dict[str, Any]:
        """处理私仓未录入的PO"""
        po_id = group.get("PO_ID")
        destination = group.get("destination", "")
        delivery_method = group.get("delivery_method", "")
        
        # 确定派送类型
        rate = None
        amount = None
        need_manual_input = True
        if "暂扣" in delivery_method:
            delivery_category = "hold"
            rate = 0
            amount = 0
            need_manual_input = False
        elif delivery_method and "客户自提" in delivery_method:
            delivery_category = "selfpickup"
        else:
            delivery_category = "selfdelivery"
        # 返回数据（不自动创建记录）
        return {
            "id": None,
            "PO_ID": po_id,
            "destination": destination,
            "zipcode": group.get("zipcode", ""),
            "delivery_method": delivery_method,
            "delivery_category": delivery_category,
            "total_pallets": group.get("total_pallets", 0),
            "total_cbm": group.get("total_cbm", 0),
            "total_weight_lbs": group.get("total_weight_lbs", 0),
            "shipping_marks": group.get("shipping_marks", ""),
            "pallet_ids": group.get("pallet_ids", []),
            "rate": rate,
            "description": '',
            "surcharges": 0,
            "note": "",
            "amount": amount,
            "is_existing": False,
            "is_previous_existing": False,
            "need_manual_input": need_manual_input,  # 私仓都需要手动录入
            "invoice_id": invoice.id,  # 记录invoice_id，用于后续创建
        }
    
    def _process_public_unbilled(
        self,
        group: Dict,
        container,
        order,
        destination,
        location,
        fee_details,
    ) -> Dict[str, Any]:
        """处理公仓未录入的PO"""
        context = {}
        po_id = group.get("PO_ID")
        delivery_method = group.get("delivery_method", "")
        warehouse = order.retrieval_id.retrieval_destination_area

        # 获取结果，如果为空则设置为0.0
        total_cbm = group.get("total_cbm")
        total_weight_lbs = group.get("total_weight_lbs")
        need_manual_input = False
        # 1. 确定派送类型
        total_pallets = group.get("total_pallets")  
        if delivery_method and any(courier in delivery_method.upper() 
                                 for courier in ["UPS", "FEDEX", "DHL", "DPD", "TNT"]):
            delivery_category = "upsdelivery"
            rate = 0
            amount = 0 
            need_manual_input = True   
        elif "暂扣" in delivery_method:
            delivery_category = "hold"
            rate = 0
            amount = 0
            need_manual_input = False             
        else:      
            if "准时达" in order.customer_name.zem_name:
                #准时达根据板子实际仓库找报价表，其他用户是根据建单
                warehouse = location.split('-')[0]

            #用转运方式计算费用
            public_key = f"{warehouse}_PUBLIC"
            if public_key not in fee_details:
                context.update({'error_messages':f'{warehouse}_PUBLIC-group-{group}未找到亚马逊沃尔玛报价表'})
                return context
            rules = fee_details.get(f"{warehouse}_PUBLIC").details
            niche_warehouse = fee_details.get(f"{warehouse}_PUBLIC").niche_warehouse
            if destination in niche_warehouse:
                is_niche_warehouse = True
            else:
                is_niche_warehouse = False
            #LA和其他的存储格式有点区别
            details = (
                {"LA_AMAZON": rules}
                if "LA" in warehouse and "LA_AMAZON" not in rules
                else rules
            )
            delivery_category = None
            rate_found = False
            for category, zones in details.items():
                for zone, locations in zones.items():
                    if destination in locations:
                        if "AMAZON" in category:
                            delivery_category = "amazon"
                            rate = zone
                            rate_found = True
                        elif "WALMART" in category:
                            delivery_category = "walmart"
                            rate = zone
                            rate_found = True
                if rate_found:
                    break
            if not rate_found:
                need_manual_input = True
                rate = 0
                amount = 0
                total_pallets = group.get("total_pallets")   
            else:            
                total_pallets = self._calculate_total_pallet(
                    total_cbm, True, is_niche_warehouse
                )               
                rate = float(rate) if rate else 0.0
                total_pallets = float(total_pallets) if total_pallets else 0.0
                amount = rate * total_pallets

        # 返回数据（不创建InvoiceItemv2记录）
        return {
            "id": None,
            "PO_ID": po_id,
            "destination": destination,
            "delivery_method": delivery_method,
            "delivery_category": delivery_category,
            "total_pallets": total_pallets,
            "total_cbm": total_cbm,
            "total_weight_lbs": total_weight_lbs,
            "shipping_marks": group.get("shipping_marks", ""),
            "pallet_ids": group.get("pallet_ids", []),
            "rate": rate,
            "description": '',
            "surcharges": 0,
            "note": "",
            "amount": amount,
            "is_existing": False,
            "is_previous_existing": False,
            "need_manual_input": need_manual_input,  
        }
    
    def _calculate_total_pallet(
        self, cbm: float, is_new_rule: bool, is_niche_warehouse: bool
    ) -> float:
        raw_p = float(cbm) / 1.8
        integer_part = int(raw_p)
        decimal_part = raw_p - integer_part
        # 本地派送的按照4.1之前的规则
        if decimal_part > 0:
            if is_new_rule:  # etd4.1之后的
                if is_niche_warehouse:
                    additional = 1 if decimal_part > 0.5 else 0.5
                else:
                    additional = 1 if decimal_part > 0.5 else 0
            else:
                if decimal_part > 0:
                    if is_niche_warehouse:
                        additional = 1
                    else:
                        additional = 1 if decimal_part > 0.9 else 0.5
            total_pallet = integer_part + additional
        elif decimal_part == 0:
            total_pallet = integer_part
        else:
            ValueError("板数计算错误")
        return total_pallet

    def _generate_items_from_existing_only(
        self,
        existing_items: Dict[str, Any],
        pallet_groups: List[Dict]
    ) -> List[Dict[str, Any]]:
        """从已有记录生成账单数据（全部已录入的情况）"""
        billing_items = []
        
        for po_id, existing_item in existing_items.items():
            # 找到对应的pallet组
            pallet_group = next((g for g in pallet_groups if g.get("PO_ID") == po_id), None)
            if pallet_group:
                item_data = self._create_item_from_existing(existing_item, pallet_group)
                billing_items.append(item_data)
        
        return billing_items
    
    def _create_item_from_existing(
        self,
        existing_item,
        pallet_group: Dict
    ) -> Dict[str, Any]:
        """从已有InvoiceItemv2记录创建账单数据"""
        is_hold = False
        if "暂扣" in pallet_group.get("delivery_method"):
            is_hold = True
        
        return {
            "id": existing_item.id,
            "PO_ID": existing_item.PO_ID,
            "destination": existing_item.warehouse_code or pallet_group.get("destination", ""),
            "zipcode": "",
            #"delivery_method": existing_item.delivery_method or pallet_group.get("delivery_method", ""),
            "delivery_method": "",
            "delivery_category": existing_item.delivery_type,
            "total_pallets": existing_item.qty,
            "total_cbm": existing_item.cbm,
            "cbm_ratio": existing_item.cbm_ratio,
            "total_weight_lbs": existing_item.weight,
            "shipping_marks": pallet_group.get("shipping_marks", ""),
            "pallet_ids": pallet_group.get("pallet_ids", []),
            "rate": existing_item.rate,
            "description": existing_item.description,
            "surcharges": existing_item.surcharges,
            "note": existing_item.note,
            "amount": existing_item.amount,
            "is_existing": True,
            "is_previous_existing": False,
            "need_manual_input": False,
            "is_hold": is_hold,
            "region": existing_item.region,
            "registered_user": existing_item.registered_user,
        }
    
    def _get_fee_details(self, order: Order, warehouse, customer_name) -> dict:
        context = {}
        quotation, quotation_error = self._get_quotation_for_order(order, 'receivable')
        if quotation_error:
            context.update({"error_messages": quotation_error})
            return context
        id = quotation.id
        if "准时达" in customer_name:
            #准时达的，如果转仓，要根据pallet实际仓库去计算报价
            fee_types = ["NJ_LOCAL", "NJ_PUBLIC", "NJ_COMBINA","SAV_PUBLIC", "SAV_COMBINA","LA_PUBLIC", "LA_COMBINA"]
        else:
            fee_types = {
                "NJ": ["NJ_LOCAL", "NJ_PUBLIC", "NJ_COMBINA"],
                "SAV": ["SAV_PUBLIC", "SAV_COMBINA"],
                "LA": ["LA_PUBLIC", "LA_COMBINA"],
            }.get(warehouse, [])

        return {
            "quotation": quotation,
            "fees": {
                fee.fee_type: fee
                for fee in FeeDetail.objects.filter(
                    quotation_id=id, fee_type__in=fee_types
                )
            }
        }
        
    def _get_pallet_groups_by_po(self, container_number: str, delivery_type: str, invoice: Invoicev2) -> tuple[list, list, dict]:
        """获取托盘数据"""
        context = {}
        error_messages = []
        base_query = Pallet.objects.filter(
            container_number__container_number=container_number,
            delivery_type=delivery_type
        ).exclude(
            PO_ID__isnull=True
        ).exclude(
            PO_ID=""
        )

        other_query = Pallet.objects.filter(
            container_number__container_number=container_number
        ).exclude(
            PO_ID__isnull=True
        ).exclude(
            PO_ID=""
        )

        group_fields = [
            "PO_ID",
            "destination",
            "zipcode",
            "delivery_method",
            "location",
            "delivery_type",
        ]

        if delivery_type == "other":
            group_fields.append("shipping_mark")

        # 按PO_ID分组统计
        pallet_groups = list(
            base_query.values(*group_fields)
            .annotate(
                total_pallets=models.Count("pallet_id"),
                total_cbm=models.Sum("cbm"),
                total_weight_lbs=models.Sum("weight_lbs"),
                pallet_ids=ArrayAgg("pallet_id"),
                shipping_marks=StringAgg("shipping_mark", delimiter=", ", distinct=True),
            ).order_by("PO_ID")
        )

        other_pallet_groups = list(
            other_query.values(*group_fields)
            .annotate(
                total_pallets=models.Count("pallet_id"),
                total_cbm=models.Sum("cbm"),
                total_weight_lbs=models.Sum("weight_lbs"),
                pallet_ids=ArrayAgg("pallet_id"),
                shipping_marks=StringAgg("shipping_mark", delimiter=", ", distinct=True),
            ).order_by("delivery_type")
        )
        if not pallet_groups:
            error_messages.append("未找到板子数据")
            context['error_messages'] = error_messages
            return [], [], context
        
        # 对每个PO组，从PackingList表中获取准确的CBM和重量数据
        for group in pallet_groups:
            po_id = group.get("PO_ID")
            shipping_marks = group.get("shipping_marks")
            if po_id:
                if delivery_type == "other":
                    try:
                        aggregated = PackingList.objects.filter(PO_ID=po_id,shipping_mark=shipping_marks).aggregate(
                            total_cbm=Sum('cbm'),
                            total_weight_lbs=Sum('total_weight_lbs')
                        )                      
                        
                        if aggregated['total_cbm'] is not None:
                            group['total_cbm'] = aggregated['total_cbm']
                        if aggregated['total_weight_lbs'] is not None:
                            group['total_weight_lbs'] = aggregated['total_weight_lbs']
                        
                    except Exception as e:
                        # 如果查询出错，不修改值
                        continue
                else:
                    
                    try:
                        aggregated = PackingList.objects.filter(PO_ID=po_id).aggregate(
                            total_cbm=Sum('cbm'),
                            total_weight_lbs=Sum('total_weight_lbs')
                        )
                        if aggregated['total_cbm'] is None and '_' in po_id and group.get('shipping_marks'):
                            # 去掉下划线再查，同时匹配 shipping_marks
                            po_id_modified = po_id.split('_', 1)[0]
                            aggregated = PackingList.objects.filter(
                                PO_ID=po_id_modified,
                                shipping_mark=shipping_marks
                            ).aggregate(
                                total_cbm=Sum('cbm'),
                                total_weight_lbs=Sum('total_weight_lbs')
                            )
                        if aggregated['total_cbm'] is not None:
                            group['total_cbm'] = aggregated['total_cbm']
                        if aggregated['total_weight_lbs'] is not None:
                            group['total_weight_lbs'] = aggregated['total_weight_lbs']
                        
                    except Exception as e:
                        # 如果查询出错，不修改值
                        continue
                    
            else:
                # 没有PO_ID的情况
                raise ValueError('pallet缺少PO_ID')
        
        if other_pallet_groups:
            container = Container.objects.get(container_number=container_number)
            # 1. 获取所有需要检查的 PO_ID
            po_ids = [group["PO_ID"] for group in other_pallet_groups]
            
            # 2. 查询 InvoiceItemv2 中相关的记录
            # 根据你的筛选条件构建查询
            invoice_items = InvoiceItemv2.objects.filter(
                container_number=container,  
                invoice_number=invoice,
                invoice_type="receivable",
                PO_ID__in=po_ids
            ).values(
                'id', 
                'PO_ID'
            )
            po_to_item_id = {item['PO_ID']: item['id'] for item in invoice_items}
            
            for group in other_pallet_groups:
                group['item_id'] = po_to_item_id.get(group["PO_ID"], None)

        if error_messages:
            context['error_messages'] = error_messages
        return pallet_groups, other_pallet_groups, context
    
    def _get_existing_activation_items(
        self,
        invoice,
        container
    ) -> Dict[str, Any]:
        activation_fee_items = InvoiceItemv2.objects.filter(
            container_number=container,
            invoice_number=invoice,
            invoice_type="receivable",
            delivery_type="activation", 
            item_category="activation_fee"
        )
        
        # 转换格式用于前端显示
        activation_fee_groups = []
        for item in activation_fee_items:
            activation_fee_groups.append({
                'id': item.id,
                'PO_ID': item.PO_ID,
                'destination': item.warehouse_code,
                'cbm': item.cbm or 0,
                'weight': item.weight or 0,
                'amount': item.amount or 0,
                'pallet': item.qty or 0,
                'note': item.note or '',
                'is_existing': True,
            })
        return activation_fee_groups

    def _get_existing_invoice_items(
            self,
            invoice,
            item_category: str
    ) -> Dict[str, Any]:
        """获取已存在的InvoiceItemv2记录，按PO_ID索引"""
        items = InvoiceItemv2.objects.filter(
            invoice_number=invoice,
            item_category=item_category,
            invoice_type="receivable"
        )

        items_without_po = []
        items_with_po = []
        for item in items:
            if item.PO_ID:
                items_with_po.append(item)
            else:
                items_without_po.append(item)

        if items_without_po:
            items_without_po = self._supplement_po_ids(invoice, items_without_po, items_with_po)

        items = InvoiceItemv2.objects.filter(
            invoice_number=invoice,
            item_category=item_category,
            invoice_type="receivable"
        )

        all_items = items_with_po + items_without_po
        # 按PO_ID建立索引
        item_dict = {}
        for item in all_items:
            if item.PO_ID:
                if item_category == "delivery_other":
                    # 对于delivery_other，使用"PO_ID-shipping_marks"作为键
                    shipping_mark = item.shipping_marks or ""  # 处理None值
                    dict_key = f"{item.PO_ID}-{shipping_mark}"
                else:
                    # 其他类别，只使用PO_ID作为键
                    dict_key = item.PO_ID
                
                item_dict[dict_key] = item

        return item_dict
    
    def _supplement_po_ids(self, invoice, items_without_po, items_with_po):
        """补充缺失的PO_ID"""
        container = invoice.container_number
        pallet_po_groups = (
            Pallet.objects
            .filter(container_number=container)
            .values("PO_ID", "destination")
            .annotate(
                cbm_sum=Sum("cbm"),
                weight_sum=Sum("weight_lbs"),
            )
        )
        # 2️⃣ destination -> [po_group]
        destination_to_po_map = defaultdict(list)
        for row in pallet_po_groups:
            destination_to_po_map[row["destination"]].append(row)
        
        # 3️⃣ items_without_po 按 destination 分组
        items_by_destination = defaultdict(list)
        for item in items_without_po:
            if item.warehouse_code:
                items_by_destination[item.warehouse_code].append(item)
        
        def po_group_distance(item, po_group):
            score = 0

            if item.cbm and po_group["cbm_sum"]:
                score += abs(item.cbm - po_group["cbm_sum"])

            if item.weight and po_group["weight_sum"]:
                score += abs(item.weight - po_group["weight_sum"])

            return score

         # 4️⃣ 记录被修改的 items
        updated_items = []

        # 5️⃣ 开始补 PO_ID
        for destination, items in items_by_destination.items():
            po_groups = destination_to_po_map.get(destination, [])

            if not po_groups:
                # pallet 里根本没有这个 destination，跳过
                continue

            # ✅ 情况一：destination 在 pallet 中唯一
            if len(po_groups) == 1:
                po_id = po_groups[0]["PO_ID"]
                for item in items:
                    item.PO_ID = po_id
                    updated_items.append(item)
                continue

            # ⚠️ 情况二：destination 对应多个 PO，用 cbm / weight 匹配
            for item in items:
                best_po = min(
                    po_groups,
                    key=lambda g: po_group_distance(item, g)
                )
                item.PO_ID = best_po["PO_ID"]
                updated_items.append(item)

        if updated_items:
            InvoiceItemv2.objects.bulk_update(updated_items, ["PO_ID"])
        return items_without_po

    def handle_invoice_warehouse_save(self, request:HttpRequest) -> Dict[str, Any]:
        """保存仓库账单"""
        context = {} 
        save_type = request.POST.get("save_type")       
        invoice_id = request.POST.get("invoice_id")
        delivery_type = request.POST.get("delivery_type")

        current_user = request.user
        username = current_user.username 
        #try:
        invoice = Invoicev2.objects.get(id=invoice_id)
        invoice_status = InvoiceStatusv2.objects.get(invoice=invoice, invoice_type="receivable")
        
        container_number = request.POST.get("container_number")
        order = Order.objects.select_related(
            "customer_name", "container_number"
        ).get(container_number__container_number=container_number)
        
        container = order.container_number  # 获取柜子对象
        container_delivery_type = getattr(container, 'delivery_type', 'mixed')
        # 设置item_category
        item_category = f"warehouse_{delivery_type}"
        
        # 费用详情
        fee_ids = request.POST.getlist("fee_id")
        descriptions = request.POST.getlist("fee_description")
        rates = request.POST.getlist("fee_rate")
        qtys = request.POST.getlist("fee_qty")
        surcharges = request.POST.getlist("fee_surcharges")
        notes = request.POST.getlist("fee_note")

        total_amount = Decimal("0.00")
        with transaction.atomic():
            # 找下仓库账单之前存的费用记录，和现在所有费用比较，差集就是前端删除的记录
            existing_items = InvoiceItemv2.objects.filter(
                invoice_number=invoice, 
                item_category=item_category
            )
            existing_ids = set(item.id for item in existing_items if item.id is not None)
            submitted_ids = set(int(fid) for fid in fee_ids if fid)  # 只包含已有的id
            to_delete_ids = existing_ids - submitted_ids
            if to_delete_ids:
                InvoiceItemv2.objects.filter(id__in=to_delete_ids).delete()

            for i in range(len(descriptions)):
                fee_id = fee_ids[i] or None
                description = descriptions[i]
                rate = Decimal(rates[i] or 0)
                qty = Decimal(qtys[i] or 0)
                surcharge = Decimal(surcharges[i] or 0)
                
                # 计算总价：总价 = 单价 * 数量 + 附加费
                amount = rate * qty + surcharge
                
                note = notes[i] or ""
                
                # 如果单价、数量和附加费都为0，则跳过
                if qty == 0 and surcharge == 0:
                    continue
                    
                total_amount += amount

                if fee_id:  # 已存在的费用项，更新
                    try:
                        item = InvoiceItemv2.objects.get(id=fee_id, invoice_number=invoice)
                        item.rate = rate
                        item.qty = qty
                        item.surcharges = surcharge
                        item.amount = amount
                        item.note = note
                        item.registered_user = username
                        item.save()
                    except InvoiceItemv2.DoesNotExist:
                        # 防止前端传了错误 id，查不到就新增
                        InvoiceItemv2.objects.create(
                            container_number=order.container_number,
                            invoice_number=invoice,
                            invoice_type="receivable",
                            item_category=item_category,
                            description=description,
                            rate=rate,
                            qty=qty,
                            surcharges=surcharge,
                            amount=amount,
                            note=note,
                            registered_user=username
                        )
                else:  # 新增费用项
                    InvoiceItemv2.objects.create(
                        container_number=order.container_number,
                        invoice_number=invoice,
                        invoice_type="receivable",
                        item_category=item_category,
                        description=description,
                        rate=rate,
                        qty=qty,
                        surcharges=surcharge,
                        amount=amount,
                        note=note,
                        registered_user=username
                    )

            # 更新账单总金额
            if delivery_type == "public":
                invoice.receivable_wh_public_amount = total_amount
                invoice_status.warehouse_public_reason = ''
            else:
                invoice.receivable_wh_other_amount = total_amount
                invoice_status.warehouse_self_reason = ''

            #计算总费用
            self._calculate_invoice_total_amount(invoice)

            # 更新仓库账单状态
            status_field = f"warehouse_{delivery_type}_status"
            setattr(invoice_status, status_field, save_type)

            if save_type == "rejected":
                reason_field = f"warehouse_{delivery_type}_reason"
                setattr(invoice_status, reason_field, request.POST.get("reject_reason", ""))
            
            # 根据柜子类型自动更新另一边的状态
            if delivery_type == "public" and container_delivery_type == "public":
                invoice_status.warehouse_other_status = "completed"
                invoice_status.delivery_other_status = "completed"

            elif delivery_type == "other" and container_delivery_type == "other":
                invoice_status.warehouse_public_status = "completed"
                invoice_status.delivery_public_status = "completed"

                
            invoice_status.save()
        delivery_type_chinese = "公仓" if delivery_type == "public" else "私仓"
        status_mapping = {
            'unstarted': '未录入',
            'in_progress': '录入中',
            'completed': '已完成',
            'rejected': '已拒绝'
        }
        status_chinese = status_mapping.get(save_type, '未知状态')
        success_msg = mark_safe(
            f"{container_number} 仓库账单保存成功！<br>"
            f"总费用: <strong>${total_amount:.2f}</strong><br>"
            f"类型: {delivery_type_chinese}<br>"
            f"状态更新为:{status_chinese}"
        )
        context["success_messages"] = success_msg
            
        return self.handle_warehouse_entry_post(request, context)

    def _calculate_invoice_total_amount(self, invoice:Invoicev2):
        def to_decimal(value, default='0.0'):
            """安全转换为 Decimal"""
            if value is None:
                return Decimal(default)
            if isinstance(value, Decimal):
                return value
            if isinstance(value, float):
                return Decimal(str(value))
            return Decimal(str(value))
        # 计算仓库总金额
        wh_public = to_decimal(invoice.receivable_wh_public_amount)
        wh_other = to_decimal(invoice.receivable_wh_other_amount)
        warehouse_total = wh_public + wh_other

        delivery_public = to_decimal(invoice.receivable_delivery_public_amount)
        delivery_other = to_decimal(invoice.receivable_delivery_other_amount)
        delivery_total = delivery_public + delivery_other

        preport_amount = to_decimal(invoice.receivable_preport_amount)
        # 更新总金额
        invoice.receivable_total_amount = preport_amount + warehouse_total + delivery_total      
        invoice.save()

    def handle_invoice_preport_save(self, request:HttpRequest) -> Dict[str, Any]:
        context = {} 
        save_type = request.POST.get("save_type")       
        invoice_id = request.POST.get("invoice_id")

        current_user = request.user 
        username = current_user.username 
 
        invoice = Invoicev2.objects.get(id=invoice_id)
        invoice_status = InvoiceStatusv2.objects.get(invoice=invoice, invoice_type="receivable")
        
        container_number = request.POST.get("container_number")
        order = Order.objects.select_related(
            "customer_name", "container_number"
        ).get(container_number__container_number=container_number)
        #费用详情
        fee_ids = request.POST.getlist("fee_id")
        descriptions = request.POST.getlist("fee_description")
        rates = request.POST.getlist("fee_rate")
        qtys = request.POST.getlist("fee_qty")
        surcharges = request.POST.getlist("fee_surcharges")
        notes = request.POST.getlist("fee_note")

        total_amount = Decimal("0.00")

        #找下港前账单之前存的费用记录，和现在所有费用比较，差集就是前端删除的记录
        existing_items = InvoiceItemv2.objects.filter(invoice_number=invoice, item_category="preport")
        existing_ids = set(item.id for item in existing_items if item.id is not None)
        submitted_ids = set(int(fid) for fid in fee_ids if fid)  # 只包含已有的id
        to_delete_ids = existing_ids - submitted_ids
        if to_delete_ids:
            InvoiceItemv2.objects.filter(id__in=to_delete_ids).delete()

        for i in range(len(descriptions)):
            fee_id = fee_ids[i] or None
            description = descriptions[i]
            rate = Decimal(rates[i] or 0)
            qty = Decimal(qtys[i] or 0)
            surcharge = Decimal(surcharges[i] or 0)

            amount = rate * qty + surcharge
            note = notes[i] or ""
            if qty == 0 and surcharge == 0:
                continue
            total_amount += amount

            if fee_id:  # 已存在的费用项，更新
                try:
                    item = InvoiceItemv2.objects.get(id=fee_id, invoice_number=invoice)
                    item.rate = rate
                    item.qty = qty
                    item.surcharges = surcharge
                    item.amount = amount
                    item.note = note
                    item.registered_user = username
                    item.save()
                except InvoiceItemv2.DoesNotExist:
                    # 防止前端传了错误 id，查不到就新增
                    InvoiceItemv2.objects.create(
                        container_number=order.container_number,
                        invoice_number=invoice,
                        invoice_type="receivable",
                        item_category="preport",
                        description=description,
                        rate=rate,
                        qty=qty,
                        surcharges=surcharge,
                        amount=amount,
                        note=note,
                        registered_user=username
                    )
            else:  # 新增费用项
                InvoiceItemv2.objects.create(
                    container_number=order.container_number,
                    invoice_number=invoice,
                    invoice_type="receivable",
                    item_category="preport",
                    description=description,
                    rate=rate,
                    qty=qty,
                    surcharges=surcharge,
                    amount=amount,
                    note=note,
                    registered_user=username
                )

            self._calculate_invoice_total_amount(invoice)
        
        # 更新港前账单状态
        invoice_status.preport_status = save_type
        if save_type == "rejected":
            invoice_status.preport_reason = request.POST.get("reject_reason", "")
        else:
            invoice_status.preport_reason = ''
        invoice_status.save()
        
        if order.order_type == "直送":
            invoice_status.warehouse_public_status = "completed"
            invoice_status.warehouse_other_status = "completed"
            invoice_status.delivery_public_status = "completed"
            invoice_status.delivery_other_status = "completed"
            invoice_status.save()
        status_mapping = {
            'pending_review': '待审核',
            'in_progress': '录入中',
            'completed': '已完成',
            'rejected': '已拒绝'
        }
        status_chinese = status_mapping.get(save_type, '未知状态')
        success_msg = mark_safe(
            f"{container_number} 柜号仓库账单保存成功！<br>"
            f"总费用: <strong>${total_amount:.2f}</strong><br>"
            f"状态更新为:{status_chinese}"
        )
        context["success_messages"] = success_msg
        
        return self.handle_preport_entry_post(request,context)

    def _extract_number(self, value):
        """从字符串里提取数字，失败则返回 0"""
        if value is None:
            return 0
        try:
            # 直接是数字
            return float(value)
        except:
            pass

        # 尝试从文本里提取数字
        match = re.search(r"[-+]?\d*\.?\d+", str(value))
        if match:
            try:
                return float(match.group())
            except:
                return 0

        return 0
        
    def query_invoice_basic(
        self,
        request,
        category: str,    
    ):
        """统一查询方法，用于五种账单类型"""
        
        warehouse = request.GET.get("warehouse", None)
        customer = request.GET.get("customer", None)

        status_field = self.CATEGORY_STATUS_FIELD[category]

        status_kwargs = {f"{status_field}__in": ["unstarted", "in_progress"]}

        # --- 基础过滤（Container） ---
        container_filter = Q()
        if warehouse:
            container_filter &= Q(retrieval_destination_precise=warehouse)

        # --- 所有应收账单对应的 InvoiceStatus ---
        qs = InvoiceStatusv2.objects.select_related(
            "invoice",
            "container_number",
            "invoice__customer",
        ).filter(
            invoice_type="receivable",
            container_number__in=Container.objects.filter(container_filter),
        )

        # =========================
        # 待录入（unstarted + in_progress）
        # =========================
        pending_input = qs.filter(**{status_field + "__in": ["unstarted", "in_progress"]})

        # =========================
        # 驳回
        # =========================
        rejected = qs.filter(**{status_field: "rejected"})

        # =========================
        # 待审核（pending_review）
        # =========================
        pending_review = qs.filter(**{status_field: "pending_review"})

        # =========================
        # 已完成（completed）
        # =========================
        completed = qs.filter(**{status_field: "completed"}).order_by("-invoice__invoice_date")

        context = {
            "pending_input": pending_input,
            "rejected": rejected,
            "pending_review": pending_review,
            "completed": completed,
            "warehouse_options": self.warehouse_options,
            "warehouse_filter": warehouse,
            "category": category,
        }

        return context
        
    def _create_invoice_and_status(self, container_number: str) -> tuple[Invoicev2, InvoiceStatusv2]:
        """创建账单和状态记录"""
        order = Order.objects.select_related(
            "customer_name", "container_number"
        ).get(container_number__container_number=container_number)
        # 创建 Invoicev2
        current_date = datetime.now().date()
        order_id = str(order.id)
        customer_id = order.customer_name.id

        # 先检查是否已经存在对应柜号的发票
        existing_invoice = Invoicev2.objects.filter(
            container_number=order.container_number
        ).first()
        
        if existing_invoice:
            # 如果发票已存在，检查对应的状态记录
            existing_status = InvoiceStatusv2.objects.filter(
                invoice=existing_invoice,
                invoice_type="receivable"
            ).first()
            
            if existing_status:
                # 两者都存在，直接返回
                return existing_invoice, existing_status
        
        invoice = Invoicev2.objects.create(
            container_number=order.container_number,
            invoice_number=f"{current_date.strftime('%Y-%m-%d').replace('-', '')}C{customer_id}{order_id}",
            created_at=current_date,
        )
        
        # 创建 InvoiceStatusv2
        invoice_status = InvoiceStatusv2.objects.create(
            container_number=order.container_number,
            invoice=invoice,
            invoice_type="receivable"
        )
        invoice.save()
        invoice_status.save()
        return invoice, invoice_status
    
    def _is_combina(self, container_number: str) -> Any:
        context = {}
        try:
            container = Container.objects.get(container_number=container_number)
        except Container.MultipleObjectsReturned:
            return False
        order = Order.objects.select_related(
            "retrieval_id", "container_number", "vessel_id"
        ).get(container_number__container_number=container_number)
        if order.order_type != "转运组合":
            return context, False, None
        if container.manually_order_type == "转运组合":
            return context, True, None 
        elif container.manually_order_type == "转运":
            return context, False, container.non_combina_reason
        
        customer = order.customer_name
        customer_name = customer.zem_name
        # 从报价表找+客服录的数据
        warehouse = order.retrieval_id.retrieval_destination_area
        vessel_etd = order.vessel_id.vessel_etd

        container_type = container.container_type
        #  基础数据统计
        plts = Pallet.objects.filter(
            container_number__container_number=container_number
        ).aggregate(
            unique_destinations=Count("destination", distinct=True),
            total_weight=Sum("weight_lbs"),
            total_cbm=Sum("cbm"),
            total_pallets=Count("id"),
        )
        plts["total_cbm"] = round(plts["total_cbm"], 2)
        plts["total_weight"] = round(plts["total_weight"], 2)
        # 获取匹配的报价表
        matching_quotation = (
            QuotationMaster.objects.filter(
                effective_date__lte=vessel_etd,
                is_user_exclusive=True,
                exclusive_user=customer_name,
                quote_type='receivable',
            )
            .order_by("-effective_date")
            .first()
        )
        if not matching_quotation:
            matching_quotation = (
                QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=False,  # 非用户专属的通用报价单
                    quote_type='receivable',
                )
                .order_by("-effective_date")
                .first()
            )
        if not matching_quotation:
            context.update({"error_messages": f"找不到{container_number}可用的报价表！"})
            return context, None, None
        # 获取组合柜规则
        try:
            stipulate_fee_detail = FeeDetail.objects.get(
                quotation_id=matching_quotation.id, fee_type="COMBINA_STIPULATE"
            )
            stipulate = stipulate_fee_detail.details
        except FeeDetail.DoesNotExist:
            context.update({
                "error_messages": f"报价表《{matching_quotation.filename}》-{matching_quotation.id}中找不到<报价表规则>分表，请截此图给技术员！"
            })
            return context, None, None
        
        combina_fee = FeeDetail.objects.get(
            quotation_id=matching_quotation.id, fee_type=f"{warehouse}_COMBINA"
        ).details
        if isinstance(combina_fee, str):
            combina_fee = json.loads(combina_fee)

        # 看是否超出组合柜限定仓点,NJ/SAV是14个
        warehouse_specific_key = f'{warehouse}_max_mixed'
        if warehouse_specific_key in stipulate.get("global_rules", {}):
            combina_threshold = stipulate["global_rules"][warehouse_specific_key]["default"]
        else:
            combina_threshold = stipulate["global_rules"]["max_mixed"]["default"]

        warehouse_specific_key1 = f'{warehouse}_bulk_threshold'
        if warehouse_specific_key1 in stipulate.get("global_rules", {}):
            uncombina_threshold = stipulate["global_rules"][warehouse_specific_key1]["default"]
        else:
            uncombina_threshold = stipulate["global_rules"]["bulk_threshold"]["default"]

        if plts["unique_destinations"] > uncombina_threshold:
            container.account_order_type = "转运"
            container.non_combina_reason = (
                f"总仓点超过{uncombina_threshold}个"
            )
            container.save()
            return context, False, f"总仓点超过{uncombina_threshold}个" # 不是组合柜

        # 按区域统计
        destinations = (
            Pallet.objects.filter(container_number__container_number=container_number)
            .values_list("destination", flat=True)
            .distinct()
        )
        plts_by_destination = (
            Pallet.objects.filter(container_number__container_number=container_number)
            .values("destination")
            .annotate(total_cbm=Sum("cbm"))
        )
        total_cbm_sum = sum(item["total_cbm"] for item in plts_by_destination)
        # 区分组合柜区域和非组合柜区域
        container_type_temp = 0 if "40" in container_type else 1
        matched_regions = self.find_matching_regions(
            plts_by_destination, combina_fee, container_type_temp, total_cbm_sum, combina_threshold
        )
        # 判断是否混区，False表示满足混区条件
        is_mix = self.is_mixed_region(
            matched_regions["matching_regions"], warehouse, vessel_etd
        )
        if is_mix:
            container.account_order_type = "转运"
            container.non_combina_reason = "混区不符合标准"
            container.save()
            return context, False, "混区不符合标准"
        
        filtered_non_destinations = [key for key in matched_regions["non_combina_dests"].keys() if "UPS" not in key]
        # 非组合柜区域
        non_combina_region_count = len(filtered_non_destinations)
        # 组合柜区域
        combina_region_count = len(matched_regions["combina_dests"])

        filtered_destinations = self._filter_ups_destinations(destinations)
        if combina_region_count + non_combina_region_count != len(filtered_destinations):
            raise ValueError(
                f"计算组合柜和非组合柜区域有误\n"
                f"组合柜目的地：{matched_regions['combina_dests']}，数量：{combina_region_count}\n"
                f"非组合柜目的地：{filtered_non_destinations}，数量：{non_combina_region_count}\n"
                f"目的地集合：{filtered_destinations}\n"
                f"目的地总数：{len(filtered_destinations)}"
            )
        sum_region_count = non_combina_region_count + combina_region_count
        if sum_region_count > uncombina_threshold:
            # 当非组合柜的区域数量超出时，不能按转运组合
            container.account_order_type = "转运"
            container.non_combina_reason = f"总区数量为{sum_region_count},要求是{uncombina_threshold}"
            container.save()
            return context, False,f"总区数量为{sum_region_count},要求是{uncombina_threshold}"
        container.non_combina_reason = None
        container.account_order_type = "转运组合"
        container.save()
        return context, True, None


    def is_mixed_region(self, matched_regions, warehouse, vessel_etd) -> bool:
        regions = list(matched_regions.keys())
        # LA仓库的特殊规则：CDEF区不能混
        if warehouse == "LA":
            if vessel_etd.month > 7 or (
                vessel_etd.month == 7 and vessel_etd.day >= 15
            ):  # 715之后没有混区限制
                return False
            if len(regions) <= 1:  # 只有一个区，就没有混区的情况
                return False
            if set(regions) == {"A区", "B区"}:  # 如果只有A区和B区，也满足混区规则
                return False
            return True
        # 其他仓库无限制
        return False
         
    def _filter_ups_destinations(self, destinations):
        """过滤掉包含UPS的目的地，支持列表和QuerySet"""
        if hasattr(destinations, '__iter__') and not isinstance(destinations, (str, dict)):
            destinations_list = list(destinations)
        else:
            destinations_list = destinations
        filtered_destinations = [dest for dest in destinations_list if 'UPS' not in str(dest) and 'FEDEX' not in str(dest)]
        return filtered_destinations
    
    def find_matching_regions(
        self,
        plts_by_destination: dict,
        combina_fee: dict,
        container_type,
        total_cbm_sum: FloatField,
        combina_threshold: int,
    ) -> dict:
        matching_regions = defaultdict(float)  # 各区的cbm总和
        des_match_quote = {}  # 各仓点的匹配详情
        destination_matches = set()  # 组合柜的仓点
        non_combina_dests = {}  # 非组合柜的仓点
        price_display = defaultdict(
            lambda: {"price": 0.0, "location": set()}
        )  # 各区的价格和仓点
        dest_cbm_list = []  # 临时存储初筛组合柜内的cbm和匹配信息

        region_counter = {}
        region_price_map = {}
        for plts in plts_by_destination:
            destination = plts["destination"]
            if ('UPS' in destination) or ('FEDEX' in destination):
                continue
            # 如果是沃尔玛的，只保留后面的名字，因为报价表里就是这么保留的
            clean_dest = destination.replace("沃尔玛", "").strip()

            if clean_dest.upper().startswith("UPS-"):
                dest = clean_dest
            else:
                dest = clean_dest.split("-")[-1].strip()

            cbm = plts["total_cbm"]
            dest_matches = []
            matched = False
            # 遍历所有区域和location
            for region, fee_data_list in combina_fee.items():           
                for fee_data in fee_data_list:
                    prices_obj = fee_data["prices"]
                    price = self._extract_price(prices_obj, container_type)
                    
                    # 如果匹配到组合柜仓点，就登记到组合柜集合中
                    if dest in fee_data["location"]:
                        # 初始化
                        if region not in region_price_map:
                            region_price_map[region] = [price]
                            region_counter[region] = 0
                            actual_region = region
                        else:
                            # 如果该 region 下已有相同价格 → 不加编号
                            found = None
                            for r_key, r_val in price_display.items():
                                if r_key.startswith(region) and r_val["price"] == price:
                                    found = r_key
                                    break
                            if found:
                                actual_region = found
                            else:                                
                                # 新价格 → 需要编号
                                region_counter[region] += 1
                                actual_region = f"{region}{region_counter[region]}"
                                region_price_map[region].append(price)

                        temp_cbm = matching_regions.get(actual_region, 0) + cbm
                        matching_regions[actual_region] = temp_cbm
                        dest_matches.append(
                            {
                                "region": actual_region,
                                "location": dest,
                                "prices": fee_data["prices"],
                                "cbm": cbm,
                            }
                        )
                        if actual_region not in price_display:
                            price_display[actual_region] = {
                                "price": price,
                                "location": set([dest]),
                            }
                        else:
                            # 不要覆盖，更新集合
                            price_display[actual_region]["location"].add(dest)
                        matched = True
            
            if not matched:
                # 非组合柜仓点
                non_combina_dests[dest] = {"cbm": cbm}
            # 记录匹配结果
            if dest_matches:
                des_match_quote[dest] = dest_matches
                # 将组合柜内的记录下来，后续方便按照cbm排序
                dest_cbm_list.append(
                    {"dest": dest, "cbm": cbm, "matches": dest_matches}
                )
                destination_matches.add(dest)
        if len(destination_matches) > combina_threshold:
            # 按cbm降序排序，将cbm大的归到非组合
            sorted_dests = sorted(dest_cbm_list, key=lambda x: x["cbm"], reverse=True)
            # 重新将排序后的前12个加入里面
            destination_matches = set()
            matching_regions = defaultdict(float)
            price_display = defaultdict(lambda: {"price": 0.0, "location": set()})
            for item in sorted_dests[:combina_threshold]:
                dest = item["dest"]
                destination_matches.add(dest)

                # 重新计算各区域的CBM总和
                for match in item["matches"]:
                    region = match["region"]
                    matching_regions[region] += item["cbm"]
                    price_display[region]["price"] = self._extract_price(match["prices"], container_type)
                    
                    price_display[region]["location"].add(dest)

            # 其余仓点转为非组合柜
            for item in sorted_dests[combina_threshold:]:
                non_combina_dests[item["dest"]] = {"cbm": item["cbm"]}
                # 将cbm大的从组合柜集合中删除
                des_match_quote.pop(item["dest"], None)

        # 下面开始计算组合柜和非组合柜各仓点占总体积的比例
        total_ratio = 0.0
        ratio_info = []

        # 处理组合柜仓点的cbm_ratio
        for dest, matches in des_match_quote.items():
            cbm = matches[0]["cbm"]  # 同一个dest的cbm在所有matches中相同
            ratio = round(cbm / total_cbm_sum, 4)
            total_ratio += ratio
            ratio_info.append((dest, ratio, cbm, True))  # 最后一个参数表示是否是组合柜
            for match in matches:
                match["cbm_ratio"] = ratio

        # 处理非组合柜仓点的cbm_ratio
        for dest, data in non_combina_dests.items():
            cbm = data["cbm"]
            ratio = round(cbm / total_cbm_sum, 4)
            total_ratio += ratio
            ratio_info.append((dest, ratio, cbm, False))
            data["cbm_ratio"] = ratio

        # 处理四舍五入导致的误差
        if abs(total_ratio - 1.0) > 0.0001:  # 考虑浮点数精度
            # 找到CBM最大的仓点
            ratio_info.sort(key=lambda x: x[2], reverse=True)
            largest_dest, largest_ratio, largest_cbm, is_combi = ratio_info[0]

            # 调整最大的仓点的ratio
            diff = 1.0 - total_ratio
            if is_combi:
                for match in des_match_quote[largest_dest]:
                    match["cbm_ratio"] = round(match["cbm_ratio"] + diff, 4)
            else:
                non_combina_dests[largest_dest]["cbm_ratio"] = round(
                    non_combina_dests[largest_dest]["cbm_ratio"] + diff, 4
                )
        return {
            "des_match_quote": des_match_quote,
            "matching_regions": matching_regions,
            "combina_dests": destination_matches,
            "non_combina_dests": non_combina_dests,
            "price_display": price_display,
        }
    
    def _extract_price(self, prices_obj, container_type):
        """
        安全地从 prices_obj 中提取数值 price：
        - 如果 prices_obj 是 dict，按键取（container_type 可为字符串或整型）。
        - 如果是 list/tuple，且 container_type 是 int，则尝试取 prices_obj[container_type]。
        若越界或该项不是数值，则回退到列表中第一个数值项。
        - 如果是单值（int/float），直接返回。
        - 其它情况返回 None。
        """
        # 优先处理 dict
        if isinstance(prices_obj, dict):
            # 允许 container_type 是 str 或 int（int 转为索引的情况不常见）
            val = prices_obj.get(container_type)
            if isinstance(val, (int, float)):
                return val
            # 如果取到的不是数字，尝试找 dict 的第一个数字值作为回退
            for v in prices_obj.values():
                if isinstance(v, (int, float)):
                    return v
            return None

        # list/tuple 按 index 选
        if isinstance(prices_obj, (list, tuple)):
            # 当 container_type 是整数索引时，优先使用该索引
            if isinstance(container_type, int):
                try:
                    candidate = prices_obj[container_type]
                    if isinstance(candidate, (int, float)):
                        return candidate
                except Exception:
                    pass
            # 回退：选第一个数字项
            first_num = next((x for x in prices_obj if isinstance(x, (int, float))), None)
            return first_num

        # 直接是数字
        if isinstance(prices_obj, (int, float)):
            return prices_obj

        # 其他（字符串等），不能作为 price
        return None
    
    def _get_quotation_for_order(self, order: Order, quote_type: str = 'receivable') :
        """获取订单对应的报价表"""
        try:
            vessel_etd = order.vessel_id.vessel_etd
            
            
            customer = order.customer_name
            customer_name = customer.zem_name
            
            # 先查找用户专属报价表
            quotation = (
                QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=True,
                    exclusive_user=customer_name,
                    quote_type=quote_type,
                )
                .order_by("-effective_date")
                .first()
            )
            
            if not quotation:
                # 查找通用报价表
                quotation = (
                    QuotationMaster.objects.filter(
                        effective_date__lte=vessel_etd,
                        is_user_exclusive=False,
                        quote_type=quote_type,
                    )
                    .order_by("-effective_date")
                    .first()
                )
            
            if quotation:
                return quotation, None
            else:
                error_msg = f"找不到生效日期在{vessel_etd}之前的{quote_type}报价表"
                return None, error_msg
                
        except Exception as e:
            error_msg = f"查询报价表时发生错误: {str(e)}"
            return None, error_msg

    def _get_fee_details_from_quotation(self, quotation: QuotationMaster, fee_type: str = "preport") :
        """从报价表中获取费用详情"""
        try:
            fee_detail = FeeDetail.objects.get(
                quotation_id=quotation.id,
                fee_type=fee_type
            )
            return fee_detail, None
        except FeeDetail.DoesNotExist:
            error_msg = f"报价表中找不到{fee_type}类型的费用详情"
            return None, error_msg
        except Exception as e:
            error_msg = f"获取费用详情时发生错误: {str(e)}"
            return None, error_msg
    
    def _combina_get_extra_fees(self, invoice):
        """获取额外费用（从InvoiceItemv2表中获取港前、仓库、派送费用）"""
        extra_fees_items = InvoiceItemv2.objects.filter(
            invoice_number=invoice,
            invoice_type="receivable"
        ).exclude(
            delivery_type="combine"  # 排除组合柜
        ).exclude(
            delivery_type="hold"     # 排除暂扣
        ).order_by('item_category', 'id')
        
        # 转换为前端需要的格式
        extra_fees = []
        category_display = dict(InvoiceItemv2.ITEM_CATEGORY_CHOICES)
        
        for item in extra_fees_items:
            fee_data = {
                'id': item.id,
                'item_category': item.item_category,
                'item_category_display': category_display.get(item.item_category, item.item_category),
                'description': item.description or '',
                'warehouse_code': item.warehouse_code or '',
                'delivery_type': item.delivery_type or '',
                'qty': float(item.qty or 0),
                'rate': float(item.rate or 0),
                'cbm': float(item.cbm or 0),
                'weight': float(item.weight or 0),
                'surcharges': float(item.surcharges or 0),
                'note': item.note or '',
                'amount': float(item.amount or 0),
                'registered_user': item.registered_user or '',
                'PO_ID': item.PO_ID or '',
            }
            extra_fees.append(fee_data)
        
        return extra_fees

    def _filter_ups_destinations(self, destinations):
        """过滤掉包含UPS的目的地，支持列表和QuerySet"""
        if hasattr(destinations, '__iter__') and not isinstance(destinations, (str, dict)):
            destinations_list = list(destinations)
        else:
            destinations_list = destinations
        filtered_destinations = [
            dest.strip() for dest in destinations_list 
            if dest is not None 
            and 'UPS' not in str(dest).upper() 
            and 'FEDEX' not in str(dest).upper()
        ]
        return list(dict.fromkeys(filtered_destinations))
        
    def _calculate_delivery_fee_cost(
        self,
        fee_details: dict,
        warehouse: str,
        plts_by_destination: list,
        is_new_rule: bool,
        over_count: float,
    ) -> str:
        is_niche_warehouse = True
        if "LA" in warehouse:
            amazon_data = fee_details.get(f"{warehouse}_PUBLIC").details
            walmart_data = None
        else:
            amazon_data = fee_details.get(f"{warehouse}_PUBLIC").details.get(
                f"{warehouse}_AMAZON"
            )
            walmart_data = fee_details.get(f"{warehouse}_PUBLIC").details.get(
                f"{warehouse}_WALMART"
            )
        if warehouse == "NJ":
            local_data = fee_details.get("NJ_LOCAL").details
        for pl in plts_by_destination:
            # 从亚马逊、沃尔玛、本地报价表中挨个找
            # 先找亚马逊
            for price, locations in amazon_data.items():
                if pl["destination"] in locations:
                    pl["price"] = price
                    break
            if pl["price"]:  # 说明这个仓点在这个报价表里，确实是不是冷门仓点
                niche_warehouse = fee_details.get(f"{warehouse}_PUBLIC").niche_warehouse
                if pl["destination"] in niche_warehouse:
                    is_niche_warehouse = True
                else:
                    is_niche_warehouse = False

            # 再找沃尔玛
            if not pl["price"] and walmart_data:
                for price, locations in walmart_data.items():
                    if str(pl["destination"]).upper() in [
                        loc.upper() for loc in locations
                    ]:
                        pl["price"] = price
                        break
                if pl["price"]:
                    niche_warehouse = fee_details.get(
                        f"{warehouse}_PUBLIC"
                    ).niche_warehouse
                    if pl["destination"] in niche_warehouse:
                        is_niche_warehouse = True
                    else:
                        is_niche_warehouse = False
            # 再找本地派送
            if not pl["price"] and warehouse == "NJ":
                destination = re.sub(r"[^0-9]", "", str(pl["destination"]))
                for price, locations in local_data.items():
                    if str(destination) in map(str, locations["zipcodes"]):
                        pl["is_fixed_price"] = (
                            True  # 表示一口价，等会就不会再乘以板数了
                        )
                        costs = locations["prices"]
                        if not over_count:
                            total_pallet = self._calculate_total_pallet(
                                pl["total_cbm"], is_new_rule, is_niche_warehouse
                            )
                        else:
                            total_pallet = over_count
                        if total_pallet <= 5:
                            pl["price"] = int(costs[0])
                        elif total_pallet >= 5:
                            pl["price"] = int(costs[1])
                        break
            if not pl["price"]:
                pl["price"] = 0
            if not over_count:
                total_pallet = self._calculate_total_pallet(
                    pl["total_cbm"], is_new_rule, is_niche_warehouse
                )
            else:
                total_pallet = over_count
            pl["total_pallet"] = total_pallet
        return plts_by_destination
    
    def _get_combina_fee_details(self, warehouse: str, vessel_etd, customer_name: str) -> dict:
        if not vessel_etd:
            context = {"error_messages": '缺少ETD时间'}
            return context
        try:
            quotation = (
                QuotationMaster.objects.filter(
                    effective_date__lte=vessel_etd,
                    is_user_exclusive=True,
                    exclusive_user=customer_name,
                    quote_type='receivable',
                )
                .order_by("-effective_date")
                .first()
            )
            if not quotation:
                quotation = (
                    QuotationMaster.objects.filter(
                        effective_date__lte=vessel_etd,
                        is_user_exclusive=False,  # 非用户专属的通用报价单
                        quote_type='receivable',
                    )
                    .order_by("-effective_date")
                    .first()
                )
            if not quotation:
                context = {"error_messages": '找不到报价表'}
                return context
            id = quotation.id
            fee_types = {
                "NJ": ["NJ_LOCAL", "NJ_PUBLIC", "NJ_COMBINA"],
                "SAV": ["SAV_PUBLIC", "SAV_COMBINA"],
                "LA": ["LA_PUBLIC", "LA_COMBINA"],
            }.get(warehouse, [])

            return {
                fee.fee_type: fee
                for fee in FeeDetail.objects.filter(
                    quotation_id=id, fee_type__in=fee_types
                )
            }
        except QuotationMaster.DoesNotExist:
            context = {"error_messages": '没有找到有效的报价表'}
            return context
    
    def _get_combina_base_fee(self, invoice, container):
        existing_items = InvoiceItemv2.objects.filter(
            container_number=container,
            invoice_number=invoice,
            item_category="delivery_public",
            invoice_type="receivable",
            delivery_type="combine"
        )
        
        if not existing_items:
            return 0.0, 0.0, []
        
        base_fee = 0.0
        combina_total_cbm = 0.0
        combina_items_by_region = {}

        for existing_item in existing_items:
            base_fee += existing_item.amount
            combina_total_cbm += existing_item.cbm

            item_data = {
                "id": existing_item.id,
                "PO_ID": existing_item.PO_ID,
                "destination": existing_item.warehouse_code,
                "zipcode": "",
                "delivery_method": "",
                "delivery_category": existing_item.delivery_type,
                "total_pallets": existing_item.qty,
                "total_cbm": existing_item.cbm,
                "total_weight_lbs": existing_item.weight,
                "rate": existing_item.rate,
                "description": existing_item.description,
                "surcharges": existing_item.surcharges,
                "note": existing_item.note,
                "amount": existing_item.amount,
                "is_existing": True,
                "is_previous_existing": False,
                "need_manual_input": False,
                "region": existing_item.region,
            }
            
            # 按区域分组
            region = existing_item.region
            combina_items_by_region.setdefault(region, []).append(item_data)

        combina_total_cbm = round(combina_total_cbm,4)
        combina_groups = []
        # 构建组合柜分组数据
        for region, items in combina_items_by_region.items():
            if items:
                price = float(items[0].get("rate", 0))
                total_cbm = sum(item.get("total_cbm", 0) for item in items)
                region_price = round(total_cbm * price / combina_total_cbm, 3)

                combina_groups.append({
                    "region": region,
                    "price": price,
                    "total_cbm": round(total_cbm, 4),
                    "region_price": region_price,
                    "destinations": list(set(item.get("destination", "") for item in items)),
                    "items": items
                })
        return base_fee, combina_total_cbm, combina_groups

    def handle_container_invoice_combina_get(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        container_number = request.GET.get("container_number")
        container = Container.objects.get(container_number=container_number)
        order = Order.objects.select_related(
            "retrieval_id", "container_number", "vessel_id"
        ).get(container_number__container_number=container_number)

        invoice_id = request.GET.get("invoice_id")
        if invoice_id and invoice_id != 'None':
            #找到要修改的那份账单
            invoice = Invoicev2.objects.get(id=invoice_id)
            invoice_status, created = InvoiceStatusv2.objects.get_or_create(
                invoice=invoice,
                invoice_type="receivable",
                defaults={
                    "container_number": order.container_number,
                    "invoice": invoice,
                }
            )
        else:
            #说明这个柜子没有创建过账单，需要创建
            invoice, invoice_status = self._create_invoice_and_status(container_number)

        context = {
            "invoice_number": invoice.invoice_number,
            "container_number": container_number,
            "start_date": request.GET.get("start_date"),
            "end_date": request.GET.get("end_date"),
        }
        # 查看是不是财务未确认状态，未确认就从报价表找+客服录的数据，确认了就从invoice_item表找
        if invoice_status.finance_status == "completed":
            invoice_item = InvoiceItemv2.objects.filter(
                invoice_number__invoice_number=invoice.invoice_number,
                invoice_type="receivable",
            )
            context["invoice"] = invoice
            context["invoice_item"] = invoice_item
            return context
        
        # 从报价表找+客服录的数据
        warehouse = order.retrieval_id.retrieval_destination_area
        vessel_etd = order.vessel_id.vessel_etd

        container_type = container.container_type
        # 1. 基础数据统计
        plts = Pallet.objects.filter(
            container_number__container_number=container_number
        ).aggregate(
            unique_destinations=Count("destination", distinct=True),
            total_weight=Sum("weight_lbs"),
            total_cbm=Sum("cbm"),
            total_pallets=Count("id"),
        )
        plts["total_cbm"] = round(plts["total_cbm"], 2)
        plts["total_weight"] = round(plts["total_weight"], 2)
        # 3. 获取匹配的报价表
        customer = order.customer_name
        customer_name = customer.zem_name
        order_type = order.order_type

        matching_quotation, quotation_error = self._get_quotation_for_order(order, 'receivable')
        if quotation_error:
            context.update({"error_messages": quotation_error})
            return context
        
        # 4. 获取费用规则
        combina_fee = FeeDetail.objects.get(
            quotation_id=matching_quotation.id, fee_type=f"{warehouse}_COMBINA"
        ).details
        customer = order.customer_name
        customer_name = customer.zem_name
        

        fee_details = self._get_combina_fee_details(warehouse, vessel_etd, customer_name)
        if isinstance(fee_details, dict) and fee_details.get("error_messages"):
            return {"error_messages": fee_details["error_messages"]}
        
        stipulate = FeeDetail.objects.get(
            quotation_id=matching_quotation.id, fee_type="COMBINA_STIPULATE"
        ).details
        if isinstance(combina_fee, str):
            combina_fee = json.loads(combina_fee)
        
        ctx, is_combina, non_combina_reason = self._is_combina(order.container_number.container_number)
        if ctx.get('error_messages'):
            context.update({"error_messages": ctx.get('error_messages')})
            return context
        
        #检查是不是组合柜
        if not is_combina:
            context.update({"error_messages": f'不满足组合柜，原因是{non_combina_reason}'})
            return context
        
        # 2. 检查基本条件
        if plts["unique_destinations"] == 0:
            context.update({"error_messages": '未录入拆柜数据'})
            return context

        warehouse_specific_key = f'{warehouse}_max_mixed'
        if warehouse_specific_key in stipulate.get("global_rules", {}):
            combina_threshold = stipulate["global_rules"][warehouse_specific_key]["default"]
        else:
            combina_threshold = stipulate["global_rules"]["max_mixed"]["default"]

        warehouse_specific_key1 = f'{warehouse}_bulk_threshold'
        if warehouse_specific_key1 in stipulate.get("global_rules", {}):
            uncombina_threshold = stipulate["global_rules"][warehouse_specific_key1]["default"]
        else:
            uncombina_threshold = stipulate["global_rules"]["bulk_threshold"]["default"]


        if (
            plts["unique_destinations"]
            > uncombina_threshold
        ):
            container.account_order_type = "转运"
            container.save()
            context["reason"] = (
                f"超过{uncombina_threshold}个仓点"
            )
            return context

        # 按区域统计
        destinations = (
            Pallet.objects.filter(container_number__container_number=container_number)
            .values_list("destination", flat=True)
            .distinct()
        )
        plts_by_destination = (
            Pallet.objects.filter(container_number__container_number=container_number)
            .values("destination")
            .annotate(total_cbm=Sum("cbm"))
        )
        pl_cbm_by_destination = (
            PackingList.objects
            .filter(container_number__container_number=container_number)
            .values("destination")
            .annotate(total_cbm=Sum("cbm"))
        )
        cbm_map = {
            item["destination"]: item["total_cbm"] or 0
            for item in pl_cbm_by_destination
        }

        for item in plts_by_destination:
            item["total_cbm"] = cbm_map.get(item["destination"], 0)
        # 这里之前是
        total_cbm_sum = sum(item["total_cbm"] for item in plts_by_destination)
        # 区分组合柜区域和非组合柜区域
        container_type_temp = 0 if "40" in container_type else 1
        matched_regions = self.find_matching_regions(
            plts_by_destination, combina_fee, container_type_temp, total_cbm_sum, combina_threshold
        )
        
        # 非组合柜区域
        filtered_non_destinations = [
            key for key in matched_regions["non_combina_dests"].keys() 
            if "UPS" not in key and "FEDEX" not in key
        ]
        temp_non_combina_region_count = len(filtered_non_destinations)
        non_combina_region_count = len(matched_regions["non_combina_dests"])
        # 组合柜区域
        combina_region_count = len(matched_regions["combina_dests"])

        filtered_destinations = self._filter_ups_destinations(destinations)
        if combina_region_count + temp_non_combina_region_count != len(filtered_destinations):
            raise ValueError(
                f"计算组合柜和非组合柜区域有误\n"
                f"组合柜目的地：{matched_regions['combina_dests']}，数量：{combina_region_count}\n"
                f"非组合柜目的地：{filtered_non_destinations}，数量：{temp_non_combina_region_count}\n"
                f"目的地集合：{filtered_destinations}\n"
                f"目的地总数：{len(filtered_destinations)}"
            )

        non_combina_cbm_ratio = round(
            sum(
                data["cbm_ratio"]
                for data in matched_regions["non_combina_dests"].values()
            ),
            4,
        )
        non_combina_cbm = round(
            sum(data["cbm"] for data in matched_regions["non_combina_dests"].values()),
            4,
        )
        sum_region_count = combina_region_count + non_combina_region_count
        if combina_region_count > combina_threshold or sum_region_count > uncombina_threshold :
            container.account_order_type = "转运"
            container.save()
            if combina_region_count > combina_threshold:
                # reason = '不满足组合柜区域要求'
                reason = f"规定{combina_threshold}组合柜区,但实际有{combina_region_count}个:matched_regions['combina_dests']，所以按照转运方式统计价格"
            elif sum_region_count > uncombina_threshold:
                reason = f"规定共{uncombina_threshold}个区，但是有组合柜{combina_threshold}个区，有非组合柜{non_combina_region_count}个：{list(matched_regions['non_combina_dests'].keys())}，所以按照转运方式统计价格"
                # reason = '不满足组合柜区域要求'
            actual_fees = self._combina_get_extra_fees(invoice)
            context["reason"] = reason
            context["extra_fees"] = actual_fees
            return context
        
        #组合柜的固定费用
        pallet_groups, other_pallet_groups, ctx = self._get_pallet_groups_by_po(container_number, "public", invoice)
        existing_items = self._get_existing_invoice_items(invoice, "delivery_public")

        # 如果所有PO都已录入，直接返回已有数据
        if existing_items:
            result_existing = self._separate_existing_items(existing_items, pallet_groups)
            combina_groups = result_existing['combina_groups']
            base_fee = result_existing['combina_info']['base_fee']        
            combina_total_cbm = result_existing['combina_info']['total_cbm']
            combina_total_cbm_ratio = result_existing['combina_info']['total_cbm_ratio']
            combina_total_weight = result_existing['combina_info']['total_weight']
            combina_total_pallets = result_existing['combina_info']['total_pallets']
        else:
            context.update({"error_messages": '操作组未录入组合柜费用'})
            return context
        # 7.2 计算基础费用
        extra_fees = {
            "overweight": 0,
            "overpallets": 0,
            "overregion_pickup": 0,
            "overregion_delivery": 0,
        }

        # 7.3 检查超限情况
        # 超重检查
        if plts["total_weight"] > stipulate["global_rules"]["weight_limit"]["default"]:
            extra_fees["overweight"] = "需人工录入"  # 实际业务中应有默认费率

        # 超板检查——确定上限的板数
        if "40" in container_type:
            std_plt = stipulate["global_rules"]["std_40ft_plts"]
            if warehouse == "LA" and "LA_std_40ft_plts" in stipulate["global_rules"]:
                std_plt = stipulate["global_rules"]["LA_std_40ft_plts"]
        else:
            std_plt = stipulate["global_rules"]["std_45ft_plts"]
            if warehouse == "LA" and "LA_std_45ft_plts" in stipulate["global_rules"]:
                std_plt = stipulate["global_rules"]["LA_std_45ft_plts"]
        exceptions_dict = std_plt["exceptions"]

        max_pallets = 0
        for exception_des, exception_plt in exceptions_dict.items():
            exception_warehouse = exception_des.split("_")[0]
            exception_regions = list(exception_des.split("_")[1])
            # 检查是否匹配当前 warehouse 和 region
            if exception_warehouse == warehouse and any(
                region in exception_regions
                for region, value in matched_regions["matching_regions"].items()
            ):
                max_pallets = exception_plt
                break
        if max_pallets == 0:
            max_pallets = std_plt["default"]
        # 处理超的板数
        # 先计算实际板数
        total_pallets = math.ceil(plts["total_cbm"] / 1.8)  # 取上限
        if total_pallets > max_pallets:
            over_count = total_pallets - max_pallets
        else:
            over_count = 0
        # 找每个仓点的单价，倒序排序，方便计算超板的（没有超板的也要查，可能前端会改板数）
        plts_by_destination = (
            Pallet.objects.filter(container_number__container_number=container_number)
            .values("destination")
            .annotate(
                price=Value(None, output_field=models.FloatField()),
                is_fixed_price=Value(False, output_field=BooleanField()),
                total_pallet=Count("id", output_field=FloatField()),
            )
        )  # 形如{'destination': 'A', 'total_cbm': 10.5，'price':31.5,'is_fixed_price':True},
        #重新去预报里查找cbm和重量
        pl_cbm_by_destination = (
            PackingList.objects
            .filter(container_number__container_number=container_number)
            .values("destination")
            .annotate(total_cbm=Sum("cbm"))
        )
        cbm_map = {
            item["destination"]: item["total_cbm"]
            for item in pl_cbm_by_destination
        }

        for item in plts_by_destination:
            item["total_cbm"] = cbm_map.get(item["destination"], 0)
        
        plts_by_destination = self._calculate_delivery_fee_cost(
            fee_details, warehouse, plts_by_destination, destinations, over_count
        )
        max_price = 0
        max_single_price = 0
        for plt_d in plts_by_destination:
            if plt_d["is_fixed_price"]:  # 一口价的不用乘板数
                max_price = max(float(plt_d["price"]), max_price)
                max_single_price = max(max_price, max_single_price)
            else:
                max_price = max(float(plt_d["price"]) * over_count, max_price)
                max_single_price = max(float(plt_d["price"]), max_single_price)
        extra_fees["overpallets"] = max_price

        # 计算非组合柜费用的提拆费和派送费
        if non_combina_region_count:
            # 提拆费，要计算下非组合柜区域占当前柜子的cbm比例*对应的提拆费
            container_type = order.container_number.container_type
            match = re.match(r"\d+", container_type)
            if match:
                pick_subkey = match.group()
                # 这个提拆费是从组合柜规则的warehouse_pricing的nonmix_40ft 45ft取
                c_type = f"nonmix_{pick_subkey}ft"
                try:
                    pickup_fee = stipulate["warehouse_pricing"][warehouse][c_type]
                except KeyError as e:
                    error_msg = f"缺少{pick_subkey}柜型的报价配置"
                    raise ValueError(error_msg)

            extra_fees["overregion_pickup"] = round(non_combina_cbm_ratio * pickup_fee, 3)
            # 派送费
            for item in matched_regions["non_combina_dests"]:
                # 计算改区域的板数
                plts_by_destination_overregion = (
                    Pallet.objects.filter(
                        container_number__container_number=container_number,
                        destination__in=matched_regions["non_combina_dests"].keys(),
                    )
                    .values("destination")
                    .annotate(
                        total_pallet=Count("id", output_field=FloatField()),
                        price=Value(None, output_field=models.FloatField()),
                        is_fixed_price=Value(False, output_field=BooleanField()),
                    )
                )
                pl_stats_by_destination = (
                    PackingList.objects
                    .filter(
                        container_number__container_number=container_number,
                        destination__in=matched_regions["non_combina_dests"].keys(),
                    )
                    .values("destination")
                    .annotate(
                        total_cbm=Sum("cbm"),
                        total_weight=Sum("total_weight_lbs"),
                    )
                )
                pl_stat_map = {
                    item["destination"]: {
                        "total_cbm": item["total_cbm"] or 0,
                        "total_weight": item["total_weight"] or 0,
                    }
                    for item in pl_stats_by_destination
                }

                for item in plts_by_destination_overregion:
                    stats = pl_stat_map.get(item["destination"], {})
                    item["total_cbm"] = stats.get("total_cbm", 0)
                    item["total_weight"] = stats.get("total_weight", 0)

                plts_by_destination_overregion = self._calculate_delivery_fee_cost(
                    fee_details,
                    warehouse,
                    plts_by_destination_overregion,
                    destinations,
                    None,
                )

                sum_price = 0
                for plt_d in plts_by_destination_overregion:
                    if plt_d["is_fixed_price"]:  # 一口价的不用乘板数
                        sum_price += float(plt_d["price"])
                    else:
                        sum_price += float(plt_d["price"]) * plt_d["total_pallet"]
            extra_fees["overregion_delivery"] = sum_price
        else:
            pickup_fee = 0
        # 超仓点的加收费用
        addition_fee = 0
        if "tiered_pricing" in stipulate:
            region_count = combina_region_count + non_combina_region_count
            if warehouse in stipulate["tiered_pricing"]:
                for rule in stipulate["tiered_pricing"][warehouse]:
                    min_points = rule.get("min_points")
                    max_points = rule.get("max_points")
                    if int(min_points) <= region_count <= int(max_points):
                        addition_fee = {
                            "min_points": int(min_points),
                            "max_points": int(max_points),
                            "add_fee": rule.get("fee"),
                        }
        else:
            addition_fee = None

        display_data = {
            # 基础信息
            "plts_by_destination": plts_by_destination,
            "container_info": {
                "number": container_number,
                "type": container_type,
                "warehouse": warehouse,
            },
            # 组合柜信息
            "combina_data": {
                "base_fee": base_fee, 
                "combina_total_cbm": combina_total_cbm, 
                "combina_total_cbm_ratio": combina_total_cbm_ratio,
                "combina_total_weight": combina_total_weight, 
                "combina_total_pallets": combina_total_pallets,
                "combina_groups": combina_groups
            },
            # 超限费用
            "extra_fees": {
                "overweight": {
                    "is_over": plts["total_weight"]
                    > stipulate["global_rules"]["weight_limit"]["default"],
                    "current_weight": plts["total_weight"],
                    "limit_weight": stipulate["global_rules"]["weight_limit"][
                        "default"
                    ],
                    "extra_weight": round(
                        plts["total_weight"]
                        - stipulate["global_rules"]["weight_limit"]["default"],
                        2,
                    ),
                    "fee": extra_fees["overweight"],
                    "input_field": True,  # 显示输入框
                },
                "overpallets": {
                    "is_over": total_pallets > max_pallets,
                    "current_pallets": total_pallets,
                    "limit_pallets": max_pallets,
                    "over_count": over_count if total_pallets > max_pallets else 0,
                    "pallet_details": [],
                    "max_price_used": None,
                    "fee": extra_fees["overpallets"],
                },
                "overregion": {
                    "pickup": {
                        "non_combina_cbm": non_combina_cbm,
                        "total_cbm": plts["total_cbm"],
                        "ratio": non_combina_cbm_ratio * 100,
                        "base_fee": pickup_fee,
                        "fee": extra_fees["overregion_pickup"],
                    },
                    "delivery": {
                        "fee": extra_fees["overregion_delivery"],
                        "details": [],
                    },
                },
                "addition_fee": addition_fee,
            },
        }

        # 填充超板费详细信息，不超板也要展示详情，因为前端可以修改超的板数
        # if total_pallets > max_pallets:
        for plt in plts_by_destination:
            display_data["extra_fees"]["overpallets"]["pallet_details"].append(
                {
                    "destination": plt["destination"],
                    "price": plt["price"],
                    "is_fixed_price": plt["is_fixed_price"],
                    "is_max_used": float(plt["price"])
                    == max_single_price,  # 标记是否被采用
                }
            )
        display_data["extra_fees"]["overpallets"]["max_price_used"] = max_price

        # 填充超区派送费详细信息
        if non_combina_region_count:
            is_overregion = True
            for plt in plts_by_destination_overregion:
                display_data["extra_fees"]["overregion"]["delivery"]["details"].append(
                    {
                        "destination": plt["destination"],
                        "pallets": plt["total_pallet"],
                        "price": plt["price"],
                        "cbm": round(plt["total_cbm"], 2),
                        "subtotal": float(plt["price"]) * plt["total_pallet"],
                    }
                )
        else:
            is_overregion = False
        total_amount = (
            base_fee
            + extra_fees["overpallets"]
            + extra_fees["overregion_pickup"]
            + extra_fees["overregion_delivery"]
        )
        # 港前-仓库-派送录入的费用显示到界面上
        actual_fees = self._combina_get_extra_fees(invoice)

        # 8. 返回结果
        context.update(
            {
                "display_data": display_data,
                "total_amount": total_amount,
                "invoice_number": invoice.invoice_number,
                "container_number": container_number,
                "is_overregion": is_overregion,
                "extra_fees": actual_fees,
                "destination_matches": matched_regions["combina_dests"],
                "non_combina_dests": list(matched_regions["non_combina_dests"].keys()),
                "container_type_temp": container_type_temp,
                "container_type": container_type,
                "order_type": order_type,
                "quotation_info": {
                    "quotation_id": matching_quotation.quotation_id,
                    "version": matching_quotation.version,
                    "effective_date": matching_quotation.effective_date,
                    "is_user_exclusive": matching_quotation.is_user_exclusive,
                    "exclusive_user": matching_quotation.exclusive_user,
                    "filename": matching_quotation.filename,  # 添加文件名
                },
                "start_date": request.GET.get("start_date"),
                "end_date": request.GET.get("end_date"),
                "is_combina": is_combina,            
            }
        )
        return context

    def handle_invoice_confirm_combina_save(
        self, request: HttpRequest
    ) -> tuple[Any, Any]:
        '''组合柜账单，财务保存'''

        container_number = request.POST.get("container_number")
        invoice_number = request.POST.get("invoice_number")
        invoice = Invoicev2.objects.get(invoice_number=invoice_number)
        overweight_fee = float(request.POST.get("overweight_fee", 0))
        overpallet_fee = float(request.POST.get("overpallet_fee", 0))
        overregion_pickup_fee = float(request.POST.get("overregion_pickup_fee", 0))
        overregion_delivery_fee = float(request.POST.get("overregion_delivery_fee", 0))
        addition_fee_str = request.POST.get("addition_fee")
        addition_fee = float(addition_fee_str) if addition_fee_str else 0

        delete_data = request.POST.get('delete_records', '{}')
        delete_dict = json.loads(delete_data)
        extra_fee_ids = delete_dict.get('extra_fee', [])
        if extra_fee_ids:
            InvoiceItemv2.objects.filter(id__in=extra_fee_ids, invoice_number=invoice).delete()
        
        total_fee = 0.0

        # 处理要更新的逻辑
        items_data_json = request.POST.get("extra_items_data", "[]")
        if items_data_json:
            items_data = json.loads(items_data_json)
            username = request.user.username
            for item in items_data:
                total_fee += item.get('amount', 0)
                item_id = item.get('id')
                if not item_id:
                    continue
                    
                # 找到记录并更新
                InvoiceItemv2.objects.filter(id=item_id, invoice_number=invoice).update(
                    item_category=item.get('item_category'),
                    description=item.get('description'),
                    warehouse_code=item.get('warehouse_code'),
                    delivery_type=item.get('delivery_type'),
                    rate=float(item.get('rate') or 0),
                    qty=float(item.get('qty') or 0),
                    cbm=float(item.get('cbm') or 0),
                    weight=float(item.get('weight') or 0),
                    surcharges=float(item.get('surcharges') or 0),
                    amount=float(item.get('amount') or 0),
                    note=item.get('note'),
                    registered_user=username,
                )

        container = Container.objects.get(container_number=container_number)
        invoice_item_data = []

        base_location = []
        #超重费
        if overweight_fee > 0:
            total_fee += overweight_fee
            overweight_extra_weight = request.POST.get("overweight_extra_weight")
            invoice_item_data.append(
                {
                    "item_category": "combina_extra_fee",
                    "container_number": container,
                    "invoice_number": invoice,
                    "invoice_type": "receivable",
                    "description": "超重费",
                    "warehouse_code": None,
                    "cbm": None,
                    "weight": overweight_extra_weight,
                    "qty": 1.0,
                    "rate": overweight_fee,
                    "amount": overweight_fee,
                    "note": None,
                }
            )
        
        #超板费
        if overpallet_fee > 0:
            total_fee += overweight_fee
            current_pallets = request.POST.get("current_pallets")
            limit_pallets = request.POST.get("limit_pallets")
            over_count = float(current_pallets) - float(limit_pallets)
            invoice_item_data.append(
                {
                    "item_category": "combina_extra_fee",
                    "container_number": container,
                    "invoice_number": invoice,
                    "invoice_type": "receivable",
                    "description": "超板费",
                    "warehouse_code": None,
                    "cbm": None,
                    "weight": None,
                    "qty": over_count,
                    "rate": overpallet_fee / over_count if over_count > 0 else 0,
                    "amount": overpallet_fee,
                    "note": None,
                }
            )

        #超区提拆费
        if overregion_pickup_fee > 0:
            total_fee += overregion_pickup_fee
            overregion_pickup_non_combina_cbm_ratio = request.POST.get(
                "overregion_pickup_non_combina_cbm_ratio"
            )
            overregion_pickup_non_combina_base_fee = request.POST.get(
                "overregion_pickup_non_combina_base_fee"
            )
            invoice_item_data.append(
                {
                    "item_category": "combina_extra_fee",
                    "container_number": container,
                    "invoice_number": invoice,
                    "invoice_type": "receivable",
                    "description": "提拆费",
                    "warehouse_code": None,
                    "cbm": None,
                    "weight": None,
                    "qty": float(overregion_pickup_non_combina_cbm_ratio) / 100,
                    "rate": overregion_pickup_non_combina_base_fee,
                    "amount": overregion_pickup_fee,
                    "note": None,
                }
            )
        
        #超区派送费
        if overregion_delivery_fee > 0:
            total_fee += overregion_delivery_fee
            overregion_delivery_destination = request.POST.getlist(
                "overregion_delivery_destination"
            )
            overregion_delivery_pallets = request.POST.getlist(
                "overregion_delivery_pallets"
            )
            overregion_delivery_cbm = request.POST.getlist("overregion_delivery_cbm")
            overregion_delivery_price = request.POST.getlist(
                "overregion_delivery_price"
            )
            overregion_delivery_subtotal = request.POST.getlist(
                "overregion_delivery_subtotal"
            )
            for i in range(len(overregion_delivery_destination)):
                if overregion_delivery_destination[i] not in base_location:
                    invoice_item_data.append(
                        {
                            "item_category": "combina_extra_fee",
                            "container_number": container,
                            "invoice_number": invoice,
                            "invoice_type": "receivable",
                            "description": "超区派送费",
                            "warehouse_code": overregion_delivery_destination[i],
                            "cbm": overregion_delivery_cbm[i],
                            "weight": None,
                            "qty": overregion_delivery_pallets[i],
                            "rate": overregion_delivery_price[i],
                            "amount": overregion_delivery_subtotal[i],
                            "note": None,
                        }
                    )

        if addition_fee:
            total_fee += overregion_delivery_fee
            invoice_item_data.append(
                {
                    "item_category": "combina_extra_fee",
                    "container_number": container,
                    "invoice_number": invoice,
                    "invoice_type": "receivable",
                    "description": "单柜超仓点费用",
                    "warehouse_code": None,
                    "cbm": None,
                    "weight": None,
                    "qty": 1,
                    "rate": addition_fee,
                    "amount": addition_fee,
                    "note": None,
                }
            )
        
        invoice_item_instances = [
            InvoiceItemv2(**inv_itm_data) for inv_itm_data in invoice_item_data
        ]
        bulk_create_with_history(invoice_item_instances, InvoiceItemv2)

        order = Order.objects.get(container_number__container_number=container_number)

        context = self._parse_invoice_excel_data(order, invoice)
        ac = Accounting()
        workbook, invoice_data = ac._generate_invoice_excel(context)
        invoice.invoice_date = invoice_data["invoice_date"]
        invoice.invoice_link = invoice_data["invoice_link"]

        #再计算一遍总数
        invoice.receivable_total_amount = total_fee
        invoice.remain_offset = total_fee
        invoice.save()
        status_obj = InvoiceStatusv2.objects.get(
            invoice=invoice,
            invoice_type='receivable'
        )
        status_obj.finance_status = "completed"
        status_obj.save()
        ctx = {'success_messages': '保存成功！'}
        return self.handle_confirm_entry_post(request,ctx)
    
    


    